"""AESA API: Multi-D allocation model endpoints.

Reference data (boundary sets, Multi-D defaults, SSP trajectories, carbon
budget options, default sharing values), per-project AESA configurations,
compute endpoint, and xlsx export.
"""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone

from fastapi import APIRouter, Depends, HTTPException, Response, UploadFile, File

from mapper.api.project_guard import verify_project_state
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from mapper.api.bom import _sanitize_filename
from mapper.api.impact import _TASKS as IMPACT_TASKS, _TASK_LOCK as IMPACT_LOCK
from mapper.api.dsm import _current_project, _get_system
from mapper.core import aesa_session_storage, aesa_storage, sharing_preset_storage
from mapper.core.aesa_engine import (
    AESAEngine,
    BUILTIN_PRINCIPLES,
    MULTI_D_DEFAULTS,
    build_carbon_budget,
    build_default_multi_d_config,
    build_default_sharing_preset,
    co2e_conversion_for_budget,
    load_boundary_sets,
    single_product_to_impact_result,
    load_carbon_budget_options,
    load_sharing_data,
    load_ssp_trajectories,
    resolve_sharing,
    suggest_method_mapping,
)
from mapper.models.aesa_schemas import (
    AESAComputeRequest,
    AESAComputeResult,
    AESAConfiguration,
    AESAConfigurationCreate,
    AESAExportRequest,
    AESASession,
    AESASessionCreate,
    AESASessionRename,
    BoundarySet,
    CategoryAssignment,
    DownscalingChain,
    DownscalingLayer,
    MethodPBMapping,
    PlanetaryBoundary,
    PrincipleDefinition,
    SharingPreset,
    SharingPresetCreate,
)
from mapper.models.bom_schemas import ImpactAssessmentResult


router = APIRouter(prefix="/aesa", tags=["aesa"])


# ─── Reference data ──────────────────────────────────────────────────────────


@router.get("/boundary-sets")
async def get_boundary_sets() -> list[BoundarySet]:
    return list(load_boundary_sets().values())


@router.get("/boundary-sets/{set_id}")
async def get_boundary_set(set_id: str) -> BoundarySet:
    sets = load_boundary_sets()
    if set_id not in sets:
        raise HTTPException(status_code=404, detail=f"Boundary set '{set_id}' not found")
    return sets[set_id]


@router.get("/multi-d-defaults")
async def get_multi_d_defaults() -> list[dict]:
    """Returns the default per-PB sharing-principle assignments."""
    return [
        {"pb_id": pb_id, "principle": principle, "justification": just}
        for pb_id, (principle, just) in MULTI_D_DEFAULTS.items()
    ]


@router.get("/sharing-data")
async def get_sharing_data() -> dict:
    return load_sharing_data()


@router.get("/ssp-trajectories")
async def get_ssp_trajectories() -> list[dict]:
    return load_ssp_trajectories()


@router.get("/carbon-budget-options")
async def get_carbon_budget_options() -> list[dict]:
    return load_carbon_budget_options()


@router.post("/method-mapping/suggest")
async def post_method_mapping_suggest(body: dict) -> list[MethodPBMapping]:
    """Body: ``{"methods": [[...], ...], "boundary_set_id": "Sala2020_EF"}``."""
    methods = body.get("methods") or []
    set_id = body.get("boundary_set_id", "Sala2020_EF")
    sets = load_boundary_sets()
    bset = sets.get(set_id)
    if bset is None:
        raise HTTPException(status_code=404, detail=f"Boundary set '{set_id}' not found")
    return suggest_method_mapping(methods, bset)


@router.get("/defaults")
async def get_defaults() -> dict:
    """Full bundle a fresh UI can load in one call: boundary set(s), Multi-D
    defaults, sharing data, SSP list, carbon budget options, plus a
    ready-built default MultiDConfig + CarbonBudgetConfig."""
    return {
        "boundary_sets": [bs.model_dump() for bs in load_boundary_sets().values()],
        "multi_d_defaults": [
            {"pb_id": pb_id, "principle": principle, "justification": just}
            for pb_id, (principle, just) in MULTI_D_DEFAULTS.items()
        ],
        "sharing_data": load_sharing_data(),
        "ssp_trajectories": load_ssp_trajectories(),
        # Attach each option's per-budget CO2→CO2e conversion so the frontend can
        # patch `co2e_conversion` when the user switches the budget option
        # (the factor is temperature/x-dependent, recomputed per option).
        "carbon_budget_options": [
            {**o, "co2e_conversion": co2e_conversion_for_budget(o).model_dump()}
            for o in load_carbon_budget_options()
        ],
        "default_multi_d": build_default_multi_d_config().model_dump(),
        "default_carbon_budget": build_carbon_budget().model_dump(),
    }


# ─── Configuration CRUD ──────────────────────────────────────────────────────


@router.get("/configurations", response_model=list[AESAConfiguration])
async def get_configurations() -> list[AESAConfiguration]:
    project = _current_project()
    raw = aesa_storage.load_all(project)
    out: list[AESAConfiguration] = []
    for r in raw:
        try:
            out.append(AESAConfiguration(**r))
        except Exception:
            continue
    out.sort(key=lambda c: c.created_at, reverse=True)
    return out


@router.get("/configurations/{config_id}", response_model=AESAConfiguration)
async def get_configuration(config_id: str) -> AESAConfiguration:
    project = _current_project()
    raw = aesa_storage.load(project, config_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="AESA configuration not found")
    return AESAConfiguration(**raw)


@router.post(
    "/configurations",
    response_model=AESAConfiguration,
    dependencies=[Depends(verify_project_state)],
)
async def post_configuration(body: AESAConfigurationCreate) -> AESAConfiguration:
    _get_system(body.mfa_system_id)  # 404 if system missing
    project = _current_project()
    config = AESAConfiguration(
        id=uuid.uuid4().hex,
        name=body.name,
        mfa_system_id=body.mfa_system_id,
        impact_mode=body.impact_mode,
        boundary_set_id=body.boundary_set_id,
        multi_d=body.multi_d,
        carbon_budget=body.carbon_budget,
        method_mapping=body.method_mapping,
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    aesa_storage.save(project, config.model_dump())
    return config


@router.put("/configurations/{config_id}", response_model=AESAConfiguration)
async def put_configuration(
    config_id: str, body: AESAConfigurationCreate,
) -> AESAConfiguration:
    project = _current_project()
    existing = aesa_storage.load(project, config_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="AESA configuration not found")
    _get_system(body.mfa_system_id)
    updated = AESAConfiguration(
        id=config_id,
        name=body.name,
        mfa_system_id=body.mfa_system_id,
        impact_mode=body.impact_mode,
        boundary_set_id=body.boundary_set_id,
        multi_d=body.multi_d,
        carbon_budget=body.carbon_budget,
        method_mapping=body.method_mapping,
        created_at=existing.get("created_at") or datetime.now(timezone.utc).isoformat(),
    )
    aesa_storage.save(project, updated.model_dump())
    return updated


@router.delete("/configurations/{config_id}")
async def delete_configuration(config_id: str) -> dict:
    project = _current_project()
    deleted = aesa_storage.delete(project, config_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="AESA configuration not found")
    return {"deleted": True}


# ─── Saved sessions (Patch 4R) ───────────────────────────────────────────────


@router.get("/sessions", response_model=list[AESASession])
async def list_sessions() -> list[AESASession]:
    project = _current_project()
    raw = aesa_session_storage.load_all(project)
    out: list[AESASession] = []
    for r in raw:
        try:
            out.append(AESASession(**r))
        except Exception:
            # Skip malformed sessions rather than 500ing the list. A
            # corrupt file (manual edit, schema drift) shouldn't block
            # access to the remaining sessions.
            continue
    return out


@router.get("/sessions/{session_id}", response_model=AESASession)
async def get_session(session_id: str) -> AESASession:
    project = _current_project()
    raw = aesa_session_storage.load(project, session_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="AESA session not found")
    return AESASession(**raw)


@router.post(
    "/sessions",
    response_model=AESASession,
    dependencies=[Depends(verify_project_state)],
)
async def create_session(body: AESASessionCreate) -> AESASession:
    project = _current_project()
    now = datetime.now(timezone.utc).isoformat()
    session = AESASession(
        id=uuid.uuid4().hex,
        name=body.name,
        project=project or "default",
        created_at=now,
        modified_at=now,
        configuration_snapshot=body.configuration_snapshot,
        result=body.result,
        upstream_ia_task_id=body.upstream_ia_task_id,
        # Patch 4T — round-trip the display filter. ``None`` (default)
        # restores as "show all" on load; an explicit list narrows the
        # view to the saved subset.
        displayed_indicators=body.displayed_indicators,
    )
    aesa_session_storage.save(project, session.model_dump())
    return session


@router.patch("/sessions/{session_id}", response_model=AESASession)
async def rename_session(session_id: str, body: AESASessionRename) -> AESASession:
    project = _current_project()
    existing = aesa_session_storage.load(project, session_id)
    if existing is None:
        raise HTTPException(status_code=404, detail="AESA session not found")
    existing["name"] = body.name
    existing["modified_at"] = datetime.now(timezone.utc).isoformat()
    aesa_session_storage.save(project, existing)
    return AESASession(**existing)


@router.delete("/sessions/{session_id}")
async def delete_session(session_id: str) -> dict:
    project = _current_project()
    deleted = aesa_session_storage.delete(project, session_id)
    if not deleted:
        raise HTTPException(status_code=404, detail="AESA session not found")
    return {"deleted": True}


# ─── Compute ─────────────────────────────────────────────────────────────────


def _resolve_impact(
    task_id: str | None, inline: ImpactAssessmentResult | None,
) -> ImpactAssessmentResult:
    if inline is not None:
        return inline
    if task_id is None:
        raise HTTPException(status_code=400, detail="impact_task_id or impact_result required")
    with IMPACT_LOCK:
        task = IMPACT_TASKS.get(task_id)
    if task is None or task.result is None:
        raise HTTPException(status_code=404, detail=f"Impact task {task_id} not found or not finished")
    return task.result


def _resolve_config(body: AESAComputeRequest) -> AESAConfiguration:
    if body.config is not None:
        return body.config
    if body.config_id is None:
        raise HTTPException(status_code=400, detail="config_id or config required")
    project = _current_project()
    raw = aesa_storage.load(project, body.config_id)
    if raw is None:
        raise HTTPException(status_code=404, detail="AESA configuration not found")
    return AESAConfiguration(**raw)


@router.post("/compute", response_model=AESAComputeResult)
async def post_compute(body: AESAComputeRequest) -> AESAComputeResult:
    from mapper.core.compute_metrics import measure_compute
    meter = measure_compute()
    config = _resolve_config(body)
    if body.single_product_result is not None:
        # Single-LCA (non-fleet) source: adapt the static single-product result
        # into the per-year impact the engine consumes (reference_year sets the
        # climate annual-allowance year). Takes precedence over task/inline.
        impact = single_product_to_impact_result(
            body.single_product_result,
            reference_year=body.reference_year,
            system_id=config.mfa_system_id,
        )
    else:
        impact = _resolve_impact(body.impact_task_id, body.impact_result)
    # The system-match check fires only when BOTH the impact and the config
    # carry a system id (fleet path, unchanged). A single-LCA source has no DSM
    # system (meta.mfa_system_id is None) → the check is skipped.
    if (impact.meta.mfa_system_id and config.mfa_system_id
            and impact.meta.mfa_system_id != config.mfa_system_id):
        raise HTTPException(
            status_code=400,
            detail="Impact result is for a different DSM system than the AESA config.",
        )
    sets = load_boundary_sets()
    bset = sets.get(config.boundary_set_id)
    if bset is None:
        raise HTTPException(
            status_code=400,
            detail=f"Boundary set '{config.boundary_set_id}' not found",
        )
    # Patch 2c — scaffold guard: a structure-only boundary set (e.g.
    # Ryberg2018_PBLCIA) is marked not-computable and/or carries null SOS
    # (pb_value). Reject with a clear message rather than crashing on a null
    # pb_value × factor or a null ef_indicator in suggest_method_mapping.
    if (not bset.computable
            or any(pb.pb_value is None for pb in bset.boundaries.values())):
        raise HTTPException(
            status_code=400,
            detail=(
                f"Boundary set '{bset.name}' is scaffolded but not yet "
                "computable: it needs a PB-LCIA characterisation method and "
                "SOS (planetary boundary) values before AESA can compute "
                "sustainability ratios against it."
            ),
        )
    # Patch 2d — inert guard: a CO2e/GHG carbon-budget basis is opt-in and stays
    # INERT until a sourced CO2→CO2e conversion is supplied. Reject rather than
    # silently computing on the CO2 budget (wrong scope) or a fabricated factor.
    cb = config.carbon_budget
    if cb is not None and cb.budget_basis == "CO2e_GHG" and cb.co2e_ratio() is None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Carbon budget set to CO2e/GHG basis but no sourced CO2→CO2e "
                "conversion supplied; supply a per-scenario conversion (ratio) "
                "or use the CO2 basis."
            ),
        )
    if body.run_sensitivity:
        result = AESAEngine.compute_with_sensitivity(impact.results, config, bset)
    else:
        result = AESAEngine.compute(impact.results, config, bset)
    result.compute_metrics = meter.build()
    return result


# ─── Sharing Presets (global) ────────────────────────────────────────────────


@router.get("/sharing-presets", response_model=list[SharingPreset])
async def get_sharing_presets() -> list[SharingPreset]:
    raw = sharing_preset_storage.load_all()
    out: list[SharingPreset] = []
    for r in raw:
        try:
            out.append(SharingPreset(**r))
        except Exception:
            continue
    return out


@router.get("/sharing-presets/{preset_id}", response_model=SharingPreset)
async def get_sharing_preset(preset_id: str) -> SharingPreset:
    raw = sharing_preset_storage.load(preset_id)
    if raw is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    return SharingPreset(**raw)


@router.post("/sharing-presets", response_model=SharingPreset)
async def post_sharing_preset(body: SharingPresetCreate) -> SharingPreset:
    now = datetime.now(timezone.utc).isoformat()
    preset = SharingPreset(
        id=uuid.uuid4().hex,
        name=body.name,
        description=body.description,
        built_in=False,
        principles=body.principles,
        category_assignments=body.category_assignments,
        chain=body.chain,
        created_at=now,
        updated_at=now,
    )
    sharing_preset_storage.save(preset.model_dump())
    return preset


@router.put("/sharing-presets/{preset_id}", response_model=SharingPreset)
async def put_sharing_preset(preset_id: str, body: SharingPresetCreate) -> SharingPreset:
    if sharing_preset_storage.is_built_in(preset_id):
        raise HTTPException(
            status_code=400,
            detail="Built-in presets are read-only. Duplicate the preset to customize it.",
        )
    existing = sharing_preset_storage.load(preset_id)
    if existing is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    updated = SharingPreset(
        id=preset_id,
        name=body.name,
        description=body.description,
        built_in=False,
        principles=body.principles,
        category_assignments=body.category_assignments,
        chain=body.chain,
        created_at=existing.get("created_at") or datetime.now(timezone.utc).isoformat(),
        updated_at=datetime.now(timezone.utc).isoformat(),
    )
    sharing_preset_storage.save(updated.model_dump())
    return updated


@router.delete("/sharing-presets/{preset_id}")
async def delete_sharing_preset(preset_id: str) -> dict:
    if sharing_preset_storage.is_built_in(preset_id):
        raise HTTPException(status_code=400, detail="Built-in presets cannot be deleted.")
    if not sharing_preset_storage.delete(preset_id):
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    return {"deleted": True}


@router.post("/sharing-presets/{preset_id}/duplicate", response_model=SharingPreset)
async def post_duplicate_preset(preset_id: str, body: dict | None = None) -> SharingPreset:
    """Duplicate any preset (including built-ins) into a new editable one."""
    raw = sharing_preset_storage.load(preset_id)
    if raw is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    src = SharingPreset(**raw)
    new_name = (body or {}).get("name") or f"{src.name} (copy)"
    now = datetime.now(timezone.utc).isoformat()
    dup = SharingPreset(
        id=uuid.uuid4().hex,
        name=new_name,
        description=src.description,
        built_in=False,
        principles=list(src.principles),
        category_assignments=list(src.category_assignments),
        chain=src.chain.model_copy(deep=True),
        created_at=now,
        updated_at=now,
    )
    sharing_preset_storage.save(dup.model_dump())
    return dup


# ─── Downscaling chain (convenience: read / update the chain on a preset) ────


@router.get("/downscaling-chain/{preset_id}", response_model=DownscalingChain)
async def get_downscaling_chain(preset_id: str) -> DownscalingChain:
    raw = sharing_preset_storage.load(preset_id)
    if raw is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    return SharingPreset(**raw).chain


@router.put("/downscaling-chain/{preset_id}", response_model=SharingPreset)
async def put_downscaling_chain(preset_id: str, chain: DownscalingChain) -> SharingPreset:
    if sharing_preset_storage.is_built_in(preset_id):
        raise HTTPException(
            status_code=400,
            detail="Built-in presets are read-only. Duplicate the preset first.",
        )
    raw = sharing_preset_storage.load(preset_id)
    if raw is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    preset = SharingPreset(**raw).model_copy(update={
        "chain": chain,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    })
    sharing_preset_storage.save(preset.model_dump())
    return preset


# ─── Sharing xlsx: template / export / import ────────────────────────────────


@router.get("/sharing/template")
async def get_sharing_template() -> Response:
    """Download an xlsx template pre-filled with the built-in Ferhati preset."""
    preset = build_default_sharing_preset()
    wb = _build_sharing_workbook(preset, include_instructions=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="aesa_sharing_template.xlsx"'},
    )


@router.get("/sharing/export/{preset_id}")
async def get_sharing_export(preset_id: str) -> Response:
    raw = sharing_preset_storage.load(preset_id)
    if raw is None:
        raise HTTPException(status_code=404, detail=f"Preset '{preset_id}' not found")
    preset = SharingPreset(**raw)
    wb = _build_sharing_workbook(preset, include_instructions=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{_sanitize_filename(preset.name, 'aesa_sharing')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/sharing/import", response_model=SharingPreset)
async def post_sharing_import(
    file: UploadFile = File(...),
    name: str | None = None,
) -> SharingPreset:
    """Parse an uploaded xlsx into a new editable preset and persist it."""
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read xlsx: {e}") from e
    try:
        preset = _parse_sharing_workbook(wb, name or (file.filename or "Imported preset"))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e)) from e
    now = datetime.now(timezone.utc).isoformat()
    preset = preset.model_copy(update={
        "id": uuid.uuid4().hex,
        "built_in": False,
        "created_at": now,
        "updated_at": now,
    })
    sharing_preset_storage.save(preset.model_dump())
    return preset


# ─── Export ──────────────────────────────────────────────────────────────────


@router.post("/export")
async def post_export(body: AESAExportRequest) -> Response:
    config = body.config
    sys_def = _get_system(config.mfa_system_id)

    wb = _build_aesa_workbook(config, body.result, sys_def.name)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{_sanitize_filename(sys_def.name, 'system')}_aesa.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_aesa_workbook(
    config: AESAConfiguration, result: AESAComputeResult, system_name: str,
) -> Workbook:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="064E3B")
    num_fmt = "0.000E+00"

    wb = Workbook()
    wb.remove(wb.active)

    def _autosize(ws) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for cell in col_cells:
                v = cell.value
                if v is None:
                    continue
                widest = max(widest, min(60, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    def _style_header(ws) -> None:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    # ── Summary ──
    ws = wb.create_sheet("Summary")
    ws.append(["System", system_name])
    ws.append(["AESA Configuration", config.name])
    ws.append(["Boundary Set", config.boundary_set_id])
    ws.append(["Impact Mode", config.impact_mode])
    # Patch 5AS — metadata header so an exported workbook is self-describing.
    ws.append(["DSM scenario", config.dsm_scenario_id or "(active)"])
    # Patch 5AT — modern sharing-preset configs have multi_d=None; only emit
    # the legacy Layer-2 row when the 2-layer Multi-D config is present.
    if config.multi_d is not None:
        ws.append(["Layer 2 (sector share)", config.multi_d.layer2_sector_share])
    if config.carbon_budget is not None:
        cb = config.carbon_budget
        ws.append(["Carbon budget", f"{cb.budget_source} — {cb.initial_budget_gt} Gt"])
        ws.append(["SSP scenario", cb.ssp_scenario])
        ws.append(["Budget horizon", f"{cb.start_year}–{cb.end_year}"])
    ws.append(["Sensitivity", "all 5 principles" if getattr(result, "sensitivity", None) else "primary only"])
    ws.append([])
    hdr_row = ws.max_row + 1
    ws.append(["Year", "Safe", "Zone of Uncertainty", "High Risk", "Assessed"])
    for cell in ws[hdr_row]:
        cell.font = header_font
        cell.fill = header_fill
    for y in result.summary_by_year:
        ws.append([y.year, y.safe, y.zone_of_uncertainty, y.high_risk, y.total_assessed])
    _autosize(ws)

    # ── Sustainability Ratios (matrix) ──
    ws = wb.create_sheet("Sustainability Ratios")
    pb_ids: list[str] = []
    pb_names: dict[str, str] = {}
    seen = set()
    for r in result.results:
        if r.pb_id not in seen:
            seen.add(r.pb_id)
            pb_ids.append(r.pb_id)
            pb_names[r.pb_id] = r.pb_name
    years_sorted = sorted({r.year for r in result.results})
    ws.append(["Year", *[pb_names[p] for p in pb_ids]])
    _style_header(ws)
    sr_lookup: dict[tuple[int, str], float] = {
        (r.year, r.pb_id): r.sr for r in result.results
    }
    for y in years_sorted:
        ws.append([y, *[sr_lookup.get((y, p), "") for p in pb_ids]])
    for row in ws.iter_rows(min_row=2, min_col=2):
        for cell in row:
            cell.number_format = "0.000"
    ws.freeze_panes = "A2"
    _autosize(ws)

    # ── Impacts vs SOS (long-form) ──
    ws = wb.create_sheet("Impacts vs SOS")
    # Patch 5AS — each per-year, per-indicator row now carries the numbers
    # behind the SR timeline: system impact (numerator), assigned SOS share
    # (denominator), SR, the total system share, and — for climate (cumulative)
    # — the allocation chain (global remaining budget → per-year global
    # allocation → × system share = allocated SOS). All read from the
    # authoritative result row; nothing recomputed in the export.
    # Patch 2d — relabel the carbon-budget chain columns as CO2e when the
    # budget basis is CO2e/GHG (the values are CO2e-scaled by compute). CO2
    # basis (default) keeps the original "(Gt)" / "(Gt/yr)" labels — no drift.
    _co2e = (config.carbon_budget is not None
             and config.carbon_budget.budget_basis == "CO2e_GHG"
             and config.carbon_budget.co2e_ratio() is not None)
    _rem_lbl = "Remaining Budget (Gt CO2e)" if _co2e else "Remaining Budget (Gt)"
    _alloc_lbl = "Global Allocation (Gt CO2e/yr)" if _co2e else "Global Allocation (Gt/yr)"
    ws.append([
        "Year", "PB ID", "PB Name", "EF Indicator", "Method",
        "Impact", "Allocated SOS", "SR", "Zone",
        "Principle", "L1 Factor", "L2 Factor", "Boundary Type", "Unit",
        "System Share", _rem_lbl, _alloc_lbl,
    ])
    _style_header(ws)
    for r in result.results:
        ws.append([
            r.year, r.pb_id, r.pb_name, r.ef_indicator, r.method_label,
            r.impact, r.allocated_sos, r.sr, r.zone,
            r.sharing_principle, r.sharing_factor_l1, r.sharing_factor_l2,
            r.boundary_type, r.unit,
            r.total_sharing_factor, r.remaining_budget_gt, r.global_allocation_gt,
        ])
    for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
        for cell in row:
            cell.number_format = num_fmt
    for row in ws.iter_rows(min_row=2, min_col=8, max_col=8):
        for cell in row:
            cell.number_format = "0.000"
    for row in ws.iter_rows(min_row=2, min_col=15, max_col=15):  # System Share
        for cell in row:
            cell.number_format = num_fmt
    for row in ws.iter_rows(min_row=2, min_col=16, max_col=17):  # Gt columns
        for cell in row:
            if cell.value is not None:
                cell.number_format = "0.00"
    ws.freeze_panes = "A2"
    _autosize(ws)

    # ── By Fuel Type (cohort breakdown) ──
    ws = wb.create_sheet("By Fuel Type")
    ws.append(["Year", "PB", "Cohort", "Impact"])
    _style_header(ws)
    for r in result.results:
        for cohort, val in r.impact_by_cohort.items():
            ws.append([r.year, r.pb_name, cohort, val])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = num_fmt
    ws.freeze_panes = "A2"
    _autosize(ws)

    # ── Multi-D Configuration ──
    # Patch 5AT — legacy 2-layer Multi-D sheet; skipped entirely for modern
    # sharing-preset configs (multi_d=None). The sharing chain is documented by
    # the preset itself elsewhere; this sheet is for the legacy config shape.
    if config.multi_d is not None:
        ws = wb.create_sheet("Multi-D Configuration")
        ws.append(["PB ID", "Principle", "Justification", "System Value", "Global Value"])
        _style_header(ws)
        for pb_id, sp in config.multi_d.layer1.items():
            ws.append([
                pb_id, sp.principle, sp.justification,
                sp.system_value, sp.global_value,
            ])
        ws.append([])
        ws.append(["Layer 2 sector share", config.multi_d.layer2_sector_share])
        ws.append(["Layer 2 source", config.multi_d.layer2_source])
        _autosize(ws)

    # ── Carbon Budget ──
    if config.carbon_budget is not None:
        ws = wb.create_sheet("Carbon Budget")
        cb = config.carbon_budget
        ws.append(["Initial budget (Gt CO2)", cb.initial_budget_gt])
        ws.append(["Source", cb.budget_source])
        ws.append(["SSP scenario", cb.ssp_scenario])
        ws.append(["Start year", cb.start_year])
        ws.append(["End year", cb.end_year])
        ws.append([])
        ws.append(["Year", "Projected global CO2 (Gt)", "Remaining budget (Gt)",
                   "Annual global allocation (Gt)"])
        for row in ws.iter_rows(min_row=7, max_row=7):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
        for y in range(cb.start_year, cb.end_year + 1):
            ws.append([
                y,
                cb.projected_emissions.get(y, 0.0),
                cb.remaining_budget(y),
                cb.annual_global_allocation(y),
            ])
        _autosize(ws)

    # ── Sensitivity (if run) ──
    if result.sensitivity:
        ws = wb.create_sheet("Sensitivity")
        ws.append(["Scenario", "Year", "PB", "SR", "Zone", "Allocated SOS"])
        _style_header(ws)
        # Baseline Multi-D first, then each uniform principle
        for r in result.results:
            ws.append([
                "Multi-D", r.year, r.pb_name, r.sr, r.zone, r.allocated_sos,
            ])
        for principle, rows in result.sensitivity.items():
            for r in rows:
                ws.append([
                    principle, r.year, r.pb_name, r.sr, r.zone, r.allocated_sos,
                ])
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
            for cell in row:
                cell.number_format = "0.000"
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=6):
            for cell in row:
                cell.number_format = num_fmt
        ws.freeze_panes = "A2"
        _autosize(ws)

    # ── Methodology Notes ──
    ws = wb.create_sheet("Methodology")
    ws.append(["Field", "Value"])
    _style_header(ws)
    ws.append(["Configuration", config.name])
    ws.append(["Boundary set", config.boundary_set_id])
    ws.append(["Impact mode", config.impact_mode])
    # Patch 5AT — legacy Multi-D rows only when multi_d is present (None for
    # sharing-preset configs).
    if config.multi_d is not None:
        ws.append(["Layer 2 (grandfathering)", f"{config.multi_d.layer2_sector_share} — {config.multi_d.layer2_source}"])
        principles = sorted({sp.principle for sp in config.multi_d.layer1.values()})
        ws.append(["Layer 1 principles used", ", ".join(principles)])
    if config.carbon_budget:
        ws.append(["Carbon budget", f"{config.carbon_budget.initial_budget_gt} Gt — {config.carbon_budget.budget_source}"])
        ws.append(["SSP scenario", config.carbon_budget.ssp_scenario])
    ws.append(["Missing PB categories",
               ", ".join(result.missing_categories) if result.missing_categories else "none"])
    ws.append(["Uncertainty", "Deterministic — Monte Carlo planned for v1.1"])
    _autosize(ws)

    return wb


# ─── Sharing preset xlsx (template / export / import helpers) ────────────────


_SHARING_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SHARING_HEADER_FILL = PatternFill("solid", fgColor="064E3B")


def _style_sharing_header(ws) -> None:
    for cell in ws[1]:
        cell.font = _SHARING_HEADER_FONT
        cell.fill = _SHARING_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _autosize_sharing(ws) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        widest = 0
        for cell in col_cells:
            v = cell.value
            if v is None:
                continue
            widest = max(widest, min(60, len(str(v))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)


def _build_sharing_workbook(preset: SharingPreset, include_instructions: bool = True) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Principles
    ws = wb.create_sheet("Principles")
    ws.append(["Principle ID", "Name", "Description"])
    _style_sharing_header(ws)
    for p in preset.principles:
        ws.append([p.id, p.name, p.description])
    _autosize_sharing(ws)

    # Sheet 2: Category Assignments
    ws = wb.create_sheet("Category Assignments")
    ws.append(["PB ID", "Principle ID", "Justification"])
    _style_sharing_header(ws)
    for a in preset.category_assignments:
        ws.append([a.pb_id, a.principle_id, a.justification])
    _autosize_sharing(ws)

    # Sheet 3: Downscaling Chain
    ws = wb.create_sheet("Downscaling Chain")
    ws.append(["Layer", "Name", "Mode", "Fixed Principle", "Description"])
    _style_sharing_header(ws)
    for layer in preset.chain.layers:
        ws.append([
            layer.layer_number, layer.name, layer.principle_mode,
            layer.fixed_principle or "", layer.description,
        ])
    _autosize_sharing(ws)

    # Sheet 4: Sharing Data
    ws = wb.create_sheet("Sharing Data")
    ws.append(["Layer", "Principle", "System Value", "Global Value", "Year", "Source"])
    _style_sharing_header(ws)
    for layer in preset.chain.layers:
        for principle_id, years in layer.data.items():
            for year, pair in sorted(years.items()):
                sys_val, glob_val = pair
                ws.append([layer.layer_number, principle_id, sys_val, glob_val, year, ""])
    _autosize_sharing(ws)

    # Sheet 5: Instructions
    if include_instructions:
        ws = wb.create_sheet("Instructions")
        rows = [
            ["AESA Sharing Preset — Template"],
            [""],
            ["This workbook defines a sharing preset: which principle applies to each"],
            ["impact category, how the downscaling chain is structured, and the data"],
            ["that parameterises each layer."],
            [""],
            ["Sheet: Principles"],
            ["  One row per sharing principle. 'Principle ID' is the short key referenced"],
            ["  elsewhere (e.g. EpC, IN, GDP). You may add custom principles."],
            [""],
            ["Sheet: Category Assignments"],
            ["  Maps each planetary boundary (PB ID) to a principle. Used by layers"],
            ["  whose Mode is 'category_specific'."],
            [""],
            ["Sheet: Downscaling Chain"],
            ["  Ordered list of layers (minimum 1)."],
            ["    Layer            : integer order (1, 2, 3, ...)."],
            ["    Name             : human-readable label."],
            ["    Mode             : 'category_specific' or 'fixed'."],
            ["    Fixed Principle  : required if Mode = 'fixed'."],
            [""],
            ["Sheet: Sharing Data"],
            ["  (Layer × Principle × Year) → (System Value, Global Value)."],
            ["  Factor = System Value / Global Value."],
            ["  One row per combination. If only one year is given for a principle"],
            ["  the engine uses it as constant across all years."],
            [""],
            ["Allocated SOS formula"],
            ["  allocated_sos(pb, year) = PB_value × ∏ layer_factor(layer, pb, year)"],
        ]
        for r in rows:
            ws.append(r)
        ws.column_dimensions["A"].width = 95

    return wb


def _norm_header(value) -> str:
    return str(value).strip().lower() if value is not None else ""


def _read_sheet_rows(ws) -> tuple[list[str], list[list]]:
    """Return (headers lowercased, data rows) ignoring fully-empty trailing rows."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    headers = [_norm_header(h) for h in rows[0]]
    data = []
    for r in rows[1:]:
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in r):
            continue
        data.append(list(r))
    return headers, data


def _col(headers: list[str], *aliases: str) -> int:
    """Find a column by any of its alias names; -1 if absent."""
    for alias in aliases:
        key = alias.strip().lower()
        if key in headers:
            return headers.index(key)
    return -1


def _parse_sharing_workbook(wb: Workbook, default_name: str) -> SharingPreset:
    required = {"Principles", "Category Assignments", "Downscaling Chain", "Sharing Data"}
    missing = required - set(wb.sheetnames)
    if missing:
        raise ValueError(f"xlsx missing required sheet(s): {sorted(missing)}")

    # Principles
    p_headers, p_rows = _read_sheet_rows(wb["Principles"])
    i_id = _col(p_headers, "principle id", "id")
    i_name = _col(p_headers, "name")
    i_desc = _col(p_headers, "description")
    if i_id < 0 or i_name < 0:
        raise ValueError("Principles sheet requires 'Principle ID' and 'Name' columns.")
    principles: list[PrincipleDefinition] = []
    for r in p_rows:
        pid = r[i_id]
        if pid is None or str(pid).strip() == "":
            continue
        principles.append(PrincipleDefinition(
            id=str(pid).strip(),
            name=str(r[i_name] or pid).strip(),
            description=str(r[i_desc] or "").strip() if i_desc >= 0 else "",
        ))
    if not principles:
        raise ValueError("Principles sheet is empty.")
    principle_ids = {p.id for p in principles}

    # Category Assignments
    a_headers, a_rows = _read_sheet_rows(wb["Category Assignments"])
    i_pb = _col(a_headers, "pb id", "impact category")
    i_pri = _col(a_headers, "principle id", "sharing principle", "principle")
    i_just = _col(a_headers, "justification")
    if i_pb < 0 or i_pri < 0:
        raise ValueError("Category Assignments sheet requires 'PB ID' and 'Principle ID' columns.")
    assignments: list[CategoryAssignment] = []
    for r in a_rows:
        pb = r[i_pb]
        pri = r[i_pri]
        if pb is None or pri is None:
            continue
        pri_s = str(pri).strip()
        if pri_s not in principle_ids:
            raise ValueError(
                f"Category '{pb}' references unknown principle '{pri_s}'. "
                f"Add it to the Principles sheet first.",
            )
        assignments.append(CategoryAssignment(
            pb_id=str(pb).strip(),
            principle_id=pri_s,
            justification=str(r[i_just] or "").strip() if i_just >= 0 else "",
        ))

    # Downscaling Chain (definitions)
    c_headers, c_rows = _read_sheet_rows(wb["Downscaling Chain"])
    i_layer = _col(c_headers, "layer")
    i_ln = _col(c_headers, "name")
    i_mode = _col(c_headers, "mode")
    i_fp = _col(c_headers, "fixed principle")
    i_cdesc = _col(c_headers, "description")
    if i_layer < 0 or i_ln < 0 or i_mode < 0:
        raise ValueError("Downscaling Chain sheet requires 'Layer', 'Name', 'Mode' columns.")
    layers_meta: dict[int, dict] = {}
    for r in c_rows:
        if r[i_layer] is None:
            continue
        try:
            num = int(r[i_layer])
        except (TypeError, ValueError) as e:
            raise ValueError(f"Layer column must be integer, got {r[i_layer]!r}") from e
        mode = str(r[i_mode]).strip().lower() if r[i_mode] is not None else ""
        if mode not in ("category_specific", "fixed"):
            raise ValueError(
                f"Layer {num}: mode must be 'category_specific' or 'fixed', got '{mode}'.",
            )
        fp = str(r[i_fp]).strip() if i_fp >= 0 and r[i_fp] else None
        if mode == "fixed":
            if not fp:
                raise ValueError(f"Layer {num}: 'Fixed Principle' required when mode=fixed.")
            if fp not in principle_ids:
                raise ValueError(f"Layer {num}: unknown principle '{fp}'.")
        layers_meta[num] = {
            "name": str(r[i_ln] or f"Layer {num}").strip(),
            "mode": mode,
            "fixed_principle": fp if mode == "fixed" else None,
            "description": str(r[i_cdesc] or "").strip() if i_cdesc >= 0 else "",
            "data": {},
        }
    if not layers_meta:
        raise ValueError("Downscaling Chain sheet is empty.")

    # Sharing Data — populate layer data
    d_headers, d_rows = _read_sheet_rows(wb["Sharing Data"])
    i_dl = _col(d_headers, "layer")
    i_dp = _col(d_headers, "principle", "principle id")
    i_sv = _col(d_headers, "system value")
    i_gv = _col(d_headers, "global value")
    i_yr = _col(d_headers, "year")
    if min(i_dl, i_dp, i_sv, i_gv, i_yr) < 0:
        raise ValueError(
            "Sharing Data sheet requires Layer, Principle, System Value, Global Value, Year columns.",
        )
    for r in d_rows:
        if r[i_dl] is None or r[i_dp] is None:
            continue
        try:
            num = int(r[i_dl])
            year = int(r[i_yr])
            sys_v = float(r[i_sv])
            glob_v = float(r[i_gv])
        except (TypeError, ValueError) as e:
            raise ValueError(f"Sharing Data row has invalid numeric cell: {r}") from e
        if num not in layers_meta:
            raise ValueError(
                f"Sharing Data references layer {num}, not defined in Downscaling Chain.",
            )
        pid = str(r[i_dp]).strip()
        if pid not in principle_ids:
            raise ValueError(f"Sharing Data row references unknown principle '{pid}'.")
        layers_meta[num]["data"].setdefault(pid, {})[year] = (sys_v, glob_v)

    # Build layers in order
    layers: list[DownscalingLayer] = []
    for num in sorted(layers_meta):
        meta = layers_meta[num]
        layers.append(DownscalingLayer(
            layer_number=num,
            name=meta["name"],
            principle_mode=meta["mode"],
            fixed_principle=meta["fixed_principle"],
            description=meta["description"],
            data=meta["data"],
        ))

    return SharingPreset(
        id="",  # caller assigns
        name=default_name,
        description="",
        built_in=False,
        principles=principles,
        category_assignments=assignments,
        chain=DownscalingChain(layers=layers),
    )
