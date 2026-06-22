# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""FastAPI router for the BOM / Archetype module + DSM × LCA pipeline (Phase 2B)."""
from __future__ import annotations

import datetime
import io
import re
import threading
import uuid

from fastapi import APIRouter, Depends, File, HTTPException, Query, Response, UploadFile

from mapper.api.project_guard import verify_project_state
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel

from mapper.api.dsm import _current_project, _get_system, _proj_results, _proj_states
from mapper.core import dsm_storage
from mapper.core.bom_engine import (
    add_child_in_roots,
    assign_ids_to_roots,
    find_node_in_roots,
    find_root_containing,
    flatten_roots,
    flatten_roots_for_year,
    generate_archetype_timeline,
    has_evolution,
    iter_all_materials,
    material_count_total,
    remove_node_in_roots,
    resolve_quantity,
    stage_to_scope,
    summarize_archetype,
    total_mass_kg,
    unlinked_count_total,
    validate_roots,
    validation_error_count,
)
from mapper.core.bom_validator import BOMValidationRow, issues_by_node_key, validate_bom
from mapper.core.bw2_wrapper import PersistentLCARunner, run_archetype_lca, run_lca_multi_method
from mapper.core.dsm_engine import all_cohort_keys
from mapper.core.material_flow_engine import compute_material_flows
from mapper.core.dsm_lca_engine import DSMLCAPipeline
from mapper.models.bom_schemas import (
    Archetype,
    ArchetypeCreate,
    ArchetypeLCARequest,
    ArchetypeLCAResult,
    ArchetypeSummary,
    ArchetypeTimeline,
    BOMNode,
    BOMNodeCreate,
    BOMNodeUpdate,
    CohortMapping,
    CohortMappingResult,
    EcoinventLink,
    FlattenedBOM,
    FlattenedMaterial,
    MaterialEvolution,
    MaterialFlowMultiRequest,
    MaterialFlowRequest,
    MaterialFlowResult,
    MaterialFlowScenarioRun,
    MultiMaterialFlowResult,
    DSMLCABatchResult,
    DSMLCARequest,
    DSMLCAResult,
    QuantityMilestone,
    ValidationReport,
)


_FILENAME_UNSAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _sanitize_filename(name: str, fallback: str = "archetype", max_len: int = 100) -> str:
    cleaned = _FILENAME_UNSAFE.sub("_", name).strip("_")
    return (cleaned or fallback)[:max_len]


# ── Folder helpers ───────────────────────────────────────────────────────────
# Folders are represented as forward-slash paths. They are implicit from any
# archetype's ``folder`` field plus any empty folders persisted in folders.json.

_FOLDER_SEGMENT = re.compile(r"^[A-Za-z0-9 _-]+$")
MAX_FOLDER_DEPTH = 5


def _normalize_folder(path: str | None) -> str | None:
    """Trim, collapse slashes, validate. Returns None for root, raises on bad input."""
    if path is None:
        return None
    p = path.strip().strip("/")
    if not p:
        return None
    segments = [s for s in p.split("/") if s != ""]
    if len(segments) > MAX_FOLDER_DEPTH:
        raise HTTPException(status_code=400, detail=f"Folder depth exceeds {MAX_FOLDER_DEPTH}")
    for seg in segments:
        if not _FOLDER_SEGMENT.match(seg):
            raise HTTPException(
                status_code=400,
                detail=f"Folder segment '{seg}' contains invalid characters (allowed: letters, digits, space, _, -).",
            )
    return "/".join(segments)


def _all_folder_paths(project: str | None = None) -> list[str]:
    """Union of folders used by archetypes + persisted empty folders."""
    p = project or _current_project()
    folders: set[str] = set(dsm_storage.load_folders(p))
    for arc in _proj_archetypes(p).values():
        if arc.folder:
            folders.add(arc.folder)
            # Include all ancestor paths.
            parts = arc.folder.split("/")
            for i in range(1, len(parts)):
                folders.add("/".join(parts[:i]))
    return sorted(folders)


router = APIRouter(tags=["bom"])


# In-memory stores — nested by bw2 project name (outer key). Archetypes persist
# to disk alongside DSM systems; cohort mappings persist with their system;
# DSM×LCA results are transient.
_archetypes: dict[str, dict[str, Archetype]] = {}
_cohort_mappings: dict[str, dict[str, CohortMapping]] = {}
# Per-system batch of results — one DSMLCAResult per method.
_dsm_lca_results: dict[str, dict[str, list[DSMLCAResult]]] = {}
_lock = threading.Lock()


def _now_iso() -> str:
    return datetime.datetime.now().isoformat()


def _proj_archetypes(project: str | None = None) -> dict[str, Archetype]:
    p = project or _current_project()
    return _archetypes.setdefault(p, {})


def _proj_cohort_mappings(project: str | None = None) -> dict[str, CohortMapping]:
    p = project or _current_project()
    return _cohort_mappings.setdefault(p, {})


def _proj_dsm_lca_results(project: str | None = None) -> dict[str, list[DSMLCAResult]]:
    p = project or _current_project()
    return _dsm_lca_results.setdefault(p, {})


def _get_archetype(arc_id: str) -> Archetype:
    arc = _proj_archetypes().get(arc_id)
    if arc is None:
        raise HTTPException(status_code=404, detail=f"Archetype '{arc_id}' not found")
    return arc


# ── Archetype CRUD ───────────────────────────────────────────────────────────


@router.post(
    "/bom/archetypes",
    response_model=Archetype,
    dependencies=[Depends(verify_project_state)],
)
async def create_archetype(body: ArchetypeCreate) -> Archetype:
    folder = _normalize_folder(body.folder)
    arc = Archetype(
        id=str(uuid.uuid4()),
        name=body.name,
        description=body.description,
        category=body.category,
        folder=folder,
        bom=body.bom,
        created_at=_now_iso(),
        updated_at=_now_iso(),
    )
    assign_ids_to_roots(arc.bom)
    project = _current_project()
    with _lock:
        _proj_archetypes(project)[arc.id] = arc  # type: ignore[index]
    dsm_storage.save_archetype(project, arc)
    return arc


@router.get("/bom/archetypes", response_model=list[ArchetypeSummary])
async def list_archetypes() -> list[ArchetypeSummary]:
    return [ArchetypeSummary(**summarize_archetype(arc)) for arc in _proj_archetypes().values()]


# NOTE: specific /bom/archetypes/<literal> routes MUST be declared before the
# parameterized {arc_id} route, otherwise FastAPI resolves the literal as an
# archetype id and returns 404. Forward references are fine — the helpers
# (_build_multi_export_workbook, _sanitize_filename, etc.) are looked up at
# call time, not import time.
@router.get("/bom/archetypes/export-all")
async def export_all_archetypes(folder: str | None = None) -> Response:
    project = _current_project()
    archetypes = list(_proj_archetypes(project).values())
    if folder:
        norm = _normalize_folder(folder)
        archetypes = [
            a for a in archetypes
            if a.folder and (a.folder == norm or a.folder.startswith(f"{norm}/"))
        ]
    if not archetypes:
        raise HTTPException(status_code=404, detail="No archetypes to export in the selected scope.")
    archetypes.sort(key=lambda a: ((a.folder or ""), a.name))
    wb = _build_multi_export_workbook(archetypes)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    tag = _sanitize_filename(folder.replace("/", "_"), "all") if folder else "all"
    filename = f"archetypes_{tag}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/bom/archetypes/{arc_id}/validation-report", response_model=ValidationReport)
async def get_archetype_validation_report(arc_id: str) -> ValidationReport:
    """Return the persisted validation report for ``arc_id``.

    A 404 means the archetype itself is missing. A 200 with ``total_rows == 0``
    and an empty ``issues`` list means the archetype was imported before
    Patch 2 (no validation) or contains no LCA-linked materials yet.
    """
    arc = _get_archetype(arc_id)
    if arc.validation_report is not None:
        return arc.validation_report
    return ValidationReport(
        total_rows=0, valid_rows=0, error_rows=0, warning_rows=0,
        project_name=_current_project(),
    )


@router.get("/bom/archetypes/{arc_id}", response_model=Archetype)
async def get_archetype(arc_id: str) -> Archetype:
    return _get_archetype(arc_id)


@router.put("/bom/archetypes/{arc_id}", response_model=Archetype)
async def update_archetype(arc_id: str, body: ArchetypeCreate) -> Archetype:
    existing = _get_archetype(arc_id)
    folder = _normalize_folder(body.folder)
    updated = Archetype(
        id=arc_id,
        name=body.name,
        description=body.description,
        category=body.category,
        folder=folder,
        bom=body.bom,
        created_at=existing.created_at,
        updated_at=_now_iso(),
    )
    assign_ids_to_roots(updated.bom)
    project = _current_project()
    with _lock:
        _proj_archetypes(project)[arc_id] = updated
    dsm_storage.save_archetype(project, updated)
    return updated


@router.delete("/bom/archetypes/{arc_id}")
async def delete_archetype(arc_id: str) -> dict[str, bool]:
    project = _current_project()
    archetypes = _proj_archetypes(project)
    if arc_id not in archetypes:
        raise HTTPException(status_code=404, detail="Archetype not found")
    with _lock:
        archetypes.pop(arc_id, None)
        # Clean any cohort mappings (in this project) that pointed at this archetype.
        mappings = _proj_cohort_mappings(project)
        for sys_id, mapping in list(mappings.items()):
            mapping.mappings = [m for m in mapping.mappings if m.archetype_id != arc_id]
            mappings[sys_id] = mapping
            dsm_storage.save_cohort_mappings(project, sys_id, mapping)
    dsm_storage.delete_archetype_file(project, arc_id)
    return {"deleted": True}


# ── Folder endpoints ─────────────────────────────────────────────────────────


class MoveArchetypeBody(BaseModel):
    archetype_id: str
    new_folder: str | None = None


class FolderCreateBody(BaseModel):
    path: str


class FolderRenameBody(BaseModel):
    old_path: str
    new_path: str


class FolderDeleteBody(BaseModel):
    path: str
    delete_archetypes: bool = False


@router.get("/bom/folders")
async def list_folders() -> list[str]:
    return _all_folder_paths()


@router.post("/bom/folders/create")
async def create_folder(body: FolderCreateBody) -> dict:
    norm = _normalize_folder(body.path)
    if norm is None:
        raise HTTPException(status_code=400, detail="Root folder cannot be explicitly created.")
    project = _current_project()
    folders = set(dsm_storage.load_folders(project))
    folders.add(norm)
    # Include all ancestors so the tree visibly expands.
    parts = norm.split("/")
    for i in range(1, len(parts)):
        folders.add("/".join(parts[:i]))
    dsm_storage.save_folders(project, sorted(folders))
    return {"path": norm, "folders": _all_folder_paths(project)}


@router.post("/bom/folders/rename")
async def rename_folder(body: FolderRenameBody) -> dict:
    old = _normalize_folder(body.old_path)
    new = _normalize_folder(body.new_path)
    if old is None or new is None:
        raise HTTPException(status_code=400, detail="Cannot rename the root folder.")
    if old == new:
        return {"renamed": 0, "folders": _all_folder_paths()}
    project = _current_project()
    archetypes = _proj_archetypes(project)
    renamed = 0
    with _lock:
        for arc in archetypes.values():
            if not arc.folder:
                continue
            if arc.folder == old or arc.folder.startswith(old + "/"):
                arc.folder = new + arc.folder[len(old):]
                arc.updated_at = _now_iso()
                dsm_storage.save_archetype(project, arc)
                renamed += 1
        # Update persisted empty-folder list.
        folders = set(dsm_storage.load_folders(project))
        remapped: set[str] = set()
        for f in folders:
            if f == old or f.startswith(old + "/"):
                remapped.add(new + f[len(old):])
            else:
                remapped.add(f)
        dsm_storage.save_folders(project, sorted(remapped))
    return {"renamed": renamed, "folders": _all_folder_paths(project)}


@router.post("/bom/folders/delete")
async def delete_folder(body: FolderDeleteBody) -> dict:
    path = _normalize_folder(body.path)
    if path is None:
        raise HTTPException(status_code=400, detail="Cannot delete the root folder.")
    project = _current_project()
    archetypes = _proj_archetypes(project)
    affected = [arc for arc in archetypes.values()
                if arc.folder == path or (arc.folder and arc.folder.startswith(path + "/"))]
    deleted_arcs = 0
    moved_arcs = 0
    with _lock:
        if body.delete_archetypes:
            # Remove those archetypes entirely.
            mappings = _proj_cohort_mappings(project)
            for arc in affected:
                if arc.id:
                    archetypes.pop(arc.id, None)
                    dsm_storage.delete_archetype_file(project, arc.id)
                    for sys_id, mapping in list(mappings.items()):
                        before = len(mapping.mappings)
                        mapping.mappings = [m for m in mapping.mappings if m.archetype_id != arc.id]
                        if len(mapping.mappings) != before:
                            dsm_storage.save_cohort_mappings(project, sys_id, mapping)
                    deleted_arcs += 1
        else:
            # Move to root.
            for arc in affected:
                arc.folder = None
                arc.updated_at = _now_iso()
                dsm_storage.save_archetype(project, arc)
                moved_arcs += 1
        # Remove the folder and any descendants from the persisted list.
        folders = set(dsm_storage.load_folders(project))
        folders = {f for f in folders if not (f == path or f.startswith(path + "/"))}
        dsm_storage.save_folders(project, sorted(folders))
    return {"deleted_archetypes": deleted_arcs, "moved_archetypes": moved_arcs, "folders": _all_folder_paths(project)}


@router.post("/bom/archetypes/move", response_model=Archetype)
async def move_archetype(body: MoveArchetypeBody) -> Archetype:
    arc = _get_archetype(body.archetype_id)
    new_folder = _normalize_folder(body.new_folder)
    arc.folder = new_folder
    arc.updated_at = _now_iso()
    project = _current_project()
    dsm_storage.save_archetype(project, arc)
    return arc


# ── BOM tree node operations ─────────────────────────────────────────────────


@router.post("/bom/archetypes/{arc_id}/nodes", response_model=Archetype)
async def add_bom_node(arc_id: str, body: BOMNodeCreate) -> Archetype:
    arc = _get_archetype(arc_id)
    ok = add_child_in_roots(arc.bom, body.parent_node_id, body.node)
    if not ok:
        raise HTTPException(
            status_code=404,
            detail=f"Parent node '{body.parent_node_id}' not found in this archetype",
        )
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return arc


@router.put("/bom/archetypes/{arc_id}/nodes/{node_id}", response_model=BOMNode)
async def update_bom_node(arc_id: str, node_id: str, body: BOMNodeUpdate) -> BOMNode:
    arc = _get_archetype(arc_id)
    node = find_node_in_roots(arc.bom, node_id)
    if node is None:
        raise HTTPException(status_code=404, detail=f"Node '{node_id}' not found")
    if body.name is not None:
        node.name = body.name
    if body.quantity is not None:
        node.quantity = body.quantity
    if body.unit is not None:
        node.unit = body.unit
    if "scope" in body.model_fields_set:
        if body.scope is None:
            node.scope = None
        elif body.scope in ("inflows", "stock", "outflows"):
            node.scope = body.scope
            # Auto-derive is_annual from scope unless the same request also
            # sets is_annual explicitly (caller override wins).
            if body.is_annual is None:
                node.is_annual = body.scope == "stock"
        else:
            raise HTTPException(status_code=400, detail=f"Invalid scope: {body.scope!r}")
    if body.is_annual is not None:
        node.is_annual = body.is_annual
    if body.ecoinvent_activity is not None:
        node.ecoinvent_activity = body.ecoinvent_activity
        # Linking an activity makes a node a material.
        if node.children:
            raise HTTPException(
                status_code=400,
                detail="Cannot link an ecoinvent activity to a component with children.",
            )
        node.node_type = "material"
    # Distinguish "field omitted" (leave evolution alone) from "explicitly null"
    # (clear evolution back to fixed). Both `null` and `{"method": "fixed"}`
    # clear the field so the caller can reset a learning_rate/milestones entry.
    if "evolution" in body.model_fields_set:
        if body.evolution is None or body.evolution.method == "fixed":
            node.evolution = None
        else:
            if node.node_type != "material":
                raise HTTPException(
                    status_code=400,
                    detail="Only material leaves can carry an evolution.",
                )
            node.evolution = body.evolution
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return node


@router.delete("/bom/archetypes/{arc_id}/nodes/{node_id}", response_model=Archetype)
async def delete_bom_node(arc_id: str, node_id: str) -> Archetype:
    arc = _get_archetype(arc_id)
    ok = remove_node_in_roots(arc.bom, node_id)
    if not ok:
        raise HTTPException(status_code=404, detail=f"Node '{node_id}' not found")
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return arc


# ── BOM analysis ─────────────────────────────────────────────────────────────


@router.get("/bom/archetypes/{arc_id}/flatten", response_model=FlattenedBOM)
async def flatten_archetype(arc_id: str, year: int | None = None) -> FlattenedBOM:
    arc = _get_archetype(arc_id)
    materials = (
        flatten_roots_for_year(arc.bom, year) if year is not None else flatten_roots(arc.bom)
    )
    return FlattenedBOM(
        archetype_id=arc_id,
        materials=materials,
        total_mass_kg=total_mass_kg(materials),
        unlinked_count=sum(1 for m in materials if m.ecoinvent_activity is None),
    )


# ── Timeline endpoints ───────────────────────────────────────────────────────


def _expand_years(
    years_csv: str | None, year_start: int | None, year_end: int | None, step: int
) -> list[int]:
    if years_csv:
        out: list[int] = []
        for tok in years_csv.split(","):
            tok = tok.strip()
            if not tok:
                continue
            try:
                out.append(int(tok))
            except ValueError:
                raise HTTPException(status_code=400, detail=f"Invalid year '{tok}' in years=")
        if not out:
            raise HTTPException(status_code=400, detail="years= was empty.")
        return sorted(set(out))
    if year_start is None or year_end is None:
        raise HTTPException(
            status_code=400,
            detail="Provide either years=a,b,c or both year_start and year_end.",
        )
    if year_start > year_end:
        raise HTTPException(status_code=400, detail="year_start must be ≤ year_end.")
    if step <= 0:
        raise HTTPException(status_code=400, detail="step must be positive.")
    return list(range(year_start, year_end + 1, step))


@router.get("/bom/archetypes/{arc_id}/timeline", response_model=ArchetypeTimeline)
async def get_archetype_timeline(
    arc_id: str,
    years: str | None = None,
    year_start: int | None = None,
    year_end: int | None = None,
    step: int = 5,
) -> ArchetypeTimeline:
    arc = _get_archetype(arc_id)
    year_list = _expand_years(years, year_start, year_end, step)
    return generate_archetype_timeline(arc, year_list)


class TimelineCompareRow(BaseModel):
    node_id: str
    name: str
    path: list[str]
    unit: str
    year_start: int
    year_end: int
    quantity_start: float
    quantity_end: float
    delta: float
    delta_pct: float | None  # None when quantity_start == 0
    has_evolution: bool


class TimelineCompareResult(BaseModel):
    archetype_id: str
    year_start: int
    year_end: int
    rows: list[TimelineCompareRow]
    total_mass_start: float
    total_mass_end: float


@router.get(
    "/bom/archetypes/{arc_id}/timeline/compare", response_model=TimelineCompareResult
)
async def compare_archetype_timeline(
    arc_id: str, year_start: int, year_end: int
) -> TimelineCompareResult:
    arc = _get_archetype(arc_id)
    if year_start > year_end:
        raise HTTPException(status_code=400, detail="year_start must be ≤ year_end.")
    timeline = generate_archetype_timeline(arc, [year_start, year_end])
    rows: list[TimelineCompareRow] = []
    for r in timeline.rows:
        q0 = r.quantities.get(year_start, 0.0)
        q1 = r.quantities.get(year_end, 0.0)
        delta = q1 - q0
        delta_pct = (delta / q0 * 100.0) if q0 else None
        rows.append(
            TimelineCompareRow(
                node_id=r.node_id,
                name=r.name,
                path=r.path,
                unit=r.unit,
                year_start=year_start,
                year_end=year_end,
                quantity_start=q0,
                quantity_end=q1,
                delta=delta,
                delta_pct=delta_pct,
                has_evolution=r.has_evolution,
            )
        )
    return TimelineCompareResult(
        archetype_id=arc_id,
        year_start=year_start,
        year_end=year_end,
        rows=rows,
        total_mass_start=timeline.total_mass_by_year.get(year_start, 0.0),
        total_mass_end=timeline.total_mass_by_year.get(year_end, 0.0),
    )


class ApplyLearningRateRequest(BaseModel):
    node_ids: list[str] | None = None  # None → all materials
    learning_rate: float | None = None  # None → reset evolution to fixed
    base_year: int = 2025
    reset: bool = False  # Explicit reset flag; also triggered by learning_rate=None.


@router.post("/bom/archetypes/{arc_id}/apply-learning-rate", response_model=Archetype)
async def apply_learning_rate(
    arc_id: str, body: ApplyLearningRateRequest
) -> Archetype:
    arc = _get_archetype(arc_id)
    target: set[str] | None = set(body.node_ids) if body.node_ids else None
    reset = body.reset or body.learning_rate is None
    touched = 0
    for m in iter_all_materials(arc.bom):
        if target is not None and (m.id or "") not in target:
            continue
        if reset:
            m.evolution = None
        else:
            m.evolution = MaterialEvolution(
                method="learning_rate",
                learning_rate=body.learning_rate,
                base_year=body.base_year,
            )
        touched += 1
    if touched == 0:
        raise HTTPException(status_code=400, detail="No materials matched the given node_ids.")
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return arc


class ApplyReboundEffectRequest(BaseModel):
    node_ids: list[str] | None = None  # None → all materials
    rebound_rate: float | None = None  # None → reset evolution to fixed
    base_year: int = 2025
    applies_to_stages: list[str] | None = None
    reset: bool = False


@router.post("/bom/archetypes/{arc_id}/apply-rebound-effect", response_model=Archetype)
async def apply_rebound_effect(
    arc_id: str, body: ApplyReboundEffectRequest
) -> Archetype:
    arc = _get_archetype(arc_id)
    target: set[str] | None = set(body.node_ids) if body.node_ids else None
    reset = body.reset or body.rebound_rate is None
    touched = 0
    for m in iter_all_materials(arc.bom):
        if target is not None and (m.id or "") not in target:
            continue
        if reset:
            m.evolution = None
        else:
            m.evolution = MaterialEvolution(
                method="rebound_effect",
                rebound_rate=body.rebound_rate,
                base_year=body.base_year,
                applies_to_stages=body.applies_to_stages,
            )
        touched += 1
    if touched == 0:
        raise HTTPException(status_code=400, detail="No materials matched the given node_ids.")
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return arc


class ApplyMilestonesRequest(BaseModel):
    node_id: str
    milestones: list[QuantityMilestone]


@router.post("/bom/archetypes/{arc_id}/apply-milestones", response_model=BOMNode)
async def apply_milestones(arc_id: str, body: ApplyMilestonesRequest) -> BOMNode:
    arc = _get_archetype(arc_id)
    node = find_node_in_roots(arc.bom, body.node_id)
    if node is None:
        raise HTTPException(status_code=404, detail=f"Node '{body.node_id}' not found")
    if node.node_type != "material":
        raise HTTPException(status_code=400, detail="Milestones only apply to material leaves.")
    if len(body.milestones) < 2:
        raise HTTPException(status_code=400, detail="Provide at least two milestones.")
    node.evolution = MaterialEvolution(
        method="milestones",
        milestones=sorted(body.milestones, key=lambda m: m.year),
    )
    arc.updated_at = _now_iso()
    dsm_storage.save_archetype(_current_project(), arc)
    return node


@router.post("/bom/archetypes/{arc_id}/lca", response_model=ArchetypeLCAResult)
async def standalone_lca(arc_id: str, body: ArchetypeLCARequest) -> ArchetypeLCAResult:
    arc = _get_archetype(arc_id)
    err_count = validation_error_count(arc.bom)
    if err_count:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "validation_failed",
                "message": (
                    f"Archetype '{arc.name}' has {err_count} row(s) with unresolved "
                    "ecoinvent links. LCA computation is blocked until they are fixed."
                ),
                "archetype_id": arc.id,
                "archetype_name": arc.name,
                "error_rows": err_count,
                "report_url": f"/api/bom/archetypes/{arc.id}/validation-report",
            },
        )
    issues = validate_roots(arc.bom)
    unlinked = [i for i in issues if "no ecoinvent activity" in i]
    if unlinked:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot run LCA: {len(unlinked)} unlinked material(s). First issue: {unlinked[0]}",
        )
    materials: list[FlattenedMaterial] = flatten_roots(arc.bom)
    # demand keyed by (db, code) → amount × external multiplier
    demand: dict[tuple[str, str], float] = {}
    for m in materials:
        if m.ecoinvent_activity is None:
            continue
        key = (m.ecoinvent_activity.database, m.ecoinvent_activity.code)
        demand[key] = demand.get(key, 0.0) + m.quantity * body.amount

    try:
        result = run_archetype_lca(demand, tuple(body.method))
    except Exception as e:  # pragma: no cover — bw2 errors bubble up
        raise HTTPException(status_code=500, detail=f"LCA failed: {e}")

    # Map per-activity scores back to per-material names.
    impact_by_material: dict[str, float] = {}
    by_activity = result["by_activity"]
    for m in materials:
        if m.ecoinvent_activity is None:
            continue
        key = (m.ecoinvent_activity.database, m.ecoinvent_activity.code)
        score = by_activity.get(key, 0.0)
        # Multiple materials can share an activity — split proportionally to qty.
        same_act_qty = sum(
            x.quantity for x in materials
            if x.ecoinvent_activity is not None
            and (x.ecoinvent_activity.database, x.ecoinvent_activity.code) == key
        )
        share = (m.quantity / same_act_qty) if same_act_qty > 0 else 0.0
        impact_by_material[m.name] = impact_by_material.get(m.name, 0.0) + score * share

    return ArchetypeLCAResult(
        archetype_id=arc_id,
        method=body.method,
        score=result["score"],
        unit=result["unit"],
        amount=body.amount,
        impact_by_material=impact_by_material,
    )


# ── Cohort mapping ───────────────────────────────────────────────────────────


@router.post(
    "/dsm/systems/{system_id}/cohort-mappings", response_model=CohortMappingResult
)
async def set_cohort_mappings(system_id: str, body: CohortMapping) -> CohortMappingResult:
    sys_def = _get_system(system_id)
    project = _current_project()
    archetypes = _proj_archetypes(project)
    valid_cohorts = set(all_cohort_keys(sys_def.dimensions))

    invalid_cohorts: list[str] = []
    invalid_archetypes: list[str] = []
    seen_cohorts: set[str] = set()
    cleaned: list = []

    for entry in body.mappings:
        if entry.cohort_key not in valid_cohorts:
            invalid_cohorts.append(entry.cohort_key)
            continue
        if entry.archetype_id not in archetypes:
            invalid_archetypes.append(entry.archetype_id)
            continue
        if entry.cohort_key in seen_cohorts:
            continue
        # Clamp scaling factor to a sane positive value.
        if entry.scaling_factor <= 0:
            entry.scaling_factor = 1.0
        seen_cohorts.add(entry.cohort_key)
        cleaned.append(entry)

    body.mappings = cleaned
    body.mfa_system_id = system_id
    _proj_cohort_mappings(project)[system_id] = body
    dsm_storage.save_cohort_mappings(project, system_id, body)

    return CohortMappingResult(
        mapped_cohorts=len(cleaned),
        unmapped_cohorts=sorted(valid_cohorts - seen_cohorts),
        invalid_cohorts=invalid_cohorts,
        invalid_archetypes=invalid_archetypes,
    )


@router.get(
    "/dsm/systems/{system_id}/cohort-mappings", response_model=CohortMapping
)
async def get_cohort_mappings(system_id: str) -> CohortMapping:
    _get_system(system_id)
    return _proj_cohort_mappings().get(
        system_id, CohortMapping(mfa_system_id=system_id, mappings=[])
    )


_COHORT_MAPPING_COLUMNS = ["fuel_type", "size", "archetype", "scaling_factor"]


def _cohort_template_workbook(
    sys_def,
    existing: CohortMapping | None = None,
    archetypes_by_id: dict[str, str] | None = None,
) -> Workbook:
    """Build a cohort-mappings xlsx template listing every valid cohort combination
    for the system. When ``existing`` is provided, pre-fills the archetype +
    scaling_factor + color cells from current mappings — supports round-trip
    export.

    Column order: nad dimensions, archetype, scaling_factor, color (Patch 4AK).
    The Color column stores per-row color overrides as ``#RRGGBB`` hex; empty
    cells mean "no override → algorithm default."
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Mappings"
    from openpyxl.styles import Alignment, Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    nads = [d for d in sys_def.dimensions if not d.is_age]
    nad_names = [d.name for d in nads]
    # Patch 4AK — Color column appended at the end.
    headers = nad_names + ["archetype", "scaling_factor", "color"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

    # Build lookups from `existing` for round-trip fill.
    entries_by_ck: dict[str, CohortMappingEntry] = {}
    row_colors: dict[str, str] = {}
    if existing is not None:
        for entry in existing.mappings:
            entries_by_ck[entry.cohort_key] = entry
        row_colors = dict(existing.row_colors or {})
    archetypes_by_id = archetypes_by_id or {}

    for ck in all_cohort_keys(sys_def.dimensions):
        parts = ck.split("|") if ck else []
        entry = entries_by_ck.get(ck)
        arc_name = ""
        sf_val: str | float = ""
        if entry is not None:
            arc_name = archetypes_by_id.get(entry.archetype_id, entry.archetype_id)
            sf_val = entry.scaling_factor
        color = row_colors.get(ck, "")
        row = list(parts) + [arc_name, sf_val, color]
        ws.append(row)

    for col_letter, width in zip(
        "ABCDEFGHIJK",
        [16, 16, 16, 16, 28, 14, 12],
    ):
        ws.column_dimensions[col_letter].width = width
    return wb


@router.get("/dsm/systems/{system_id}/cohort-mappings/template")
async def get_cohort_mappings_template(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    # Patch 4AK — round-trip export: include current mappings + row colors
    # if any. Existing single-system callers (the LCA Architect "blank
    # template" path) still work because the file structure is identical;
    # the cells are just filled in.
    project = _current_project()
    existing = _proj_cohort_mappings(project).get(system_id)
    archetypes = _proj_archetypes(project)
    archetypes_by_id = {arc_id: arc.name for arc_id, arc in archetypes.items()}
    wb = _cohort_template_workbook(sys_def, existing, archetypes_by_id)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{_sanitize_filename(sys_def.name, 'system')}_cohort_mappings_template.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _parse_cohort_upload(data: bytes, filename: str, nad_names: list[str]) -> list[dict]:
    """Parse xlsx or csv file into a list of {dim1: ..., dim2: ..., archetype, scaling_factor} dicts.

    Accepts any column order; requires archetype + scaling_factor + at least one
    dimension column (matched case-insensitively against non-age dimension names).
    """
    lower = filename.lower()
    rows: list[list] = []
    if lower.endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        for r in ws.iter_rows(values_only=True):
            rows.append(list(r))
    else:
        import csv as _csv
        text = data.decode("utf-8-sig", errors="replace").splitlines()
        for r in _csv.reader(text):
            rows.append(list(r))

    if not rows:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    header = [str(c or "").strip().lower() for c in rows[0]]
    if "archetype" not in header:
        raise HTTPException(status_code=400, detail="Missing required column 'archetype'.")
    if "scaling_factor" not in header:
        raise HTTPException(status_code=400, detail="Missing required column 'scaling_factor'.")

    # Map each non-age dimension name → column index (case-insensitive).
    lower_nads = {n.lower(): n for n in nad_names}
    dim_cols: dict[str, int] = {}
    for idx, h in enumerate(header):
        if h in lower_nads:
            dim_cols[lower_nads[h]] = idx
    if len(dim_cols) == 0:
        raise HTTPException(
            status_code=400,
            detail=f"No dimension columns recognized. Expected one of: {', '.join(nad_names)}.",
        )
    arc_idx = header.index("archetype")
    scale_idx = header.index("scaling_factor")
    # Patch 4AK — optional Color column. Absent in pre-4AK templates,
    # which still parse cleanly (color defaults to '').
    color_idx = header.index("color") if "color" in header else -1

    out: list[dict] = []
    for raw in rows[1:]:
        if not raw or all((c is None or str(c).strip() == "") for c in raw):
            continue
        # Pad row if shorter than header.
        while len(raw) < len(header):
            raw.append(None)
        entry: dict = {}
        for name, idx in dim_cols.items():
            val = raw[idx]
            entry[name] = str(val).strip() if val is not None else ""
        arc = raw[arc_idx]
        entry["archetype"] = str(arc).strip() if arc is not None else ""
        sf = raw[scale_idx]
        try:
            entry["scaling_factor"] = float(sf) if sf not in (None, "") else 1.0
        except (TypeError, ValueError):
            entry["scaling_factor"] = 1.0
        if color_idx >= 0:
            cval = raw[color_idx]
            entry["color"] = str(cval).strip() if cval is not None else ""
        else:
            entry["color"] = ""
        out.append(entry)
    return out


# Patch 4AK — strict #RRGGBB hex matcher for the Color column.
_HEX_RE = re.compile(r"^#[0-9a-fA-F]{6}$")


def _normalize_color(raw: str) -> tuple[str | None, bool]:
    """Returns (hex_or_None, is_error). Empty / 'auto' → (None, False).
    Valid hex → (normalized lowercase hex, False). Otherwise → (None, True).
    """
    s = (raw or "").strip()
    if s == "" or s.lower() == "auto":
        return None, False
    if _HEX_RE.match(s):
        return s.lower(), False
    return None, True


@router.post(
    "/dsm/systems/{system_id}/cohort-mappings/upload",
    response_model=CohortMappingResult,
)
async def upload_cohort_mappings(
    system_id: str, file: UploadFile = File(...)
) -> CohortMappingResult:
    """Upload cohort mappings from xlsx/csv. Dimension columns (e.g. fuel_type,
    size) plus ``archetype`` and ``scaling_factor`` are required. Empty archetype
    cells mark the cohort as unmapped. Replaces all existing mappings.
    """
    sys_def = _get_system(system_id)
    project = _current_project()
    archetypes = _proj_archetypes(project)
    nads = [d for d in sys_def.dimensions if not d.is_age]
    nad_names = [d.name for d in nads]

    data = await file.read()
    parsed = _parse_cohort_upload(data, file.filename or "upload.xlsx", nad_names)

    # Build name→id map for archetype lookup (case-insensitive).
    name_to_id: dict[str, str] = {}
    for arc_id, arc in archetypes.items():
        name_to_id[arc.name.lower()] = arc_id

    entries = []
    invalid_cohorts: list[str] = []
    invalid_archetypes: list[str] = []
    # Patch 4AK — Color column tracking.
    row_colors: dict[str, str] = {}
    invalid_row_colors: list[str] = []
    seen: set[str] = set()
    valid_keys = set(all_cohort_keys(sys_def.dimensions))

    for row in parsed:
        arc_label = (row.get("archetype") or "").strip()
        if not arc_label:
            continue  # Unmapped row — skip silently.
        arc_id = name_to_id.get(arc_label.lower())
        if arc_id is None and arc_label in archetypes:
            arc_id = arc_label  # Accept raw id too.
        if arc_id is None:
            invalid_archetypes.append(arc_label)
            continue
        ck = "|".join(row.get(n, "") for n in nad_names)
        if ck not in valid_keys:
            invalid_cohorts.append(ck)
            continue
        if ck in seen:
            continue
        sf = row.get("scaling_factor", 1.0)
        if not isinstance(sf, (int, float)) or sf <= 0:
            sf = 1.0
        seen.add(ck)
        entries.append({"cohort_key": ck, "archetype_id": arc_id, "scaling_factor": float(sf)})
        # Patch 4AK — parse Color column. Errors recorded but don't abort
        # the row's other data (archetype + scale still saved).
        color_raw = row.get("color", "") or ""
        normalized, is_error = _normalize_color(color_raw)
        if is_error:
            invalid_row_colors.append(f"{ck}: {color_raw}")
        elif normalized is not None:
            row_colors[ck] = normalized

    from mapper.models.bom_schemas import CohortMappingEntry
    mapping = CohortMapping(
        mfa_system_id=system_id,
        mappings=[CohortMappingEntry(**e) for e in entries],
        row_colors=row_colors,
    )
    _proj_cohort_mappings(project)[system_id] = mapping
    dsm_storage.save_cohort_mappings(project, system_id, mapping)

    return CohortMappingResult(
        mapped_cohorts=len(entries),
        unmapped_cohorts=sorted(valid_keys - seen),
        invalid_cohorts=sorted(set(invalid_cohorts)),
        invalid_archetypes=sorted(set(invalid_archetypes)),
        invalid_row_colors=sorted(set(invalid_row_colors)),
    )


# ── Combined DSM × LCA ───────────────────────────────────────────────────────


@router.post("/dsm/systems/{system_id}/dsm-lca", response_model=DSMLCABatchResult)
async def run_dsm_lca(system_id: str, body: DSMLCARequest) -> DSMLCABatchResult:
    from mapper.core.compute_metrics import measure_compute
    meter = measure_compute()
    _get_system(system_id)
    project = _current_project()
    sim = _proj_results(project).get(system_id)
    if sim is None:
        raise HTTPException(
            status_code=400,
            detail="No simulation results yet. Run /dsm/systems/{id}/simulate first.",
        )
    mapping = _proj_cohort_mappings(project).get(system_id)
    if mapping is None or not mapping.mappings:
        raise HTTPException(
            status_code=400,
            detail="No cohort mappings set. POST /dsm/systems/{id}/cohort-mappings first.",
        )

    # Normalize methods: prefer ``methods`` (list); fall back to legacy ``method``.
    method_lists: list[list[str]]
    if body.methods:
        method_lists = [list(m) for m in body.methods if m]
    elif body.method:
        method_lists = [list(body.method)]
    else:
        raise HTTPException(status_code=400, detail="At least one method is required.")
    if not method_lists:
        raise HTTPException(status_code=400, detail="At least one method is required.")
    method_tuples = [tuple(m) for m in method_lists]

    if body.year_start is not None and body.year_end is not None and body.year_start > body.year_end:
        raise HTTPException(status_code=400, detail="year_start must be ≤ year_end.")

    archetypes = _proj_archetypes(project)
    # Validate every mapped archetype still exists and has all materials linked.
    cohort_to_archetype: dict[str, tuple[str, float]] = {}
    for entry in mapping.mappings:
        arc = archetypes.get(entry.archetype_id)
        if arc is None:
            raise HTTPException(
                status_code=400,
                detail=f"Archetype '{entry.archetype_id}' (mapped to {entry.cohort_key}) was deleted.",
            )
        unlinked = sum(1 for m in iter_all_materials(arc.bom) if m.ecoinvent_activity is None)
        if unlinked:
            raise HTTPException(
                status_code=400,
                detail=f"Archetype '{arc.name}' has {unlinked} unlinked material(s).",
            )
        cohort_to_archetype[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    # Resolve parameter expressions against the selected set (if any).
    from mapper.api import parameters as _parameters
    from mapper.core.parameter_engine import ParameterEngine, ParameterError
    engine: ParameterEngine | None = None
    if body.parameter_set_id:
        pset = _parameters.get_parameter_set(body.parameter_set_id, project)
        if pset is None:
            raise HTTPException(
                status_code=400,
                detail=f"Parameter set '{body.parameter_set_id}' not found",
            )
        engine = ParameterEngine(pset.parameters)

    # Discover dependent subsystems with user-defined cohort mappings so their
    # BOM impacts can be aggregated into the total (matches /impact/calculate).
    from mapper.api import subsystems as _subs
    from mapper.core.dsm_lca_engine import (
        aggregate_subsystem_results,
        build_subsystem_cohort_mapping,
    )
    from mapper.core.subsystem_engine import compute_dependent_subsystem

    dep_subs = _subs.get_subsystems_for_system(system_id, project)
    sub_sim_results: dict[str, object] = {}
    sub_cohort_mappings: dict[str, dict[str, tuple[str, float]]] = {}
    setup_warnings: list[str] = []
    for sub_id, sub in dep_subs.items():
        if not sub.dependency_rules:
            continue
        sub_mapping, unmapped = build_subsystem_cohort_mapping(sub)
        if unmapped:
            setup_warnings.append(
                f"Subsystem '{sub.name}': {len(unmapped)} unmapped archetype"
                f"{'s' if len(unmapped) != 1 else ''} excluded from calculation: "
                f"{', '.join(unmapped)}"
            )
        if not sub_mapping:
            continue
        for aid in {bom_id for bom_id, _ in sub_mapping.values()}:
            arc = archetypes.get(aid)
            if arc is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Dependent subsystem '{sub.name}' is mapped to archetype "
                        f"'{aid}' which does not exist in the BOM library."
                    ),
                )
            unlinked = sum(1 for m in iter_all_materials(arc.bom) if m.ecoinvent_activity is None)
            if unlinked:
                raise HTTPException(
                    status_code=400,
                    detail=f"Archetype '{arc.name}' has {unlinked} unlinked material(s).",
                )
        try:
            sub_sim = compute_dependent_subsystem(sub, _get_system(system_id), sim, engine)
        except (ParameterError, ValueError) as e:
            raise HTTPException(
                status_code=400, detail=f"Dependent subsystem '{sub.name}': {e}"
            )
        sub_sim_results[sub_id] = sub_sim
        sub_cohort_mappings[sub_id] = sub_mapping

    try:
        persistent = PersistentLCARunner()

        def _pipeline(sim_result, cohort_map):
            return DSMLCAPipeline(
                simulation_result=sim_result,
                archetypes=archetypes,
                cohort_mappings=cohort_map,
                methods=method_tuples,
                lca_runner=persistent,
                year_start=body.year_start,
                year_end=body.year_end,
                parameter_engine=engine,
            )

        primary_results = _pipeline(sim, cohort_to_archetype).calculate(body.scope)
        if sub_sim_results:
            results_by_subsystem: dict[str, list[DSMLCAResult]] = {
                system_id: primary_results,
            }
            for sub_id, sub_sim in sub_sim_results.items():
                results_by_subsystem[sub_id] = _pipeline(
                    sub_sim, sub_cohort_mappings[sub_id]
                ).calculate(body.scope)
            results = aggregate_subsystem_results(results_by_subsystem)
            for r in results:
                r.mfa_system_id = system_id
        else:
            results = primary_results
    except ParameterError as e:
        raise HTTPException(status_code=400, detail=f"Parameter resolution failed: {e}")
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:  # pragma: no cover
        raise HTTPException(status_code=500, detail=f"DSM×LCA failed: {e}")

    _proj_dsm_lca_results(project)[system_id] = results
    return DSMLCABatchResult(
        results=results,
        methods_calculated=len(results),
        year_start=body.year_start,
        year_end=body.year_end,
        warnings=setup_warnings,
        compute_metrics=meter.build(),
    )


@router.get("/dsm/systems/{system_id}/dsm-lca", response_model=DSMLCABatchResult)
async def get_dsm_lca(system_id: str) -> DSMLCABatchResult:
    _get_system(system_id)
    res = _proj_dsm_lca_results().get(system_id)
    if res is None:
        raise HTTPException(status_code=404, detail="No DSM × LCA results yet.")
    return DSMLCABatchResult(results=res, methods_calculated=len(res))


# ── DSM × LCA Excel export ───────────────────────────────────────────────────


def _short_method_label(method: list[str]) -> str:
    return method[-1] if method else "method"


def _build_mfa_lca_workbook(
    system_name: str,
    results: list[DSMLCAResult],
    scope: str,
    selected_year: int | None,
    cohort_mapping: CohortMapping | None,
    archetypes: dict[str, Archetype],
    sim_counts: dict[int, dict[str, float]] | None,
    dims: list | None,
    *,
    elapsed_seconds: float | None = None,
    sim_result=None,
) -> Workbook:
    """Build a comprehensive XLSX workbook with 9 sheets for Impact Assessment
    results. Designed for easy analysis in Excel (pivot tables, filtering).
    """
    import datetime

    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    # ── Styles ─────────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    meta_font = Font(bold=True, color="374151")
    meta_val_font = Font(color="374151")

    def _num_fmt(v: float) -> str:
        """Pick number format for a value."""
        if v == 0:
            return "0.00"
        a = abs(v)
        if a < 0.01 or a > 1e6:
            return "0.00E+00"
        return "#,##0.00"

    SCI_FMT = "0.00E+00"
    INT_FMT = "#,##0"
    PCT_FMT = "0.00"

    mapping_by_cohort: dict[str, tuple[str, float]] = {}
    if cohort_mapping is not None:
        for entry in cohort_mapping.mappings:
            mapping_by_cohort[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    nonage_dim_names: list[str] = []
    if dims is not None:
        for d in dims:
            name = getattr(d, "name", None) or (d.get("name") if isinstance(d, dict) else None)
            is_age = getattr(d, "is_age", None)
            if is_age is None and isinstance(d, dict):
                is_age = d.get("is_age", False)
            if name and not is_age:
                nonage_dim_names.append(name)
    dim_headers = [d.capitalize() for d in nonage_dim_names] or ["Cohort"]
    n_dims = len(dim_headers)

    def _split_cohort(ck: str) -> list[str]:
        parts = ck.split("|")
        out = parts[:n_dims] if nonage_dim_names else [ck]
        while len(out) < n_dims:
            out.append("")
        return out

    def _style_header(ws, row_num: int = 1) -> None:
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autosize(ws, max_width: int = 40, sample_rows: int = 50) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for i, cell in enumerate(col_cells):
                if i >= sample_rows + 1:
                    break
                v = cell.value
                if v is not None:
                    widest = max(widest, min(max_width, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    def _apply_sci(ws, min_row, min_col, max_col) -> None:
        for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.number_format = SCI_FMT

    # Collect common data
    labels = [_short_method_label(r.method) for r in results]
    units = [r.unit for r in results]
    years_set: set[int] = set()
    for r in results:
        for yr in r.years:
            years_set.add(yr.year)
    years_list = sorted(years_set)
    year_start_val = years_list[0] if years_list else None
    year_end_val = years_list[-1] if years_list else None
    stages_included = results[0].stages_included if results else []
    cohort_keys = sorted({
        ck for r in results for yr in r.years for ck in yr.impact_by_cohort
    })

    # Precompute archetype name per cohort
    cohort_arc_name: dict[str, str] = {}
    cohort_arc_scale: dict[str, float] = {}
    for ck in cohort_keys:
        arc_id, scale = mapping_by_cohort.get(ck, ("", 1.0))
        cohort_arc_name[ck] = archetypes[arc_id].name if arc_id in archetypes else ""
        cohort_arc_scale[ck] = scale

    # Unique archetype names used
    arc_names_set = sorted({n for n in cohort_arc_name.values() if n})

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1: Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3ECFCF"

    scope_labels = {"inflows": "Manufacturing", "stock": "Operation", "outflows": "End of Life", "all": "Full lifecycle"}
    meta_rows = [
        ("Project", system_name),
        ("Scope", scope_labels.get(scope, scope)),
        ("Stages included", ", ".join(stages_included) if stages_included else "—"),
        ("Year range", f"{year_start_val}–{year_end_val}" if year_start_val else "—"),
        ("Indicators calculated", len(results)),
        ("Cohorts", len(cohort_keys)),
        ("Archetypes", len(arc_names_set)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if elapsed_seconds is not None:
        m, s = divmod(int(elapsed_seconds), 60)
        meta_rows.append(("Calculation time", f"{m}m {s}s" if m else f"{s}s"))
    meta_rows.append(("MApper version", "1.0"))

    for label, value in meta_rows:
        ws.append([label, value])
        ws[ws.max_row][0].font = meta_font
        ws[ws.max_row][1].font = meta_val_font

    ws.append([])
    ws.append(["Indicator", "Method path", "Unit", "Cumulative impact", "Peak year", "Peak impact"])
    _style_header(ws, ws.max_row)
    data_start = ws.max_row + 1
    for res in results:
        ws.append([
            _short_method_label(res.method),
            " › ".join(res.method),
            res.unit,
            res.summary.total_impact,
            res.summary.peak_year,
            res.summary.peak_impact,
        ])
    _apply_sci(ws, data_start, 4, 4)
    _apply_sci(ws, data_start, 6, 6)

    _autosize(ws)
    ws.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2: Annual totals
    # ══════════════════════════════════════════════════════════════════════════
    ws_at = wb.create_sheet("Annual totals")
    ws_at.sheet_properties.tabColor = "4A90D9"
    header = ["Year"] + [f"{l} ({u})" for l, u in zip(labels, units)]
    ws_at.append(header)
    _style_header(ws_at)
    for y in years_list:
        row: list = [y]
        for r in results:
            yr = next((v for v in r.years if v.year == y), None)
            row.append(yr.total_impact if yr else 0.0)
        ws_at.append(row)
    _apply_sci(ws_at, 2, 2, len(header))
    ws_at.freeze_panes = "B2"
    _autosize(ws_at)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3: By indicator (annual, cumulative, YoY %)
    # ══════════════════════════════════════════════════════════════════════════
    ws_bi = wb.create_sheet("By indicator")
    ws_bi.sheet_properties.tabColor = "4A90D9"
    # Build header: Year | for each indicator: Annual | Cumulative | YoY %
    bi_header: list[str] = ["Year"]
    for l, u in zip(labels, units):
        bi_header.append(f"{l} ({u})")
        bi_header.append(f"{l} cumulative")
        bi_header.append(f"{l} YoY %")
    ws_bi.append(bi_header)
    _style_header(ws_bi)

    # Pre-compute year→total for each result
    year_totals_by_result: list[dict[int, float]] = []
    for r in results:
        d: dict[int, float] = {}
        for yr in r.years:
            d[yr.year] = yr.total_impact
        year_totals_by_result.append(d)

    for y in years_list:
        row: list = [y]
        for ri, r in enumerate(results):
            annual = year_totals_by_result[ri].get(y, 0.0)
            cumul = sum(year_totals_by_result[ri].get(yy, 0.0) for yy in years_list if yy <= y)
            prev_y = y - 1
            prev_val = year_totals_by_result[ri].get(prev_y, 0.0)
            yoy = ((annual - prev_val) / abs(prev_val) * 100.0) if prev_val else None
            row.extend([annual, cumul, yoy if yoy is not None else ""])
        ws_bi.append(row)

    for col_idx in range(2, len(bi_header) + 1):
        # Annual and cumulative columns get sci format, YoY gets %
        col_offset = (col_idx - 2) % 3  # 0=annual, 1=cumul, 2=yoy
        fmt = SCI_FMT if col_offset < 2 else PCT_FMT
        for row in ws_bi.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    ws_bi.freeze_panes = "B2"
    _autosize(ws_bi)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4: By fuel type (first non-age dimension)
    # ══════════════════════════════════════════════════════════════════════════
    if results and n_dims >= 1:
        ws_ft = wb.create_sheet("By fuel type")
        ws_ft.sheet_properties.tabColor = "4A90D9"
        primary_dim = dim_headers[0]
        ft_header = ["Year", primary_dim, "Vehicle count"] + [f"{l} ({u})" for l, u in zip(labels, units)]
        ws_ft.append(ft_header)
        _style_header(ws_ft)

        # Collect unique primary-dim values
        primary_vals: set[str] = set()
        for ck in cohort_keys:
            primary_vals.add(_split_cohort(ck)[0])
        primary_sorted = sorted(primary_vals)

        for y in years_list:
            ft_rows: list[tuple] = []
            for pv in primary_sorted:
                # Aggregate across cohorts that share this primary dim
                total_count = 0.0
                impacts: list[float] = [0.0] * len(results)
                for ck in cohort_keys:
                    if _split_cohort(ck)[0] != pv:
                        continue
                    sc = (sim_counts or {}).get(y, {})
                    total_count += sc.get(ck, 0.0)
                    for ri, r in enumerate(results):
                        yr = next((v for v in r.years if v.year == y), None)
                        if yr:
                            impacts[ri] += yr.impact_by_cohort.get(ck, 0.0)
                ft_rows.append((y, pv, total_count, *impacts))
            # Sort by total impact of first indicator desc within this year
            ft_rows.sort(key=lambda r: abs(r[3]) if len(r) > 3 else 0, reverse=True)
            for r in ft_rows:
                ws_ft.append(list(r))

        for row in ws_ft.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = INT_FMT
        _apply_sci(ws_ft, 2, 4, len(ft_header))
        ws_ft.freeze_panes = "C2"
        _autosize(ws_ft)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 5: By cohort (full granularity)
    # ══════════════════════════════════════════════════════════════════════════
    if results:
        ws_coh = wb.create_sheet("By cohort")
        ws_coh.sheet_properties.tabColor = "4A90D9"
        # Header: Year | dims... | Archetype | Scale | Count | per-vehicle + total per indicator
        coh_header = ["Year"] + dim_headers + ["Archetype", "Scale", "Vehicle count"]
        for l, u in zip(labels, units):
            coh_header.append(f"{l} per vehicle ({u})")
            coh_header.append(f"{l} total ({u})")
        ws_coh.append(coh_header)
        _style_header(ws_coh)

        for y in years_list:
            sc = (sim_counts or {}).get(y, {})
            for ck in cohort_keys:
                dim_vals = _split_cohort(ck)
                count = sc.get(ck, 0.0)
                row: list = [y] + dim_vals + [
                    cohort_arc_name.get(ck, ""),
                    cohort_arc_scale.get(ck, 1.0),
                    count,
                ]
                for ri, r in enumerate(results):
                    yr = next((v for v in r.years if v.year == y), None)
                    impact = yr.impact_by_cohort.get(ck, 0.0) if yr else 0.0
                    per_v = (impact / count) if count else 0.0
                    row.extend([per_v, impact])
                ws_coh.append(row)

        count_col = 1 + n_dims + 3  # Vehicle count column
        for row in ws_coh.iter_rows(min_row=2, min_col=count_col, max_col=count_col):
            for cell in row:
                cell.number_format = INT_FMT
        _apply_sci(ws_coh, 2, count_col + 1, len(coh_header))
        ws_coh.freeze_panes = f"{get_column_letter(n_dims + 2)}2"
        _autosize(ws_coh)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 6: By archetype
    # ══════════════════════════════════════════════════════════════════════════
    if results and arc_names_set:
        ws_arc = wb.create_sheet("By archetype")
        ws_arc.sheet_properties.tabColor = "4A90D9"
        arc_header = ["Year", "Archetype", "Vehicle count"] + [f"{l} ({u})" for l, u in zip(labels, units)]
        ws_arc.append(arc_header)
        _style_header(ws_arc)

        for y in years_list:
            sc = (sim_counts or {}).get(y, {})
            for an in arc_names_set:
                total_count = 0.0
                impacts: list[float] = [0.0] * len(results)
                for ck in cohort_keys:
                    if cohort_arc_name.get(ck) != an:
                        continue
                    total_count += sc.get(ck, 0.0)
                    for ri, r in enumerate(results):
                        yr = next((v for v in r.years if v.year == y), None)
                        if yr:
                            impacts[ri] += yr.impact_by_cohort.get(ck, 0.0)
                ws_arc.append([y, an, total_count, *impacts])

        for row in ws_arc.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = INT_FMT
        _apply_sci(ws_arc, 2, 4, len(arc_header))
        ws_arc.freeze_panes = "C2"
        _autosize(ws_arc)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 7: By stage
    # ══════════════════════════════════════════════════════════════════════════
    if results and stages_included:
        ws_stg = wb.create_sheet("By stage")
        ws_stg.sheet_properties.tabColor = "4A90D9"
        if scope == "all" and len(stages_included) > 1:
            # For full lifecycle, show per-stage column. We don't have per-stage
            # breakdown in the results, so show the included stages as info.
            stg_header = ["Year", "Stage"] + [f"{l} ({u})" for l, u in zip(labels, units)]
            ws_stg.append(stg_header)
            _style_header(ws_stg)
            ws_stg.append([])
            ws_stg.append(["Per-stage breakdown requires running each scope separately."])
            ws_stg.append(["Stages included in this calculation: " + ", ".join(stages_included)])
            ws_stg.append([])
            ws_stg.append(["Showing fleet-level totals (all stages combined):"])
            ws_stg.append([])
            # Fall back to total per year
            for y in years_list:
                row: list = [y, ", ".join(stages_included)]
                for r in results:
                    yr = next((v for v in r.years if v.year == y), None)
                    row.append(yr.total_impact if yr else 0.0)
                ws_stg.append(row)
        else:
            # Single scope — one stage grouping
            stage_name = stages_included[0] if stages_included else scope_labels.get(scope, scope)
            stg_header = ["Year", "Stage"] + [f"{l} ({u})" for l, u in zip(labels, units)]
            ws_stg.append(stg_header)
            _style_header(ws_stg)
            for y in years_list:
                row = [y, stage_name]
                for r in results:
                    yr = next((v for v in r.years if v.year == y), None)
                    row.append(yr.total_impact if yr else 0.0)
                ws_stg.append(row)

        _apply_sci(ws_stg, 2, 3, len(stg_header))
        ws_stg.freeze_panes = "C2"
        _autosize(ws_stg)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 8: Cohort mappings
    # ══════════════════════════════════════════════════════════════════════════
    if cohort_mapping is not None and cohort_mapping.mappings:
        ws_cm = wb.create_sheet("Cohort mappings")
        ws_cm.sheet_properties.tabColor = "4A90D9"
        cm_header = dim_headers + ["Archetype", "Scaling factor"]
        ws_cm.append(cm_header)
        _style_header(ws_cm)
        for entry in cohort_mapping.mappings:
            dim_vals = _split_cohort(entry.cohort_key)
            arc_name = archetypes[entry.archetype_id].name if entry.archetype_id in archetypes else entry.archetype_id
            ws_cm.append(dim_vals + [arc_name, entry.scaling_factor])
        ws_cm.freeze_panes = "A2"
        _autosize(ws_cm)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 9: DSM fleet data
    # ══════════════════════════════════════════════════════════════════════════
    if sim_result is not None:
        ws_dsm = wb.create_sheet("DSM fleet data")
        ws_dsm.sheet_properties.tabColor = "4A90D9"

        # Collect all primary-dim values from stock for per-type columns
        primary_dim = dim_headers[0] if dim_headers else "Cohort"
        all_primaries: set[str] = set()
        for yr in sim_result.years:
            for ck in yr.stock:
                all_primaries.add(_split_cohort(ck)[0])
        primary_sorted_dsm = sorted(all_primaries)

        mfa_header = ["Year", "Total stock", "Total inflows", "Total outflows"]
        for pv in primary_sorted_dsm:
            mfa_header.append(f"Stock: {pv}")
        ws_dsm.append(mfa_header)
        _style_header(ws_dsm)

        for yr in sim_result.years:
            total_stock = sum(yr.stock.values())
            total_in = sum(yr.inflow.values())
            total_out = sum(yr.outflow.values())
            row: list = [yr.year, total_stock, total_in, total_out]
            for pv in primary_sorted_dsm:
                s = sum(v for ck, v in yr.stock.items() if _split_cohort(ck)[0] == pv)
                row.append(s)
            ws_dsm.append(row)

        for row in ws_dsm.iter_rows(min_row=2, min_col=2, max_col=len(mfa_header)):
            for cell in row:
                cell.number_format = INT_FMT
        ws_dsm.freeze_panes = "B2"
        _autosize(ws_dsm)

    return wb


@router.post("/dsm/systems/{system_id}/dsm-lca/export")
async def export_dsm_lca(system_id: str, year: int | None = None) -> Response:
    sys_def = _get_system(system_id)
    project = _current_project()
    results = _proj_dsm_lca_results(project).get(system_id)
    if not results:
        raise HTTPException(status_code=400, detail="No DSM × LCA results to export — run a calculation first.")

    mapping = _proj_cohort_mappings(project).get(system_id)
    sim = _proj_results(project).get(system_id)
    sim_counts: dict[int, dict[str, float]] = {}
    if sim is not None:
        scope = results[0].scope
        for yr in sim.years:
            if scope == "inflows":
                sim_counts[yr.year] = dict(yr.inflow)
            elif scope == "outflows":
                sim_counts[yr.year] = dict(yr.outflow)
            else:
                sim_counts[yr.year] = dict(yr.stock)

    wb = _build_mfa_lca_workbook(
        system_name=sys_def.name,
        results=results,
        scope=results[0].scope,
        selected_year=year,
        cohort_mapping=mapping,
        archetypes=_proj_archetypes(project),
        sim_counts=sim_counts,
        dims=list(sys_def.dimensions),
        sim_result=sim,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    import datetime as _dt
    scope = results[0].scope
    scope_tags = {"inflows": "Manufacturing", "stock": "Operation", "outflows": "End_of_Life", "all": "Full_lifecycle"}
    filename = f"MApper_Impact_{_sanitize_filename(sys_def.name, 'system')}_{scope_tags.get(scope, scope)}_{_dt.date.today().isoformat()}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Material Flows ──────────────────────────────────────────────────────────


@router.post("/dsm/systems/{system_id}/material-flows", response_model=MaterialFlowResult)
async def material_flows(system_id: str, body: MaterialFlowRequest) -> MaterialFlowResult:
    from mapper.core.compute_metrics import measure_compute
    meter = measure_compute()
    _get_system(system_id)
    project = _current_project()
    # Patch 4M — when ``dsm_scenario_id`` is set, run a fresh sim for
    # that scenario via the same helper Impact Assessment uses for
    # multi-DSM fan-out (Patch 2E.1). When unset, read the cached
    # active-scenario sim — full backward compat with the pre-Patch-4M
    # behavior every existing single-scenario caller relies on.
    if body.dsm_scenario_id is not None:
        from mapper.api.dsm import simulate_for_scenario
        sim = simulate_for_scenario(system_id, body.dsm_scenario_id)
    else:
        sim = _proj_results(project).get(system_id)
    if sim is None:
        raise HTTPException(
            status_code=400,
            detail="No simulation results yet. Run /dsm/systems/{id}/simulate first.",
        )
    mapping = _proj_cohort_mappings(project).get(system_id)
    if mapping is None or not mapping.mappings:
        raise HTTPException(
            status_code=400,
            detail="No cohort mappings set. Configure cohort mappings in Impact Assessment first.",
        )
    if body.year_start is not None and body.year_end is not None and body.year_start > body.year_end:
        raise HTTPException(status_code=400, detail="year_start must be <= year_end.")

    # Patch 4M — parameter scenario resolution. When the scenario name
    # is set (and not "Base"), apply the parameter engine to clone each
    # archetype with resolved quantity expressions before flowing
    # through ``compute_material_flows``. Mirrors single-product LCA's
    # ``calculate_archetype`` (lca.py) pattern. ``None`` / "Base" keeps
    # base values — full backward compat.
    archetypes_raw = _proj_archetypes(project)
    archetypes = archetypes_raw
    if body.parameter_scenario is not None and body.parameter_scenario != "Base":
        from mapper.api.parameters import _table_for
        from mapper.core.bom_engine import resolve_archetype_with_engine
        from mapper.core.parameter_engine import ParameterEngine, ParameterError
        table = _table_for(project)
        if body.parameter_scenario not in table.list_scenarios():
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Parameter scenario '{body.parameter_scenario}' not found in active table"
                ),
            )
        try:
            engine = ParameterEngine(table, scenario=body.parameter_scenario)
            archetypes = {
                aid: resolve_archetype_with_engine(arc, engine)
                for aid, arc in archetypes_raw.items()
            }
        except ParameterError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Parameter resolution failed: {e}",
            )
    cohort_to_archetype: dict[str, tuple[str, float]] = {}
    for entry in mapping.mappings:
        if entry.archetype_id not in archetypes:
            continue
        cohort_to_archetype[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    sys_def = _get_system(system_id)

    # Discover dependent subsystems and compute their flows too.
    from mapper.api import subsystems as _subs
    from mapper.core.dsm_lca_engine import build_subsystem_cohort_mapping
    from mapper.core.subsystem_engine import compute_dependent_subsystem
    from mapper.models.bom_schemas import SubsystemRef

    dep_subs = _subs.get_subsystems_for_system(system_id, project)
    setup_warnings: list[str] = []
    sub_runs: list[tuple[str, str, object, dict[str, tuple[str, float]]]] = []
    for sub_id, sub in dep_subs.items():
        if not sub.dependency_rules:
            continue
        sub_mapping, unmapped = build_subsystem_cohort_mapping(sub)
        if unmapped:
            setup_warnings.append(
                f"Subsystem '{sub.name}': {len(unmapped)} unmapped archetype"
                f"{'s' if len(unmapped) != 1 else ''} excluded from material flows: "
                f"{', '.join(unmapped)}"
            )
        if not sub_mapping:
            continue
        missing = [aid for aid, _ in sub_mapping.values() if aid not in archetypes]
        if missing:
            setup_warnings.append(
                f"Subsystem '{sub.name}': archetypes not in library, skipped: {', '.join(sorted(set(missing)))}"
            )
            continue
        try:
            sub_sim = compute_dependent_subsystem(sub, sys_def, sim, None)
        except ValueError as e:
            setup_warnings.append(f"Subsystem '{sub.name}': {e}")
            continue
        sub_runs.append((sub_id, sub.name, sub_sim, sub_mapping))

    try:
        primary = compute_material_flows(
            sim=sim,
            archetypes=archetypes,
            cohort_mappings=cohort_to_archetype,
            scope=body.scope,
            year_start=body.year_start,
            year_end=body.year_end,
            group_by=body.group_by,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    subsystems_meta: list[SubsystemRef] = []
    if sub_runs:
        subsystems_meta.append(SubsystemRef(id=system_id, name=sys_def.name))

    merged_materials = list(primary.materials)
    stage_set: set[str] = set(primary.stages_included)
    stages_out: list[str] = list(primary.stages_included)
    year_min = primary.year_start
    year_max = primary.year_end

    for sub_id, sub_name, sub_sim, sub_mapping in sub_runs:
        try:
            sub_result = compute_material_flows(
                sim=sub_sim,
                archetypes=archetypes,
                cohort_mappings=sub_mapping,
                scope=body.scope,
                year_start=body.year_start,
                year_end=body.year_end,
                group_by=body.group_by,
            )
        except ValueError as e:
            setup_warnings.append(f"Subsystem '{sub_name}': {e}")
            continue
        subsystems_meta.append(SubsystemRef(id=sub_id, name=sub_name))
        for m in sub_result.materials:
            m.subsystem_id = sub_id
            m.subsystem_name = sub_name
            merged_materials.append(m)
        for s in sub_result.stages_included:
            if s not in stage_set:
                stage_set.add(s)
                stages_out.append(s)
        if sub_result.materials:
            year_min = min(year_min, sub_result.year_start)
            year_max = max(year_max, sub_result.year_end)

    merged_materials.sort(key=lambda s: sum(s.values.values()), reverse=True)

    # Unit-count context from the DSM simulation result. Maps scope → which
    # YearResult surface to read. "all" folds everything into stock (the
    # operating-fleet count) as a sensible default for the full-lifecycle view.
    scope_attr = {
        "inflows": "inflow",
        "stock": "stock",
        "outflows": "outflow",
        "all": "stock",
    }.get(primary.scope, "stock")
    system_units_by_year: dict[int, float] = {}
    # Key by archetype *name* (not id) so frontend can index it with
    # ``MaterialSeries.name`` under group_by="archetype".
    archetype_units_by_year: dict[str, dict[int, float]] = {}
    for yr in sim.years:
        if yr.year < year_min or yr.year > year_max:
            continue
        per_cohort = getattr(yr, scope_attr, {}) or {}
        system_units_by_year[yr.year] = float(sum(per_cohort.values()))
        for ck, count in per_cohort.items():
            archetype_id = cohort_to_archetype.get(ck, (ck, 1.0))[0]
            arc = archetypes.get(archetype_id)
            arc_key = arc.name if arc else archetype_id
            bucket = archetype_units_by_year.setdefault(arc_key, {})
            bucket[yr.year] = bucket.get(yr.year, 0.0) + float(count)

    return MaterialFlowResult(
        scope=primary.scope,
        stages_included=stages_out,
        year_start=year_min,
        year_end=year_max,
        group_by=primary.group_by,
        materials=merged_materials,
        elapsed_seconds=primary.elapsed_seconds,
        subsystems=subsystems_meta,
        warnings=setup_warnings,
        compute_metrics=meter.build(),
        unit_name=sys_def.unit_name or "units",
        system_units_by_year=system_units_by_year,
        archetype_units_by_year=archetype_units_by_year,
    )


@router.post(
    "/dsm/systems/{system_id}/material-flows-multi",
    response_model=MultiMaterialFlowResult,
)
async def material_flows_multi(
    system_id: str, body: MaterialFlowMultiRequest,
) -> MultiMaterialFlowResult:
    """Patch 4M — multi-axis fan-out for Material Flows.

    One axis at a time (axisConflict). Server-side sync loop:
    each scenario in the chosen axis spawns a single
    ``compute_material_flows`` call (same code path as the legacy
    single-result endpoint, just with the per-scenario in-task
    fields set on the cloned request body). Returns the assembled
    envelope; the frontend renders a scenario tab bar above the
    result table to switch between runs.

    LCI scenarios are deliberately not an axis here — MFA tracks
    physical material throughputs, which don't depend on the
    background LCI database. See CLAUDE.md "Material Flows axes".
    """
    import time as _time
    dsm_ids = body.dsm_scenario_ids or []
    param_ids = body.parameter_scenarios or []
    if dsm_ids and param_ids:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot fan out across DSM and parameter axes "
                "simultaneously (axisConflict). Pick one axis at a time."
            ),
        )
    if not dsm_ids and not param_ids:
        raise HTTPException(
            status_code=400,
            detail=(
                "At least one of dsm_scenario_ids / parameter_scenarios "
                "must be non-empty."
            ),
        )

    axis = "dsm" if dsm_ids else "parameter"
    runs: list[MaterialFlowScenarioRun] = []
    t0 = _time.perf_counter()

    if axis == "dsm":
        # Resolve scenario labels up front so the envelope is
        # self-contained for the frontend tab bar — no follow-up
        # round trip to the DSM scenario list endpoint.
        from mapper.api.dsm import _get_or_create_state
        state = _get_or_create_state(system_id)
        label_by_id: dict[str, str] = {}
        for sc in state.scenarios:
            label_by_id[sc.id] = sc.name
        for sid in dsm_ids:
            if sid not in label_by_id:
                raise HTTPException(
                    status_code=404,
                    detail=f"DSM scenario '{sid}' not found on system '{system_id}'.",
                )
            sub_body = MaterialFlowRequest(
                scope=body.scope,
                year_start=body.year_start,
                year_end=body.year_end,
                group_by=body.group_by,
                dsm_scenario_id=sid,
                parameter_scenario=None,
            )
            sub_result = await material_flows(system_id, sub_body)
            runs.append(MaterialFlowScenarioRun(
                axis="dsm",
                scenario_id=sid,
                scenario_label=label_by_id[sid],
                result=sub_result,
            ))
    else:  # parameter axis
        for pid in param_ids:
            sub_body = MaterialFlowRequest(
                scope=body.scope,
                year_start=body.year_start,
                year_end=body.year_end,
                group_by=body.group_by,
                dsm_scenario_id=None,
                parameter_scenario=pid,
            )
            sub_result = await material_flows(system_id, sub_body)
            runs.append(MaterialFlowScenarioRun(
                axis="parameter",
                scenario_id=pid,
                scenario_label=pid,  # parameter scenario name == its label
                result=sub_result,
            ))

    elapsed = _time.perf_counter() - t0
    return MultiMaterialFlowResult(axis=axis, runs=runs, elapsed_seconds=elapsed)


@router.get("/dsm/systems/{system_id}/material-flows/export")
async def export_material_flows(
    system_id: str,
    scope: str = "stock",
    year_start: int | None = None,
    year_end: int | None = None,
) -> Response:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    sys_def = _get_system(system_id)
    project = _current_project()
    sim = _proj_results(project).get(system_id)
    if sim is None:
        raise HTTPException(status_code=400, detail="No simulation results.")
    mapping = _proj_cohort_mappings(project).get(system_id)
    if mapping is None or not mapping.mappings:
        raise HTTPException(status_code=400, detail="No cohort mappings set.")

    archetypes = _proj_archetypes(project)
    cohort_to_archetype: dict[str, tuple[str, float]] = {}
    for entry in mapping.mappings:
        if entry.archetype_id not in archetypes:
            continue
        cohort_to_archetype[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    result = compute_material_flows(
        sim=sim,
        archetypes=archetypes,
        cohort_mappings=cohort_to_archetype,
        scope=scope,
        year_start=year_start,
        year_end=year_end,
        group_by="material",
    )

    # Pre-compute unit counts aligned to the export scope — used on the Summary
    # sheet and the long-format "By Archetype" sheet.
    scope_attr = {
        "inflows": "inflow",
        "stock": "stock",
        "outflows": "outflow",
        "all": "stock",
    }.get(scope, "stock")
    unit_label = (sys_def.unit_name or "units").strip() or "units"
    system_units_by_year_exp: dict[int, float] = {}
    archetype_units_by_year_exp: dict[str, dict[int, float]] = {}
    for yr in sim.years:
        if yr.year < result.year_start or yr.year > result.year_end:
            continue
        per_cohort = getattr(yr, scope_attr, {}) or {}
        system_units_by_year_exp[yr.year] = float(sum(per_cohort.values()))
        for ck, count in per_cohort.items():
            archetype_id = cohort_to_archetype.get(ck, (ck, 1.0))[0]
            bucket = archetype_units_by_year_exp.setdefault(archetype_id, {})
            bucket[yr.year] = bucket.get(yr.year, 0.0) + float(count)

    # ── Build workbook ──────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2D8A8A")
    num_fmt = "#,##0.00"

    def _auto(ws) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for cell in col_cells:
                if cell.value is not None:
                    widest = max(widest, min(60, len(str(cell.value))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Summary
    ws = wb.create_sheet("Summary", 0)
    ws.append(["Material Flows Export"])
    ws.append([])
    ws.append(["System", sys_def.name])
    ws.append(["Scope", result.scope])
    ws.append(["Years", f"{result.year_start} - {result.year_end}"])
    ws.append(["Stages included", ", ".join(result.stages_included)])
    ws.append(["Materials", len(result.materials)])
    ws.append([f"Unit label ({unit_label})", unit_label])
    total_units_start = system_units_by_year_exp.get(result.year_start, 0.0)
    total_units_end = system_units_by_year_exp.get(result.year_end, 0.0)
    ws.append([f"Total {unit_label} ({result.year_start})", total_units_start])
    ws.append([f"Total {unit_label} ({result.year_end})", total_units_end])
    ws.append(["Elapsed (s)", result.elapsed_seconds])
    for cell in ws[1]:
        if cell.value:
            cell.font = Font(bold=True, size=14)
    _auto(ws)

    # Sheet 2: Annual Totals (material × year pivot)
    years = sorted({yr for m in result.materials for yr in m.values})
    if years:
        ws = wb.create_sheet("Annual Totals")
        header = ["Material", "Unit", "Stage", "Component"] + [str(y) for y in years] + ["Total"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for m in result.materials:
            row = [m.name, m.unit, m.stage, m.component]
            total = 0.0
            for y in years:
                v = m.values.get(y, 0.0)
                row.append(v)
                total += v
            row.append(total)
            ws.append(row)
        ws.freeze_panes = "E2"
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=len(header)):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 3: By Component (aggregate)
    comp_result = compute_material_flows(
        sim=sim, archetypes=archetypes, cohort_mappings=cohort_to_archetype,
        scope=scope, year_start=year_start, year_end=year_end, group_by="component",
    )
    if years and comp_result.materials:
        ws = wb.create_sheet("By Component")
        header = ["Component", "Unit"] + [str(y) for y in years] + ["Total"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for m in comp_result.materials:
            row: list = [m.name, m.unit]
            total = 0.0
            for y in years:
                v = m.values.get(y, 0.0)
                row.append(v)
                total += v
            row.append(total)
            ws.append(row)
        ws.freeze_panes = "C2"
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=len(header)):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 4: By Archetype
    if years:
        ws = wb.create_sheet("By Archetype")
        header = ["Material", "Archetype", "Unit"] + [str(y) for y in years] + ["Total"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for m in result.materials:
            for arc_name, arc_years in sorted(m.by_archetype.items()):
                row: list = [m.name, arc_name, m.unit]
                total = 0.0
                for y in years:
                    v = arc_years.get(y, 0.0)
                    row.append(v)
                    total += v
                row.append(total)
                ws.append(row)
        ws.freeze_panes = "D2"
        for row in ws.iter_rows(min_row=2, min_col=4, max_col=len(header)):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 5: Evolution Rates
    evolving = [m for m in result.materials if m.evolution_method]
    if evolving:
        ws = wb.create_sheet("Evolution Rates")
        header = ["Material", "Unit", "Method", "Rate"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for m in evolving:
            rate_str = f"{m.evolution_rate:+.2%}" if m.evolution_rate is not None else ""
            ws.append([m.name, m.unit, m.evolution_method, rate_str])
        _auto(ws)

    # Sheet 6: By Material (long) — one row per (year, material).
    if years:
        ws = wb.create_sheet("By Material (long)")
        header = ["Year", "Scope", "Material", "Unit", "Stage", "Component", "Value"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for y in years:
            for m in result.materials:
                v = m.values.get(y, 0.0)
                if v == 0:
                    continue
                ws.append([y, result.scope, m.name, m.unit, m.stage, m.component, v])
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 7: By Component (long)
    if years and comp_result.materials:
        ws = wb.create_sheet("By Component (long)")
        header = ["Year", "Scope", "Component", "Unit", "Value"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for y in years:
            for m in comp_result.materials:
                v = m.values.get(y, 0.0)
                if v == 0:
                    continue
                ws.append([y, result.scope, m.name, m.unit, v])
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 8: By Stage (long) — kg per stage per year, plus a system-wide
    # stage split. Stage grouping folds all materials into the stage buckets
    # the compute engine emits.
    stage_result = compute_material_flows(
        sim=sim, archetypes=archetypes, cohort_mappings=cohort_to_archetype,
        scope=scope, year_start=year_start, year_end=year_end, group_by="stage",
    )
    if years and stage_result.materials:
        ws = wb.create_sheet("By Stage (long)")
        header = ["Year", "Scope", "Stage", "Unit", "Value"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for y in years:
            for m in stage_result.materials:
                v = m.values.get(y, 0.0)
                if v == 0:
                    continue
                ws.append([y, result.scope, m.name, m.unit, v])
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 9: By Archetype (long) — year × archetype × kg PLUS unit counts
    # from the DSM simulation. One "Units" column captures the product-level
    # count for the selected scope (inflows/stock/outflows).
    if years:
        ws = wb.create_sheet("By Archetype (long)")
        header = ["Year", "Scope", "Archetype", "Material", "Unit", "Value", f"Units ({unit_label})"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for y in years:
            for m in result.materials:
                for arc_name, arc_years in sorted(m.by_archetype.items()):
                    v = arc_years.get(y, 0.0)
                    if v == 0:
                        continue
                    arc_units = archetype_units_by_year_exp.get(arc_name, {}).get(y, 0.0)
                    ws.append([y, result.scope, arc_name, m.name, m.unit, v, arc_units])
        ws.freeze_panes = "A2"
        for row in ws.iter_rows(min_row=2, min_col=6, max_col=7):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    # Sheet 10: Unit Counts — system-wide product counts per year for the
    # selected scope. Small sheet, but makes the "context" explicit so users
    # don't have to cross-reference the DSM simulation export.
    if system_units_by_year_exp:
        ws = wb.create_sheet("Unit Counts")
        header = ["Year", "Scope", f"System {unit_label}"]
        ws.append(header)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        for y in sorted(system_units_by_year_exp):
            ws.append([y, scope, system_units_by_year_exp[y]])
        for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
            for cell in row:
                cell.number_format = num_fmt
        _auto(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{_sanitize_filename(sys_def.name, 'material_flows')}_flows_{scope}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Useful for cleanup if the user deletes the parent DSM system from another router.
def purge_system(system_id: str) -> None:
    # Called from dsm.delete_system after the system is already removed from the
    # current project's stores. Clean any cohort mappings / DSM-LCA results for
    # this system across every project (the system id is unique per project in
    # practice, but cheap to sweep all).
    for proj_map in _cohort_mappings.values():
        proj_map.pop(system_id, None)
    for proj_res in _dsm_lca_results.values():
        proj_res.pop(system_id, None)


# ── Excel export / import ────────────────────────────────────────────────────
# Each row represents a BOM node. "Parent" is the direct parent component's
# name (scoped to the current Stage). Empty parent means the row IS a stage
# root — Name becomes the stage root's name.


_BOM_COLUMNS = [
    "Stage",
    "Scope",
    "Parent",
    "Name",
    "Type",
    "Quantity",
    "Unit",
    "Ecoinvent Database",
    "Ecoinvent Code",
    "Ecoinvent Name",
    "Ecoinvent Location",
    "Evolution Method",
    "Learning Rate",
    "Rebound Rate",
    "Base Year",
    "Milestone Years",
    "Milestone Values",
    "Rebound Applies To Stages",
]


def _format_milestones(ms: list[QuantityMilestone]) -> tuple[str, str]:
    """Serialize milestones as two parallel ``;``-separated strings."""
    ordered = sorted(ms, key=lambda m: m.year)
    years = ";".join(str(m.year) for m in ordered)
    values = ";".join(f"{m.quantity:g}" for m in ordered)
    return years, values


def _parse_milestones(years_str: str, values_str: str) -> list[QuantityMilestone]:
    """Inverse of ``_format_milestones``. Empty input returns []."""
    if not years_str or not values_str:
        return []
    ys = [t.strip() for t in years_str.replace(",", ";").split(";") if t.strip()]
    vs = [t.strip() for t in values_str.replace(",", ";").split(";") if t.strip()]
    out: list[QuantityMilestone] = []
    for y, v in zip(ys, vs):
        try:
            out.append(QuantityMilestone(year=int(float(y)), quantity=float(v)))
        except (TypeError, ValueError):
            continue
    return sorted(out, key=lambda m: m.year)


def _walk_for_export(
    node: BOMNode,
    stage: str,
    parent_name: str,
    rows: list[list],
    stage_scope: str = "",
) -> None:
    link = node.ecoinvent_activity
    ev = node.evolution
    # Materials always advertise an evolution method ("fixed" when absent) so
    # readers can tell at a glance that the field is wired; components leave
    # the column empty since they can't carry evolution.
    if node.node_type == "material":
        ev_method = ev.method if ev else "fixed"
    else:
        ev_method = ""
    ev_lr = ev.learning_rate if (ev and ev.method == "learning_rate" and ev.learning_rate is not None) else ""
    ev_rb = ev.rebound_rate if (ev and ev.method == "rebound_effect" and ev.rebound_rate is not None) else ""
    ev_base = ev.base_year if (ev and ev.method in ("learning_rate", "rebound_effect")) else ""
    ms_years, ms_values = "", ""
    if ev and ev.method == "milestones" and ev.milestones:
        ms_years, ms_values = _format_milestones(ev.milestones)
    rb_stages = (
        ";".join(ev.applies_to_stages)
        if (ev and ev.method == "rebound_effect" and ev.applies_to_stages)
        else ""
    )
    # Scope is only emitted on the stage root row (parent_name == "").
    scope_cell = stage_scope if not parent_name else ""
    # Preserve parameter expressions across round-trip: if the node was
    # defined with an expression, emit the expression string rather than the
    # last-resolved numeric value.
    quantity_cell = node.quantity_expression if node.quantity_expression else node.quantity
    rows.append([
        stage,
        scope_cell,
        parent_name,
        node.name,
        node.node_type,
        quantity_cell,
        node.unit,
        link.database if link else "",
        link.code if link else "",
        link.name if link else "",
        link.location if link else "",
        ev_method,
        ev_lr,
        ev_rb,
        ev_base,
        ms_years,
        ms_values,
        rb_stages,
    ])
    if node.children:
        for child in node.children:
            _walk_for_export(child, stage, node.name, rows, stage_scope)


def _build_export_workbook(arc: Archetype) -> Workbook:
    wb = Workbook()
    bom_ws = wb.active
    bom_ws.title = "BOM"
    bom_ws.append(_BOM_COLUMNS)

    rows: list[list] = []
    for root in arc.bom:
        # Stage root itself has empty parent; children use root.name as parent.
        _walk_for_export(root, root.name, "", rows, stage_scope=root.scope or "")
    for r in rows:
        bom_ws.append(r)

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Field", "Value"])
    summary_ws.append(["Name", arc.name])
    summary_ws.append(["Description", arc.description or ""])
    summary_ws.append(["Category", arc.category or ""])
    summary_ws.append(["Stages", len(arc.bom)])
    summary_ws.append(["Materials", material_count_total(arc.bom)])
    summary_ws.append(["Unlinked materials", unlinked_count_total(arc.bom)])
    flat = flatten_roots(arc.bom)
    summary_ws.append(["Total mass (kg)", total_mass_kg(flat)])

    # Timeline sheet — only written when at least one material has evolution,
    # so static BOMs stay single-sheet-simple.
    if has_evolution(arc.bom):
        years = list(range(2025, 2051, 5))
        timeline = generate_archetype_timeline(arc, years)
        tl_ws = wb.create_sheet("Timeline")
        tl_ws.append(["Stage", "Material", "Unit", "Method", *[str(y) for y in years]])
        for row in timeline.rows:
            stage = row.path[0] if row.path else ""
            node = find_node_in_roots(arc.bom, row.node_id)
            method = (node.evolution.method if (node and node.evolution) else "fixed")
            tl_ws.append([
                stage,
                row.name,
                row.unit,
                method,
                *[row.quantities.get(y, 0.0) for y in years],
            ])
        tl_ws.append([])
        tl_ws.append(["Total mass (kg)", "", "", "", *[timeline.total_mass_by_year.get(y, 0.0) for y in years]])
    return wb


@router.get("/bom/archetypes/{arc_id}/export")
async def export_archetype(arc_id: str) -> Response:
    arc = _get_archetype(arc_id)
    wb = _build_export_workbook(arc)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = _sanitize_filename(arc.name) + "_bom.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Multi-archetype export ───────────────────────────────────────────────────


_MULTI_BOM_COLUMNS = ["archetype_name", *_BOM_COLUMNS]


def _build_multi_export_workbook(archetypes: list[Archetype]) -> Workbook:
    wb = Workbook()
    # Archetypes metadata
    meta_ws = wb.active
    meta_ws.title = "Archetypes"
    meta_ws.append(["archetype_name", "folder", "description"])
    for arc in archetypes:
        meta_ws.append([arc.name, arc.folder or "", arc.description or ""])

    # Combined BOM
    bom_ws = wb.create_sheet("BOM")
    bom_ws.append(_MULTI_BOM_COLUMNS)
    for arc in archetypes:
        per_arc_rows: list[list] = []
        for root in arc.bom:
            _walk_for_export(root, root.name, "", per_arc_rows, stage_scope=root.scope or "")
        for r in per_arc_rows:
            bom_ws.append([arc.name, *r])

    # Scaling Reference — reuse the template content.
    scaling = wb.create_sheet("Scaling Reference")
    scaling.append(["Cohort Key", "Archetype", "Scaling Factor", "Notes"])
    scaling.append(["Usage", "Apply these factors when mapping cohorts to an archetype."])
    scaling.append(["Formula", "demand = count × scaling_factor × material_quantity"])

    # Instructions
    ins = wb.create_sheet("Instructions")
    ins.append(["Column", "Notes"])
    ins.append(["archetype_name (Archetypes sheet)", "Unique within the file. Referenced by every row in the BOM sheet."])
    ins.append(["folder (Archetypes sheet)", "Forward-slash path (e.g. 'Group_1/Type_A'). Empty = root. Max depth 5."])
    ins.append(["description (Archetypes sheet)", "Optional description."])
    ins.append(["archetype_name (BOM sheet)", "Points at an archetype in the Archetypes sheet."])
    ins.append(["Stage / Scope / Parent / Name / Type", "Same semantics as the single-archetype format. 'Scope' is 'inflows' | 'stock' | 'outflows' and is set ONLY on the stage root row (Parent empty). Empty = fall back to keyword matching."])
    ins.append(["Evolution Method", "One of 'fixed', 'learning_rate', 'rebound_effect', 'milestones'. Materials only."])
    ins.append(["Rebound Rate", "For 'rebound_effect': annual fractional *increase* in consumption (e.g. 0.02 = +2%/yr)."])
    ins.append(["Rebound Applies To Stages", "Optional for 'rebound_effect': ';'-separated stage names limiting where rebound is applied. Empty = all stages."])
    ins.append(["Milestone Years / Values", "Two ';'-separated parallel lists. Linear interpolation between."])

    return wb


@router.get("/bom/template")
async def download_bom_template() -> Response:
    wb = Workbook()
    meta_ws = wb.active
    meta_ws.title = "Archetypes"
    meta_ws.append(["archetype_name", "folder", "description"])
    # Generic placeholders. MApper is domain-agnostic — examples cover three
    # archetypes with different stage shapes so users see the structure, not a
    # specific case study.
    meta_ws.append(["Product_A", "Group_1/Type_A", "Replace with your archetype (e.g. wind turbine, vehicle, building, device)."])
    meta_ws.append(["Product_B", "Group_1/Type_B", "A second archetype illustrating per-year stock-phase consumption."])
    meta_ws.append(["Asset_C", "Group_2/Type_C", "An archetype with no stock-phase consumption (manufacture + EOL only)."])

    bom_ws = wb.create_sheet("BOM")
    bom_ws.append(_MULTI_BOM_COLUMNS)

    def row(arc: str, *values):
        bom_ws.append([arc, *values])

    # Column order after archetype_name: Stage, Scope, Parent, Name, Type,
    # Quantity, Unit, Ecoinvent Database/Code/Name/Location, Evolution Method,
    # Learning Rate, Rebound Rate, Base Year, Milestone Years, Milestone Values,
    # Rebound Applies To Stages.
    # Scope is set on the stage ROOT row (Parent empty); child rows leave it blank.

    # Product_A — full lifecycle with all four evolution methods illustrated.
    row("Product_A", "Manufacturing", "inflows", "", "Manufacturing", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_A", "Manufacturing", "", "Manufacturing", "Steel", "material", 900, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_STEEL", "steel production", "GLO",
        "fixed", "", "", "", "", "", "")
    row("Product_A", "Manufacturing", "", "Manufacturing", "Aluminium", "material", 120, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_ALU", "aluminium production", "GLO",
        "learning_rate", -0.02, "", 2025, "", "", "")
    row("Product_A", "Key Component", "inflows", "", "Key Component", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_A", "Key Component", "", "Key Component", "Functional material", "material", 300, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_FUNCMAT", "functional material, generic", "GLO",
        "milestones", "", "", "", "2025;2035;2050", "300;230;180", "")
    row("Product_A", "Use Phase", "stock", "", "Use Phase", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_A", "Use Phase", "", "Use Phase", "Energy consumption", "material", 2550, "kWh",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_ELEC_MIX", "market for electricity, low voltage", "EU",
        "rebound_effect", "", 0.015, 2025, "", "", "Use Phase")
    row("Product_A", "End of Life", "outflows", "", "End of Life", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_A", "End of Life", "", "End of Life", "Treatment, generic", "material", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")

    # Product_B — rebound with applies_to_stages filter + maintenance stage.
    row("Product_B", "Manufacturing", "inflows", "", "Manufacturing", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_B", "Manufacturing", "", "Manufacturing", "Steel", "material", 1100, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_STEEL", "steel production", "GLO",
        "fixed", "", "", "", "", "", "")
    row("Product_B", "Use Phase", "stock", "", "Use Phase", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_B", "Use Phase", "", "Use Phase", "Fuel consumption", "material", 727, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_FUEL", "fuel, generic", "GLO",
        "rebound_effect", "", 0.02, 2025, "", "", "Use Phase")
    row("Product_B", "Maintenance", "stock", "", "Maintenance", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_B", "Maintenance", "", "Maintenance", "Replacement part", "material", 13, "kg", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_B", "End of Life", "outflows", "", "End of Life", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Product_B", "End of Life", "", "End of Life", "Treatment, generic", "material", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")

    # Asset_C — no stock-phase consumption (manufacture + EOL only).
    row("Asset_C", "Housing", "inflows", "", "Housing", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Asset_C", "Housing", "", "Housing", "Steel enclosure", "material", 250, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_STEEL", "steel production", "GLO",
        "fixed", "", "", "", "", "", "")
    row("Asset_C", "Electronics", "inflows", "", "Electronics", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Asset_C", "Electronics", "", "Electronics", "Power electronics", "material", 60, "kg",
        "ecoinvent-3.9.1-cutoff", "EXAMPLE_CODE_ELEC", "electronics, unspecified", "GLO",
        "learning_rate", -0.03, "", 2025, "", "", "")
    row("Asset_C", "End of Life", "outflows", "", "End of Life", "component", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")
    row("Asset_C", "End of Life", "", "End of Life", "Treatment, e-waste", "material", 1, "piece", "", "", "", "", "", "", "", "", "", "", "")

    instructions = wb.create_sheet("Instructions")
    instructions.append(["Column", "Notes"])
    instructions.append(["archetype_name (Archetypes sheet)", "Unique per file. Referenced by every row in the BOM sheet."])
    instructions.append(["folder (Archetypes sheet)", "Forward-slash path, e.g. 'Group_1/Type_A'. Empty = root. Max depth 5. Allowed characters: letters, digits, space, _, -"])
    instructions.append(["description (Archetypes sheet)", "Optional description."])
    instructions.append(["archetype_name (BOM sheet, first column)", "Points at an archetype defined in the Archetypes sheet."])
    instructions.append(["Stage", "Life cycle stage name. Examples: Manufacturing, Use Phase, Maintenance, End of Life — but you can name them anything that fits your system. One row per unique Stage with empty Parent acts as that stage's root component."])
    instructions.append(["Scope", "Explicit DSM scope for the stage: 'inflows' (manufacturing / one-time), 'stock' (per-year use-phase / maintenance), 'outflows' (end-of-life). Set ONLY on the stage root row; child rows leave it blank. If empty, the system falls back to keyword-matching on the stage name."])
    instructions.append(["Parent", "Direct parent component name within the same Stage. Empty for stage roots."])
    instructions.append(["Name", "Node name."])
    instructions.append(["Type", "'component' or 'material'. Materials are leaves and may link to ecoinvent. A 'unit' of an archetype is whatever your system tracks: a discrete item (vehicle, building, machine, device) or a continuous quantity (kg, kWh, m³)."])
    instructions.append(["Quantity / Unit", "Numeric quantity (e.g. 53.0) OR a parameter expression (e.g. material_mass * 0.35). Expressions resolve against the active ParameterSet at pipeline time. Operators: + - * / ** ( ). Functions: min, max, abs, round, sum. Use whatever unit fits the material (kg, kWh, m³, kg-CO₂eq, …)."])
    instructions.append(["Ecoinvent *", "Populate Database + Code for materials to link them. Name and Location are informational."])
    instructions.append(["Evolution Method", "One of 'fixed' (or blank), 'learning_rate', 'rebound_effect', 'milestones'. Materials only."])
    instructions.append(["Learning Rate / Base Year", "For 'learning_rate': Quantity(year) = Quantity × (1 + rate)^(year − base_year). Typically negative (efficiency gain)."])
    instructions.append(["Rebound Rate / Base Year", "For 'rebound_effect': annual fractional *increase* in consumption (e.g. 0.02 = +2%/yr). Typically positive. Mutually exclusive with Learning Rate and Milestones. Example: a product becomes more efficient (less consumption per unit of service) but users compensate by using it more — net consumption grows ~2%/yr in the early transition years."])
    instructions.append(["Rebound Applies To Stages", "Optional filter for rebound_effect: semicolon-separated list of stage names (e.g. 'Use Phase' or 'Use Phase;Maintenance'). If empty, rebound applies regardless of stage."])
    instructions.append(["Milestone Years / Values", "For 'milestones': two ';'-separated parallel lists. Linear interpolation between."])

    scaling = wb.create_sheet("Scaling Reference")
    scaling.append(["Cohort Key", "Archetype", "Scaling Factor", "Notes"])
    scaling.append(["Product_A|Class_1", "Product_A", 1.00, "Base size"])
    scaling.append(["Product_A|Class_2", "Product_A", 1.25, "Mid size"])
    scaling.append(["Product_A|Class_3", "Product_A", 1.50, "Large size"])
    scaling.append([])
    scaling.append(["Usage", "Apply these factors when mapping cohorts to an archetype."])
    scaling.append(["Formula", "demand = count × scaling_factor × material_quantity"])
    scaling.append(["Cohort key format", "Pipe-delimited concatenation of your DSM dimension labels (e.g. 'Type_A|Class_1' if your system has 'type' and 'class' dimensions)."])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="mapper_archetypes_template.xlsx"'},
    )


class BOMImportResult(BaseModel):
    id: str
    name: str
    stages: int
    materials: int
    linked: int
    unlinked: int
    warnings: list[str]


def _parse_bom_workbook(
    wb: Workbook,
    archetype_filter: str | None = None,
) -> tuple[list[BOMNode], list[str], list["BOMValidationRow"]]:
    """Parse the BOM sheet into stage-root BOMNode trees.

    Returns ``(roots, parse_warnings, validation_rows)``. The third element is
    a flat list of every material row's bw2 link context (db, code, name,
    location) tagged with its Excel row index — fed into ``validate_bom`` by
    the import endpoint. The validator runs once per workbook (Patch 2).

    If ``archetype_filter`` is set, only rows whose ``archetype_name`` column
    matches are included (for multi-archetype workbooks). The column is
    optional when ``archetype_filter`` is None (single-archetype format)."""
    if "BOM" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Missing 'BOM' sheet in workbook.")
    ws = wb["BOM"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="BOM sheet is empty.")
    header = [str(c or "").strip() for c in header_row]
    required = {"Stage", "Parent", "Name", "Type", "Quantity", "Unit"}
    missing = required - set(header)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"BOM sheet missing required columns: {sorted(missing)}",
        )
    idx = {col: header.index(col) for col in header}
    arc_col_idx: int | None = None
    for key in ("archetype_name", "Archetype", "Archetype Name"):
        if key in idx:
            arc_col_idx = idx[key]
            break

    def col(row: tuple, name: str, default=""):
        i = idx.get(name)
        if i is None or i >= len(row):
            return default
        v = row[i]
        return default if v is None else v

    warnings: list[str] = []
    validation_rows: list[BOMValidationRow] = []
    archetype_label = archetype_filter or ""
    # stage_name -> root BOMNode
    stages: dict[str, BOMNode] = {}
    # (stage, node_name) -> node map — parent lookup is scoped to its stage.
    # Assumes names are unique within a stage; duplicates trigger a warning.
    node_by_name: dict[tuple[str, str], BOMNode] = {}

    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    if archetype_filter is not None and arc_col_idx is not None:
        rows = [
            r for r in all_rows
            if len(r) > arc_col_idx and r[arc_col_idx] is not None
            and str(r[arc_col_idx]).strip() == archetype_filter
        ]
    else:
        rows = all_rows
    # Process stage roots (empty Parent) first so children resolve.
    rows.sort(key=lambda r: 0 if not str(col(r, "Parent") or "").strip() else 1)

    def _ensure_stage_root(stage_name: str, explicit_scope: str | None = None) -> BOMNode:
        if stage_name in stages:
            root = stages[stage_name]
            # Late-arriving explicit scope from a root row processed out of order
            # upgrades the placeholder's scope and is_annual.
            if explicit_scope and not root.scope:
                root.scope = explicit_scope
                root.is_annual = explicit_scope == "stock"
            return root
        resolved = stage_to_scope(stage_name, explicit_scope)
        stage_root = BOMNode(
            name=stage_name, node_type="component", quantity=1, unit="piece",
            scope=explicit_scope or None,
            is_annual=(resolved == "stock"),
            children=[],
        )
        stages[stage_name] = stage_root
        node_by_name[(stage_name, stage_name)] = stage_root
        return stage_root

    for r_idx, row in enumerate(rows, start=2):
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        stage = str(col(row, "Stage") or "").strip()
        parent_name = str(col(row, "Parent") or "").strip()
        name = str(col(row, "Name") or "").strip()
        scope_cell = str(col(row, "Scope") or "").strip().lower()
        explicit_scope: str | None = None
        if scope_cell:
            if scope_cell in ("inflows", "stock", "outflows"):
                explicit_scope = scope_cell
            else:
                warnings.append(
                    f"Row {r_idx}: invalid Scope '{scope_cell}'; must be inflows|stock|outflows. Ignored."
                )
        node_type = str(col(row, "Type") or "").strip().lower() or "component"
        raw_qty = col(row, "Quantity", 1)
        qty: float = 1.0
        qty_expr: str | None = None
        if raw_qty is None or (isinstance(raw_qty, str) and not raw_qty.strip()):
            qty = 1.0
        elif isinstance(raw_qty, (int, float)) and not isinstance(raw_qty, bool):
            qty = float(raw_qty)
        else:
            # String cell: try numeric first (Excel sometimes stores numbers
            # as text), then treat as a parameter expression.
            s = str(raw_qty).strip()
            try:
                qty = float(s)
            except ValueError:
                qty_expr = s
                # Leave ``qty`` at 1.0 as a safe pre-resolution placeholder; the
                # pipeline will re-resolve against the active ParameterSet.
                # If the expression is a plain parameter name and no set is
                # active at import time, we still get 1.0 here — documented
                # behaviour: the editor shows "1 (unresolved)" until a set is
                # picked.
                warnings.append(
                    f"Row {r_idx}: Quantity '{s}' stored as expression; resolved at pipeline time."
                )
        unit = str(col(row, "Unit") or "unit").strip() or "unit"

        if not stage or not name:
            warnings.append(f"Row {r_idx}: missing Stage or Name; skipped.")
            continue
        if node_type not in ("component", "material"):
            warnings.append(f"Row {r_idx}: unknown Type '{node_type}'; defaulted to 'component'.")
            node_type = "component"

        link: EcoinventLink | None = None
        evolution: MaterialEvolution | None = None
        if node_type == "material":
            db = str(col(row, "Ecoinvent Database") or "").strip()
            code = str(col(row, "Ecoinvent Code") or "").strip()
            ec_name = str(col(row, "Ecoinvent Name") or "").strip()
            ec_loc = str(col(row, "Ecoinvent Location") or "").strip()
            if db and code:
                link = EcoinventLink(
                    database=db,
                    code=code,
                    name=ec_name,
                    location=ec_loc,
                )
            # Always emit a validation row for materials — even if one of
            # (db, code) is empty, the validator distinguishes "abstract row"
            # (both blank) from "structural error" (one set, one blank).
            if db or code:
                validation_rows.append(BOMValidationRow(
                    archetype=archetype_label,
                    stage=stage,
                    row_idx=r_idx,
                    name=name,
                    database=db or None,
                    code=code or None,
                    ecoinvent_name=ec_name,
                    ecoinvent_location=ec_loc,
                ))

            ev_method = str(col(row, "Evolution Method") or "").strip().lower()
            if ev_method == "learning_rate":
                try:
                    lr = float(col(row, "Learning Rate", 0) or 0)
                except (TypeError, ValueError):
                    lr = 0.0
                    warnings.append(f"Row {r_idx}: invalid Learning Rate; defaulted to 0.")
                try:
                    base_year = int(float(col(row, "Base Year", 2025) or 2025))
                except (TypeError, ValueError):
                    base_year = 2025
                evolution = MaterialEvolution(
                    method="learning_rate", learning_rate=lr, base_year=base_year
                )
            elif ev_method == "rebound_effect":
                try:
                    rb = float(col(row, "Rebound Rate", 0) or 0)
                except (TypeError, ValueError):
                    rb = 0.0
                    warnings.append(f"Row {r_idx}: invalid Rebound Rate; defaulted to 0.")
                try:
                    base_year = int(float(col(row, "Base Year", 2025) or 2025))
                except (TypeError, ValueError):
                    base_year = 2025
                stages_cell = str(col(row, "Rebound Applies To Stages") or "").strip()
                applies_to: list[str] | None = None
                if stages_cell:
                    parts = [t.strip() for t in stages_cell.replace(",", ";").split(";")]
                    parts = [p for p in parts if p]
                    applies_to = parts or None
                evolution = MaterialEvolution(
                    method="rebound_effect",
                    rebound_rate=rb,
                    base_year=base_year,
                    applies_to_stages=applies_to,
                )
            elif ev_method == "milestones":
                ms = _parse_milestones(
                    str(col(row, "Milestone Years") or ""),
                    str(col(row, "Milestone Values") or ""),
                )
                if len(ms) >= 2:
                    evolution = MaterialEvolution(method="milestones", milestones=ms)
                else:
                    warnings.append(f"Row {r_idx}: milestones need ≥2 year/value pairs; ignored.")
            elif ev_method and ev_method != "fixed":
                warnings.append(f"Row {r_idx}: unknown Evolution Method '{ev_method}'; ignored.")

        node = BOMNode(
            name=name,
            node_type=node_type,
            quantity=qty,
            quantity_expression=qty_expr,
            unit=unit,
            children=[] if node_type == "component" else None,
            ecoinvent_activity=link,
            evolution=evolution,
        )

        if not parent_name:
            # Stage root row. If Name == Stage, this row IS the stage root;
            # otherwise make a stage root and attach as top-level child.
            if name == stage and stage not in stages:
                node.scope = explicit_scope or None
                node.is_annual = stage_to_scope(stage, explicit_scope) == "stock"
                stages[stage] = node
                node_by_name[(stage, stage)] = node
                continue
            parent = _ensure_stage_root(stage, explicit_scope)
            parent.children = parent.children or []
            parent.children.append(node)
            node_by_name[(stage, name)] = node
            continue

        # Non-root row: find parent by (stage, parent_name).
        if stage not in stages:
            _ensure_stage_root(stage, explicit_scope)
            warnings.append(
                f"Row {r_idx}: stage '{stage}' was not defined explicitly; created a default stage root."
            )

        parent = node_by_name.get((stage, parent_name))
        if parent is None:
            warnings.append(
                f"Row {r_idx}: parent '{parent_name}' not found in stage '{stage}'; attached to stage root instead."
            )
            parent = stages[stage]
        if parent.node_type == "material":
            parent.node_type = "component"
            parent.ecoinvent_activity = None
        parent.children = parent.children or []
        parent.children.append(node)
        if (stage, name) in node_by_name:
            warnings.append(
                f"Row {r_idx}: duplicate node name '{name}' in stage '{stage}'; later rows cannot reference it unambiguously."
            )
        node_by_name[(stage, name)] = node

    roots = list(stages.values())
    assign_ids_to_roots(roots)
    return roots, warnings, validation_rows


class MultiImportArchetypeSummary(BaseModel):
    id: str
    name: str
    folder: str | None
    stages: int
    materials: int
    linked: int
    unlinked: int
    action: str = "created"  # "created" | "updated"
    # Patch 2: counts surfaced in the post-upload UI banner.
    validation_error_rows: int = 0
    validation_warning_rows: int = 0


class MultiImportResult(BaseModel):
    format: str  # "single" | "multi"
    mode: str = "merge"  # "merge" | "replace"
    created: int
    updated: int = 0
    folders_created: int
    archetypes: list[MultiImportArchetypeSummary]
    warnings: list[str]
    # Patch 2: per-archetype validation reports keyed by archetype id. The UI
    # renders these in the post-upload modal; the union of all errors here is
    # what blocks LCA compute.
    validation_reports: dict[str, ValidationReport] = {}


def _read_archetypes_sheet(wb: Workbook) -> list[dict]:
    """Parse the Archetypes sheet into a list of dicts with keys:
    archetype_name, folder, description. Raises HTTPException on malformed."""
    ws = wb["Archetypes"]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="Archetypes sheet is empty.")
    header = [str(c or "").strip() for c in header_row]
    if "archetype_name" not in header:
        raise HTTPException(status_code=400, detail="Archetypes sheet missing 'archetype_name' column.")
    idx = {c: header.index(c) for c in header}

    def pick(row: tuple, key: str) -> str:
        i = idx.get(key)
        if i is None or i >= len(row) or row[i] is None:
            return ""
        return str(row[i]).strip()

    out: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        name = pick(row, "archetype_name")
        if not name:
            continue
        out.append({
            "name": name,
            "folder": pick(row, "folder") or None,
            "description": pick(row, "description") or None,
        })
    return out


def _apply_validation_to_archetype(arc: Archetype, report: ValidationReport) -> None:
    """Stamp ``validation_status`` / ``validation_message`` onto every material
    node in ``arc.bom`` based on the report, and attach the report to ``arc``.

    Mutates ``arc`` in place. Resets all material nodes to ``"ok"`` first so
    re-imports clear stale status from a previous validation."""
    by_node = issues_by_node_key(report)
    for node in iter_all_materials(arc.bom):
        # Clear any stale status from a prior import.
        node.validation_status = "ok"
        node.validation_message = None
        # The validator keys on (archetype, stage, name); the stage here is
        # the root BOMNode's name. We need the path from root to material.
        # Find the matching archetype root by walking the tree:
    # Build (stage, name) → node lookup once per archetype.
    node_lookup: dict[tuple[str, str], BOMNode] = {}
    for stage_root in arc.bom:
        for child in iter_all_materials([stage_root]):
            node_lookup[(stage_root.name, child.name)] = child
    for (arc_name, stage, name), bucket in by_node.items():
        if arc_name and arc_name != arc.name:
            continue
        node = node_lookup.get((stage, name))
        if node is None:
            continue
        # Pick the worst severity for this row's status.
        has_error = any(i.severity == "error" for i in bucket)
        node.validation_status = "error" if has_error else "warning"
        node.validation_message = "; ".join(i.message for i in bucket)
    arc.validation_report = report


@router.post(
    "/bom/archetypes/import",
    response_model=MultiImportResult,
    dependencies=[Depends(verify_project_state)],
)
async def import_archetype(
    file: UploadFile = File(...),
    mode: str = Query("merge", pattern="^(merge|replace)$"),
) -> MultiImportResult:
    try:
        data = await file.read()
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    project = _current_project()
    warnings: list[str] = []
    touched: list[tuple[Archetype, str]] = []  # (archetype, "created" | "updated")

    # In replace mode, wipe the existing library up front. Merge mode preserves
    # everything not referenced by the file.
    if mode == "replace":
        with _lock:
            existing = dict(_proj_archetypes(project))
            _proj_archetypes(project).clear()
        for arc_id in existing:
            try:
                dsm_storage.delete_archetype_file(project, arc_id)
            except Exception:  # pragma: no cover — filesystem best-effort
                pass

    # Build a name → existing archetype index (merge mode only) so we can upsert.
    name_to_existing: dict[str, Archetype] = {}
    if mode == "merge":
        for a in _proj_archetypes(project).values():
            name_to_existing[a.name] = a

    def _upsert(
        name: str,
        description: str | None,
        category: str | None,
        folder: str | None,
        roots: list,
    ) -> tuple[Archetype, str]:
        existing = name_to_existing.get(name) if mode == "merge" else None
        if existing is not None and existing.id:
            arc = Archetype(
                id=existing.id,
                name=name,
                description=description if description is not None else existing.description,
                category=category if category is not None else existing.category,
                folder=folder if folder is not None else existing.folder,
                bom=roots,
                created_at=existing.created_at or _now_iso(),
                updated_at=_now_iso(),
            )
            action = "updated"
        else:
            arc = Archetype(
                id=str(uuid.uuid4()),
                name=name,
                description=description,
                category=category,
                folder=folder,
                bom=roots,
                created_at=_now_iso(),
                updated_at=_now_iso(),
            )
            action = "created"
        with _lock:
            _proj_archetypes(project)[arc.id] = arc  # type: ignore[index]
        dsm_storage.save_archetype(project, arc)
        return arc, action

    # Per-archetype validation rows accumulated across the workbook so we can
    # run the bw2 validator once at the end (cheap-first ordering relies on the
    # per-(db,code) cache surviving across archetypes — the same upstream
    # activity is referenced by many archetypes' BOMs).
    arc_to_validation_rows: dict[str, list[BOMValidationRow]] = {}

    if "Archetypes" in wb.sheetnames:
        # ── Multi-archetype format ──
        meta = _read_archetypes_sheet(wb)
        if not meta:
            raise HTTPException(status_code=400, detail="Archetypes sheet has no rows.")
        seen_names: set[str] = set()
        for m in meta:
            if m["name"] in seen_names:
                raise HTTPException(
                    status_code=400,
                    detail=f"Duplicate archetype_name '{m['name']}' in Archetypes sheet.",
                )
            seen_names.add(m["name"])

        for m in meta:
            roots, wrn, vrows = _parse_bom_workbook(wb, archetype_filter=m["name"])
            if not roots:
                warnings.append(f"Archetype '{m['name']}': no BOM rows found; skipped.")
                continue
            for w in wrn:
                warnings.append(f"[{m['name']}] {w}")
            folder = _normalize_folder(m["folder"])
            arc, action = _upsert(m["name"], m["description"], None, folder, roots)
            touched.append((arc, action))
            arc_to_validation_rows[arc.id or ""] = vrows
        fmt = "multi"
    else:
        # ── Legacy single-archetype format (Summary sheet optional). ──
        roots, warnings, single_vrows = _parse_bom_workbook(wb)
        name = file.filename.rsplit(".", 1)[0] if file.filename else "Imported archetype"
        description: str | None = None
        category: str | None = None
        if "Summary" in wb.sheetnames:
            sws = wb["Summary"]
            for row in sws.iter_rows(min_row=2, values_only=True):
                if not row or row[0] is None:
                    continue
                key = str(row[0]).strip().lower()
                val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if key == "name" and val:
                    name = val
                elif key == "description":
                    description = val or None
                elif key == "category":
                    category = val or None
        arc, action = _upsert(name, description, category, None, roots)
        touched.append((arc, action))
        # Re-tag rows with the resolved archetype name (parser used "" since
        # there's no archetype_name column in single-archetype workbooks).
        for vr in single_vrows:
            vr.archetype = arc.name
        arc_to_validation_rows[arc.id or ""] = single_vrows
        fmt = "single"

    # ── Validate every imported archetype against bw2data (Patch 2) ──────
    # One pass per archetype keeps the per-row error messages accurate; the
    # bw2 result cache lives inside each call but is cheap to rebuild.
    for arc, _action in touched:
        vrows = arc_to_validation_rows.get(arc.id or "", [])
        report = validate_bom(vrows, project_name=project) if vrows else ValidationReport(
            total_rows=0, valid_rows=0, error_rows=0, warning_rows=0,
            project_name=project,
        )
        _apply_validation_to_archetype(arc, report)
        # Persist the updated archetype JSON (now with per-row status + report).
        dsm_storage.save_archetype(project, arc)

    # Auto-register any folders that were introduced.
    existing_folders = set(dsm_storage.load_folders(project))
    new_folders: set[str] = set()
    for arc, _ in touched:
        if arc.folder and arc.folder not in existing_folders:
            new_folders.add(arc.folder)
            parts = arc.folder.split("/")
            for i in range(1, len(parts)):
                new_folders.add("/".join(parts[:i]))
    if new_folders:
        dsm_storage.save_folders(project, sorted(existing_folders | new_folders))

    summaries: list[MultiImportArchetypeSummary] = []
    validation_reports: dict[str, ValidationReport] = {}
    created_count = 0
    updated_count = 0
    for arc, action in touched:
        mats = material_count_total(arc.bom)
        unlinked = unlinked_count_total(arc.bom)
        report = arc.validation_report
        err_rows = report.error_rows if report else 0
        warn_rows = report.warning_rows if report else 0
        summaries.append(MultiImportArchetypeSummary(
            id=arc.id or "",
            name=arc.name,
            folder=arc.folder,
            stages=len(arc.bom),
            materials=mats,
            linked=mats - unlinked,
            unlinked=unlinked,
            action=action,
            validation_error_rows=err_rows,
            validation_warning_rows=warn_rows,
        ))
        if report and arc.id:
            validation_reports[arc.id] = report
        if action == "updated":
            updated_count += 1
        else:
            created_count += 1

    return MultiImportResult(
        format=fmt,
        mode=mode,
        created=created_count,
        updated=updated_count,
        folders_created=len(new_folders),
        archetypes=summaries,
        warnings=warnings,
        validation_reports=validation_reports,
    )
