# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Shared per-cohort Excel-sheet construction for Impact Assessment exports.

The STATIC fleet export (``mapper.api.bom._build_mfa_lca_workbook``) and the
PROSPECTIVE multi-LCI export (``mapper.api.impact._build_multi_scenario_workbook``)
both need a "By cohort" and "By subsystem" sheet with the SAME cohort-label
resolution: strip the ``<uuid>::`` subsystem prefix (no raw UUID in any cell),
resolve the owning System name, and resolve the archetype (primary OR subsystem
mapping). Two divergent copies of that logic are how the label bugs kept
recurring, so it lives here once and both builders call it.

The only difference between the two exports is an optional leading
``LCI Scenario`` column (multi-LCI), so the writers take a list of
``(scenario_label | None, results)`` — one entry (label ``None``) for the static
single-scenario case, N entries for multi-LCI.
"""

from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from typing import Any, Literal

from fastapi import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger("mapper.export")

XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# What a workbook IS, which decides how a demo project marks it. Required at
# every call site — there is deliberately no default.
#
#   "data"       results/records a user reads. On a demo project the fictional-
#                data warning is stamped into row 1 of every sheet, because the
#                file travels and nothing else in it says the numbers are fake.
#   "round_trip" an artefact MApper READS BACK — templates, settings exports.
#                Stamping one puts a row above the header and the importer then
#                rejects the file the app just produced. These get the DEMO_
#                filename prefix only; they carry configuration, not fictional
#                measurements.
#
# This was previously `template: bool = False`, i.e. opt-in, and the default was
# wrong for round-trippable files. Two separate endpoints silently inherited the
# stamp and broke on re-import (DSM upload templates, then the AESA config
# export). Making it a required argument means a new export cannot inherit
# either behaviour by accident — it has to say which it is.
ExportKind = Literal["data", "round_trip"]

# Stamped into row 1 of every sheet of every export taken from the demo
# project. A spreadsheet is the one artefact that leaves the application: it
# gets emailed, archived and opened months later with no memory of where it
# came from, and nothing in a grid of numbers says "fictional". The in-app
# banner and the "(FICTIONAL DATA)" name suffixes do not travel with the file,
# so the warning has to be inside the workbook itself.
DEMO_EXPORT_WARNING = (
    "\u26a0 SYNTHETIC DEMO DATA \u2014 every value in this workbook is fictional. "
    "Produced by MApper's demo project to exercise the software without an "
    "ecoinvent licence. This is NOT an environmental assessment \u2014 do not "
    "cite, publish or make decisions from these numbers."
)
_DEMO_FONT = Font(bold=True, color="FFFFFF", size=11)
_DEMO_FILL = PatternFill("solid", fgColor="C00000")

SCI_FMT = "0.00E+00"
INT_FMT = "#,##0"
HEADER_FILL_HEX = "3ECFCF"
COHORT_TAB_HEX = "4A90D9"

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor=HEADER_FILL_HEX)


# ── The one exit for every Excel export ──────────────────────────────────────


def active_project_is_demo() -> bool:
    """True when the active Brightway2 project is MApper's synthetic demo.

    Keyed off the project identity (``demo_project.is_demo_project``), not the
    filename or the sheet contents — a filename heuristic would miss a renamed
    download and would false-positive on a real project someone called "demo".
    Never raises: a broken bw2 state must not block an export.
    """
    try:
        import bw2data as bd

        from mapper.core.demo_project import is_demo_project

        return is_demo_project(bd.projects.current)
    except Exception:  # pragma: no cover - defensive
        logger.warning("export: could not determine demo status", exc_info=True)
        return False


def stamp_demo_warning(wb: Workbook) -> None:
    """Insert the fictional-data warning as row 1 of every sheet.

    ``insert_rows(1)`` shifts existing content down but does NOT move
    ``freeze_panes``, so any frozen row is re-pointed one row lower here —
    otherwise the frozen header would end up showing the warning instead of the
    column titles. Exports use no merged cells or charts (which openpyxl also
    fails to shift), so nothing else needs fixing up.
    """
    for ws in wb.worksheets:
        frozen = ws.freeze_panes
        ws.insert_rows(1)

        cell = ws.cell(row=1, column=1, value=DEMO_EXPORT_WARNING)
        cell.font = _DEMO_FONT
        cell.fill = _DEMO_FILL
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[1].height = 24

        # Fill the banner colour across the used width so it reads as a band
        # rather than one coloured cell.
        for col in range(2, max(ws.max_column, 1) + 1):
            ws.cell(row=1, column=col).fill = _DEMO_FILL

        if frozen:
            m = re.match(r"^([A-Z]+)(\d+)$", str(frozen))
            if m:
                ws.freeze_panes = f"{m.group(1)}{int(m.group(2)) + 1}"


def excel_response(
    wb: Workbook,
    filename: str,
    *,
    kind: ExportKind,
    is_demo: bool | None = None,
) -> Response:
    """Serialise ``wb`` and return it as an .xlsx download.

    Every Excel export in MApper returns through here. Centralising it is what
    makes the demo warning impossible to forget: a new export surface gets the
    stamping by construction rather than by remembering to add it.

    When the active project is the demo, the workbook is stamped AND the
    filename gains a ``DEMO_`` prefix, so the file is identifiable in a
    downloads folder without being opened.

    ``is_demo`` overrides the automatic check (tests, and callers that already
    know).
    """
    demo = active_project_is_demo() if is_demo is None else is_demo
    if demo:
        if kind == "data":
            stamp_demo_warning(wb)
        if not filename.startswith("DEMO_"):
            filename = f"DEMO_{filename}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def excel_response_from_bytes(
    data: bytes,
    filename: str,
    *,
    kind: ExportKind,
    is_demo: bool | None = None,
) -> Response:
    """Same contract as :func:`excel_response`, for callers holding raw bytes.

    Several export endpoints delegate to a builder that already serialised the
    workbook. Rather than reshape those builders, this re-opens the bytes —
    but ONLY when the demo is active. On a real project the bytes are returned
    exactly as produced, so the normal export path is byte-for-byte unchanged
    and carries no round-trip risk.
    """
    demo = active_project_is_demo() if is_demo is None else is_demo
    if demo and kind == "round_trip":
        # See ExportKind: MApper reads this file back, so prefix only.
        if not filename.startswith("DEMO_"):
            filename = f"DEMO_{filename}"
    elif demo:
        try:
            from openpyxl import load_workbook

            wb = load_workbook(io.BytesIO(data))
            stamp_demo_warning(wb)
            out = io.BytesIO()
            wb.save(out)
            data = out.getvalue()
            if not filename.startswith("DEMO_"):
                filename = f"DEMO_{filename}"
        except Exception:
            # A stamping failure must not cost the user their export; the
            # filename prefix still marks it.
            logger.warning("export: demo stamping failed", exc_info=True)
            if not filename.startswith("DEMO_"):
                filename = f"DEMO_{filename}"

    return Response(
        content=data,
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def style_header(ws, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, max_width: int = 40, sample_rows: int = 50) -> None:
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        widest = 0
        for i, cell in enumerate(col_cells):
            if i >= sample_rows + 1:
                break
            v = cell.value
            if v is not None:
                widest = max(widest, min(max_width, len(str(v))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)


def apply_sci(ws, min_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.number_format = SCI_FMT


@dataclass
class CohortResolver:
    """Resolves a (possibly subsystem-prefixed) cohort key to its readable
    parts. Shared so the static and prospective exports can never diverge on
    UUID stripping / System / Archetype resolution."""

    system_name: str
    system_id: str | None
    dim_headers: list[str]
    n_dims: int
    _nonage_names: list[str]
    _mapping_by_cohort: dict[str, tuple[str, float]]
    _sub_by_id: dict[str, dict]
    _archetypes: dict[str, Any]

    def split_prefix(self, ck: str) -> tuple[str | None, str]:
        """(owner_id, cohort_suffix). owner_id None for a non-prefixed key."""
        if "::" in ck:
            pid, rest = ck.split("::", 1)
            return pid, rest
        return None, ck

    def is_subsystem(self, ck: str) -> bool:
        pid, _ = self.split_prefix(ck)
        return pid is not None and pid != self.system_id

    def system_for(self, ck: str) -> str:
        """Human-readable owning-system name for a (possibly prefixed) cohort."""
        pid, _ = self.split_prefix(ck)
        if pid is None or pid == self.system_id:
            return self.system_name
        return self._sub_by_id.get(pid, {}).get("name") or self.system_name

    def resolve_arc(self, ck: str) -> tuple[str, float]:
        """(archetype_id, scale) — primary mapping (bare, even under a
        ``<system_id>::`` prefix) OR the owning subsystem's mapping."""
        pid, rest = self.split_prefix(ck)
        if pid is not None and pid != self.system_id:
            return self._sub_by_id.get(pid, {}).get("mappings", {}).get(rest, ("", 1.0))
        return self._mapping_by_cohort.get(rest, ("", 1.0))

    def arc_name(self, ck: str) -> str:
        arc_id, _ = self.resolve_arc(ck)
        arc = self._archetypes.get(arc_id)
        return getattr(arc, "name", "") if arc is not None else ""

    def arc_scale(self, ck: str) -> float:
        _, scale = self.resolve_arc(ck)
        return scale

    def split_cohort(self, ck: str) -> list[str]:
        """Readable cohort split into dim columns — UUID prefix stripped."""
        display = ck.split("::", 1)[-1] if "::" in ck else ck
        parts = display.split("|")
        out = parts[: self.n_dims] if self._nonage_names else [display]
        while len(out) < self.n_dims:
            out.append("")
        return out


def build_cohort_resolver(
    *,
    system_name: str,
    system_id: str | None,
    cohort_mapping: Any | None,
    subsystems: list | None,
    archetypes: dict[str, Any],
    dims: list | None,
) -> CohortResolver:
    mapping_by_cohort: dict[str, tuple[str, float]] = {}
    if cohort_mapping is not None:
        for entry in cohort_mapping.mappings:
            mapping_by_cohort[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    nonage_names: list[str] = []
    for d in dims or []:
        name = getattr(d, "name", None) or (d.get("name") if isinstance(d, dict) else None)
        is_age = getattr(d, "is_age", None)
        if is_age is None and isinstance(d, dict):
            is_age = d.get("is_age", False)
        if name and not is_age:
            nonage_names.append(name)
    dim_headers = [d.capitalize() for d in nonage_names] or ["Cohort"]

    return CohortResolver(
        system_name=system_name,
        system_id=system_id,
        dim_headers=dim_headers,
        n_dims=len(dim_headers),
        _nonage_names=nonage_names,
        _mapping_by_cohort=mapping_by_cohort,
        _sub_by_id={s["id"]: s for s in (subsystems or [])},
        _archetypes=archetypes,
    )


def _union_years(scenarios: list[tuple[str | None, list]]) -> list[int]:
    ys: set[int] = set()
    for _, results in scenarios:
        for r in results:
            for yr in r.years:
                ys.add(yr.year)
    return sorted(ys)


def _union_cohorts(scenarios: list[tuple[str | None, list]]) -> list[str]:
    cs: set[str] = set()
    for _, results in scenarios:
        for r in results:
            for yr in r.years:
                cs.update(yr.impact_by_cohort.keys())
    return sorted(cs)


def write_by_cohort_sheet(
    wb: Workbook,
    scenarios: list[tuple[str | None, list]],
    resolver: CohortResolver,
    sim_counts: dict[int, dict[str, float]] | None,
    labels: list[str],
    units: list[str],
    *,
    scenario_header: str = "LCI Scenario",
) -> None:
    """"By cohort" sheet: one row per (scenario ×) year × cohort. Columns:
    Year [| <scenario_header>] | System | dims… | Archetype | Scale |
    Vehicle count | (<indicator> per vehicle, <indicator> total)×N.

    ``scenarios`` = ``[(label | None, results)]`` — a single ``None`` entry
    reproduces the static single-scenario sheet EXACTLY (no scenario column);
    N labelled entries add the leading scenario column (shape (a): filterable,
    single sheet). No-op when there are no results.
    """
    if not any(results for _, results in scenarios):
        return
    has_scen = any(lbl is not None for lbl, _ in scenarios)
    ws = wb.create_sheet("By cohort")
    ws.sheet_properties.tabColor = COHORT_TAB_HEX

    header: list[str] = ["Year"]
    if has_scen:
        header.append(scenario_header)
    header += ["System"] + resolver.dim_headers + ["Archetype", "Scale", "Vehicle count"]
    for l, u in zip(labels, units):
        header.append(f"{l} per vehicle ({u})")
        header.append(f"{l} total ({u})")
    ws.append(header)
    style_header(ws)

    for lbl, results in scenarios:
        if not results:
            continue
        years_list = _union_years([(lbl, results)])
        cohort_keys = _union_cohorts([(lbl, results)])
        for y in years_list:
            sc = (sim_counts or {}).get(y, {})
            for ck in cohort_keys:
                count = sc.get(ck, 0.0)
                arc_name = resolver.arc_name(ck) or "—"
                row: list = [y]
                if has_scen:
                    row.append(lbl)
                row += [resolver.system_for(ck)] + resolver.split_cohort(ck) + [
                    arc_name, resolver.arc_scale(ck), count,
                ]
                for r in results:
                    yr = next((v for v in r.years if v.year == y), None)
                    impact = yr.impact_by_cohort.get(ck, 0.0) if yr else 0.0
                    per_v = (impact / count) if count else 0.0
                    row.extend([per_v, impact])
                ws.append(row)

    scen_off = 1 if has_scen else 0
    # Year(1) [+ scenario] + System(1) + dims + Archetype + Scale = count col next.
    count_col = 2 + scen_off + resolver.n_dims + 3
    for row in ws.iter_rows(min_row=2, min_col=count_col, max_col=count_col):
        for cell in row:
            cell.number_format = INT_FMT
    apply_sci(ws, 2, count_col + 1, len(header))
    ws.freeze_panes = f"{get_column_letter(resolver.n_dims + 3 + scen_off)}2"
    autosize(ws)


def write_by_subsystem_sheet(
    wb: Workbook,
    scenarios: list[tuple[str | None, list]],
    resolver: CohortResolver,
    sim_counts: dict[int, dict[str, float]] | None,
    labels: list[str],
    units: list[str],
    *,
    scenario_header: str = "LCI Scenario",
) -> bool:
    """"By subsystem" sheet: subsystem-owned cohorts only. Columns:
    Year [| <scenario_header>] | Subsystem | Dependent archetype |
    BOM archetype | Scale | Unit count | <indicator>×N. Returns True when the
    sheet was written (i.e. at least one subsystem cohort contributed)."""
    if not any(results for _, results in scenarios):
        return False
    all_cohorts = _union_cohorts(scenarios)
    if not any(resolver.is_subsystem(ck) for ck in all_cohorts):
        return False
    has_scen = any(lbl is not None for lbl, _ in scenarios)
    ws = wb.create_sheet("By subsystem")
    ws.sheet_properties.tabColor = COHORT_TAB_HEX

    header: list[str] = ["Year"]
    if has_scen:
        header.append(scenario_header)
    header += ["Subsystem", "Dependent archetype", "BOM archetype", "Scale", "Unit count"]
    header += [f"{l} ({u})" for l, u in zip(labels, units)]
    ws.append(header)
    style_header(ws)

    for lbl, results in scenarios:
        if not results:
            continue
        years_list = _union_years([(lbl, results)])
        sub_cohort_keys = [ck for ck in _union_cohorts([(lbl, results)]) if resolver.is_subsystem(ck)]
        for y in years_list:
            sc = (sim_counts or {}).get(y, {})
            for ck in sub_cohort_keys:
                _pid, rest = resolver.split_prefix(ck)
                count = sc.get(ck, 0.0)
                row: list = [y]
                if has_scen:
                    row.append(lbl)
                row += [
                    resolver.system_for(ck),
                    rest,
                    resolver.arc_name(ck) or "—",
                    resolver.arc_scale(ck),
                    count,
                ]
                for r in results:
                    yr = next((v for v in r.years if v.year == y), None)
                    row.append(yr.impact_by_cohort.get(ck, 0.0) if yr else 0.0)
                ws.append(row)

    scen_off = 1 if has_scen else 0
    unit_col = 6 + scen_off  # Unit count column
    for row in ws.iter_rows(min_row=2, min_col=unit_col, max_col=unit_col):
        for cell in row:
            cell.number_format = INT_FMT
    apply_sci(ws, 2, unit_col + 1, len(header))
    ws.freeze_panes = "C2"
    autosize(ws)
    return True
