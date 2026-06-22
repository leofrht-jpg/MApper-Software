# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""FastAPI router for the DSM dynamic stock module (Phase 2A)."""
from __future__ import annotations

import copy
import datetime
import io
import os
import re
import threading
import uuid
from typing import Optional

import bw2data
from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import Response

from mapper.api.project_guard import verify_project_state

from mapper.core.dsm_engine import (
    DEFAULT_WEIBULL_SCALE,
    DEFAULT_WEIBULL_SHAPE,
    SUPPORTED_EXTS,
    DynamicStockModel,
    aggregate_stock_template_xlsx,
    all_cohort_keys,
    cohort_key_to_dict,
    inflow_template_xlsx,
    non_age_dimensions,
    outflow_template_xlsx,
    parse_aggregate_stock_file,
    parse_label_file,
    parse_inflow_file,
    parse_outflow_file,
    parse_stock_file,
    parse_stock_target_file,
    stock_target_template_xlsx,
    stock_template_xlsx,
    survival_curve,
)
from mapper.core import dsm_storage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from mapper.models.dsm_schemas import (
    BASE_SCENARIO_ID,
    DimensionDef,
    DSMScalingRule,
    DSMScalingRuleList,
    DSMScenario,
    InflowData,
    InflowUploadResult,
    MaterializedDSMState,
    ModeConfig,
    DSMSystemState,
    MultiScenarioSimulationResult,
    OutflowData,
    OutflowUploadResult,
    ScalingRuleSetResult,
    SimulateScenariosRequest,
    SimulationResult,
    SimulationSummary,
    StockTargetData,
    StockTargetUploadResult,
    StockUploadResult,
    SurvivalConfig,
    SurvivalConfigList,
    SurvivalPreviewPoint,
    SurvivalSetResult,
    SystemDefinition,
    SystemSummary,
    SystemUpdateResponse,
    TimeHorizon,
    YearResult,
    get_base_scenario,
    get_scenario,
    materialize_scenario,
)


router = APIRouter(prefix="/dsm", tags=["dsm"])

# In-memory stores, hydrated from disk at startup. All stores are nested by
# bw2 project name: outer key = project, inner key = system id.
_systems: dict[str, dict[str, SystemDefinition]] = {}
_states: dict[str, dict[str, DSMSystemState]] = {}
_results: dict[str, dict[str, SimulationResult]] = {}
# Last multi-scenario run per (project, system_id). Populated by the
# /simulate-scenarios endpoint and consumed by /export so a single download
# contains every scenario × case result. Cleared whenever state mutates so we
# don't export a stale scenario set.
_multi_results: dict[str, dict[str, "MultiScenarioSimulationResult"]] = {}
_lock = threading.Lock()


def _current_project() -> str:
    return bw2data.projects.current


def _proj_systems(project: str | None = None) -> dict[str, SystemDefinition]:
    p = project or _current_project()
    return _systems.setdefault(p, {})


def _proj_states(project: str | None = None) -> dict[str, DSMSystemState]:
    p = project or _current_project()
    return _states.setdefault(p, {})


def _proj_results(project: str | None = None) -> dict[str, SimulationResult]:
    p = project or _current_project()
    return _results.setdefault(p, {})


def _proj_multi_results(project: str | None = None) -> dict[str, "MultiScenarioSimulationResult"]:
    p = project or _current_project()
    return _multi_results.setdefault(p, {})


def hydrate_from_disk() -> None:
    """Load persisted systems/states/results into the in-memory stores."""
    systems, states, results, mappings, archetypes, subsystems, sub_results = (
        dsm_storage.load_all()
    )
    _systems.update(systems)
    _states.update(states)
    _results.update(results)
    # Cohort mappings + archetypes live in bom.py — install them there.
    try:
        from mapper.api import bom as _bom
        if mappings:
            _bom._cohort_mappings.update(mappings)
        if archetypes:
            _bom._archetypes.update(archetypes)
    except Exception:
        pass
    # Subsystems live in subsystems.py — install them there.
    try:
        from mapper.api import subsystems as _subs
        if subsystems:
            _subs._subsystems.update(subsystems)
        if sub_results:
            _subs._subsystem_results.update(sub_results)
    except Exception:
        pass


def _now_iso() -> str:
    return datetime.datetime.now().isoformat()


_FILENAME_UNSAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _sanitize_filename(name: str, fallback: str = "system", max_len: int = 100) -> str:
    """Produce a filesystem-safe base name for downloads.

    Replaces whitespace with underscores, strips non-alphanumeric chars (keeping
    ``._-``), collapses repeats, and trims to ``max_len``.
    """
    cleaned = (name or "").strip().replace(" ", "_")
    cleaned = _FILENAME_UNSAFE.sub("", cleaned).strip("._-")
    if not cleaned:
        return fallback
    return cleaned[:max_len]


def _get_system(system_id: str) -> SystemDefinition:
    sys_def = _proj_systems().get(system_id)
    if not sys_def:
        raise HTTPException(status_code=404, detail=f"System '{system_id}' not found")
    return sys_def


def _get_or_create_state(system_id: str) -> DSMSystemState:
    states = _proj_states()
    state = states.get(system_id)
    if state is None:
        state = DSMSystemState(system_id=system_id)
        states[system_id] = state
    # Ensure a Base scenario always exists — keeps the rest of the code simple.
    if not state.scenarios:
        state.scenarios.append(
            DSMScenario(id=BASE_SCENARIO_ID, name="Base", is_base=True)
        )
    if state.active_scenario_id is None:
        state.active_scenario_id = get_base_scenario(state).id
    return state


def _target_scenario(state: DSMSystemState, scenario_id: str | None) -> DSMScenario:
    """Resolve a write target: explicit id > active scenario > Base.

    Raises 404 when the caller passes an id that doesn't exist.
    """
    if scenario_id is not None:
        try:
            return get_scenario(state, scenario_id)
        except KeyError:
            raise HTTPException(
                status_code=404, detail=f"Scenario '{scenario_id}' not found."
            )
    return get_scenario(state, state.active_scenario_id)


def _slot(scenario: DSMScenario, name: str, default_factory):
    """Return scenario's slot, initializing from default when still inherited.

    Once a slot has been explicitly written we stop inheriting from Base for
    it, so uploading to a non-base scenario creates a proper override.
    """
    val = getattr(scenario, name)
    if val is None:
        val = default_factory()
        setattr(scenario, name, val)
    return val


def _materialized(state: DSMSystemState, scenario_id: str | None = None) -> MaterializedDSMState:
    return materialize_scenario(state, scenario_id)


def _validate_definition(definition: SystemDefinition) -> None:
    if definition.time_horizon.end_year < definition.time_horizon.start_year:
        raise HTTPException(status_code=400, detail="end_year must be ≥ start_year")
    nads = non_age_dimensions(definition.dimensions)
    if not nads:
        raise HTTPException(
            status_code=400,
            detail="At least one non-age dimension is required.",
        )
    seen_names: set[str] = set()
    for d in nads:
        if not d.name:
            raise HTTPException(status_code=400, detail="Every dimension needs a name.")
        if d.name in seen_names:
            raise HTTPException(status_code=400, detail=f"Duplicate dimension name: {d.name}")
        seen_names.add(d.name)
        if not d.labels:
            raise HTTPException(
                status_code=400,
                detail=f"Dimension '{d.name}' must have at least one label.",
            )


def _summary(definition: SystemDefinition) -> SystemSummary:
    return SystemSummary(
        id=definition.id or "",
        name=definition.name,
        description=definition.description,
        time_horizon=definition.time_horizon,
        dimension_count=len(non_age_dimensions(definition.dimensions)),
        cohort_count=len(all_cohort_keys(definition.dimensions)),
        created_at=definition.created_at or _now_iso(),
    )


# ── System CRUD ──────────────────────────────────────────────────────────────


@router.post(
    "/systems",
    response_model=SystemDefinition,
    dependencies=[Depends(verify_project_state)],
)
async def create_system(body: SystemDefinition) -> SystemDefinition:
    _validate_definition(body)
    project = _current_project()
    with _lock:
        sid = str(uuid.uuid4())
        body.id = sid
        body.created_at = _now_iso()
        _proj_systems(project)[sid] = body
        _proj_states(project)[sid] = DSMSystemState(system_id=sid)
    dsm_storage.save_system(project, body)
    dsm_storage.save_state(project, sid, _proj_states(project)[sid])
    return body


@router.get("/systems", response_model=list[SystemSummary])
async def list_systems() -> list[SystemSummary]:
    return [_summary(s) for s in _proj_systems().values()]


@router.get("/systems/{system_id}", response_model=SystemDefinition)
async def get_system(system_id: str) -> SystemDefinition:
    return _get_system(system_id)


def _pair_renames(old_labels: list[str], new_labels: list[str]) -> dict[str, str]:
    """Heuristic rename detection per dimension.

    Labels in the set intersection are unchanged. Labels only in old are "removed";
    labels only in new are "added". When the counts match, pair them by order and
    treat as renames; otherwise report add/remove (no rename pairing).
    """
    common = set(old_labels) & set(new_labels)
    old_only = [l for l in old_labels if l not in common]
    new_only = [l for l in new_labels if l not in common]
    if old_only and len(old_only) == len(new_only):
        return dict(zip(old_only, new_only))
    return {}


def _migrate_state(
    old_def: SystemDefinition,
    new_def: SystemDefinition,
    state: DSMSystemState,
) -> tuple[DSMSystemState, list[str]]:
    """Best-effort migration of DSM state data across a dimension change.

    Strategy:
      - Match non-age dims by machine name.
      - Per-dim heuristic rename pairing when label counts are equal.
      - Orphaned rows (refs to removed labels or removed dims) are dropped + counted.
      - Dim add/remove or reorder → cohort_key format changes, translate via name map;
        if the new key can't be fully resolved, drop that row.
    """
    warnings: list[str] = []
    old_nads = non_age_dimensions(old_def.dimensions)
    new_nads = non_age_dimensions(new_def.dimensions)
    old_by_name = {d.name: d for d in old_nads}
    new_by_name = {d.name: d for d in new_nads}

    removed_dims = [n for n in old_by_name if n not in new_by_name]
    added_dims = [n for n in new_by_name if n not in old_by_name]
    if removed_dims:
        warnings.append(f"Removed dimension(s): {', '.join(removed_dims)}.")
    if added_dims:
        warnings.append(f"Added dimension(s) with no prior data: {', '.join(added_dims)}.")

    # Build per-dim label translation maps (old_label -> new_label or None)
    label_trans: dict[str, dict[str, str | None]] = {}
    for name, old_dim in old_by_name.items():
        if name not in new_by_name:
            continue
        new_dim = new_by_name[name]
        renames = _pair_renames(old_dim.labels, new_dim.labels)
        tmap: dict[str, str | None] = {}
        for lab in old_dim.labels:
            new_lab = renames.get(lab, lab)
            tmap[lab] = new_lab if new_lab in new_dim.labels else None
        label_trans[name] = tmap
        # Surface per-dim warnings
        renamed_n = len(renames)
        dropped = [l for l, nl in tmap.items() if nl is None]
        added_labels = [l for l in new_dim.labels if l not in old_dim.labels and l not in renames.values()]
        if renamed_n:
            pairs = ", ".join(f"'{o}'→'{n}'" for o, n in renames.items())
            warnings.append(f"Renamed label(s) in {name}: {pairs}.")
        if dropped:
            warnings.append(f"Removed label(s) from {name}: {', '.join(dropped)}.")
        if added_labels:
            warnings.append(f"Added label(s) to {name}: {', '.join(added_labels)}.")

    # Added dims need a default label value to extend old keys; pick first label.
    added_defaults: dict[str, str | None] = {}
    for name in added_dims:
        labels = new_by_name[name].labels
        added_defaults[name] = labels[0] if labels else None

    def translate_cohort_key(old_ck: str) -> str | None:
        parts = old_ck.split("|") if old_ck else []
        if len(parts) != len(old_nads):
            return None
        # Map old values into a dict keyed by old dim name
        old_vals: dict[str, str] = {}
        for i, dim in enumerate(old_nads):
            old_vals[dim.name] = parts[i]
        # Build new cohort components in new dim order
        new_parts: list[str] = []
        for new_dim in new_nads:
            if new_dim.name in old_by_name:
                old_label = old_vals.get(new_dim.name)
                if old_label is None:
                    return None
                translated = label_trans.get(new_dim.name, {}).get(old_label)
                if translated is None:
                    return None
                new_parts.append(translated)
            else:
                default = added_defaults.get(new_dim.name)
                if default is None:
                    return None
                new_parts.append(default)
        return "|".join(new_parts)

    horizon_len = new_def.time_horizon.length
    year_range = set(new_def.time_horizon.years)

    # Aggregate per-slot drop counts across all scenarios so we only surface
    # one warning per data kind instead of one per scenario.
    orphaned_stock = 0
    orphaned_inflow = 0
    orphaned_outflows = 0
    orphaned_targets = 0
    orphaned_modes = 0

    for scenario in state.scenarios:
        if scenario.initial_stock is not None:
            new_stock: dict[str, float] = {}
            for full_key, count in scenario.initial_stock.items():
                cohort_part, _, age_part = full_key.rpartition("|")
                try:
                    age = int(age_part)
                except ValueError:
                    orphaned_stock += 1
                    continue
                new_ck = translate_cohort_key(cohort_part)
                if new_ck is None or age >= horizon_len:
                    orphaned_stock += 1
                    continue
                new_stock[f"{new_ck}|{age}"] = new_stock.get(f"{new_ck}|{age}", 0.0) + count
            scenario.initial_stock = new_stock

        if scenario.inflows is not None:
            new_inflows: list[InflowData] = []
            for inflow in scenario.inflows:
                if inflow.year not in year_range:
                    orphaned_inflow += len(inflow.counts)
                    continue
                new_counts: dict[str, float] = {}
                for ck, count in inflow.counts.items():
                    new_ck = translate_cohort_key(ck)
                    if new_ck is None:
                        orphaned_inflow += 1
                        continue
                    new_counts[new_ck] = new_counts.get(new_ck, 0.0) + count
                if new_counts:
                    new_inflows.append(InflowData(year=inflow.year, counts=new_counts))
            scenario.inflows = new_inflows

        if scenario.stock_targets is not None:
            new_targets: list[StockTargetData] = []
            for tgt in scenario.stock_targets:
                if tgt.year not in year_range:
                    orphaned_targets += len(tgt.counts)
                    continue
                new_counts = {}
                for ck, count in tgt.counts.items():
                    new_ck = translate_cohort_key(ck)
                    if new_ck is None:
                        orphaned_targets += 1
                        continue
                    new_counts[new_ck] = new_counts.get(new_ck, 0.0) + count
                if new_counts:
                    new_targets.append(StockTargetData(year=tgt.year, counts=new_counts))
            scenario.stock_targets = new_targets

        if scenario.outflows is not None:
            new_outflows: list[OutflowData] = []
            for of in scenario.outflows:
                if of.year not in year_range:
                    orphaned_outflows += len(of.counts)
                    continue
                new_counts = {}
                for ck, count in of.counts.items():
                    new_ck = translate_cohort_key(ck)
                    if new_ck is None:
                        orphaned_outflows += 1
                        continue
                    new_counts[new_ck] = new_counts.get(new_ck, 0.0) + count
                new_age_counts: dict[str, float] = {}
                for full_key, count in of.cohort_age_counts.items():
                    cohort_part, _, age_part = full_key.rpartition("|")
                    try:
                        age = int(age_part)
                    except ValueError:
                        orphaned_outflows += 1
                        continue
                    new_ck = translate_cohort_key(cohort_part)
                    if new_ck is None or age >= horizon_len:
                        orphaned_outflows += 1
                        continue
                    key = f"{new_ck}|{age}"
                    new_age_counts[key] = new_age_counts.get(key, 0.0) + count
                if new_counts or new_age_counts:
                    new_outflows.append(
                        OutflowData(year=of.year, counts=new_counts, cohort_age_counts=new_age_counts)
                    )
            scenario.outflows = new_outflows

        if scenario.mode_configs is not None:
            new_mode_configs: list[ModeConfig] = []
            for cfg in scenario.mode_configs:
                new_filters = {}
                keep = True
                for dim_name, label in cfg.dimension_filters.items():
                    if dim_name not in new_by_name:
                        keep = False
                        break
                    tmap = label_trans.get(dim_name, {})
                    new_label = tmap.get(label, label if label in new_by_name[dim_name].labels else None)
                    if new_label is None:
                        keep = False
                        break
                    new_filters[dim_name] = new_label
                if keep:
                    cfg.dimension_filters = new_filters
                    new_mode_configs.append(cfg)
                else:
                    orphaned_modes += 1
            scenario.mode_configs = new_mode_configs

    if orphaned_stock:
        warnings.append(f"Dropped {orphaned_stock} orphaned stock row(s) during migration.")
    if orphaned_inflow:
        warnings.append(f"Dropped {orphaned_inflow} orphaned inflow cohort row(s) during migration.")
    if orphaned_targets:
        warnings.append(f"Dropped {orphaned_targets} stock-target row(s) during migration.")
    if orphaned_outflows:
        warnings.append(f"Dropped {orphaned_outflows} manual-outflow row(s) during migration.")
    if orphaned_modes:
        warnings.append(f"Dropped {orphaned_modes} mode config(s) that referenced removed labels/dimensions.")

    # Migrate survival_configs (dim filters)
    new_configs: list[SurvivalConfig] = []
    orphaned_survival = 0
    for cfg in state.survival_configs:
        new_filters: dict[str, str] = {}
        keep = True
        for dim_name, label in cfg.dimension_filters.items():
            if dim_name not in new_by_name:
                keep = False
                break
            tmap = label_trans.get(dim_name, {})
            new_label = tmap.get(label, label if label in new_by_name[dim_name].labels else None)
            if new_label is None:
                keep = False
                break
            new_filters[dim_name] = new_label
        if keep:
            cfg.dimension_filters = new_filters
            new_configs.append(cfg)
        else:
            orphaned_survival += 1
    state.survival_configs = new_configs
    if orphaned_survival:
        warnings.append(f"Dropped {orphaned_survival} survival config(s) that referenced removed labels/dimensions.")

    return state, warnings


@router.put("/systems/{system_id}", response_model=SystemUpdateResponse)
async def update_system(system_id: str, body: SystemDefinition) -> SystemUpdateResponse:
    existing = _get_system(system_id)
    _validate_definition(body)
    project = _current_project()
    with _lock:
        body.id = system_id
        body.created_at = existing.created_at
        state = _get_or_create_state(system_id)
        migrated_state, warnings = _migrate_state(existing, body, state)
        _proj_systems(project)[system_id] = body
        _proj_states(project)[system_id] = migrated_state
        # Dimension/label/horizon changes invalidate prior simulation results.
        _proj_results(project).pop(system_id, None)
        _proj_multi_results(project).pop(system_id, None)
    dsm_storage.save_system(project, body)
    dsm_storage.save_state(project, system_id, migrated_state)
    dsm_storage.clear_results(project, system_id)
    # Dimension/label changes invalidate dependent-subsystem results too.
    try:
        from mapper.api import subsystems as _subs
        _subs.invalidate_results(system_id)
    except Exception:
        pass
    dsm_storage.clear_subsystem_results(project, system_id)
    return SystemUpdateResponse(system=body, warnings=warnings)


@router.delete("/systems/{system_id}")
async def delete_system(system_id: str) -> dict[str, bool]:
    project = _current_project()
    if system_id not in _proj_systems(project):
        raise HTTPException(status_code=404, detail="System not found")
    with _lock:
        _proj_systems(project).pop(system_id, None)
        _proj_states(project).pop(system_id, None)
        _proj_results(project).pop(system_id, None)
        _proj_multi_results(project).pop(system_id, None)
    dsm_storage.delete_system_dir(project, system_id)
    try:
        from mapper.api import bom as _bom
        _bom.purge_system(system_id)
    except Exception:
        pass
    return {"deleted": True}


# ── Scenario CRUD ────────────────────────────────────────────────────────────


_SCENARIO_ID_OK = re.compile(r"^[a-z0-9][a-z0-9_-]{0,62}$")


class ScenarioCreate(BaseModel):
    """Body for creating a new DSM scenario.

    ``id`` is optional; the server derives a url-safe one from ``name`` when
    omitted. ``copy_from`` seeds all data slots from an existing scenario —
    useful for "fork Base and edit the inflows" flows.
    """
    id: str | None = None
    name: str
    description: str | None = None
    copy_from: str | None = None


_CLEARABLE_SLOTS = {
    "initial_stock",
    "inflows",
    "stock_targets",
    "outflows",
    "mode_configs",
    "scaling_rules",
}


class ScenarioUpdate(BaseModel):
    """Partial update for a scenario.

    ``clear_slots`` resets the named data slots to ``None`` so the scenario
    inherits them from Base again. Has no effect on the Base scenario (which
    has nothing to inherit from). Unknown slot names are rejected.
    """
    name: str | None = None
    description: str | None = None
    clear_slots: list[str] | None = None


class ScenarioList(BaseModel):
    scenarios: list[DSMScenario]
    active_scenario_id: str | None


def _slugify_scenario_id(name: str) -> str:
    base = re.sub(r"[^a-z0-9_-]+", "-", (name or "").lower()).strip("-_")
    return (base or "scenario")[:63]


def _unique_scenario_id(state: DSMSystemState, candidate: str) -> str:
    taken = {s.id for s in state.scenarios}
    if candidate not in taken:
        return candidate
    i = 2
    while f"{candidate}-{i}" in taken:
        i += 1
    return f"{candidate}-{i}"


@router.get("/systems/{system_id}/scenarios", response_model=ScenarioList)
async def list_scenarios(system_id: str) -> ScenarioList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    return ScenarioList(
        scenarios=state.scenarios,
        active_scenario_id=state.active_scenario_id,
    )


@router.post("/systems/{system_id}/scenarios", response_model=DSMScenario)
async def create_scenario(system_id: str, body: ScenarioCreate) -> DSMScenario:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    sid = body.id or _slugify_scenario_id(body.name)
    if not _SCENARIO_ID_OK.match(sid):
        raise HTTPException(
            status_code=400,
            detail="Scenario id must be lowercase letters, digits, '-' or '_' (1–63 chars).",
        )
    if sid == BASE_SCENARIO_ID:
        raise HTTPException(
            status_code=400, detail="'base' is reserved for the Base scenario.",
        )
    sid = _unique_scenario_id(state, sid)
    source: DSMScenario | None = None
    if body.copy_from:
        try:
            source = get_scenario(state, body.copy_from)
        except KeyError:
            raise HTTPException(
                status_code=404, detail=f"copy_from scenario '{body.copy_from}' not found.",
            )
    now = _now_iso()
    new = DSMScenario(
        id=sid,
        name=body.name.strip() or sid,
        description=body.description,
        is_base=False,
        created_at=now,
        updated_at=now,
    )
    if source is not None:
        # Deep-copy slot values from the source so edits don't bleed through.
        new.initial_stock = (
            dict(source.initial_stock) if source.initial_stock is not None else None
        )
        new.inflows = (
            [i.model_copy(deep=True) for i in source.inflows]
            if source.inflows is not None else None
        )
        new.stock_targets = (
            [t.model_copy(deep=True) for t in source.stock_targets]
            if source.stock_targets is not None else None
        )
        new.outflows = (
            [o.model_copy(deep=True) for o in source.outflows]
            if source.outflows is not None else None
        )
        new.mode_configs = (
            [c.model_copy(deep=True) for c in source.mode_configs]
            if source.mode_configs is not None else None
        )
        new.scaling_rules = (
            [r.model_copy(deep=True) for r in source.scaling_rules]
            if source.scaling_rules is not None else None
        )
    state.scenarios.append(new)
    dsm_storage.save_state(_current_project(), system_id, state)
    return new


@router.patch("/systems/{system_id}/scenarios/{scenario_id}", response_model=DSMScenario)
async def update_scenario(
    system_id: str, scenario_id: str, body: ScenarioUpdate
) -> DSMScenario:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    try:
        scenario = get_scenario(state, scenario_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Scenario '{scenario_id}' not found.")
    if body.name is not None:
        scenario.name = body.name.strip() or scenario.name
    if body.description is not None:
        scenario.description = body.description or None
    if body.clear_slots:
        unknown = [s for s in body.clear_slots if s not in _CLEARABLE_SLOTS]
        if unknown:
            raise HTTPException(
                status_code=400,
                detail=f"Unknown slot(s): {', '.join(unknown)}. "
                f"Valid: {sorted(_CLEARABLE_SLOTS)}.",
            )
        if not scenario.is_base:
            for slot in body.clear_slots:
                setattr(scenario, slot, None)
    scenario.updated_at = _now_iso()
    dsm_storage.save_state(_current_project(), system_id, state)
    return scenario


@router.delete("/systems/{system_id}/scenarios/{scenario_id}")
async def delete_scenario(system_id: str, scenario_id: str) -> dict[str, str]:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    try:
        scenario = get_scenario(state, scenario_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Scenario '{scenario_id}' not found.")
    if scenario.is_base:
        raise HTTPException(status_code=400, detail="The Base scenario cannot be deleted.")
    state.scenarios = [s for s in state.scenarios if s.id != scenario_id]
    if state.active_scenario_id == scenario_id:
        state.active_scenario_id = get_base_scenario(state).id
    dsm_storage.save_state(_current_project(), system_id, state)
    return {"deleted": scenario_id}


@router.post(
    "/systems/{system_id}/scenarios/{scenario_id}/activate",
    response_model=ScenarioList,
)
async def activate_scenario(system_id: str, scenario_id: str) -> ScenarioList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    try:
        get_scenario(state, scenario_id)
    except KeyError:
        raise HTTPException(status_code=404, detail=f"Scenario '{scenario_id}' not found.")
    state.active_scenario_id = scenario_id
    dsm_storage.save_state(_current_project(), system_id, state)
    return ScenarioList(
        scenarios=state.scenarios,
        active_scenario_id=state.active_scenario_id,
    )


_INHERITABLE_SLOTS = (
    "initial_stock", "inflows", "stock_targets",
    "outflows", "mode_configs", "scaling_rules",
)


@router.post(
    "/systems/{system_id}/scenarios/{new_base_id}/promote-to-base",
    response_model=ScenarioList,
)
async def promote_to_base(system_id: str, new_base_id: str) -> ScenarioList:
    """Promote an existing scenario to Base.

    Before swapping the ``is_base`` flag we walk every other scenario and copy
    the current-Base value into any ``None`` slot — so scenarios that were
    inheriting don't silently switch fallback to the new Base. Inheritance is
    thus *flattened* across the fleet at promotion time; users can re-null
    slots later via ``clear_slots`` if they want to inherit from the new Base.
    """
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    try:
        new_base = get_scenario(state, new_base_id)
    except KeyError:
        raise HTTPException(
            status_code=404, detail=f"Scenario '{new_base_id}' not found.",
        )
    old_base = get_base_scenario(state)
    if new_base.id == old_base.id:
        return ScenarioList(
            scenarios=state.scenarios,
            active_scenario_id=state.active_scenario_id,
        )

    now = _now_iso()
    # Flatten inheritance: every non-target scenario materializes its current
    # effective value (own override → old-Base fallback) into an explicit slot.
    for scenario in state.scenarios:
        if scenario.id == new_base.id:
            continue
        changed = False
        for slot in _INHERITABLE_SLOTS:
            if getattr(scenario, slot) is None:
                setattr(scenario, slot, copy.deepcopy(getattr(old_base, slot)))
                changed = True
        if changed:
            scenario.updated_at = now

    old_base.is_base = False
    old_base.updated_at = now
    new_base.is_base = True
    new_base.updated_at = now

    project = _current_project()
    dsm_storage.save_state(project, system_id, state)
    # Base change invalidates cached simulation results — subsystem aggregates
    # and the multi-scenario export cache must rebuild on next run.
    _proj_results(project).pop(system_id, None)
    _proj_multi_results(project).pop(system_id, None)
    try:
        from mapper.api import subsystems as _subs
        _subs.invalidate_results(system_id)
    except Exception:
        pass
    dsm_storage.clear_subsystem_results(project, system_id)

    return ScenarioList(
        scenarios=state.scenarios,
        active_scenario_id=state.active_scenario_id,
    )


# ── State endpoints ──────────────────────────────────────────────────────────


@router.get("/systems/{system_id}/state", response_model=DSMSystemState)
async def get_state(system_id: str) -> DSMSystemState:
    _get_system(system_id)
    return _get_or_create_state(system_id)


class SystemSettingsPatch(BaseModel):
    """Partial update for per-system state flags that aren't owned by a dedicated endpoint."""
    integer_units: Optional[bool] = None


class SystemSettings(BaseModel):
    integer_units: bool


@router.patch(
    "/systems/{system_id}/settings", response_model=SystemSettings
)
async def patch_settings(system_id: str, body: SystemSettingsPatch) -> SystemSettings:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    if body.integer_units is not None:
        state.integer_units = body.integer_units
        dsm_storage.save_state(_current_project(), system_id, state)
    return SystemSettings(integer_units=state.integer_units)


# ── Label file parser (CSV / XLSX) ───────────────────────────────────────────


@router.post("/parse-labels")
async def parse_labels(
    file: UploadFile = File(...),
    expected_dimension: str = Form(...),
    valid_dimensions: str | None = Form(None),
) -> dict[str, list[str]]:
    """Parse a `{expected_dimension}_labels.{csv,xlsx}` file into labels.

    `valid_dimensions` is an optional comma-separated list of sibling dimension
    names used to improve error messages when the uploaded file is for a
    different (or unrecognized) dimension.
    """
    filename = file.filename or ""
    valid: list[str] | None = None
    if valid_dimensions:
        valid = [s.strip() for s in valid_dimensions.split(",") if s.strip()]
    raw = await file.read()
    try:
        labels = parse_label_file(raw, filename, expected_dimension, valid)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"labels": labels}


# ── Stock upload ─────────────────────────────────────────────────────────────


def _check_ext(filename: str) -> None:
    ext = os.path.splitext(filename or "")[1].lower()
    if ext not in SUPPORTED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.",
        )


@router.post("/systems/{system_id}/stock/upload", response_model=StockUploadResult)
async def upload_stock(
    system_id: str,
    file: UploadFile = File(...),
    scenario_id: str | None = Query(None),
) -> StockUploadResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        parsed, rows = parse_stock_file(raw, filename, sys_def.dimensions)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    scenario.initial_stock = parsed
    dsm_storage.save_state(_current_project(), system_id, state)
    cohort_keys: set[str] = set()
    total = 0.0
    for full_key, count in parsed.items():
        ck = full_key.rsplit("|", 1)[0]
        cohort_keys.add(ck)
        total += count
    return StockUploadResult(
        rows_parsed=rows, cohorts_found=len(cohort_keys), total_items=total
    )


# ── Inflow upload ────────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/inflows/upload", response_model=InflowUploadResult)
async def upload_inflows(
    system_id: str,
    file: UploadFile = File(...),
    scenario_id: str | None = Query(None),
) -> InflowUploadResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        inflows, rows = parse_inflow_file(
            raw, filename, sys_def.dimensions, sys_def.time_horizon.years
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    scenario.inflows = inflows
    dsm_storage.save_state(_current_project(), system_id, state)
    total = sum(sum(inf.counts.values()) for inf in inflows)
    return InflowUploadResult(
        years_parsed=len(inflows), rows_parsed=rows, total_inflows=total
    )


# ── Stock targets upload (Mode B) ────────────────────────────────────────────


@router.post(
    "/systems/{system_id}/stock-targets/upload",
    response_model=StockTargetUploadResult,
)
async def upload_stock_targets(
    system_id: str,
    file: UploadFile = File(...),
    scenario_id: str | None = Query(None),
) -> StockTargetUploadResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        targets, rows = parse_stock_target_file(
            raw, filename, sys_def.dimensions, sys_def.time_horizon.years
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    scenario.stock_targets = targets
    dsm_storage.save_state(_current_project(), system_id, state)
    total = sum(sum(t.counts.values()) for t in targets)
    return StockTargetUploadResult(
        years_parsed=len(targets), rows_parsed=rows, total_targets=total
    )


# ── Outflow upload (manual mode) ─────────────────────────────────────────────


@router.post(
    "/systems/{system_id}/outflows/upload", response_model=OutflowUploadResult
)
async def upload_outflows(
    system_id: str,
    file: UploadFile = File(...),
    scenario_id: str | None = Query(None),
) -> OutflowUploadResult:
    """Upload user-provided annual outflows for manual mode.

    Format mirrors the inflow CSV (``year, <dims>, count``). An optional
    ``age`` or ``birth_year`` column unlocks cohort-specific outflows; when
    absent, the engine allocates FIFO from the oldest cohort.
    """
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        outflows, rows, cohort_specific = parse_outflow_file(
            raw, filename, sys_def.dimensions, sys_def.time_horizon.years
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    scenario.outflows = outflows
    dsm_storage.save_state(_current_project(), system_id, state)
    total = sum(sum(o.counts.values()) for o in outflows)
    return OutflowUploadResult(
        years_parsed=len(outflows),
        rows_parsed=rows,
        total_outflows=total,
        cohort_specific=cohort_specific,
    )


# ── Aggregate-format stock upload (Phase 4) ──────────────────────────────────


@router.post(
    "/systems/{system_id}/stock/upload-aggregate", response_model=StockUploadResult
)
async def upload_stock_aggregate(
    system_id: str,
    file: UploadFile = File(...),
    shape: float = DEFAULT_WEIBULL_SHAPE,
    scale: float = DEFAULT_WEIBULL_SCALE,
    max_age: int = 25,
    scenario_id: str | None = Query(None),
) -> StockUploadResult:
    """Upload aggregate-format initial stock (no ``age`` column).

    Server applies Weibull reverse decomposition to spread each row's total
    across synthetic age cohorts. Used when age-distributed registration data
    is unavailable (e.g., wind-turbine case study).
    """
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    if shape <= 0 or scale <= 0 or max_age <= 0:
        raise HTTPException(
            status_code=400, detail="shape, scale, and max_age must be > 0"
        )
    try:
        parsed, rows = parse_aggregate_stock_file(
            raw, filename, sys_def.dimensions, shape, scale, max_age
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    scenario.initial_stock = parsed
    dsm_storage.save_state(_current_project(), system_id, state)
    cohort_keys: set[str] = set()
    total = 0.0
    for full_key, count in parsed.items():
        ck = full_key.rsplit("|", 1)[0]
        cohort_keys.add(ck)
        total += count
    return StockUploadResult(
        rows_parsed=rows, cohorts_found=len(cohort_keys), total_items=total
    )


# ── Mode configs (Mode B per-cohort mode assignment) ─────────────────────────


class ModeConfigList(BaseModel):
    configs: list[ModeConfig]


class ModeConfigSetResult(BaseModel):
    configs_set: int


@router.get("/systems/{system_id}/mode-configs", response_model=ModeConfigList)
async def get_mode_configs(
    system_id: str,
    scenario_id: str | None = Query(None),
) -> ModeConfigList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    # Reads use the resolved view so callers see Base fallback for inherited slots.
    view = _materialized(state, scenario_id)
    return ModeConfigList(configs=view.mode_configs)


@router.put("/systems/{system_id}/mode-configs", response_model=ModeConfigSetResult)
async def set_mode_configs(
    system_id: str,
    body: ModeConfigList,
    scenario_id: str | None = Query(None),
) -> ModeConfigSetResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    for cfg in body.configs:
        for k, v in cfg.dimension_filters.items():
            if k not in label_index:
                raise HTTPException(
                    status_code=400, detail=f"Unknown dimension '{k}' in mode filter."
                )
            if v not in label_index[k]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Label '{v}' not in dimension '{k}' (allowed: {sorted(label_index[k])}).",
                )
    scenario.mode_configs = body.configs
    dsm_storage.save_state(_current_project(), system_id, state)
    return ModeConfigSetResult(configs_set=len(body.configs))


# ── Scaling rules (parameter-driven scenario scaling) ───────────────────────


def _validate_scaling_rule(
    rule: DSMScalingRule,
    label_index: dict[str, set[str]],
) -> None:
    from mapper.core.parameter_engine import ParameterEngine, ParameterError
    for k, v in rule.dimension_filters.items():
        if k not in label_index:
            raise HTTPException(
                status_code=400, detail=f"Unknown dimension '{k}' in scaling rule filter."
            )
        if v not in label_index[k]:
            raise HTTPException(
                status_code=400,
                detail=f"Label '{v}' not in dimension '{k}' (allowed: {sorted(label_index[k])}).",
            )
    # Probe the expression with placeholder base/year values drawn from the
    # current parameter table so missing-parameter errors surface at create
    # time instead of mid-simulation.
    from mapper.api import parameters as _params
    table = _params._table_for()
    engine = ParameterEngine(table)
    try:
        engine.resolve(rule.expression, extra_vars={"base": 1.0, "year": 0.0})
    except ParameterError as e:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid scaling expression '{rule.expression}': {e}",
        )


@router.get(
    "/systems/{system_id}/scaling-rules", response_model=DSMScalingRuleList
)
async def get_scaling_rules(
    system_id: str,
    scenario_id: str | None = Query(None),
) -> DSMScalingRuleList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    view = _materialized(state, scenario_id)
    return DSMScalingRuleList(rules=view.scaling_rules)


@router.put(
    "/systems/{system_id}/scaling-rules", response_model=ScalingRuleSetResult
)
async def set_scaling_rules(
    system_id: str,
    body: DSMScalingRuleList,
    scenario_id: str | None = Query(None),
) -> ScalingRuleSetResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    seen_ids: set[str] = set()
    for rule in body.rules:
        if rule.id in seen_ids:
            raise HTTPException(
                status_code=400, detail=f"Duplicate scaling rule id: '{rule.id}'",
            )
        seen_ids.add(rule.id)
        _validate_scaling_rule(rule, label_index)
    scenario.scaling_rules = body.rules
    dsm_storage.save_state(_current_project(), system_id, state)
    return ScalingRuleSetResult(rules_set=len(body.rules))


@router.post(
    "/systems/{system_id}/scaling-rules", response_model=DSMScalingRule
)
async def create_scaling_rule(
    system_id: str,
    rule: DSMScalingRule,
    scenario_id: str | None = Query(None),
) -> DSMScalingRule:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    rules = _slot(scenario, "scaling_rules", list)
    if any(r.id == rule.id for r in rules):
        raise HTTPException(
            status_code=400, detail=f"Scaling rule id '{rule.id}' already exists.",
        )
    _validate_scaling_rule(rule, label_index)
    rules.append(rule)
    dsm_storage.save_state(_current_project(), system_id, state)
    return rule


@router.put(
    "/systems/{system_id}/scaling-rules/{rule_id}", response_model=DSMScalingRule
)
async def update_scaling_rule(
    system_id: str,
    rule_id: str,
    rule: DSMScalingRule,
    scenario_id: str | None = Query(None),
) -> DSMScalingRule:
    if rule.id != rule_id:
        raise HTTPException(
            status_code=400,
            detail=f"Path id '{rule_id}' does not match body id '{rule.id}'.",
        )
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    _validate_scaling_rule(rule, label_index)
    rules = scenario.scaling_rules or []
    for i, r in enumerate(rules):
        if r.id == rule_id:
            rules[i] = rule
            scenario.scaling_rules = rules
            dsm_storage.save_state(_current_project(), system_id, state)
            return rule
    raise HTTPException(status_code=404, detail=f"Scaling rule '{rule_id}' not found.")


@router.delete("/systems/{system_id}/scaling-rules/{rule_id}")
async def delete_scaling_rule(
    system_id: str,
    rule_id: str,
    scenario_id: str | None = Query(None),
) -> dict[str, str]:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    scenario = _target_scenario(state, scenario_id)
    rules = scenario.scaling_rules or []
    original = len(rules)
    rules = [r for r in rules if r.id != rule_id]
    if len(rules) == original:
        raise HTTPException(status_code=404, detail=f"Scaling rule '{rule_id}' not found.")
    scenario.scaling_rules = rules
    dsm_storage.save_state(_current_project(), system_id, state)
    return {"deleted": rule_id}


# ── Survival ─────────────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/survival", response_model=SurvivalSetResult)
async def set_survival(system_id: str, body: SurvivalConfigList) -> SurvivalSetResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    for cfg in body.configs:
        for k, v in cfg.dimension_filters.items():
            if k not in label_index:
                raise HTTPException(
                    status_code=400, detail=f"Unknown dimension '{k}' in survival filter."
                )
            if v not in label_index[k]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Label '{v}' not in dimension '{k}' (allowed: {sorted(label_index[k])}).",
                )
    state.survival_configs = body.configs
    dsm_storage.save_state(_current_project(), system_id, state)
    return SurvivalSetResult(configs_set=len(body.configs))


@router.get("/systems/{system_id}/survival", response_model=SurvivalConfigList)
async def get_survival(system_id: str) -> SurvivalConfigList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    return SurvivalConfigList(configs=state.survival_configs)


@router.get(
    "/systems/{system_id}/survival/preview", response_model=list[SurvivalPreviewPoint]
)
async def preview_survival(
    system_id: str,
    shape: float = DEFAULT_WEIBULL_SHAPE,
    scale: float = DEFAULT_WEIBULL_SCALE,
    max_age: Optional[int] = None,
) -> list[SurvivalPreviewPoint]:
    sys_def = _get_system(system_id)
    if shape <= 0 or scale <= 0:
        raise HTTPException(status_code=400, detail="shape and scale must be > 0")
    horizon_len = sys_def.time_horizon.length
    cap = max_age if max_age is not None and max_age > 0 else horizon_len
    return survival_curve(shape, scale, cap)


# ── Simulation ──────────────────────────────────────────────────────────────


def _engine_for_scenario(scenario: str | None):
    """Build a :class:`ParameterEngine` for the current project's table.

    Returns ``None`` when no parameter table exists (scaling rules will be a
    no-op — base data flows through unchanged).
    """
    from mapper.api import parameters as _params
    from mapper.core.parameter_engine import ParameterEngine
    table = _params._table_for()
    if not table.parameters and not table.scenarios:
        return None
    return ParameterEngine(table, scenario=scenario)


def simulate_for_scenario(system_id: str, scenario_id: str | None) -> SimulationResult:
    """Simulate a DSM scenario fresh and return the result *without*
    polluting the cached active-scenario sim or persisted storage.

    Used by the Impact Assessment multi-DSM fan-out (Patch 2E.1) so each
    impact run can compute against an arbitrary scenario without changing
    what the DSM Dashboard considers "active".

    Resolution: ``scenario_id`` is materialized via ``materialize_scenario``;
    when ``None`` the active scenario (or Base) is used. Raises
    ``HTTPException(404)`` when the id doesn't exist on the system.
    """
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    if scenario_id is not None:
        try:
            get_scenario(state, scenario_id)
        except KeyError:
            raise HTTPException(
                status_code=404,
                detail=f"DSM scenario '{scenario_id}' not found on system '{system_id}'.",
            )
    view = _materialized(state, scenario_id)
    try:
        engine = _engine_for_scenario(None) if view.scaling_rules else None
        model = DynamicStockModel(sys_def, view, parameter_engine=engine)
        return model.simulate()
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"DSM scenario simulation failed: {e}"
        )


@router.post("/systems/{system_id}/simulate", response_model=SimulationResult)
async def simulate(
    system_id: str,
    scenario_id: str | None = Query(None),
) -> SimulationResult:
    from mapper.core.compute_metrics import measure_compute
    meter = measure_compute()
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    view = _materialized(state, scenario_id)
    try:
        engine = _engine_for_scenario(None) if view.scaling_rules else None
        model = DynamicStockModel(sys_def, view, parameter_engine=engine)
        result = model.simulate()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Simulation failed: {e}")
    project = _current_project()
    _proj_results(project)[system_id] = result
    # Single-scenario simulate supersedes any prior multi-scenario export cache.
    _proj_multi_results(project).pop(system_id, None)
    dsm_storage.save_results(project, system_id, result)
    # Primary result changed → dependent subsystem results are stale.
    try:
        from mapper.api import subsystems as _subs
        _subs.invalidate_results(system_id)
    except Exception:
        pass
    dsm_storage.clear_subsystem_results(project, system_id)
    result.compute_metrics = meter.build()
    return result


@router.post(
    "/systems/{system_id}/simulate-scenarios",
    response_model=MultiScenarioSimulationResult,
)
async def simulate_scenarios(
    system_id: str, body: SimulateScenariosRequest
) -> MultiScenarioSimulationResult:
    """Run the cross-product of DSM scenarios × sensitivity cases.

    Each (scenario, case) pair materializes the scenario's data with Base
    inheritance, then runs the engine with a ParameterEngine bound to the
    case. Results are keyed ``"{scenario_id}|{case}"``; when only one axis
    varies the key degrades to just that axis for compatibility.

    When the materialized state has no scaling rules, every case produces
    byte-identical output per scenario — we compute once and alias the Base
    result across cases to avoid redundant work.
    """
    from mapper.api import parameters as _params
    from mapper.core.compute_metrics import measure_compute
    from mapper.core.parameter_engine import ParameterEngine

    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)

    scenario_ids = body.scenario_ids or [
        state.active_scenario_id or get_base_scenario(state).id
    ]
    cases = body.cases or ["Base"]

    existing_ids = {s.id for s in state.scenarios}
    unknown_scenarios = [s for s in scenario_ids if s not in existing_ids]
    if unknown_scenarios:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown DSM scenario(s): {', '.join(unknown_scenarios)}. "
            f"Available: {sorted(existing_ids)}.",
        )

    table = _params._table_for()
    valid_cases = set(table.list_scenarios())
    unknown_cases = [c for c in cases if c not in valid_cases]
    if unknown_cases:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown case(s): {', '.join(unknown_cases)}. "
            f"Available: {sorted(valid_cases)}.",
        )

    meter = measure_compute()
    results: dict[str, SimulationResult] = {}
    warnings: list[str] = []

    single_scenario = len(scenario_ids) == 1
    single_case = len(cases) == 1

    def _key(sid: str, case: str) -> str:
        if single_scenario and not single_case:
            return case
        if single_case and not single_scenario:
            return sid
        return f"{sid}|{case}"

    for sid in scenario_ids:
        view = _materialized(state, sid)
        # When no rules reference parameters, every case collapses to Base —
        # compute once and alias.
        if not view.scaling_rules:
            try:
                model = DynamicStockModel(sys_def, view, parameter_engine=None)
                base_result = model.simulate()
            except Exception as e:
                raise HTTPException(
                    status_code=500,
                    detail=f"Simulation failed for scenario '{sid}': {e}",
                )
            for case in cases:
                results[_key(sid, case)] = base_result
            if len(cases) > 1:
                warnings.append(
                    f"Scenario '{sid}' has no scaling rules — all cases produce identical results."
                )
            continue
        for case in cases:
            try:
                engine = ParameterEngine(table, scenario=case)
                model = DynamicStockModel(sys_def, view, parameter_engine=engine)
                results[_key(sid, case)] = model.simulate()
            except Exception as e:
                raise HTTPException(
                    status_code=500,
                    detail=f"Simulation failed for scenario '{sid}', case '{case}': {e}",
                )

    # Persist the first (scenario_id, case) pair as the "primary" single-scenario
    # result so downstream Impact Assessment / AESA stay consistent without
    # needing a scenario-aware storage refactor.
    primary = results[_key(scenario_ids[0], cases[0])]
    project = _current_project()
    _proj_results(project)[system_id] = primary
    dsm_storage.save_results(project, system_id, primary)
    try:
        from mapper.api import subsystems as _subs
        _subs.invalidate_results(system_id)
    except Exception:
        pass
    dsm_storage.clear_subsystem_results(project, system_id)
    primary.compute_metrics = meter.build()
    multi = MultiScenarioSimulationResult(
        system_id=system_id, scenarios=results, warnings=warnings,
    )
    # Cache so /export can include every scenario × case in one workbook.
    _proj_multi_results(project)[system_id] = multi
    return multi


@router.get("/systems/{system_id}/results", response_model=SimulationResult)
async def get_results(system_id: str) -> SimulationResult:
    _get_system(system_id)
    res = _proj_results().get(system_id)
    if res is None:
        raise HTTPException(
            status_code=404, detail="No simulation results yet. Run /simulate first."
        )
    return res


# ── Export ───────────────────────────────────────────────────────────────────


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4B4690")  # blue/purple
_COUNT_FMT = "#,##0.00"
_SHEET_COLORS = {
    "Summary": "4B4690",
    "Stock by Year": "5B7DB1",
    "Inflows by Year": "3E8E7E",
    "Outflows by Year": "A65A5A",
    "Natural Outflows": "B07050",
    "Forced Retirement": "8B3A3A",
    "Stock Targets": "6B5DA1",
    "Age Distribution": "8064A2",
    "Mass Balance": "2F4F7F",
}


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_width(ws, headers: list[str], sample_rows: int = 20) -> None:
    """Set approximate column widths based on header + a sampling of rows."""
    max_row = min(ws.max_row, sample_rows + 1)
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(len(str(header)), 10)
        for row_idx in range(2, max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            width = max(width, min(len(str(v)), 40))
        ws.column_dimensions[letter].width = width + 2


def _apply_count_format(ws, col_indices: list[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        for col_idx in col_indices:
            if col_idx <= len(row):
                row[col_idx - 1].number_format = _COUNT_FMT


# ── Scenario-aware sheet helpers ────────────────────────────────────────────
#
# Every export — single or multi — prefixes every data row with a "Scenario"
# column (and a "Case" column when sensitivity cases are also varied) so the
# output shape is consistent and the workbook round-trips. Each sheet builder
# takes an ``entries`` list of ``(scenario_label, case_label_or_None, …)``
# tuples and appends all rows in one pass.


class _ExportEntry(BaseModel):
    scenario_id: str
    scenario_label: str        # DSMScenario.name, falls back to id
    case_label: str | None     # None when cases axis collapsed to a single default
    scenario: DSMScenario      # raw (for slot-ownership inspection in Summary)
    state: MaterializedDSMState  # already resolved for this scenario
    result: SimulationResult

    model_config = {"arbitrary_types_allowed": True}


def _prefix_headers(include_case: bool) -> list[str]:
    return ["Scenario"] + (["Case"] if include_case else [])


def _write_prefix(ws, row: int, scenario_label: str, case_label: str | None) -> int:
    """Write Scenario (+ optional Case) columns. Return next free col index."""
    ws.cell(row=row, column=1, value=scenario_label)
    if case_label is not None:
        ws.cell(row=row, column=2, value=case_label)
        return 3
    return 2


def _build_summary_sheet(
    ws,
    sys_def: SystemDefinition,
    entries: list[_ExportEntry],
    multi_warnings: list[str],
) -> None:
    """Build the Summary sheet.

    Legacy sections (System meta → Dimensions → Survival configuration →
    Mode configuration) come first so the ``_parse_summary_sheet`` importer
    keeps working on round-trip. Per-scenario totals + config blocks are
    appended afterwards.
    """
    ws.title = "Summary"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Summary"]

    bold = Font(bold=True)
    header12 = Font(bold=True, size=12)
    shared = entries[0].state  # survival/mode live on the state shared by cases.
    integer_units = any(e.state.integer_units for e in entries)
    include_case = any(e.case_label is not None for e in entries)

    # ── Key/value meta (import-system reads these labels) ──────────────────
    rows: list[tuple[str, object]] = [
        ("System name", sys_def.name),
        ("Description", sys_def.description or ""),
        ("Unit name", sys_def.unit_name or "units"),
        ("Start year", sys_def.time_horizon.start_year),
        ("End year", sys_def.time_horizon.end_year),
        ("Horizon length (years)", sys_def.time_horizon.length),
        ("Created at", sys_def.created_at or ""),
        ("Integer units", "yes" if integer_units else "no"),
        ("Scenarios exported", len(entries)),
        ("", ""),
    ]
    r = 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)) and label.startswith(("Total", "Net")):
            ws.cell(row=r, column=2).number_format = _COUNT_FMT
        r += 1

    # ── Dimensions (must appear before Scenario-config blocks for importer) ─
    ws.cell(row=r, column=1, value="Dimensions").font = header12
    r += 1
    for i, h in enumerate(["Name", "Display name", "Labels"], start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1
    for d in sys_def.dimensions:
        ws.cell(row=r, column=1, value=d.name)
        ws.cell(row=r, column=2, value=d.display_name or d.name)
        if d.is_age:
            ws.cell(
                row=r, column=3,
                value=f"auto-generated 0 – {sys_def.time_horizon.length - 1}",
            )
        else:
            ws.cell(row=r, column=3, value=", ".join(d.labels))
        r += 1
    r += 1

    # ── Survival configuration (system-wide, same for every scenario) ──────
    ws.cell(row=r, column=1, value="Survival configuration").font = header12
    r += 1
    surv_headers = ["Filter", "Method", "Weibull shape", "Weibull scale", "Custom points"]
    for i, h in enumerate(surv_headers, start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1
    if not shared.survival_configs:
        ws.cell(row=r, column=1, value="(default)")
        ws.cell(row=r, column=2, value="weibull")
        ws.cell(row=r, column=3, value=DEFAULT_WEIBULL_SHAPE)
        ws.cell(row=r, column=4, value=DEFAULT_WEIBULL_SCALE)
        r += 1
    else:
        for cfg in shared.survival_configs:
            filter_str = (
                ", ".join(f"{k}={v}" for k, v in cfg.dimension_filters.items())
                if cfg.dimension_filters else "(default)"
            )
            ws.cell(row=r, column=1, value=filter_str)
            ws.cell(row=r, column=2, value=cfg.method)
            ws.cell(row=r, column=3, value=cfg.weibull_shape)
            ws.cell(row=r, column=4, value=cfg.weibull_scale)
            if cfg.custom_curve:
                ws.cell(
                    row=r, column=5,
                    value="; ".join(f"{p.age}:{p.survival_rate}" for p in cfg.custom_curve),
                )
            r += 1
    r += 1

    # ── Mode configuration (base scenario only for back-compat) ────────────
    ws.cell(row=r, column=1, value="Mode configuration").font = header12
    r += 1
    ws.cell(row=r, column=1, value="Filter").font = bold
    ws.cell(row=r, column=2, value="Mode").font = bold
    r += 1
    base_entry = next((e for e in entries if e.scenario.is_base), entries[0])
    base_modes = base_entry.state.mode_configs
    if not base_modes:
        ws.cell(row=r, column=1, value="(default)")
        ws.cell(row=r, column=2, value="survival_inflow")
        r += 1
    else:
        for cfg in base_modes:
            filter_str = (
                ", ".join(f"{k}={v}" for k, v in cfg.dimension_filters.items())
                if cfg.dimension_filters else "(default)"
            )
            ws.cell(row=r, column=1, value=filter_str)
            ws.cell(row=r, column=2, value=cfg.mode)
            r += 1
    r += 1

    # ── Per-scenario totals ────────────────────────────────────────────────
    ws.cell(row=r, column=1, value="Per-scenario totals").font = header12
    r += 1
    totals_headers = _prefix_headers(include_case) + [
        "Stock @ start", "Stock @ end",
        "Total inflows", "Total outflows", "Net change", "Warnings",
    ]
    for i, h in enumerate(totals_headers, start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1
    for e in entries:
        col = _write_prefix(ws, r, e.scenario_label, e.case_label)
        s = e.result.summary
        ws.cell(row=r, column=col, value=s.total_stock_start).number_format = _COUNT_FMT; col += 1
        ws.cell(row=r, column=col, value=s.total_stock_end).number_format = _COUNT_FMT; col += 1
        ws.cell(row=r, column=col, value=s.total_inflows).number_format = _COUNT_FMT; col += 1
        ws.cell(row=r, column=col, value=s.total_outflows).number_format = _COUNT_FMT; col += 1
        ws.cell(row=r, column=col, value=s.total_inflows - s.total_outflows).number_format = _COUNT_FMT; col += 1
        ws.cell(row=r, column=col, value="; ".join(s.warnings) if s.warnings else "")
        r += 1
    r += 1

    # ── Per-scenario configuration (slot ownership + rules) ────────────────
    _SLOT_NAMES = (
        "initial_stock", "inflows", "stock_targets",
        "outflows", "mode_configs", "scaling_rules",
    )
    ws.cell(row=r, column=1, value="Scenario configuration").font = header12
    r += 1
    cfg_headers = (
        ["Scenario", "Is base"]
        + [f"Slot: {s}" for s in _SLOT_NAMES]
        + ["Mode configs", "Scaling rules", "Warnings"]
    )
    for i, h in enumerate(cfg_headers, start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1
    seen_sids: set[str] = set()
    for e in entries:
        if e.scenario_id in seen_sids:
            continue
        seen_sids.add(e.scenario_id)
        sc = e.scenario
        col = 1
        ws.cell(row=r, column=col, value=e.scenario_label); col += 1
        ws.cell(row=r, column=col, value="yes" if sc.is_base else "no"); col += 1
        for slot in _SLOT_NAMES:
            raw = getattr(sc, slot, None)
            ws.cell(
                row=r, column=col,
                value="owned" if (sc.is_base or raw is not None) else "inherited",
            )
            col += 1
        modes = "; ".join(
            f"{'/'.join(f'{k}={v}' for k, v in (m.dimension_filters or {}).items()) or '(default)'}"
            f"→{m.mode}"
            for m in e.state.mode_configs
        ) or "(default)"
        ws.cell(row=r, column=col, value=modes); col += 1
        rules = "; ".join(
            f"{rl.name}[{rl.applies_to}]={rl.expression}" for rl in e.state.scaling_rules
        ) or "(none)"
        ws.cell(row=r, column=col, value=rules); col += 1
        warns = "; ".join(e.result.summary.warnings) if e.result.summary.warnings else ""
        ws.cell(row=r, column=col, value=warns)
        r += 1
    r += 1

    # ── Run-level warnings (cross-scenario) ────────────────────────────────
    if multi_warnings:
        ws.cell(row=r, column=1, value="Run warnings").font = header12
        r += 1
        for w in multi_warnings:
            ws.cell(row=r, column=1, value=w)
            r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 36


def _dim_headers(nads) -> list[str]:
    return [d.display_name or d.name for d in nads]


def _cohort_values(cohort_key: str, sys_def: SystemDefinition, nads) -> list[str]:
    d = cohort_key_to_dict(cohort_key, sys_def.dimensions)
    return [d.get(dim.name, "") for dim in nads]


def _build_cohort_year_sheet(
    ws,
    title: str,
    sys_def: SystemDefinition,
    entries: list[_ExportEntry],
    value_key: str,
    value_header: str,
    include_zeros: bool,
    include_case: bool,
) -> None:
    ws.title = title
    ws.sheet_properties.tabColor = _SHEET_COLORS.get(title, "808080")
    nads = non_age_dimensions(sys_def.dimensions)
    dim_headers = _dim_headers(nads)
    headers = _prefix_headers(include_case) + ["Year"] + dim_headers + [value_header]
    _write_header(ws, headers)

    r = 2
    for e in entries:
        for yr in e.result.years:
            data = getattr(yr, value_key)
            for ck, count in data.items():
                if not include_zeros and count == 0:
                    continue
                col = _write_prefix(ws, r, e.scenario_label, e.case_label)
                ws.cell(row=r, column=col, value=yr.year); col += 1
                for v in _cohort_values(ck, sys_def, nads):
                    ws.cell(row=r, column=col, value=v); col += 1
                ws.cell(row=r, column=col, value=float(count)).number_format = _COUNT_FMT
                r += 1

    # Per-scenario yearly totals for the stock sheet — helpful at a glance.
    if value_key == "stock":
        r += 1
        bold = Font(bold=True)
        ws.cell(row=r, column=1, value="Yearly totals").font = bold
        r += 1
        tot_headers = _prefix_headers(include_case) + ["Year", "Total Stock"]
        for i, h in enumerate(tot_headers, start=1):
            ws.cell(row=r, column=i, value=h).font = bold
        r += 1
        for e in entries:
            for yr in e.result.years:
                col = _write_prefix(ws, r, e.scenario_label, e.case_label)
                ws.cell(row=r, column=col, value=yr.year); col += 1
                total = sum(yr.stock.values())
                ws.cell(row=r, column=col, value=total).number_format = _COUNT_FMT
                r += 1

    _auto_width(ws, headers)


def _build_age_sheet(
    ws, sys_def: SystemDefinition, entries: list[_ExportEntry], include_case: bool,
) -> None:
    ws.title = "Age Distribution"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Age Distribution"]
    nads = non_age_dimensions(sys_def.dimensions)
    headers = _prefix_headers(include_case) + ["Year"] + _dim_headers(nads) + ["Age", "Count"]
    _write_header(ws, headers)

    r = 2
    for e in entries:
        for yr in e.result.years:
            for ck, by_age in yr.stock_by_age.items():
                values = _cohort_values(ck, sys_def, nads)
                for age, count in sorted(by_age.items()):
                    if count == 0:
                        continue
                    col = _write_prefix(ws, r, e.scenario_label, e.case_label)
                    ws.cell(row=r, column=col, value=yr.year); col += 1
                    for v in values:
                        ws.cell(row=r, column=col, value=v); col += 1
                    ws.cell(row=r, column=col, value=age); col += 1
                    ws.cell(row=r, column=col, value=float(count)).number_format = _COUNT_FMT
                    r += 1
    _auto_width(ws, headers)


def _build_mass_balance_sheet(
    ws, entries: list[_ExportEntry], include_case: bool,
) -> None:
    ws.title = "Mass Balance"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Mass Balance"]
    headers = _prefix_headers(include_case) + [
        "Year", "Total Stock", "Total Inflow", "Total Outflow", "Net Change",
    ]
    _write_header(ws, headers)
    r = 2
    for e in entries:
        for yr in e.result.years:
            col = _write_prefix(ws, r, e.scenario_label, e.case_label)
            ws.cell(row=r, column=col, value=yr.year); col += 1
            total_stock = sum(yr.stock.values())
            total_in = sum(yr.inflow.values())
            total_out = sum(yr.outflow.values())
            ws.cell(row=r, column=col, value=total_stock).number_format = _COUNT_FMT; col += 1
            ws.cell(row=r, column=col, value=total_in).number_format = _COUNT_FMT; col += 1
            ws.cell(row=r, column=col, value=total_out).number_format = _COUNT_FMT; col += 1
            ws.cell(row=r, column=col, value=total_in - total_out).number_format = _COUNT_FMT
            r += 1
    _auto_width(ws, headers)


def _build_stock_targets_sheet(
    ws, sys_def: SystemDefinition, entries: list[_ExportEntry], include_case: bool,
) -> None:
    ws.title = "Stock Targets"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Stock Targets"]
    nads = non_age_dimensions(sys_def.dimensions)
    headers = _prefix_headers(include_case) + ["Year"] + _dim_headers(nads) + ["Target Stock"]
    _write_header(ws, headers)
    r = 2
    # Targets live on the scenario (shared across cases). Emit once per
    # scenario_id to avoid duplicating the same rows across cases.
    seen_sids: set[str] = set()
    for e in entries:
        if e.scenario_id in seen_sids:
            continue
        seen_sids.add(e.scenario_id)
        for tgt in e.state.stock_targets:
            for ck, count in tgt.counts.items():
                if count == 0:
                    continue
                col = _write_prefix(ws, r, e.scenario_label, e.case_label)
                ws.cell(row=r, column=col, value=tgt.year); col += 1
                for v in _cohort_values(ck, sys_def, nads):
                    ws.cell(row=r, column=col, value=v); col += 1
                ws.cell(row=r, column=col, value=float(count)).number_format = _COUNT_FMT
                r += 1
    _auto_width(ws, headers)


def _has_any_entries(entries: list[_ExportEntry], key: str) -> bool:
    for e in entries:
        for yr in e.result.years:
            data = getattr(yr, key, None) or {}
            if any(v for v in data.values()):
                return True
    return False


def _build_cohort_detail_sheet(
    ws, sys_def: SystemDefinition, entries: list[_ExportEntry], include_case: bool,
) -> None:
    """Year × birth_year × cohort dimensions × surviving count, per scenario.

    ``birth_year = year - age`` is derived from ``stock_by_age``. Complements
    the age-indexed sheet by letting users track individual build-year cohorts
    through time (useful for cohort-based LCA and scrappage analysis).
    """
    ws.title = "Cohort Detail"
    ws.sheet_properties.tabColor = _SHEET_COLORS.get("Cohort Detail", "9a7f5a")
    nads = non_age_dimensions(sys_def.dimensions)
    headers = (
        _prefix_headers(include_case)
        + ["Year", "Birth year"] + _dim_headers(nads) + ["Surviving count"]
    )
    _write_header(ws, headers)
    r = 2
    for e in entries:
        for yr in e.result.years:
            for ck, by_age in yr.stock_by_age.items():
                values = _cohort_values(ck, sys_def, nads)
                for age, count in sorted(by_age.items()):
                    if count == 0:
                        continue
                    col = _write_prefix(ws, r, e.scenario_label, e.case_label)
                    ws.cell(row=r, column=col, value=yr.year); col += 1
                    ws.cell(row=r, column=col, value=yr.year - age); col += 1
                    for v in values:
                        ws.cell(row=r, column=col, value=v); col += 1
                    ws.cell(row=r, column=col, value=float(count)).number_format = _COUNT_FMT
                    r += 1
    _auto_width(ws, headers)


def _build_scaling_rules_sheet(
    ws, entries: list[_ExportEntry],
) -> None:
    """Per-scenario export of scaling rule definitions (expression strings).

    Full per-year resolved factors require a case-bound parameter engine and
    are out of scope here; this lists each rule so users know which
    expressions were attached to which target for each scenario.
    """
    ws.title = "Scaling Rules"
    ws.sheet_properties.tabColor = _SHEET_COLORS.get("Scaling Rules", "7a8fc7")
    headers = ["Scenario", "Rule ID", "Name", "Applies to", "Filter", "Expression"]
    _write_header(ws, headers)
    r = 2
    seen_sids: set[str] = set()
    for e in entries:
        if e.scenario_id in seen_sids:
            continue
        seen_sids.add(e.scenario_id)
        for rule in e.state.scaling_rules:
            ws.cell(row=r, column=1, value=e.scenario_label)
            ws.cell(row=r, column=2, value=rule.id)
            ws.cell(row=r, column=3, value=rule.name)
            ws.cell(row=r, column=4, value=rule.applies_to)
            filter_str = (
                ", ".join(f"{k}={v}" for k, v in rule.dimension_filters.items())
                if rule.dimension_filters else "(all)"
            )
            ws.cell(row=r, column=5, value=filter_str)
            ws.cell(row=r, column=6, value=rule.expression)
            r += 1
    _auto_width(ws, headers)


def _build_export_workbook(
    sys_def: SystemDefinition,
    entries: list[_ExportEntry],
    multi_warnings: list[str] | None = None,
) -> bytes:
    """Build an xlsx workbook for one or more (scenario, case) results.

    Every sheet prefixes rows with a ``Scenario`` column (and ``Case`` when
    sensitivity cases vary) so the format is consistent between single- and
    multi-scenario exports and can round-trip through the importer.
    """
    assert entries, "At least one entry is required"
    include_case = any(e.case_label is not None for e in entries)
    multi_warnings = multi_warnings or []

    wb = Workbook()
    _build_summary_sheet(wb.active, sys_def, entries, multi_warnings)

    _build_cohort_year_sheet(
        wb.create_sheet(), "Stock by Year", sys_def, entries,
        value_key="stock", value_header="Total Stock",
        include_zeros=False, include_case=include_case,
    )
    _build_cohort_year_sheet(
        wb.create_sheet(), "Inflows by Year", sys_def, entries,
        value_key="inflow", value_header="Inflow Count",
        include_zeros=False, include_case=include_case,
    )
    _build_cohort_year_sheet(
        wb.create_sheet(), "Outflows by Year", sys_def, entries,
        value_key="outflow", value_header="Outflow Count",
        include_zeros=False, include_case=include_case,
    )
    # Mode B breakdown sheets — emitted only when populated.
    if _has_any_entries(entries, "natural_outflow"):
        _build_cohort_year_sheet(
            wb.create_sheet(), "Natural Outflows", sys_def, entries,
            value_key="natural_outflow", value_header="Natural Outflow",
            include_zeros=False, include_case=include_case,
        )
    if _has_any_entries(entries, "forced_retirement"):
        _build_cohort_year_sheet(
            wb.create_sheet(), "Forced Retirement", sys_def, entries,
            value_key="forced_retirement", value_header="Forced Retirement",
            include_zeros=False, include_case=include_case,
        )
    if _has_any_entries(entries, "manual_outflow"):
        _build_cohort_year_sheet(
            wb.create_sheet(), "Manual Outflows", sys_def, entries,
            value_key="manual_outflow", value_header="Manual Outflow",
            include_zeros=False, include_case=include_case,
        )
    if any(e.state.stock_targets for e in entries):
        _build_stock_targets_sheet(wb.create_sheet(), sys_def, entries, include_case)
    _build_cohort_detail_sheet(wb.create_sheet(), sys_def, entries, include_case)
    _build_age_sheet(wb.create_sheet(), sys_def, entries, include_case)
    _build_mass_balance_sheet(wb.create_sheet(), entries, include_case)
    if any(e.state.scaling_rules for e in entries):
        _build_scaling_rules_sheet(wb.create_sheet(), entries)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _decompose_multi_keys(
    keys: list[str], state: DSMSystemState,
) -> list[tuple[str, str, str | None]]:
    """Split ``MultiScenarioSimulationResult`` keys back to (key, sid, case).

    The engine encodes keys as ``"{sid}|{case}"`` when both axes vary, or as
    just the varying axis when the other collapses to a single default. This
    walker inverts that: if ``|`` is present → both; else if the value names a
    known scenario → sid-only axis (case_label = None); else → case-only axis
    (sid = active scenario).
    """
    scenario_ids = {s.id for s in state.scenarios}
    default_sid = state.active_scenario_id or (
        next(iter(scenario_ids), None) if scenario_ids else None
    )
    out: list[tuple[str, str, str | None]] = []
    for k in keys:
        if "|" in k:
            sid, case = k.split("|", 1)
            out.append((k, sid, case))
        elif k in scenario_ids:
            out.append((k, k, None))
        else:
            out.append((k, default_sid or k, k))
    return out


def _entries_from_multi(
    sys_def: SystemDefinition,
    state: DSMSystemState,
    multi: MultiScenarioSimulationResult,
) -> list[_ExportEntry]:
    keys = list(multi.scenarios.keys())
    decoded = _decompose_multi_keys(keys, state)
    by_id = {s.id: s for s in state.scenarios}
    entries: list[_ExportEntry] = []
    for key, sid, case in decoded:
        scenario = by_id.get(sid)
        if scenario is None:
            # Defensive: if the scenario was renamed/removed between run and
            # export, synthesize a stub so we still emit every result.
            scenario = DSMScenario(id=sid, name=sid, is_base=False)
        entries.append(_ExportEntry(
            scenario_id=sid,
            scenario_label=scenario.name or sid,
            case_label=case,
            scenario=scenario,
            state=_materialized(state, sid),
            result=multi.scenarios[key],
        ))
    return entries


@router.get("/systems/{system_id}/export")
async def export_results(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    project = _current_project()

    multi = _proj_multi_results(project).get(system_id)
    if multi is not None and multi.scenarios:
        entries = _entries_from_multi(sys_def, state, multi)
        content = _build_export_workbook(sys_def, entries, multi.warnings)
        n_scenarios = len({e.scenario_id for e in entries})
        n_keys = len(entries)
        today = datetime.date.today().isoformat()
        # Scenario count drives the filename; when cases also vary we still
        # pluralize on scenarios (the dominant axis for the user).
        label = f"{n_keys}scenarios" if n_keys > 1 else "1scenario"
        fname = f"mapper_dsm_export_{label}_{today}.xlsx"
        if n_scenarios == 1 and n_keys == 1:
            # Single-scenario single-case: keep the friendlier legacy name.
            fname = f"{_sanitize_filename(sys_def.name)}_simulation.xlsx"
    else:
        result = _proj_results(project).get(system_id)
        if result is None:
            raise HTTPException(
                status_code=400,
                detail="No simulation results available. Run /simulate first.",
            )
        active = _target_scenario(state, state.active_scenario_id)
        entry = _ExportEntry(
            scenario_id=active.id,
            scenario_label=active.name or "Base",
            case_label=None,
            scenario=active,
            state=_materialized(state),
            result=result,
        )
        content = _build_export_workbook(sys_def, [entry], [])
        fname = f"{_sanitize_filename(sys_def.name)}_simulation.xlsx"

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


def _cohort_export_rows(
    sys_def: SystemDefinition, result: SimulationResult,
) -> tuple[list[str], list[list[object]]]:
    """Long-format flatten of a DSM result: ONE row per (year × cohort) across
    ALL years and ALL cohorts. Columns: Year, <non-age dims, display name>,
    Stock, Inflow, Outflow, Net. Within each year, cohorts are sorted by Stock
    descending (mirrors the dashboard "Cohorts in {year}" table); years
    ascending. Net = Inflow − Outflow. Pure (no I/O) → unit-testable."""
    dims = non_age_dimensions(sys_def.dimensions)
    headers = ["Year"] + [d.display_name or d.name for d in dims] + ["Stock", "Inflow", "Outflow", "Net"]
    out: list[list[object]] = []
    for yr in sorted(result.years, key=lambda y: y.year):
        cks = set(yr.stock) | set(yr.inflow) | set(yr.outflow)
        triples: list[tuple[float, list[object]]] = []
        for ck in cks:
            dvals = cohort_key_to_dict(ck, sys_def.dimensions)
            stock = yr.stock.get(ck, 0.0)
            inflow = yr.inflow.get(ck, 0.0)
            outflow = yr.outflow.get(ck, 0.0)
            row: list[object] = [yr.year] + [dvals.get(d.name, "") for d in dims]
            row += [stock, inflow, outflow, inflow - outflow]
            triples.append((stock, row))
        triples.sort(key=lambda t: t[0], reverse=True)
        out.extend(row for _, row in triples)
    return headers, out


@router.get("/systems/{system_id}/cohorts/export")
async def export_cohorts(system_id: str) -> Response:
    """Export the FULL per-year cohort data ("Cohorts in {year}" box, all years)
    to xlsx — one long-format sheet, one row per (year × cohort), ALL cohorts ×
    ALL years, columns: Year + per-dimension + Stock / Inflow / Outflow / Net.
    Reads the active simulation result (run /simulate first)."""
    sys_def = _get_system(system_id)
    project = _current_project()
    result = _proj_results(project).get(system_id)
    if result is None:
        raise HTTPException(
            status_code=400,
            detail="No simulation results available. Run /simulate first.",
        )
    headers, rows = _cohort_export_rows(sys_def, result)
    wb = Workbook()
    ws = wb.active
    ws.title = "Cohorts by Year"
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    years = [y.year for y in result.years]
    span = f"{min(years)}-{max(years)}" if years else "all"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cohorts_{span}.xlsx"'},
    )


# ── Import from previously exported Excel ───────────────────────────────────


def _dim_header_to_name_map(sys_def: SystemDefinition) -> dict[str, str]:
    """Map both display_name and machine name → machine name."""
    m: dict[str, str] = {}
    for d in non_age_dimensions(sys_def.dimensions):
        m[d.name] = d.name
        if d.display_name:
            m[d.display_name] = d.name
    return m


def _build_cohort_key(row_values: dict[str, str], sys_def: SystemDefinition) -> str | None:
    nads = non_age_dimensions(sys_def.dimensions)
    parts: list[str] = []
    for d in nads:
        v = row_values.get(d.name)
        if v is None or v == "":
            return None
        parts.append(str(v))
    return "|".join(parts)


def _locate_year_col(headers: list[str]) -> int:
    """Return the 0-based index of the ``Year`` column.

    The new export prefixes every data row with ``Scenario`` (+ optional
    ``Case``); older workbooks had ``Year`` in column A. This finds whichever
    layout the file uses. Raises ``ValueError`` when absent.
    """
    for i, h in enumerate(headers):
        if h == "Year":
            return i
    raise ValueError("No 'Year' column found in sheet header.")


def _pick_import_scenario(ws, scenario_col: int | None) -> str | None:
    """For multi-scenario workbooks, pick the scenario whose rows we restore.

    Prefers a scenario literally named ``Base``; otherwise picks the first
    distinct scenario label encountered. Returns ``None`` when the sheet has
    no scenario column (legacy single-scenario export).
    """
    if scenario_col is None:
        return None
    labels: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            continue
        if row is None or all(c is None or c == "" for c in row):
            break
        if scenario_col < len(row) and row[scenario_col] is not None:
            lbl = str(row[scenario_col])
            if lbl and lbl not in labels:
                labels.append(lbl)
    for lbl in labels:
        if lbl.lower() == "base":
            return lbl
    return labels[0] if labels else None


def _parse_sheet_rows(ws, sys_def: SystemDefinition) -> list[tuple[int, str, float]]:
    """Parse a "X by Year" sheet into [(year, cohort_key, value)] rows.

    Handles both legacy (Year in col A) and scenario-prefixed exports
    (Scenario/Case cols precede Year). For multi-scenario workbooks only the
    first distinct scenario's rows are restored — the SimulationResult store
    holds a single scenario.
    """
    rows: list[tuple[int, str, float]] = []
    header_map = _dim_header_to_name_map(sys_def)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return rows
    headers = [str(c) if c is not None else "" for c in all_rows[0]]
    try:
        year_col = _locate_year_col(headers)
    except ValueError:
        return rows
    scenario_col = 0 if year_col > 0 else None
    keep_scenario = _pick_import_scenario(ws, scenario_col)

    for row in all_rows[1:]:
        if row is None or all(c is None or c == "" for c in row):
            break
        if scenario_col is not None and keep_scenario is not None:
            cell = row[scenario_col] if scenario_col < len(row) else None
            if cell is None or str(cell) != keep_scenario:
                continue
        if year_col >= len(row) or row[year_col] is None:
            break
        try:
            year = int(row[year_col])
        except (TypeError, ValueError):
            break
        value_cell = row[-1]
        try:
            value = float(value_cell) if value_cell is not None else 0.0
        except (TypeError, ValueError):
            continue
        vals: dict[str, str] = {}
        # Dimension columns sit strictly between Year and the final value col.
        for i, header in enumerate(headers[year_col + 1:-1], start=year_col + 1):
            machine = header_map.get(header, header)
            cell = row[i] if i < len(row) else None
            vals[machine] = "" if cell is None else str(cell)
        ck = _build_cohort_key(vals, sys_def)
        if ck is None:
            continue
        rows.append((year, ck, value))
    return rows


def _parse_age_sheet(ws, sys_def: SystemDefinition) -> list[tuple[int, str, int, float]]:
    """Parse the Age Distribution sheet into [(year, cohort_key, age, count)].

    Scenario/Case prefix-aware: see :func:`_parse_sheet_rows`.
    """
    rows: list[tuple[int, str, int, float]] = []
    header_map = _dim_header_to_name_map(sys_def)
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return rows
    headers = [str(c) if c is not None else "" for c in all_rows[0]]
    try:
        year_col = _locate_year_col(headers)
    except ValueError:
        return rows
    scenario_col = 0 if year_col > 0 else None
    keep_scenario = _pick_import_scenario(ws, scenario_col)

    for row in all_rows[1:]:
        if row is None or all(c is None or c == "" for c in row):
            break
        if scenario_col is not None and keep_scenario is not None:
            cell = row[scenario_col] if scenario_col < len(row) else None
            if cell is None or str(cell) != keep_scenario:
                continue
        try:
            year = int(row[year_col])
            age = int(row[-2])
            count = float(row[-1])
        except (TypeError, ValueError, IndexError):
            continue
        vals: dict[str, str] = {}
        # Dimension columns are Year+1 .. -2 (excluding Age, Count).
        for i, header in enumerate(headers[year_col + 1:-2], start=year_col + 1):
            machine = header_map.get(header, header)
            cell = row[i] if i < len(row) else None
            vals[machine] = "" if cell is None else str(cell)
        ck = _build_cohort_key(vals, sys_def)
        if ck is None:
            continue
        rows.append((year, ck, age, count))
    return rows


def _reconstruct_simulation(
    sys_def: SystemDefinition, wb
) -> tuple[SimulationResult, list[str], list[str]]:
    """Parse an exported workbook back into a SimulationResult.

    Returns ``(result, warnings, cohorts)``.
    """
    warnings: list[str] = []

    def _read(title: str) -> list[tuple[int, str, float]]:
        if title not in wb.sheetnames:
            warnings.append(f"Missing sheet: {title}")
            return []
        return _parse_sheet_rows(wb[title], sys_def)

    stock_rows = _read("Stock by Year")
    inflow_rows = _read("Inflows by Year")
    outflow_rows = _read("Outflows by Year")
    age_rows: list[tuple[int, str, int, float]] = []
    if "Age Distribution" in wb.sheetnames:
        age_rows = _parse_age_sheet(wb["Age Distribution"], sys_def)
    else:
        warnings.append("Missing sheet: Age Distribution")

    def _by_year(rows: list[tuple[int, str, float]]) -> dict[int, dict[str, float]]:
        out: dict[int, dict[str, float]] = {}
        for year, ck, val in rows:
            out.setdefault(year, {})[ck] = val
        return out

    stock_by_year = _by_year(stock_rows)
    inflow_by_year = _by_year(inflow_rows)
    outflow_by_year = _by_year(outflow_rows)

    stock_by_age_by_year: dict[int, dict[str, dict[int, float]]] = {}
    for year, ck, age, cnt in age_rows:
        stock_by_age_by_year.setdefault(year, {}).setdefault(ck, {})[age] = cnt

    all_years = sorted(set(stock_by_year) | set(inflow_by_year) | set(outflow_by_year))
    if not all_years:
        raise HTTPException(status_code=400, detail="No simulation rows found in workbook.")

    cohorts: set[str] = set()
    year_results: list[YearResult] = []
    for y in all_years:
        stock = stock_by_year.get(y, {})
        inflow = inflow_by_year.get(y, {})
        outflow = outflow_by_year.get(y, {})
        stock_by_age = stock_by_age_by_year.get(y, {})
        cohorts.update(stock)
        cohorts.update(inflow)
        cohorts.update(outflow)
        year_results.append(
            YearResult(
                year=y,
                stock=stock,
                stock_by_age=stock_by_age,
                inflow=inflow,
                outflow=outflow,
                outflow_by_age={},
            )
        )

    total_in = sum(sum(r.inflow.values()) for r in year_results)
    total_out = sum(sum(r.outflow.values()) for r in year_results)
    summary = SimulationSummary(
        total_stock_start=sum(year_results[0].stock.values()),
        total_stock_end=sum(year_results[-1].stock.values()),
        total_inflows=total_in,
        total_outflows=total_out,
    )

    sid = sys_def.id or ""
    return (
        SimulationResult(system_id=sid, years=year_results, summary=summary),
        warnings,
        sorted(cohorts),
    )


class ImportResult(BaseModel):
    years_imported: int
    cohorts_found: int
    warnings: list[str] = []


@router.post("/systems/{system_id}/import-simulation", response_model=ImportResult)
async def import_simulation(system_id: str, file: UploadFile = File(...)) -> ImportResult:
    """Restore a simulation result from a previously exported .xlsx for an existing system."""
    sys_def = _get_system(system_id)
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file exported from MApper.")
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    result, warnings, cohorts = _reconstruct_simulation(sys_def, wb)
    result.system_id = system_id

    project = _current_project()
    with _lock:
        _proj_results(project)[system_id] = result
    dsm_storage.save_results(project, system_id, result)

    return ImportResult(
        years_imported=len(result.years),
        cohorts_found=len(cohorts),
        warnings=warnings,
    )


def _parse_summary_sheet(ws) -> dict[str, object]:
    """Extract key-value pairs from the Summary sheet."""
    kv: dict[str, object] = {}
    dims: list[DimensionDef] = []
    survival: list[SurvivalConfig] = []

    rows = list(ws.iter_rows(values_only=True))
    # key-value section is everything until row "Dimensions"
    i = 0
    while i < len(rows):
        r = rows[i]
        if not r:
            i += 1
            continue
        label = r[0]
        if label == "Dimensions":
            i += 1
            break
        if label is None or label == "":
            i += 1
            continue
        kv[str(label)] = r[1] if len(r) > 1 else None
        i += 1

    # dimensions table: rows with header [Name, Display name, Labels] then data rows
    # skip header row
    if i < len(rows):
        i += 1
    while i < len(rows):
        r = rows[i]
        if not r or r[0] is None or r[0] == "" or r[0] == "Survival configuration":
            break
        name = str(r[0])
        display = str(r[1]) if len(r) > 1 and r[1] is not None else name
        labels_cell = r[2] if len(r) > 2 else ""
        labels_str = str(labels_cell) if labels_cell is not None else ""
        is_age = labels_str.startswith("auto-generated")
        labels = [] if is_age else [l.strip() for l in labels_str.split(",") if l.strip()]
        dims.append(DimensionDef(name=name, display_name=display, labels=labels, is_age=is_age))
        i += 1

    # Skip to survival section header
    while i < len(rows) and (not rows[i] or rows[i][0] != "Survival configuration"):
        i += 1
    if i < len(rows):
        i += 2  # skip section title + header row
    while i < len(rows):
        r = rows[i]
        if not r or r[0] is None or r[0] == "":
            break
        filter_str = str(r[0])
        method = str(r[1]) if len(r) > 1 and r[1] else "weibull"
        shape = r[2] if len(r) > 2 else None
        scale = r[3] if len(r) > 3 else None
        filters: dict[str, str] = {}
        if filter_str and filter_str != "(default)":
            for part in filter_str.split(","):
                if "=" in part:
                    k, v = part.split("=", 1)
                    filters[k.strip()] = v.strip()
        survival.append(
            SurvivalConfig(
                dimension_filters=filters,
                method=method,
                weibull_shape=float(shape) if isinstance(shape, (int, float)) else None,
                weibull_scale=float(scale) if isinstance(scale, (int, float)) else None,
            )
        )
        i += 1

    kv["__dimensions__"] = dims
    kv["__survival__"] = survival
    return kv


@router.post("/import-system", response_model=SystemDefinition)
async def import_system(file: UploadFile = File(...)) -> SystemDefinition:
    """Create a brand new system (with state + results) from an exported workbook."""
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file exported from MApper.")
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    if "Summary" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook missing required Summary sheet.")
    meta = _parse_summary_sheet(wb["Summary"])

    name = str(meta.get("System name") or "Imported system")
    description = str(meta.get("Description") or "") or None
    try:
        start_year = int(meta.get("Start year"))  # type: ignore[arg-type]
        end_year = int(meta.get("End year"))  # type: ignore[arg-type]
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Summary sheet missing valid Start/End year.")

    dims: list[DimensionDef] = meta.get("__dimensions__") or []  # type: ignore[assignment]
    if not [d for d in dims if not d.is_age]:
        raise HTTPException(status_code=400, detail="No dimensions found in Summary sheet.")
    survival_configs: list[SurvivalConfig] = meta.get("__survival__") or []  # type: ignore[assignment]

    # Ensure age dim is present.
    if not any(d.is_age for d in dims):
        dims.append(DimensionDef(name="age", display_name="Age", labels=[], is_age=True))

    new_def = SystemDefinition(
        id=None,
        name=name,
        description=description,
        time_horizon=TimeHorizon(start_year=start_year, end_year=end_year),
        dimensions=dims,
    )
    _validate_definition(new_def)

    project = _current_project()
    with _lock:
        sid = str(uuid.uuid4())
        new_def.id = sid
        new_def.created_at = _now_iso()
        _proj_systems(project)[sid] = new_def
        state = DSMSystemState(
            system_id=sid,
            survival_configs=survival_configs,
            scenarios=[DSMScenario(id=BASE_SCENARIO_ID, name="Base", is_base=True)],
            active_scenario_id=BASE_SCENARIO_ID,
        )
        _proj_states(project)[sid] = state

    # Reconstruct results using the workbook
    result, warnings, _cohorts = _reconstruct_simulation(new_def, wb)
    result.system_id = sid
    base = get_base_scenario(state)

    # Rebuild initial_stock from year=start_year age distribution
    start_yr_result = next((y for y in result.years if y.year == start_year), None)
    if start_yr_result is not None:
        initial: dict[str, float] = {}
        for ck, by_age in start_yr_result.stock_by_age.items():
            for age, cnt in by_age.items():
                initial[f"{ck}|{age}"] = cnt
        base.initial_stock = initial

    # Rebuild inflows
    inflows: list[InflowData] = []
    for yr in result.years:
        if yr.inflow:
            inflows.append(InflowData(year=yr.year, counts=dict(yr.inflow)))
    base.inflows = inflows

    with _lock:
        _proj_states(project)[sid] = state
        _proj_results(project)[sid] = result

    dsm_storage.save_system(project, new_def)
    dsm_storage.save_state(project, sid, state)
    dsm_storage.save_results(project, sid, result)

    return new_def


# ── CSV templates ───────────────────────────────────────────────────────────


_XLSX_MEDIA_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


@router.post("/systems/{system_id}/templates/stock")
async def template_stock(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    data = stock_template_xlsx(sys_def.dimensions)
    fname = f"stock_template_{_sanitize_filename(sys_def.name)}.xlsx"
    return Response(
        content=data,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/systems/{system_id}/templates/inflows")
async def template_inflows(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    data = inflow_template_xlsx(sys_def.dimensions, sys_def.time_horizon.years)
    fname = f"inflow_template_{_sanitize_filename(sys_def.name)}.xlsx"
    return Response(
        content=data,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/systems/{system_id}/templates/stock-targets")
async def template_stock_targets(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    data = stock_target_template_xlsx(sys_def.dimensions, sys_def.time_horizon.years)
    fname = f"stock_target_template_{_sanitize_filename(sys_def.name)}.xlsx"
    return Response(
        content=data,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/systems/{system_id}/templates/outflows")
async def template_outflows(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    data = outflow_template_xlsx(sys_def.dimensions, sys_def.time_horizon.years)
    fname = f"outflow_template_{_sanitize_filename(sys_def.name)}.xlsx"
    return Response(
        content=data,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/systems/{system_id}/templates/stock-aggregate")
async def template_stock_aggregate(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    data = aggregate_stock_template_xlsx(sys_def.dimensions)
    fname = f"stock_aggregate_template_{_sanitize_filename(sys_def.name)}.xlsx"
    return Response(
        content=data,
        media_type=_XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
