"""Unified Impact Assessment endpoints.

Single entry point that can run DSM × LCA in two modes:
- ``static``: one LCI for the whole horizon (existing DSMLCAPipeline).
  BOMs may still evolve year-to-year via MaterialEvolution.
- ``projected``: year-matched prospective databases from premise-generated
  scenarios (ProjectedDSMLCAPipeline).

Runs as a background task so long calculations (projected × many years ×
many methods) don't block the API. Progress is streamed over a WebSocket.
"""
from __future__ import annotations

import asyncio
import io
import logging
import threading
import time
import uuid
from typing import Any

import bw2data
from fastapi import APIRouter, HTTPException, Response, WebSocket, WebSocketDisconnect

from mapper.api import tasks as task_registry
from mapper.api.tasks import CancelledOperation
from mapper.models.schemas import (
    ArchetypeLCACalculateResult,
    CancelledTaskResponse,
    MultiProductExportRequest,
    MultiProductLCAResult,
    SingleProductComparisonExportRequest,
    SingleProductProspectiveExportRequest,
    SingleProductProspectiveRunPayload,
    SingleProductStaticExportRequest,
)
from mapper.api.bom import (
    _build_mfa_lca_workbook,
    _current_project,
    _proj_archetypes,
    _proj_cohort_mappings,
    _proj_results,
    _sanitize_filename,
)
from mapper.api.dsm import _get_or_create_state, _get_system, simulate_for_scenario
from mapper.models.dsm_schemas import BASE_SCENARIO_ID, get_scenario
from mapper.core import plca_storage
from mapper.core.bom_engine import iter_all_materials, validation_error_count
from mapper.core.bw2_wrapper import PersistentLCARunner, run_lca_multi_method
from mapper.core.dsm_lca_engine import (
    DSMLCAPipeline,
    ProjectedDSMLCAPipeline,
    aggregate_subsystem_results,
    build_subsystem_cohort_mapping,
    resolve_database_for_year,
)
from mapper.core.subsystem_engine import compute_dependent_subsystem
from mapper.models.bom_schemas import (
    ImpactAssessmentMeta,
    ImpactAssessmentRequest,
    ImpactAssessmentResult,
    ImpactComparePoint,
    ImpactCompareMethodResult,
    ImpactCompareRequest,
    ImpactCompareResult,
    ImpactExportRequest,
    DSMLCAResult,
    DSMScenarioImpactResult,
    MultiDSMImpactResult,
    MultiParamImpactResult,
    MultiPairedImpactResult,
    MultiScenarioImpactResult,
    MultiScenarioProjectedImpactResult,
    PairedScenarioImpactResult,
    ParamScenarioImpactResult,
    ProspectiveScenarioRef,
    ScenarioImpactResult,
    ScenarioProjectedResult,
)
from openpyxl.styles import Alignment, Font, PatternFill


router = APIRouter(prefix="/impact", tags=["impact"])


# ── Task registry (mirrors plca.py) ───────────────────────────────────────────


class _TaskState:
    def __init__(self) -> None:
        self.stage: str = "queued"
        self.pct: float = 0.0
        self.done: bool = False
        self.error: str | None = None
        # Single-scenario projected/static runs store ``ImpactAssessmentResult``;
        # multi-scenario projected runs store ``MultiScenarioProjectedImpactResult``
        # under the same task_id.
        self.result: ImpactAssessmentResult | MultiScenarioProjectedImpactResult | None = None
        self.subscribers: list[asyncio.Queue] = []
        self.cancelled: bool = False


_TASKS: dict[str, _TaskState] = {}
_TASK_LOCK = threading.Lock()


def _notify_all(task: _TaskState, payload: dict[str, Any]) -> None:
    for q in list(task.subscribers):
        try:
            q.put_nowait(payload)
        except asyncio.QueueFull:
            pass


# ── Helpers ───────────────────────────────────────────────────────────────────


def _resolve_prospective_dbs(project: str, scenario) -> list[tuple[str, int]]:
    """Return [(db_name, year), ...] for every registered DB that matches the
    given base/iam/ssp triple."""
    log = logging.getLogger(__name__)
    registry = plca_storage.load_registry(project)
    iam = scenario.iam.lower()
    existing = set(bw2data.databases)
    log.info(
        "pLCA resolve: project=%s, scenario={base_db=%r, iam=%r, ssp=%r}, "
        "registry_entries=%d, bw2_databases=%d",
        project, scenario.base_db, scenario.iam, scenario.ssp,
        len(registry), len(existing),
    )
    log.debug("  bw2data.databases: %s", sorted(existing))
    log.debug("  registry raw: %s", registry)

    out: list[tuple[str, int]] = []
    rejected: list[tuple[str, str]] = []  # (name, reason)
    for entry in registry:
        name = entry.get("name") or "?"
        if entry.get("base_db") != scenario.base_db:
            rejected.append((name, f"base_db={entry.get('base_db')!r} != {scenario.base_db!r}"))
            continue
        if (entry.get("iam") or "").lower() != iam:
            rejected.append((name, f"iam={entry.get('iam')!r} != {iam!r}"))
            continue
        if entry.get("ssp") != scenario.ssp:
            rejected.append((name, f"ssp={entry.get('ssp')!r} != {scenario.ssp!r}"))
            continue
        if not name or name not in existing:
            rejected.append((name, "name not in bw2data.databases (was DB deleted?)"))
            continue
        try:
            out.append((name, int(entry.get("year"))))
        except (TypeError, ValueError):
            rejected.append((name, f"bad year={entry.get('year')!r}"))
            continue
    out.sort(key=lambda p: p[1])
    log.info("pLCA resolve: matched %d DB(s): %s", len(out), out)
    if rejected:
        log.info("pLCA resolve: rejected %d entries: %s", len(rejected), rejected)
    return out


def _year_to_database_map(
    sim_years: list[int],
    prospective_dbs: list[tuple[str, int]],
    year_start: int | None,
    year_end: int | None,
) -> dict[int, str]:
    out: dict[int, str] = {}
    for y in sim_years:
        if year_start is not None and y < year_start:
            continue
        if year_end is not None and y > year_end:
            continue
        match = resolve_database_for_year(y, prospective_dbs)
        if match is not None:
            out[y] = match[0]
    return out


# ── Endpoints ─────────────────────────────────────────────────────────────────


@router.post("/calculate")
async def post_calculate(body: ImpactAssessmentRequest) -> dict[str, str]:
    mode = (body.mode or "").lower()
    if mode not in {"static", "projected"}:
        raise HTTPException(status_code=400, detail="mode must be 'static' or 'projected'")
    _get_system(body.mfa_system_id)
    project = _current_project()

    # Resolve the DSM scenario to use. When ``dsm_scenario_id`` is set we
    # simulate that scenario fresh (multi-DSM fan-out path, Patch 2E.1). When
    # unset we fall back to the cached active-scenario sim — backward compat
    # for every existing single-scenario caller.
    state = _get_or_create_state(body.mfa_system_id)
    if body.dsm_scenario_id is not None:
        try:
            scen_obj = get_scenario(state, body.dsm_scenario_id)
        except KeyError:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"DSM scenario '{body.dsm_scenario_id}' not found on "
                    f"system '{body.mfa_system_id}'."
                ),
            )
        resolved_dsm_scenario_id = scen_obj.id
        sim = simulate_for_scenario(body.mfa_system_id, body.dsm_scenario_id)
    else:
        resolved_dsm_scenario_id = state.active_scenario_id or BASE_SCENARIO_ID
        sim = _proj_results(project).get(body.mfa_system_id)
        if sim is None:
            raise HTTPException(
                status_code=400,
                detail="No simulation results yet. Run /dsm/systems/{id}/simulate first.",
            )
    mapping = _proj_cohort_mappings(project).get(body.mfa_system_id)
    if mapping is None or not mapping.mappings:
        raise HTTPException(
            status_code=400,
            detail="No cohort mappings set. POST /dsm/systems/{id}/cohort-mappings first.",
        )
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required.")
    if body.year_start is not None and body.year_end is not None and body.year_start > body.year_end:
        raise HTTPException(status_code=400, detail="year_start must be ≤ year_end.")

    archetypes = _proj_archetypes(project)
    cohort_to_archetype: dict[str, tuple[str, float]] = {}
    for entry in mapping.mappings:
        arc = archetypes.get(entry.archetype_id)
        if arc is None:
            raise HTTPException(
                status_code=400,
                detail=f"Archetype '{entry.archetype_id}' (mapped to {entry.cohort_key}) was deleted.",
            )
        err_rows = validation_error_count(arc.bom)
        if err_rows:
            raise HTTPException(
                status_code=422,
                detail={
                    "error": "validation_failed",
                    "message": (
                        f"Archetype '{arc.name}' has {err_rows} row(s) with unresolved "
                        "ecoinvent links. LCA computation is blocked until they are fixed."
                    ),
                    "archetype_id": arc.id,
                    "archetype_name": arc.name,
                    "error_rows": err_rows,
                    "report_url": f"/api/bom/archetypes/{arc.id}/validation-report",
                },
            )
        unlinked = sum(1 for m in iter_all_materials(arc.bom) if m.ecoinvent_activity is None)
        if unlinked:
            raise HTTPException(
                status_code=400,
                detail=f"Archetype '{arc.name}' has {unlinked} unlinked material(s).",
            )
        cohort_to_archetype[entry.cohort_key] = (entry.archetype_id, entry.scaling_factor)

    method_tuples = [tuple(m) for m in body.methods if m]
    if not method_tuples:
        raise HTTPException(status_code=400, detail="At least one method is required.")

    # Resolve parameter expressions against the selected set (if any).
    from mapper.api import parameters as _parameters
    from mapper.core.parameter_engine import ParameterEngine, ParameterError
    param_engine: ParameterEngine | None = None
    if body.parameter_set_id:
        pset = _parameters.get_parameter_set(body.parameter_set_id, project)
        if pset is None:
            raise HTTPException(
                status_code=400,
                detail=f"Parameter set '{body.parameter_set_id}' not found",
            )
        param_engine = ParameterEngine(pset.parameters)

    # Discover dependent subsystems of this primary system. Each yields its own
    # SimulationResult (computed against the primary sim) plus a user-defined
    # cohort→BOM-archetype mapping. Unmapped dependent archetypes are skipped
    # with a warning rather than failing the whole run.
    from mapper.api import subsystems as _subs
    dep_subs = _subs.get_subsystems_for_system(body.mfa_system_id, project)
    sub_sim_results: dict[str, Any] = {}
    sub_cohort_mappings: dict[str, dict[str, tuple[str, float]]] = {}
    setup_warnings: list[str] = []
    for sub_id, sub in dep_subs.items():
        if not sub.dependency_rules:
            continue
        mapping, unmapped = build_subsystem_cohort_mapping(sub)
        if unmapped:
            setup_warnings.append(
                f"Subsystem '{sub.name}': {len(unmapped)} unmapped archetype"
                f"{'s' if len(unmapped) != 1 else ''} excluded from calculation: "
                f"{', '.join(unmapped)}"
            )
        if not mapping:
            # Nothing in this subsystem is mapped — skip it entirely.
            continue
        for aid in {bom_id for bom_id, _ in mapping.values()}:
            arc = archetypes.get(aid)
            if arc is None:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Dependent subsystem '{sub.name}' is mapped to archetype "
                        f"'{aid}' which does not exist in the BOM library."
                    ),
                )
            err_rows = validation_error_count(arc.bom)
            if err_rows:
                raise HTTPException(
                    status_code=422,
                    detail={
                        "error": "validation_failed",
                        "message": (
                            f"Archetype '{arc.name}' (subsystem '{sub.name}') has "
                            f"{err_rows} row(s) with unresolved ecoinvent links. "
                            "LCA computation is blocked until they are fixed."
                        ),
                        "archetype_id": arc.id,
                        "archetype_name": arc.name,
                        "error_rows": err_rows,
                        "report_url": f"/api/bom/archetypes/{arc.id}/validation-report",
                    },
                )
            unlinked = sum(1 for m in iter_all_materials(arc.bom) if m.ecoinvent_activity is None)
            if unlinked:
                raise HTTPException(
                    status_code=400,
                    detail=f"Archetype '{arc.name}' has {unlinked} unlinked material(s).",
                )
        try:
            sub_sim = compute_dependent_subsystem(
                sub, _get_system(body.mfa_system_id), sim, param_engine
            )
        except (ParameterError, ValueError) as e:
            raise HTTPException(
                status_code=400,
                detail=f"Dependent subsystem '{sub.name}': {e}",
            )
        sub_sim_results[sub_id] = sub_sim
        sub_cohort_mappings[sub_id] = mapping

    # Normalize the projected-mode scenario list:
    #   ``lci_scenarios`` (new) takes precedence over ``scenario`` (legacy).
    #   ``len == 1`` collapses to single-scenario semantics so the legacy
    #   single-shape response stays untouched. ``len > 1`` enters the multi
    #   path which returns a :class:`MultiScenarioProjectedImpactResult`.
    lci_scenarios_list: list[ProspectiveScenarioRef] = []
    if mode == "projected":
        if body.lci_scenarios:
            lci_scenarios_list = list(body.lci_scenarios)
        elif body.scenario is not None:
            lci_scenarios_list = [body.scenario]
        else:
            raise HTTPException(
                status_code=400,
                detail="Projected mode requires a scenario (base_db, iam, ssp) or lci_scenarios list.",
            )

    multi_lci_mode = mode == "projected" and len(lci_scenarios_list) > 1

    # Pre-resolve per-scenario prospective DBs so we 400 fast on missing
    # databases (rather than mid-worker after spending wall time on the
    # earlier scenarios). ``year_to_db`` is the single-scenario shortcut for
    # the legacy ``meta.year_to_database`` field.
    per_scenario_prospective: dict[int, list[tuple[str, int]]] = {}
    per_scenario_year_to_db: dict[int, dict[int, str]] = {}
    prospective_dbs: list[tuple[str, int]] = []
    year_to_db: dict[int, str] = {}
    if mode == "projected":
        sim_years = [yr.year for yr in sim.years]
        for idx, sc in enumerate(lci_scenarios_list):
            dbs = _resolve_prospective_dbs(project, sc)
            if not dbs:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"No prospective databases found for scenario "
                        f"{sc.base_db} / {sc.iam} / {sc.ssp}. "
                        "Generate them via /plca/generate first."
                    ),
                )
            per_scenario_prospective[idx] = dbs
            per_scenario_year_to_db[idx] = _year_to_database_map(
                sim_years, dbs, body.year_start, body.year_end,
            )
        if not multi_lci_mode:
            prospective_dbs = per_scenario_prospective[0]
            year_to_db = per_scenario_year_to_db[0]

    task_id = uuid.uuid4().hex
    task = _TaskState()
    with _TASK_LOCK:
        _TASKS[task_id] = task
    task_registry.register(task_id)

    loop = asyncio.get_running_loop()

    def _publish(stage: str, pct: float) -> None:
        task.stage = stage
        task.pct = pct
        loop.call_soon_threadsafe(
            _notify_all, task, {"type": "progress", "stage": stage, "pct": pct}
        )

    def _run() -> None:
        t0 = time.perf_counter()
        try:
            _publish("preparing", 0.02)
            in_range_years = [
                yr.year for yr in sim.years
                if (body.year_start is None or yr.year >= body.year_start)
                and (body.year_end is None or yr.year <= body.year_end)
            ]
            total_years = max(len(in_range_years), 1)
            if body.scope == "all":
                scope_labels = ["Manufacturing", "Operation", "End of Life"]
            else:
                scope_labels = [_SCOPE_LABELS.get(body.scope, body.scope)]
            n_subs = 1 + len(sub_sim_results)

            # Per-scenario worker. Returns the ``ImpactAssessmentResult`` for
            # one (base_db, iam, ssp). Pct is mapped into the slice
            # [pct_lo, pct_hi] so multi-scenario progress advances smoothly
            # across the whole job.
            def _calc_one_scenario(
                scenario: ProspectiveScenarioRef | None,
                pct_lo: float,
                pct_hi: float,
                slice_label: str,
            ) -> tuple[ImpactAssessmentResult, dict[int, str]]:
                if task_registry.is_cancelled(task_id):
                    raise CancelledOperation(task_id)
                # Each scenario gets its own factorization (different
                # technosphere matrix) — that's the cost of swapping LCI.
                persistent_local = PersistentLCARunner()

                def _publish_slice(stage: str, pct: float) -> None:
                    scaled = pct_lo + (pct_hi - pct_lo) * max(0.0, min(1.0, pct))
                    prefix = f"{slice_label} · " if slice_label else ""
                    _publish(f"{prefix}{stage}", scaled)

                runner_local = _progress_runner(
                    _publish_slice, in_range_years, scope_labels, persistent_local,
                    subsystem_count=n_subs,
                    task_id=task_id,
                )

                # Per-scenario prospective DBs (already pre-resolved upstream).
                if mode == "projected":
                    sc_idx = lci_scenarios_list.index(scenario) if scenario else 0
                    prosp = per_scenario_prospective[sc_idx]
                    yr_to_db = per_scenario_year_to_db[sc_idx]
                else:
                    prosp = []
                    yr_to_db = {}

                def _make_pipeline_local(sim_result, cohort_map):
                    if mode == "projected":
                        return ProjectedDSMLCAPipeline(
                            simulation_result=sim_result,
                            archetypes=archetypes,
                            cohort_mappings=cohort_map,
                            methods=method_tuples,
                            lca_runner=runner_local,
                            year_start=body.year_start,
                            year_end=body.year_end,
                            parameter_engine=param_engine,
                            prospective_dbs=prosp,
                            fallback_base_db=body.base_db,
                        )
                    return DSMLCAPipeline(
                        simulation_result=sim_result,
                        archetypes=archetypes,
                        cohort_mappings=cohort_map,
                        methods=method_tuples,
                        lca_runner=runner_local,
                        year_start=body.year_start,
                        year_end=body.year_end,
                        parameter_engine=param_engine,
                    )

                primary_pipeline = _make_pipeline_local(sim, cohort_to_archetype)
                primary_results = primary_pipeline.calculate(body.scope)

                if sub_sim_results:
                    results_by_subsystem: dict[str, list] = {
                        body.mfa_system_id: primary_results,
                    }
                    for sub_id, sub_sim in sub_sim_results.items():
                        sub_pipeline = _make_pipeline_local(sub_sim, sub_cohort_mappings[sub_id])
                        results_by_subsystem[sub_id] = sub_pipeline.calculate(body.scope)
                    sc_results = aggregate_subsystem_results(results_by_subsystem)
                    for r in sc_results:
                        r.mfa_system_id = body.mfa_system_id
                else:
                    sc_results = primary_results

                logging.getLogger(__name__).info(
                    "LCA scenario complete (%s): factorizations=%d, redo_lci=%d, "
                    "method_switches=%d",
                    slice_label or "single",
                    persistent_local.factorizations,
                    persistent_local.redo_calls,
                    persistent_local.method_switches,
                )

                sc_meta = ImpactAssessmentMeta(
                    mode=mode,
                    mfa_system_id=body.mfa_system_id,
                    scope=body.scope,
                    year_start=body.year_start,
                    year_end=body.year_end,
                    base_db=body.base_db,
                    scenario=scenario,
                    parameter_set_id=body.parameter_set_id,
                    dsm_scenario_id=resolved_dsm_scenario_id,
                    year_to_database=yr_to_db,
                    warnings=setup_warnings,
                )
                sc_out = ImpactAssessmentResult(
                    task_id=task_id, meta=sc_meta, results=sc_results,
                    elapsed_seconds=None,  # filled in by caller for single mode
                )
                return sc_out, yr_to_db

            label = "projected" if mode == "projected" else "static"
            if multi_lci_mode:
                _publish(
                    f"running {label} LCA ({len(lci_scenarios_list)} LCI scenarios × "
                    f"{total_years} years × {len(method_tuples)} method(s) "
                    f"× {n_subs} subsystem{'s' if n_subs != 1 else ''})",
                    0.1,
                )
                # Sequential scenario loop. Each scenario gets a slice of
                # [0.10, 0.95]. Cancellation is honored at scenario boundaries
                # AND inside each LCA call (per-iteration ``_progress_runner``).
                scenario_results: list[ScenarioProjectedResult] = []
                base_lo = 0.10
                base_hi = 0.95
                step = (base_hi - base_lo) / max(len(lci_scenarios_list), 1)
                for k, sc in enumerate(lci_scenarios_list):
                    if task_registry.is_cancelled(task_id):
                        raise CancelledOperation(task_id)
                    pct_lo = base_lo + step * k
                    pct_hi = base_lo + step * (k + 1)
                    slice_label = f"scenario {k + 1}/{len(lci_scenarios_list)} · {sc.iam.upper()}/{sc.ssp}"
                    sc_result, _ = _calc_one_scenario(sc, pct_lo, pct_hi, slice_label)
                    scenario_results.append(
                        ScenarioProjectedResult(scenario=sc, result=sc_result)
                    )
                elapsed = round(time.perf_counter() - t0, 2)
                wrapper_meta = ImpactAssessmentMeta(
                    mode=mode,
                    mfa_system_id=body.mfa_system_id,
                    scope=body.scope,
                    year_start=body.year_start,
                    year_end=body.year_end,
                    base_db=body.base_db,
                    scenario=None,
                    parameter_set_id=body.parameter_set_id,
                    dsm_scenario_id=resolved_dsm_scenario_id,
                    year_to_database={},
                    warnings=setup_warnings,
                )
                out_multi = MultiScenarioProjectedImpactResult(
                    task_id=task_id,
                    meta=wrapper_meta,
                    scenarios=scenario_results,
                    elapsed_seconds=elapsed,
                )
                task.result = out_multi
                task.stage = "done"
                task.pct = 1.0
                task.done = True
                loop.call_soon_threadsafe(
                    _notify_all,
                    task,
                    {
                        "type": "done",
                        "result_type": "multi_scenario_projected",
                        "scenarios_calculated": len(scenario_results),
                        "elapsed_seconds": elapsed,
                    },
                )
                return

            _publish(
                f"running {label} LCA ({total_years} years × {len(method_tuples)} method(s) "
                f"× {n_subs} subsystem{'s' if n_subs != 1 else ''})",
                0.1,
            )
            single_scenario = lci_scenarios_list[0] if lci_scenarios_list else None
            single_result, single_yr_to_db = _calc_one_scenario(
                single_scenario, 0.10, 0.95, "",
            )
            elapsed = round(time.perf_counter() - t0, 2)
            single_result.elapsed_seconds = elapsed
            task.result = single_result
            task.stage = "done"
            task.pct = 1.0
            task.done = True
            loop.call_soon_threadsafe(
                _notify_all,
                task,
                {
                    "type": "done",
                    "methods_calculated": len(single_result.results),
                    "year_to_database": single_yr_to_db,
                    "elapsed_seconds": elapsed,
                },
            )
        except CancelledOperation:
            task.cancelled = True
            task.done = True
            task.stage = "cancelled"
            # Result intentionally left None — partial year-by-year
            # ImpactAssessmentResult would mislead the trajectory chart.
            loop.call_soon_threadsafe(
                _notify_all, task, {"type": "cancelled", "task_id": task_id}
            )
        except Exception as exc:  # pragma: no cover
            task.error = str(exc)
            task.done = True
            loop.call_soon_threadsafe(
                _notify_all, task, {"type": "error", "error": str(exc)}
            )
        finally:
            task_registry.unregister(task_id)

    threading.Thread(target=_run, daemon=True).start()
    return {"task_id": task_id}


_SCOPE_LABELS = {
    "inflows": "Manufacturing",
    "stock": "Operation",
    "outflows": "End of Life",
    "all": "Full lifecycle",
}


def _progress_runner(
    publish,
    years: list[int],
    scope_labels: list[str],
    persistent: PersistentLCARunner | None = None,
    subsystem_count: int = 1,
    task_id: str | None = None,
):
    """Wrap a ``PersistentLCARunner`` (or fall back to ``run_lca_multi_method``)
    and emit a progress tick per (scope, year) pair.

    When *persistent* is provided, the runner reuses a single LU
    factorization of the technosphere matrix across all years — turning
    ~0.5 s per year into ~1 ms back-substitution.

    The pipeline iterates scopes sequentially (``_ATOMIC_SCOPES`` order:
    inflows → stock → outflows), each looping over all in-range years,
    so the (scope, year) pair can be derived from the call counter.
    """
    lca_fn = persistent or run_lca_multi_method
    counter = {"n": 0}
    total_years = max(len(years), 1)
    total = total_years * max(len(scope_labels), 1) * max(subsystem_count, 1)

    def runner(demand, method_tuples):
        if task_id is not None and task_registry.is_cancelled(task_id):
            raise CancelledOperation(task_id)
        counter["n"] += 1
        n = counter["n"]
        pct = 0.1 + 0.85 * min(1.0, n / total)
        idx = n - 1
        scope_idx = min(idx // total_years, len(scope_labels) - 1)
        year_idx = idx % total_years
        year = years[year_idx] if years else None
        scope_label = scope_labels[scope_idx] if scope_labels else None
        if year is not None and scope_label and len(scope_labels) > 1:
            stage = f"{year} · {scope_label} ({n}/{total})"
        elif year is not None and scope_label:
            stage = f"year {year} · {scope_label} ({n}/{total})"
        elif year is not None:
            stage = f"year {year} ({n}/{total})"
        else:
            stage = f"{n}/{total}"
        publish(stage, pct)
        return lca_fn(demand, method_tuples)

    return runner


@router.post("/calculate-scenarios")
async def post_calculate_scenarios(body: ImpactAssessmentRequest) -> dict[str, dict[str, str]]:
    """Launch the impact pipeline once per scenario.

    Two fan-out axes are supported, **mutually exclusive**:

    - ``scenarios`` — parameter-table scenario names (Patch 2C). One task per
      scenario, each with its own ``parameter_set_id``.
    - ``dsm_scenario_ids`` — DSM scenario ids (Patch 2E.1). One task per id,
      each simulating that DSM scenario fresh inside the worker.

    Returns ``{scenarios: {key: task_id, ...}}`` where ``key`` is either the
    parameter scenario name OR the DSM scenario id depending on which axis
    fanned out. The client polls ``/impact/results/{task_id}`` per task and
    assembles the appropriate envelope (``MultiParamImpactResult`` or
    ``MultiDSMImpactResult``) before exporting.

    Falls back to single-scenario behaviour (one task on ``"Base"``) when
    neither field is set — callers get a uniform shape.

    The 3-way axisConflict rule (LCI × DSM × Parameter, at most one >1) is
    mirrored here: 400 if both ``scenarios`` and ``dsm_scenario_ids`` are
    non-empty. Multi-LCI on top of either is rejected by ``post_calculate``
    when it sees ``len(lci_scenarios) > 1`` plus a multi-axis fan-out parent.
    """
    fan_param = bool(body.scenarios)
    fan_dsm = bool(body.dsm_scenario_ids)
    fan_paired = bool(body.paired_scenarios)
    # Multi-LCI is "in-task" (single task, sequential per scenario) and is
    # already rejected by post_calculate when paired with a multi-axis parent.
    # Here we only check the parallel-fan-out axes.
    multi_axes = sum([fan_param, fan_dsm, fan_paired])
    if multi_axes > 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot fan out multiple axes simultaneously "
                "(parameter, DSM, paired DSM×LCI). Pick one at a time "
                "(axisConflict rule)."
            ),
        )

    if fan_paired:
        # Paired DSM × LCI fan-out (Patch 2F): one task per pair, each task
        # carries BOTH ``dsm_scenario_id`` (singular, drives a fresh sim
        # inside the worker) AND ``scenario`` (singular ProspectiveScenarioRef
        # for the year-matched LCI lookup). The pair key is
        # ``"<dsm_id>::<base_db>::<iam>::<ssp>"`` — deterministic so the
        # frontend can compute it and look up assignments without round-trip.
        # Multi-LCI ``lci_scenarios`` is forced to None on per-pair bodies so
        # the worker's single-scenario path runs (each pair has exactly one
        # LCI ref by definition).
        # Pre-validate uniqueness across the whole pair list before launching
        # any task — otherwise a duplicate pair late in the list would leave
        # earlier successful tasks orphaned with no way for the client to
        # collect them.
        pair_keys: list[str] = []
        seen_keys: set[str] = set()
        for pair in body.paired_scenarios or []:
            ref = pair.lci_scenario
            key = f"{pair.dsm_scenario_id}::{ref.base_db}::{ref.iam}::{ref.ssp}"
            if key in seen_keys:
                # Frontend validates and forbids duplicates inline; this is
                # defence in depth.
                raise HTTPException(
                    status_code=400,
                    detail=f"Duplicate pair: {key}",
                )
            seen_keys.add(key)
            pair_keys.append(key)

        out_paired: dict[str, str] = {}
        for pair, key in zip(body.paired_scenarios or [], pair_keys):
            sid = pair.dsm_scenario_id
            ref = pair.lci_scenario
            per_pair = body.model_copy(update={
                "scenarios": None,
                "dsm_scenario_ids": None,
                "paired_scenarios": None,
                "lci_scenarios": None,
                "dsm_scenario_id": sid,
                "scenario": ref,
            })
            response = await post_calculate(per_pair)
            out_paired[key] = response["task_id"]
        return {"scenarios": out_paired}

    if fan_dsm:
        # DSM-axis fan-out: one task per DSM scenario id; each task simulates
        # its scenario fresh via ``ImpactAssessmentRequest.dsm_scenario_id``.
        out: dict[str, str] = {}
        for sid in body.dsm_scenario_ids or []:
            per_scenario = body.model_copy(update={
                "scenarios": None,
                "dsm_scenario_ids": None,
                "dsm_scenario_id": sid,
            })
            response = await post_calculate(per_scenario)
            out[sid] = response["task_id"]
        return {"scenarios": out}

    scenarios = body.scenarios or [body.parameter_set_id or "Base"]
    if not scenarios:
        raise HTTPException(status_code=400, detail="scenarios must be non-empty")

    out_param: dict[str, str] = {}
    for scen in scenarios:
        per_scenario = body.model_copy(update={
            "scenarios": None,
            "parameter_set_id": scen,
        })
        response = await post_calculate(per_scenario)
        out_param[scen] = response["task_id"]
    return {"scenarios": out_param}


@router.get(
    "/results/{task_id}",
    response_model=ImpactAssessmentResult | MultiScenarioProjectedImpactResult | CancelledTaskResponse,
)
async def get_results(
    task_id: str,
) -> ImpactAssessmentResult | MultiScenarioProjectedImpactResult | CancelledTaskResponse:
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Unknown task id")
    if task.cancelled:
        return CancelledTaskResponse(task_id=task_id)
    if task.error:
        raise HTTPException(status_code=500, detail=task.error)
    if not task.done or task.result is None:
        raise HTTPException(status_code=425, detail="Task not finished yet")
    return task.result


@router.post("/compare", response_model=ImpactCompareResult)
async def post_compare(body: ImpactCompareRequest) -> ImpactCompareResult:
    with _TASK_LOCK:
        t_static = _TASKS.get(body.static_task_id)
        t_proj = _TASKS.get(body.projected_task_id)
    if t_static is None or t_static.result is None:
        raise HTTPException(status_code=404, detail="Static task not found or not finished.")
    if t_proj is None or t_proj.result is None:
        raise HTTPException(status_code=404, detail="Projected task not found or not finished.")
    if isinstance(t_static.result, MultiScenarioProjectedImpactResult) or isinstance(
        t_proj.result, MultiScenarioProjectedImpactResult
    ):
        raise HTTPException(
            status_code=400,
            detail=(
                "Static-vs-Projected comparison is not supported for multi-scenario "
                "LCI runs. Pick one LCI scenario at a time when comparing."
            ),
        )
    static = t_static.result
    proj = t_proj.result

    if static.meta.mfa_system_id != proj.meta.mfa_system_id:
        raise HTTPException(status_code=400, detail="Comparison requires both runs on the same DSM system.")
    if static.meta.scope != proj.meta.scope:
        raise HTTPException(status_code=400, detail="Comparison requires both runs in the same scope.")

    static_by_method: dict[tuple, DSMLCAResult] = {tuple(r.method): r for r in static.results}
    proj_by_method: dict[tuple, DSMLCAResult] = {tuple(r.method): r for r in proj.results}

    methods_out: list[ImpactCompareMethodResult] = []
    for mkey, s_res in static_by_method.items():
        p_res = proj_by_method.get(mkey)
        if p_res is None:
            continue
        s_years = {y.year: y.total_impact for y in s_res.years}
        p_years = {y.year: y.total_impact for y in p_res.years}
        all_years = sorted(set(s_years) | set(p_years))
        points: list[ImpactComparePoint] = []
        total_s = 0.0
        total_p = 0.0
        for y in all_years:
            sv = s_years.get(y, 0.0)
            pv = p_years.get(y, 0.0)
            delta = pv - sv
            dpct = (delta / sv * 100.0) if sv else None
            points.append(ImpactComparePoint(
                year=y, static_impact=sv, projected_impact=pv, delta=delta, delta_pct=dpct,
            ))
            total_s += sv
            total_p += pv
        total_delta = total_p - total_s
        methods_out.append(ImpactCompareMethodResult(
            method=list(mkey),
            method_label=s_res.method_label or " › ".join(mkey),
            unit=p_res.unit or s_res.unit,
            points=points,
            total_static=total_s,
            total_projected=total_p,
            total_delta=total_delta,
            total_delta_pct=(total_delta / total_s * 100.0) if total_s else None,
        ))

    return ImpactCompareResult(
        mfa_system_id=static.meta.mfa_system_id,
        scope=static.meta.scope,
        methods=methods_out,
    )


def _resolve_export_result(
    task_id: str | None, inline: ImpactAssessmentResult | None
) -> ImpactAssessmentResult:
    if inline is not None:
        return inline
    if task_id is None:
        raise HTTPException(status_code=400, detail="Either task_id or inline result is required.")
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None or task.result is None:
        raise HTTPException(status_code=404, detail=f"Task {task_id} not found or not finished.")
    if isinstance(task.result, MultiScenarioProjectedImpactResult):
        raise HTTPException(
            status_code=400,
            detail=(
                "Task is a multi-scenario projected run. Pass `multi_result` (or "
                "fetch via /impact/results) and route through the multi-scenario "
                "export path."
            ),
        )
    return task.result


def _resolve_multi_export_result(
    task_id: str | None,
    inline: MultiScenarioProjectedImpactResult | None,
) -> MultiScenarioProjectedImpactResult | None:
    """Resolve a multi-scenario result either from an inline payload or a
    registered task. Returns ``None`` if neither side carries one — caller
    falls back to single-scenario export."""
    if inline is not None:
        return inline
    if task_id is None:
        return None
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None or task.result is None:
        return None
    if isinstance(task.result, MultiScenarioProjectedImpactResult):
        return task.result
    return None


def _append_compare_sheet(
    wb,
    static: ImpactAssessmentResult,
    projected: ImpactAssessmentResult,
) -> None:
    if static.meta.mfa_system_id != projected.meta.mfa_system_id:
        return
    if static.meta.scope != projected.meta.scope:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="374151")
    num_fmt = "0.00E+00"

    ws = wb.create_sheet("Static vs Projected")
    ws.append([f"Scope: {static.meta.scope}   ·   Static task: {static.task_id}   ·   Projected task: {projected.task_id}"])
    ws["A1"].font = Font(italic=True, color="6B7280")

    static_by_method: dict[tuple, DSMLCAResult] = {tuple(r.method): r for r in static.results}
    first_method_block = True

    for p_res in projected.results:
        mkey = tuple(p_res.method)
        s_res = static_by_method.get(mkey)
        if s_res is None:
            continue

        if not first_method_block:
            ws.append([])
        first_method_block = False

        ws.append([f"Method: {' › '.join(mkey)}   Unit: {p_res.unit or s_res.unit}"])
        ws[ws.max_row][0].font = Font(bold=True)

        header = ["Year", f"Static ({s_res.unit})", f"Projected ({p_res.unit})", "Δ", "Δ %"]
        ws.append(header)
        for cell in ws[ws.max_row]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        s_years = {y.year: y.total_impact for y in s_res.years}
        p_years = {y.year: y.total_impact for y in p_res.years}
        years = sorted(set(s_years) | set(p_years))
        total_s = 0.0
        total_p = 0.0
        data_start = ws.max_row + 1
        for y in years:
            sv = s_years.get(y, 0.0)
            pv = p_years.get(y, 0.0)
            delta = pv - sv
            dpct = (delta / abs(sv) * 100.0) if sv else None
            ws.append([y, sv, pv, delta, dpct if dpct is not None else ""])
            total_s += sv
            total_p += pv
        total_delta = total_p - total_s
        total_pct = (total_delta / abs(total_s) * 100.0) if total_s else None
        ws.append(["Total", total_s, total_p, total_delta, total_pct if total_pct is not None else ""])
        ws[ws.max_row][0].font = Font(bold=True)

        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = num_fmt
        for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, min_col=5, max_col=5):
            for cell in row:
                cell.number_format = "0.00"


def _scenario_label(sc: ProspectiveScenarioRef) -> str:
    """Compact human-readable LCI scenario label, e.g. ``REMIND/SSP2-PkBudg1150``."""
    return f"{sc.iam.upper()}/{sc.ssp}"


def _build_multi_scenario_workbook(
    system_name: str,
    multi_result: MultiScenarioProjectedImpactResult,
    sys_def,
    sim_counts: dict[int, dict[str, float]] | None,
    sim_result,
    archetypes: dict,
    cohort_mapping,
):
    """Build a multi-LCI-scenario Excel workbook.

    Layout:
      - Sheet "Summary" — one row per (LCI Scenario × Indicator) with cumulative
        impact, peak year, peak impact.
      - Sheet "Annual totals" — Year, LCI Scenario, then one column per indicator.
      - Sheet "By indicator" — Year, LCI Scenario, then triplets per indicator
        (Annual / Cumulative / YoY %).
      - Sheet "LCI Scenarios" — index of scenarios (base_db / iam / ssp).

    Per-cohort and contribution-style sheets are deliberately omitted in the
    multi-scenario export to keep the file readable. Users wanting per-cohort
    detail should re-run the scenario standalone.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from mapper.api.bom import _short_method_label
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    SCI_FMT = "0.00E+00"
    PCT_FMT = "0.00"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    meta_font = Font(bold=True, color="374151")

    def _style_header(ws, row_num: int = 1) -> None:
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autosize(ws, max_width: int = 40) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for i, cell in enumerate(col_cells):
                if i >= 51:
                    break
                v = cell.value
                if v is not None:
                    widest = max(widest, min(max_width, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    scenarios = multi_result.scenarios
    if not scenarios:
        # Edge case — should not happen but build an empty wb gracefully.
        wb.create_sheet("Summary")
        return wb

    # All scenarios share methods/years (driven by same request body), so we
    # take indicator labels from the first scenario.
    first_results = scenarios[0].result.results
    labels = [_short_method_label(r.method) for r in first_results]
    units = [r.unit for r in first_results]
    year_set: set[int] = set()
    for s in scenarios:
        for r in s.result.results:
            for yr in r.years:
                year_set.add(yr.year)
    years_list = sorted(year_set)

    scope_labels = {
        "inflows": "Manufacturing",
        "stock": "Operation",
        "outflows": "End of Life",
        "all": "Full lifecycle",
    }

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3ECFCF"

    meta_rows = [
        ("Project", system_name),
        ("Mode", "Projected · multi-LCI-scenario"),
        ("Scope", scope_labels.get(multi_result.meta.scope, multi_result.meta.scope)),
        ("Year range", f"{years_list[0]}–{years_list[-1]}" if years_list else "—"),
        ("LCI scenarios", len(scenarios)),
        ("Indicators", len(first_results)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if multi_result.elapsed_seconds is not None:
        m, s = divmod(int(multi_result.elapsed_seconds), 60)
        meta_rows.append(
            ("Calculation time", f"{m}m {s}s" if m else f"{s}s")
        )

    for label, value in meta_rows:
        ws.append([label, value])
        ws[ws.max_row][0].font = meta_font

    ws.append([])
    ws.append([
        "LCI Scenario", "Base DB", "IAM", "SSP",
        "Indicator", "Method path", "Unit",
        "Cumulative impact", "Peak year", "Peak impact",
    ])
    _style_header(ws, ws.max_row)
    summary_start = ws.max_row + 1
    for s in scenarios:
        sc = s.scenario
        sc_label = _scenario_label(sc)
        for r in s.result.results:
            ws.append([
                sc_label,
                sc.base_db,
                sc.iam,
                sc.ssp,
                _short_method_label(r.method),
                " › ".join(r.method),
                r.unit,
                r.summary.total_impact,
                r.summary.peak_year,
                r.summary.peak_impact,
            ])
    # Sci format on Cumulative + Peak impact (cols 8, 10)
    for col_idx in (8, 10):
        for row in ws.iter_rows(min_row=summary_start, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = SCI_FMT
    ws.freeze_panes = "A" + str(summary_start)
    _autosize(ws)

    # ── Sheet 2: Annual totals ────────────────────────────────────────────────
    ws_at = wb.create_sheet("Annual totals")
    ws_at.sheet_properties.tabColor = "4A90D9"
    header = ["Year", "LCI Scenario"] + [f"{l} ({u})" for l, u in zip(labels, units)]
    ws_at.append(header)
    _style_header(ws_at)

    for s in scenarios:
        sc_label = _scenario_label(s.scenario)
        result_by_method = {tuple(r.method): r for r in s.result.results}
        for y in years_list:
            row: list = [y, sc_label]
            for r in first_results:
                mkey = tuple(r.method)
                src = result_by_method.get(mkey)
                yr = next((v for v in src.years if v.year == y), None) if src else None
                row.append(yr.total_impact if yr else 0.0)
            ws_at.append(row)
    for row in ws_at.iter_rows(min_row=2, min_col=3, max_col=len(header)):
        for cell in row:
            cell.number_format = SCI_FMT
    ws_at.freeze_panes = "C2"
    _autosize(ws_at)

    # ── Sheet 3: By indicator (Annual / Cumulative / YoY %) ───────────────────
    ws_bi = wb.create_sheet("By indicator")
    ws_bi.sheet_properties.tabColor = "4A90D9"
    bi_header: list[str] = ["Year", "LCI Scenario"]
    for l, u in zip(labels, units):
        bi_header.append(f"{l} ({u})")
        bi_header.append(f"{l} cumulative")
        bi_header.append(f"{l} YoY %")
    ws_bi.append(bi_header)
    _style_header(ws_bi)

    for s in scenarios:
        sc_label = _scenario_label(s.scenario)
        result_by_method = {tuple(r.method): r for r in s.result.results}
        # Pre-compute year→total per indicator for this scenario
        per_method_year_total: list[dict[int, float]] = []
        for r in first_results:
            mkey = tuple(r.method)
            src = result_by_method.get(mkey)
            d: dict[int, float] = {}
            if src is not None:
                for yr in src.years:
                    d[yr.year] = yr.total_impact
            per_method_year_total.append(d)

        for y in years_list:
            row: list = [y, sc_label]
            for ri in range(len(first_results)):
                annual = per_method_year_total[ri].get(y, 0.0)
                cumul = sum(
                    per_method_year_total[ri].get(yy, 0.0)
                    for yy in years_list if yy <= y
                )
                prev = per_method_year_total[ri].get(y - 1, 0.0)
                yoy = ((annual - prev) / abs(prev) * 100.0) if prev else None
                row.extend([annual, cumul, yoy if yoy is not None else ""])
            ws_bi.append(row)

    for col_idx in range(3, len(bi_header) + 1):
        col_offset = (col_idx - 3) % 3  # 0=annual, 1=cumul, 2=yoy
        fmt = SCI_FMT if col_offset < 2 else PCT_FMT
        for row in ws_bi.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    ws_bi.freeze_panes = "C2"
    _autosize(ws_bi)

    # ── Sheet 4: LCI Scenarios index ──────────────────────────────────────────
    ws_idx = wb.create_sheet("LCI Scenarios")
    ws_idx.sheet_properties.tabColor = "9CA3AF"
    ws_idx.append(["LCI Scenario", "Base DB", "IAM", "SSP"])
    _style_header(ws_idx)
    for s in scenarios:
        sc = s.scenario
        ws_idx.append([_scenario_label(sc), sc.base_db, sc.iam, sc.ssp])
    ws_idx.freeze_panes = "A2"
    _autosize(ws_idx)

    return wb


def _resolve_varying_parameters(
    scenario_names: list[str],
) -> tuple[list[str], dict[str, dict[str, float]]]:
    """For the given parameter-table scenario names, return:

      - the ordered list of parameter names whose value *varies* across at
        least two of the named scenarios (parameters that are constant
        across all selected scenarios are dropped — they don't distinguish
        anything in the index sheet);
      - a ``{scenario_name -> {param_name -> value}}`` map covering only the
        varying parameters.

    Reads the active project's :class:`ParameterTable` via
    ``mapper.core.parameter_storage``. If no table is loaded (project has no
    parameters defined) returns ``([], {scen: {} for scen in scenario_names})``
    so the caller can still render an index sheet — just with no parameter
    columns.
    """
    from mapper.core import parameter_storage

    project = _current_project()
    table = parameter_storage.load_parameter_table(project)
    if table is None or not table.parameters:
        return [], {scen: {} for scen in scenario_names}

    full: dict[str, dict[str, float]] = {
        scen: table.resolve_all(scen) for scen in scenario_names
    }
    # A parameter is "varying" iff at least two scenarios produce different
    # values. Use the parameter ordering from ``table.parameters`` for
    # determinism — alpha-sorting would lose any user-meaningful ordering.
    varying: list[str] = []
    for pname in table.parameters:
        values = {full[scen].get(pname) for scen in scenario_names}
        if len(values) > 1:
            varying.append(pname)

    trimmed: dict[str, dict[str, float]] = {
        scen: {p: full[scen][p] for p in varying if p in full[scen]}
        for scen in scenario_names
    }
    return varying, trimmed


def _build_multi_param_workbook(
    system_name: str,
    multi_param_result: MultiParamImpactResult,
):
    """Build a multi-parameter Excel workbook (sibling to the multi-LCI
    builder).

    Layout — symmetric with ``_build_multi_scenario_workbook``:
      - Sheet "Summary" — meta block + one row per (Sensitivity case ×
        Indicator) with cumulative impact, peak year, peak impact.
      - Sheet "Annual totals" — Year, Sensitivity case, then one column per
        indicator.
      - Sheet "By indicator" — Year, Sensitivity case, then triplets per
        indicator (Annual / Cumulative / YoY %).
      - Sheet "Parameter Scenarios" — index of scenarios. Lists only the
        parameters whose values *vary* across the selected scenarios (a
        parameter that's constant doesn't distinguish anything).

    Per-cohort and contribution-style sheets are deliberately omitted to keep
    the file readable. Re-run a single scenario for that detail.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from mapper.api.bom import _short_method_label
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    SCI_FMT = "0.00E+00"
    PCT_FMT = "0.00"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    meta_font = Font(bold=True, color="374151")

    def _style_header(ws, row_num: int = 1) -> None:
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autosize(ws, max_width: int = 40) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for i, cell in enumerate(col_cells):
                if i >= 51:
                    break
                v = cell.value
                if v is not None:
                    widest = max(widest, min(max_width, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    scenarios = multi_param_result.scenarios
    if not scenarios:
        wb.create_sheet("Summary")
        return wb

    # Iterate scenarios in submission order (matches the in-app tab order).
    # Indicator labels come from the first scenario; all scenarios share the
    # same method list since they're spawned from one request body.
    first_results = scenarios[0].result.results
    labels = [_short_method_label(r.method) for r in first_results]
    units = [r.unit for r in first_results]

    year_set: set[int] = set()
    for s in scenarios:
        for r in s.result.results:
            for yr in r.years:
                year_set.add(yr.year)
    years_list = sorted(year_set)

    scope_labels = {
        "inflows": "Manufacturing",
        "stock": "Operation",
        "outflows": "End of Life",
        "all": "Full lifecycle",
    }

    mode_label = (
        "Static · multi-parameter"
        if multi_param_result.meta.mode == "static"
        else "Projected · multi-parameter"
    )

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3ECFCF"

    meta_rows = [
        ("Project", system_name),
        ("Mode", mode_label),
        ("Scope", scope_labels.get(multi_param_result.meta.scope, multi_param_result.meta.scope)),
        ("Year range", f"{years_list[0]}–{years_list[-1]}" if years_list else "—"),
        ("Sensitivity cases", len(scenarios)),
        ("Indicators", len(first_results)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if multi_param_result.elapsed_seconds is not None:
        m, s_sec = divmod(int(multi_param_result.elapsed_seconds), 60)
        meta_rows.append(
            ("Calculation time", f"{m}m {s_sec}s" if m else f"{s_sec}s")
        )

    for label, value in meta_rows:
        ws.append([label, value])
        ws[ws.max_row][0].font = meta_font

    ws.append([])
    ws.append([
        "Sensitivity case",
        "Indicator", "Method path", "Unit",
        "Cumulative impact", "Peak year", "Peak impact",
    ])
    _style_header(ws, ws.max_row)
    summary_start = ws.max_row + 1
    for s in scenarios:
        for r in s.result.results:
            ws.append([
                s.scenario,
                _short_method_label(r.method),
                " › ".join(r.method),
                r.unit,
                r.summary.total_impact,
                r.summary.peak_year,
                r.summary.peak_impact,
            ])
    # Sci format on Cumulative + Peak impact (cols 5, 7)
    for col_idx in (5, 7):
        for row in ws.iter_rows(min_row=summary_start, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = SCI_FMT
    ws.freeze_panes = "A" + str(summary_start)
    _autosize(ws)

    # ── Sheet 2: Annual totals ────────────────────────────────────────────────
    ws_at = wb.create_sheet("Annual totals")
    ws_at.sheet_properties.tabColor = "4A90D9"
    header = ["Year", "Sensitivity case"] + [f"{l} ({u})" for l, u in zip(labels, units)]
    ws_at.append(header)
    _style_header(ws_at)

    for s in scenarios:
        result_by_method = {tuple(r.method): r for r in s.result.results}
        for y in years_list:
            row: list = [y, s.scenario]
            for r in first_results:
                mkey = tuple(r.method)
                src = result_by_method.get(mkey)
                yr = next((v for v in src.years if v.year == y), None) if src else None
                row.append(yr.total_impact if yr else 0.0)
            ws_at.append(row)
    for row in ws_at.iter_rows(min_row=2, min_col=3, max_col=len(header)):
        for cell in row:
            cell.number_format = SCI_FMT
    ws_at.freeze_panes = "C2"
    _autosize(ws_at)

    # ── Sheet 3: By indicator (Annual / Cumulative / YoY %) ───────────────────
    ws_bi = wb.create_sheet("By indicator")
    ws_bi.sheet_properties.tabColor = "4A90D9"
    bi_header: list[str] = ["Year", "Sensitivity case"]
    for l, u in zip(labels, units):
        bi_header.append(f"{l} ({u})")
        bi_header.append(f"{l} cumulative")
        bi_header.append(f"{l} YoY %")
    ws_bi.append(bi_header)
    _style_header(ws_bi)

    for s in scenarios:
        result_by_method = {tuple(r.method): r for r in s.result.results}
        per_method_year_total: list[dict[int, float]] = []
        for r in first_results:
            mkey = tuple(r.method)
            src = result_by_method.get(mkey)
            d: dict[int, float] = {}
            if src is not None:
                for yr in src.years:
                    d[yr.year] = yr.total_impact
            per_method_year_total.append(d)

        for y in years_list:
            row: list = [y, s.scenario]
            for ri in range(len(first_results)):
                annual = per_method_year_total[ri].get(y, 0.0)
                cumul = sum(
                    per_method_year_total[ri].get(yy, 0.0)
                    for yy in years_list if yy <= y
                )
                prev = per_method_year_total[ri].get(y - 1, 0.0)
                yoy = ((annual - prev) / abs(prev) * 100.0) if prev else None
                row.extend([annual, cumul, yoy if yoy is not None else ""])
            ws_bi.append(row)

    for col_idx in range(3, len(bi_header) + 1):
        col_offset = (col_idx - 3) % 3  # 0=annual, 1=cumul, 2=yoy
        fmt = SCI_FMT if col_offset < 2 else PCT_FMT
        for row in ws_bi.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    ws_bi.freeze_panes = "C2"
    _autosize(ws_bi)

    # ── Sheet 4: Parameter Scenarios index ────────────────────────────────────
    ws_idx = wb.create_sheet("Parameter Scenarios")
    ws_idx.sheet_properties.tabColor = "9CA3AF"
    scenario_names = [s.scenario for s in scenarios]
    varying, values_by_scen = _resolve_varying_parameters(scenario_names)

    if varying:
        ws_idx.append(["Sensitivity case", *varying])
        _style_header(ws_idx)
        for scen in scenario_names:
            row = [scen]
            scen_values = values_by_scen.get(scen, {})
            for p in varying:
                row.append(scen_values.get(p, ""))
            ws_idx.append(row)
        # Numeric format for the parameter-value columns.
        for col_idx in range(2, 2 + len(varying)):
            for row in ws_idx.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = "0.######"
    else:
        # No parameter table loaded, or all selected scenarios share identical
        # values for every parameter. Render a stub so the sheet still
        # documents what was selected.
        ws_idx.append(["Sensitivity case"])
        _style_header(ws_idx)
        for scen in scenario_names:
            ws_idx.append([scen])
        ws_idx.append([])
        note = (
            "No parameters vary across the selected sensitivity cases. The "
            "selection still produces distinct results if BOM expressions "
            "reference parameters — but no parameter values differ between "
            "the named scenarios in the active project's parameter table."
        )
        ws_idx.append([note])
        ws_idx[ws_idx.max_row][0].font = Font(italic=True, color="6B7280")

    ws_idx.freeze_panes = "B2"
    _autosize(ws_idx)

    return wb


def _dsm_scenario_summary_stats(
    result: ImpactAssessmentResult,
) -> dict[str, float | int | None]:
    """Per-DSM-scenario summary stats derived from the impact result itself
    (no re-simulation needed).

    DSM scenarios are structurally opaque — their distinguishing data lives in
    nested slots (initial_stock dict, inflows list, mode_configs list, …) that
    can't be flattened to "parameter X = value Y" rows. The multi-parameter
    ``_resolve_varying_parameters`` pattern doesn't translate. Instead the
    index sheet shows what *actually differed* across scenarios in their
    simulation output: fleet counts at the year boundaries and the peak
    year.

    All stats are derived from ``result.results[0].years[].count_by_cohort``
    (count semantics depend on ``meta.scope`` — inflows / stock / outflows).
    Returns a dict; missing values are ``None`` (rendered as blank in the
    sheet) when the result is empty.
    """
    if not result.results or not result.results[0].years:
        return {
            "first_year": None,
            "last_year": None,
            "count_first": None,
            "count_last": None,
            "peak_year": None,
            "peak_count": None,
            "cohorts": 0,
        }

    years = result.results[0].years
    first = years[0]
    last = years[-1]
    fleet = [(yr.year, sum(yr.count_by_cohort.values())) for yr in years]
    peak_year, peak_count = max(fleet, key=lambda t: t[1]) if fleet else (None, None)

    cohort_keys: set[str] = set()
    for yr in years:
        cohort_keys.update(yr.count_by_cohort.keys())

    return {
        "first_year": first.year,
        "last_year": last.year,
        "count_first": sum(first.count_by_cohort.values()),
        "count_last": sum(last.count_by_cohort.values()),
        "peak_year": peak_year,
        "peak_count": peak_count,
        "cohorts": len(cohort_keys),
    }


def _build_multi_dsm_workbook(
    system_name: str,
    multi_dsm_result: MultiDSMImpactResult,
):
    """Build a multi-DSM Excel workbook (sibling to the multi-parameter and
    multi-LCI builders).

    Layout — symmetric with ``_build_multi_param_workbook``:

      - Sheet "Summary" — meta block + one row per (DSM scenario × Indicator)
        with cumulative impact, peak year, peak impact.
      - Sheet "Annual totals" — Year, DSM scenario, then one column per
        indicator.
      - Sheet "By indicator" — Year, DSM scenario, then triplets per
        indicator (Annual / Cumulative / YoY %).
      - Sheet "DSM Scenarios" — index. DSM scenarios are structurally opaque
        (no flat parameter table to varying-filter), so the index lists
        per-scenario simulation summary stats derived from the impact
        result: fleet count at the year boundaries, peak year + count, and
        number of distinct cohorts.

    Per-cohort and contribution-style sheets are deliberately omitted (same
    convention as multi-LCI / multi-param). Re-run a single DSM scenario for
    that detail.

    Discriminator column on every data sheet: ``"DSM scenario"`` (matches
    the UI chip label).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from mapper.api.bom import _short_method_label
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    SCI_FMT = "0.00E+00"
    PCT_FMT = "0.00"
    INT_FMT = "0"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    meta_font = Font(bold=True, color="374151")

    def _style_header(ws, row_num: int = 1) -> None:
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autosize(ws, max_width: int = 40) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for i, cell in enumerate(col_cells):
                if i >= 51:
                    break
                v = cell.value
                if v is not None:
                    widest = max(widest, min(max_width, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    scenarios = multi_dsm_result.scenarios
    if not scenarios:
        wb.create_sheet("Summary")
        return wb

    # Iterate DSM scenarios in submission order (matches the in-app tab order).
    # Indicator labels come from the first scenario; all share the same method
    # list since they're spawned from one request body.
    first_results = scenarios[0].result.results
    labels = [_short_method_label(r.method) for r in first_results]
    units = [r.unit for r in first_results]

    year_set: set[int] = set()
    for s in scenarios:
        for r in s.result.results:
            for yr in r.years:
                year_set.add(yr.year)
    years_list = sorted(year_set)

    scope_labels = {
        "inflows": "Manufacturing",
        "stock": "Operation",
        "outflows": "End of Life",
        "all": "Full lifecycle",
    }

    mode_label = (
        "Static · multi-DSM"
        if multi_dsm_result.meta.mode == "static"
        else "Projected · multi-DSM"
    )

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3ECFCF"

    meta_rows = [
        ("Project", system_name),
        ("Mode", mode_label),
        ("Scope", scope_labels.get(multi_dsm_result.meta.scope, multi_dsm_result.meta.scope)),
        ("Year range", f"{years_list[0]}–{years_list[-1]}" if years_list else "—"),
        ("DSM scenarios", len(scenarios)),
        ("Indicators", len(first_results)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if multi_dsm_result.elapsed_seconds is not None:
        m, s_sec = divmod(int(multi_dsm_result.elapsed_seconds), 60)
        meta_rows.append(
            ("Calculation time", f"{m}m {s_sec}s" if m else f"{s_sec}s")
        )

    for label, value in meta_rows:
        ws.append([label, value])
        ws[ws.max_row][0].font = meta_font

    ws.append([])
    ws.append([
        "DSM scenario",
        "Indicator", "Method path", "Unit",
        "Cumulative impact", "Peak year", "Peak impact",
    ])
    _style_header(ws, ws.max_row)
    summary_start = ws.max_row + 1
    for s in scenarios:
        for r in s.result.results:
            ws.append([
                s.scenario_name,
                _short_method_label(r.method),
                " › ".join(r.method),
                r.unit,
                r.summary.total_impact,
                r.summary.peak_year,
                r.summary.peak_impact,
            ])
    for col_idx in (5, 7):
        for row in ws.iter_rows(min_row=summary_start, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = SCI_FMT
    ws.freeze_panes = "A" + str(summary_start)
    _autosize(ws)

    # ── Sheet 2: Annual totals ────────────────────────────────────────────────
    ws_at = wb.create_sheet("Annual totals")
    ws_at.sheet_properties.tabColor = "4A90D9"
    header = ["Year", "DSM scenario"] + [f"{l} ({u})" for l, u in zip(labels, units)]
    ws_at.append(header)
    _style_header(ws_at)

    for s in scenarios:
        result_by_method = {tuple(r.method): r for r in s.result.results}
        for y in years_list:
            row: list = [y, s.scenario_name]
            for r in first_results:
                mkey = tuple(r.method)
                src = result_by_method.get(mkey)
                yr = next((v for v in src.years if v.year == y), None) if src else None
                row.append(yr.total_impact if yr else 0.0)
            ws_at.append(row)
    for row in ws_at.iter_rows(min_row=2, min_col=3, max_col=len(header)):
        for cell in row:
            cell.number_format = SCI_FMT
    ws_at.freeze_panes = "C2"
    _autosize(ws_at)

    # ── Sheet 3: By indicator (Annual / Cumulative / YoY %) ───────────────────
    ws_bi = wb.create_sheet("By indicator")
    ws_bi.sheet_properties.tabColor = "4A90D9"
    bi_header: list[str] = ["Year", "DSM scenario"]
    for l, u in zip(labels, units):
        bi_header.append(f"{l} ({u})")
        bi_header.append(f"{l} cumulative")
        bi_header.append(f"{l} YoY %")
    ws_bi.append(bi_header)
    _style_header(ws_bi)

    for s in scenarios:
        result_by_method = {tuple(r.method): r for r in s.result.results}
        per_method_year_total: list[dict[int, float]] = []
        for r in first_results:
            mkey = tuple(r.method)
            src = result_by_method.get(mkey)
            d: dict[int, float] = {}
            if src is not None:
                for yr in src.years:
                    d[yr.year] = yr.total_impact
            per_method_year_total.append(d)

        for y in years_list:
            row: list = [y, s.scenario_name]
            for ri in range(len(first_results)):
                annual = per_method_year_total[ri].get(y, 0.0)
                cumul = sum(
                    per_method_year_total[ri].get(yy, 0.0)
                    for yy in years_list if yy <= y
                )
                prev = per_method_year_total[ri].get(y - 1, 0.0)
                yoy = ((annual - prev) / abs(prev) * 100.0) if prev else None
                row.extend([annual, cumul, yoy if yoy is not None else ""])
            ws_bi.append(row)

    for col_idx in range(3, len(bi_header) + 1):
        col_offset = (col_idx - 3) % 3
        fmt = SCI_FMT if col_offset < 2 else PCT_FMT
        for row in ws_bi.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    ws_bi.freeze_panes = "C2"
    _autosize(ws_bi)

    # ── Sheet 4: DSM Scenarios index ──────────────────────────────────────────
    # DSM scenarios are structurally opaque (initial_stock / inflows /
    # stock_targets / outflows / mode_configs / scaling_rules — all nested
    # data, no flat numeric parameters). Index lists per-scenario simulation
    # summary stats derived from the impact result.
    ws_idx = wb.create_sheet("DSM Scenarios")
    ws_idx.sheet_properties.tabColor = "9CA3AF"
    ws_idx.append([
        "DSM scenario", "Scenario ID",
        "First year", "Fleet at first year",
        "Last year", "Fleet at last year",
        "Peak year", "Peak fleet count",
        "Cohorts active",
    ])
    _style_header(ws_idx)

    any_stats = False
    for s in scenarios:
        stats = _dsm_scenario_summary_stats(s.result)
        if stats["first_year"] is not None:
            any_stats = True
        ws_idx.append([
            s.scenario_name,
            s.scenario_id,
            stats["first_year"] if stats["first_year"] is not None else "",
            stats["count_first"] if stats["count_first"] is not None else "",
            stats["last_year"] if stats["last_year"] is not None else "",
            stats["count_last"] if stats["count_last"] is not None else "",
            stats["peak_year"] if stats["peak_year"] is not None else "",
            stats["peak_count"] if stats["peak_count"] is not None else "",
            stats["cohorts"],
        ])
    # Numeric formats: years as integers, fleet counts as fixed (DSM counts
    # can be fractional in non-integer-units mode).
    for col_idx in (3, 5, 7):  # Year columns
        for row in ws_idx.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = INT_FMT
    for col_idx in (4, 6, 8):  # Fleet count columns
        for row in ws_idx.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.##"
    for row in ws_idx.iter_rows(min_row=2, min_col=9, max_col=9):  # Cohorts active
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = INT_FMT

    if not any_stats:
        # All scenarios produced empty results — render a stub note so the
        # sheet still documents the selection meaningfully.
        ws_idx.append([])
        note = (
            "No simulation data available for the selected DSM scenarios. "
            "Re-run the impact assessment if scenarios were not yet computed."
        )
        ws_idx.append([note])
        ws_idx[ws_idx.max_row][0].font = Font(italic=True, color="6B7280")

    _autosize(ws_idx)

    return wb


def _paired_short_label(dsm_name: str, ref: ProspectiveScenarioRef) -> str:
    """Pair label for chart legends and discriminator columns.

    Format: ``<dsm_name> × <iam>/<ssp>`` — drops base_db (usually a single
    constant within one project's pairs) to keep the label compact for
    repeated use on every data row.
    """
    return f"{dsm_name} × {ref.iam}/{ref.ssp}"


def _paired_full_label(dsm_name: str, ref: ProspectiveScenarioRef) -> str:
    """Verbose pair label for tooltips and the index sheet.

    Format: ``<dsm_name> stock × <base_db>/<iam>/<ssp>`` — includes base_db
    so a future cross-project workbook stays unambiguous.
    """
    return f"{dsm_name} stock × {ref.base_db}/{ref.iam}/{ref.ssp}"


def _build_multi_paired_workbook(
    system_name: str,
    multi_paired_result: MultiPairedImpactResult,
):
    """Build a multi-paired Excel workbook (sibling to the multi-LCI,
    multi-parameter, and multi-DSM builders).

    Layout — symmetric with ``_build_multi_dsm_workbook``:

      - Sheet "Summary" — meta block + one row per (Pair × Indicator).
      - Sheet "Annual totals" — Year, Pair, then one column per indicator.
      - Sheet "By indicator" — Year, Pair, then triplets per indicator
        (Annual / Cumulative / YoY %).
      - Sheet "Pairs" — index. One row per pair listing the DSM scenario
        side (id + name + simulation summary stats) and the LCI side
        (base_db / iam / ssp + full label).

    Per-cohort and contribution-style sheets are deliberately omitted (same
    convention as the other multi-axis builders). Re-run a single pair for
    that detail.

    Discriminator column on every data sheet: ``"Pair"`` (short label,
    e.g. ``SSP1 × remind/SSP1-PkBudg1150``).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from mapper.api.bom import _short_method_label
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    SCI_FMT = "0.00E+00"
    PCT_FMT = "0.00"
    INT_FMT = "0"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    meta_font = Font(bold=True, color="374151")

    def _style_header(ws, row_num: int = 1) -> None:
        for cell in ws[row_num]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

    def _autosize(ws, max_width: int = 40) -> None:
        for col_idx, col_cells in enumerate(ws.columns, start=1):
            widest = 0
            for i, cell in enumerate(col_cells):
                if i >= 51:
                    break
                v = cell.value
                if v is not None:
                    widest = max(widest, min(max_width, len(str(v))))
            ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)

    pairs = multi_paired_result.scenarios
    if not pairs:
        wb.create_sheet("Summary")
        return wb

    # Submission order from the request body is preserved through the
    # envelope — matches the in-app pair list order.
    first_results = pairs[0].result.results
    labels = [_short_method_label(r.method) for r in first_results]
    units = [r.unit for r in first_results]

    year_set: set[int] = set()
    for p in pairs:
        for r in p.result.results:
            for yr in r.years:
                year_set.add(yr.year)
    years_list = sorted(year_set)

    scope_labels = {
        "inflows": "Manufacturing",
        "stock": "Operation",
        "outflows": "End of Life",
        "all": "Full lifecycle",
    }

    short_label_by_pair: dict[int, str] = {
        idx: _paired_short_label(p.dsm_scenario_name, p.lci_scenario)
        for idx, p in enumerate(pairs)
    }

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_properties.tabColor = "3ECFCF"

    meta_rows = [
        ("Project", system_name),
        ("Mode", "Paired DSM × LCI (co-varying)"),
        ("Scope", scope_labels.get(multi_paired_result.meta.scope, multi_paired_result.meta.scope)),
        ("Year range", f"{years_list[0]}–{years_list[-1]}" if years_list else "—"),
        ("Pairs", len(pairs)),
        ("Indicators", len(first_results)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    if multi_paired_result.elapsed_seconds is not None:
        m, s_sec = divmod(int(multi_paired_result.elapsed_seconds), 60)
        meta_rows.append(
            ("Calculation time", f"{m}m {s_sec}s" if m else f"{s_sec}s")
        )

    for label, value in meta_rows:
        ws.append([label, value])
        ws[ws.max_row][0].font = meta_font

    ws.append([])
    ws.append([
        "Pair",
        "Indicator", "Method path", "Unit",
        "Cumulative impact", "Peak year", "Peak impact",
    ])
    _style_header(ws, ws.max_row)
    summary_start = ws.max_row + 1
    for idx, p in enumerate(pairs):
        plabel = short_label_by_pair[idx]
        for r in p.result.results:
            ws.append([
                plabel,
                _short_method_label(r.method),
                " › ".join(r.method),
                r.unit,
                r.summary.total_impact,
                r.summary.peak_year,
                r.summary.peak_impact,
            ])
    for col_idx in (5, 7):
        for row in ws.iter_rows(min_row=summary_start, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = SCI_FMT
    ws.freeze_panes = "A" + str(summary_start)
    _autosize(ws)

    # ── Sheet 2: Annual totals ────────────────────────────────────────────────
    ws_at = wb.create_sheet("Annual totals")
    ws_at.sheet_properties.tabColor = "4A90D9"
    header = ["Year", "Pair"] + [f"{l} ({u})" for l, u in zip(labels, units)]
    ws_at.append(header)
    _style_header(ws_at)

    for idx, p in enumerate(pairs):
        plabel = short_label_by_pair[idx]
        result_by_method = {tuple(r.method): r for r in p.result.results}
        for y in years_list:
            row: list = [y, plabel]
            for r in first_results:
                mkey = tuple(r.method)
                src = result_by_method.get(mkey)
                yr = next((v for v in src.years if v.year == y), None) if src else None
                row.append(yr.total_impact if yr else 0.0)
            ws_at.append(row)
    for row in ws_at.iter_rows(min_row=2, min_col=3, max_col=len(header)):
        for cell in row:
            cell.number_format = SCI_FMT
    ws_at.freeze_panes = "C2"
    _autosize(ws_at)

    # ── Sheet 3: By indicator (Annual / Cumulative / YoY %) ───────────────────
    ws_bi = wb.create_sheet("By indicator")
    ws_bi.sheet_properties.tabColor = "4A90D9"
    bi_header: list[str] = ["Year", "Pair"]
    for l, u in zip(labels, units):
        bi_header.append(f"{l} ({u})")
        bi_header.append(f"{l} cumulative")
        bi_header.append(f"{l} YoY %")
    ws_bi.append(bi_header)
    _style_header(ws_bi)

    for idx, p in enumerate(pairs):
        plabel = short_label_by_pair[idx]
        result_by_method = {tuple(r.method): r for r in p.result.results}
        per_method_year_total: list[dict[int, float]] = []
        for r in first_results:
            mkey = tuple(r.method)
            src = result_by_method.get(mkey)
            d: dict[int, float] = {}
            if src is not None:
                for yr in src.years:
                    d[yr.year] = yr.total_impact
            per_method_year_total.append(d)

        for y in years_list:
            row: list = [y, plabel]
            for ri in range(len(first_results)):
                annual = per_method_year_total[ri].get(y, 0.0)
                cumul = sum(
                    per_method_year_total[ri].get(yy, 0.0)
                    for yy in years_list if yy <= y
                )
                prev = per_method_year_total[ri].get(y - 1, 0.0)
                yoy = ((annual - prev) / abs(prev) * 100.0) if prev else None
                row.extend([annual, cumul, yoy if yoy is not None else ""])
            ws_bi.append(row)

    for col_idx in range(3, len(bi_header) + 1):
        col_offset = (col_idx - 3) % 3
        fmt = SCI_FMT if col_offset < 2 else PCT_FMT
        for row in ws_bi.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = fmt
    ws_bi.freeze_panes = "C2"
    _autosize(ws_bi)

    # ── Sheet 4: Pairs index ──────────────────────────────────────────────────
    # Self-contained pair description: short label (matches data-sheet column),
    # full label (verbose), DSM identity + simulation summary stats, LCI ref
    # components broken out. No re-simulate; everything derived from the
    # envelope.
    ws_idx = wb.create_sheet("Pairs")
    ws_idx.sheet_properties.tabColor = "9CA3AF"
    ws_idx.append([
        "Pair", "Full label",
        "DSM scenario", "DSM scenario ID",
        "First year", "Fleet at first year",
        "Last year", "Fleet at last year",
        "Peak year", "Peak fleet count",
        "Cohorts active",
        "LCI base DB", "LCI IAM", "LCI SSP",
    ])
    _style_header(ws_idx)

    any_stats = False
    for idx, p in enumerate(pairs):
        stats = _dsm_scenario_summary_stats(p.result)
        if stats["first_year"] is not None:
            any_stats = True
        ws_idx.append([
            short_label_by_pair[idx],
            _paired_full_label(p.dsm_scenario_name, p.lci_scenario),
            p.dsm_scenario_name,
            p.dsm_scenario_id,
            stats["first_year"] if stats["first_year"] is not None else "",
            stats["count_first"] if stats["count_first"] is not None else "",
            stats["last_year"] if stats["last_year"] is not None else "",
            stats["count_last"] if stats["count_last"] is not None else "",
            stats["peak_year"] if stats["peak_year"] is not None else "",
            stats["peak_count"] if stats["peak_count"] is not None else "",
            stats["cohorts"],
            p.lci_scenario.base_db,
            p.lci_scenario.iam,
            p.lci_scenario.ssp,
        ])
    # Year columns as integers; fleet counts allow fractional; cohorts integer.
    for col_idx in (5, 7, 9):
        for row in ws_idx.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = INT_FMT
    for col_idx in (6, 8, 10):
        for row in ws_idx.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.##"
    for row in ws_idx.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = INT_FMT

    if not any_stats:
        ws_idx.append([])
        note = (
            "No simulation data available for the selected pairs. "
            "Re-run the impact assessment if pairs were not yet computed."
        )
        ws_idx.append([note])
        ws_idx[ws_idx.max_row][0].font = Font(italic=True, color="6B7280")

    _autosize(ws_idx)

    return wb


@router.post("/export")
async def post_export(body: ImpactExportRequest) -> Response:
    # Multi-parameter (Static or Projected) and multi-LCI runs go through
    # sibling builders with a discriminator column on every data sheet.
    # Single-scenario keeps the rich legacy 9-sheet workbook path unchanged.
    # The 3-way axisConflict rule on the frontend prevents both multi-modes
    # from being set simultaneously; if a client violates that we 400.
    if body.multi_param_result is not None and body.multi_result is not None:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot export multi-parameter and multi-LCI envelopes in one "
                "request. Pick one axis at a time (axisConflict rule)."
            ),
        )
    multi_axis_envelopes = sum(
        1 for x in (
            body.multi_param_result, body.multi_result,
            body.multi_dsm_result, body.multi_paired_result,
        )
        if x is not None
    )
    if multi_axis_envelopes > 1:
        raise HTTPException(
            status_code=400,
            detail=(
                "Cannot export multiple multi-axis envelopes in one request. "
                "Pick one axis at a time (axisConflict rule)."
            ),
        )
    if body.multi_paired_result is not None:
        mpr = body.multi_paired_result
        if not mpr.scenarios:
            raise HTTPException(
                status_code=400,
                detail="multi_paired_result must contain at least one pair.",
            )
        sys_def = _get_system(mpr.meta.mfa_system_id)

        wb = _build_multi_paired_workbook(
            system_name=sys_def.name, multi_paired_result=mpr,
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import datetime
        scope_labels_fname = {
            "inflows": "Manufacturing",
            "stock": "Operation",
            "outflows": "End_of_Life",
            "all": "Full_lifecycle",
        }
        scope_tag = scope_labels_fname.get(mpr.meta.scope, mpr.meta.scope)
        date_tag = datetime.date.today().isoformat()
        filename = (
            f"MApper_Impact_MultiPaired_"
            f"{_sanitize_filename(sys_def.name, 'system')}_{scope_tag}_{date_tag}.xlsx"
        )
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if body.multi_dsm_result is not None:
        md = body.multi_dsm_result
        if not md.scenarios:
            raise HTTPException(
                status_code=400,
                detail="multi_dsm_result must contain at least one scenario.",
            )
        sys_def = _get_system(md.meta.mfa_system_id)

        wb = _build_multi_dsm_workbook(system_name=sys_def.name, multi_dsm_result=md)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import datetime
        scope_labels_fname = {
            "inflows": "Manufacturing",
            "stock": "Operation",
            "outflows": "End_of_Life",
            "all": "Full_lifecycle",
        }
        scope_tag = scope_labels_fname.get(md.meta.scope, md.meta.scope)
        date_tag = datetime.date.today().isoformat()
        filename = (
            f"MApper_Impact_MultiDSM_"
            f"{_sanitize_filename(sys_def.name, 'system')}_{scope_tag}_{date_tag}.xlsx"
        )
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    project = _current_project()
    archetypes = _proj_archetypes(project)

    if body.multi_param_result is not None:
        mp = body.multi_param_result
        if not mp.scenarios:
            raise HTTPException(
                status_code=400,
                detail="multi_param_result must contain at least one scenario.",
            )
        sys_def = _get_system(mp.meta.mfa_system_id)

        wb = _build_multi_param_workbook(system_name=sys_def.name, multi_param_result=mp)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import datetime
        scope_labels_fname = {
            "inflows": "Manufacturing",
            "stock": "Operation",
            "outflows": "End_of_Life",
            "all": "Full_lifecycle",
        }
        scope_tag = scope_labels_fname.get(mp.meta.scope, mp.meta.scope)
        date_tag = datetime.date.today().isoformat()
        filename = (
            f"MApper_Impact_MultiParam_"
            f"{_sanitize_filename(sys_def.name, 'system')}_{scope_tag}_{date_tag}.xlsx"
        )
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    multi_result = _resolve_multi_export_result(body.task_id, body.multi_result)

    if multi_result is not None:
        sys_def = _get_system(multi_result.meta.mfa_system_id)
        mapping = _proj_cohort_mappings(project).get(multi_result.meta.mfa_system_id)
        sim = _proj_results(project).get(multi_result.meta.mfa_system_id)
        sim_counts: dict[int, dict[str, float]] = {}
        if sim is not None:
            scope = multi_result.meta.scope
            for yr in sim.years:
                if scope == "inflows":
                    sim_counts[yr.year] = dict(yr.inflow)
                elif scope == "outflows":
                    sim_counts[yr.year] = dict(yr.outflow)
                else:
                    sim_counts[yr.year] = dict(yr.stock)

        wb = _build_multi_scenario_workbook(
            system_name=sys_def.name,
            multi_result=multi_result,
            sys_def=sys_def,
            sim_counts=sim_counts,
            sim_result=sim,
            archetypes=archetypes,
            cohort_mapping=mapping,
        )

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        import datetime
        scope_labels_fname = {
            "inflows": "Manufacturing",
            "stock": "Operation",
            "outflows": "End_of_Life",
            "all": "Full_lifecycle",
        }
        scope_tag = scope_labels_fname.get(multi_result.meta.scope, multi_result.meta.scope)
        date_tag = datetime.date.today().isoformat()
        filename = (
            f"MApper_Impact_MultiLCI_"
            f"{_sanitize_filename(sys_def.name, 'system')}_{scope_tag}_{date_tag}.xlsx"
        )
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    # Single-scenario path (legacy).
    result = _resolve_export_result(body.task_id, body.result)
    compare_partner: ImpactAssessmentResult | None = None
    if body.compare_task_id is not None or body.compare_result is not None:
        compare_partner = _resolve_export_result(body.compare_task_id, body.compare_result)

    sys_def = _get_system(result.meta.mfa_system_id)
    mapping = _proj_cohort_mappings(project).get(result.meta.mfa_system_id)

    sim = _proj_results(project).get(result.meta.mfa_system_id)
    sim_counts: dict[int, dict[str, float]] = {}
    if sim is not None:
        scope = result.meta.scope
        for yr in sim.years:
            if scope == "inflows":
                sim_counts[yr.year] = dict(yr.inflow)
            elif scope == "outflows":
                sim_counts[yr.year] = dict(yr.outflow)
            else:
                sim_counts[yr.year] = dict(yr.stock)

    wb = _build_mfa_lca_workbook(
        system_name=sys_def.name,
        results=result.results,
        scope=result.meta.scope,
        selected_year=body.year,
        cohort_mapping=mapping,
        archetypes=archetypes,
        sim_counts=sim_counts,
        dims=list(sys_def.dimensions),
        elapsed_seconds=result.elapsed_seconds,
        sim_result=sim,
    )

    # Decide pairing for the Static-vs-Projected sheet.
    if compare_partner is not None:
        if result.meta.mode == "static":
            _append_compare_sheet(wb, static=result, projected=compare_partner)
        else:
            _append_compare_sheet(wb, static=compare_partner, projected=result)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import datetime
    scope_labels = {"inflows": "Manufacturing", "stock": "Operation", "outflows": "End_of_Life", "all": "Full_lifecycle"}
    scope_tag = scope_labels.get(result.meta.scope, result.meta.scope)
    date_tag = datetime.date.today().isoformat()
    filename = f"MApper_Impact_{_sanitize_filename(sys_def.name, 'system')}_{scope_tag}_{date_tag}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Single-product Impact Assessment exports (Patch 4G) ─────────────────────
#
# Three sibling builders, one per single-product sub-tab. Pattern matches
# the multi-axis builders earlier in this file:
#   - per-axis discriminator column on every data sheet
#   - first sheet is metadata/Configuration
#   - sci format on numeric columns, header style cyan/white
#
# Each builder is narrow: Static covers methods + (optional) sensitivity
# fan-out + (optional) stage breakdown. Prospective covers the
# (database × year × method) cube. Comparison reads both halves and
# computes Δ = P − S (matches the on-screen sign convention from
# Patches 4E/4F). Per-tab files keep each sub-tab's export shareable in
# isolation — no mega-workbook. See CLAUDE.md "Don't bundle Static +
# Prospective + Comparison into one mega-workbook."


_SP_SCOPE_LABELS = {
    "inflows": "Manufacturing",
    "stock": "Operation",
    "outflows": "End of Life",
    "all": "Full lifecycle",
}
_SP_SCOPE_FNAME = {
    "inflows": "Manufacturing",
    "stock": "Operation",
    "outflows": "End_of_Life",
    "all": "Full_lifecycle",
}
_SP_SCI_FMT = "0.00E+00"
_SP_PCT_FMT = "0.00"
_SP_HEADER_FONT = Font(bold=True, color="FFFFFF")
_SP_HEADER_FILL = PatternFill("solid", fgColor="3ECFCF")
_SP_META_FONT = Font(bold=True, color="374151")


def _sp_style_header(ws, row_num: int = 1) -> None:
    for cell in ws[row_num]:
        cell.font = _SP_HEADER_FONT
        cell.fill = _SP_HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _sp_autosize(ws, max_width: int = 40) -> None:
    from openpyxl.utils import get_column_letter
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        widest = 0
        for i, cell in enumerate(col_cells):
            if i >= 51:
                break
            v = cell.value
            if v is not None:
                widest = max(widest, min(max_width, len(str(v))))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, widest + 2)


def _sp_format_stage_amounts(stage_amounts: dict[str, float]) -> str:
    """Compact one-liner for the Configuration sheet.
    Stage amounts are insertion-ordered in the result so we preserve
    that order — it carries lifecycle semantics (Manufacturing first,
    EoL last on conventional BOMs)."""
    if not stage_amounts:
        return "—"
    return " · ".join(
        f"{k} {v:g}" for k, v in stage_amounts.items()
    )


def _sp_stage_amount_meta_rows(meta) -> list[tuple[str, Any]]:
    """Patch 5K+ — Configuration rows for stage-amount provenance: preset +
    lifetime. ``meta`` is an optional StageAmountsMeta; when absent (old
    client) both render as "—". Labels match the multi-item "Stage amounts"
    sheet ("Preset", "Lifetime (yr)") so single- and multi-item exports read
    consistently. Inserted right after the existing "Stage amounts" row; the
    per-stage amounts themselves continue to come from the result echo."""
    return [
        ("Preset", meta.preset if meta else "—"),
        ("Lifetime (yr)", meta.lifetime if meta else "—"),
    ]


def _sp_short_method(method: list[str]) -> str:
    """Short label for an LCIA method tuple. The full path goes in a
    sibling column when needed; this is what the user sees as the row
    name."""
    if not method:
        return "?"
    return method[-1]


def _build_single_product_static_workbook(
    archetype_name: str,
    scope: str,
    scenarios: list[tuple[str, ArchetypeLCACalculateResult]],
    stage_amounts_meta=None,
):
    """Build a single-product Static Background workbook.

    Layout:
      - Sheet "Configuration" — archetype, scope, stage amounts, methods,
        sensitivity cases, computation timestamp.
      - Sheet "Total impacts" — row=method, columns=value/unit. If
        multi-parameter (len(scenarios) > 1), one value column per
        sensitivity case (Base / user-named / ...).
      - Sheet "Stage breakdown" — only present when ``scope == 'all'``
        AND at least one scenario carries a non-empty stage_breakdown.
        Row=method, columns=Manufacturing / Use / Maintenance / EoL
        per stage scaled by stage_amounts.

    Multi-parameter detail: every numeric column on the data sheets gets
    its scenario-prefix when N > 1. Single-scenario keeps the simpler
    "value | unit" shape so the workbook reads like a normal report.
    """
    from openpyxl import Workbook
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    if not scenarios:
        wb.create_sheet("Configuration")
        return wb

    is_multi = len(scenarios) > 1
    primary = scenarios[0][1]  # use first as the canonical method list
    method_list = primary.results

    # ── Sheet 1: Configuration ──────────────────────────────────────────
    ws_cfg = wb.create_sheet("Configuration")
    ws_cfg.sheet_properties.tabColor = "3ECFCF"

    sensitivity_label = (
        ", ".join(label for label, _ in scenarios)
        if is_multi else scenarios[0][0]
    )
    cfg_rows: list[tuple[str, Any]] = [
        ("Archetype", archetype_name),
        ("Mode", "Single product · Static Background"),
        ("Scope", _SP_SCOPE_LABELS.get(scope, scope)),
        ("Stage amounts", _sp_format_stage_amounts(primary.stage_amounts)),
        *_sp_stage_amount_meta_rows(stage_amounts_meta),
        ("Stages included", ", ".join(primary.stages_included)),
        ("Indicators", len(method_list)),
        ("Sensitivity cases", sensitivity_label),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("Elapsed (s)", f"{primary.elapsed_seconds:.2f}"),
    ]
    if primary.parameter_scenario:
        cfg_rows.append(("Parameter scenario (active)", primary.parameter_scenario))
    if primary.warnings:
        cfg_rows.append(("Warnings", f"{len(primary.warnings)} (see in-app log)"))

    for label, value in cfg_rows:
        ws_cfg.append([label, value])
        ws_cfg[ws_cfg.max_row][0].font = _SP_META_FONT
    _sp_autosize(ws_cfg)

    # ── Sheet 2: Total impacts ──────────────────────────────────────────
    ws_tot = wb.create_sheet("Total impacts")
    ws_tot.sheet_properties.tabColor = "4A90D9"

    if is_multi:
        header = ["Indicator", "Method path", "Unit"] + [label for label, _ in scenarios]
        ws_tot.append(header)
        _sp_style_header(ws_tot)

        # Index method tuples per scenario for stable lookup. We anchor
        # row order on the first scenario; methods missing in others are
        # written as empty cells (rather than 0) to make the gap visible.
        per_scenario: list[dict[tuple[str, ...], float]] = []
        per_scenario_labels: list[dict[tuple[str, ...], str]] = []
        for _label, sr in scenarios:
            per_scenario.append({tuple(r.method): r.score for r in sr.results})
            per_scenario_labels.append({tuple(r.method): r.unit for r in sr.results})

        for r in method_list:
            mkey = tuple(r.method)
            row: list = [_sp_short_method(r.method), " › ".join(r.method), r.unit]
            for sc_idx in range(len(scenarios)):
                row.append(per_scenario[sc_idx].get(mkey, ""))
            ws_tot.append(row)
        # Sci format on every scenario column (4 = first scenario col)
        for col_idx in range(4, 4 + len(scenarios)):
            for row in ws_tot.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = _SP_SCI_FMT
    else:
        ws_tot.append(["Indicator", "Method path", "Score", "Unit"])
        _sp_style_header(ws_tot)
        for r in method_list:
            ws_tot.append([
                _sp_short_method(r.method),
                " › ".join(r.method),
                r.score,
                r.unit,
            ])
            ws_tot.cell(row=ws_tot.max_row, column=3).number_format = _SP_SCI_FMT
    ws_tot.freeze_panes = "A2"
    _sp_autosize(ws_tot)

    # ── Sheet 3: Stage breakdown (only when scope='all' and present) ────
    has_breakdown = scope == "all" and any(
        sr.stage_breakdown for _label, sr in scenarios
    )
    if has_breakdown:
        ws_sb = wb.create_sheet("Stage breakdown")
        ws_sb.sheet_properties.tabColor = "C29CFF"

        # Discover the stage column order from the first scenario that
        # has a breakdown — stages are insertion-ordered in the BOM, so
        # this preserves lifecycle ordering (Manufacturing → ... → EoL).
        stage_order: list[str] = []
        seen: set[str] = set()
        for _label, sr in scenarios:
            if not sr.stage_breakdown:
                continue
            for _method_label, by_stage in sr.stage_breakdown.items():
                for stage in by_stage.keys():
                    if stage not in seen:
                        seen.add(stage)
                        stage_order.append(stage)

        if is_multi:
            header = ["Sensitivity case", "Indicator", "Method path", "Unit"] + stage_order + ["Total"]
        else:
            header = ["Indicator", "Method path", "Unit"] + stage_order + ["Total"]
        ws_sb.append(header)
        _sp_style_header(ws_sb)

        for label, sr in scenarios:
            if not sr.stage_breakdown:
                continue
            # Map method tuple → ArchetypeLCAMethodResult for unit lookup
            unit_by_label = {
                _sp_short_method(r.method): r.unit for r in sr.results
            }
            path_by_label = {
                _sp_short_method(r.method): " › ".join(r.method) for r in sr.results
            }
            for method_label, by_stage in sr.stage_breakdown.items():
                row_total = sum(by_stage.values())
                row: list = []
                if is_multi:
                    row.append(label)
                row.extend([
                    method_label,
                    path_by_label.get(method_label, "—"),
                    unit_by_label.get(method_label, ""),
                ])
                for stage in stage_order:
                    row.append(by_stage.get(stage, 0.0))
                row.append(row_total)
                ws_sb.append(row)

        # Sci format every numeric column (stages + Total).
        first_num_col = 5 if is_multi else 4
        last_num_col = first_num_col + len(stage_order)  # +Total
        for col_idx in range(first_num_col, last_num_col + 1):
            for row in ws_sb.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = _SP_SCI_FMT
        ws_sb.freeze_panes = "A2"
        _sp_autosize(ws_sb)

    return wb


def _build_single_product_prospective_workbook(
    archetype_name: str,
    scope: str,
    runs: list[SingleProductProspectiveRunPayload],
    stage_amounts_meta=None,
):
    """Build a single-product Prospective Background workbook.

    Layout:
      - Sheet "Configuration" — archetype, scope, stage amounts, selected
        databases (one line per (iam, ssp, year)), year range, methods.
      - Sheet "Time series (wide)" — row=(database, method), columns
        per year. Most readable for users skimming a trajectory shape.
      - Sheet "Time series (long)" — one row per (database, year,
        method), single value column. Friendly for downstream tooling
        (pandas, R) that prefers long-format input.
      - Sheet "Stage breakdown by year" — only when scope='all' and at
        least one run carries a stage_breakdown. Row=(database, year,
        method), columns per stage + Total. Stages CAN evolve with
        prospective LCI (different LCI = different background = different
        per-stage decomposition), so keeping the year on the row axis is
        load-bearing.

    Year axis is the union across all runs, sorted ascending. Missing
    cells are blank (not 0) — explicit absence reads better than a
    silent zero.
    """
    from openpyxl import Workbook
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    if not runs:
        wb.create_sheet("Configuration")
        return wb

    primary = runs[0].result
    method_list = primary.results

    # Sort runs by (iam, ssp, year). Stable order for the wide sheet's
    # row axis and the configuration listing.
    sorted_runs = sorted(
        runs,
        key=lambda r: (r.iam, r.ssp, r.year if r.year is not None else 0),
    )

    # Year axis = union of run years, sorted ascending.
    year_set: set[int] = set()
    for r in sorted_runs:
        if r.year is not None:
            year_set.add(r.year)
    years = sorted(year_set)

    # ── Sheet 1: Configuration ──────────────────────────────────────────
    ws_cfg = wb.create_sheet("Configuration")
    ws_cfg.sheet_properties.tabColor = "3ECFCF"

    cfg_rows: list[tuple[str, Any]] = [
        ("Archetype", archetype_name),
        ("Mode", "Single product · Prospective Background"),
        ("Scope", _SP_SCOPE_LABELS.get(scope, scope)),
        ("Stage amounts", _sp_format_stage_amounts(primary.stage_amounts)),
        *_sp_stage_amount_meta_rows(stage_amounts_meta),
        ("Year range", f"{years[0]}–{years[-1]}" if years else "—"),
        ("Indicators", len(method_list)),
        ("Databases", len(sorted_runs)),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in cfg_rows:
        ws_cfg.append([label, value])
        ws_cfg[ws_cfg.max_row][0].font = _SP_META_FONT

    ws_cfg.append([])
    ws_cfg.append(["Database", "IAM", "SSP", "Year"])
    _sp_style_header(ws_cfg, ws_cfg.max_row)
    for r in sorted_runs:
        ws_cfg.append([r.db_name, r.iam, r.ssp, r.year if r.year is not None else "—"])
    _sp_autosize(ws_cfg)

    # ── Sheet 2: Time series (wide) ─────────────────────────────────────
    ws_w = wb.create_sheet("Time series (wide)")
    ws_w.sheet_properties.tabColor = "4A90D9"

    header_w = ["Database", "IAM", "SSP", "Indicator", "Method path", "Unit"] + [
        str(y) for y in years
    ]
    ws_w.append(header_w)
    _sp_style_header(ws_w)

    # Build a (db_name, method_tuple) → {year: score} index for the wide
    # layout. Each (iam, ssp) trajectory typically appears N times (once
    # per year run).
    per_db_year_score: dict[str, dict[tuple[str, ...], dict[int, float]]] = {}
    db_meta: dict[str, tuple[str, str]] = {}
    for r in sorted_runs:
        db_meta[r.db_name] = (r.iam, r.ssp)
        scores = per_db_year_score.setdefault(r.db_name, {})
        for m in r.result.results:
            mkey = tuple(m.method)
            year_dict = scores.setdefault(mkey, {})
            if r.year is not None:
                year_dict[r.year] = m.score

    # Row order: sorted dbs × method_list order from the primary run.
    # When trajectories share the same (iam, ssp), this groups all years
    # for one trajectory together.
    seen_dbs: list[str] = []
    seen_set: set[str] = set()
    for r in sorted_runs:
        if r.db_name not in seen_set:
            seen_set.add(r.db_name)
            seen_dbs.append(r.db_name)

    for db_name in seen_dbs:
        iam, ssp = db_meta[db_name]
        for r in method_list:
            mkey = tuple(r.method)
            row: list = [
                db_name, iam, ssp,
                _sp_short_method(r.method),
                " › ".join(r.method),
                r.unit,
            ]
            year_scores = per_db_year_score.get(db_name, {}).get(mkey, {})
            for y in years:
                row.append(year_scores.get(y, ""))
            ws_w.append(row)

    for col_idx in range(7, 7 + len(years)):
        for row in ws_w.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _SP_SCI_FMT
    ws_w.freeze_panes = "G2"
    _sp_autosize(ws_w)

    # ── Sheet 3: Time series (long) ─────────────────────────────────────
    ws_l = wb.create_sheet("Time series (long)")
    ws_l.sheet_properties.tabColor = "4A90D9"

    ws_l.append([
        "Database", "IAM", "SSP", "Year",
        "Indicator", "Method path", "Score", "Unit",
    ])
    _sp_style_header(ws_l)

    for r in sorted_runs:
        for m in r.result.results:
            ws_l.append([
                r.db_name, r.iam, r.ssp, r.year if r.year is not None else "—",
                _sp_short_method(m.method),
                " › ".join(m.method),
                m.score,
                m.unit,
            ])
            ws_l.cell(row=ws_l.max_row, column=7).number_format = _SP_SCI_FMT
    ws_l.freeze_panes = "A2"
    _sp_autosize(ws_l)

    # ── Sheet 4: Stage breakdown by year (when meaningful) ──────────────
    has_breakdown = scope == "all" and any(
        r.result.stage_breakdown for r in sorted_runs
    )
    if has_breakdown:
        ws_sb = wb.create_sheet("Stage breakdown by year")
        ws_sb.sheet_properties.tabColor = "C29CFF"

        # Discover stage order from the first run with a breakdown.
        stage_order: list[str] = []
        seen_stages: set[str] = set()
        for r in sorted_runs:
            if not r.result.stage_breakdown:
                continue
            for _ml, by_stage in r.result.stage_breakdown.items():
                for stage in by_stage.keys():
                    if stage not in seen_stages:
                        seen_stages.add(stage)
                        stage_order.append(stage)

        header = [
            "Database", "IAM", "SSP", "Year",
            "Indicator", "Method path", "Unit",
        ] + stage_order + ["Total"]
        ws_sb.append(header)
        _sp_style_header(ws_sb)

        for r in sorted_runs:
            if not r.result.stage_breakdown:
                continue
            unit_by_label = {
                _sp_short_method(m.method): m.unit for m in r.result.results
            }
            path_by_label = {
                _sp_short_method(m.method): " › ".join(m.method)
                for m in r.result.results
            }
            for method_label, by_stage in r.result.stage_breakdown.items():
                total = sum(by_stage.values())
                row: list = [
                    r.db_name, r.iam, r.ssp,
                    r.year if r.year is not None else "—",
                    method_label,
                    path_by_label.get(method_label, "—"),
                    unit_by_label.get(method_label, ""),
                ]
                for stage in stage_order:
                    row.append(by_stage.get(stage, 0.0))
                row.append(total)
                ws_sb.append(row)

        first_num = 8  # one past Unit
        last_num = first_num + len(stage_order)  # + Total
        for col_idx in range(first_num, last_num + 1):
            for row in ws_sb.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    cell.number_format = _SP_SCI_FMT
        ws_sb.freeze_panes = "A2"
        _sp_autosize(ws_sb)

    return wb


def _build_single_product_comparison_workbook(
    archetype_name: str,
    scope: str,
    static_result: ArchetypeLCACalculateResult,
    projected_runs: list[SingleProductProspectiveRunPayload],
    stage_amounts_meta=None,
):
    """Build a single-product Comparison workbook.

    Sign convention: ``Δ = P − S`` (matches the on-screen Δ chart from
    Patches 4E/4F). Positive Δ = worsening, negative Δ = improvement.

    Layout:
      - Sheet "Configuration" — archetype, scope, static stage amounts,
        prospective databases.
      - Sheet "Comparison data" — one row per (database, year, method).
        Columns: Static, Projected, Δ, Δ% = (P-S)/|S| × 100.
      - Sheet "Cumulative summary per trajectory" — one row per ((iam,
        ssp), method). Columns: years counted, average Δ, peak Δ year,
        peak Δ value, cumulative Δ over the trajectory.

    The Δ sign flip on the in-app legend (down=improvement) is a
    presentation choice; the workbook leaves Δ raw so analysts can pick
    their own direction. Δ% is the magnitude-normalized read for
    comparing across indicators with very different units.
    """
    from openpyxl import Workbook
    import datetime

    wb = Workbook()
    wb.remove(wb.active)

    # Static methods table — anchor for Δ math.
    static_by_method: dict[tuple[str, ...], float] = {
        tuple(r.method): r.score for r in static_result.results
    }
    static_unit_by_method: dict[tuple[str, ...], str] = {
        tuple(r.method): r.unit for r in static_result.results
    }
    static_label_by_method: dict[tuple[str, ...], str] = {
        tuple(r.method): _sp_short_method(r.method) for r in static_result.results
    }
    static_path_by_method: dict[tuple[str, ...], str] = {
        tuple(r.method): " › ".join(r.method) for r in static_result.results
    }

    sorted_runs = sorted(
        projected_runs,
        key=lambda r: (r.iam, r.ssp, r.year if r.year is not None else 0),
    )

    # ── Sheet 1: Configuration ──────────────────────────────────────────
    ws_cfg = wb.create_sheet("Configuration")
    ws_cfg.sheet_properties.tabColor = "3ECFCF"

    years_set: set[int] = set()
    for r in sorted_runs:
        if r.year is not None:
            years_set.add(r.year)
    years = sorted(years_set)

    cfg_rows: list[tuple[str, Any]] = [
        ("Archetype", archetype_name),
        ("Mode", "Single product · Comparison (Projected − Static)"),
        ("Scope", _SP_SCOPE_LABELS.get(scope, scope)),
        ("Stage amounts", _sp_format_stage_amounts(static_result.stage_amounts)),
        *_sp_stage_amount_meta_rows(stage_amounts_meta),
        ("Year range", f"{years[0]}–{years[-1]}" if years else "—"),
        ("Indicators (Static)", len(static_result.results)),
        ("Prospective runs", len(sorted_runs)),
        ("Sign convention", "Δ = P − S (positive = worsening, negative = improvement)"),
        ("Calculation date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, value in cfg_rows:
        ws_cfg.append([label, value])
        ws_cfg[ws_cfg.max_row][0].font = _SP_META_FONT

    ws_cfg.append([])
    ws_cfg.append(["Database", "IAM", "SSP", "Year"])
    _sp_style_header(ws_cfg, ws_cfg.max_row)
    for r in sorted_runs:
        ws_cfg.append([r.db_name, r.iam, r.ssp, r.year if r.year is not None else "—"])
    _sp_autosize(ws_cfg)

    # ── Sheet 2: Comparison data (long) ─────────────────────────────────
    ws_d = wb.create_sheet("Comparison data")
    ws_d.sheet_properties.tabColor = "4A90D9"

    ws_d.append([
        "Database", "IAM", "SSP", "Year",
        "Indicator", "Method path", "Unit",
        "Static (S)", "Projected (P)", "Δ (P − S)", "Δ %",
    ])
    _sp_style_header(ws_d)

    for r in sorted_runs:
        for m in r.result.results:
            mkey = tuple(m.method)
            s_val = static_by_method.get(mkey)
            if s_val is None:
                # Method present in projected but not in static — rare,
                # surface as blank columns + flag in path so it's
                # findable but doesn't break the sheet.
                ws_d.append([
                    r.db_name, r.iam, r.ssp,
                    r.year if r.year is not None else "—",
                    _sp_short_method(m.method),
                    " › ".join(m.method) + "  (not in Static)",
                    m.unit,
                    "", m.score, "", "",
                ])
                continue
            p_val = m.score
            delta = p_val - s_val
            pct = (delta / abs(s_val) * 100.0) if s_val != 0 else None
            ws_d.append([
                r.db_name, r.iam, r.ssp,
                r.year if r.year is not None else "—",
                _sp_short_method(m.method),
                " › ".join(m.method),
                m.unit,
                s_val, p_val, delta,
                pct if pct is not None else "",
            ])

    # Sci format on S/P/Δ; pct format on Δ%.
    for col_idx in (8, 9, 10):
        for row in ws_d.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _SP_SCI_FMT
    for row in ws_d.iter_rows(min_row=2, min_col=11, max_col=11):
        for cell in row:
            cell.number_format = _SP_PCT_FMT
    ws_d.freeze_panes = "A2"
    _sp_autosize(ws_d)

    # ── Sheet 3: Cumulative summary per trajectory ──────────────────────
    # Sheet name kept ≤ 31 chars (Excel hard limit). The longer phrase
    # "Cumulative summary per trajectory" is in the Configuration sheet.
    ws_s = wb.create_sheet("Cumulative summary per traj.")
    ws_s.sheet_properties.tabColor = "C29CFF"

    ws_s.append([
        "IAM", "SSP", "Indicator", "Method path", "Unit",
        "Years counted", "Average Δ", "Peak Δ year", "Peak Δ",
        "Cumulative Δ",
    ])
    _sp_style_header(ws_s)

    # Group projected runs by (iam, ssp) trajectory.
    traj_runs: dict[tuple[str, str], list[SingleProductProspectiveRunPayload]] = {}
    for r in sorted_runs:
        traj_runs.setdefault((r.iam, r.ssp), []).append(r)

    for (iam, ssp), runs_for_traj in sorted(traj_runs.items()):
        # For each method in static, compute the trajectory's Δ stats.
        for sm in static_result.results:
            mkey = tuple(sm.method)
            deltas: list[tuple[int | None, float]] = []
            for r in runs_for_traj:
                p_match = next(
                    (m for m in r.result.results if tuple(m.method) == mkey),
                    None,
                )
                if p_match is None:
                    continue
                deltas.append((r.year, p_match.score - sm.score))
            if not deltas:
                continue
            d_values = [d for _y, d in deltas]
            n = len(d_values)
            avg = sum(d_values) / n
            cumulative = sum(d_values)
            # Peak Δ = most positive (worsening). Negative Δ = improvement.
            peak_year, peak_val = max(deltas, key=lambda t: t[1])
            ws_s.append([
                iam, ssp,
                _sp_short_method(sm.method),
                " › ".join(sm.method),
                sm.unit,
                n,
                avg,
                peak_year if peak_year is not None else "—",
                peak_val,
                cumulative,
            ])

    for col_idx in (7, 9, 10):  # Average Δ, Peak Δ, Cumulative Δ
        for row in ws_s.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _SP_SCI_FMT
    ws_s.freeze_panes = "A2"
    _sp_autosize(ws_s)

    return wb


def _sp_filename(kind: str, archetype_name: str) -> str:
    """Filename pattern matches the per-axis system-mode exports —
    ``MApper_Impact_<Axis>_<discriminator>_<date>.xlsx``. ``kind`` is
    the per-tab discriminator (``SingleProduct_Static`` etc.); the
    archetype name is sanitised so spaces and slashes don't escape into
    the filename."""
    import datetime
    safe = _sanitize_filename(archetype_name, "archetype")
    date_tag = datetime.date.today().isoformat()
    return f"MApper_Impact_{kind}_{safe}_{date_tag}.xlsx"


def _sp_xlsx_response(wb, filename: str) -> Response:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/export-single-product-static")
async def post_export_single_product_static(
    body: SingleProductStaticExportRequest,
) -> Response:
    if not body.scenarios:
        raise HTTPException(
            status_code=400,
            detail="scenarios must contain at least one entry",
        )
    wb = _build_single_product_static_workbook(
        archetype_name=body.archetype_name,
        scope=body.scope,
        scenarios=[(s.label, s.result) for s in body.scenarios],
        stage_amounts_meta=body.stage_amounts_meta,
    )
    filename = _sp_filename("SingleProduct_Static", body.archetype_name)
    return _sp_xlsx_response(wb, filename)


@router.post("/export-single-product-prospective")
async def post_export_single_product_prospective(
    body: SingleProductProspectiveExportRequest,
) -> Response:
    if not body.runs:
        raise HTTPException(
            status_code=400,
            detail="runs must contain at least one entry",
        )
    wb = _build_single_product_prospective_workbook(
        archetype_name=body.archetype_name,
        scope=body.scope,
        runs=body.runs,
        stage_amounts_meta=body.stage_amounts_meta,
    )
    filename = _sp_filename("SingleProduct_Prospective", body.archetype_name)
    return _sp_xlsx_response(wb, filename)


@router.post("/export-single-product-comparison")
async def post_export_single_product_comparison(
    body: SingleProductComparisonExportRequest,
) -> Response:
    if not body.projected_runs:
        raise HTTPException(
            status_code=400,
            detail=(
                "projected_runs must contain at least one entry — "
                "Comparison requires both sides to have results."
            ),
        )
    wb = _build_single_product_comparison_workbook(
        archetype_name=body.archetype_name,
        scope=body.scope,
        static_result=body.static_result,
        projected_runs=body.projected_runs,
        stage_amounts_meta=body.stage_amounts_meta,
    )
    filename = _sp_filename("SingleProduct_Comparison", body.archetype_name)
    return _sp_xlsx_response(wb, filename)


@router.websocket("/ws/{task_id}")
async def ws_progress(websocket: WebSocket, task_id: str) -> None:
    await websocket.accept()
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None:
        await websocket.send_json({"type": "error", "error": "Unknown task id"})
        await websocket.close()
        return

    queue: asyncio.Queue = asyncio.Queue(maxsize=256)
    task.subscribers.append(queue)

    await websocket.send_json({"type": "progress", "stage": task.stage, "pct": task.pct})
    if task.done:
        if task.cancelled:
            await websocket.send_json({"type": "cancelled", "task_id": task_id})
        elif task.error:
            await websocket.send_json({"type": "error", "error": task.error})
        else:
            res = task.result
            if isinstance(res, MultiScenarioProjectedImpactResult):
                await websocket.send_json({
                    "type": "done",
                    "result_type": "multi_scenario_projected",
                    "scenarios_calculated": len(res.scenarios),
                    "elapsed_seconds": res.elapsed_seconds,
                })
            else:
                await websocket.send_json({
                    "type": "done",
                    "methods_calculated": len(res.results) if res else 0,
                    "year_to_database": res.meta.year_to_database if res else {},
                })
        await websocket.close()
        return

    try:
        while True:
            payload = await queue.get()
            await websocket.send_json(payload)
            if payload.get("type") in ("done", "error", "cancelled"):
                break
    except WebSocketDisconnect:
        pass
    finally:
        try:
            task.subscribers.remove(queue)
        except ValueError:
            pass
        task_registry.maybe_cancel_on_last_subscriber_leave(
            task_id,
            remaining_subscribers=len(task.subscribers),
            task_done=task.done,
        )
        try:
            await websocket.close()
        except Exception:
            pass


# ── Multi-product LCA comparison export (Patch 4AG.4) ──────────────────────────


def _mp_unique_method_labels(result: MultiProductLCAResult) -> list[tuple[str, str]]:
    """Walk successful items and collect unique (method_label, unit)
    pairs in first-seen order. Failed items contribute nothing. The
    `unit` is paired so the wide-shape sheet headers can show
    "method (unit)" for unambiguous readout — methods with the same
    label across different LCIA packages would otherwise collide."""
    seen: dict[str, str] = {}
    order: list[str] = []
    for item in result.items:
        if item.status != "success":
            continue
        method_results = (
            item.archetype_result.results if item.archetype_result else
            item.activity_result.results if item.activity_result else []
        )
        for m in method_results:
            if m.method_label not in seen:
                seen[m.method_label] = m.unit
                order.append(m.method_label)
    return [(label, seen[label]) for label in order]


def _build_multi_product_workbook(body: MultiProductExportRequest):
    """Build the comparison workbook for a multi-product LCA run.

    Sheets (in order):
      1. Configuration — items list + shared compute settings + timestamp
      1b. Stage amounts — per archetype item: preset + lifetime + per-stage
         amounts (Patch 5J, reproducibility). Emitted when ≥1 archetype item.
      2. Comparison (wide) — rows=items, columns=methods; one row per item
         including failed items (which show "—" for method values plus an
         "error_message" trailing column)
      3. Comparison (long) — one row per (item, method); for downstream
         tooling (pandas, R) that prefers tidy long format
      4. Stage breakdown sheets — one per ARCHETYPE item that carries
         `stage_breakdown` (i.e. scope='all' archetype items). Only
         emitted when at least one archetype item has stage data.
      5. Errors — one row per FAILED item with type, label, error_message.
         Only emitted when result.error_count > 0.

    Filename pattern: ``MApper_MultiProduct_Comparison_<date>.xlsx``.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment

    result = body.result
    method_pairs = _mp_unique_method_labels(result)  # [(label, unit), ...]
    method_labels = [m for m, _ in method_pairs]

    wb = Workbook()

    # ── 1. Configuration sheet ─────────────────────────────────────
    ws_cfg = wb.active
    ws_cfg.title = "Configuration"
    ws_cfg.append(["MApper Multi-Product LCA Comparison"])
    ws_cfg.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_cfg.append([])
    # Meta block — one key per row.
    import datetime
    computed_at = body.computed_at or datetime.datetime.now(datetime.timezone.utc).isoformat()
    meta_rows: list[tuple[str, str]] = [
        ("Computed at", computed_at),
        ("Scope", body.scope),
        ("Compute database", body.compute_database or "base ecoinvent"),
        ("Items requested", str(len(result.items))),
        ("Items succeeded", str(result.success_count)),
        ("Items failed", str(result.error_count)),
        ("Methods", ", ".join(method_labels) if method_labels else "—"),
        ("Compute elapsed (s)", f"{result.elapsed_seconds:.2f}"),
    ]
    for k, v in meta_rows:
        ws_cfg.append([k, v])
        ws_cfg.cell(row=ws_cfg.max_row, column=1).font = _SP_META_FONT
    # Items table
    ws_cfg.append([])
    ws_cfg.append(["Items"])
    ws_cfg.cell(row=ws_cfg.max_row, column=1).font = Font(bold=True, size=12)
    items_header_row = ws_cfg.max_row + 1
    ws_cfg.append(["#", "Type", "Identifier", "Label", "Status", "Notes"])
    _sp_style_header(ws_cfg, items_header_row)
    for i, item in enumerate(result.items, start=1):
        notes = ""
        if item.type == "archetype" and item.archetype_result:
            stages = item.archetype_result.stages_included or []
            notes = "stages: " + ", ".join(stages) if stages else ""
        elif item.type == "activity":
            notes = "single-activity demand"
        ws_cfg.append([
            i, item.type, item.item_id, item.label, item.status, notes,
        ])
    _sp_autosize(ws_cfg)

    # ── 1b. Stage amounts sheet (per archetype item) ───────────────
    # Reproducibility (Patch 5J): records preset + lifetime + per-stage
    # amounts for every archetype item so the run can be reconstructed from
    # the export alone. Activities have no BOM stages → skipped. preset and
    # lifetime come from the threaded `stage_amounts_meta` (frontend-only
    # concepts); the per-stage amounts fall back to the result-echoed
    # `stage_amounts` when no meta is supplied (older clients). Emitted only
    # when at least one archetype item exists. Placed right after
    # Configuration (sheets are accessed by name, so order is non-load-bearing).
    meta_map = body.stage_amounts_meta or {}
    arch_items = [it for it in result.items if it.type == "archetype"]
    if arch_items:
        # Resolved per-item amounts: prefer the threaded meta, else the
        # result's echoed amounts. Drives both the column set and the cells.
        def _item_amounts(it) -> dict[str, float]:
            m = meta_map.get(it.item_id)
            if m and m.amounts:
                return m.amounts
            if it.archetype_result and it.archetype_result.stage_amounts:
                return it.archetype_result.stage_amounts
            return {}

        # Stage columns = union across items, first-seen order (lifecycle
        # order is preserved within each item's insertion-ordered dict).
        stage_cols: list[str] = []
        for it in arch_items:
            for stage in _item_amounts(it).keys():
                if stage not in stage_cols:
                    stage_cols.append(stage)

        ws_sa = wb.create_sheet("Stage amounts", index=1)
        ws_sa.append(["Stage amounts (per item)"])
        ws_sa.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws_sa.append([])
        sa_header_row = ws_sa.max_row + 1
        ws_sa.append(["#", "Item", "Preset", "Lifetime (yr)", *stage_cols])
        _sp_style_header(ws_sa, sa_header_row)
        for i, it in enumerate(arch_items, start=1):
            m = meta_map.get(it.item_id)
            preset = m.preset if m else "—"
            lifetime = m.lifetime if m else "—"
            amounts = _item_amounts(it)
            row = [i, it.label, preset, lifetime, *[amounts.get(s, "") for s in stage_cols]]
            ws_sa.append(row)
        _sp_autosize(ws_sa)

    # ── 1c. Vintages sheet (per activity item) ─────────────────────
    # Per-item-vintage provenance: each activity item is computed against ITS
    # OWN database (base ecoinvent or a premise SSP×year vintage). This records
    # which DB/SSP/year each item used so the run is reproducible from the
    # export alone — mirrors the Stage amounts sheet (5J) for activity mode.
    # preset-like display fields (label/iam/ssp/year) come from the threaded
    # `activity_vintage_meta`; the database falls back to the item_id when no
    # meta is supplied (older clients). Emitted only when ≥1 activity item.
    vintage_meta = body.activity_vintage_meta or {}
    act_items = [it for it in result.items if it.type == "activity"]
    if act_items:
        ws_v = wb.create_sheet("Vintages", index=(2 if arch_items else 1))
        ws_v.append(["Vintages (per activity item)"])
        ws_v.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws_v.append([])
        v_header_row = ws_v.max_row + 1
        ws_v.append(["#", "Item", "Vintage", "Database", "Base database", "IAM", "SSP", "Year"])
        _sp_style_header(ws_v, v_header_row)
        for i, it in enumerate(act_items, start=1):
            m = vintage_meta.get(it.item_id)
            # Fallback: derive the database from the item_id ("{database}|{code}").
            db_from_id = it.item_id.split("|", 1)[0] if "|" in it.item_id else it.item_id
            ws_v.append([
                i,
                it.label,
                (m.label if m else "—") or "—",
                (m.database if m and m.database else db_from_id),
                (m.base_database if m else None) or "—",
                (m.iam if m else None) or "—",
                (m.ssp if m else None) or "—",
                (m.year if m and m.year is not None else "—"),
            ])
        _sp_autosize(ws_v)

    # ── 2. Comparison (wide) ───────────────────────────────────────
    ws_wide = wb.create_sheet("Comparison (wide)")
    header = ["#", "Type", "Item"] + [f"{lbl} ({unit})" for lbl, unit in method_pairs] + ["Error"]
    ws_wide.append(header)
    _sp_style_header(ws_wide)
    for i, item in enumerate(result.items, start=1):
        method_results = (
            item.archetype_result.results if item.archetype_result else
            item.activity_result.results if item.activity_result else []
        )
        by_label = {m.method_label: m.score for m in method_results}
        row = [i, item.type, item.label]
        for label, _unit in method_pairs:
            v = by_label.get(label)
            row.append(v if v is not None else "—")
        row.append(item.error_message or "")
        ws_wide.append(row)
        # Format method-score cells in scientific notation. Skip
        # non-numeric "—" cells (failed items contribute strings).
        for col_idx in range(4, 4 + len(method_pairs)):
            cell = ws_wide.cell(row=ws_wide.max_row, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.000E+00"
    _sp_autosize(ws_wide)

    # ── 3. Comparison (long) ───────────────────────────────────────
    ws_long = wb.create_sheet("Comparison (long)")
    ws_long.append(["Item", "Type", "Method", "Score", "Unit"])
    _sp_style_header(ws_long)
    for item in result.items:
        if item.status != "success":
            continue
        method_results = (
            item.archetype_result.results if item.archetype_result else
            item.activity_result.results if item.activity_result else []
        )
        for m in method_results:
            ws_long.append([item.label, item.type, m.method_label, m.score, m.unit])
            ws_long.cell(row=ws_long.max_row, column=4).number_format = "0.000E+00"
    _sp_autosize(ws_long)

    # ── 4. Stage breakdown sheets (per archetype item, when present) ─
    # One sheet per archetype item that carries a `stage_breakdown`
    # (i.e. scope='all' archetype items per Patch 4B). Skipped when
    # no item carries stage data — keeps the workbook minimal.
    for item in result.items:
        if item.type != "archetype" or not item.archetype_result:
            continue
        sb = item.archetype_result.stage_breakdown
        if not sb:
            continue
        # Stable sheet name: prefix "SB_" + label, truncated to 27
        # chars (openpyxl limit is 31; allow 4 for prefix).
        sheet_name = f"SB_{item.label}"[:31]
        # Avoid duplicate sheet names across items with similar labels.
        n = 1
        base = sheet_name
        while sheet_name in wb.sheetnames:
            n += 1
            sheet_name = f"{base[:27]}_{n}"
        ws_sb = wb.create_sheet(sheet_name)
        ws_sb.append([f"Stage breakdown — {item.label}"])
        ws_sb.cell(row=1, column=1).font = Font(bold=True, size=12)
        ws_sb.append([])
        # Collect all stages across methods in first-seen order.
        stage_order: list[str] = []
        for stages in sb.values():
            for stage in stages.keys():
                if stage not in stage_order:
                    stage_order.append(stage)
        header_row = ["Method", "Unit"] + stage_order + ["Total"]
        ws_sb.append(header_row)
        _sp_style_header(ws_sb, ws_sb.max_row)
        method_results = item.archetype_result.results
        for m in method_results:
            method_stages = sb.get(m.method_label, {})
            stage_values = [method_stages.get(s, 0.0) for s in stage_order]
            total = sum(stage_values)
            ws_sb.append([m.method_label, m.unit, *stage_values, total])
            row_idx = ws_sb.max_row
            for col_idx in range(3, 3 + len(stage_order) + 1):
                ws_sb.cell(row=row_idx, column=col_idx).number_format = "0.000E+00"
        _sp_autosize(ws_sb)

    # ── 5. Errors sheet (when partial / total failure) ─────────────
    if result.error_count > 0:
        ws_err = wb.create_sheet("Errors")
        ws_err.append(["Item", "Type", "Identifier", "Error message"])
        _sp_style_header(ws_err)
        for item in result.items:
            if item.status != "error":
                continue
            ws_err.append([
                item.label, item.type, item.item_id,
                item.error_message or "(no detail)",
            ])
            ws_err.cell(row=ws_err.max_row, column=4).alignment = Alignment(wrap_text=True)
        _sp_autosize(ws_err)

    return wb


@router.post("/export-multi-product")
async def post_export_multi_product(body: MultiProductExportRequest) -> Response:
    """Export endpoint for the multi-product LCA comparison (Patch
    4AG.4). Accepts the full result envelope plus compute-config
    metadata; returns the workbook as an xlsx download."""
    if not body.result.items:
        raise HTTPException(
            status_code=400,
            detail="result.items must contain at least one entry",
        )
    wb = _build_multi_product_workbook(body)
    import datetime
    date_tag = datetime.date.today().isoformat()
    filename = f"MApper_MultiProduct_Comparison_{date_tag}.xlsx"
    return _sp_xlsx_response(wb, filename)
