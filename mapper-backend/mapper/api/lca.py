import asyncio
import datetime
import io
import logging
import threading
import time
import uuid
from typing import Any, NamedTuple

import bw2data

logger = logging.getLogger(__name__)
from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.responses import Response

from mapper.core.bw2_wrapper import (
    MultiDBPersistentRunner,
    PersistentLCARunner,
    get_biosphere_contributions,
    get_contributions,
    get_recursive_contribution_tree,
    get_supply_chain,
    parse_activity_key,
    run_lca,
)
import bw2calc
from mapper.core.bom_engine import (
    filter_roots_by_scope,
    flatten_roots,
    flatten_roots_for_scope,
    stage_to_scope,
    stages_in_scope,
)
from mapper.core.dsm_lca_engine import (
    blend_method_scores,
    resolve_bracket,
    resolve_database_for_year,
)
from mapper.core.tasks import Task, create_task, get_task, run_in_thread
from mapper.api import tasks as task_registry
from mapper.api.tasks import CancelledOperation
from mapper.models.schemas import (
    ActivityContribution,
    CancelledTaskResponse,
    ActivityDemandItem,
    ActivityLCAMethodResult,
    ActivityLCARequest,
    ActivityLCAResult,
    ArchetypeLCACalculateRequest,
    ArchetypeLCACalculateResult,
    ArchetypeLCAExportRequest,
    ArchetypeLCAMethodResult,
    ArchetypeTrajectoryMethodScore,
    ArchetypeTrajectoryRequest,
    ArchetypeTrajectoryResult,
    ArchetypeTrajectoryYear,
    BiosphereContributionItem,
    ContributionAnalysisExportRequest,
    ContributionAnalysisRequest,
    ContributionAnalysisResult,
    ContributionTreeNode,
    ContributionsResponse,
    LCACalculateRequest,
    LCAResult,
    MaterialContribution,
    MultiProductItemResult,
    MultiProductLCARequest,
    MultiProductLCAResult,
    MultiYearContributionExportRequest,
    MultiYearContributionRequest,
    MultiYearContributionResult,
    MultiYearContributionTaskStarted,
    MultiYearEvolutionItem,
    MultiYearTrajectoryPoint,
    SankeyData,
    StageContribution,
    TaskStartedResponse,
)
from mapper.ws.progress import stream_task_progress

router = APIRouter()

# In-memory store for LCA results keyed by task_id
_lca_results: dict[str, dict] = {}


def _lca_worker(
    task: Task,
    functional_unit_key: str,
    amount: float,
    method_tuple: list[str],
) -> None:
    if task_registry.is_cancelled(task.task_id):
        raise CancelledOperation(task.task_id)
    task.update("building_matrix", 0.1, "Building technosphere matrix…")
    result = run_lca(functional_unit_key, amount, method_tuple)

    if task_registry.is_cancelled(task.task_id):
        raise CancelledOperation(task.task_id)
    task.update("solving", 0.4, "Solving linear system…")
    # LCI already done inside run_lca

    task.update("characterizing", 0.6, "Characterizing impacts…")
    # LCIA already done inside run_lca

    if task_registry.is_cancelled(task.task_id):
        raise CancelledOperation(task.task_id)
    task.update("analyzing", 0.8, "Analyzing contributions…")
    lca_obj = result.pop("lca_object")
    activity_key = result.pop("activity_key")

    # Contribution analysis can fail for certain activities (biosphere flows,
    # zero-score results, etc.).  Return the score even if analysis fails.
    try:
        contributions = get_contributions(lca_obj, result["score"])
    except Exception:
        contributions = {"items": [], "rest_amount": result["score"], "rest_percentage": 100.0}
    try:
        supply_chain = get_supply_chain(lca_obj, method=method_tuple, depth=3)
    except Exception:
        logger.exception("legacy LCA: get_supply_chain failed")
        supply_chain = {"nodes": [], "links": [], "total_nodes_discovered": 0, "truncated": False}

    _lca_results[task.task_id] = {
        "result": {
            "task_id": task.task_id,
            "method": method_tuple,
            "functional_unit_name": result["functional_unit_name"],
            "functional_unit_amount": amount,
            "score": result["score"],
            "unit": result["unit"],
            "calculated_at": datetime.datetime.now().isoformat(),
        },
        "contributions": contributions,
        "supply_chain": supply_chain,
    }

    task.update("done", 1.0, "Calculation complete.")


@router.post("/lca/calculate", response_model=TaskStartedResponse)
async def calculate_lca(body: LCACalculateRequest) -> TaskStartedResponse:
    try:
        # Validate key exists
        key = parse_activity_key(body.functional_unit.key)
        act = bw2data.get_activity(key)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid functional unit: {e}")

    # Biosphere flows cannot be used as functional units
    if "biosphere" in str(key[0]).lower():
        raise HTTPException(
            status_code=400,
            detail="Biosphere flows cannot be used as functional units. Select a technosphere activity from ecoinvent.",
        )

    task = create_task()
    task_registry.register(task.task_id)

    def _wrapped(t: Task, *a, **kw):
        try:
            return _lca_worker(t, *a, **kw)
        finally:
            task_registry.unregister(t.task_id)

    run_in_thread(
        task,
        _wrapped,
        body.functional_unit.key,
        body.functional_unit.amount,
        body.method,
    )
    return TaskStartedResponse(task_id=task.task_id, status="started")


@router.websocket("/ws/lca/{task_id}")
async def ws_lca_progress(websocket: WebSocket, task_id: str) -> None:
    await stream_task_progress(websocket, task_id)


@router.get(
    "/lca/results/{task_id}",
    response_model=LCAResult | CancelledTaskResponse,
)
async def get_lca_result(task_id: str) -> LCAResult | CancelledTaskResponse:
    data = _lca_results.get(task_id)
    if not data:
        task = get_task(task_id)
        if task and task.status == "cancelled":
            return CancelledTaskResponse(task_id=task_id)
        if task and task.status == "error":
            raise HTTPException(status_code=500, detail=task.error)
        raise HTTPException(status_code=404, detail="Result not ready or not found")
    return LCAResult(**data["result"])


@router.get("/lca/results/{task_id}/contributions", response_model=ContributionsResponse)
async def get_lca_contributions(task_id: str, limit: int = 10) -> ContributionsResponse:
    data = _lca_results.get(task_id)
    if not data:
        raise HTTPException(status_code=404, detail="Result not found")
    contrib = data["contributions"]
    return ContributionsResponse(**contrib)


@router.get("/lca/results/{task_id}/supply-chain", response_model=SankeyData)
async def get_supply_chain_data(task_id: str) -> SankeyData:
    data = _lca_results.get(task_id)
    if not data:
        raise HTTPException(status_code=404, detail="Result not found")
    return SankeyData(**data["supply_chain"])


# ── Multi-Activity LCA Calculator ──────────────────────────────────────────────


@router.post("/lca/calculate-activities", response_model=ActivityLCAResult)
async def calculate_activity_lca(body: ActivityLCARequest) -> ActivityLCAResult:
    """LCA for one or more technosphere activities with per-activity contribution."""
    t0 = time.perf_counter()

    if not body.activities:
        raise HTTPException(status_code=400, detail="At least one activity is required")
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required")

    # Validate all activities exist and are technosphere
    activity_meta: dict[tuple[str, str], dict] = {}
    for item in body.activities:
        if "biosphere" in item.database.lower():
            raise HTTPException(
                status_code=400,
                detail=f"Biosphere flows cannot be used as functional units. "
                       f"Remove '{item.code}' from the list.",
            )
        key = (item.database, item.code)
        try:
            act = bw2data.get_activity(key)
        except Exception:
            raise HTTPException(status_code=400, detail=f"Activity not found: {key}")
        activity_meta[key] = {
            "name": act.get("reference product", act.get("name", "")),
            "location": str(act.get("location", "")),
            "unit": act.get("unit", ""),
        }

    method_tuples = [tuple(ml) for ml in body.methods]

    # Build total demand
    total_demand: dict[tuple[str, str], float] = {}
    for item in body.activities:
        key = (item.database, item.code)
        total_demand[key] = total_demand.get(key, 0.0) + item.amount

    runner = PersistentLCARunner()

    try:
        # Total scores — 1 factorization + method switches
        total_scores = runner(total_demand, method_tuples)

        # Per-activity scores — reuses UMFPACK factorization
        activity_scores: dict[tuple[str, str], dict[tuple, float]] = {}
        for act_key, act_amount in total_demand.items():
            single = {act_key: act_amount}
            scores = runner(single, method_tuples)
            activity_scores[act_key] = {mt: sc for mt, (sc, _u) in scores.items()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LCA calculation failed: {e}")

    results: list[ActivityLCAMethodResult] = []
    for method_list, mt in zip(body.methods, method_tuples):
        total_score, unit = total_scores[mt]

        contribs: list[ActivityContribution] = []
        for act_key, act_amount in total_demand.items():
            impact = activity_scores.get(act_key, {}).get(mt, 0.0)
            if abs(impact) < 1e-20:
                continue
            meta = activity_meta[act_key]
            contribs.append(ActivityContribution(
                name=meta["name"],
                location=meta["location"],
                database=act_key[0],
                code=act_key[1],
                demand_amount=act_amount,
                demand_unit=meta["unit"],
                impact=impact,
                percentage=(abs(impact) / abs(total_score) * 100) if total_score else 0.0,
            ))

        contribs.sort(key=lambda c: abs(c.impact), reverse=True)
        label = method_list[-1] if method_list else ""

        results.append(ActivityLCAMethodResult(
            method=method_list,
            method_label=label,
            score=total_score,
            unit=unit,
            contributions=contribs,
        ))

    elapsed = round(time.perf_counter() - t0, 2)
    return ActivityLCAResult(results=results, elapsed_seconds=elapsed)


# ── Archetype LCA Calculator ────────────────────────────────────────────────


class _ArchetypeDemand(NamedTuple):
    """Source-DB demand bundle shared by the discrete + continuous-horizon
    single-product paths."""
    arc: Any
    stages: list[str]
    effective_amounts: dict[str, float]
    linked: list
    method_tuples: list[tuple]
    total_demand: dict[tuple[str, str], float]


def _build_archetype_source_demand(
    *,
    archetype_id: str,
    scope: str,
    amount: float,
    stage_amounts: dict[str, float],
    methods: list[list[str]],
    parameter_scenario: str | None,
) -> _ArchetypeDemand:
    """Shared source-DB demand builder for the single-product archetype LCA
    paths (discrete ``calculate_archetype_lca`` + the continuous-horizon
    trajectory endpoint). Resolves the archetype (optionally through a parameter
    scenario), filters by scope, applies per-stage amounts, flattens, and
    aggregates the linked materials into a source-DB-keyed ``(db, code) →
    amount`` demand. Behavior-preserving extraction — raises the same
    HTTPExceptions as the inline code it replaced; the discrete path's numbers
    stay byte-identical (guarded by the existing archetype-LCA tests)."""
    from mapper.api.bom import _get_archetype
    from mapper.api.parameters import _table_for
    from mapper.core.bom_engine import flatten_bom, resolve_archetype_with_engine
    from mapper.core.parameter_engine import ParameterEngine, ParameterError

    arc = _get_archetype(archetype_id)  # raises 404 if not found

    if not methods or len(methods) == 0:
        raise HTTPException(status_code=400, detail="At least one method is required")

    if scope not in ("inflows", "stock", "outflows", "all"):
        raise HTTPException(status_code=400, detail=f"Invalid scope: {scope}")

    # Resolve parameter expressions in the BOM if a scenario is requested.
    # Single-product mode in Impact Assessment uses this for parameter
    # sensitivity fan-out. Backward compat: scenario None → table's base
    # values, identical to pre-Patch behavior.
    if parameter_scenario is not None:
        table = _table_for()
        if parameter_scenario not in (None, "Base") and parameter_scenario not in table.list_scenarios():
            raise HTTPException(
                status_code=400,
                detail=f"Parameter scenario '{parameter_scenario}' not found in active table",
            )
        try:
            engine = ParameterEngine(table, scenario=parameter_scenario)
            arc = resolve_archetype_with_engine(arc, engine)
        except ParameterError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Parameter resolution failed: {e}",
            )

    # Filter roots by scope
    scope_roots = filter_roots_by_scope(arc.bom, scope)
    stages = [r.name for r in scope_roots]

    # Resolve per-stage amounts.  When stage_amounts is provided, each stage
    # gets its own multiplier (e.g. Manufacturing=1, Use Phase=15).
    # Falls back to the flat `amount` field for backward compatibility.
    effective_amounts: dict[str, float] = {}
    if stage_amounts:
        for r in scope_roots:
            effective_amounts[r.name] = stage_amounts.get(r.name, 1.0)
    else:
        for r in scope_roots:
            effective_amounts[r.name] = amount

    # Flatten each stage separately and apply its amount multiplier
    all_materials = []
    for root in scope_roots:
        flat = flatten_bom(root)
        stage_amt = effective_amounts.get(root.name, 1.0)
        for m in flat:
            m._stage_amount = stage_amt  # type: ignore[attr-defined]
            all_materials.append(m)

    # Collect linked materials
    linked = [m for m in all_materials if m.ecoinvent_activity is not None]
    if not linked:
        raise HTTPException(
            status_code=400,
            detail="No linked materials in this scope. Link ecoinvent activities to materials first.",
        )

    method_tuples = [tuple(ml) for ml in methods]

    # Build total demand: (db, code) → amount, applying per-stage multipliers.
    # Keys are kept in source-DB form here; `_translate_demand_to_database`
    # re-keys to a prospective DB at the call site if requested.
    total_demand: dict[tuple[str, str], float] = {}
    for m in linked:
        key = (m.ecoinvent_activity.database, m.ecoinvent_activity.code)  # type: ignore[union-attr]
        stage_amt = m._stage_amount  # type: ignore[attr-defined]
        total_demand[key] = total_demand.get(key, 0.0) + m.quantity * stage_amt

    return _ArchetypeDemand(arc, stages, effective_amounts, linked, method_tuples, total_demand)


@router.post("/lca/calculate-archetype", response_model=ArchetypeLCACalculateResult)
async def calculate_archetype_lca(body: ArchetypeLCACalculateRequest) -> ArchetypeLCACalculateResult:
    t0 = time.perf_counter()

    bundle = _build_archetype_source_demand(
        archetype_id=body.archetype_id,
        scope=body.scope,
        amount=body.amount,
        stage_amounts=body.stage_amounts,
        methods=body.methods,
        parameter_scenario=body.parameter_scenario,
    )
    arc = bundle.arc
    stages = bundle.stages
    effective_amounts = bundle.effective_amounts
    linked = bundle.linked
    method_tuples = bundle.method_tuples
    total_demand = bundle.total_demand

    # Build a source→translated key map. Each source (db, code) pair gets
    # mapped once via `_translate_demand_to_database`, which emits any
    # missing-activity warnings. The map drives both the aggregate runner
    # call and the per-material contribution share loop below.
    translation_map: dict[tuple[str, str], tuple[str, str]] = {}
    warnings: list[str] = []
    if body.compute_database:
        single_keys = {k: 1.0 for k in total_demand.keys()}
        translated_singles, warnings = _translate_demand_to_database(single_keys, body.compute_database)
        # `_translate_demand_to_database` aggregates by translated key. To
        # recover the source→translated correspondence (the helper drops
        # the source key after translation), iterate per-key with warnings
        # silenced to avoid duplicate emission.
        for src_key in total_demand.keys():
            one, _ = _translate_demand_to_database({src_key: 1.0}, body.compute_database)
            translation_map[src_key] = next(iter(one.keys()))
        del translated_singles  # bulk-translated set used only for warning emission
    else:
        for src_key in total_demand.keys():
            translation_map[src_key] = src_key

    # Aggregate translated demand for the totals run.
    translated_demand: dict[tuple[str, str], float] = {}
    for src_key, amount in total_demand.items():
        tkey = translation_map[src_key]
        translated_demand[tkey] = translated_demand.get(tkey, 0.0) + amount

    runner = PersistentLCARunner()

    try:
        total_scores = runner(translated_demand, method_tuples)

        activity_scores: dict[tuple[str, str], dict[tuple, float]] = {}
        for src_key, act_amount in total_demand.items():
            tkey = translation_map[src_key]
            single = {tkey: act_amount}
            scores = runner(single, method_tuples)
            activity_scores[src_key] = {mt: sc for mt, (sc, _u) in scores.items()}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"LCA calculation failed: {e}")

    # Assemble results per method
    results: list[ArchetypeLCAMethodResult] = []
    # Stage breakdown is populated only when scope == "all" — specific-stage
    # scopes already filter to one stage so a breakdown would be redundant.
    # Shape: {method_label: {stage_name: score}}.
    stage_breakdown: dict[str, dict[str, float]] | None = (
        {} if body.scope == "all" else None
    )
    for method_list, mt in zip(body.methods, method_tuples):
        total_score, unit = total_scores[mt]

        contribs: list[MaterialContribution] = []
        per_stage: dict[str, float] = {}
        for m in linked:
            key = (m.ecoinvent_activity.database, m.ecoinvent_activity.code)  # type: ignore[union-attr]
            stage_amt = m._stage_amount  # type: ignore[attr-defined]
            act_score = activity_scores.get(key, {}).get(mt, 0.0)
            # Share proportionally among materials sharing same activity
            same_act_qty = sum(
                x.quantity * x._stage_amount  # type: ignore[attr-defined]
                for x in linked
                if (x.ecoinvent_activity.database, x.ecoinvent_activity.code) == key  # type: ignore[union-attr]
            )
            share = (m.quantity * stage_amt / same_act_qty) if same_act_qty > 0 else 0.0
            impact = act_score * share
            stage_name = m.path[0] if m.path else ""
            if stage_breakdown is not None and stage_name:
                per_stage[stage_name] = per_stage.get(stage_name, 0.0) + impact
            if abs(impact) < 1e-20:
                continue
            contribs.append(MaterialContribution(
                name=m.name,
                stage=stage_name,
                component=m.path[1] if len(m.path) > 2 else "",
                quantity=m.quantity * stage_amt,
                unit=m.unit,
                impact=impact,
                percentage=(abs(impact) / abs(total_score) * 100) if total_score else 0.0,
            ))

        contribs.sort(key=lambda c: abs(c.impact), reverse=True)
        label = method_list[-1] if method_list else ""

        if stage_breakdown is not None:
            stage_breakdown[label] = per_stage

        results.append(ArchetypeLCAMethodResult(
            method=method_list,
            method_label=label,
            score=total_score,
            unit=unit,
            contributions=contribs,
        ))

    elapsed = round(time.perf_counter() - t0, 2)

    return ArchetypeLCACalculateResult(
        archetype_id=body.archetype_id,
        archetype_name=arc.name,
        scope=body.scope,
        amount=body.amount,
        stage_amounts=effective_amounts,
        stages_included=stages,
        results=results,
        elapsed_seconds=elapsed,
        compute_database=body.compute_database,
        parameter_scenario=body.parameter_scenario,
        warnings=warnings,
        stage_breakdown=stage_breakdown,
    )


# ── Single-product continuous-horizon trajectory (Stage B.1) ─────────────────


def _trajectory_year_scores(
    *,
    total_demand: dict[tuple[str, str], float],
    method_tuples: list[tuple],
    anchors: list[tuple[str, int]],
    temporal_mode: str,
    runner: Any,
    year_start: int | None = None,
    year_end: int | None = None,
    translate=None,
) -> tuple[list[tuple[int, dict[tuple, tuple[float, str]]]], list[str]]:
    """Per-year TOTALS across a prospective trajectory's anchor span.

    Steps annually over ``min..max`` anchor year (optionally narrowed to
    ``[year_start, year_end]`` but never beyond the span — no extrapolation,
    mirroring the system-level path). For each year:

    - ``block``       → nearest-earlier anchor DB (``resolve_database_for_year``),
      one solve.
    - ``interpolate`` → ``resolve_bracket``: at an anchor (or clamped at a span
      end) it's a single solve; strictly between anchors it blends the two
      bracketing solves via ``blend_method_scores`` (the shared core helper, so
      block == interpolate AT anchors, and the curve passes through the discrete
      single-DB values).

    ``runner`` is a ``MultiDBPersistentRunner`` (each anchor DB factorized once);
    ``translate`` re-keys the source demand to an anchor DB. Both are injectable
    so the logic is testable with a fake runner (no real bw2 solves). Returns
    ``([(year, {method: (score, unit)}), ...], warnings)``."""
    if translate is None:
        translate = _translate_demand_to_database
    if not anchors:
        return [], []
    years_sorted = sorted(y for _, y in anchors)
    lo, hi = years_sorted[0], years_sorted[-1]
    start = lo if year_start is None else max(lo, year_start)
    end = hi if year_end is None else min(hi, year_end)

    warnings: list[str] = []
    seen_warn: set[str] = set()

    def _solve_for_db(db_name: str) -> dict[tuple, tuple[float, str]]:
        translated, warns = translate(total_demand, db_name)
        for w in warns:
            if w not in seen_warn:
                seen_warn.add(w)
                warnings.append(w)
        return runner(translated, method_tuples)

    out: list[tuple[int, dict[tuple, tuple[float, str]]]] = []
    for year in range(start, end + 1):
        if temporal_mode == "block":
            picked = resolve_database_for_year(year, anchors)
            scores = _solve_for_db(picked[0])
        else:  # interpolate (default)
            bracket = resolve_bracket(year, anchors)
            if bracket.upper_db is None:
                # EXACT anchor or CLAMP at a span end → single solve.
                scores = _solve_for_db(bracket.lower_db)
            else:
                scores_a = _solve_for_db(bracket.lower_db)
                scores_b = _solve_for_db(bracket.upper_db)
                scores = blend_method_scores(
                    scores_a, scores_b, bracket.frac, method_tuples
                )
        out.append((year, scores))
    return out, warnings


@router.post("/lca/calculate-archetype-trajectory", response_model=ArchetypeTrajectoryResult)
async def calculate_archetype_trajectory(body: ArchetypeTrajectoryRequest) -> ArchetypeTrajectoryResult:
    """Continuous-horizon single-product LCA: ONE archetype computed year-by-year
    across a prospective trajectory's premise anchors (the single-product
    analogue of system-level projected impact). TOTALS ONLY per year — for the
    per-activity stage/material breakdown of any single year, call
    ``/lca/calculate-archetype`` with that year's ``compute_database``. Reuses
    the 6A/6B primitives (``resolve_bracket`` / ``blend_method_scores``) and one
    ``MultiDBPersistentRunner`` per trajectory so each anchor DB factorizes once."""
    from mapper.core import plca_storage

    t0 = time.perf_counter()

    bundle = _build_archetype_source_demand(
        archetype_id=body.archetype_id,
        scope=body.scope,
        amount=body.amount,
        stage_amounts=body.stage_amounts,
        methods=body.methods,
        parameter_scenario=body.parameter_scenario,
    )

    project = bw2data.projects.current
    anchors = plca_storage.resolve_prospective_dbs(
        project, body.base_db, body.iam, body.ssp,
    )
    anchor_years = sorted(y for _, y in anchors)

    warnings: list[str] = []
    # Degenerate guard: ≤1 anchor → no span to interpolate across. Don't crash;
    # return what we can (a single point for 1 anchor, nothing for 0) plus a
    # clear no-curve warning the frontend can surface.
    if len(anchors) == 0:
        warnings.append(
            f"No prospective databases found for {body.base_db} · {body.iam} · "
            f"{body.ssp}. Generate premise databases for this trajectory first."
        )
    elif len(anchors) == 1:
        warnings.append(
            "Only one prospective anchor available for this trajectory — showing "
            "a single point (no curve). Generate more years to interpolate."
        )

    try:
        runner = MultiDBPersistentRunner()
        per_year, solve_warnings = _trajectory_year_scores(
            total_demand=bundle.total_demand,
            method_tuples=bundle.method_tuples,
            anchors=anchors,
            temporal_mode=body.temporal_mode,
            runner=runner,
            year_start=body.year_start,
            year_end=body.year_end,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Trajectory LCA failed: {e}")
    warnings.extend(solve_warnings)

    # Assemble the TOTALS-only per-year envelope.
    years: list[ArchetypeTrajectoryYear] = []
    for year, scores in per_year:
        method_scores: list[ArchetypeTrajectoryMethodScore] = []
        for method_list, mt in zip(body.methods, bundle.method_tuples):
            score, unit = scores.get(mt, (0.0, ""))
            method_scores.append(ArchetypeTrajectoryMethodScore(
                method=method_list,
                method_label=method_list[-1] if method_list else "",
                score=score,
                unit=unit,
            ))
        years.append(ArchetypeTrajectoryYear(year=year, method_scores=method_scores))

    elapsed = round(time.perf_counter() - t0, 2)
    return ArchetypeTrajectoryResult(
        archetype_id=body.archetype_id,
        archetype_name=bundle.arc.name,
        scope=body.scope,
        base_db=body.base_db,
        iam=body.iam,
        ssp=body.ssp,
        temporal_mode=body.temporal_mode,
        parameter_scenario=body.parameter_scenario,
        anchor_years=anchor_years,
        years=years,
        elapsed_seconds=elapsed,
        warnings=warnings,
    )


# ── Export archetype LCA results to XLSX ──────────────────────────────────────


def _build_lca_export_workbook(data: list[ArchetypeLCACalculateResult]):  # noqa: C901
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt = '#,##0.0000'
    pct_fmt = '0.0"%"'

    def _style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    def _auto_width(ws) -> None:
        for col in ws.columns:
            mx = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                mx = max(mx, len(val))
            ws.column_dimensions[letter].width = min(mx + 3, 40)

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for d in data:
        rows.append(("", ""))
        rows.append(("Archetype", d.archetype_name))
        rows.append(("Scope", d.scope))
        rows.append(("Amount", d.amount))
        rows.append(("Stages", ", ".join(d.stages_included)))
        rows.append(("Elapsed (s)", d.elapsed_seconds))
        rows.append(("Indicators", len(d.results)))
        for r in d.results:
            rows.append(("  Method", " › ".join(r.method)))
    for row in rows:
        ws.append(row)
    _auto_width(ws)

    # ── Sheet 2: Results ──
    ws2 = wb.create_sheet("Results")
    ws2.append(["Archetype", "Method", "Category", "Indicator", "Score", "Unit"])
    _style_header(ws2, 6)
    for d in data:
        for r in d.results:
            ws2.append([
                d.archetype_name,
                r.method[0] if r.method else "",
                r.method[1] if len(r.method) > 1 else "",
                r.method[-1] if r.method else "",
                r.score,
                r.unit,
            ])
            ws2.cell(row=ws2.max_row, column=5).number_format = num_fmt
    _auto_width(ws2)

    # ── Sheet 3: Contributions by material ──
    ws3 = wb.create_sheet("Contributions by material")
    # Build headers: Material | Stage | Component | then per-indicator: Qty | Unit | Impact | %
    indicators = data[0].results if data else []
    headers = ["Archetype", "Material", "Stage", "Component"]
    for r in indicators:
        label = r.method[-1] if r.method else "?"
        headers += [f"{label} Qty", f"{label} Unit", f"{label} Impact ({r.unit})", f"{label} %"]
    ws3.append(headers)
    _style_header(ws3, len(headers))

    for d in data:
        # Collect all material names across indicators
        all_names: list[str] = []
        seen: set[str] = set()
        for r in d.results:
            for c in r.contributions:
                if c.name not in seen:
                    seen.add(c.name)
                    all_names.append(c.name)
        # Build lookup: indicator_idx → {material_name → contribution}
        lookup: dict[int, dict[str, MaterialContribution]] = {}
        for idx, r in enumerate(d.results):
            lookup[idx] = {c.name: c for c in r.contributions}
        for mat_name in all_names:
            row_data: list = [d.archetype_name, mat_name, "", ""]
            first_contrib = None
            for idx in range(len(d.results)):
                c = lookup.get(idx, {}).get(mat_name)
                if c:
                    if not first_contrib:
                        first_contrib = c
                    row_data += [c.quantity, c.unit, c.impact, c.percentage]
                else:
                    row_data += ["", "", 0, 0]
            if first_contrib:
                row_data[2] = first_contrib.stage
                row_data[3] = first_contrib.component
            ws3.append(row_data)
    _auto_width(ws3)

    # ── Sheet 4: Contributions by stage ──
    ws4 = wb.create_sheet("Contributions by stage")
    headers4 = ["Archetype", "Stage"]
    for r in (data[0].results if data else []):
        label = r.method[-1] if r.method else "?"
        headers4.append(f"{label} ({r.unit})")
    ws4.append(headers4)
    _style_header(ws4, len(headers4))

    for d in data:
        stage_sums: dict[str, list[float]] = {}
        for idx, r in enumerate(d.results):
            for c in r.contributions:
                stage = c.stage or "(unknown)"
                if stage not in stage_sums:
                    stage_sums[stage] = [0.0] * len(d.results)
                stage_sums[stage][idx] += c.impact
        for stage_name, values in stage_sums.items():
            ws4.append([d.archetype_name, stage_name] + values)
    _auto_width(ws4)

    return wb


# ── Contribution Analysis (Single-Product LCA) ─────────────────────────────

# Cache key = (target_descriptor, method_tuple, scope, year, compute_database).
# Cutoff/depth are presentation-layer filters applied to the cached deepest
# tree. compute_database is part of the key because translating activity keys
# to a prospective DB produces a genuinely different LCI — the score, top
# flows, and tree all change.
_contribution_cache: dict[tuple, dict] = {}


def _translate_demand_to_database(
    demand: dict[tuple[str, str], float],
    compute_database: str | None,
) -> tuple[dict[tuple[str, str], float], list[str]]:
    """Re-key a demand dict to run against ``compute_database`` (a premise-
    generated prospective DB, typically). For each key, look up the same
    ``code`` inside ``compute_database``. If found, use the translated key.
    If not found (the activity wasn't carried over by premise, was renamed,
    or was dropped), fall back to the original key and record a warning so
    the frontend can flag the result as partially translated.

    When ``compute_database`` is None or matches every demand key's source
    DB, returns the demand unchanged with no warnings.
    """
    if not compute_database:
        return demand, []

    # Skip translation when every key already lives in compute_database.
    if all(src_db == compute_database for (src_db, _code) in demand):
        return demand, []

    if compute_database not in bw2data.databases:
        return demand, [
            f"compute_database '{compute_database}' not found in this project; "
            "computing against source databases instead."
        ]

    translated: dict[tuple[str, str], float] = {}
    warnings: list[str] = []
    for (src_db, code), amount in demand.items():
        if src_db == compute_database:
            translated[(src_db, code)] = translated.get((src_db, code), 0.0) + amount
            continue
        candidate = (compute_database, code)
        try:
            bw2data.get_activity(candidate)
        except Exception:
            translated[(src_db, code)] = translated.get((src_db, code), 0.0) + amount
            warnings.append(
                f"Activity {src_db}/{code} not found in {compute_database}; "
                "fell back to source database for this key."
            )
            continue
        translated[candidate] = translated.get(candidate, 0.0) + amount
    return translated, warnings


def _build_archetype_demand(
    archetype_id: str,
    scope: str,
    stage_amounts: dict[str, float] | None,
) -> tuple[
    dict[tuple[str, str], float],
    str,
    list[str],
    dict[str, dict[tuple[str, str], float]],
]:
    """Flatten an archetype BOM (scope-filtered) into a (db, code) → amount map.

    Returns ``(combined_demand, archetype_name, stages_included, per_stage_demand)``.

    ``per_stage_demand`` is keyed by stage name (Manufacturing, Use Phase, …)
    and used by the contribution-analysis path to compute a per-stage
    breakdown via cheap ``redo_lci`` calls on the same factorized matrix.
    The aggregated ``combined_demand`` equals the union of these stage
    demands and is what the main LCA solves against.
    """
    from mapper.api.bom import _get_archetype
    from mapper.core.bom_engine import flatten_bom

    arc = _get_archetype(archetype_id)
    if scope not in ("inflows", "stock", "outflows", "all"):
        raise HTTPException(status_code=400, detail=f"Invalid scope: {scope}")

    scope_roots = filter_roots_by_scope(arc.bom, scope)
    stages = [r.name for r in scope_roots]
    eff: dict[str, float] = {}
    if stage_amounts:
        for r in scope_roots:
            eff[r.name] = stage_amounts.get(r.name, 1.0)
    else:
        for r in scope_roots:
            eff[r.name] = 1.0

    demand: dict[tuple[str, str], float] = {}
    per_stage: dict[str, dict[tuple[str, str], float]] = {}
    for root in scope_roots:
        flat = flatten_bom(root)
        mult = eff.get(root.name, 1.0)
        stage_demand: dict[tuple[str, str], float] = {}
        for m in flat:
            if m.ecoinvent_activity is None:
                continue
            key = (m.ecoinvent_activity.database, m.ecoinvent_activity.code)
            qty = m.quantity * mult
            demand[key] = demand.get(key, 0.0) + qty
            stage_demand[key] = stage_demand.get(key, 0.0) + qty
        per_stage[root.name] = stage_demand

    if not demand:
        raise HTTPException(
            status_code=400,
            detail="No linked materials in this scope. Link ecoinvent activities first.",
        )
    return demand, arc.name, stages, per_stage


def _tree_to_schema(tree: dict) -> ContributionTreeNode:
    return ContributionTreeNode(
        name=tree.get("name", ""),
        key=tree.get("key", ""),
        location=tree.get("location", ""),
        amount=tree.get("amount", 0.0),
        unit=tree.get("unit", ""),
        score=tree.get("score", 0.0),
        unit_score=tree.get("unit_score", ""),
        percentage=tree.get("percentage", 0.0),
        children=[_tree_to_schema(c) for c in tree.get("children", [])],
    )


def _compute_contribution_analysis(
    body: ContributionAnalysisRequest,
    *,
    runner: PersistentLCARunner | None = None,
) -> ContributionAnalysisResult:
    """Synchronous core of the contribution-analysis endpoint. Reused by the
    multi-year endpoint, which loops over per-year requests in a worker
    thread. Side effects: populates ``_contribution_cache``.

    ``runner`` (optional): a shared :class:`PersistentLCARunner` so the
    technosphere LU factorization is reused — across the main LCA + the
    recursive contribution tree (within one year), and across years (in
    static mode, or as fallback in projected mode where the DB changes
    per year, the runner still saves the duplicate tree-builder factorization).
    When ``None``, a fresh runner is built locally.
    """
    t0 = time.perf_counter()

    if body.target_type not in ("activity", "archetype"):
        raise HTTPException(status_code=400, detail="target_type must be 'activity' or 'archetype'")
    if not body.method:
        raise HTTPException(status_code=400, detail="method is required")

    method_tuple = tuple(body.method)

    # Build demand + label
    if body.target_type == "activity":
        if not body.database or not body.code:
            raise HTTPException(status_code=400, detail="database and code are required for activity target")
        if "biosphere" in body.database.lower():
            raise HTTPException(
                status_code=400,
                detail="Biosphere flows cannot be used as functional units.",
            )
        try:
            act = bw2data.get_activity((body.database, body.code))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Activity not found: {e}")
        demand: dict[tuple[str, str], float] = {(body.database, body.code): float(body.amount)}
        target_label = act.get("reference product", act.get("name", ""))
        target_descriptor: tuple = ("activity", body.database, body.code, float(body.amount))
        scope = "all"
        per_stage_demand: dict[str, dict[tuple[str, str], float]] = {}
    else:
        if not body.archetype_id:
            raise HTTPException(status_code=400, detail="archetype_id is required for archetype target")
        demand, target_label, _stages, per_stage_demand = _build_archetype_demand(
            body.archetype_id, body.scope, body.stage_amounts
        )
        sa_tag = tuple(sorted((body.stage_amounts or {}).items()))
        target_descriptor = ("archetype", body.archetype_id, body.scope, sa_tag)
        scope = body.scope

    # Apply prospective-database translation before computing or hitting the
    # cache. Warnings (e.g. fallbacks for missing keys) are surfaced on the
    # response so the frontend can flag partial translations.
    demand, translation_warnings = _translate_demand_to_database(
        demand, body.compute_database
    )

    cache_key = (target_descriptor, method_tuple, scope, body.year, body.compute_database)
    cached = _contribution_cache.get(cache_key)

    # Recompute when cache miss OR when caller asks for a deeper/more-detailed
    # view than what's cached (cutoff lower or depth higher).
    needs_recompute = (
        cached is None
        or body.cutoff < cached.get("cutoff", float("inf"))
        or body.max_depth > cached.get("max_depth", -1)
        or body.limit > cached.get("limit", -1)
        or body.max_nodes > cached.get("max_nodes", -1)
        # Pre-by_stage cache entry on an archetype target — recompute so we
        # populate the per-stage breakdown.
        or ("by_stage" not in (cached or {}) and bool(per_stage_demand))
        # Pre-truncated-sankey cache entry — recompute so the SankeyData
        # carries ``total_nodes_discovered`` and ``truncated``.
        or ("total_nodes_discovered" not in (cached or {}).get("sankey", {}))
    )

    if needs_recompute:
        try:
            # Per-phase timings. Logged in one summary line at the end so the
            # multi-year trace produces one row per year.
            phase: dict[str, float] = {}

            # One persistent runner per call (or shared by the caller in
            # multi-year mode). Solves the technosphere matrix once and
            # reuses the factorization for the recursive tree.
            local_runner = runner if runner is not None else PersistentLCARunner()
            t_p = time.perf_counter()
            facts_before = local_runner.factorizations
            scores = local_runner(demand, [method_tuple])
            phase["solve"] = time.perf_counter() - t_p
            phase["factorized"] = float(local_runner.factorizations - facts_before)
            score = float(scores[method_tuple][0])
            method_unit = scores[method_tuple][1] or bw2data.methods.get(
                method_tuple, {}
            ).get("unit", "")
            lca = local_runner._lca  # underlying bw2calc.LCA, state = aggregate demand

            t_p = time.perf_counter()
            try:
                techno = get_contributions(lca, score, limit=body.limit)
            except Exception:
                logger.exception("contribution-analysis: get_contributions failed")
                techno = {"items": [], "rest_amount": score, "rest_percentage": 100.0}
            phase["techno"] = time.perf_counter() - t_p

            t_p = time.perf_counter()
            try:
                bio = get_biosphere_contributions(lca, score, limit=body.limit)
            except Exception:
                logger.exception(
                    "contribution-analysis: get_biosphere_contributions failed "
                    "(target=%s db=%s code=%s method=%s)",
                    body.target_type, body.database, body.code, method_tuple,
                )
                bio = {"items": [], "rest_amount": 0.0, "rest_percentage": 0.0}
            phase["bio"] = time.perf_counter() - t_p

            # Shared unit-score cache for sankey BFS + recursive tree builder.
            # Linear LCA: score(act, x) = x × unit_score(act). Both functions
            # need unit scores; sharing one dict between them collapses thousands
            # of redundant sub-LCAs into one call per unique activity. Scope is
            # per call, which means per (target, method, database) — safe to
            # discard at year boundary in multi-year runs.
            unit_score_cache: dict[tuple[str, str], float] = {}

            t_p = time.perf_counter()
            try:
                sankey = get_supply_chain(
                    lca,
                    method=method_tuple,
                    runner=local_runner,
                    depth=min(body.max_depth, 4),
                    max_nodes=body.max_nodes,
                    unit_score_cache=unit_score_cache,
                )
            except Exception:
                logger.exception("contribution-analysis: get_supply_chain failed")
                sankey = {"nodes": [], "links": [], "total_nodes_discovered": 0, "truncated": False}
            phase["sankey"] = time.perf_counter() - t_p

            # Per-stage breakdown for archetype targets. One ``redo_lci`` per
            # stage on the same factorized matrix — back-substitution only
            # (~15 ms each) when the runner is hot. Activity targets have no
            # inherent stages so by_stage stays empty.
            t_p = time.perf_counter()
            by_stage: list[dict] = []
            if per_stage_demand:
                try:
                    for stage_name, raw_stage_demand in per_stage_demand.items():
                        translated, _ = _translate_demand_to_database(
                            raw_stage_demand, body.compute_database
                        )
                        if not translated:
                            by_stage.append({
                                "stage": stage_name, "score": 0.0,
                                "unit": method_unit, "percentage": 0.0,
                            })
                            continue
                        stage_scores = local_runner(translated, [method_tuple])
                        sc = float(stage_scores[method_tuple][0])
                        pct = (abs(sc) / abs(score) * 100.0) if score else 0.0
                        by_stage.append({
                            "stage": stage_name, "score": sc,
                            "unit": method_unit, "percentage": pct,
                        })
                    # Numerical sanity check: per-stage scores must sum to the
                    # aggregate. LCA is linear in the demand vector, so any
                    # discrepancy beyond floating-point noise indicates a bug.
                    stage_sum = sum(s["score"] for s in by_stage)
                    if score:
                        rel_err = abs(stage_sum - score) / abs(score)
                        if rel_err > 1e-9:
                            logger.warning(
                                "contribution-analysis: per-stage scores sum=%g "
                                "differs from aggregate=%g by rel=%g",
                                stage_sum, score, rel_err,
                            )
                except Exception:
                    logger.exception("contribution-analysis: by_stage failed")
                    by_stage = []
            phase["by_stage"] = time.perf_counter() - t_p

            t_p = time.perf_counter()
            try:
                # Pass the same runner so the tree's many sub-LCAs reuse the
                # factorization instead of rebuilding it from scratch. Pass
                # the same unit-score cache the sankey BFS just populated —
                # many of the dominant upstream activities will already be
                # characterised, so the tree builder hits the cache instead
                # of triggering fresh sub-LCAs.
                tree = get_recursive_contribution_tree(
                    demand, method_tuple,
                    cutoff=body.cutoff,
                    max_depth=body.max_depth,
                    runner=local_runner,
                    unit_score_cache=unit_score_cache,
                )
            except Exception as e:
                tree = {
                    "name": f"(tree unavailable: {e})",
                    "key": "", "amount": 0.0, "unit": "", "score": score,
                    "unit_score": method_unit, "percentage": 100.0, "children": [],
                }
            phase["tree"] = time.perf_counter() - t_p

            logger.info(
                "[CA-phases] db=%s year=%s "
                "solve=%.2fs (refac=%.0f) techno=%.2fs bio=%.2fs "
                "sankey=%.2fs by_stage=%.2fs tree=%.2fs "
                "unit_score_cache=%d biosphere_items=%d score=%g",
                body.compute_database, body.year,
                phase["solve"], phase["factorized"], phase["techno"], phase["bio"],
                phase["sankey"], phase["by_stage"], phase["tree"],
                len(unit_score_cache),
                len(bio.get("items", [])), score,
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Contribution analysis failed: {e}")

        cached = {
            "score": score,
            "method_unit": method_unit,
            "techno": techno,
            "bio": bio,
            "sankey": sankey,
            "tree": tree,
            "by_stage": by_stage,
            "max_nodes": body.max_nodes,
            "cutoff": body.cutoff,
            "max_depth": body.max_depth,
            "limit": body.limit,
        }
        _contribution_cache[cache_key] = cached

    elapsed = round(time.perf_counter() - t0, 3)

    from mapper import __version__ as _mapper_version

    return ContributionAnalysisResult(
        target_type=body.target_type,
        target_label=target_label,
        method=list(method_tuple),
        method_unit=cached["method_unit"],
        score=cached["score"],
        scope=scope,
        year=body.year,
        compute_database=body.compute_database,
        top_technosphere=ContributionsResponse(**cached["techno"]),
        top_biosphere=[BiosphereContributionItem(**i) for i in cached["bio"]["items"]],
        biosphere_rest_amount=cached["bio"]["rest_amount"],
        biosphere_rest_percentage=cached["bio"]["rest_percentage"],
        supply_chain_sankey=SankeyData(**cached["sankey"]),
        supply_chain_tree=_tree_to_schema(cached["tree"]),
        by_stage=[StageContribution(**s) for s in cached.get("by_stage", [])],
        cutoff=body.cutoff,
        max_depth=body.max_depth,
        elapsed_seconds=elapsed,
        warnings=list(translation_warnings),
        computed_at=datetime.datetime.now(datetime.timezone.utc).isoformat(),
        mapper_version=_mapper_version,
    )


@router.post("/lca/contribution-analysis", response_model=ContributionAnalysisResult)
async def calculate_contribution_analysis(
    body: ContributionAnalysisRequest,
) -> ContributionAnalysisResult:
    return _compute_contribution_analysis(body)


@router.get("/lca/prospective-years")
async def list_prospective_years(database: str) -> dict:
    """List the years for which a prospective database exists in the current
    bw2 project, given an IAM × pathway pattern.

    ``database`` may be either:
      - the bare pattern, e.g. ``ecoinvent-3.10-cutoff_premise_remind_ssp2-pkbudg1150``
      - a fully-qualified single-year name with a trailing ``_<year>``, in
        which case the year suffix is stripped and the rest is treated as
        the pattern.

    Years are detected by matching siblings ``{pattern}_<4-digit-year>``
    against the project's database list. Static (non-prospective) inputs
    return an empty list and the caller should treat all user-chosen years
    as available — the static base DB doesn't gate years.
    """
    import re

    if not database:
        raise HTTPException(status_code=400, detail="database is required")

    # Strip trailing _<year> if the caller passed a fully-qualified name.
    pattern = re.sub(r"_(\d{4})$", "", database)

    # Static / non-premise database → no year gating.
    if "_premise_" not in pattern:
        return {"pattern": pattern, "available_years": [], "is_prospective": False}

    needle = re.compile(rf"^{re.escape(pattern)}_(\d{{4}})$")
    years: list[int] = []
    for db_name in bw2data.databases:
        m = needle.match(db_name)
        if m:
            years.append(int(m.group(1)))
    years.sort()
    return {"pattern": pattern, "available_years": years, "is_prospective": True}


# ── Multi-year contribution analysis ──────────────────────────────────────


class _MYTaskState:
    """In-memory state for one multi-year contribution-analysis task. Same
    pattern as ``mapper.api.plca._TaskState`` (worker thread + WS subscribers
    queue) — extracted here because it carries a different result shape."""

    def __init__(self, planned_years: list[int]) -> None:
        self.planned_years = planned_years
        self.stage: str = "queued"
        self.pct: float = 0.0
        self.done: bool = False
        self.error: str | None = None
        self.result: MultiYearContributionResult | None = None
        self.subscribers: list[asyncio.Queue] = []
        self.cancelled: bool = False


_MY_TASKS: dict[str, _MYTaskState] = {}
_MY_TASK_LOCK = threading.Lock()


def _my_notify_all(task: _MYTaskState, payload: dict[str, Any]) -> None:
    for q in list(task.subscribers):
        try:
            q.put_nowait(payload)
        except asyncio.QueueFull:
            pass


def _build_multi_year_result(
    body: MultiYearContributionRequest,
    per_year_results: dict[int, ContributionAnalysisResult],
    elapsed: float,
) -> MultiYearContributionResult:
    """Aggregate per-year ContributionAnalysisResult dicts into a single
    multi-year result with trajectory + evolution views ready for charting."""
    from mapper import __version__ as _mapper_version

    sorted_years = sorted(per_year_results)

    trajectory = [
        MultiYearTrajectoryPoint(
            year=y,
            score=per_year_results[y].score,
            compute_database=per_year_results[y].compute_database,
            has_warnings=bool(per_year_results[y].warnings),
        )
        for y in sorted_years
    ]

    # Evolution: union of top-N keys across all years; per-year score for each.
    evolution_index: dict[str, MultiYearEvolutionItem] = {}
    for y in sorted_years:
        r = per_year_results[y]
        for it in r.top_technosphere.items:
            ev = evolution_index.get(it.activity_key)
            if ev is None:
                ev = MultiYearEvolutionItem(
                    activity_key=it.activity_key,
                    activity_name=it.activity_name,
                    location=it.location,
                    unit=it.unit,
                )
                evolution_index[it.activity_key] = ev
            ev.by_year[str(y)] = it.amount
    # Fill missing years with 0 so chart lines stay continuous.
    for ev in evolution_index.values():
        for y in sorted_years:
            ev.by_year.setdefault(str(y), 0.0)
    # Stable order: descending by mean contribution across years.
    evolution = sorted(
        evolution_index.values(),
        key=lambda e: -sum(e.by_year.values()) / max(len(e.by_year), 1),
    )

    aggregated_warnings: list[str] = []
    for y in sorted_years:
        for w in per_year_results[y].warnings:
            aggregated_warnings.append(f"[{y}] {w}")

    first = per_year_results[sorted_years[0]]

    return MultiYearContributionResult(
        target_type=first.target_type,
        target_label=first.target_label,
        method=list(first.method),
        method_unit=first.method_unit,
        compute_database_pattern=body.compute_database_pattern,
        years=sorted_years,
        results={str(y): per_year_results[y] for y in sorted_years},
        trajectory=trajectory,
        evolution=evolution,
        cutoff=body.cutoff,
        max_depth=body.max_depth,
        elapsed_seconds=elapsed,
        warnings=aggregated_warnings,
        computed_at=datetime.datetime.now(datetime.timezone.utc).isoformat(),
        mapper_version=_mapper_version,
    )


@router.post(
    "/lca/contribution-analysis/multi-year",
    response_model=MultiYearContributionTaskStarted,
)
async def start_multi_year_contribution(
    body: MultiYearContributionRequest,
) -> MultiYearContributionTaskStarted:
    """Start a multi-year contribution-analysis run. Returns a task_id; the
    full result streams via WebSocket /api/ws/lca/multi-year/{task_id} and
    is also retrievable from GET /api/lca/contribution-analysis/multi-year/
    {task_id} once done."""
    if not body.years:
        raise HTTPException(status_code=400, detail="years must be a non-empty list")
    if body.target_type not in ("activity", "archetype"):
        raise HTTPException(
            status_code=400, detail="target_type must be 'activity' or 'archetype'"
        )
    if not body.method:
        raise HTTPException(status_code=400, detail="method is required")

    sorted_years = sorted(set(int(y) for y in body.years))
    pattern = body.compute_database_pattern
    compute_databases = (
        [f"{pattern}_{y}" for y in sorted_years] if pattern else [""] * len(sorted_years)
    )

    task_id = uuid.uuid4().hex
    task = _MYTaskState(planned_years=sorted_years)
    with _MY_TASK_LOCK:
        _MY_TASKS[task_id] = task
    task_registry.register(task_id)

    loop = asyncio.get_running_loop()

    def _safe_notify(payload: dict[str, Any]) -> None:
        # The runner thread can outlive the request loop in unit tests where
        # ``asyncio.run()`` closes the loop right after the await completes.
        # Drop the notification rather than crash the thread.
        try:
            loop.call_soon_threadsafe(_my_notify_all, task, payload)
        except RuntimeError:
            pass

    def _run() -> None:
        t0 = time.perf_counter()
        per_year: dict[int, ContributionAnalysisResult] = {}
        # Shared runner across all years.
        #   - Static (no pattern): same matrix every year → one factorization total.
        #     Each subsequent year is just back-substitution (~15 ms).
        #   - Projected (pattern set): different DB per year → runner falls back
        #     to a rebuild on each year, but still saves the duplicate
        #     factorization that the recursive tree builder used to do.
        shared_runner = PersistentLCARunner()
        logger.info(
            "[multi-year] start years=%s pattern=%s target=%s archetype=%s "
            "method=%s",
            sorted_years, pattern, body.target_type, body.archetype_id,
            body.method,
        )
        try:
            per_year_elapsed: dict[int, float] = {}
            for i, year in enumerate(sorted_years):
                if task_registry.is_cancelled(task_id):
                    raise CancelledOperation(task_id)
                task.stage = f"year {year} ({i + 1}/{len(sorted_years)})"
                task.pct = i / len(sorted_years) * 100.0
                _safe_notify(
                    {"type": "progress", "stage": task.stage, "pct": task.pct, "year": year}
                )
                per_year_body = ContributionAnalysisRequest(
                    target_type=body.target_type,
                    database=body.database,
                    code=body.code,
                    amount=body.amount,
                    archetype_id=body.archetype_id,
                    scope=body.scope,
                    stage_amounts=body.stage_amounts,
                    year=year,
                    compute_database=(f"{pattern}_{year}" if pattern else None),
                    method=body.method,
                    limit=body.limit,
                    cutoff=body.cutoff,
                    max_depth=body.max_depth,
                    max_nodes=body.max_nodes,
                )
                t_year = time.perf_counter()
                per_year[year] = _compute_contribution_analysis(
                    per_year_body, runner=shared_runner
                )
                yr_elapsed = time.perf_counter() - t_year
                per_year_elapsed[year] = yr_elapsed
                logger.info(
                    "[multi-year] year=%d elapsed=%.2fs running_total=%.2fs",
                    year, yr_elapsed, time.perf_counter() - t0,
                )
            elapsed = round(time.perf_counter() - t0, 3)
            logger.info(
                "[multi-year] done total=%.2fs per_year=%s "
                "factorizations=%d redo_calls=%d method_switches=%d",
                elapsed,
                {y: round(v, 2) for y, v in per_year_elapsed.items()},
                shared_runner.factorizations,
                shared_runner.redo_calls,
                shared_runner.method_switches,
            )
            task.result = _build_multi_year_result(body, per_year, elapsed)
            task.stage = "done"
            task.pct = 100.0
            task.done = True
            _safe_notify({"type": "done", "task_id": task_id})
        except CancelledOperation:
            # Cancellation is an expected outcome; result intentionally not
            # built/persisted (partial per-year data would be misleading on
            # the trajectory chart). The "cancelled" frame mirrors the
            # "done"/"error" frames so the WS protocol stays consistent.
            logger.info(
                "[multi-year] cancelled task=%s after %d/%d years",
                task_id, len(per_year), len(sorted_years),
            )
            task.cancelled = True
            task.done = True
            task.stage = "cancelled"
            _safe_notify({"type": "cancelled", "task_id": task_id})
        except HTTPException as exc:
            task.error = exc.detail if isinstance(exc.detail, str) else str(exc.detail)
            task.done = True
            _safe_notify({"type": "error", "error": task.error})
        except Exception as exc:
            logger.exception("multi-year contribution analysis failed")
            task.error = str(exc)
            task.done = True
            _safe_notify({"type": "error", "error": str(exc)})
        finally:
            task_registry.unregister(task_id)

    threading.Thread(target=_run, daemon=True).start()
    return MultiYearContributionTaskStarted(
        task_id=task_id,
        planned_years=sorted_years,
        compute_databases=compute_databases,
    )


@router.get(
    "/lca/contribution-analysis/multi-year/{task_id}",
    response_model=MultiYearContributionResult | CancelledTaskResponse,
)
async def get_multi_year_contribution(
    task_id: str,
) -> MultiYearContributionResult | CancelledTaskResponse:
    with _MY_TASK_LOCK:
        task = _MY_TASKS.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Unknown task id")
    if not task.done:
        raise HTTPException(status_code=409, detail=f"Task not finished: {task.stage}")
    if task.cancelled:
        return CancelledTaskResponse(task_id=task_id)
    if task.error:
        raise HTTPException(status_code=500, detail=task.error)
    if task.result is None:
        raise HTTPException(status_code=500, detail="Task finished without a result")
    return task.result


@router.websocket("/ws/lca/multi-year/{task_id}")
async def ws_multi_year_progress(websocket: WebSocket, task_id: str) -> None:
    await websocket.accept()
    with _MY_TASK_LOCK:
        task = _MY_TASKS.get(task_id)
    if task is None:
        await websocket.send_json({"type": "error", "error": "Unknown task id"})
        await websocket.close()
        return

    queue: asyncio.Queue = asyncio.Queue(maxsize=256)
    task.subscribers.append(queue)

    await websocket.send_json({"type": "progress", "stage": task.stage, "pct": task.pct})
    if task.done:
        if task.cancelled:
            await websocket.send_json({"type": "cancelled", "task_id": task_id})
        elif task.error:
            await websocket.send_json({"type": "error", "error": task.error})
        else:
            await websocket.send_json({"type": "done", "task_id": task_id})
        await websocket.close()
        return

    try:
        while True:
            payload = await queue.get()
            await websocket.send_json(payload)
            if payload.get("type") in ("done", "error", "cancelled"):
                break
    except WebSocketDisconnect:
        pass
    finally:
        try:
            task.subscribers.remove(queue)
        except ValueError:
            pass
        # Disconnect-cancel: if no live subscribers remain on a still-
        # running task (and we're past the open-WS grace window), cancel it.
        # Covers tab close / refresh / network drop. Explicit POST cancel
        # follows a different path; this is just the implicit signal.
        task_registry.maybe_cancel_on_last_subscriber_leave(
            task_id,
            remaining_subscribers=len(task.subscribers),
            task_done=task.done,
        )
        try:
            await websocket.close()
        except Exception:
            pass


def _build_contribution_workbook(result: ContributionAnalysisResult):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    def _auto_width(ws) -> None:
        for col in ws.columns:
            mx = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                mx = max(mx, len(val))
            ws.column_dimensions[letter].width = min(mx + 3, 50)

    method_label = " › ".join(result.method)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    _style_header(ws, 2)
    ws.append(["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Target type", result.target_type])
    ws.append(["Target", result.target_label])
    ws.append(["Scope", result.scope])
    ws.append(["Year", result.year if result.year is not None else "—"])
    ws.append(["Method", method_label])
    ws.append(["Score", result.score])
    ws.append(["Unit", result.method_unit])
    ws.append(["Cutoff", result.cutoff])
    ws.append(["Max depth", result.max_depth])
    ws.append(["Elapsed (s)", result.elapsed_seconds])
    _auto_width(ws)

    # Sheet 2: Top Technosphere
    ws2 = wb.create_sheet("Top Technosphere")
    ws2.append(["Activity", "Location", "Key", "Impact", "Unit", "%"])
    _style_header(ws2, 6)
    for it in result.top_technosphere.items:
        ws2.append([it.activity_name, it.location, it.activity_key, it.amount, it.unit, it.percentage])
    ws2.append(["Rest", "", "", result.top_technosphere.rest_amount, result.method_unit,
                result.top_technosphere.rest_percentage])
    _auto_width(ws2)

    # Sheet 3: Top Biosphere
    ws3 = wb.create_sheet("Top Biosphere")
    ws3.append(["Flow", "Compartment", "Subcompartment", "Inventory amount", "Inv unit",
                "Impact", "Unit", "%"])
    _style_header(ws3, 8)
    for b in result.top_biosphere:
        ws3.append([
            b.flow_name, b.compartment, b.subcompartment,
            b.inventory_amount, b.inventory_unit,
            b.amount, b.unit, b.percentage,
        ])
    ws3.append(["Rest", "", "", "", "", result.biosphere_rest_amount, result.method_unit,
                result.biosphere_rest_percentage])
    _auto_width(ws3)

    # Sheet 4: Supply Chain Tree (flat with depth column)
    ws4 = wb.create_sheet("Supply Chain Tree")
    ws4.append(["Depth", "Name", "Location", "Key", "Amount", "Unit", "Score", "Score unit", "%"])
    _style_header(ws4, 9)

    def walk(node: ContributionTreeNode, depth: int = 0) -> None:
        ws4.append([
            depth,
            ("  " * depth) + node.name,
            node.location,
            node.key,
            node.amount,
            node.unit,
            node.score,
            node.unit_score,
            node.percentage,
        ])
        for c in node.children:
            walk(c, depth + 1)

    walk(result.supply_chain_tree, 0)
    _auto_width(ws4)

    # Sheet 5: By Stage (only meaningful for archetype targets — best-effort)
    ws5 = wb.create_sheet("By Stage")
    ws5.append(["Stage", "Note"])
    _style_header(ws5, 2)
    if result.target_type == "archetype":
        ws5.append([result.scope, "Per-stage breakdown available in /lca/calculate-archetype export."])
    else:
        ws5.append(["—", "Not applicable for activity targets."])
    _auto_width(ws5)

    # Sheet 6: Methodology
    ws6 = wb.create_sheet("Methodology")
    ws6.append(["Topic", "Details"])
    _style_header(ws6, 2)
    method_notes = [
        ("Top technosphere", "bw2analyzer.ContributionAnalysis.annotated_top_processes"),
        ("Top biosphere", "bw2analyzer.ContributionAnalysis.annotated_top_emissions"),
        ("Supply chain Sankey", "BFS over technosphere exchanges (raw exchange amounts)"),
        ("Supply chain tree", "Recursive sub-LCA per node — characterised impact at each branch"),
        ("Cutoff", f"{result.cutoff} (fraction of root score; branches below are pruned)"),
        ("Max depth", f"{result.max_depth} (hard recursion cap)"),
        ("LCIA method", method_label),
        ("Score unit", result.method_unit),
    ]
    for row in method_notes:
        ws6.append(row)
    _auto_width(ws6)

    return wb


@router.post("/lca/contribution-analysis/export")
async def export_contribution_analysis(body: ContributionAnalysisExportRequest) -> Response:
    wb = _build_contribution_workbook(body.result)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (body.result.target_label or "target").replace(" ", "-")[:40]
    date_tag = datetime.date.today().isoformat()
    filename = f"MApper_LCA_Contribution_{safe}_{date_tag}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _build_multi_year_workbook(result: MultiYearContributionResult):
    """Multi-year contribution workbook. Layout:
      • Summary — target, method, years, computed_at, mapper_version
      • Trajectory — Year, Score, Database, Has warnings
      • Evolution — one row per contributor × union of years (top-N union)
      • Per-year sheets — same shape as single-year export, named ``Y{year}``
      • Warnings — flat list (per-year prefixed)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="3ECFCF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _style_header(ws, ncols: int) -> None:
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.freeze_panes = "A2"

    def _auto_width(ws) -> None:
        for col in ws.columns:
            mx = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                mx = max(mx, len(val))
            ws.column_dimensions[letter].width = min(mx + 3, 50)

    method_label = " › ".join(result.method)

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    _style_header(ws, 2)
    ws.append(["Generated", datetime.datetime.now().strftime("%Y-%m-%d %H:%M")])
    ws.append(["Target type", result.target_type])
    ws.append(["Target", result.target_label])
    ws.append(["Method", method_label])
    ws.append(["Method unit", result.method_unit])
    ws.append(["Compute DB pattern", result.compute_database_pattern or "—"])
    ws.append(["Years", ", ".join(str(y) for y in result.years)])
    ws.append(["Cutoff", result.cutoff])
    ws.append(["Max depth", result.max_depth])
    ws.append(["Elapsed (s)", result.elapsed_seconds])
    ws.append(["Computed at (UTC)", result.computed_at or "—"])
    ws.append(["MApper version", result.mapper_version or "—"])
    _auto_width(ws)

    # Trajectory
    ws_t = wb.create_sheet("Trajectory")
    ws_t.append(["Year", "Score", f"Unit ({result.method_unit})", "Database", "Has warnings"])
    _style_header(ws_t, 5)
    for p in result.trajectory:
        ws_t.append([p.year, p.score, result.method_unit, p.compute_database or "—", p.has_warnings])
    _auto_width(ws_t)

    # Evolution
    ws_e = wb.create_sheet("Evolution")
    year_cols = [str(y) for y in result.years]
    ws_e.append(["Activity", "Location", "Key", "Unit", *year_cols])
    _style_header(ws_e, 4 + len(year_cols))
    for ev in result.evolution:
        row = [ev.activity_name, ev.location, ev.activity_key, ev.unit]
        row.extend(ev.by_year.get(y, 0.0) for y in year_cols)
        ws_e.append(row)
    _auto_width(ws_e)

    # Per-year detail sheets
    for y in result.years:
        per = result.results.get(str(y))
        if per is None:
            continue
        ws_y = wb.create_sheet(f"Y{y}")
        ws_y.append(["Field", "Value"])
        _style_header(ws_y, 2)
        ws_y.append(["Year", y])
        ws_y.append(["Score", per.score])
        ws_y.append(["Unit", per.method_unit])
        ws_y.append(["Database", per.compute_database or "—"])
        ws_y.append([])
        ws_y.append(["Top technosphere"])
        ws_y.append(["Activity", "Location", "Key", "Impact", "Unit", "%"])
        for it in per.top_technosphere.items:
            ws_y.append([it.activity_name, it.location, it.activity_key, it.amount, it.unit, it.percentage])
        ws_y.append(["Rest", "", "", per.top_technosphere.rest_amount, per.method_unit,
                     per.top_technosphere.rest_percentage])
        ws_y.append([])
        ws_y.append(["Top biosphere"])
        ws_y.append(["Flow", "Compartment", "Subcompartment", "Inv amount", "Inv unit",
                     "Impact", "Unit", "%"])
        for b in per.top_biosphere:
            ws_y.append([b.flow_name, b.compartment, b.subcompartment,
                         b.inventory_amount, b.inventory_unit,
                         b.amount, b.unit, b.percentage])
        _auto_width(ws_y)

    # Warnings
    ws_w = wb.create_sheet("Warnings")
    ws_w.append(["#", "Message"])
    _style_header(ws_w, 2)
    if result.warnings:
        for i, w in enumerate(result.warnings, start=1):
            ws_w.append([i, w])
    else:
        ws_w.append(["—", "No warnings."])
    _auto_width(ws_w)

    return wb


@router.post("/lca/contribution-analysis/multi-year/export")
async def export_multi_year_contribution(
    body: MultiYearContributionExportRequest,
) -> Response:
    wb = _build_multi_year_workbook(body.result)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = (body.result.target_label or "target").replace(" ", "-")[:40]
    span = (
        f"{body.result.years[0]}-{body.result.years[-1]}"
        if body.result.years else "multiyear"
    )
    date_tag = datetime.date.today().isoformat()
    filename = f"MApper_LCA_Trajectory_{safe}_{span}_{date_tag}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/lca/export-archetype")
async def export_archetype_lca(body: ArchetypeLCAExportRequest) -> Response:
    wb = _build_lca_export_workbook(body.results)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    names = [d.archetype_name for d in body.results]
    safe = "_".join(n.replace(" ", "-") for n in names[:3]) or "archetype"
    scope = body.results[0].scope if body.results else "all"
    date_tag = datetime.date.today().isoformat()
    filename = f"MApper_LCA_{safe}_{scope}_{date_tag}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Multi-product LCA comparison (Patch 4AG.1) ─────────────────────────────────


def _activity_label(database: str, code: str) -> str:
    """Best-effort human-readable label for an activity. Falls back to
    the bare code when bw2data lookup fails (e.g. activity missing in
    the active project). Used to populate the per-item label in the
    multi-product response so charts and tables can render a name
    even when the underlying compute fails."""
    try:
        act = bw2data.get_activity((database, code))
        return act.get("reference product", act.get("name", "")) or code
    except Exception:
        return code


@router.post("/lca/calculate-multi-product", response_model=MultiProductLCAResult)
async def calculate_multi_product_lca(body: MultiProductLCARequest) -> MultiProductLCAResult:
    """Compute N independent LCAs (mixed archetype + activity items)
    for side-by-side comparison.

    Distinct from ``/lca/calculate-activities``, which treats N
    activities as a SINGLE combined demand (contributions sum to a
    total). This endpoint treats N items as N SEPARATE LCAs — each
    item's result is independent, side-by-side comparable across
    items.

    Per-item error isolation: a failure on one item does not abort
    the fan-out. The failing item's response slot carries
    ``status="error"`` with an ``error_message``; the remaining
    items still compute. The aggregate ``success_count`` /
    ``error_count`` fields summarise the run.
    """
    t0 = time.perf_counter()

    if not body.items:
        raise HTTPException(status_code=400, detail="At least one item is required")
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required")

    results: list[MultiProductItemResult] = []
    success_count = 0
    error_count = 0

    for item in body.items:
        # Discriminated dispatch — Pydantic 2 narrows the union by
        # `type`. Each branch builds the corresponding single-product
        # request body and calls the existing endpoint's async
        # handler directly (FastAPI handlers are plain async functions
        # under the hood, callable in-process without HTTP round-trip).
        if item.type == "archetype":
            try:
                arc_body = ArchetypeLCACalculateRequest(
                    archetype_id=item.archetype_id,
                    scope=body.scope,
                    methods=body.methods,
                    compute_database=body.compute_database,
                    parameter_scenario=item.parameter_scenario,
                    stage_amounts=item.stage_amounts,
                )
                arc_res = await calculate_archetype_lca(arc_body)
                results.append(MultiProductItemResult(
                    type="archetype",
                    item_id=item.archetype_id,
                    label=arc_res.archetype_name,
                    status="success",
                    archetype_result=arc_res,
                ))
                success_count += 1
            except HTTPException as e:
                results.append(MultiProductItemResult(
                    type="archetype",
                    item_id=item.archetype_id,
                    label=item.archetype_id,
                    status="error",
                    error_message=str(e.detail),
                ))
                error_count += 1
            except Exception as e:  # defensive — any uncaught exception is an item-level failure
                results.append(MultiProductItemResult(
                    type="archetype",
                    item_id=item.archetype_id,
                    label=item.archetype_id,
                    status="error",
                    error_message=f"{type(e).__name__}: {e}",
                ))
                error_count += 1

        else:  # item.type == "activity"
            # `item_id` is unique per (activity × vintage) because the
            # vintage's DB name is part of the key — two vintages of one
            # activity get distinct ids/colors. The label is vintage-aware
            # so they don't collide on a chart axis.
            item_id = f"{item.database}|{item.code}"
            act_label = _activity_label(item.database, item.code)
            if item.vintage_label:
                act_label = f"{act_label} [{item.vintage_label}]"
            try:
                act_body = ActivityLCARequest(
                    activities=[ActivityDemandItem(
                        database=item.database, code=item.code, amount=item.amount,
                    )],
                    methods=body.methods,
                )
                # Per-item DB: `calculate_activity_lca` resolves the activity
                # in `item.database` directly (premise preserves codes), so
                # the LCA runs against THAT vintage's technosphere. A premise
                # SSP×year DB name here computes against that prospective DB.
                act_res = await calculate_activity_lca(act_body)
                results.append(MultiProductItemResult(
                    type="activity",
                    item_id=item_id,
                    label=act_label,
                    status="success",
                    activity_result=act_res,
                ))
                success_count += 1
            except HTTPException as e:
                results.append(MultiProductItemResult(
                    type="activity",
                    item_id=item_id,
                    label=act_label,
                    status="error",
                    error_message=str(e.detail),
                ))
                error_count += 1
            except Exception as e:
                results.append(MultiProductItemResult(
                    type="activity",
                    item_id=item_id,
                    label=act_label,
                    status="error",
                    error_message=f"{type(e).__name__}: {e}",
                ))
                error_count += 1

    elapsed = round(time.perf_counter() - t0, 2)
    return MultiProductLCAResult(
        items=results,
        elapsed_seconds=elapsed,
        success_count=success_count,
        error_count=error_count,
    )
