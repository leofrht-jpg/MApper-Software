"""FastAPI router for the MFA dynamic stock module (Phase 2A)."""
from __future__ import annotations

import datetime
import io
import os
import re
import threading
import uuid
from typing import Optional

import bw2data
from fastapi import APIRouter, File, HTTPException, UploadFile
from fastapi.responses import Response

from mapper.core.mfa_engine import (
    DEFAULT_WEIBULL_SCALE,
    DEFAULT_WEIBULL_SHAPE,
    SUPPORTED_EXTS,
    DynamicStockModel,
    all_cohort_keys,
    cohort_key_to_dict,
    inflow_template_csv,
    non_age_dimensions,
    parse_first_column_labels,
    parse_inflow_file,
    parse_stock_file,
    stock_template_csv,
    survival_curve,
)
from mapper.core import mfa_storage
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from mapper.models.mfa_schemas import (
    DimensionDef,
    InflowData,
    InflowUploadResult,
    MFASystemState,
    SimulationResult,
    SimulationSummary,
    StockUploadResult,
    SurvivalConfig,
    SurvivalConfigList,
    SurvivalPreviewPoint,
    SurvivalSetResult,
    SystemDefinition,
    SystemSummary,
    SystemUpdateResponse,
    TimeHorizon,
    YearResult,
)


router = APIRouter(prefix="/mfa", tags=["mfa"])

# In-memory stores, hydrated from disk at startup. All stores are nested by
# bw2 project name: outer key = project, inner key = system id.
_systems: dict[str, dict[str, SystemDefinition]] = {}
_states: dict[str, dict[str, MFASystemState]] = {}
_results: dict[str, dict[str, SimulationResult]] = {}
_lock = threading.Lock()


def _current_project() -> str:
    return bw2data.projects.current


def _proj_systems(project: str | None = None) -> dict[str, SystemDefinition]:
    p = project or _current_project()
    return _systems.setdefault(p, {})


def _proj_states(project: str | None = None) -> dict[str, MFASystemState]:
    p = project or _current_project()
    return _states.setdefault(p, {})


def _proj_results(project: str | None = None) -> dict[str, SimulationResult]:
    p = project or _current_project()
    return _results.setdefault(p, {})


def hydrate_from_disk() -> None:
    """Load persisted systems/states/results into the in-memory stores."""
    systems, states, results, mappings, archetypes = mfa_storage.load_all()
    _systems.update(systems)
    _states.update(states)
    _results.update(results)
    # Cohort mappings + archetypes live in bom.py — install them there.
    try:
        from mapper.api import bom as _bom
        if mappings:
            _bom._cohort_mappings.update(mappings)
        if archetypes:
            _bom._archetypes.update(archetypes)
    except Exception:
        pass


def _now_iso() -> str:
    return datetime.datetime.now().isoformat()


_FILENAME_UNSAFE = re.compile(r"[^A-Za-z0-9._-]+")


def _sanitize_filename(name: str, fallback: str = "system", max_len: int = 100) -> str:
    """Produce a filesystem-safe base name for downloads.

    Replaces whitespace with underscores, strips non-alphanumeric chars (keeping
    ``._-``), collapses repeats, and trims to ``max_len``.
    """
    cleaned = (name or "").strip().replace(" ", "_")
    cleaned = _FILENAME_UNSAFE.sub("", cleaned).strip("._-")
    if not cleaned:
        return fallback
    return cleaned[:max_len]


def _get_system(system_id: str) -> SystemDefinition:
    sys_def = _proj_systems().get(system_id)
    if not sys_def:
        raise HTTPException(status_code=404, detail=f"System '{system_id}' not found")
    return sys_def


def _get_or_create_state(system_id: str) -> MFASystemState:
    states = _proj_states()
    state = states.get(system_id)
    if state is None:
        state = MFASystemState(system_id=system_id)
        states[system_id] = state
    return state


def _validate_definition(definition: SystemDefinition) -> None:
    if definition.time_horizon.end_year < definition.time_horizon.start_year:
        raise HTTPException(status_code=400, detail="end_year must be ≥ start_year")
    nads = non_age_dimensions(definition.dimensions)
    if not nads:
        raise HTTPException(
            status_code=400,
            detail="At least one non-age dimension is required.",
        )
    seen_names: set[str] = set()
    for d in nads:
        if not d.name:
            raise HTTPException(status_code=400, detail="Every dimension needs a name.")
        if d.name in seen_names:
            raise HTTPException(status_code=400, detail=f"Duplicate dimension name: {d.name}")
        seen_names.add(d.name)
        if not d.labels:
            raise HTTPException(
                status_code=400,
                detail=f"Dimension '{d.name}' must have at least one label.",
            )


def _summary(definition: SystemDefinition) -> SystemSummary:
    return SystemSummary(
        id=definition.id or "",
        name=definition.name,
        description=definition.description,
        time_horizon=definition.time_horizon,
        dimension_count=len(non_age_dimensions(definition.dimensions)),
        cohort_count=len(all_cohort_keys(definition.dimensions)),
        created_at=definition.created_at or _now_iso(),
    )


# ── System CRUD ──────────────────────────────────────────────────────────────


@router.post("/systems", response_model=SystemDefinition)
async def create_system(body: SystemDefinition) -> SystemDefinition:
    _validate_definition(body)
    project = _current_project()
    with _lock:
        sid = str(uuid.uuid4())
        body.id = sid
        body.created_at = _now_iso()
        _proj_systems(project)[sid] = body
        _proj_states(project)[sid] = MFASystemState(system_id=sid)
    mfa_storage.save_system(project, body)
    mfa_storage.save_state(project, sid, _proj_states(project)[sid])
    return body


@router.get("/systems", response_model=list[SystemSummary])
async def list_systems() -> list[SystemSummary]:
    return [_summary(s) for s in _proj_systems().values()]


@router.get("/systems/{system_id}", response_model=SystemDefinition)
async def get_system(system_id: str) -> SystemDefinition:
    return _get_system(system_id)


def _pair_renames(old_labels: list[str], new_labels: list[str]) -> dict[str, str]:
    """Heuristic rename detection per dimension.

    Labels in the set intersection are unchanged. Labels only in old are "removed";
    labels only in new are "added". When the counts match, pair them by order and
    treat as renames; otherwise report add/remove (no rename pairing).
    """
    common = set(old_labels) & set(new_labels)
    old_only = [l for l in old_labels if l not in common]
    new_only = [l for l in new_labels if l not in common]
    if old_only and len(old_only) == len(new_only):
        return dict(zip(old_only, new_only))
    return {}


def _migrate_state(
    old_def: SystemDefinition,
    new_def: SystemDefinition,
    state: MFASystemState,
) -> tuple[MFASystemState, list[str]]:
    """Best-effort migration of MFA state data across a dimension change.

    Strategy:
      - Match non-age dims by machine name.
      - Per-dim heuristic rename pairing when label counts are equal.
      - Orphaned rows (refs to removed labels or removed dims) are dropped + counted.
      - Dim add/remove or reorder → cohort_key format changes, translate via name map;
        if the new key can't be fully resolved, drop that row.
    """
    warnings: list[str] = []
    old_nads = non_age_dimensions(old_def.dimensions)
    new_nads = non_age_dimensions(new_def.dimensions)
    old_by_name = {d.name: d for d in old_nads}
    new_by_name = {d.name: d for d in new_nads}

    removed_dims = [n for n in old_by_name if n not in new_by_name]
    added_dims = [n for n in new_by_name if n not in old_by_name]
    if removed_dims:
        warnings.append(f"Removed dimension(s): {', '.join(removed_dims)}.")
    if added_dims:
        warnings.append(f"Added dimension(s) with no prior data: {', '.join(added_dims)}.")

    # Build per-dim label translation maps (old_label -> new_label or None)
    label_trans: dict[str, dict[str, str | None]] = {}
    for name, old_dim in old_by_name.items():
        if name not in new_by_name:
            continue
        new_dim = new_by_name[name]
        renames = _pair_renames(old_dim.labels, new_dim.labels)
        tmap: dict[str, str | None] = {}
        for lab in old_dim.labels:
            new_lab = renames.get(lab, lab)
            tmap[lab] = new_lab if new_lab in new_dim.labels else None
        label_trans[name] = tmap
        # Surface per-dim warnings
        renamed_n = len(renames)
        dropped = [l for l, nl in tmap.items() if nl is None]
        added_labels = [l for l in new_dim.labels if l not in old_dim.labels and l not in renames.values()]
        if renamed_n:
            pairs = ", ".join(f"'{o}'→'{n}'" for o, n in renames.items())
            warnings.append(f"Renamed label(s) in {name}: {pairs}.")
        if dropped:
            warnings.append(f"Removed label(s) from {name}: {', '.join(dropped)}.")
        if added_labels:
            warnings.append(f"Added label(s) to {name}: {', '.join(added_labels)}.")

    # Added dims need a default label value to extend old keys; pick first label.
    added_defaults: dict[str, str | None] = {}
    for name in added_dims:
        labels = new_by_name[name].labels
        added_defaults[name] = labels[0] if labels else None

    def translate_cohort_key(old_ck: str) -> str | None:
        parts = old_ck.split("|") if old_ck else []
        if len(parts) != len(old_nads):
            return None
        # Map old values into a dict keyed by old dim name
        old_vals: dict[str, str] = {}
        for i, dim in enumerate(old_nads):
            old_vals[dim.name] = parts[i]
        # Build new cohort components in new dim order
        new_parts: list[str] = []
        for new_dim in new_nads:
            if new_dim.name in old_by_name:
                old_label = old_vals.get(new_dim.name)
                if old_label is None:
                    return None
                translated = label_trans.get(new_dim.name, {}).get(old_label)
                if translated is None:
                    return None
                new_parts.append(translated)
            else:
                default = added_defaults.get(new_dim.name)
                if default is None:
                    return None
                new_parts.append(default)
        return "|".join(new_parts)

    # Migrate initial_stock (keyed by f"{cohort_key}|{age}")
    new_stock: dict[str, float] = {}
    orphaned_stock = 0
    horizon_len = new_def.time_horizon.length
    for full_key, count in state.initial_stock.items():
        cohort_part, _, age_part = full_key.rpartition("|")
        try:
            age = int(age_part)
        except ValueError:
            orphaned_stock += 1
            continue
        new_ck = translate_cohort_key(cohort_part)
        if new_ck is None or age >= horizon_len:
            orphaned_stock += 1
            continue
        new_stock[f"{new_ck}|{age}"] = new_stock.get(f"{new_ck}|{age}", 0.0) + count
    state.initial_stock = new_stock
    if orphaned_stock:
        warnings.append(f"Dropped {orphaned_stock} orphaned stock row(s) during migration.")

    # Migrate inflows
    new_inflows: list[InflowData] = []
    orphaned_inflow = 0
    year_range = set(new_def.time_horizon.years)
    for inflow in state.inflows:
        if inflow.year not in year_range:
            orphaned_inflow += len(inflow.counts)
            continue
        new_counts: dict[str, float] = {}
        for ck, count in inflow.counts.items():
            new_ck = translate_cohort_key(ck)
            if new_ck is None:
                orphaned_inflow += 1
                continue
            new_counts[new_ck] = new_counts.get(new_ck, 0.0) + count
        if new_counts:
            new_inflows.append(InflowData(year=inflow.year, counts=new_counts))
    state.inflows = new_inflows
    if orphaned_inflow:
        warnings.append(f"Dropped {orphaned_inflow} orphaned inflow cohort row(s) during migration.")

    # Migrate survival_configs (dim filters)
    new_configs: list[SurvivalConfig] = []
    orphaned_survival = 0
    for cfg in state.survival_configs:
        new_filters: dict[str, str] = {}
        keep = True
        for dim_name, label in cfg.dimension_filters.items():
            if dim_name not in new_by_name:
                keep = False
                break
            tmap = label_trans.get(dim_name, {})
            new_label = tmap.get(label, label if label in new_by_name[dim_name].labels else None)
            if new_label is None:
                keep = False
                break
            new_filters[dim_name] = new_label
        if keep:
            cfg.dimension_filters = new_filters
            new_configs.append(cfg)
        else:
            orphaned_survival += 1
    state.survival_configs = new_configs
    if orphaned_survival:
        warnings.append(f"Dropped {orphaned_survival} survival config(s) that referenced removed labels/dimensions.")

    return state, warnings


@router.put("/systems/{system_id}", response_model=SystemUpdateResponse)
async def update_system(system_id: str, body: SystemDefinition) -> SystemUpdateResponse:
    existing = _get_system(system_id)
    _validate_definition(body)
    project = _current_project()
    with _lock:
        body.id = system_id
        body.created_at = existing.created_at
        state = _get_or_create_state(system_id)
        migrated_state, warnings = _migrate_state(existing, body, state)
        _proj_systems(project)[system_id] = body
        _proj_states(project)[system_id] = migrated_state
        # Dimension/label/horizon changes invalidate prior simulation results.
        _proj_results(project).pop(system_id, None)
    mfa_storage.save_system(project, body)
    mfa_storage.save_state(project, system_id, migrated_state)
    mfa_storage.clear_results(project, system_id)
    return SystemUpdateResponse(system=body, warnings=warnings)


@router.delete("/systems/{system_id}")
async def delete_system(system_id: str) -> dict[str, bool]:
    project = _current_project()
    if system_id not in _proj_systems(project):
        raise HTTPException(status_code=404, detail="System not found")
    with _lock:
        _proj_systems(project).pop(system_id, None)
        _proj_states(project).pop(system_id, None)
        _proj_results(project).pop(system_id, None)
    mfa_storage.delete_system_dir(project, system_id)
    try:
        from mapper.api import bom as _bom
        _bom.purge_system(system_id)
    except Exception:
        pass
    return {"deleted": True}


# ── State endpoints ──────────────────────────────────────────────────────────


@router.get("/systems/{system_id}/state", response_model=MFASystemState)
async def get_state(system_id: str) -> MFASystemState:
    _get_system(system_id)
    return _get_or_create_state(system_id)


# ── Label file parser (CSV / XLSX) ───────────────────────────────────────────


@router.post("/parse-labels")
async def parse_labels(file: UploadFile = File(...)) -> dict[str, list[str]]:
    """Extract unique first-column values from a .csv or .xlsx file.

    Used by the dimension editor to support Excel label imports without
    bundling a JS xlsx library.
    """
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        labels = parse_first_column_labels(raw, filename)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    return {"labels": labels}


# ── Stock upload ─────────────────────────────────────────────────────────────


def _check_ext(filename: str) -> None:
    ext = os.path.splitext(filename or "")[1].lower()
    if ext not in SUPPORTED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.",
        )


@router.post("/systems/{system_id}/stock/upload", response_model=StockUploadResult)
async def upload_stock(system_id: str, file: UploadFile = File(...)) -> StockUploadResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        parsed, rows = parse_stock_file(raw, filename, sys_def.dimensions)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    state.initial_stock = parsed
    mfa_storage.save_state(_current_project(), system_id, state)
    cohort_keys: set[str] = set()
    total = 0.0
    for full_key, count in parsed.items():
        ck = full_key.rsplit("|", 1)[0]
        cohort_keys.add(ck)
        total += count
    return StockUploadResult(
        rows_parsed=rows, cohorts_found=len(cohort_keys), total_items=total
    )


# ── Inflow upload ────────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/inflows/upload", response_model=InflowUploadResult)
async def upload_inflows(system_id: str, file: UploadFile = File(...)) -> InflowUploadResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    filename = file.filename or ""
    _check_ext(filename)
    raw = await file.read()
    try:
        inflows, rows = parse_inflow_file(
            raw, filename, sys_def.dimensions, sys_def.time_horizon.years
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    state.inflows = inflows
    mfa_storage.save_state(_current_project(), system_id, state)
    total = sum(sum(inf.counts.values()) for inf in inflows)
    return InflowUploadResult(
        years_parsed=len(inflows), rows_parsed=rows, total_inflows=total
    )


# ── Survival ─────────────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/survival", response_model=SurvivalSetResult)
async def set_survival(system_id: str, body: SurvivalConfigList) -> SurvivalSetResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    label_index = {d.name: set(d.labels) for d in non_age_dimensions(sys_def.dimensions)}
    for cfg in body.configs:
        for k, v in cfg.dimension_filters.items():
            if k not in label_index:
                raise HTTPException(
                    status_code=400, detail=f"Unknown dimension '{k}' in survival filter."
                )
            if v not in label_index[k]:
                raise HTTPException(
                    status_code=400,
                    detail=f"Label '{v}' not in dimension '{k}' (allowed: {sorted(label_index[k])}).",
                )
    state.survival_configs = body.configs
    mfa_storage.save_state(_current_project(), system_id, state)
    return SurvivalSetResult(configs_set=len(body.configs))


@router.get("/systems/{system_id}/survival", response_model=SurvivalConfigList)
async def get_survival(system_id: str) -> SurvivalConfigList:
    _get_system(system_id)
    state = _get_or_create_state(system_id)
    return SurvivalConfigList(configs=state.survival_configs)


@router.get(
    "/systems/{system_id}/survival/preview", response_model=list[SurvivalPreviewPoint]
)
async def preview_survival(
    system_id: str,
    shape: float = DEFAULT_WEIBULL_SHAPE,
    scale: float = DEFAULT_WEIBULL_SCALE,
    max_age: Optional[int] = None,
) -> list[SurvivalPreviewPoint]:
    sys_def = _get_system(system_id)
    if shape <= 0 or scale <= 0:
        raise HTTPException(status_code=400, detail="shape and scale must be > 0")
    horizon_len = sys_def.time_horizon.length
    cap = max_age if max_age is not None and max_age > 0 else horizon_len
    return survival_curve(shape, scale, cap)


# ── Simulation ──────────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/simulate", response_model=SimulationResult)
async def simulate(system_id: str) -> SimulationResult:
    sys_def = _get_system(system_id)
    state = _get_or_create_state(system_id)
    try:
        model = DynamicStockModel(sys_def, state)
        result = model.simulate()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Simulation failed: {e}")
    project = _current_project()
    _proj_results(project)[system_id] = result
    mfa_storage.save_results(project, system_id, result)
    return result


@router.get("/systems/{system_id}/results", response_model=SimulationResult)
async def get_results(system_id: str) -> SimulationResult:
    _get_system(system_id)
    res = _proj_results().get(system_id)
    if res is None:
        raise HTTPException(
            status_code=404, detail="No simulation results yet. Run /simulate first."
        )
    return res


# ── Export ───────────────────────────────────────────────────────────────────


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4B4690")  # blue/purple
_COUNT_FMT = "#,##0.00"
_SHEET_COLORS = {
    "Summary": "4B4690",
    "Stock by Year": "5B7DB1",
    "Inflows by Year": "3E8E7E",
    "Outflows by Year": "A65A5A",
    "Age Distribution": "8064A2",
    "Mass Balance": "2F4F7F",
}


def _write_header(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"


def _auto_width(ws, headers: list[str], sample_rows: int = 20) -> None:
    """Set approximate column widths based on header + a sampling of rows."""
    max_row = min(ws.max_row, sample_rows + 1)
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(len(str(header)), 10)
        for row_idx in range(2, max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            width = max(width, min(len(str(v)), 40))
        ws.column_dimensions[letter].width = width + 2


def _apply_count_format(ws, col_indices: list[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        for col_idx in col_indices:
            if col_idx <= len(row):
                row[col_idx - 1].number_format = _COUNT_FMT


def _build_summary_sheet(ws, sys_def: SystemDefinition, state: MFASystemState, summary) -> None:
    ws.title = "Summary"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Summary"]

    bold = Font(bold=True)
    rows: list[tuple[str, object]] = [
        ("System name", sys_def.name),
        ("Description", sys_def.description or ""),
        ("Start year", sys_def.time_horizon.start_year),
        ("End year", sys_def.time_horizon.end_year),
        ("Horizon length (years)", sys_def.time_horizon.length),
        ("Created at", sys_def.created_at or ""),
        ("", ""),
        ("Total stock at start", summary.total_stock_start),
        ("Total stock at end", summary.total_stock_end),
        ("Total inflows", summary.total_inflows),
        ("Total outflows", summary.total_outflows),
        ("Net change", summary.total_inflows - summary.total_outflows),
        ("", ""),
    ]
    r = 1
    for label, value in rows:
        ws.cell(row=r, column=1, value=label).font = bold
        ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)) and label.startswith(("Total", "Net")):
            ws.cell(row=r, column=2).number_format = _COUNT_FMT
        r += 1

    # Dimensions section
    ws.cell(row=r, column=1, value="Dimensions").font = Font(bold=True, size=12)
    r += 1
    ws.cell(row=r, column=1, value="Name").font = bold
    ws.cell(row=r, column=2, value="Display name").font = bold
    ws.cell(row=r, column=3, value="Labels").font = bold
    r += 1
    for d in sys_def.dimensions:
        ws.cell(row=r, column=1, value=d.name)
        ws.cell(row=r, column=2, value=d.display_name or d.name)
        if d.is_age:
            ws.cell(
                row=r, column=3,
                value=f"auto-generated 0 – {sys_def.time_horizon.length - 1}",
            )
        else:
            ws.cell(row=r, column=3, value=", ".join(d.labels))
        r += 1
    r += 1

    # Survival section
    ws.cell(row=r, column=1, value="Survival configuration").font = Font(bold=True, size=12)
    r += 1
    headers = ["Filter", "Method", "Weibull shape", "Weibull scale", "Custom points"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=r, column=i, value=h).font = bold
    r += 1
    if not state.survival_configs:
        ws.cell(row=r, column=1, value="(default)")
        ws.cell(row=r, column=2, value="weibull")
        ws.cell(row=r, column=3, value=DEFAULT_WEIBULL_SHAPE)
        ws.cell(row=r, column=4, value=DEFAULT_WEIBULL_SCALE)
        r += 1
    else:
        for cfg in state.survival_configs:
            filter_str = (
                ", ".join(f"{k}={v}" for k, v in cfg.dimension_filters.items())
                if cfg.dimension_filters else "(default)"
            )
            ws.cell(row=r, column=1, value=filter_str)
            ws.cell(row=r, column=2, value=cfg.method)
            ws.cell(row=r, column=3, value=cfg.weibull_shape)
            ws.cell(row=r, column=4, value=cfg.weibull_scale)
            if cfg.custom_curve:
                ws.cell(
                    row=r, column=5,
                    value="; ".join(f"{p.age}:{p.survival_rate}" for p in cfg.custom_curve),
                )
            r += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 36


def _dim_headers(nads) -> list[str]:
    return [d.display_name or d.name for d in nads]


def _cohort_values(cohort_key: str, sys_def: SystemDefinition, nads) -> list[str]:
    d = cohort_key_to_dict(cohort_key, sys_def.dimensions)
    return [d.get(dim.name, "") for dim in nads]


def _build_cohort_year_sheet(
    ws,
    title: str,
    sys_def: SystemDefinition,
    result: SimulationResult,
    value_key: str,
    value_header: str,
    include_zeros: bool,
) -> None:
    ws.title = title
    ws.sheet_properties.tabColor = _SHEET_COLORS[title]
    nads = non_age_dimensions(sys_def.dimensions)
    dim_headers = _dim_headers(nads)
    headers = ["Year"] + dim_headers + [value_header]
    _write_header(ws, headers)

    r = 2
    for yr in result.years:
        data = getattr(yr, value_key)
        for ck, count in data.items():
            if not include_zeros and count == 0:
                continue
            ws.cell(row=r, column=1, value=yr.year)
            values = _cohort_values(ck, sys_def, nads)
            for i, v in enumerate(values, start=2):
                ws.cell(row=r, column=i, value=v)
            ws.cell(row=r, column=len(headers), value=float(count)).number_format = _COUNT_FMT
            r += 1

    # For stock, append a yearly total per dimension-less bucket at the end.
    if value_key == "stock":
        r += 1
        bold = Font(bold=True)
        ws.cell(row=r, column=1, value="Yearly totals").font = bold
        r += 1
        ws.cell(row=r, column=1, value="Year").font = bold
        ws.cell(row=r, column=2, value="Total Stock").font = bold
        r += 1
        for yr in result.years:
            ws.cell(row=r, column=1, value=yr.year)
            total = sum(yr.stock.values())
            ws.cell(row=r, column=2, value=total).number_format = _COUNT_FMT
            r += 1

    _auto_width(ws, headers)


def _build_age_sheet(ws, sys_def: SystemDefinition, result: SimulationResult) -> None:
    ws.title = "Age Distribution"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Age Distribution"]
    nads = non_age_dimensions(sys_def.dimensions)
    dim_headers = _dim_headers(nads)
    headers = ["Year"] + dim_headers + ["Age", "Count"]
    _write_header(ws, headers)

    r = 2
    for yr in result.years:
        for ck, by_age in yr.stock_by_age.items():
            values = _cohort_values(ck, sys_def, nads)
            for age, count in sorted(by_age.items()):
                if count == 0:
                    continue
                ws.cell(row=r, column=1, value=yr.year)
                for i, v in enumerate(values, start=2):
                    ws.cell(row=r, column=i, value=v)
                ws.cell(row=r, column=len(headers) - 1, value=age)
                ws.cell(row=r, column=len(headers), value=float(count)).number_format = _COUNT_FMT
                r += 1
    _auto_width(ws, headers)


def _build_mass_balance_sheet(ws, result: SimulationResult) -> None:
    ws.title = "Mass Balance"
    ws.sheet_properties.tabColor = _SHEET_COLORS["Mass Balance"]
    headers = ["Year", "Total Stock", "Total Inflow", "Total Outflow", "Net Change"]
    _write_header(ws, headers)

    for i, yr in enumerate(result.years, start=2):
        total_stock = sum(yr.stock.values())
        total_in = sum(yr.inflow.values())
        total_out = sum(yr.outflow.values())
        ws.cell(row=i, column=1, value=yr.year)
        ws.cell(row=i, column=2, value=total_stock).number_format = _COUNT_FMT
        ws.cell(row=i, column=3, value=total_in).number_format = _COUNT_FMT
        ws.cell(row=i, column=4, value=total_out).number_format = _COUNT_FMT
        ws.cell(row=i, column=5, value=total_in - total_out).number_format = _COUNT_FMT

    _auto_width(ws, headers)


def _build_export_workbook(
    sys_def: SystemDefinition, state: MFASystemState, result: SimulationResult
) -> bytes:
    wb = Workbook()
    # First sheet: summary (replace the default sheet)
    _build_summary_sheet(wb.active, sys_def, state, result.summary)

    _build_cohort_year_sheet(
        wb.create_sheet(), "Stock by Year", sys_def, result,
        value_key="stock", value_header="Total Stock", include_zeros=False,
    )
    _build_cohort_year_sheet(
        wb.create_sheet(), "Inflows by Year", sys_def, result,
        value_key="inflow", value_header="Inflow Count", include_zeros=False,
    )
    _build_cohort_year_sheet(
        wb.create_sheet(), "Outflows by Year", sys_def, result,
        value_key="outflow", value_header="Outflow Count", include_zeros=False,
    )
    _build_age_sheet(wb.create_sheet(), sys_def, result)
    _build_mass_balance_sheet(wb.create_sheet(), result)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/systems/{system_id}/export")
async def export_results(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    result = _proj_results().get(system_id)
    if result is None:
        raise HTTPException(
            status_code=400,
            detail="No simulation results available. Run /simulate first.",
        )
    state = _get_or_create_state(system_id)
    content = _build_export_workbook(sys_def, state, result)
    fname = f"{_sanitize_filename(sys_def.name)}_simulation.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Import from previously exported Excel ───────────────────────────────────


def _dim_header_to_name_map(sys_def: SystemDefinition) -> dict[str, str]:
    """Map both display_name and machine name → machine name."""
    m: dict[str, str] = {}
    for d in non_age_dimensions(sys_def.dimensions):
        m[d.name] = d.name
        if d.display_name:
            m[d.display_name] = d.name
    return m


def _build_cohort_key(row_values: dict[str, str], sys_def: SystemDefinition) -> str | None:
    nads = non_age_dimensions(sys_def.dimensions)
    parts: list[str] = []
    for d in nads:
        v = row_values.get(d.name)
        if v is None or v == "":
            return None
        parts.append(str(v))
    return "|".join(parts)


def _parse_sheet_rows(ws, sys_def: SystemDefinition) -> list[tuple[int, str, float]]:
    """Parse a "X by Year" sheet into [(year, cohort_key, value)] rows.

    Stops at blank row (end of data) to skip any trailing totals section.
    """
    rows: list[tuple[int, str, float]] = []
    header_map = _dim_header_to_name_map(sys_def)
    headers: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        if row is None or all(c is None or c == "" for c in row):
            break
        if not row or row[0] is None:
            break
        try:
            year = int(row[0])
        except (TypeError, ValueError):
            break
        value_cell = row[-1]
        try:
            value = float(value_cell) if value_cell is not None else 0.0
        except (TypeError, ValueError):
            continue
        vals: dict[str, str] = {}
        for i, header in enumerate(headers[1:-1], start=1):
            machine = header_map.get(header, header)
            cell = row[i] if i < len(row) else None
            vals[machine] = "" if cell is None else str(cell)
        ck = _build_cohort_key(vals, sys_def)
        if ck is None:
            continue
        rows.append((year, ck, value))
    return rows


def _parse_age_sheet(ws, sys_def: SystemDefinition) -> list[tuple[int, str, int, float]]:
    """Parse the Age Distribution sheet into [(year, cohort_key, age, count)]."""
    rows: list[tuple[int, str, int, float]] = []
    header_map = _dim_header_to_name_map(sys_def)
    headers: list[str] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row_idx == 1:
            headers = [str(c) if c is not None else "" for c in row]
            continue
        if row is None or all(c is None or c == "" for c in row):
            break
        try:
            year = int(row[0])
            age = int(row[-2])
            count = float(row[-1])
        except (TypeError, ValueError):
            continue
        vals: dict[str, str] = {}
        # dim columns are at indices 1 .. len-2 (excluding Year, Age, Count)
        for i, header in enumerate(headers[1:-2], start=1):
            machine = header_map.get(header, header)
            cell = row[i] if i < len(row) else None
            vals[machine] = "" if cell is None else str(cell)
        ck = _build_cohort_key(vals, sys_def)
        if ck is None:
            continue
        rows.append((year, ck, age, count))
    return rows


def _reconstruct_simulation(
    sys_def: SystemDefinition, wb
) -> tuple[SimulationResult, list[str], list[str]]:
    """Parse an exported workbook back into a SimulationResult.

    Returns ``(result, warnings, cohorts)``.
    """
    warnings: list[str] = []

    def _read(title: str) -> list[tuple[int, str, float]]:
        if title not in wb.sheetnames:
            warnings.append(f"Missing sheet: {title}")
            return []
        return _parse_sheet_rows(wb[title], sys_def)

    stock_rows = _read("Stock by Year")
    inflow_rows = _read("Inflows by Year")
    outflow_rows = _read("Outflows by Year")
    age_rows: list[tuple[int, str, int, float]] = []
    if "Age Distribution" in wb.sheetnames:
        age_rows = _parse_age_sheet(wb["Age Distribution"], sys_def)
    else:
        warnings.append("Missing sheet: Age Distribution")

    def _by_year(rows: list[tuple[int, str, float]]) -> dict[int, dict[str, float]]:
        out: dict[int, dict[str, float]] = {}
        for year, ck, val in rows:
            out.setdefault(year, {})[ck] = val
        return out

    stock_by_year = _by_year(stock_rows)
    inflow_by_year = _by_year(inflow_rows)
    outflow_by_year = _by_year(outflow_rows)

    stock_by_age_by_year: dict[int, dict[str, dict[int, float]]] = {}
    for year, ck, age, cnt in age_rows:
        stock_by_age_by_year.setdefault(year, {}).setdefault(ck, {})[age] = cnt

    all_years = sorted(set(stock_by_year) | set(inflow_by_year) | set(outflow_by_year))
    if not all_years:
        raise HTTPException(status_code=400, detail="No simulation rows found in workbook.")

    cohorts: set[str] = set()
    year_results: list[YearResult] = []
    for y in all_years:
        stock = stock_by_year.get(y, {})
        inflow = inflow_by_year.get(y, {})
        outflow = outflow_by_year.get(y, {})
        stock_by_age = stock_by_age_by_year.get(y, {})
        cohorts.update(stock)
        cohorts.update(inflow)
        cohorts.update(outflow)
        year_results.append(
            YearResult(
                year=y,
                stock=stock,
                stock_by_age=stock_by_age,
                inflow=inflow,
                outflow=outflow,
                outflow_by_age={},
            )
        )

    total_in = sum(sum(r.inflow.values()) for r in year_results)
    total_out = sum(sum(r.outflow.values()) for r in year_results)
    summary = SimulationSummary(
        total_stock_start=sum(year_results[0].stock.values()),
        total_stock_end=sum(year_results[-1].stock.values()),
        total_inflows=total_in,
        total_outflows=total_out,
    )

    sid = sys_def.id or ""
    return (
        SimulationResult(system_id=sid, years=year_results, summary=summary),
        warnings,
        sorted(cohorts),
    )


class ImportResult(BaseModel):
    years_imported: int
    cohorts_found: int
    warnings: list[str] = []


@router.post("/systems/{system_id}/import-simulation", response_model=ImportResult)
async def import_simulation(system_id: str, file: UploadFile = File(...)) -> ImportResult:
    """Restore a simulation result from a previously exported .xlsx for an existing system."""
    sys_def = _get_system(system_id)
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file exported from MApper.")
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    result, warnings, cohorts = _reconstruct_simulation(sys_def, wb)
    result.system_id = system_id

    project = _current_project()
    with _lock:
        _proj_results(project)[system_id] = result
    mfa_storage.save_results(project, system_id, result)

    return ImportResult(
        years_imported=len(result.years),
        cohorts_found=len(cohorts),
        warnings=warnings,
    )


def _parse_summary_sheet(ws) -> dict[str, object]:
    """Extract key-value pairs from the Summary sheet."""
    kv: dict[str, object] = {}
    dims: list[DimensionDef] = []
    survival: list[SurvivalConfig] = []

    rows = list(ws.iter_rows(values_only=True))
    # key-value section is everything until row "Dimensions"
    i = 0
    while i < len(rows):
        r = rows[i]
        if not r:
            i += 1
            continue
        label = r[0]
        if label == "Dimensions":
            i += 1
            break
        if label is None or label == "":
            i += 1
            continue
        kv[str(label)] = r[1] if len(r) > 1 else None
        i += 1

    # dimensions table: rows with header [Name, Display name, Labels] then data rows
    # skip header row
    if i < len(rows):
        i += 1
    while i < len(rows):
        r = rows[i]
        if not r or r[0] is None or r[0] == "" or r[0] == "Survival configuration":
            break
        name = str(r[0])
        display = str(r[1]) if len(r) > 1 and r[1] is not None else name
        labels_cell = r[2] if len(r) > 2 else ""
        labels_str = str(labels_cell) if labels_cell is not None else ""
        is_age = labels_str.startswith("auto-generated")
        labels = [] if is_age else [l.strip() for l in labels_str.split(",") if l.strip()]
        dims.append(DimensionDef(name=name, display_name=display, labels=labels, is_age=is_age))
        i += 1

    # Skip to survival section header
    while i < len(rows) and (not rows[i] or rows[i][0] != "Survival configuration"):
        i += 1
    if i < len(rows):
        i += 2  # skip section title + header row
    while i < len(rows):
        r = rows[i]
        if not r or r[0] is None or r[0] == "":
            break
        filter_str = str(r[0])
        method = str(r[1]) if len(r) > 1 and r[1] else "weibull"
        shape = r[2] if len(r) > 2 else None
        scale = r[3] if len(r) > 3 else None
        filters: dict[str, str] = {}
        if filter_str and filter_str != "(default)":
            for part in filter_str.split(","):
                if "=" in part:
                    k, v = part.split("=", 1)
                    filters[k.strip()] = v.strip()
        survival.append(
            SurvivalConfig(
                dimension_filters=filters,
                method=method,
                weibull_shape=float(shape) if isinstance(shape, (int, float)) else None,
                weibull_scale=float(scale) if isinstance(scale, (int, float)) else None,
            )
        )
        i += 1

    kv["__dimensions__"] = dims
    kv["__survival__"] = survival
    return kv


@router.post("/import-system", response_model=SystemDefinition)
async def import_system(file: UploadFile = File(...)) -> SystemDefinition:
    """Create a brand new system (with state + results) from an exported workbook."""
    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".xlsx", ".xls"):
        raise HTTPException(status_code=400, detail="Upload an .xlsx file exported from MApper.")
    raw = await file.read()
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    if "Summary" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Workbook missing required Summary sheet.")
    meta = _parse_summary_sheet(wb["Summary"])

    name = str(meta.get("System name") or "Imported system")
    description = str(meta.get("Description") or "") or None
    try:
        start_year = int(meta.get("Start year"))  # type: ignore[arg-type]
        end_year = int(meta.get("End year"))  # type: ignore[arg-type]
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Summary sheet missing valid Start/End year.")

    dims: list[DimensionDef] = meta.get("__dimensions__") or []  # type: ignore[assignment]
    if not [d for d in dims if not d.is_age]:
        raise HTTPException(status_code=400, detail="No dimensions found in Summary sheet.")
    survival_configs: list[SurvivalConfig] = meta.get("__survival__") or []  # type: ignore[assignment]

    # Ensure age dim is present.
    if not any(d.is_age for d in dims):
        dims.append(DimensionDef(name="age", display_name="Age", labels=[], is_age=True))

    new_def = SystemDefinition(
        id=None,
        name=name,
        description=description,
        time_horizon=TimeHorizon(start_year=start_year, end_year=end_year),
        dimensions=dims,
    )
    _validate_definition(new_def)

    project = _current_project()
    with _lock:
        sid = str(uuid.uuid4())
        new_def.id = sid
        new_def.created_at = _now_iso()
        _proj_systems(project)[sid] = new_def
        state = MFASystemState(system_id=sid, survival_configs=survival_configs)
        _proj_states(project)[sid] = state

    # Reconstruct results using the workbook
    result, warnings, _cohorts = _reconstruct_simulation(new_def, wb)
    result.system_id = sid

    # Rebuild initial_stock from year=start_year age distribution
    start_yr_result = next((y for y in result.years if y.year == start_year), None)
    if start_yr_result is not None:
        initial: dict[str, float] = {}
        for ck, by_age in start_yr_result.stock_by_age.items():
            for age, cnt in by_age.items():
                initial[f"{ck}|{age}"] = cnt
        state.initial_stock = initial

    # Rebuild inflows
    inflows: list[InflowData] = []
    for yr in result.years:
        if yr.inflow:
            inflows.append(InflowData(year=yr.year, counts=dict(yr.inflow)))
    state.inflows = inflows

    with _lock:
        _proj_states(project)[sid] = state
        _proj_results(project)[sid] = result

    mfa_storage.save_system(project, new_def)
    mfa_storage.save_state(project, sid, state)
    mfa_storage.save_results(project, sid, result)

    return new_def


# ── CSV templates ───────────────────────────────────────────────────────────


@router.post("/systems/{system_id}/templates/stock")
async def template_stock(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    csv_text = stock_template_csv(sys_def.dimensions)
    fname = f"stock_template_{_sanitize_filename(sys_def.name)}.csv"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/systems/{system_id}/templates/inflows")
async def template_inflows(system_id: str) -> Response:
    sys_def = _get_system(system_id)
    csv_text = inflow_template_csv(sys_def.dimensions, sys_def.time_horizon.years)
    fname = f"inflow_template_{_sanitize_filename(sys_def.name)}.csv"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
