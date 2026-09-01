# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""``POST /lca/monte-carlo`` -- single-product uncertainty propagation.

Task-registry + WebSocket + cancellation, the pattern ``plca`` and ``impact``
use. Cancellation is checked EVERY iteration: at ~0.066 s each, Stop lands
within one iteration and feels instant.

The methodology -- what is sampled, the per-iteration order of operations, and
why the iterative solver is the right one -- lives in
``mapper.core.monte_carlo_engine``. Read that before changing this file.
"""

from __future__ import annotations

import asyncio
import threading
import time
import uuid
from typing import Any

import numpy as np
from fastapi import APIRouter, HTTPException, WebSocket, WebSocketDisconnect

from mapper.api import tasks as task_registry
from mapper.api.tasks import CancelledOperation
from mapper.core import monte_carlo_engine as mce
from mapper.core.monte_carlo_engine import (
    UncertaintyConfigError,
    collect_param_draws,
    collect_row_draws,
    lognormal_factor,
    summarize,
    variance_shares,
)
from mapper.models.bom_schemas import MaterialPedigreeLibrary
from mapper.models.schemas import (
    ArchetypeLCAMethodDistribution,
    ItemDistribution,
    MonteCarloExportRequest,
    MonteCarloMultiExportRequest,
    MonteCarloMultiRequest,
    MonteCarloMultiResult,
    MonteCarloRequest,
    MonteCarloResult,
    MaterialScoringScope,
    MonteCarloStartResponse,
    PairwiseDifference,
    PedigreeCoverage,
    ScoredInput,
    PedigreeTableResponse,
    UnscoredMaterial,
    VarianceContributor,
)

router = APIRouter()

MAX_ITERATIONS = 20_000


class _TaskState:
    def __init__(self) -> None:
        self.stage: str = "queued"
        self.pct: float = 0.0
        self.done: bool = False
        self.error: str | None = None
        self.cancelled: bool = False
        self.result: MonteCarloResult | None = None
        self.multi_result: "MonteCarloMultiResult | None" = None
        self.subscribers: list[asyncio.Queue] = []


_TASKS: dict[str, _TaskState] = {}
_TASK_LOCK = threading.Lock()


def _notify_all(task: _TaskState, payload: dict[str, Any]) -> None:
    for q in list(task.subscribers):
        try:
            q.put_nowait(payload)
        except asyncio.QueueFull:
            pass


def _emit(task: _TaskState, stage: str, pct: float) -> None:
    task.stage = stage
    task.pct = pct
    _notify_all(task, {"type": "progress", "stage": stage, "pct": pct})


# ── The run ───────────────────────────────────────────────────────────────────


def _run_monte_carlo(
    body: MonteCarloRequest,
    task: _TaskState,
    task_id: str,
) -> MonteCarloResult:
    """Draw ``body.iterations`` samples and summarise them per indicator."""
    import bw2calc

    from mapper.api.lca import _build_archetype_source_demand, _translate_demand_to_database
    from mapper.api.parameters import _table_for
    from mapper.core.bom_engine import (
        filter_roots_by_scope,
        flatten_root_with_amounts,
        resolve_archetype_with_engine,
    )
    from mapper.core.bw2_wrapper import PersistentLCARunner
    from mapper.core.parameter_engine import ParameterEngine

    t_start = time.perf_counter()
    _emit(task, "preparing", 0.01)

    # The deterministic run, reusing the ordinary single-product builder. This
    # both validates the request (same 4xx as /lca/calculate-archetype) and
    # gives the point score the distribution is shown against.
    bundle = _build_archetype_source_demand(
        archetype_id=body.archetype_id,
        scope=body.scope,
        amount=body.amount,
        stage_amounts=body.stage_amounts,
        methods=body.methods,
        parameter_scenario=body.parameter_scenario,
        basis_amounts=body.basis_amounts,
    )
    method_tuples = bundle.method_tuples
    base_demand, warnings = _translate_demand_to_database(
        bundle.total_demand, body.compute_database
    )
    if not base_demand:
        raise HTTPException(
            status_code=400,
            detail="No linked materials resolved against the selected database.",
        )

    runner = PersistentLCARunner()
    det_scores = runner(base_demand, method_tuples)

    # What actually carries uncertainty. `collect_row_draws` is where the
    # expression-row rule is enforced.
    table = _table_for()
    referenced = _referenced_parameters(bundle.arc)
    from mapper.core.material_pedigree_storage import load_library

    library = load_library(_current_project())
    try:
        row_draws = collect_row_draws(bundle.linked, library.entries)
        param_draws = collect_param_draws(table, referenced)
    except UncertaintyConfigError as e:
        raise HTTPException(status_code=400, detail=str(e))

    seed = body.seed if body.seed is not None else int(uuid.uuid4().int % (2**31 - 1))
    rng = np.random.default_rng(seed)
    n = body.iterations

    # One MonteCarloLCA per method would resample the background independently
    # per indicator, which would decorrelate indicators that share an
    # inventory. One chain, characterised against every method, keeps them on
    # the same draw -- and costs ~0.2 ms per extra indicator.
    mc = bw2calc.MonteCarloLCA(base_demand, method_tuples[0], seed=seed)
    next(mc)  # loads data, builds the RNGs
    cf_samplers = _method_cf_samplers(mc, method_tuples, seed)

    samples: dict[tuple, list[float]] = {m: [] for m in method_tuples}
    # Keyed by node_id for rows -- the SAME key the draws use -- because two
    # rows can share a NAME. Keyed by name it collided: one list was created
    # for the pair, both rows appended to it every iteration, it reached 2n
    # entries, and the `len(v) == n` filter below dropped it. `variance_shares`
    # then renormalised, so the table still summed to 100 % with a contributor
    # silently missing. MAp-test's `Fuel Station` repeats six names.
    #
    # Parameters stay keyed by name: `table.parameters` is a dict, so a
    # parameter name is unique by construction.
    input_draws: dict[str, list[float]] = {}
    display_by_key: dict[str, tuple[str, str]] = {}   # key -> (kind, display name)
    sigma_by_key: dict[str, float] = {}
    if body.variance_contributions:
        # `setdefault` de-duplicates while preserving first-seen order. A
        # duplicate node_id is degenerate (`factors` below is a dict, so the
        # two rows would share one factor anyway) but must not produce two
        # appends per iteration -- that overflow is what this patch removes.
        for r in row_draws:
            k = f"row::{r.node_id}"
            input_draws.setdefault(k, [])
            display_by_key.setdefault(k, ("row", r.name))
            sigma_by_key.setdefault(k, r.sigma)
        for p in param_draws:
            k = f"param::{p.name}"
            input_draws.setdefault(k, [])
            display_by_key.setdefault(k, ("parameter", p.name))
            sigma_by_key.setdefault(k, p.sigma)

    base_materials = _linked_with_amounts(
        bundle.arc, body.scope, bundle.effective_amounts, body.basis_amounts
    )
    base_values = table.resolve_all(body.parameter_scenario, 2025)

    # Source-db key -> the key actually present in the technosphere being
    # solved against. Derived ONCE: an activity link never changes across
    # iterations (only its quantity does), and `_translate_demand_to_database`
    # does bw2data lookups, so re-deriving it per iteration would cost more
    # than the solve it feeds.
    key_map = _translation_map(
        {(m.ecoinvent_activity.database, m.ecoinvent_activity.code) for m, _ in base_materials},
        body.compute_database,
    )

    # Nothing in the foreground varies -> the demand is constant, so skip
    # rebuilding it entirely and let the run vary the background alone. This is
    # a legitimate configuration (and the default before any row or parameter
    # is tagged), not a degenerate one.
    static_demand = None if (row_draws or param_draws) else base_demand

    for i in range(n):
        if task_registry.is_cancelled(task_id):
            raise CancelledOperation(task_id)

        # ── 1. draw each PARAMETER once ──────────────────────────────────────
        if param_draws:
            drawn = dict(base_values)
            for p in param_draws:
                f = lognormal_factor(rng, p.sigma)
                drawn[p.name] = base_values.get(p.name, p.base_value) * f
                if body.variance_contributions:
                    input_draws[f"param::{p.name}"].append(f)
            # ── 2. re-resolve every expression against those draws ───────────
            arc_i = resolve_archetype_with_engine(bundle.arc, ParameterEngine(drawn))
            materials_i = _linked_with_amounts(
                arc_i, body.scope, bundle.effective_amounts, body.basis_amounts
            )
        else:
            materials_i = base_materials

        # ── 3. apply per-row uncertainty to LITERAL rows only ────────────────
        # `collect_row_draws` has already rejected any expression row carrying
        # its own uncertainty, so nothing here can double-count a driver. The
        # factors are keyed by node_id and applied DURING demand construction,
        # because several rows may share one ecoinvent code and scaling the
        # aggregated entry would leak a row's factor onto its neighbours.
        factors: dict[str, float] = {}
        for r in row_draws:
            f = lognormal_factor(rng, r.sigma)
            factors[r.node_id] = f
        if body.variance_contributions:
            # Recorded FROM `factors`, so the accumulator cannot disagree with
            # the sampling: exactly one append per key per iteration, whatever
            # `factors` ended up holding. The draw order above is untouched, so
            # the RNG stream -- and every score -- is unchanged.
            for nid, f in factors.items():
                input_draws[f"row::{nid}"].append(f)

        demand_i = static_demand if static_demand is not None else _aggregate(
            materials_i, factors, key_map
        )

        mc.demand = demand_i
        mc.build_demand_array()
        next(mc)                      # resamples A, B and C, then solves
        # Per-flow totals as a matvec. Equivalent to summing the rows of the
        # materialised inventory (B * diag(supply)), but without touching the
        # 4.7k x 23k sparse product a second time.
        flow_totals = mc.biosphere_matrix * mc.supply_array
        for m in method_tuples:
            rows, cf_rng = cf_samplers[m]
            samples[m].append(float(cf_rng.next() @ flow_totals[rows]))

        if (i + 1) % 10 == 0 or i + 1 == n:
            _emit(task, f"iteration {i + 1}/{n}", 0.02 + 0.96 * (i + 1) / n)

    if task_registry.is_cancelled(task_id):
        raise CancelledOperation(task_id)
    _emit(task, "summarising", 0.99)

    distributions: list[ArchetypeLCAMethodDistribution] = []
    for m in method_tuples:
        s = summarize(samples[m])
        det, unit = det_scores.get(m, (0.0, ""))
        distributions.append(
            ArchetypeLCAMethodDistribution(
                method=list(m),
                method_label=" | ".join(m[1:]) or m[0],
                unit=unit,
                deterministic=det,
                n_iterations=n,
                seed=seed,
                samples=samples[m] if body.keep_samples else None,
                **s,
            )
        )

    contributors: list[VarianceContributor] = []
    if body.variance_contributions and input_draws:
        primary = samples[method_tuples[0]]
        # Every key now carries exactly n entries by construction, so this
        # filter is a belt-and-braces guard rather than the load-bearing thing
        # it accidentally became when names collided.
        arrays = {k: np.asarray(v) for k, v in input_draws.items() if len(v) == n}
        for key, share in variance_shares(arrays, primary):
            kind, name = display_by_key.get(key, ("row", key.partition("::")[2]))
            contributors.append(
                VarianceContributor(
                    name=name,
                    kind=kind,
                    share=share,
                    gsd2=mce.gsd2_from_sigma(sigma_by_key.get(key, 0.0)),
                )
            )

    scored: list[ScoredInput] = []
    for r in row_draws:
        unc = _uncertainty_for_row(bundle.linked, r.node_id, library.entries)
        scored.append(ScoredInput(
            name=r.name, kind="row",
            pedigree=dict((unc.pedigree or {}) if unc else {}),
            basic_variance=(unc.basic_variance if unc else 0.0),
            gsd2=mce.gsd2_from_sigma(r.sigma), inherited=r.inherited,
        ))
    for p in param_draws:
        pu = getattr(table.parameters.get(p.name), "uncertainty", None)
        scored.append(ScoredInput(
            name=p.name, kind="parameter",
            pedigree=dict((pu.pedigree or {}) if pu else {}),
            basic_variance=(pu.basic_variance if pu else 0.0),
            gsd2=mce.gsd2_from_sigma(p.sigma),
        ))

    return MonteCarloResult(
        archetype_id=body.archetype_id,
        archetype_name=bundle.arc.name,
        scope=body.scope,
        n_iterations=n,
        seed=seed,
        elapsed_seconds=round(time.perf_counter() - t_start, 3),
        compute_database=body.compute_database,
        parameter_scenario=body.parameter_scenario,
        distributions=distributions,
        contributors=contributors,
        rows_with_uncertainty=len(row_draws),
        rows_inherited=sum(1 for r in row_draws if r.inherited),
        parameters_with_uncertainty=len(param_draws),
        scored_inputs=scored,
        warnings=warnings,
    )


# ── Helpers ───────────────────────────────────────────────────────────────────


def _uncertainty_for_row(materials: Any, node_id: str, library: dict) -> Any:
    """The RowUncertainty a given node actually used: its own, else the
    material-library entry it inherited."""
    for m in materials:
        if getattr(m, "node_id", None) != node_id:
            continue
        return m.uncertainty or library.get(m.name)
    return None


def _referenced_parameters(arc: Any) -> set[str]:
    """Every parameter name appearing in any of the archetype's expressions."""
    import re

    names: set[str] = set()
    ident = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")

    def walk(n: Any) -> None:
        expr = getattr(n, "quantity_expression", None)
        if expr:
            names.update(ident.findall(expr))
        for lever in (getattr(n, "global_levers", None) or []):
            names.add(lever)
        for c in (getattr(n, "children", None) or []):
            walk(c)

    for root in arc.bom:
        walk(root)
    return names


def _linked_with_amounts(
    arc: Any,
    scope: str,
    effective_amounts: dict[str, float],
    basis_amounts: dict[str, float] | None,
) -> list[tuple[Any, float]]:
    """Flatten to (material, stage_amount) pairs, in scope.

    Mirrors the tail of ``_build_archetype_source_demand`` -- the same
    ``flatten_root_with_amounts`` pairing -- so a per-iteration demand is built
    exactly the way the deterministic one is. Kept as pairs rather than an
    aggregated dict so a per-row factor can be applied before aggregation.
    """
    from mapper.core.bom_engine import filter_roots_by_scope, flatten_root_with_amounts

    out: list[tuple[Any, float]] = []
    for root in filter_roots_by_scope(arc.bom, scope):
        for m, amt in flatten_root_with_amounts(root, effective_amounts, basis_amounts):
            if m.ecoinvent_activity is not None:
                out.append((m, amt))
    return out


def _aggregate(
    materials: list[tuple[Any, float]],
    factors: dict[str, float],
    key_map: dict[tuple[str, str], tuple[str, str] | None],
) -> dict[tuple[str, str], float]:
    """Sum materials into a demand, applying per-node factors BEFORE the sum.

    Before, not after: several rows can link one ecoinvent code, and scaling
    the aggregated entry would apply one row's draw to every other row sharing
    it. ``key_map`` re-keys to the database actually being solved against; a
    key that maps to ``None`` is absent there and is dropped, matching
    ``_translate_demand_to_database``.
    """
    demand: dict[tuple[str, str], float] = {}
    for m, amt in materials:
        link = m.ecoinvent_activity
        target = key_map.get((link.database, link.code))
        if target is None:
            continue
        q = m.quantity * amt * factors.get(m.node_id, 1.0)
        demand[target] = demand.get(target, 0.0) + q
    return demand


def _translation_map(
    keys: set[tuple[str, str]],
    compute_database: str | None,
) -> dict[tuple[str, str], tuple[str, str] | None]:
    """Map each source-db key to its key in ``compute_database`` (or itself)."""
    from mapper.api.lca import _translate_demand_to_database

    if not compute_database:
        return {k: k for k in keys}
    out: dict[tuple[str, str], tuple[str, str] | None] = {}
    for k in keys:
        translated, _ = _translate_demand_to_database({k: 1.0}, compute_database)
        out[k] = next(iter(translated), None) if translated else None
    return out


def _method_cf_samplers(mc: Any, method_tuples: list[tuple], seed: int) -> dict:
    """One characterisation-factor RNG per method, all on the same inventory.

    Two traps here, both hit on the way in.

    ``load_lcia_data()`` reads ``self.method_filepath``, NOT ``self.method`` --
    so assigning ``mc.method = m`` and reloading silently returns the ORIGINAL
    method's factors. That produced sixteen identical distributions on the
    first end-to-end run. ``switch_method`` is the call that recomputes the
    filepath, so it is the one to use.

    And ``switch_method`` replaces ``cf_params`` while the chain's ``cf_rng``
    still holds the previous method's array, which makes the next draw raise
    "Incompatible data & indices". The chain is therefore switched back to its
    original method and its ``cf_rng`` rebuilt before any iteration runs.

    Characterisation is diagonal, so a score is a dot product of the sampled
    factors with the per-flow inventory totals -- no need to rebuild a sparse
    matrix per method per iteration. Every method gets its OWN sampled factors
    this way, rather than the chain's one method being sampled and the other
    fifteen held fixed.
    """
    from stats_arrays.random import MCRandomNumberGenerator

    per: dict = {}
    original = mc.method
    for m in method_tuples:
        mc.switch_method(m)
        per[m] = (mc.cf_params["row"].copy(), MCRandomNumberGenerator(mc.cf_params, seed=seed))
    mc.switch_method(original)
    mc.cf_rng = MCRandomNumberGenerator(mc.cf_params, seed=seed)
    return per


# ── Endpoints ─────────────────────────────────────────────────────────────────


@router.get("/lca/pedigree", response_model=PedigreeTableResponse)
async def get_pedigree_table() -> PedigreeTableResponse:
    """Serve the pedigree constants to the UI.

    The scoring UI needs the factors to show a live GSD^2 as a user picks
    scores. Serving them keeps ONE table: a hard-coded copy in the frontend
    would drift silently, since both copies would keep producing plausible
    numbers.
    """
    from mapper.core.pedigree import (
        DEFAULT_BASIC_VARIANCE,
        INDICATORS,
        UNCERTAINTY_FACTORS,
    )

    return PedigreeTableResponse(
        indicators=list(INDICATORS),
        factors={k: list(v) for k, v in UNCERTAINTY_FACTORS.items()},
        default_basic_variance=DEFAULT_BASIC_VARIANCE,
        convention="sigma_i^2 = [ln(f_i) / 2]^2 — the factor is a 95% range, hence the /2.",
    )




# ── Paired multi-item ─────────────────────────────────────────────────────────


def _run_monte_carlo_multi(
    body: MonteCarloMultiRequest,
    task: _TaskState,
    task_id: str,
) -> MonteCarloMultiResult:
    """One sampled world per iteration, every item solved against it.

    PAIRED IS THE ONLY MODE. Independent draws let a shared driver take two
    different values in the same iteration, so a difference that is
    structurally near-certain reads as noise: measured on Battery Circularity,
    A and A0 share 23 parameters and all 27 ecoinvent activities, and
    independent sampling widened sd(A-A0) by 6.8x and pushed the 95% interval
    across zero -- reporting "not distinguishable" where the paired run says A
    is lower in 100% of iterations. It is also CHEAPER, because sampling the
    matrices is the expensive part and this does it once per iteration rather
    than once per item.

    The marginals are unaffected, and that is asserted rather than assumed:
    the RNG sequence depends only on the seed, not on which demand is solved
    against it, so item i's per-iteration scores match a single-item run at the
    same seed. See ``test_paired_marginals_match_single_item``.
    """
    import bw2calc

    from mapper.api.lca import _build_archetype_source_demand, _translate_demand_to_database
    from mapper.api.parameters import _table_for
    from mapper.core.bom_engine import resolve_archetype_with_engine
    from mapper.core.bw2_wrapper import PersistentLCARunner
    from mapper.core.material_pedigree_storage import load_library
    from mapper.core.parameter_engine import ParameterEngine

    t_start = time.perf_counter()
    _emit(task, "preparing", 0.01)

    if not body.archetype_ids:
        raise HTTPException(status_code=400, detail="At least one item is required")
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required")

    method_tuples = [tuple(m) for m in body.methods]
    warnings: list[str] = []
    items: list[dict] = []
    runner = PersistentLCARunner()
    table = _table_for()
    library = load_library(_current_project())
    referenced: set[str] = set()

    for aid in body.archetype_ids:
        bundle = _build_archetype_source_demand(
            archetype_id=aid, scope=body.scope, amount=1.0,
            stage_amounts=body.stage_amounts.get(aid, {}),
            methods=body.methods, parameter_scenario=body.parameter_scenario,
            basis_amounts=body.basis_amounts,
        )
        demand, w = _translate_demand_to_database(bundle.total_demand, body.compute_database)
        warnings.extend(w)
        if not demand:
            raise HTTPException(
                status_code=400,
                detail=f"'{bundle.arc.name}' has no linked materials in this scope.",
            )
        # Everything the per-iteration demand rebuild needs, gathered ONCE.
        # ``key_map`` in particular: an activity LINK never changes across
        # iterations (only its quantity does) and building it does bw2data
        # lookups, so per-iteration would cost more than the solve it feeds.
        base_materials = _linked_with_amounts(
            bundle.arc, body.scope, bundle.effective_amounts, body.basis_amounts
        )
        try:
            row_draws = collect_row_draws(bundle.linked, library.entries)
        except UncertaintyConfigError as e:
            raise HTTPException(status_code=400, detail=str(e))
        items.append({
            "id": aid, "name": bundle.arc.name, "demand": demand,
            "det": runner(demand, method_tuples),
            "samples": {m: [] for m in method_tuples},
            "arc": bundle.arc,
            "effective_amounts": bundle.effective_amounts,
            "base_materials": base_materials,
            "row_draws": row_draws,
            "key_map": _translation_map(
                {
                    (m.ecoinvent_activity.database, m.ecoinvent_activity.code)
                    for m, _ in base_materials
                },
                body.compute_database,
            ),
        })
        referenced |= _referenced_parameters(bundle.arc)

    # ── The shared foreground draw ──────────────────────────────────────────
    #
    # Parameters are collected across the UNION of every item's referenced
    # names and drawn ONCE per iteration, then applied to every item. That is
    # what makes the foreground half of the pairing work: A and A0 share 23
    # parameters, and drawing them per item would let one shared driver take
    # two values in the same iteration -- exactly the decorrelation the paired
    # mode exists to remove, reintroduced one layer down.
    #
    # Row draws stay PER ITEM (keyed by node_id) because a row belongs to one
    # archetype; two archetypes sharing a material NAME are still two rows, as
    # in the single-item path.
    try:
        param_draws = collect_param_draws(table, referenced)
    except UncertaintyConfigError as e:
        raise HTTPException(status_code=400, detail=str(e))

    seed = body.seed if body.seed is not None else int(uuid.uuid4().int % (2**31 - 1))
    n = body.iterations
    rng = np.random.default_rng(seed)
    base_values = table.resolve_all(body.parameter_scenario, 2025)

    # Nothing in any foreground varies -> every demand is constant, so skip the
    # rebuild entirely and let the run vary the background alone. Same
    # short-circuit as the single-item path, and the default before anything is
    # scored.
    foreground_varies = bool(param_draws) or any(it["row_draws"] for it in items)

    mc = bw2calc.MonteCarloLCA(items[0]["demand"], method_tuples[0], seed=seed)
    next(mc)
    cf_samplers = _method_cf_samplers(mc, method_tuples, seed)

    for i in range(n):
        # Cancellation stops the WHOLE job, not one item: the check is at the
        # iteration boundary, before any item is solved, so a stop never leaves
        # a half-populated world where some items have i draws and others i-1.
        if task_registry.is_cancelled(task_id):
            raise CancelledOperation(task_id)

        # ── 1. draw each PARAMETER once, for the WHOLE iteration ────────
        # Before any item is touched, so every item resolves against the same
        # world. Drawing inside the item loop would give one shared driver two
        # values per iteration.
        demands_i: list[dict] | None = None
        if foreground_varies:
            drawn = dict(base_values)
            for pdraw in param_draws:
                f = lognormal_factor(rng, pdraw.sigma)
                drawn[pdraw.name] = base_values.get(pdraw.name, pdraw.base_value) * f
            engine_i = ParameterEngine(drawn) if param_draws else None

            demands_i = []
            for it in items:
                # ── 2. re-resolve expressions against those draws ───────────
                if engine_i is not None:
                    arc_i = resolve_archetype_with_engine(it["arc"], engine_i)
                    materials_i = _linked_with_amounts(
                        arc_i, body.scope, it["effective_amounts"], body.basis_amounts
                    )
                else:
                    materials_i = it["base_materials"]
                # ── 3. per-row factors, LITERAL rows only, keyed by node_id ──
                factors = {
                    r.node_id: lognormal_factor(rng, r.sigma) for r in it["row_draws"]
                }
                demands_i.append(_aggregate(materials_i, factors, it["key_map"]))

        # ── 4. advance the sampled world ONCE, then solve every item ────────
        #
        # ``next(mc)`` is what resamples A, B and C -- it must fire exactly
        # once per iteration or the items stop sharing a world, which is the
        # decorrelation this mode exists to remove. So item 0 is NOT special:
        # its demand is pushed in BEFORE the advance (the chain solves for
        # whatever demand is loaded when ``next`` runs), and every later item
        # re-solves the same matrices via ``lci_calculation``.
        if demands_i is not None:
            mc.demand = demands_i[0]
            mc.build_demand_array()
        next(mc)                       # ONE sampled world for this iteration
        for idx, it in enumerate(items):
            if idx > 0:
                # Same sampled matrices, next demand.
                mc.demand = demands_i[idx] if demands_i is not None else it["demand"]
                mc.build_demand_array()
                mc.lci_calculation()
            flow_totals = mc.biosphere_matrix * mc.supply_array
            # Reset the per-iteration CF cache ONCE, before the method loop.
            # Resetting inside it left only the LAST method's draw, so item 2
            # raised KeyError(<first method tuple>) -- the reported crash.
            if idx == 0:
                items[0]["_cf"] = {}
            for m in method_tuples:
                rows, cf_rng = cf_samplers[m]
                # The CF draw belongs to the ITERATION, not the item: every
                # item must be characterised with the same sampled factors or
                # the difference reacquires the decorrelation pairing removes.
                if idx == 0:
                    vals = cf_rng.next()
                    items[0]["_cf"][m] = vals
                else:
                    vals = items[0]["_cf"][m]
                it["samples"][m].append(float(vals @ flow_totals[rows]))
        # Restore item 0's demand so the next iteration's warm start is stable.
        # (Superseded at the top of the next iteration when the foreground
        # varies -- kept so the constant-foreground path is unchanged.)
        if len(items) > 1:
            mc.demand = items[0]["demand"]
            mc.build_demand_array()

        if (i + 1) % 10 == 0 or i + 1 == n:
            _emit(task, f"iteration {i + 1}/{n}", 0.02 + 0.96 * (i + 1) / n)

    if task_registry.is_cancelled(task_id):
        raise CancelledOperation(task_id)
    _emit(task, "summarising", 0.99)

    out_items: list[ItemDistribution] = []
    for it in items:
        dists = []
        for m in method_tuples:
            st = summarize(it["samples"][m])
            det, unit = it["det"].get(m, (0.0, ""))
            dists.append(ArchetypeLCAMethodDistribution(
                method=list(m), method_label=" | ".join(m[1:]) or m[0], unit=unit,
                deterministic=det, n_iterations=n, seed=seed,
                samples=it["samples"][m] if body.keep_samples else None, **st,
            ))
        out_items.append(ItemDistribution(
            archetype_id=it["id"], archetype_name=it["name"], distributions=dists))

    diffs = _pairwise_differences(items, method_tuples)

    return MonteCarloMultiResult(
        scope=body.scope, n_iterations=n, seed=seed,
        elapsed_seconds=round(time.perf_counter() - t_start, 3),
        compute_database=body.compute_database,
        parameter_scenario=body.parameter_scenario,
        items=out_items, differences=diffs, warnings=warnings,
    )


def _pairwise_differences(items: list[dict], method_tuples: list[tuple]) -> list[PairwiseDifference]:
    """Every ordered pair in comparison order, per indicator."""
    out: list[PairwiseDifference] = []
    for i in range(len(items)):
        for j in range(i + 1, len(items)):
            a, b = items[i], items[j]
            for m in method_tuples:
                sa = np.asarray(a["samples"][m])
                sb = np.asarray(b["samples"][m])
                d = sa - sb
                q = np.percentile(d, [2.5, 25, 50, 75, 97.5])
                corr = 0.0
                if sa.size > 1 and sa.std() > 0 and sb.std() > 0:
                    corr = float(np.corrcoef(sa, sb)[0, 1])
                det_a = a["det"].get(m, (0.0, ""))
                det_b = b["det"].get(m, (0.0, ""))
                out.append(PairwiseDifference(
                    method=list(m), method_label=" | ".join(m[1:]) or m[0],
                    unit=det_a[1],
                    a_id=a["id"], a_name=a["name"], b_id=b["id"], b_name=b["name"],
                    deterministic=det_a[0] - det_b[0],
                    median=float(q[2]), mean=float(d.mean()),
                    p2_5=float(q[0]), p25=float(q[1]), p75=float(q[3]), p97_5=float(q[4]),
                    fraction_a_lower=float(np.mean(sa < sb)),
                    correlation=corr,
                ))
    return out


# ── Excel export ──────────────────────────────────────────────────────────────

#: Field acronym for the shared filename scheme, alongside LCA / pLCA / AESA /
#: DSM / MFA. Single-product Monte Carlo has no DSM system, so the archetype
#: name takes the ``system`` slot and there are never subsystems.
MC_DOMAIN = "MC"

#: Repeated in the Summary because a distribution read without it is
#: over-confident: roughly 12% of ecoinvent's non-production exchanges carry no
#: uncertainty distribution and are sampled as fixed.
LOWER_BOUND_NOTE = (
    "LOWER BOUND. About 12% of ecoinvent's technosphere exchanges carry no "
    "uncertainty distribution and are sampled as FIXED, so their contribution "
    "to the spread is missing from every figure in this workbook."
)


def _scope_label(scope: str) -> str:
    """Reuse Impact Assessment's labels so one scope never reads two ways."""
    from mapper.api.impact import _SCOPE_LABELS

    return _SCOPE_LABELS.get(scope, "Full Lifecycle" if scope == "all" else scope)


# ── Dispersion labels, stated wherever a number appears ──────────────────────
# This workbook is a reproducibility record, and a bare "GSD2" header is
# exactly what let two different statistics travel under one name: the input
# side used exp(2*sigma) while the Monte Carlo output used exp(1.96*sigma_hat).
# Both now mean exp(2*sigma), and every sheet that prints one says so.
_GSD2_NOTE = (
    "GSD2 = exp(2*sigma), the squared geometric standard deviation. This is "
    "ecoinvent's convention and the same definition used for scored inputs, so "
    "an input GSD2 and an output GSD2 are the same statistic. The 95% interval "
    "spans APPROXIMATELY median / and x this (2 standing in for 1.96)."
)
_DISPERSION_NOTE = (
    "95% dispersion factor = p97.5 / median, read straight off the percentiles. "
    "It assumes nothing about the distribution's shape, which matters because a "
    "sum of lognormals is not lognormal; it is the exact figure where GSD2 is "
    "the lognormal approximation."
)
_MIGRATION_NOTE = (
    "MIGRATION: workbooks exported before 2026-08-31 reported exp(1.96*sigma) "
    "in a column headed GSD2. To convert such a figure without re-running, "
    "raise it to the power 2/1.96: GSD2 = reported ** 1.020408. The correction "
    "is +0.2% at 1.10, +0.6% at 1.37, +1.4% at 2.00, and is always upward."
)


def _build_monte_carlo_workbook(
    result: MonteCarloResult,
    coverage: PedigreeCoverage | None,
) -> "Workbook":
    from openpyxl import Workbook

    from mapper.api.cohort_export import apply_sci, autosize, style_header

    wb = Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    style_header(ws)

    rows: list[tuple[str, object]] = [
        ("Archetype", result.archetype_name),
        ("Sensitivity case", result.parameter_scenario or "Base"),
        ("Scope", _scope_label(result.scope)),
        ("Background database", result.compute_database or "base ecoinvent"),
        ("Iterations", result.n_iterations),
        # Not optional. A Monte Carlo result nobody can reproduce is not a
        # research output, and the seed is the whole of the reproduction.
        ("Seed", result.seed),
        ("Elapsed (s)", round(result.elapsed_seconds, 2)),
        ("", ""),
        ("SCORING PROVENANCE", ""),
        ("Rows scored", result.rows_with_uncertainty),
        ("  ...inheriting a material score", result.rows_inherited),
        ("Parameters scored", result.parameters_with_uncertainty),
    ]
    if coverage is not None:
        rows += [
            ("Materials scored (project)",
             f"{coverage.materials_scored} of {coverage.materials_total}"),
            ("Materials scored (this archetype)",
             f"{coverage.archetype_materials_scored} of {coverage.archetype_materials_total}"),
            ("Impact-weighted coverage",
             f"{coverage.impact_share * 100:.1f}% of {coverage.method_label}"),
            ("Coverage basis",
             "Weighted by impact, not row count: the share of THIS archetype's "
             "total |impact| carried by rows whose uncertainty was scored."),
        ]
    else:
        rows.append(
            ("Impact-weighted coverage",
             "not recorded for this export — re-export with coverage available")
        )
    if result.rows_with_uncertainty == 0 and result.parameters_with_uncertainty == 0:
        rows.append(
            ("Foreground uncertainty",
             "NONE scored. This run varied the BACKGROUND only, so the spread "
             "reflects ecoinvent's own exchange uncertainty and nothing about "
             "the foreground data.")
        )
    rows += [("", ""), ("Caveat", LOWER_BOUND_NOTE)]
    if result.warnings:
        rows.append(("Warnings", " | ".join(result.warnings)))

    for k, v in rows:
        ws.append([k, v])
    autosize(ws)

    # ── Distributions ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Distributions")
    ws.append([
        "Indicator", "Unit", "Deterministic", "Median", "Median / deterministic",
        "Mean", "p2.5", "p25", "p75", "p97.5",
        "GSD2 = exp(2*sigma)", "95% dispersion factor = p97.5 / median",
        "Iterations", "Seed",
    ])
    style_header(ws)
    for d in result.distributions:
        ratio = (d.median / d.deterministic) if d.deterministic else None
        ws.append([
            d.method_label, d.unit, d.deterministic, d.median,
            round(ratio, 4) if ratio is not None else "n/a",
            d.mean, d.p2_5, d.p25, d.p75, d.p97_5,
            round(d.gsd2, 4) if d.gsd2 else "n/a",
            round(d.dispersion_95, 4) if d.dispersion_95 else "n/a",
            d.n_iterations, d.seed,
        ])
    if result.distributions:
        # Scores span orders of magnitude across indicators; the ratio and the
        # two dispersion columns are small and stay readable as plain numbers.
        apply_sci(ws, min_row=2, min_col=3, max_col=4)
        apply_sci(ws, min_row=2, min_col=6, max_col=10)
    ws.append([])
    ws.append([_GSD2_NOTE])
    ws.append([_DISPERSION_NOTE])
    ws.append([_MIGRATION_NOTE])
    autosize(ws)

    # ── Variance contribution ─────────────────────────────────────────────────
    ws = wb.create_sheet("Variance contribution")
    ws.append(["Input", "Kind", "Share of spread", "GSD2 = exp(2*sigma)"])
    style_header(ws)
    for c in result.contributors:
        ws.append([c.name, c.kind, round(c.share, 6), round(c.gsd2, 4)])
    if not result.contributors:
        ws.append([
            "No foreground input carried uncertainty, so the whole spread comes "
            "from the background database.", "", "", "",
        ])
    else:
        ws.append([])
        ws.append([
            "Shares are an approximate attribution (squared rank correlation, "
            "normalised) — the inputs are not orthogonal.", "", "", "",
        ])
    ws.append([])
    ws.append([_GSD2_NOTE])
    autosize(ws)

    # ── Pedigree scores ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Pedigree scores")
    ws.append([
        "Input", "Kind", "Source",
        "Reliability", "Completeness", "Temporal", "Geographical",
        "Technological", "Basic variance", "GSD2 = exp(2*sigma)",
    ])
    style_header(ws)
    for si in result.scored_inputs:
        ped = si.pedigree or {}
        ws.append([
            si.name, si.kind,
            "material library" if si.inherited else "own",
            ped.get("reliability", ""),
            ped.get("completeness", ""),
            ped.get("temporal correlation", ""),
            ped.get("geographical correlation", ""),
            ped.get("further technological correlation", ""),
            si.basic_variance, round(si.gsd2, 4),
        ])
    if not result.scored_inputs:
        ws.append([
            "Nothing was scored. Every foreground row and parameter was held at "
            "its resolved value.", "", "", "", "", "", "", "", "", "",
        ])
    else:
        ws.append([])
        ws.append([
            "Scores compose as sigma_i^2 = [ln(factor_i)/2]^2 using the "
            "Weidema/Frischknecht factors ecoinvent applied to the background, "
            "so a foreground score and a background exchange share one matrix.",
            "", "", "", "", "", "", "", "", "",
        ])
    ws.append([])
    ws.append([_GSD2_NOTE])
    autosize(ws)

    # ── Samples ───────────────────────────────────────────────────────────────
    # Written as a NOTE sheet when the draws were not retained, never omitted.
    # An absent sheet is ambiguous with "this build does not produce one", and
    # a reader comparing two workbooks would have no way to tell which.
    ws = wb.create_sheet("Samples")
    with_samples = [d for d in result.distributions if d.samples]
    if not with_samples:
        ws.append(["Samples were not retained for this run."])
        style_header(ws)
        ws.append([
            "The run was launched with keep_samples disabled, so the "
            "per-iteration values were summarised and discarded. Re-run with "
            "samples retained to populate this sheet; the percentiles on "
            "Distributions are unaffected."
        ])
    else:
        ws.append(["Iteration"] + [d.method_label for d in with_samples])
        style_header(ws)
        n = max(len(d.samples or []) for d in with_samples)
        for i in range(n):
            ws.append(
                [i + 1]
                + [(d.samples[i] if d.samples and i < len(d.samples) else None)
                   for d in with_samples]
            )
        apply_sci(ws, min_row=2, min_col=2, max_col=1 + len(with_samples))
    autosize(ws)

    return wb


@router.post("/lca/monte-carlo/export")
async def post_monte_carlo_export(body: MonteCarloExportRequest) -> Any:
    from mapper.api.bom import build_export_filename
    from mapper.api.cohort_export import excel_response

    wb = _build_monte_carlo_workbook(body.result, body.coverage)
    # Single-product: the archetype takes the system slot, never a subsystem.
    filename = build_export_filename(body.result.archetype_name, [], MC_DOMAIN)
    return excel_response(wb, filename, kind="data")




def _build_monte_carlo_multi_workbook(result: MonteCarloMultiResult) -> "Workbook":
    """Multi-item paired results.

    The Item column sits alongside Sensitivity case, matching how every other
    multi-axis export carries its discriminator. Pairwise differences get their
    OWN sheet rather than extra columns: they are the headline of a paired run,
    not an annotation on the per-item rows, and one row per (pair x indicator)
    does not fit beside one row per (item x indicator).
    """
    from openpyxl import Workbook

    from mapper.api.cohort_export import apply_sci, autosize, style_header

    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws.append(["Field", "Value"])
    style_header(ws)
    for k, v in [
        ("Items", len(result.items)),
        ("Sensitivity case", result.parameter_scenario or "Base"),
        ("Scope", _scope_label(result.scope)),
        ("Background database", result.compute_database or "base ecoinvent"),
        ("Iterations", result.n_iterations),
        ("Seed", result.seed),
        ("Elapsed (s)", round(result.elapsed_seconds, 2)),
        ("", ""),
        ("Sampling", "PAIRED. One draw set per iteration, applied to every "
                     "item, so the pairwise differences are meaningful. "
                     "Marginals are unaffected."),
        ("Caveat", LOWER_BOUND_NOTE),
    ]:
        ws.append([k, v])
    if result.warnings:
        ws.append(["Warnings", " | ".join(result.warnings)])
    autosize(ws)

    ws = wb.create_sheet("Distributions")
    ws.append([
        "Item", "Sensitivity case", "Indicator", "Unit", "Deterministic", "Median",
        "Median / deterministic", "Mean", "p2.5", "p25", "p75", "p97.5",
        "GSD2 = exp(2*sigma)", "95% dispersion factor = p97.5 / median",
        "Iterations", "Seed",
    ])
    style_header(ws)
    case = result.parameter_scenario or "Base"
    for it in result.items:
        for d in it.distributions:
            ratio = (d.median / d.deterministic) if d.deterministic else None
            ws.append([
                it.archetype_name, case, d.method_label, d.unit,
                d.deterministic, d.median,
                round(ratio, 4) if ratio is not None else "n/a",
                d.mean, d.p2_5, d.p25, d.p75, d.p97_5,
                round(d.gsd2, 4) if d.gsd2 else "n/a",
                round(d.dispersion_95, 4) if d.dispersion_95 else "n/a",
                d.n_iterations, d.seed,
            ])
    if result.items:
        apply_sci(ws, min_row=2, min_col=5, max_col=6)
        apply_sci(ws, min_row=2, min_col=8, max_col=12)
    ws.append([])
    ws.append([_GSD2_NOTE])
    ws.append([_DISPERSION_NOTE])
    ws.append([_MIGRATION_NOTE])
    autosize(ws)

    ws = wb.create_sheet("Pairwise differences")
    ws.append([
        "A", "B", "Indicator", "Unit", "Deterministic (A-B)", "Median (A-B)",
        "Mean", "p2.5", "p25", "p75", "p97.5",
        "A lower in", "Correlation(A,B)",
    ])
    style_header(ws)
    for d in result.differences:
        ws.append([
            d.a_name, d.b_name, d.method_label, d.unit,
            d.deterministic, d.median, d.mean, d.p2_5, d.p25, d.p75, d.p97_5,
            f"{d.fraction_a_lower * 100:.1f}%", round(d.correlation, 4),
        ])
    if result.differences:
        apply_sci(ws, min_row=2, min_col=5, max_col=11)
        ws.append([])
        ws.append([
            "Correlation is INFORMATIVE, not a warning. A weakly correlated "
            "pair gives a genuinely wide difference and that is correct; the "
            "correlation says where the precision comes from.",
        ])
    else:
        ws.append(["A single item has no pairwise difference."])
    autosize(ws)

    ws = wb.create_sheet("Samples")
    cols: list[tuple[str, list[float]]] = []
    for it in result.items:
        for d in it.distributions:
            if d.samples:
                cols.append((f"{it.archetype_name} — {d.method_label}", d.samples))
    if not cols:
        ws.append(["Samples were not retained for this run."])
        style_header(ws)
        ws.append([
            "Re-run with samples retained to populate this sheet; the "
            "percentiles on Distributions are unaffected."
        ])
    else:
        ws.append(["Iteration"] + [c[0] for c in cols])
        style_header(ws)
        n = max(len(c[1]) for c in cols)
        for i in range(n):
            ws.append([i + 1] + [(c[1][i] if i < len(c[1]) else None) for c in cols])
        apply_sci(ws, min_row=2, min_col=2, max_col=1 + len(cols))
    autosize(ws)

    return wb


@router.post("/lca/monte-carlo/multi/export")
async def post_monte_carlo_multi_export(body: MonteCarloMultiExportRequest) -> Any:
    from mapper.api.bom import build_export_filename
    from mapper.api.cohort_export import excel_response

    wb = _build_monte_carlo_multi_workbook(body.result)
    names = [i.archetype_name for i in body.result.items]
    head = names[0] if names else "comparison"
    filename = build_export_filename(head, names[1:], MC_DOMAIN)
    return excel_response(wb, filename, kind="data")


# ── Material pedigree library ────────────────────────────────────────────────


def _current_project() -> str:
    import bw2data

    return bw2data.projects.current


@router.get("/lca/material-pedigree", response_model=MaterialPedigreeLibrary)
async def get_material_pedigree() -> MaterialPedigreeLibrary:
    from mapper.core.material_pedigree_storage import load_library

    return load_library(_current_project())


@router.put("/lca/material-pedigree", response_model=MaterialPedigreeLibrary)
async def put_material_pedigree(body: MaterialPedigreeLibrary) -> MaterialPedigreeLibrary:
    from mapper.core.material_pedigree_storage import load_library, save_library

    save_library(_current_project(), body)
    return load_library(_current_project())


@router.get("/lca/material-pedigree/materials", response_model=MaterialScoringScope)
async def list_project_materials() -> MaterialScoringScope:
    """The rows of the scoring table, plus what it cannot reach.

    SPLICED. A composed archetype's children are part of what it computes, so
    they must be part of what it can score -- and the impact-weighted
    denominator has always been spliced (it comes from
    ``_build_archetype_source_demand``). Walking the unspliced tree here made
    the count and the percentage disagree on a composed archetype.

    Expression rows are excluded from ``materials`` because they can never
    carry their own score, but they are COUNTED, so the UI can say where the
    uncertainty actually lives instead of showing a short list with no
    explanation.
    """
    return _project_scoring_scope()


def _project_scoring_scope() -> MaterialScoringScope:
    from mapper.api.bom import _proj_archetypes

    arcs = _proj_archetypes()
    names: set[str] = set()
    expr_names: set[str] = set()
    literal_rows = 0
    expr_rows = 0
    for arc in arcs.values():
        for root in _spliced_roots(arc, arcs):
            literal_rows, expr_rows = _collect_names(
                root, names, expr_names, literal_rows, expr_rows
            )
    return MaterialScoringScope(
        materials=sorted(names),
        expression_rows=expr_rows,
        expression_names=len(expr_names),
        literal_rows=literal_rows,
        archetypes=len(arcs),
    )


def _spliced_roots(arc: Any, registry: dict) -> list:
    """``arc.bom`` with any ``includes`` resolved.

    A dangling or cyclic reference must not take the whole materials list down
    -- the archetype is broken and compute will say so loudly; this endpoint
    degrades to that archetype's own rows.
    """
    from mapper.core.bom_engine import ArchetypeCompositionError, splice_includes

    if not getattr(arc, "includes", None):
        return arc.bom
    try:
        return splice_includes(arc, registry).bom
    except ArchetypeCompositionError:
        return arc.bom


def _collect_names(
    node: Any,
    out: set[str],
    expr_out: set[str],
    literal_rows: int,
    expr_rows: int,
) -> tuple[int, int]:
    if node.node_type == "material":
        if node.quantity_expression:
            expr_out.add(node.name)
            expr_rows += 1
        else:
            out.add(node.name)
            literal_rows += 1
    for c in (node.children or []):
        literal_rows, expr_rows = _collect_names(
            c, out, expr_out, literal_rows, expr_rows
        )
    return literal_rows, expr_rows


@router.get("/lca/material-pedigree/coverage", response_model=PedigreeCoverage)
async def get_pedigree_coverage(
    archetype_id: str,
    method: str,
    scope: str = "all",
    compute_database: str | None = None,
) -> PedigreeCoverage:
    """How much of one archetype's impact rests on scored data.

    ``method`` is the ``|``-joined tuple, e.g.
    ``EF v3.1|climate change|global warming potential (GWP100)``.

    The impact weighting is the point. A row count reports clicking; this
    reports how much of the ANSWER is assessed, which is where the next hour of
    scoring is worth spending and what makes a reported GSD^2 legible.
    """
    from mapper.api.lca import _build_archetype_source_demand, _translate_demand_to_database
    from mapper.core.bw2_wrapper import PersistentLCARunner
    from mapper.core.material_pedigree_storage import load_library

    method_tuple = tuple(method.split("|"))
    if len(method_tuple) < 2:
        raise HTTPException(status_code=400, detail="method must be a '|'-joined tuple")

    bundle = _build_archetype_source_demand(
        archetype_id=archetype_id, scope=scope, amount=1.0, stage_amounts={},
        methods=[list(method_tuple)], parameter_scenario=None,
    )
    library = load_library(_current_project())

    # Per-(db, code) UNIT score, one back-substitution each after the first
    # factorization -- so contributions cost ~0.02 s per distinct activity
    # rather than a solve per row.
    runner = PersistentLCARunner()
    unit: dict[tuple[str, str], float] = {}
    for m in bundle.linked:
        link = m.ecoinvent_activity
        key = (link.database, link.code)
        if key in unit:
            continue
        d, _ = _translate_demand_to_database({key: 1.0}, compute_database)
        if not d:
            unit[key] = 0.0
            continue
        scores = runner(d, [method_tuple])
        unit[key] = scores.get(method_tuple, (0.0, ""))[0]

    unit_label = ""
    if bundle.linked:
        first = bundle.linked[0].ecoinvent_activity
        d, _ = _translate_demand_to_database({(first.database, first.code): 1.0}, compute_database)
        if d:
            unit_label = runner(d, [method_tuple]).get(method_tuple, (0.0, ""))[1]

    # |impact| per material NAME, and whether that row is scored.
    by_name: dict[str, float] = {}
    scored_names: set[str] = set()
    for m in bundle.linked:
        if m.quantity_expression:
            # Not scoreable -- it inherits from its parameters. Counting it as
            # unscored would understate coverage and point the user at a row
            # they cannot fix.
            continue
        link = m.ecoinvent_activity
        amt = getattr(m, "_stage_amount", 1.0)
        contrib = abs(m.quantity * amt * unit[(link.database, link.code)])
        by_name[m.name] = by_name.get(m.name, 0.0) + contrib
        if m.uncertainty is not None or m.name in library.entries:
            scored_names.add(m.name)

    total = sum(by_name.values())
    covered = sum(v for k, v in by_name.items() if k in scored_names)
    top = sorted(
        ((k, v) for k, v in by_name.items() if k not in scored_names),
        key=lambda kv: -kv[1],
    )[:10]

    # The SAME spliced walk the materials list uses, so the count and the
    # percentage can never be computed over different row sets.
    scope = _project_scoring_scope()
    project_names = set(scope.materials)

    return PedigreeCoverage(
        materials_total=len(project_names),
        materials_scored=len(set(library.entries) & project_names),
        archetype_materials_total=len(by_name),
        archetype_materials_scored=len(scored_names),
        # None, not 0.0, when there is nothing here to score: every row is a
        # parameter expression. 0% would tell the user they are missing
        # something fixable.
        impact_share=(covered / total) if by_name and total > 0 else None,
        method_label=" | ".join(method_tuple[1:]) or method_tuple[0],
        unit=unit_label,
        top_unscored=[
            UnscoredMaterial(name=k, share=(v / total) if total else 0.0, impact=v)
            for k, v in top if v > 0
        ],
    )


def _validate_methods_registered(methods: list[list[str]]) -> None:
    """Delegates to the shared api-layer validator. Both MC routes share
    ``_method_cf_samplers``, so both call this."""
    from mapper.api.method_validation import validate_methods_registered

    validate_methods_registered(methods)


@router.post("/lca/monte-carlo", response_model=MonteCarloStartResponse)
async def post_monte_carlo(body: MonteCarloRequest) -> MonteCarloStartResponse:
    from mapper.api.parameters import validate_parameter_scenarios

    validate_parameter_scenarios(body.parameter_scenario)
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required")
    if body.iterations < 1 or body.iterations > MAX_ITERATIONS:
        raise HTTPException(
            status_code=400,
            detail=f"iterations must be between 1 and {MAX_ITERATIONS}",
        )
    _validate_methods_registered(body.methods)

    task_id = str(uuid.uuid4())
    task = _TaskState()
    with _TASK_LOCK:
        _TASKS[task_id] = task
    task_registry.register(task_id)

    def work() -> None:
        try:
            task.result = _run_monte_carlo(body, task, task_id)
            task.done = True
            _notify_all(task, {"type": "done", "task_id": task_id})
        except CancelledOperation:
            task.cancelled = True
            task.done = True
            _notify_all(task, {"type": "cancelled", "task_id": task_id})
        except HTTPException as e:
            task.error = str(e.detail)
            task.done = True
            _notify_all(task, {"type": "error", "error": task.error})
        except Exception as e:  # noqa: BLE001 - surfaced to the client
            task.error = f"{type(e).__name__}: {e}"
            task.done = True
            _notify_all(task, {"type": "error", "error": task.error})
        finally:
            task_registry.unregister(task_id)

    # A plain daemon thread, as `plca` and `impact` do for the same shape.
    # NOT `core.tasks.run_in_thread`: that drives a `core.tasks.Task`
    # (status/finish/fail) and calls `fn(task, ...)`, whereas this route owns a
    # WS-oriented `_TaskState` and a zero-arg closure. Passing the closure
    # alone raised `run_in_thread() missing 1 required positional argument`
    # on every call, so the endpoint 500'd before any sampling began.
    threading.Thread(target=work, daemon=True).start()
    return MonteCarloStartResponse(task_id=task_id)



@router.post("/lca/monte-carlo/multi", response_model=MonteCarloStartResponse)
async def post_monte_carlo_multi(body: MonteCarloMultiRequest) -> MonteCarloStartResponse:
    from mapper.api.parameters import validate_parameter_scenarios

    validate_parameter_scenarios(body.parameter_scenario)
    if not body.archetype_ids:
        raise HTTPException(status_code=400, detail="At least one item is required")
    if not body.methods:
        raise HTTPException(status_code=400, detail="At least one method is required")
    if body.iterations < 1 or body.iterations > MAX_ITERATIONS:
        raise HTTPException(
            status_code=400, detail=f"iterations must be between 1 and {MAX_ITERATIONS}")
    _validate_methods_registered(body.methods)

    task_id = str(uuid.uuid4())
    task = _TaskState()
    with _TASK_LOCK:
        _TASKS[task_id] = task
    task_registry.register(task_id)

    def work() -> None:
        try:
            task.multi_result = _run_monte_carlo_multi(body, task, task_id)
            task.done = True
            _notify_all(task, {"type": "done", "task_id": task_id})
        except CancelledOperation:
            # ONE task id for the whole job, so a single cancel stops every
            # item rather than leaving the rest running.
            task.cancelled = True
            task.done = True
            _notify_all(task, {"type": "cancelled", "task_id": task_id})
        except HTTPException as e:
            task.error = str(e.detail)
            task.done = True
            _notify_all(task, {"type": "error", "error": task.error})
        except Exception as e:  # noqa: BLE001 - surfaced to the client
            task.error = f"{type(e).__name__}: {e}"
            task.done = True
            _notify_all(task, {"type": "error", "error": task.error})
        finally:
            task_registry.unregister(task_id)

    threading.Thread(target=work, daemon=True).start()
    return MonteCarloStartResponse(task_id=task_id)


@router.get("/lca/monte-carlo/multi/{task_id}", response_model=MonteCarloMultiResult | dict)
async def get_monte_carlo_multi(task_id: str) -> Any:
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Unknown task id")
    if task.cancelled:
        return {"cancelled": True, "task_id": task_id}
    if task.error:
        raise HTTPException(status_code=500, detail=task.error)
    if not task.done or task.multi_result is None:
        raise HTTPException(status_code=409, detail="Task is still running")
    return task.multi_result


@router.get("/lca/monte-carlo/{task_id}", response_model=MonteCarloResult | dict)
async def get_monte_carlo(task_id: str) -> Any:
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None:
        raise HTTPException(status_code=404, detail="Unknown task id")
    if task.cancelled:
        return {"cancelled": True, "task_id": task_id}
    if task.error:
        raise HTTPException(status_code=500, detail=task.error)
    if not task.done or task.result is None:
        raise HTTPException(status_code=409, detail="Task is still running")
    return task.result


@router.websocket("/lca/monte-carlo/ws/{task_id}")
async def ws_monte_carlo(websocket: WebSocket, task_id: str) -> None:
    await websocket.accept()
    with _TASK_LOCK:
        task = _TASKS.get(task_id)
    if task is None:
        await websocket.send_json({"type": "error", "error": "Unknown task id"})
        await websocket.close()
        return

    queue: asyncio.Queue = asyncio.Queue(maxsize=256)
    task.subscribers.append(queue)
    await websocket.send_json({"type": "progress", "stage": task.stage, "pct": task.pct})
    if task.done:
        if task.cancelled:
            await websocket.send_json({"type": "cancelled", "task_id": task_id})
        elif task.error:
            await websocket.send_json({"type": "error", "error": task.error})
        else:
            await websocket.send_json({"type": "done", "task_id": task_id})
        await websocket.close()
        return

    try:
        while True:
            payload = await queue.get()
            await websocket.send_json(payload)
            if payload.get("type") in ("done", "error", "cancelled"):
                break
    except WebSocketDisconnect:
        pass
    finally:
        try:
            task.subscribers.remove(queue)
        except ValueError:
            pass
        task_registry.maybe_cancel_on_last_subscriber_leave(
            task_id,
            remaining_subscribers=len(task.subscribers),
            task_done=task.done,
        )
        try:
            await websocket.close()
        except Exception:
            pass
