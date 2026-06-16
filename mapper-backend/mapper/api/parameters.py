"""FastAPI router for the parameter table + scenarios.

Each brightway2 project has exactly one :class:`ParameterTable`. The table is
a grid of parameters (rows) × scenarios (columns): the ``Base`` column is
implicit and always present; user-defined scenario columns hold per-parameter
overrides. Missing override entries inherit from Base.

The in-memory registry ``_tables`` is the source of truth during a session;
every write is mirrored to JSON on disk via
:mod:`mapper.core.parameter_storage`.

Legacy ``/sets/*`` routes are kept as 301 redirects so old clients don't
die — the redirect targets and the synthesized ``ParameterSet`` returned by
:func:`get_parameter_set` let the DSM-LCA pipeline address any scenario by
name.
"""
from __future__ import annotations

import io
import re
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, Response, UploadFile
from fastapi.responses import RedirectResponse

from mapper.api.project_guard import verify_project_state
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from mapper.api.dsm import _current_project
from mapper.core import parameter_storage
from mapper.core.parameter_engine import (
    ParameterEngine,
    ParameterError,
    validate_parameter_name,
)
from mapper.models.parameter_schemas import (
    Parameter,
    ParameterSet,
    ParameterSetSummary,
    ParameterTable,
    ParameterTableUpdate,
    ResolveRequest,
    ResolveResult,
    ScenarioCreate,
    ScenarioRename,
    ValidateRequest,
    ValidateResult,
)


router = APIRouter(prefix="/parameters", tags=["parameters"])


# ── In-memory registry ──────────────────────────────────────────────────────

# ``{project_name -> ParameterTable}``. Populated on startup by
# :func:`install_parameters` from :func:`parameter_storage.load_all`.
_tables: dict[str, ParameterTable] = {}


def install_parameters(data: dict[str, ParameterTable]) -> None:
    """Replace the in-memory table registry (called from main startup)."""
    _tables.clear()
    _tables.update(data)


def _table_for(project: str | None = None) -> ParameterTable:
    p = project or _current_project()
    return _tables.setdefault(p, ParameterTable())


def _persist(project: str, table: ParameterTable) -> None:
    _tables[project] = table
    parameter_storage.save_parameter_table(project, table)


def _now() -> str:
    return datetime.utcnow().isoformat()


# ── Public accessor used by DSM-LCA, BOM, subsystems, impact ────────────────


def get_parameter_set(
    scenario_or_id: str | None, project: str | None = None
) -> ParameterSet | None:
    """Return a synthesized :class:`ParameterSet` for ``scenario_or_id``.

    ``scenario_or_id`` is treated as a scenario name (``"Base"`` or any entry
    in ``table.scenarios``). Legacy callers that pass the old UUID ``set_id``
    now get ``None`` — they should migrate to scenario names.
    """
    if scenario_or_id is None:
        return None
    table = _table_for(project)
    if scenario_or_id not in table.list_scenarios():
        return None
    resolved = table.resolve_all(scenario_or_id)
    params: list[Parameter] = []
    for name, value in resolved.items():
        src = table.parameters[name]
        params.append(Parameter(
            name=name,
            base_value=value,
            unit=src.unit,
            description=src.description,
            category=src.category,
        ))
    return ParameterSet(
        id=scenario_or_id,
        name=scenario_or_id,
        parameters=params,
        created_at=table.created_at,
        updated_at=table.updated_at,
    )


# ── ParameterTable CRUD ─────────────────────────────────────────────────────


def _validate_params(params: dict[str, Parameter]) -> None:
    for key, p in params.items():
        err = validate_parameter_name(p.name)
        if err:
            raise HTTPException(status_code=400, detail=err)
        if p.name != key:
            raise HTTPException(
                status_code=400,
                detail=f"Parameter dict key '{key}' does not match name '{p.name}'",
            )
        try:
            float(p.base_value)
        except (TypeError, ValueError):
            raise HTTPException(
                status_code=400,
                detail=f"Parameter '{p.name}' has non-numeric base_value",
            )
        for scen, v in p.scenario_overrides.items():
            try:
                float(v)
            except (TypeError, ValueError):
                raise HTTPException(
                    status_code=400,
                    detail=f"Parameter '{p.name}' has non-numeric override for scenario '{scen}'",
                )


@router.get("/table", response_model=ParameterTable)
async def get_table() -> ParameterTable:
    return _table_for()


@router.put("/table", response_model=ParameterTable)
async def put_table(body: ParameterTableUpdate) -> ParameterTable:
    project = _current_project()
    table = _table_for(project)
    new_params = body.parameters if body.parameters is not None else table.parameters
    new_scenarios = body.scenarios if body.scenarios is not None else table.scenarios
    new_categories = body.categories if body.categories is not None else table.categories
    _validate_params(new_params)
    # Scrub overrides against the known scenario list — any extras get dropped.
    scen_set = set(new_scenarios)
    cleaned: dict[str, Parameter] = {}
    for name, p in new_params.items():
        cleaned[name] = p.model_copy(update={
            "scenario_overrides": {
                s: v for s, v in p.scenario_overrides.items() if s in scen_set
            },
        })
    # Normalize categories: trim, drop empties, dedup while preserving order.
    seen: set[str] = set()
    norm_categories: list[str] = []
    for c in new_categories:
        t = (c or "").strip()
        if t and t not in seen:
            seen.add(t)
            norm_categories.append(t)
    updated = ParameterTable(
        parameters=cleaned,
        scenarios=list(new_scenarios),
        categories=norm_categories,
        created_at=table.created_at or _now(),
        updated_at=_now(),
    )
    _persist(project, updated)
    return updated


# ── Scenario columns ────────────────────────────────────────────────────────


@router.post(
    "/table/scenarios",
    response_model=ParameterTable,
    dependencies=[Depends(verify_project_state)],
)
async def create_scenario(body: ScenarioCreate) -> ParameterTable:
    project = _current_project()
    table = _table_for(project)
    name = body.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Scenario name is required")
    if name == ParameterTable.BASE_SCENARIO:
        raise HTTPException(
            status_code=400,
            detail=f"'{ParameterTable.BASE_SCENARIO}' is reserved",
        )
    if name in table.scenarios:
        raise HTTPException(status_code=400, detail=f"Scenario '{name}' already exists")

    new_scenarios = [*table.scenarios, name]
    new_params: dict[str, Parameter] = {}
    for pname, p in table.parameters.items():
        overrides = dict(p.scenario_overrides)
        if body.copy_from and body.copy_from in table.list_scenarios():
            src_val = table.resolve(pname, body.copy_from)
            # Only store as override if it differs from base (keep cells empty
            # when they'd just echo the Base value).
            if src_val != p.base_value:
                overrides[name] = src_val
        new_params[pname] = p.model_copy(update={"scenario_overrides": overrides})

    updated = ParameterTable(
        parameters=new_params,
        scenarios=new_scenarios,
        categories=list(table.categories),
        created_at=table.created_at,
        updated_at=_now(),
    )
    _persist(project, updated)
    return updated


@router.delete("/table/scenarios/{name}", response_model=ParameterTable)
async def delete_scenario(name: str) -> ParameterTable:
    project = _current_project()
    table = _table_for(project)
    if name == ParameterTable.BASE_SCENARIO:
        raise HTTPException(status_code=400, detail="Cannot delete the Base scenario")
    if name not in table.scenarios:
        raise HTTPException(status_code=404, detail=f"Scenario '{name}' not found")
    new_params = {
        pname: p.model_copy(update={
            "scenario_overrides": {s: v for s, v in p.scenario_overrides.items() if s != name},
        })
        for pname, p in table.parameters.items()
    }
    updated = ParameterTable(
        parameters=new_params,
        scenarios=[s for s in table.scenarios if s != name],
        categories=list(table.categories),
        created_at=table.created_at,
        updated_at=_now(),
    )
    _persist(project, updated)
    return updated


@router.patch("/table/scenarios", response_model=ParameterTable)
async def rename_scenario(body: ScenarioRename) -> ParameterTable:
    project = _current_project()
    table = _table_for(project)
    if body.old_name == ParameterTable.BASE_SCENARIO or body.new_name == ParameterTable.BASE_SCENARIO:
        raise HTTPException(status_code=400, detail="Cannot rename to/from Base")
    if body.old_name not in table.scenarios:
        raise HTTPException(status_code=404, detail=f"Scenario '{body.old_name}' not found")
    if body.new_name in table.scenarios:
        raise HTTPException(status_code=400, detail=f"Scenario '{body.new_name}' already exists")
    new_params = {
        pname: p.model_copy(update={
            "scenario_overrides": {
                (body.new_name if s == body.old_name else s): v
                for s, v in p.scenario_overrides.items()
            },
        })
        for pname, p in table.parameters.items()
    }
    updated = ParameterTable(
        parameters=new_params,
        scenarios=[body.new_name if s == body.old_name else s for s in table.scenarios],
        categories=list(table.categories),
        created_at=table.created_at,
        updated_at=_now(),
    )
    _persist(project, updated)
    return updated


# ── Expression resolve / validate ───────────────────────────────────────────


def _engine_for_scenario(scenario: str | None) -> ParameterEngine:
    return ParameterEngine(_table_for(), scenario=scenario)


@router.post("/resolve", response_model=ResolveResult)
async def resolve(body: ResolveRequest) -> ResolveResult:
    scen = body.scenario or body.parameter_set_id  # accept legacy field
    engine = _engine_for_scenario(scen)
    refs = sorted(ParameterEngine.find_references(body.expression))
    try:
        value = engine.resolve(body.expression)
    except ParameterError as e:
        return ResolveResult(expression=body.expression, value=None, error=str(e), references=refs)
    return ResolveResult(expression=body.expression, value=value, error=None, references=refs)


@router.post("/validate", response_model=ValidateResult)
async def validate(body: ValidateRequest) -> ValidateResult:
    scen = body.scenario or body.parameter_set_id
    engine = _engine_for_scenario(scen)
    out: list[ResolveResult] = []
    for expr in body.expressions:
        refs = sorted(ParameterEngine.find_references(expr))
        try:
            v = engine.resolve(expr)
            out.append(ResolveResult(expression=expr, value=v, error=None, references=refs))
        except ParameterError as e:
            out.append(ResolveResult(expression=expr, value=None, error=str(e), references=refs))
    return ValidateResult(results=out)


# ── Excel template / import / export ────────────────────────────────────────


_BASE_COLS = ["Name", "Base Value", "Unit", "Description", "Category"]


def _style_header(ws, n_cols: int) -> None:
    header_fill = PatternFill(start_color="FF4F46E5", end_color="FF4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _build_workbook(table: ParameterTable | None) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Parameters"

    scenarios = table.scenarios if table else []
    headers = [*_BASE_COLS, *scenarios]
    ws.append(headers)
    _style_header(ws, len(headers))

    if table and table.parameters:
        for p in table.parameters.values():
            row = [p.name, p.base_value, p.unit or "", p.description or "", p.category or ""]
            for s in scenarios:
                row.append(p.scenario_overrides.get(s, ""))
            ws.append(row)
    else:
        ws.append(["battery_mass_lfp", 250, "kg", "LFP battery pack mass", "Battery"])
        ws.append(["battery_mass_nmc811", 230, "kg", "NMC811 battery pack mass", "Battery"])
        ws.append(["dk_annual_km", 15000, "km/yr", "Danish average annual driving distance", "Use Phase"])
        ws.append(["electricity_consumption_bev", 0.17, "kWh/km", "BEV electricity consumption per km", "Use Phase"])

    widths = [28, 12, 12, 48, 16] + [14] * len(scenarios)
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    instructions = wb.create_sheet("Instructions")
    instructions.append(["Field", "Notes"])
    _style_header(instructions, 2)
    rows = [
        ["Name", "Unique identifier; snake_case (lowercase letters, digits, underscore). Cannot be min/max/abs/round/sum."],
        ["Base Value", "Numeric value used for the Base scenario (and inherited by scenario columns left empty)."],
        ["Unit", "Optional descriptive unit (kg, kWh/km, km/yr, …)."],
        ["Description", "Optional human-readable note."],
        ["Category", "Optional grouping label (Battery, Drivetrain, Use Phase, …)."],
        ["<Scenario columns>", "One column per scenario. Leave a cell blank to inherit the Base Value; enter a number to override."],
        ["", ""],
        ["Using parameters in BOMs", "In a BOM's Quantity cell, write the parameter name or an expression like 'battery_mass_lfp * 0.35'. Supported: + - * / ** ( ) and min, max, abs, round, sum."],
    ]
    for r in rows:
        instructions.append(r)
    instructions.column_dimensions["A"].width = 24
    instructions.column_dimensions["B"].width = 110

    return wb


def _sanitize_filename(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", (name or "parameters").strip())
    return cleaned or "parameters"


@router.get("/template")
async def download_template() -> Response:
    wb = _build_workbook(None)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="parameters_template.xlsx"'},
    )


@router.get("/export")
async def export_table() -> Response:
    table = _table_for()
    wb = _build_workbook(table)
    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="parameters.xlsx"'},
    )


@router.post(
    "/import",
    response_model=ParameterTable,
    dependencies=[Depends(verify_project_state)],
)
async def import_table(file: UploadFile = File(...), mode: str = "replace") -> ParameterTable:
    """Import a parameters workbook.

    ``mode=replace`` overwrites the existing table; ``mode=merge`` overlays
    rows from the file onto the current table (adds new rows, overwrites
    values for rows that already exist, leaves untouched rows alone).
    """
    if mode not in ("replace", "merge"):
        raise HTTPException(status_code=400, detail=f"Unknown mode '{mode}' (expected replace|merge)")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read workbook: {e}")

    ws = None
    for candidate in ("Parameters", "parameters"):
        if candidate in wb.sheetnames:
            ws = wb[candidate]
            break
    if ws is None:
        ws = wb.worksheets[0]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(status_code=400, detail="Empty workbook")
    header = [str(c).strip() if c is not None else "" for c in header_row]

    # "Value" is accepted as a legacy alias for "Base Value".
    name_idx = header.index("Name") if "Name" in header else -1
    base_idx = (
        header.index("Base Value") if "Base Value" in header
        else header.index("Value") if "Value" in header
        else -1
    )
    if name_idx < 0 or base_idx < 0:
        raise HTTPException(
            status_code=400,
            detail=f"Sheet must have 'Name' and 'Base Value' (or 'Value') columns; got {header}",
        )

    def idx(key: str) -> int:
        return header.index(key) if key in header else -1

    unit_i, desc_i, cat_i = idx("Unit"), idx("Description"), idx("Category")
    known_fixed = {"Name", "Base Value", "Value", "Unit", "Description", "Category"}
    scenario_columns = [
        (i, h) for i, h in enumerate(header)
        if h and h not in known_fixed
    ]

    new_params: dict[str, Parameter] = {}
    warnings: list[str] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        pname = row[name_idx] if name_idx < len(row) else None
        if pname is None or str(pname).strip() == "":
            continue
        pname_s = str(pname).strip()
        err = validate_parameter_name(pname_s)
        if err:
            warnings.append(f"Row {r_idx}: {err}; skipped.")
            continue
        if pname_s in new_params:
            warnings.append(f"Row {r_idx}: duplicate '{pname_s}'; skipped.")
            continue
        try:
            base_val = float(row[base_idx])
        except (TypeError, ValueError, IndexError):
            warnings.append(f"Row {r_idx}: non-numeric base value for '{pname_s}'; skipped.")
            continue
        unit_v = row[unit_i] if 0 <= unit_i < len(row) else None
        desc_v = row[desc_i] if 0 <= desc_i < len(row) else None
        cat_v = row[cat_i] if 0 <= cat_i < len(row) else None

        overrides: dict[str, float] = {}
        for col_i, scen_name in scenario_columns:
            if col_i >= len(row):
                continue
            cell = row[col_i]
            if cell is None or (isinstance(cell, str) and cell.strip() == ""):
                continue  # blank => inherit Base
            try:
                overrides[scen_name] = float(cell)
            except (TypeError, ValueError):
                warnings.append(
                    f"Row {r_idx}: non-numeric override for '{pname_s}' in scenario '{scen_name}'; skipped."
                )

        new_params[pname_s] = Parameter(
            name=pname_s,
            base_value=base_val,
            unit=str(unit_v).strip() if unit_v else None,
            description=str(desc_v).strip() if desc_v else None,
            category=str(cat_v).strip() if cat_v else None,
            scenario_overrides=overrides,
        )

    scen_names = [h for _, h in scenario_columns]

    project = _current_project()
    existing = _table_for(project)
    if mode == "merge":
        merged = {**existing.parameters, **new_params}
        merged_scenarios = list(dict.fromkeys([*existing.scenarios, *scen_names]))
        table = ParameterTable(
            parameters=merged,
            scenarios=merged_scenarios,
            categories=list(existing.categories),
            created_at=existing.created_at or _now(),
            updated_at=_now(),
        )
    else:  # replace
        table = ParameterTable(
            parameters=new_params,
            scenarios=scen_names,
            categories=[],
            created_at=existing.created_at or _now(),
            updated_at=_now(),
        )
    _persist(project, table)
    return table


# ── Legacy /sets/* → 301 redirects ──────────────────────────────────────────
#
# Old frontend clients still address per-variant sets by id. Redirect them to
# the single-table endpoint so they get a coherent error path during the
# transition (and so tests can confirm the migration surface).


@router.get("/sets", response_model=list[ParameterSetSummary])
async def list_sets_legacy() -> list[ParameterSetSummary]:
    """Shim returning the table's scenarios as pseudo-sets."""
    table = _table_for()
    out: list[ParameterSetSummary] = []
    for scen in table.list_scenarios():
        cats = sorted({p.category for p in table.parameters.values() if p.category})
        out.append(ParameterSetSummary(
            id=scen,
            name=scen,
            parameter_count=len(table.parameters),
            categories=cats,
            created_at=table.created_at or "",
            updated_at=table.updated_at or "",
        ))
    return out


@router.get("/sets/{set_id}")
async def get_set_legacy(set_id: str) -> Response:
    pset = get_parameter_set(set_id)
    if pset is None:
        raise HTTPException(status_code=404, detail=f"Scenario '{set_id}' not found")
    return Response(
        content=pset.model_dump_json(),
        media_type="application/json",
    )


@router.post("/sets")
async def create_set_legacy() -> Response:
    return RedirectResponse(url="/api/parameters/table/scenarios", status_code=307)


@router.put("/sets/{set_id}")
async def update_set_legacy(set_id: str) -> Response:
    return RedirectResponse(url="/api/parameters/table", status_code=307)


@router.delete("/sets/{set_id}")
async def delete_set_legacy(set_id: str) -> Response:
    return RedirectResponse(
        url=f"/api/parameters/table/scenarios/{set_id}", status_code=307
    )
