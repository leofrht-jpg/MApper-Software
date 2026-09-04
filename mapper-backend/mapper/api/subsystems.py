# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""FastAPI router for Subsystems (coupled product populations).

Primary subsystems are synthesized from ``SystemDefinition`` at read time and
are not persisted. Only dependent subsystems are stored.

Layout under the per-system storage dir::

    {system_id}/subsystems.json         # dependent subsystems only
    {system_id}/subsystem_results.json  # last compute output per dependent

Dependent-subsystem results are invalidated whenever the primary simulation
result is cleared (see :mod:`mapper.api.dsm`).
"""
from __future__ import annotations

import threading
import uuid

import bw2data
from mapper.core.upload_guard import refuse_if_nothing_resolved
from fastapi import APIRouter, Depends, File, Header, HTTPException, UploadFile
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel

from mapper.api.dsm import (
    _get_system,
    _proj_results,
    _proj_systems,
    _sanitize_filename,
)
from mapper.api.parameters import _table_for, get_parameter_set
from mapper.core import dsm_storage
from mapper.core.dsm_engine import (
    dependent_stock_template_csv,
    inflow_template_csv,
    non_age_dimensions,
    outflow_template_csv,
    parse_dependent_stock_file,
    parse_inflow_file,
    parse_outflow_file,
)
from mapper.core.parameter_engine import ParameterEngine, ParameterError
from mapper.core.subsystem_engine import (
    compute_dependent_subsystem,
    compute_subsystem_result,
    validate_dependency_rule,
)
from mapper.models.dsm_schemas import SimulationResult
from mapper.models.subsystem_schemas import (
    DependencyRule,
    Subsystem,
    SubsystemList,
    SubsystemSummary,
)


def _reconcile_project_from_header(
    x_mapper_project: str | None = Header(default=None),
) -> None:
    """Router dependency — reconcile the backend's active bw2 project to the
    client's declared ``X-Mapper-Project`` when they differ.

    Fixes the project-state desync: a backend restart resets
    ``bw2data.projects.current`` to ``default`` while the user is still on
    another project (e.g. ``MAp-test``). Subsystem (and their parent-system)
    lookups scope by ``_current_project()`` → the in-memory store, so they miss
    and 404 with "Subsystem not found" even though the record exists on disk
    under the real project. Honoring the client's declared project (validated
    against the known bw2 projects) makes the whole subsystem surface — reads
    AND writes, including the parent-system lookup via ``_get_system`` — operate
    on the project the user is actually viewing.

    Runs BEFORE the handler, so ``_current_project()`` inside the handler
    already returns the reconciled value. Header absent/empty or naming an
    unknown project → no-op (bw2 current is used, preserving curl/test and
    non-browser behavior). Subsystem endpoints carry no 409 write-guard
    contract, so — unlike ``verify_project_state`` — this reconciles rather
    than rejecting."""
    if not x_mapper_project or not x_mapper_project.strip():
        return
    if x_mapper_project == bw2data.projects.current:
        return
    try:
        known = {p.name for p in bw2data.projects}
    except Exception:  # noqa: BLE001 — never let reconciliation break a request
        return
    if x_mapper_project in known:
        bw2data.projects.set_current(x_mapper_project)


from mapper.api.cohort_export import excel_response_from_bytes


router = APIRouter(
    prefix="/dsm",
    tags=["subsystems"],
    dependencies=[Depends(_reconcile_project_from_header)],
)

# In-memory stores, hydrated from disk at startup.
# ``{project -> {system_id -> {subsystem_id -> Subsystem}}}``. Only dependent
# subsystems live here; the primary is synthesized from SystemDefinition.
_subsystems: dict[str, dict[str, dict[str, Subsystem]]] = {}
# ``{project -> {system_id -> {subsystem_id -> SimulationResult}}}``.
_subsystem_results: dict[str, dict[str, dict[str, SimulationResult]]] = {}
_lock = threading.Lock()


def _current_project() -> str:
    return bw2data.projects.current


def _proj_subs(project: str | None = None) -> dict[str, dict[str, Subsystem]]:
    p = project or _current_project()
    return _subsystems.setdefault(p, {})


def _proj_sub_results(project: str | None = None) -> dict[str, dict[str, SimulationResult]]:
    p = project or _current_project()
    return _subsystem_results.setdefault(p, {})


def _sys_subs(system_id: str, project: str | None = None) -> dict[str, Subsystem]:
    return _proj_subs(project).setdefault(system_id, {})


def _sys_sub_results(system_id: str, project: str | None = None) -> dict[str, SimulationResult]:
    return _proj_sub_results(project).setdefault(system_id, {})


def invalidate_results(system_id: str, project: str | None = None) -> None:
    """Drop cached dependent-subsystem simulation results for ``system_id``.

    Called from the DSM router when the primary system definition or its
    simulation output changes.
    """
    res = _proj_sub_results(project)
    res.pop(system_id, None)


# ── Primary synthesis ───────────────────────────────────────────────────────


def _synthesize_primary(system_id: str) -> Subsystem:
    sys_def = _get_system(system_id)
    return Subsystem(
        id=system_id,
        name=sys_def.name,
        type="primary",
        dimensions=list(sys_def.dimensions),
        depends_on=None,
        dependency_rules=[],
    )


def _summary_for(sub: Subsystem, result: SimulationResult | None = None) -> SubsystemSummary:
    if sub.type == "primary":
        # Primary archetype count = number of non-age cohort combinations.
        from mapper.core.dsm_engine import all_cohort_keys
        cohort_count = len(all_cohort_keys(sub.dimensions))
        return SubsystemSummary(
            id=sub.id,
            name=sub.name,
            type="primary",
            dimension_count=len(non_age_dimensions(sub.dimensions)),
            archetype_count=cohort_count,
            rule_count=0,
            depends_on=None,
        )
    # Dependent: archetype count = unique ids across rules.
    archetype_ids = {r.dependent_archetype_id for r in sub.dependency_rules}
    return SubsystemSummary(
        id=sub.id,
        name=sub.name,
        type="dependent",
        dimension_count=len(non_age_dimensions(sub.dimensions)),
        archetype_count=len(archetype_ids),
        rule_count=len(sub.dependency_rules),
        depends_on=sub.depends_on,
    )


# ── Validation ──────────────────────────────────────────────────────────────


def _validate_subsystem_shape(body: Subsystem, system_id: str) -> None:
    if body.type != "dependent":
        raise HTTPException(
            status_code=400,
            detail="Only dependent subsystems can be created/updated via this API.",
        )
    if not body.name or not body.name.strip():
        raise HTTPException(status_code=400, detail="Subsystem name is required.")
    if body.depends_on and body.depends_on != system_id:
        raise HTTPException(
            status_code=400,
            detail=f"depends_on '{body.depends_on}' must match system_id '{system_id}'.",
        )
    nads = non_age_dimensions(body.dimensions)
    seen: set[str] = set()
    for d in nads:
        if not d.name:
            raise HTTPException(status_code=400, detail="Each dimension needs a name.")
        if d.name in seen:
            raise HTTPException(
                status_code=400, detail=f"Duplicate dimension name '{d.name}'"
            )
        seen.add(d.name)
        if not d.labels:
            raise HTTPException(
                status_code=400,
                detail=f"Dimension '{d.name}' must have at least one label.",
            )

    primary = _get_system(system_id)
    rule_ids: set[str] = set()
    for rule in body.dependency_rules:
        errors = validate_dependency_rule(rule, primary.dimensions)
        if errors:
            raise HTTPException(
                status_code=400,
                detail=f"Rule {rule.id or '?'}: {'; '.join(errors)}",
            )
        if rule.id and rule.id in rule_ids:
            raise HTTPException(
                status_code=400, detail=f"Duplicate rule id '{rule.id}'"
            )
        if rule.id:
            rule_ids.add(rule.id)


# ── Routes: listing ─────────────────────────────────────────────────────────


@router.get("/systems/{system_id}/subsystems", response_model=SubsystemList)
async def list_subsystems(system_id: str) -> SubsystemList:
    """Return the synthesized primary subsystem plus all stored dependents."""
    primary = _synthesize_primary(system_id)
    deps = list(_sys_subs(system_id).values())
    return SubsystemList(subsystems=[primary] + deps)


@router.get(
    "/systems/{system_id}/subsystems/summary", response_model=list[SubsystemSummary]
)
async def list_subsystem_summaries(system_id: str) -> list[SubsystemSummary]:
    primary = _synthesize_primary(system_id)
    deps = list(_sys_subs(system_id).values())
    return [_summary_for(primary)] + [_summary_for(s) for s in deps]


@router.get(
    "/systems/{system_id}/subsystems/{subsystem_id}", response_model=Subsystem
)
async def get_subsystem(system_id: str, subsystem_id: str) -> Subsystem:
    if subsystem_id == system_id:
        return _synthesize_primary(system_id)
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    return sub


# ── Routes: CRUD ────────────────────────────────────────────────────────────


def _persist_subs(project: str, system_id: str) -> None:
    dsm_storage.save_subsystems(project, system_id, _sys_subs(system_id, project))


def _persist_sub_results(project: str, system_id: str) -> None:
    results = _sys_sub_results(system_id, project)
    if results:
        dsm_storage.save_subsystem_results(project, system_id, results)
    else:
        dsm_storage.clear_subsystem_results(project, system_id)


@router.post("/systems/{system_id}/subsystems", response_model=Subsystem)
async def create_subsystem(system_id: str, body: Subsystem) -> Subsystem:
    # Verify the primary exists before accepting the dependent.
    _get_system(system_id)
    # Assign server-side ids where missing.
    body.id = str(uuid.uuid4())
    body.depends_on = system_id
    body.type = "dependent"
    for rule in body.dependency_rules:
        if not rule.id:
            rule.id = str(uuid.uuid4())
    _validate_subsystem_shape(body, system_id)
    project = _current_project()
    with _lock:
        _sys_subs(system_id, project)[body.id] = body
        _sys_sub_results(system_id, project).pop(body.id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return body


@router.put(
    "/systems/{system_id}/subsystems/{subsystem_id}", response_model=Subsystem
)
async def update_subsystem(
    system_id: str, subsystem_id: str, body: Subsystem
) -> Subsystem:
    if subsystem_id == system_id:
        raise HTTPException(
            status_code=400,
            detail="Primary subsystems are defined by the system itself. Edit the system instead.",
        )
    subs = _sys_subs(system_id)
    if subsystem_id not in subs:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    body.id = subsystem_id
    body.depends_on = system_id
    body.type = "dependent"
    for rule in body.dependency_rules:
        if not rule.id:
            rule.id = str(uuid.uuid4())
    _validate_subsystem_shape(body, system_id)
    project = _current_project()
    with _lock:
        subs[subsystem_id] = body
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return body


@router.delete("/systems/{system_id}/subsystems/{subsystem_id}")
async def delete_subsystem(system_id: str, subsystem_id: str) -> dict[str, bool]:
    if subsystem_id == system_id:
        raise HTTPException(
            status_code=400, detail="Cannot delete a primary subsystem."
        )
    project = _current_project()
    subs = _sys_subs(system_id, project)
    if subsystem_id not in subs:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    with _lock:
        subs.pop(subsystem_id, None)
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return {"deleted": True}


# ── Routes: initial stock ───────────────────────────────────────────────────


class InitialStockUploadResult(BaseModel):
    archetypes_found: int
    total_items: float
    rows_parsed: int


def _check_stock_ext(filename: str) -> None:
    import os
    ext = os.path.splitext(filename)[1].lower()
    if ext not in (".csv", ".xlsx", ".xls"):
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.",
        )


@router.post(
    "/systems/{system_id}/subsystems/{subsystem_id}/stock/upload",
    response_model=InitialStockUploadResult,
)
async def upload_subsystem_initial_stock(
    system_id: str,
    subsystem_id: str,
    file: UploadFile = File(...),
) -> InitialStockUploadResult:
    if subsystem_id == system_id:
        raise HTTPException(
            status_code=400,
            detail="Primary initial stock uploads use /systems/{id}/stock/upload.",
        )
    subs = _sys_subs(system_id)
    sub = subs.get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    filename = file.filename or ""
    _check_stock_ext(filename)
    raw = await file.read()
    try:
        parsed, rows = parse_dependent_stock_file(raw, filename, sub.dimensions)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))

    project = _current_project()
    with _lock:
        # Guard: a well-formed file with ZERO data rows parses clean and would
        # otherwise be written straight over live data. Not safe by accident --
        # the parsers' ValueError is a COLUMN check, which a correct-header /
        # no-rows file sails through. See core/upload_guard.
        refuse_if_nothing_resolved(rows_seen=rows, resolved=len(parsed), what="initial-stock")
        sub.initial_stock = parsed
        # Invalidate cached results — they depended on the old base-year floor.
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return InitialStockUploadResult(
        archetypes_found=len(parsed),
        total_items=float(sum(parsed.values())),
        rows_parsed=rows,
    )


@router.delete(
    "/systems/{system_id}/subsystems/{subsystem_id}/stock",
)
async def clear_subsystem_initial_stock(
    system_id: str, subsystem_id: str
) -> dict[str, bool]:
    if subsystem_id == system_id:
        raise HTTPException(status_code=400, detail="Primary system. Use the DSM API.")
    subs = _sys_subs(system_id)
    sub = subs.get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    project = _current_project()
    with _lock:
        sub.initial_stock = {}
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return {"cleared": True}


@router.post(
    "/systems/{system_id}/subsystems/{subsystem_id}/stock/template",
)
async def template_subsystem_stock(system_id: str, subsystem_id: str) -> Response:
    if subsystem_id == system_id:
        raise HTTPException(status_code=400, detail="Primary template. Use the DSM API.")
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    csv_text = dependent_stock_template_csv(sub.dimensions)
    # Static, human-readable filename — no UUID / system_id / timestamp.
    fname = "initial_stock_template.xlsx"
    return Response(
        content=csv_text,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Routes: manual-mode inflows / outflows ──────────────────────────────────
#
# Manual subsystems simulate independently from their OWN uploaded inflows/
# outflows (same CSV/XLSX schema as a primary system). Stored on the subsystem
# as ``manual_inflows`` / ``manual_outflows`` (cohort → {year → count}).


class ManualFlowUploadResult(BaseModel):
    cohorts_found: int
    rows_parsed: int


def _require_dependent(system_id: str, subsystem_id: str) -> Subsystem:
    if subsystem_id == system_id:
        raise HTTPException(status_code=400, detail="Primary flows use the DSM API.")
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    return sub


def _flows_to_manual(rows: list) -> dict[str, dict[int, float]]:
    """InflowData/OutflowData list (year, counts={cohort:count}) → cohort→{year→count}."""
    manual: dict[str, dict[int, float]] = {}
    for row in rows:
        for cohort, count in (row.counts or {}).items():
            if count:
                manual.setdefault(cohort, {})[int(row.year)] = float(count)
    return manual


@router.post(
    "/systems/{system_id}/subsystems/{subsystem_id}/manual-inflows/upload",
    response_model=ManualFlowUploadResult,
)
async def upload_manual_inflows(
    system_id: str, subsystem_id: str, file: UploadFile = File(...)
) -> ManualFlowUploadResult:
    sub = _require_dependent(system_id, subsystem_id)
    _check_stock_ext(file.filename or "")
    raw = await file.read()
    years = _get_system(system_id).time_horizon.years
    try:
        inflows, rows = parse_inflow_file(raw, file.filename or "", sub.dimensions, years)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    manual = _flows_to_manual(inflows)
    project = _current_project()
    with _lock:
        # Guard: a well-formed file with ZERO data rows parses clean and would
        # otherwise be written straight over live data. Not safe by accident --
        # the parsers' ValueError is a COLUMN check, which a correct-header /
        # no-rows file sails through. See core/upload_guard.
        refuse_if_nothing_resolved(rows_seen=rows, resolved=len(manual), what="manual-inflow")
        sub.manual_inflows = manual
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return ManualFlowUploadResult(cohorts_found=len(manual), rows_parsed=rows)


@router.post(
    "/systems/{system_id}/subsystems/{subsystem_id}/manual-outflows/upload",
    response_model=ManualFlowUploadResult,
)
async def upload_manual_outflows(
    system_id: str, subsystem_id: str, file: UploadFile = File(...)
) -> ManualFlowUploadResult:
    sub = _require_dependent(system_id, subsystem_id)
    _check_stock_ext(file.filename or "")
    raw = await file.read()
    years = _get_system(system_id).time_horizon.years
    try:
        outflows, rows, _ = parse_outflow_file(raw, file.filename or "", sub.dimensions, years)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    manual = _flows_to_manual(outflows)
    project = _current_project()
    with _lock:
        # Guard: a well-formed file with ZERO data rows parses clean and would
        # otherwise be written straight over live data. Not safe by accident --
        # the parsers' ValueError is a COLUMN check, which a correct-header /
        # no-rows file sails through. See core/upload_guard.
        refuse_if_nothing_resolved(rows_seen=rows, resolved=len(manual), what="manual-outflow")
        sub.manual_outflows = manual
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return ManualFlowUploadResult(cohorts_found=len(manual), rows_parsed=rows)


@router.post("/systems/{system_id}/subsystems/{subsystem_id}/manual-inflows/template")
async def template_manual_inflows(system_id: str, subsystem_id: str) -> Response:
    sub = _require_dependent(system_id, subsystem_id)
    years = _get_system(system_id).time_horizon.years
    csv_text = inflow_template_csv(sub.dimensions, years)
    fname = f"inflows_template_{_sanitize_filename(sub.name)}.csv"
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.post("/systems/{system_id}/subsystems/{subsystem_id}/manual-outflows/template")
async def template_manual_outflows(system_id: str, subsystem_id: str) -> Response:
    sub = _require_dependent(system_id, subsystem_id)
    years = _get_system(system_id).time_horizon.years
    csv_text = outflow_template_csv(sub.dimensions, years)
    fname = f"outflows_template_{_sanitize_filename(sub.name)}.csv"
    return Response(content=csv_text, media_type="text/csv",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.delete("/systems/{system_id}/subsystems/{subsystem_id}/manual-inflows")
async def clear_manual_inflows(system_id: str, subsystem_id: str) -> dict[str, bool]:
    sub = _require_dependent(system_id, subsystem_id)
    project = _current_project()
    with _lock:
        sub.manual_inflows = {}
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return {"cleared": True}


@router.delete("/systems/{system_id}/subsystems/{subsystem_id}/manual-outflows")
async def clear_manual_outflows(system_id: str, subsystem_id: str) -> dict[str, bool]:
    sub = _require_dependent(system_id, subsystem_id)
    project = _current_project()
    with _lock:
        sub.manual_outflows = {}
        _sys_sub_results(system_id, project).pop(subsystem_id, None)
    _persist_subs(project, system_id)
    _persist_sub_results(project, system_id)
    return {"cleared": True}


# ── Routes: rule validation ─────────────────────────────────────────────────


class RuleValidationResult(BaseModel):
    ok: bool
    errors: list[str]


@router.post(
    "/systems/{system_id}/subsystems/validate-rule",
    response_model=RuleValidationResult,
)
async def validate_rule(system_id: str, rule: DependencyRule) -> RuleValidationResult:
    primary = _get_system(system_id)
    errors = validate_dependency_rule(rule, primary.dimensions)
    # Also try to parse the expression against the currently active params
    # so users see syntax errors in the UI without waiting for a compute.
    engine = ParameterEngine()
    try:
        engine.resolve(
            rule.expression,
            extra_vars={"filtered_stock": 0.0, "total_primary_stock": 0.0, "year": 0.0},
        )
    except ParameterError as e:
        errors.append(str(e))
    return RuleValidationResult(ok=not errors, errors=errors)


# ── Routes: dependency-rules Excel template + bulk import ────────────────────
#
# Template + Upload for dependency rules, mirroring the Initial stock
# template/upload convention. Scoped to a dependent subsystem via the
# ``subsystem_id`` query param (rules live on a subsystem, not the system).


class DependencyRuleImportError(BaseModel):
    row: int          # 1-based sheet row number (header is row 1)
    field: str
    message: str


def _dep_filter_columns(primary_dims: list) -> list[str]:
    """One ``filter_<dim>`` column per primary non-age dimension (e.g. filter_f)."""
    return [f"filter_{d.name}" for d in non_age_dimensions(primary_dims)]


def _dep_rules_workbook(sub: Subsystem, primary_dims: list):
    """Populated .xlsx: a Rules sheet (headers + example rows) + a locked
    Reference sheet auto-populated from live subsystem/primary/parameter data."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from mapper.core.dsm_engine import all_cohort_keys

    filter_cols = _dep_filter_columns(primary_dims)
    headers = ["rule_number", "dependent_archetype", "description", *filter_cols, "expression"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"
    ws.append(headers)
    hfill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="left", vertical="center")

    def _example(n: int, dep: str, desc: str, f_vals: list[str], expr: str) -> list:
        row: list = [n, dep, desc]
        for i in range(len(filter_cols)):
            row.append(f_vals[i] if i < len(f_vals) else "")
        row.append(expr)
        return row

    # Illustrative example rows (the user replaces them). Values match the
    # reference case study; they document the format, not the exact data.
    ws.append(_example(1, "Residential AC Charger|Default", "1 charger per BEV",
                       ["BEV-LFP,BEV-NCA,BEV-NMC532", "Small,Sedan,SUV"], "filtered_stock"))
    ws.append(_example(2, "Public DC Charger|Default", "fast chargers for large BEVs",
                       ["BEV-NMC622,BEV-NMC811", "SUV"], "filtered_stock * 0.1"))

    widths = [12, 32, 30, *[24] * len(filter_cols), 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Reference sheet — read-only, auto-populated from live data.
    ref = wb.create_sheet("Reference")

    def _section(title: str, values: list[str]) -> None:
        ref.append([title])
        ref.cell(row=ref.max_row, column=1).font = Font(bold=True)
        for v in values:
            ref.append([v])
        ref.append([])

    _section("Valid dependent_archetype values", all_cohort_keys(sub.dimensions))
    for d in non_age_dimensions(primary_dims):
        _section(f"Valid {d.name} values (filter_{d.name}, comma-separated; 'all' = no filter)", list(d.labels))
    pset = get_parameter_set("Base")
    param_names = [p.name for p in pset.parameters] if pset else []
    _section("Expression variables", ["filtered_stock", "total_primary_stock", "year", *param_names])
    ref.column_dimensions["A"].width = 48
    ref.protection.sheet = True  # locked / read-only

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/systems/{system_id}/dependency-rules/template")
async def dependency_rules_template(system_id: str, subsystem_id: str) -> Response:
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    primary = _get_system(system_id)
    data = _dep_rules_workbook(sub, primary.dimensions)
    return excel_response_from_bytes(data, "dependency_rules_template.xlsx", kind="round_trip")


@router.post("/systems/{system_id}/dependency-rules/import")
async def import_dependency_rules(
    system_id: str,
    subsystem_id: str,
    file: UploadFile = File(...),
) -> JSONResponse:
    """Parse + validate a filled template. Returns 422 with per-row/field errors
    if ANY row is invalid (no partial import); 200 with the parsed rules
    otherwise. Does NOT save — the client confirms the destructive replace and
    then calls the existing update endpoint."""
    import io
    import os

    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    primary = _get_system(system_id)

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext != ".xlsx":
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext or '(none)'}'. Upload an .xlsx file "
                   "(export the template first).",
        )
    raw = await file.read()

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}")
    if "Rules" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail='The workbook has no sheet named "Rules".')
    ws = wb["Rules"]

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=400, detail="The Rules sheet is empty.")
    header = [str(h).strip().lower() if h is not None else "" for h in all_rows[0]]
    col = {name: i for i, name in enumerate(header) if name}
    for req in ("dependent_archetype", "expression"):
        if req not in col:
            raise HTTPException(status_code=400, detail=f'The Rules sheet is missing the "{req}" column.')
    filter_cols = {name[len("filter_"):]: i for name, i in col.items() if name.startswith("filter_")}

    from mapper.core.dsm_engine import all_cohort_keys

    valid_archetypes = set(all_cohort_keys(sub.dimensions))

    def _cell(r: tuple, key: str) -> str:
        i = col.get(key)
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()

    parsed: list[DependencyRule] = []
    errors: list[DependencyRuleImportError] = []

    for ridx, r in enumerate(all_rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in r):
            continue  # skip blank rows
        dep = _cell(r, "dependent_archetype")
        expr = _cell(r, "expression")
        desc = _cell(r, "description") or None
        driver_filter: dict[str, list[str]] = {}
        for dim_name, fi in filter_cols.items():
            val = str(r[fi]).strip() if fi < len(r) and r[fi] is not None else ""
            if not val or val.lower() == "all":
                continue
            driver_filter[dim_name] = [v.strip() for v in val.split(",") if v.strip()]

        rule = DependencyRule(
            id="", dependent_archetype_id=dep, driver_filter=driver_filter,
            expression=expr, description=desc,
        )

        if not dep:
            errors.append(DependencyRuleImportError(row=ridx, field="dependent_archetype", message="dependent_archetype is required"))
        elif dep not in valid_archetypes:
            errors.append(DependencyRuleImportError(row=ridx, field="dependent_archetype", message=f"'{dep}' is not a valid archetype in this subsystem"))
        # Reuse the shared validator for driver_filter (dimension/value checks).
        for msg in validate_dependency_rule(rule, primary.dimensions):
            if "driver_filter" in msg:
                errors.append(DependencyRuleImportError(row=ridx, field="filter", message=msg))
        # Reuse the same expression-syntax check as the "Validate" button.
        if not expr:
            errors.append(DependencyRuleImportError(row=ridx, field="expression", message="expression is required"))
        else:
            try:
                ParameterEngine().resolve(
                    expr, extra_vars={"filtered_stock": 0.0, "total_primary_stock": 0.0, "year": 0.0},
                )
            except ParameterError as e:
                errors.append(DependencyRuleImportError(row=ridx, field="expression", message=str(e)))

        parsed.append(rule)

    if not parsed and not errors:
        raise HTTPException(status_code=400, detail="No data rows found in the Rules sheet.")

    if errors:
        return JSONResponse(
            status_code=422,
            content={"ok": False, "errors": [e.model_dump() for e in errors]},
        )
    return JSONResponse(
        status_code=200,
        content={"ok": True, "rules": [r.model_dump() for r in parsed]},
    )


# ── Routes: subsystem cohort mapping template / import ───────────────────────
# Mirrors the primary system's cohort-mapping Template/Upload (bom.py) and the
# sibling dependency-rules Template/Import above: a populated .xlsx template with
# one row per cohort key + a Reference sheet of valid BOM archetypes; a validate-
# only import that rejects the whole file on ANY invalid row (no partial import).
# The client confirms the destructive replace and saves via the existing
# subsystem update endpoint.


_COHORT_MAP_HEADERS = ["dependent_archetype", "bom_archetype", "scale", "color"]


def _cohort_mapping_workbook(sub: Subsystem, archetype_names: list[str]):
    """Populated .xlsx: a Mappings sheet with every cohort key of ``sub`` in the
    ``dependent_archetype`` column (bom_archetype + scale left blank for the user)
    plus a locked Reference sheet listing valid BOM archetype names."""
    import io

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    from mapper.core.dsm_engine import all_cohort_keys

    wb = Workbook()
    ws = wb.active
    ws.title = "Mappings"
    ws.append(_COHORT_MAP_HEADERS)
    hfill = PatternFill(start_color="374151", end_color="374151", fill_type="solid")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hfill
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.freeze_panes = "A2"

    # Pre-populate the dependent_archetype column with the subsystem's cohort
    # keys — the same rows shown in the modal. bom_archetype + scale + color
    # stay blank for the user to fill.
    for ck in all_cohort_keys(sub.dimensions):
        ws.append([ck, "", "", ""])

    for col_letter, width in zip("ABCD", [32, 32, 12, 12]):
        ws.column_dimensions[col_letter].width = width

    ref = wb.create_sheet("Reference")

    def _section(title: str, values: list[str]) -> None:
        ref.append([title])
        ref.cell(row=ref.max_row, column=1).font = Font(bold=True)
        for v in values:
            ref.append([v])
        ref.append([])

    _section("Valid bom_archetype values (leave blank to leave a cohort unmapped)", archetype_names)
    _section("Valid dependent_archetype values", all_cohort_keys(sub.dimensions))
    _section("color (optional)", ["Hex color for this cohort's chart series, e.g. #2dd4bf (#rrggbb).",
                                  "Leave blank to use the app's default color."])
    ref.column_dimensions["A"].width = 48
    ref.protection.sheet = True

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@router.get("/systems/{system_id}/cohort-mapping/template")
async def cohort_mapping_template(system_id: str, subsystem_id: str) -> Response:
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    from mapper.api.bom import _proj_archetypes

    archetype_names = sorted(a.name for a in _proj_archetypes().values())
    data = _cohort_mapping_workbook(sub, archetype_names)
    # Filename: cohort_mapping_<subsystem_name>_template.xlsx — spaces→_, lowered.
    safe = _sanitize_filename(sub.name, "subsystem").lower()
    return excel_response_from_bytes(data, f"cohort_mapping_{safe}_template.xlsx", kind="round_trip")


@router.post("/systems/{system_id}/cohort-mapping/import")
async def import_cohort_mapping(
    system_id: str,
    subsystem_id: str,
    file: UploadFile = File(...),
) -> JSONResponse:
    """Parse + validate a filled cohort-mapping template. Returns 422 with
    per-row/field errors if ANY row is invalid (no partial import); 200 with the
    parsed ``{cohort_key: {archetype_id, scaling_factor}}`` map otherwise. Does
    NOT save — the client confirms the destructive replace and then saves the
    subsystem via the existing update endpoint."""
    import io
    import os

    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")

    filename = file.filename or ""
    ext = os.path.splitext(filename)[1].lower()
    if ext != ".xlsx":
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext or '(none)'}'. Upload an .xlsx file "
                   "(export the template first).",
        )
    raw = await file.read()

    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True)
    except Exception as e:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Could not read the Excel file: {e}")
    if "Mappings" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail='The workbook has no sheet named "Mappings".')
    ws = wb["Mappings"]

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        raise HTTPException(status_code=400, detail="The Mappings sheet is empty.")
    header = [str(h).strip().lower() if h is not None else "" for h in all_rows[0]]
    col = {name: i for i, name in enumerate(header) if name}
    for req in ("dependent_archetype", "bom_archetype"):
        if req not in col:
            raise HTTPException(status_code=400, detail=f'The Mappings sheet is missing the "{req}" column.')

    from mapper.api.bom import _proj_archetypes
    from mapper.core.dsm_engine import all_cohort_keys

    valid_cohorts = set(all_cohort_keys(sub.dimensions))
    archetypes = _proj_archetypes()
    name_to_id: dict[str, str] = {a.name.lower(): aid for aid, a in archetypes.items()}

    def _cell(r: tuple, key: str) -> str:
        i = col.get(key)
        if i is None or i >= len(r) or r[i] is None:
            return ""
        return str(r[i]).strip()

    mappings: dict[str, dict] = {}
    errors: list[DependencyRuleImportError] = []
    seen: set[str] = set()

    for ridx, r in enumerate(all_rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in r):
            continue  # skip blank rows
        dep = _cell(r, "dependent_archetype")
        bom = _cell(r, "bom_archetype")
        scale_raw = _cell(r, "scale")
        color_raw = _cell(r, "color")

        if not dep:
            errors.append(DependencyRuleImportError(row=ridx, field="dependent_archetype", message="dependent_archetype is required"))
            continue
        if dep not in valid_cohorts:
            errors.append(DependencyRuleImportError(row=ridx, field="dependent_archetype", message=f"'{dep}' is not a valid cohort key for this subsystem"))
        if dep in seen:
            errors.append(DependencyRuleImportError(row=ridx, field="dependent_archetype", message=f"Duplicate row for '{dep}'"))
        seen.add(dep)

        # bom_archetype: blank => unmapped (valid). Otherwise must resolve.
        arc_id: str | None = None
        if bom:
            arc_id = name_to_id.get(bom.lower())
            if arc_id is None and bom in archetypes:
                arc_id = bom  # accept raw id too
            if arc_id is None:
                errors.append(DependencyRuleImportError(row=ridx, field="bom_archetype", message=f"'{bom}' is not a known BOM archetype in this project"))

        # scale: default 1.0 when absent; must be a positive number otherwise.
        scale = 1.0
        if scale_raw:
            try:
                scale = float(scale_raw)
            except (TypeError, ValueError):
                errors.append(DependencyRuleImportError(row=ridx, field="scale", message=f"'{scale_raw}' is not a number"))
                scale = 1.0
            else:
                if scale <= 0:
                    errors.append(DependencyRuleImportError(row=ridx, field="scale", message="scale must be a positive number"))

        # color: blank/absent → None (app default). Otherwise must be valid hex.
        from mapper.api.bom import _normalize_color
        color, color_err = _normalize_color(color_raw)
        if color_err:
            errors.append(DependencyRuleImportError(row=ridx, field="color", message=f"'{color_raw}' is not a valid hex color (e.g. #2dd4bf)"))

        # Only mapped rows contribute to the saved mapping (unmapped = omitted).
        if arc_id is not None:
            mappings[dep] = {"archetype_id": arc_id, "scaling_factor": scale, "color": color}

    if not seen and not errors:
        raise HTTPException(status_code=400, detail="No data rows found in the Mappings sheet.")

    if errors:
        return JSONResponse(
            status_code=422,
            content={"ok": False, "errors": [e.model_dump() for e in errors]},
        )
    return JSONResponse(
        status_code=200,
        content={"ok": True, "mappings": mappings},
    )


# ── Routes: compute ─────────────────────────────────────────────────────────


class ComputeRequest(BaseModel):
    parameter_set_id: str | None = None


class ComputeAllResponse(BaseModel):
    subsystem_results: dict[str, SimulationResult]


def _require_primary_result(system_id: str) -> SimulationResult:
    result = _proj_results().get(system_id)
    if result is None:
        raise HTTPException(
            status_code=400,
            detail="No primary simulation yet. Run /dsm/systems/{id}/simulate first.",
        )
    return result


def _table_kwargs(parameter_set_id: str | None) -> dict:
    """The RAW table + scenario for ``compute_subsystem_result``.

    ``_build_engine`` below returns a PRE-RESOLVED engine: it goes through
    ``get_parameter_set``, which calls ``table.resolve_all(scenario)`` with no
    year, so a keyframed parameter is already flattened to its ``base_value``
    before the engine exists. Handing the raw table alongside lets the
    dependency-rule loop re-resolve per simulation year.

    Never raises -- an unresolvable table simply means no year-varying
    resolution, which is the pre-existing behaviour.
    """
    try:
        return {
            "parameter_table": _table_for(),
            "parameter_scenario": parameter_set_id or None,
        }
    except Exception:
        return {}


def _build_engine(parameter_set_id: str | None) -> ParameterEngine:
    if not parameter_set_id:
        return ParameterEngine()
    pset = get_parameter_set(parameter_set_id)
    if pset is None:
        raise HTTPException(
            status_code=404,
            detail=f"Parameter set '{parameter_set_id}' not found",
        )
    return ParameterEngine(pset.parameters)


@router.post(
    "/systems/{system_id}/subsystems/{subsystem_id}/compute",
    response_model=SimulationResult,
)
async def compute_subsystem(
    system_id: str,
    subsystem_id: str,
    body: ComputeRequest | None = None,
) -> SimulationResult:
    if subsystem_id == system_id:
        # Primary is computed via the DSM /simulate endpoint.
        return _require_primary_result(system_id)
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    primary_def = _get_system(system_id)
    engine = _build_engine(body.parameter_set_id if body else None)
    # Manual-mode subsystems simulate independently — no primary sim required.
    primary_result = None if sub.mode == "manual" else _require_primary_result(system_id)
    try:
        result = compute_subsystem_result(
            sub, primary_def, primary_result, engine,
            **_table_kwargs(body.parameter_set_id if body else None),
        )
    except ParameterError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    project = _current_project()
    with _lock:
        _sys_sub_results(system_id, project)[subsystem_id] = result
    _persist_sub_results(project, system_id)
    return result


@router.post(
    "/systems/{system_id}/subsystems/compute-all", response_model=ComputeAllResponse
)
async def compute_all_subsystems(
    system_id: str, body: ComputeRequest | None = None
) -> ComputeAllResponse:
    """Run every dependent subsystem against the current primary simulation.

    Returns a map of ``{subsystem_id → SimulationResult}``. Errors in one
    subsystem are surfaced with its id in the message but abort the batch.
    """
    primary_def = _get_system(system_id)
    primary_result = _require_primary_result(system_id)
    subs = _sys_subs(system_id)
    engine = _build_engine(body.parameter_set_id if body else None)
    results: dict[str, SimulationResult] = {}
    for sub_id, sub in subs.items():
        try:
            results[sub_id] = compute_subsystem_result(
                sub, primary_def, primary_result, engine,
                **_table_kwargs(body.parameter_set_id if body else None),
            )
        except ParameterError as e:
            raise HTTPException(
                status_code=400,
                detail=f"Subsystem '{sub.name}' ({sub_id}): {e}",
            )
        except ValueError as e:
            raise HTTPException(status_code=400, detail=str(e))
    project = _current_project()
    with _lock:
        _sys_sub_results(system_id, project).clear()
        _sys_sub_results(system_id, project).update(results)
    _persist_sub_results(project, system_id)
    return ComputeAllResponse(subsystem_results=results)


@router.get(
    "/systems/{system_id}/subsystems/{subsystem_id}/results",
    response_model=SimulationResult,
)
async def get_subsystem_result(system_id: str, subsystem_id: str) -> SimulationResult:
    if subsystem_id == system_id:
        return _require_primary_result(system_id)
    res = _sys_sub_results(system_id).get(subsystem_id)
    if res is None:
        raise HTTPException(
            status_code=404,
            detail="No results yet. Run /compute for this subsystem first.",
        )
    return res


# ── Subsystem DSM Excel export ──────────────────────────────────────────────
#
# The subsystem's compute payload is a SimulationResult (stock + flows), same
# shape as a primary system — NOT impact. Subsystem IMPACT already flows into
# the Impact Assessment exports' "By subsystem" sheet, so this is deliberately
# the DSM (stock/flow) side and must not duplicate that. The primary DSM export
# (dsm._build_export_workbook) is scenario/dimension-centric with a different
# sheet set (Age Distribution, Mass Balance, Scaling Rules, …); the subsystem
# export is archetype-centric with its own sheets (Cohort mappings, Dependency
# rules, aggregate flows), so it's a focused builder here — but it REUSES the
# shared build_export_filename helper and the same UUID-strip + archetype-name
# resolution convention (no raw <uuid>:: in any cell).


def _year_range(result) -> tuple[int | None, int | None]:
    ys = [yr.year for yr in result.years]
    return (min(ys), max(ys)) if ys else (None, None)


def _build_subsystem_dsm_workbook(
    *,
    scope: str,
    primary_def,
    primary_result,
    primary_initial_stock: dict,
    primary_cohort_mapping,
    subsystem: Subsystem,
    subsystem_result: SimulationResult,
    archetypes: dict,
) -> bytes:
    """Build the subsystem DSM workbook. ``scope`` ∈ {"combined", "subsystem"}.

    combined → every data sheet carries a leading "System" column distinguishing
    primary rows from subsystem rows (the By-cohort export convention). subsystem
    → subsystem rows only, no System column.
    """
    import io as _io

    from openpyxl import Workbook

    from mapper.api.cohort_export import INT_FMT, autosize, style_header

    combined = scope == "combined"

    # Defensive UUID-prefix strip (a subsystem's OWN result keys are unprefixed;
    # dependent_archetype_id is a user-selected readable name). Same convention
    # as cohort_export._split_cohort — no raw <uuid>:: in any cell.
    def _disp(ck: str) -> str:
        return ck.split("::", 1)[-1]

    sub_map = subsystem.cohort_mappings or {}

    def sub_arc_name(ck: str) -> str:
        m = sub_map.get(ck)
        arc = archetypes.get(m.archetype_id) if m else None
        return getattr(arc, "name", "") if arc else ""

    prim_map: dict[str, tuple[str, float]] = {}
    if primary_cohort_mapping is not None:
        for e in primary_cohort_mapping.mappings:
            prim_map[e.cohort_key] = (e.archetype_id, e.scaling_factor)

    def prim_arc_name(ck: str) -> str:
        aid = prim_map.get(ck, ("", 1.0))[0]
        arc = archetypes.get(aid)
        return getattr(arc, "name", "") if arc else ""

    prim_name = primary_def.name
    sub_name = subsystem.name
    wb = Workbook()

    # ── Summary ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    y0, y1 = _year_range(subsystem_result)
    sub_cohorts = sorted({ck for yr in subsystem_result.years for ck in yr.stock})
    dim_str = ", ".join(
        d.display_name or d.name for d in non_age_dimensions(subsystem.dimensions)
    ) or "—"
    rows: list[tuple] = [("Export scope", "Main system + subsystem" if combined else "Subsystem only")]
    if combined:
        prim_cohorts = sorted({ck for yr in primary_result.years for ck in yr.stock}) if primary_result else []
        rows += [("Primary system", prim_name), ("Primary cohort count", len(prim_cohorts))]
    rows += [
        ("Subsystem", sub_name),
        ("Dimensions", dim_str),
        ("Mode", subsystem.mode),
        ("Year range", f"{y0}–{y1}" if y0 is not None else "—"),
        ("Cohort count", len(sub_cohorts)),
        ("Unit", subsystem.unit_name),
    ]
    ws.append(["Field", "Value"])
    style_header(ws)
    for r in rows:
        ws.append(list(r))
    autosize(ws)

    # ── Stock over time ──────────────────────────────────────────────────
    ws = wb.create_sheet("Stock over time")
    header = (["System"] if combined else []) + ["Year", "Cohort", "Archetype", "Count"]
    ws.append(header)
    style_header(ws)

    def _stock_rows(result, sysname, arc_fn):
        for yr in result.years:
            for ck in sorted(yr.stock):
                ws.append(([sysname] if combined else []) + [yr.year, _disp(ck), arc_fn(ck) or "—", yr.stock[ck]])

    if combined and primary_result:
        _stock_rows(primary_result, prim_name, prim_arc_name)
    _stock_rows(subsystem_result, sub_name, sub_arc_name)
    for row in ws.iter_rows(min_row=2, min_col=len(header), max_col=len(header)):
        for c in row:
            c.number_format = INT_FMT
    autosize(ws)

    # ── Inflows and outflows (aggregate — matches the two-series chart) ───
    ws = wb.create_sheet("Inflows and outflows")
    header = (["System"] if combined else []) + ["Year", "Total inflow", "Total outflow"]
    ws.append(header)
    style_header(ws)

    def _flow_rows(result, sysname):
        for yr in result.years:
            ws.append(([sysname] if combined else []) + [yr.year, sum(yr.inflow.values()), sum(yr.outflow.values())])

    if combined and primary_result:
        _flow_rows(primary_result, prim_name)
    _flow_rows(subsystem_result, sub_name)
    first_num = (2 if combined else 1) + 1
    for row in ws.iter_rows(min_row=2, min_col=first_num, max_col=len(header)):
        for c in row:
            c.number_format = INT_FMT
    autosize(ws)

    # ── Initial stock (base-year floor) ──────────────────────────────────
    ws = wb.create_sheet("Initial stock")
    header = (["System"] if combined else []) + ["Cohort", "Count"]
    ws.append(header)
    style_header(ws)
    if combined:
        for ck, v in sorted((primary_initial_stock or {}).items()):
            ws.append([prim_name, _disp(ck), v])
    for ck, v in sorted((subsystem.initial_stock or {}).items()):
        ws.append(([sub_name] if combined else []) + [_disp(ck), v])
    for row in ws.iter_rows(min_row=2, min_col=len(header), max_col=len(header)):
        for c in row:
            c.number_format = INT_FMT
    autosize(ws)

    # ── Cohort mappings (incl. colour) ───────────────────────────────────
    ws = wb.create_sheet("Cohort mappings")
    header = (["System"] if combined else []) + ["Cohort", "Archetype", "Scale", "Color"]
    ws.append(header)
    style_header(ws)
    if combined and primary_cohort_mapping is not None:
        row_colors = getattr(primary_cohort_mapping, "row_colors", {}) or {}
        for e in primary_cohort_mapping.mappings:
            ws.append([prim_name, _disp(e.cohort_key), prim_arc_name(e.cohort_key) or "—",
                       e.scaling_factor, row_colors.get(e.cohort_key, "") or ""])
    for ck, m in sorted(sub_map.items()):
        ws.append(([sub_name] if combined else []) + [_disp(ck), sub_arc_name(ck) or "—",
                   m.scaling_factor, m.color or ""])
    autosize(ws)

    # ── Dependency rules (rules mode only) ───────────────────────────────
    if subsystem.mode == "rules" and subsystem.dependency_rules:
        ws = wb.create_sheet("Dependency rules")
        ws.append(["Dependent archetype", "Driver filter", "Expression", "Description"])
        style_header(ws)
        for r in subsystem.dependency_rules:
            df = "; ".join(f"{k}={','.join(v)}" for k, v in (r.driver_filter or {}).items())
            ws.append([_disp(r.dependent_archetype_id), df, r.expression, r.description or ""])
        autosize(ws)

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.get("/systems/{system_id}/subsystems/{subsystem_id}/export")
async def export_subsystem_results(
    system_id: str, subsystem_id: str, scope: str = "subsystem"
) -> Response:
    """Excel export of a dependent subsystem's DSM (stock/flow) results.

    ``scope=subsystem`` → subsystem is the subject (Fueling_Infrastructure_DSM.xlsx);
    ``scope=combined``  → main system + subsystem (Car_Fleet+Fueling_Infrastructure_DSM.xlsx).
    """
    if scope not in ("combined", "subsystem"):
        raise HTTPException(status_code=400, detail="scope must be 'combined' or 'subsystem'")
    sub = _sys_subs(system_id).get(subsystem_id)
    if sub is None:
        raise HTTPException(status_code=404, detail="Subsystem not found")
    project = _current_project()
    sub_result = _sys_sub_results(system_id, project).get(subsystem_id)
    if sub_result is None:
        raise HTTPException(
            status_code=400,
            detail="No subsystem results. Compute the subsystem first.",
        )
    primary_def = _get_system(system_id)

    from mapper.api.bom import (
        _proj_archetypes,
        _proj_cohort_mappings,
        build_export_filename,
    )
    from mapper.api.dsm import _get_or_create_state, _materialized

    archetypes = _proj_archetypes(project)
    primary_cohort_mapping = _proj_cohort_mappings(project).get(system_id)
    primary_result = None
    primary_initial_stock: dict = {}
    if scope == "combined":
        primary_result = _proj_results(project).get(system_id)
        if primary_result is None:
            raise HTTPException(
                status_code=400,
                detail=(
                    "No primary simulation available for the combined export. Run the "
                    "main system's simulation first, or export the subsystem only."
                ),
            )
        primary_initial_stock = dict(_materialized(_get_or_create_state(system_id)).initial_stock or {})

    content = _build_subsystem_dsm_workbook(
        scope=scope,
        primary_def=primary_def,
        primary_result=primary_result,
        primary_initial_stock=primary_initial_stock,
        primary_cohort_mapping=primary_cohort_mapping,
        subsystem=sub,
        subsystem_result=sub_result,
        archetypes=archetypes,
    )
    # Shared filename scheme, domain "DSM". Combined → primary is the subject
    # with the subsystem appended; subsystem-only → the subsystem IS the subject
    # (expressible via the existing helper: name + no subsystems).
    if scope == "combined":
        fname = build_export_filename(primary_def.name, [sub.name], "DSM")
    else:
        fname = build_export_filename(sub.name, [], "DSM")

    return excel_response_from_bytes(content, fname, kind="data")


# ── Public accessors (used by DSM-LCA aggregation) ──────────────────────────


def get_subsystems_for_system(
    system_id: str, project: str | None = None
) -> dict[str, Subsystem]:
    """Return the stored dependent subsystems for ``system_id``."""
    return dict(_sys_subs(system_id, project))


def get_subsystem_results_for_system(
    system_id: str, project: str | None = None
) -> dict[str, SimulationResult]:
    """Return cached dependent-subsystem simulation results."""
    return dict(_sys_sub_results(system_id, project))
