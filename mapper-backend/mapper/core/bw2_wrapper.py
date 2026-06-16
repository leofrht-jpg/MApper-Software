import ast
import datetime
import io
import logging
import shutil
import tarfile
import tempfile
from pathlib import Path
from typing import Any

import bw2analyzer
import bw2calc
import bw2data

logger = logging.getLogger(__name__)


# ── Phase 0 ──────────────────────────────────────────────────────────────────

def get_current_project() -> str:
    return bw2data.projects.current


def list_projects() -> list[dict]:
    current = get_current_project()
    return [
        {"name": p.name, "is_current": p.name == current}
        for p in bw2data.projects
    ]


def switch_project(name: str) -> str:
    bw2data.projects.set_current(name)
    return get_current_project()


def create_project(name: str) -> str:
    """Create a new empty brightway2 project and switch to it."""
    name = (name or "").strip()
    if not name:
        raise ValueError("Project name is required")
    existing = {p.name for p in bw2data.projects}
    if name in existing:
        raise ValueError(f"Project '{name}' already exists")
    bw2data.projects.set_current(name)
    return name


def duplicate_project(source: str, new_name: str) -> str:
    """Copy ``source`` project into ``new_name`` and leave ``new_name`` active."""
    new_name = (new_name or "").strip()
    if not new_name:
        raise ValueError("New project name is required")
    existing = {p.name for p in bw2data.projects}
    if source not in existing:
        raise ValueError(f"Source project '{source}' does not exist")
    if new_name in existing:
        raise ValueError(f"Project '{new_name}' already exists")
    bw2data.projects.set_current(source)
    bw2data.projects.copy_project(new_name, switch=True)
    return new_name


def delete_project(name: str) -> str:
    """Delete ``name``. Refuses to delete the only remaining project.

    Returns the currently active project name afterwards.
    """
    projects = [p.name for p in bw2data.projects]
    if name not in projects:
        raise ValueError(f"Project '{name}' does not exist")
    if len(projects) <= 1:
        raise ValueError("Cannot delete the last remaining project")
    current = get_current_project()
    if current == name:
        fallback = next(p for p in projects if p != name)
        bw2data.projects.set_current(fallback)
    bw2data.projects.delete_project(name, delete_dir=True)
    return get_current_project()


def export_project(name: str) -> bytes:
    """Return a tar.gz of the project's data directory."""
    projects = [p.name for p in bw2data.projects]
    if name not in projects:
        raise ValueError(f"Project '{name}' does not exist")
    original = get_current_project()
    try:
        bw2data.projects.set_current(name)
        src = Path(bw2data.projects.dir)
        buf = io.BytesIO()
        with tarfile.open(fileobj=buf, mode="w:gz") as tf:
            tf.add(str(src), arcname=name)
        return buf.getvalue()
    finally:
        if original != name and original in {p.name for p in bw2data.projects}:
            bw2data.projects.set_current(original)


def import_project(data: bytes) -> str:
    """Create a new project from an exported tar.gz, disambiguating the name
    if a project with the same name already exists."""
    buf = io.BytesIO(data)
    with tempfile.TemporaryDirectory() as tmp:
        try:
            with tarfile.open(fileobj=buf, mode="r:gz") as tf:
                tf.extractall(tmp)
        except tarfile.TarError as e:
            raise ValueError(f"Invalid project archive: {e}")
        roots = [p for p in Path(tmp).iterdir() if p.is_dir()]
        if not roots:
            raise ValueError("Archive does not contain a project folder")
        src = roots[0]
        base_name = src.name
        existing = {p.name for p in bw2data.projects}
        new_name = base_name
        i = 2
        while new_name in existing:
            new_name = f"{base_name} ({i})"
            i += 1
        original = get_current_project()
        # Initialize an empty project directory we can overwrite.
        bw2data.projects.set_current(new_name)
        target = Path(bw2data.projects.dir)
        # Switch away so SQLite handles on the new project are released.
        if original in {p.name for p in bw2data.projects} and original != new_name:
            bw2data.projects.set_current(original)
        for item in src.iterdir():
            dest = target / item.name
            if item.is_dir():
                if dest.exists():
                    shutil.rmtree(dest)
                shutil.copytree(item, dest)
            else:
                if dest.exists():
                    dest.unlink()
                shutil.copy2(item, dest)
        bw2data.projects.set_current(new_name)
        return new_name


def list_databases() -> list[dict]:
    from mapper.core import plca_storage

    try:
        project = get_current_project()
    except Exception:
        project = ""
    registry = {e.get("name"): e for e in plca_storage.load_registry(project)} if project else {}

    results = []
    for name in bw2data.databases:
        meta = bw2data.databases[name]
        try:
            records = len(bw2data.Database(name))
        except Exception:
            records = 0
        modified = meta.get("modified", None)
        pmeta = registry.get(name)
        results.append({
            "name": name,
            "records": records,
            "modified": modified,
            "is_prospective": pmeta is not None,
            "prospective_meta": pmeta,
        })
    return results


# ── Phase 1A: Activities ──────────────────────────────────────────────────────

def _activity_to_summary(act) -> dict:
    return {
        "key": str(act.key),
        "code": act.get("code", ""),
        "name": act.get("name", ""),
        "location": str(act.get("location", "")),
        "unit": act.get("unit", ""),
        "product": act.get("reference product", act.get("name", "")),
        "database": act.get("database", ""),
    }


_SORT_KEYS = {
    "name_asc":     (lambda a: (a.get("name", "") or "").lower(), False),
    "name_desc":    (lambda a: (a.get("name", "") or "").lower(), True),
    "location_asc": (lambda a: (str(a.get("location", "")) or "").lower(), False),
    "unit_asc":     (lambda a: (a.get("unit", "") or "").lower(), False),
}


def get_activities(
    database_name: str,
    offset: int = 0,
    limit: int = 50,
    search: str | None = None,
    locations: list[str] | None = None,
    units: list[str] | None = None,
    sort_by: str = "name_asc",
) -> tuple[list[dict], int]:
    db = bw2data.Database(database_name)
    q = search.lower().strip() if search else ""
    loc_set = {l for l in (locations or []) if l}
    unit_set = {u for u in (units or []) if u}

    def _matches(act) -> bool:
        if q and not (
            q in (act.get("name", "") or "").lower()
            or q in str(act.get("location", "") or "").lower()
            or q in (act.get("reference product", "") or "").lower()
        ):
            return False
        if loc_set and str(act.get("location", "") or "") not in loc_set:
            return False
        if unit_set and (act.get("unit", "") or "") not in unit_set:
            return False
        return True

    matches = [a for a in db if _matches(a)]

    # Sorting. "relevance" keeps db-order when there's a search query (simple
    # substring match — no fancy ranking), otherwise falls back to name_asc.
    if sort_by == "relevance":
        if not q:
            sort_by = "name_asc"
    if sort_by in _SORT_KEYS:
        keyfn, reverse = _SORT_KEYS[sort_by]
        matches.sort(key=keyfn, reverse=reverse)

    total = len(matches)
    page = matches[offset : offset + limit]
    return [_activity_to_summary(a) for a in page], total


def search_all_activities(
    search: str,
    limit: int = 50,
    technosphere_only: bool = False,
) -> list[dict]:
    """Search across ALL databases in the current project. Returns up to *limit* results.

    Uses ``Database.search()`` when available (Whoosh-indexed — sub-100 ms
    even for 23 k activities).  Falls back to linear scan if search() fails.

    When *technosphere_only* is True, biosphere databases are skipped
    (biosphere flows cannot be used as LCA functional units).
    """
    q = search.strip()
    if not q:
        return []

    results: list[dict] = []
    for db_name in bw2data.databases:
        if technosphere_only and "biosphere" in db_name.lower():
            continue
        db = bw2data.Database(db_name)
        remaining = limit - len(results)
        if remaining <= 0:
            break
        try:
            # Whoosh-indexed search — fast
            hits = db.search(q, limit=remaining)
            for act in hits:
                results.append(_activity_to_summary(act))
                if len(results) >= limit:
                    return results
        except Exception:
            # Fallback: linear scan
            ql = q.lower()
            for act in db:
                name = (act.get("name", "") or "").lower()
                product = (act.get("reference product", "") or "").lower()
                location = str(act.get("location", "") or "").lower()
                if ql in name or ql in product or ql in location:
                    results.append(_activity_to_summary(act))
                    if len(results) >= limit:
                        return results
    return results


# Cache: (db_name, signature) -> {locations, units}. Signature is derived from
# the bw2 database's record count + last-modified timestamp so the cache
# invalidates automatically after imports/updates.
_distinct_cache: dict[tuple[str, str], dict[str, list[str]]] = {}


def get_distinct_values(database_name: str) -> dict[str, list[str]]:
    md = bw2data.databases.get(database_name, {}) or {}
    sig = f"{md.get('number', 0)}|{md.get('modified', '')}"
    key = (database_name, sig)
    cached = _distinct_cache.get(key)
    if cached is not None:
        return cached
    db = bw2data.Database(database_name)
    locs: set[str] = set()
    units: set[str] = set()
    for a in db:
        loc = str(a.get("location", "") or "")
        u = a.get("unit", "") or ""
        if loc:
            locs.add(loc)
        if u:
            units.add(u)
    result = {
        "locations": sorted(locs, key=str.lower),
        "units": sorted(units, key=str.lower),
    }
    _distinct_cache[key] = result
    return result


def get_activity_detail(database_name: str, code: str) -> dict:
    act = bw2data.get_activity((database_name, code))
    exchanges = []
    for exc in act.exchanges():
        try:
            inp = bw2data.get_activity(exc.input.key)
            exchanges.append({
                "input_key": str(exc.input.key),
                "input_name": inp.get("name", ""),
                "input_location": str(inp.get("location", "")),
                "input_unit": inp.get("unit", ""),
                "input_database": inp.get("database", ""),
                "amount": float(exc.get("amount", 0)),
                "type": exc.get("type", "technosphere"),
            })
        except Exception:
            continue

    metadata = {
        k: str(v)
        for k, v in act.items()
        if k not in {"name", "location", "unit", "reference product", "database", "code"}
        and not k.startswith("_")
    }

    return {
        **_activity_to_summary(act),
        "exchanges": exchanges,
        "metadata": metadata,
    }


# ── Batch export details (for CSV/xlsx selection export) ─────────────────────

_EXPORT_COLUMNS = [
    "database", "code", "name", "reference_product", "location", "unit",
    "classifications", "comment", "production_amount",
    "technosphere_count", "biosphere_count", "activity_type",
]


def _format_classifications(raw: Any) -> str:
    if not raw:
        return ""
    if isinstance(raw, list):
        parts = []
        for item in raw:
            if isinstance(item, (list, tuple)) and len(item) >= 2:
                parts.append(f"{item[0]}: {item[1]}")
            else:
                parts.append(str(item))
        return "; ".join(parts)
    return str(raw)


def _activity_export_row(act: Any) -> dict:
    d = act.as_dict() if hasattr(act, "as_dict") else dict(act)
    techno = 0
    bio = 0
    for exc in act.exchanges():
        t = exc.get("type", "")
        if t == "technosphere":
            techno += 1
        elif t == "biosphere":
            bio += 1
    try:
        prod_amount = float(d.get("production amount", 0) or 0)
    except (TypeError, ValueError):
        prod_amount = 0.0
    return {
        "database": d.get("database", "") or "",
        "code": d.get("code", "") or "",
        "name": d.get("name", "") or "",
        "reference_product": d.get("reference product", "") or "",
        "location": str(d.get("location", "") or ""),
        "unit": d.get("unit", "") or "",
        "classifications": _format_classifications(d.get("classifications")),
        "comment": d.get("comment", "") or "",
        "production_amount": prod_amount,
        "technosphere_count": techno,
        "biosphere_count": bio,
        "activity_type": d.get("activity type", "") or "",
    }


def get_activities_export_details(database_name: str, codes: list[str]) -> list[dict]:
    rows = []
    for code in codes:
        try:
            act = bw2data.get_activity((database_name, code))
            rows.append(_activity_export_row(act))
        except Exception:
            continue
    return rows


def build_selection_csv(rows: list[dict]) -> bytes:
    import csv
    import re

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\n")
    writer.writerow(_EXPORT_COLUMNS)
    for r in rows:
        comment = str(r.get("comment", "") or "")
        comment = re.sub(r"\s+", " ", comment).strip()
        if len(comment) > 500:
            comment = comment[:497] + "..."
        writer.writerow([
            r.get("database", ""),
            r.get("code", ""),
            r.get("name", ""),
            r.get("reference_product", ""),
            r.get("location", ""),
            r.get("unit", ""),
            r.get("classifications", ""),
            comment,
            r.get("production_amount", ""),
            r.get("technosphere_count", ""),
            r.get("biosphere_count", ""),
            r.get("activity_type", ""),
        ])
    return buf.getvalue().encode("utf-8")


def build_selection_xlsx(rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Activities"

    ws.append(_EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    comment_col = _EXPORT_COLUMNS.index("comment") + 1

    for r in rows:
        ws.append([r.get(h, "") for h in _EXPORT_COLUMNS])

    # Column widths: auto-estimate capped at 40, with explicit widths for wide text cols.
    for col_idx, header in enumerate(_EXPORT_COLUMNS, start=1):
        letter = get_column_letter(col_idx)
        if header == "comment":
            ws.column_dimensions[letter].width = 60
        elif header == "classifications":
            ws.column_dimensions[letter].width = 50
        elif header == "name" or header == "reference_product":
            ws.column_dimensions[letter].width = 40
        else:
            max_len = len(header)
            for r in rows:
                v = str(r.get(header, "") or "")
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[letter].width = min(max_len + 2, 30)

    # Wrap text on the comment column so long descriptions stay readable.
    for row in ws.iter_rows(min_row=2, min_col=comment_col, max_col=comment_col):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_methods() -> list[dict]:
    families: dict[str, dict[str, list[dict]]] = {}
    for method_tuple in bw2data.methods:
        if len(method_tuple) < 3:
            continue
        family, category, indicator = method_tuple[0], method_tuple[1], method_tuple[2]
        families.setdefault(family, {}).setdefault(category, []).append(
            {"indicator": indicator, "tuple": list(method_tuple)}
        )

    result = []
    for family, cats in sorted(families.items()):
        result.append({
            "family": family,
            "categories": [
                {"category": cat, "indicators": sorted(inds, key=lambda x: x["indicator"])}
                for cat, inds in sorted(cats.items())
            ],
        })
    return result


# ── Phase 1C: LCA ─────────────────────────────────────────────────────────────

def parse_activity_key(key_str: str) -> tuple[str, str]:
    """Parse a key string like \"('db', 'code')\" into a (db, code) tuple."""
    try:
        parsed = ast.literal_eval(key_str)
        return tuple(parsed)
    except Exception:
        raise ValueError(f"Cannot parse activity key: {key_str!r}")


def run_lca(functional_unit_key: str, amount: float, method_tuple: list[str]) -> dict[str, Any]:
    key = parse_activity_key(functional_unit_key)
    method = tuple(method_tuple)

    lca = bw2calc.LCA({key: amount}, method)
    lca.lci()
    lca.lcia()

    act = bw2data.get_activity(key)
    unit = bw2data.methods[method].get("unit", "unknown") if method in bw2data.methods else "unknown"

    return {
        "score": float(lca.score),
        "unit": unit,
        "functional_unit_name": act.get("reference product", act.get("name", "")),
        "lca_object": lca,
        "activity_key": key,
    }


def run_lca_from_demand(
    demand: dict[tuple[str, str], float], method_tuple: tuple
) -> tuple[float, str]:
    """Run an LCA from a pre-built demand dict (key=(database, code) → amount).

    Used by the DSM × LCA pipeline to do one LCA call per year over an
    aggregated material demand vector.
    """
    if not demand:
        return 0.0, ""
    method = tuple(method_tuple)
    fu = {bw2data.get_activity(k): float(v) for k, v in demand.items()}
    lca = bw2calc.LCA(fu, method)
    lca.lci()
    lca.lcia()
    unit = (
        bw2data.methods[method].get("unit", "unknown")
        if method in bw2data.methods
        else "unknown"
    )
    return float(lca.score), unit


def _method_unit(method: tuple) -> str:
    if method in bw2data.methods:
        return bw2data.methods[method].get("unit", "unknown")
    return "unknown"


def run_lca_multi_method(
    demand: dict[tuple[str, str], float], method_tuples: list[tuple]
) -> dict[tuple, tuple[float, str]]:
    """Compute LCIA scores for many methods against the same demand.

    Solves the technosphere matrix **once** (via ``lca.lci()``) and then
    switches the characterisation matrix for each additional method — an order
    of magnitude faster than rebuilding per method. Falls back to fresh LCAs
    per method if ``switch_method`` is unavailable in the installed bw2calc.
    """
    if not method_tuples:
        return {}
    unique: list[tuple] = []
    seen: set[tuple] = set()
    for m in method_tuples:
        mt = tuple(m)
        if mt in seen:
            continue
        seen.add(mt)
        unique.append(mt)
    if not demand:
        return {m: (0.0, _method_unit(m)) for m in unique}

    fu = {bw2data.get_activity(k): float(v) for k, v in demand.items()}
    first = unique[0]
    lca = bw2calc.LCA(fu, first)
    lca.lci()
    lca.lcia()
    out: dict[tuple, tuple[float, str]] = {first: (float(lca.score), _method_unit(first))}

    switch = getattr(lca, "switch_method", None)
    for mt in unique[1:]:
        if callable(switch):
            try:
                switch(mt)
                lca.lcia()
                out[mt] = (float(lca.score), _method_unit(mt))
                continue
            except Exception:
                pass  # fall through to rebuild
        # Fallback: full rebuild for this method.
        lca2 = bw2calc.LCA(fu, mt)
        lca2.lci()
        lca2.lcia()
        out[mt] = (float(lca2.score), _method_unit(mt))
    return out


def _patch_umfpack_import() -> bool:
    """Fix broken ``scikits.umfpack`` import on numpy ≥ 1.25.

    ``scikits.umfpack`` tries ``from numpy.testing import Tester`` which was
    removed.  We inject a no-op stub so the C extension can still load.
    Returns ``True`` if UMFPACK is usable after patching.
    """
    import numpy.testing as _nt

    if not hasattr(_nt, "Tester"):

        class _DummyTester:  # pragma: no cover
            def __init__(self, *a: Any, **kw: Any) -> None:
                pass

            def test(self, *a: Any, **kw: Any) -> None:
                pass

        _nt.Tester = _DummyTester  # type: ignore[attr-defined]

    try:
        from scikits.umfpack import UMFPACK_A, UmfpackContext  # noqa: F401

        return True
    except Exception:
        return False


# Module-level flag: try once, cache the result.
_UMFPACK_OK = _patch_umfpack_import()


class PersistentLCARunner:
    """Reusable multi-method LCA runner that caches the LU factorization.

    On the first call the technosphere matrix is loaded and factorized via
    UMFPACK (≈ 1.6 s for ecoinvent 3.10, 23 k × 23 k).  Every subsequent
    call reuses that factorization — only back-substitution (≈ 0.015 s) and
    characterization (≈ 0.001 s) are repeated.

    Performance comparison for 26 years × 3 scopes × 8 indicators:

        Before:  78 × spsolve       ≈ 78 × 1.7 s ≈ 133 s
        After :  1 × UMFPACK factor + 77 × back-sub ≈ 1.6 + 77 × 0.015 ≈ 3 s

    Falls back to ``spsolve``-per-call if UMFPACK is unavailable.

    Thread safety: **not** thread-safe.  Create one instance per task.
    """

    def __init__(self) -> None:
        self._lca: bw2calc.LCA | None = None
        self._umf_solver: Any | None = None          # UmfpackContext
        self._tech_csc: Any | None = None             # cached CSC matrix
        self._activity_cache: dict[tuple[str, str], Any] = {}
        # Diagnostics (read-only from outside)
        self.factorizations: int = 0
        self.redo_calls: int = 0
        self.method_switches: int = 0

    # ── internal helpers ───────────────────────────────────────────────────

    def _get_activity(self, key: tuple[str, str]) -> Any:
        """Cached ``bw2data.get_activity`` lookup."""
        if key not in self._activity_cache:
            self._activity_cache[key] = bw2data.get_activity(key)
        return self._activity_cache[key]

    def _to_fu(self, demand: dict[tuple[str, str], float]) -> dict:
        return {self._get_activity(k): float(v) for k, v in demand.items()}

    def _factorize_umfpack(self) -> bool:
        """Build an UMFPACK solver from the current technosphere matrix.

        Returns ``True`` on success, ``False`` if UMFPACK is unavailable.
        On success, ``self._lca.solver`` is set so that ``redo_lci`` and
        ``lci_calculation`` use the fast back-substitution path.
        """
        if not _UMFPACK_OK:
            return False
        try:
            import warnings

            from scikits.umfpack import UMFPACK_A, UmfpackContext

            A = self._lca.technosphere_matrix.tocsc()  # type: ignore[union-attr]
            umf = UmfpackContext("di")
            with warnings.catch_warnings():
                warnings.filterwarnings("ignore", ".*singular matrix.*")
                umf.numeric(A)
            # Store references so the GC doesn't collect them.
            self._umf_solver = umf
            self._tech_csc = A
            # Inject a solver callable that bw2calc's solve_linear_system
            # will find (it checks ``hasattr(self, 'solver')``).
            self._lca.solver = lambda b: umf.solve(  # type: ignore[union-attr]
                UMFPACK_A, A, b, autoTranspose=True
            )
            return True
        except Exception:
            return False

    # ── public interface (same signature as run_lca_multi_method) ──────────

    def __call__(
        self,
        demand: dict[tuple[str, str], float],
        method_tuples: list[tuple],
    ) -> dict[tuple, tuple[float, str]]:
        if not method_tuples:
            return {}
        # Deduplicate methods, preserving order
        unique: list[tuple] = []
        seen: set[tuple] = set()
        for m in method_tuples:
            mt = tuple(m)
            if mt not in seen:
                seen.add(mt)
                unique.append(mt)
        if not demand:
            return {m: (0.0, _method_unit(m)) for m in unique}

        fu = self._to_fu(demand)
        first_method = unique[0]

        if self._lca is None:
            # ── First call: load matrices + UMFPACK factorization ─────
            self._lca = bw2calc.LCA(fu, first_method)
            self._lca.lci()                     # loads matrices, spsolve
            self._factorize_umfpack()            # cache UMFPACK factors
            self._lca.lcia()
            self.factorizations += 1
        else:
            # ── Subsequent call: reuse factorized matrix ──────────────
            # redo_lci → build_demand_array + lci_calculation →
            # solve_linear_system → self.solver (UMFPACK back-sub ~15 ms)
            try:
                self._lca.redo_lci(fu)
            except Exception:
                # Activity not in product_dict → rebuild from scratch.
                self._lca = bw2calc.LCA(fu, first_method)
                self._lca.lci()
                self._factorize_umfpack()
                self._lca.lcia()
                self.factorizations += 1
                out: dict[tuple, tuple[float, str]] = {
                    first_method: (float(self._lca.score), _method_unit(first_method))
                }
                for mt in unique[1:]:
                    self._lca.switch_method(mt)
                    self._lca.redo_lcia()
                    out[mt] = (float(self._lca.score), _method_unit(mt))
                    self.method_switches += 1
                return out

            self._lca.switch_method(first_method)
            self._lca.redo_lcia()
            self.redo_calls += 1

        out = {first_method: (float(self._lca.score), _method_unit(first_method))}

        for mt in unique[1:]:
            self._lca.switch_method(mt)
            self._lca.redo_lcia()
            out[mt] = (float(self._lca.score), _method_unit(mt))
            self.method_switches += 1

        return out


def run_archetype_lca(
    demand: dict[tuple[str, str], float], method_tuple: tuple
) -> dict[str, Any]:
    """Standalone LCA for a flattened archetype BOM. Returns score + per-activity
    contributions.

    Per-material attribution: we run individual LCAs per activity. For a single
    archetype this is a small number of materials (typically <30) so the
    cost is acceptable; the detail is what makes the UI useful.
    """
    if not demand:
        return {"score": 0.0, "unit": "", "by_activity": {}}
    method = tuple(method_tuple)
    by_activity: dict[tuple[str, str], float] = {}
    for key, amount in demand.items():
        if amount == 0:
            continue
        act = bw2data.get_activity(key)
        lca = bw2calc.LCA({act: float(amount)}, method)
        lca.lci()
        lca.lcia()
        by_activity[key] = float(lca.score)
    unit = (
        bw2data.methods[method].get("unit", "unknown")
        if method in bw2data.methods
        else "unknown"
    )
    return {"score": sum(by_activity.values()), "unit": unit, "by_activity": by_activity}


def get_contributions(lca, total_score: float, limit: int = 10) -> dict[str, Any]:
    ca = bw2analyzer.ContributionAnalysis()
    top = ca.annotated_top_processes(lca, limit=limit + 5)

    items = []
    top_total = 0.0
    for score, _, act in top[:limit]:
        pct = (abs(score) / abs(total_score) * 100) if total_score else 0
        items.append({
            "activity_name": act.get("name", ""),
            "activity_key": str(act.key),
            "location": str(act.get("location", "")),
            "amount": float(score),
            "unit": bw2data.methods.get(lca.method, {}).get("unit", ""),
            "percentage": round(pct, 2),
        })
        top_total += float(score)

    rest = total_score - top_total
    rest_pct = (abs(rest) / abs(total_score) * 100) if total_score else 0

    return {
        "items": items,
        "rest_amount": float(rest),
        "rest_percentage": round(rest_pct, 2),
    }


def get_biosphere_contributions(
    lca, total_score: float, limit: int = 10
) -> dict[str, Any]:
    """Top biosphere flows contributing to the LCIA score.

    Computes per-flow scores by summing rows of ``lca.characterized_inventory``
    (a sparse biosphere-flows × technosphere-activities matrix). Replaces
    ``bw2analyzer.annotated_top_emissions`` because that helper passes
    ``numpy.float64`` indices into scipy sparse matrices — newer scipy
    rejects this with "Inexact indices into sparse matrices are not allowed",
    silently returning an empty list.
    """
    import numpy as np

    method_unit = bw2data.methods.get(lca.method, {}).get("unit", "")

    if not hasattr(lca, "characterized_inventory") or lca.characterized_inventory is None:
        logger.warning("biosphere CA: lca.characterized_inventory missing — did lcia() run?")
        return {"items": [], "rest_amount": float(total_score), "rest_percentage": 0.0}

    try:
        row_scores = np.asarray(lca.characterized_inventory.sum(axis=1)).ravel()
    except Exception as e:
        logger.warning("biosphere CA: failed to sum characterized_inventory: %s", e)
        return {"items": [], "rest_amount": float(total_score), "rest_percentage": 0.0}

    nonzero = int((row_scores != 0).sum())
    if nonzero == 0:
        logger.info(
            "biosphere CA: characterized_inventory has no non-zero biosphere rows "
            "(score=%s, method=%s) — method may not characterise any flows in this inventory",
            total_score, lca.method,
        )
        return {"items": [], "rest_amount": float(total_score), "rest_percentage": 0.0}

    # Top N by absolute contribution.
    order = np.argsort(np.abs(row_scores))[::-1]
    top_idx = order[:limit]

    # Reverse biosphere dict: matrix-row → biosphere flow key.
    try:
        rev_bio = {v: k for k, v in lca.biosphere_dict.items()}
    except AttributeError:
        # Older bw2calc versions had reverse_dict() returning (ra, rp, rb).
        try:
            _, _, rev_bio = lca.reverse_dict()
        except Exception as e:
            logger.warning("biosphere CA: cannot build reverse biosphere dict: %s", e)
            return {"items": [], "rest_amount": float(total_score), "rest_percentage": 0.0}

    items: list[dict[str, Any]] = []
    top_total = 0.0
    inventory = getattr(lca, "inventory", None)
    for raw_idx in top_idx:
        idx = int(raw_idx)  # critical: scipy sparse rejects numpy.float64
        score = float(row_scores[idx])
        if score == 0:
            continue
        flow_key = rev_bio.get(idx)
        if flow_key is None:
            continue
        try:
            flow = bw2data.get_activity(flow_key)
        except Exception:
            continue
        try:
            inv_amount = float(inventory[idx, :].sum()) if inventory is not None else 0.0
        except Exception:
            inv_amount = 0.0
        try:
            categories = list(flow.get("categories") or [])
        except Exception:
            categories = []
        pct = (abs(score) / abs(total_score) * 100) if total_score else 0.0
        items.append({
            "flow_name": flow.get("name", ""),
            "flow_key": str(flow.key),
            "categories": categories,
            "compartment": categories[0] if categories else "",
            "subcompartment": categories[1] if len(categories) > 1 else "",
            "inventory_amount": inv_amount,
            "inventory_unit": flow.get("unit", ""),
            "amount": score,
            "unit": method_unit,
            "percentage": round(pct, 2),
        })
        top_total += score

    rest = float(total_score) - top_total
    rest_pct = (abs(rest) / abs(total_score) * 100) if total_score else 0.0
    return {
        "items": items,
        "rest_amount": float(rest),
        "rest_percentage": round(rest_pct, 2),
    }


def get_recursive_contribution_tree(
    demand: dict[tuple[str, str], float],
    method: tuple,
    cutoff: float = 0.005,
    max_depth: int = 6,
    *,
    runner: "PersistentLCARunner | None" = None,
    unit_score_cache: dict[tuple[str, str], float] | None = None,
) -> dict[str, Any]:
    """Recursive impact-propagation tree.

    Distinct from ``get_supply_chain`` (which does BFS over technosphere
    exchanges directly): this walks the supply chain and at each node runs
    a sub-LCA so the impact propagated to children is the *characterised*
    contribution, not the raw exchange amount.

    Parameters
    ----------
    demand: functional unit (single root activity in normal use, but a
        multi-activity demand dict is accepted for archetype mode).
    method: LCIA method tuple.
    cutoff: minimum fraction of the root score below which a branch is
        truncated. Use the lowest value the caller might want — shallower
        views are derivable from a deeper tree without recomputing.
    max_depth: hard cap on recursion depth.

    Returns a tree::

        {
            "name": "<root>",
            "key": "<db>|<code>",
            "amount": <demand>,
            "unit": "<unit>",
            "score": <total characterised>,
            "unit_score": "<method unit>",
            "percentage": 100.0,
            "children": [ {...same shape...}, ... ],
        }
    """
    if not demand:
        return {
            "name": "(empty demand)",
            "key": "",
            "amount": 0.0,
            "unit": "",
            "score": 0.0,
            "unit_score": "",
            "percentage": 0.0,
            "children": [],
        }

    method_unit = _method_unit(method)

    # Use a persistent runner so the technosphere LU factorization is built
    # once and every sub-LCA is back-substitution only (~15 ms instead of
    # ~1.6 s). Critical for non-trivial trees on ecoinvent. Callers can pass
    # an existing runner to share the factorization across multiple
    # contribution-tree builds (e.g. multi-year trajectories).
    if runner is None:
        runner = PersistentLCARunner()

    method_t = tuple(method)

    # Unit-score memoization. Linear LCA: score(act, x) = x × unit_score(act).
    # Caching unit scores instead of (act, amount) tuples collapses thousands
    # of redo_calls (one per upstream demand of the same activity at different
    # amounts) into one call per unique activity. The cache scope is implicitly
    # (method, database) because both are fixed for the duration of the call;
    # callers can pass their own dict to share the cache with sibling builders
    # (e.g. ``get_supply_chain``) within the same year, and a fresh cache must
    # be passed for each new (method, database) combination.
    if unit_score_cache is None:
        unit_score_cache = {}

    def unit_score(act_key: tuple[str, str]) -> float:
        cached = unit_score_cache.get(act_key)
        if cached is not None:
            return cached
        try:
            out = runner({act_key: 1.0}, [method_t])
            s = float(out[method_t][0])
        except Exception:
            s = 0.0
        unit_score_cache[act_key] = s
        return s

    def score_for(act_key: tuple[str, str], amount: float) -> float:
        return float(amount) * unit_score(act_key)

    # Aggregate root score = denominator for cutoff filtering.
    out = runner(dict(demand), [method_t])
    root_score = float(out[method_t][0])
    abs_root = abs(root_score) if root_score else 1.0

    # Memoise per-(activity, amount, depth) so revisited subgraphs aren't recomputed.
    memo: dict[tuple[tuple[str, str], float, int], dict[str, Any]] = {}

    def expand(act_key: tuple[str, str], amount: float, depth: int,
               precomputed_score: float | None = None) -> dict[str, Any]:
        cache_key = (act_key, round(amount, 12), depth)
        if cache_key in memo:
            return memo[cache_key]

        act = bw2data.get_activity(act_key)
        score = precomputed_score if precomputed_score is not None else score_for(act_key, amount)
        pct = (abs(score) / abs_root * 100.0) if abs_root else 0.0

        node: dict[str, Any] = {
            "name": act.get("reference product", act.get("name", "")),
            "key": f"{act_key[0]}|{act_key[1]}",
            "location": str(act.get("location", "")),
            "amount": float(amount),
            "unit": act.get("unit", ""),
            "score": score,
            "unit_score": method_unit,
            "percentage": round(pct, 4),
            "children": [],
        }

        if depth >= max_depth or abs(score) <= 0:
            memo[cache_key] = node
            return node

        # Walk technosphere exchanges; recurse into each whose propagated
        # contribution clears the cutoff.
        children: list[dict[str, Any]] = []
        for exc in act.technosphere():
            try:
                child_key = exc.input.key
            except Exception:
                continue
            exc_amount = float(exc.get("amount", 0.0)) * float(amount)
            if exc_amount == 0:
                continue
            try:
                child_score = score_for(child_key, exc_amount)
            except Exception:
                continue
            if abs(child_score) / abs_root < cutoff:
                continue
            children.append(expand(child_key, exc_amount, depth + 1, precomputed_score=child_score))

        children.sort(key=lambda n: abs(n["score"]), reverse=True)
        node["children"] = children
        memo[cache_key] = node
        return node

    if len(demand) == 1:
        single_key, single_amt = next(iter(demand.items()))
        return expand(single_key, float(single_amt), 0)

    children = [expand(k, float(v), 0) for k, v in demand.items()]
    children.sort(key=lambda n: abs(n["score"]), reverse=True)
    return {
        "name": "(aggregated demand)",
        "key": "",
        "location": "",
        "amount": float(sum(demand.values())),
        "unit": "",
        "score": root_score,
        "unit_score": method_unit,
        "percentage": 100.0,
        "children": children,
    }


def get_supply_chain(
    lca,
    *,
    method,
    runner: "PersistentLCARunner | None" = None,
    depth: int = 3,
    max_nodes: int = 200,
    unit_score_cache: dict[tuple[str, str], float] | None = None,
) -> dict[str, Any]:
    """Build a Sankey graph by traversing the supply chain up to ``depth``
    levels. The graph is **acyclic by construction** (Sankey can't render
    cycles) and link values are **characterised impacts** in the active
    method's unit, matching the Tree view.

    Cycle handling
    --------------
    Each node is assigned a BFS level on first discovery. When processing an
    exchange ``u → v``:
      - ``v`` unvisited → add at ``level_u + 1``, emit edge.
      - ``v`` already at ``level_u + 1`` → aggregate edge value (same node
        reached via two distinct paths at the same depth — both are valid
        forward edges).
      - ``v`` at any other level (≤ ``level_u``, or > ``level_u + 1``) → drop
        the edge. Same-level and back-edges are intentionally dropped to keep
        the Sankey acyclic and layered, *not* because they're invalid in a
        general supply-chain graph; they're displayed implicitly via the
        already-discovered shorter path.

    Link value
    ----------
    For each edge ``u → v`` with exchange amount ``a`` (per unit of ``u``)
    and propagated upstream demand ``a × Q_u`` (where ``Q_u`` is what we've
    propagated into ``u``), the link value is::

        |a × Q_u × s_v|

    where ``s_v`` is the unit characterised score of ``v`` (impact per
    functional unit of ``v`` for ``method``). Single unit across the whole
    graph (the LCIA method's unit), no mixed-unit widths.

    Truncation
    ----------
    If the discovered graph has more than ``max_nodes`` nodes, prune by
    expanding from the root in best-first order (highest-value outgoing edge
    first) until the budget is met. The returned dict carries
    ``total_nodes_discovered`` and ``truncated`` so the UI can annotate.
    """
    from collections import deque
    import heapq

    if runner is None:
        runner = PersistentLCARunner()
    method_t = tuple(method)

    root_key = list(lca.demand.keys())[0]
    root_act = bw2data.get_activity(root_key)
    root_amount = float(lca.demand[root_key])

    def node_id(key: tuple) -> str:
        return f"{key[0]}_{key[1]}"

    nodes: dict[str, dict] = {}
    levels: dict[str, int] = {}
    # (src_id, tgt_id) → aggregated impact (always positive).
    link_values: dict[tuple[str, str], float] = {}

    # Cache unit scores so revisiting an activity at the same/forward level
    # doesn't refactor the matrix or re-back-substitute (~15 ms per call).
    # Caller can share this dict with ``get_recursive_contribution_tree`` so
    # the tree builder doesn't recompute unit scores for activities the
    # Sankey BFS already characterised. Cache scope is per (method, database).
    if unit_score_cache is None:
        unit_score_cache = {}

    def unit_score(act_key: tuple[str, str]) -> float:
        cached = unit_score_cache.get(act_key)
        if cached is not None:
            return cached
        try:
            out = runner({act_key: 1.0}, [method_t])
            s = float(out[method_t][0])
        except Exception:
            s = 0.0
        unit_score_cache[act_key] = s
        return s

    root_id = node_id(root_act.key)
    nodes[root_id] = {
        "id": root_id,
        "name": root_act.get("reference product", root_act.get("name", "")),
        "location": str(root_act.get("location", "")),
    }
    levels[root_id] = 0

    # Queue holds (activity, level, propagated_demand) — the demand is what
    # multiplies the per-unit exchange amounts into absolute upstream flow.
    queue: deque = deque([(root_act, 0, root_amount)])

    while queue:
        act, level, q_in = queue.popleft()
        if level >= depth:
            continue
        src_id = node_id(act.key)
        for exc in act.technosphere():
            try:
                child_key = exc.input.key
                child_act = bw2data.get_activity(child_key)
            except Exception:
                continue
            try:
                exc_per_unit = float(exc.get("amount", 0.0))
            except Exception:
                continue
            propagated = exc_per_unit * q_in
            if propagated == 0.0:
                continue
            tgt_id = node_id(child_act.key)
            tgt_level = levels.get(tgt_id)

            if tgt_level is None:
                # First time we see ``child`` — it's a forward edge.
                child_unit_score = unit_score(child_key)
                impact = abs(propagated * child_unit_score)
                if impact == 0.0:
                    continue
                nodes[tgt_id] = {
                    "id": tgt_id,
                    "name": child_act.get("reference product", child_act.get("name", "")),
                    "location": str(child_act.get("location", "")),
                }
                levels[tgt_id] = level + 1
                link_values[(src_id, tgt_id)] = link_values.get((src_id, tgt_id), 0.0) + impact
                queue.append((child_act, level + 1, propagated))
            elif tgt_level == level + 1:
                # Same forward layer reached via a different parent — aggregate.
                child_unit_score = unit_score(child_key)
                impact = abs(propagated * child_unit_score)
                if impact == 0.0:
                    continue
                link_values[(src_id, tgt_id)] = link_values.get((src_id, tgt_id), 0.0) + impact
            # Else: back-edge or skip-level cross-edge — drop to keep Sankey
            # layered and acyclic. The shorter path that already reached ``v``
            # carries its contribution.

    total_nodes_discovered = len(nodes)
    truncated = total_nodes_discovered > max_nodes

    if truncated:
        # Best-first expansion from root: at each step, follow the highest-
        # value outgoing edge into an unvisited node. Yields the subgraph
        # that captures the largest impact-bearing chains, not the shallowest.
        adj: dict[str, list[tuple[str, float]]] = {}
        for (s, t), v in link_values.items():
            adj.setdefault(s, []).append((t, v))

        kept: set[str] = {root_id}
        # Heap entries: (-value, target_id) — negate for max-heap via heapq.
        heap: list[tuple[float, str]] = []
        for t, v in adj.get(root_id, ()):
            heapq.heappush(heap, (-v, t))

        while heap and len(kept) < max_nodes:
            neg_v, t = heapq.heappop(heap)
            if t in kept:
                continue
            kept.add(t)
            for tt, vv in adj.get(t, ()):
                if tt not in kept:
                    heapq.heappush(heap, (-vv, tt))

        out_nodes = [nodes[n] for n in nodes if n in kept]
        out_links = [
            {"source": s, "target": t, "value": v}
            for (s, t), v in link_values.items()
            if s in kept and t in kept
        ]
    else:
        out_nodes = list(nodes.values())
        out_links = [
            {"source": s, "target": t, "value": v}
            for (s, t), v in link_values.items()
        ]

    return {
        "nodes": out_nodes,
        "links": out_links,
        "total_nodes_discovered": total_nodes_discovered,
        "truncated": truncated,
    }
