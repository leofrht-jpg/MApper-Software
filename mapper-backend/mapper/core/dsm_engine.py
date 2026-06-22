# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Pure-Python dynamic stock model engine for MApper's DSM module.

This module is intentionally framework-agnostic: no FastAPI, no global state,
no IO of its own beyond CSV/XLSX parsing helpers. The ``DynamicStockModel`` class
takes a system definition + state and returns a fully-resolved
``SimulationResult``.

Mass balance per year:
    stock(t+1) = stock(t) - outflow(t) + inflow(t)

Cohort tracking is done by keeping ``stock_by_age[cohort_key][age] = count``.
A cohort_key is the pipe-delimited concatenation of non-age dimension labels
in the order the dimensions were declared on the system (e.g. ``BEV|Small``).
"""
from __future__ import annotations

import csv
import io
import math
import os
import re
from itertools import product

from openpyxl import Workbook, load_workbook

from mapper.core.parameter_engine import ParameterEngine
from mapper.models.dsm_schemas import (
    DimensionDef,
    DSMMode,
    DSMScalingRule,
    DSMSystemState,
    InflowData,
    MaterializedDSMState,
    ModeConfig,
    OutflowData,
    ScalingTarget,
    materialize_scenario,
    SimulationResult,
    SimulationSummary,
    StockTargetData,
    SurvivalConfig,
    SurvivalPreviewPoint,
    SystemDefinition,
    YearResult,
)


COHORT_SEP = "|"
DEFAULT_WEIBULL_SHAPE = 4.0
DEFAULT_WEIBULL_SCALE = 15.0


# ── Math helpers ─────────────────────────────────────────────────────────────


def weibull_survival(age: float, shape: float, scale: float) -> float:
    """S(a) = exp(-(a/scale)^shape).

    S(0) = 1.0; S decreases monotonically as age grows.
    """
    if age <= 0:
        return 1.0
    if scale <= 0 or shape <= 0:
        return 1.0
    return math.exp(-((age / scale) ** shape))


def weibull_hazard(age: float, shape: float, scale: float) -> float:
    """Conditional probability of failure at integer ``age``, given survival to ``age - 1``.

    h(a) = 1 - S(a)/S(a-1). Returns 0 when ``age <= 0``.
    """
    if age <= 0:
        return 0.0
    s_prev = weibull_survival(age - 1, shape, scale)
    if s_prev <= 0:
        return 1.0
    s_curr = weibull_survival(age, shape, scale)
    h = 1.0 - (s_curr / s_prev)
    return max(0.0, min(1.0, h))


def survival_curve(shape: float, scale: float, max_age: int) -> list[SurvivalPreviewPoint]:
    """Return S(a) and h(a) at each integer age in [0, max_age]."""
    out: list[SurvivalPreviewPoint] = []
    for a in range(max_age + 1):
        out.append(
            SurvivalPreviewPoint(
                age=a,
                survival_rate=weibull_survival(a, shape, scale),
                hazard_rate=weibull_hazard(a, shape, scale),
            )
        )
    return out


def custom_curve_survival(curve: list[tuple[int, float]], age: int) -> float:
    """Linear interpolation between custom (age, survival_rate) points."""
    if not curve:
        return 1.0
    pts = sorted(curve, key=lambda p: p[0])
    if age <= pts[0][0]:
        return pts[0][1]
    if age >= pts[-1][0]:
        return pts[-1][1]
    for (a0, s0), (a1, s1) in zip(pts, pts[1:]):
        if a0 <= age <= a1:
            if a1 == a0:
                return s0
            t = (age - a0) / (a1 - a0)
            return s0 + t * (s1 - s0)
    return pts[-1][1]


def custom_curve_hazard(curve: list[tuple[int, float]], age: int) -> float:
    if age <= 0:
        return 0.0
    s_prev = custom_curve_survival(curve, age - 1)
    if s_prev <= 0:
        return 1.0
    s_curr = custom_curve_survival(curve, age)
    return max(0.0, min(1.0, 1.0 - (s_curr / s_prev)))


def largest_remainder_round(
    values: dict, target_total: float | None = None
) -> dict:
    """Round a mapping of ``key → float`` to integer-valued floats, preserving
    the rounded total.

    Uses the `largest remainder method <https://en.wikipedia.org/wiki/Largest_remainders_method>`_:
    every value is floored, then the residual ``round(Σv) - Σ floor(v)`` units
    are distributed one-by-one to the entries with the largest fractional
    parts. Ties on fractional part resolve by sort order of the keys, so the
    result is fully deterministic.

    ``target_total`` pins the post-rounding sum; when omitted it defaults to
    ``round(sum(values))``. Returns ``{}`` on empty input. Negative values are
    left untouched — they're not expected in this engine, and ``math.floor``
    plus remainder handling would give surprising results.
    """
    if not values:
        return {}
    items = list(values.items())
    floors: dict = {k: math.floor(v) for k, v in items}
    target = round(sum(values.values())) if target_total is None else round(target_total)
    to_distribute = int(target - sum(floors.values()))
    if to_distribute > 0:
        # Sort descending by fractional part; keys break ties deterministically.
        ranked = sorted(
            items,
            key=lambda kv: (-(kv[1] - math.floor(kv[1])), str(kv[0])),
        )
        for k, _ in ranked[:to_distribute]:
            floors[k] = floors[k] + 1
    elif to_distribute < 0:
        # Remove units from the smallest fractional parts (entries closest to
        # their floor). Applies when target_total is explicitly below Σ floor.
        ranked = sorted(
            items,
            key=lambda kv: (kv[1] - math.floor(kv[1]), str(kv[0])),
        )
        for k, _ in ranked[: -to_distribute]:
            if floors[k] > 0:
                floors[k] = floors[k] - 1
    # Return with float values to match the engine's dict[str|int, float] typing.
    return {k: float(v) for k, v in floors.items()}


def weibull_reverse_age_decomposition(
    total: float,
    shape: float = DEFAULT_WEIBULL_SHAPE,
    scale: float = DEFAULT_WEIBULL_SCALE,
    max_age: int = 25,
    survival_floor: float = 1e-3,
    integer_units: bool = False,
) -> dict[int, float]:
    """Split ``total`` over integer ages 1..max_age using Weibull survival weights.

    Initial stock represents pre-existing units (age ≥ 1). Age=0 is reserved
    for new arrivals and must be supplied via the inflows CSV at year t₀.
    Returns ``{age: count}`` with keys ≥ 1. If ``total <= 0`` returns an
    empty dict.
    """
    if total <= 0:
        return {}
    weights: list[tuple[int, float]] = []
    for age in range(1, max_age + 1):
        w = weibull_survival(age, shape, scale)
        if w < survival_floor and weights:
            break
        weights.append((age, w))
    tw = sum(w for _, w in weights)
    if tw <= 0:
        # Degenerate case (e.g. scale → 0): collapse to the youngest available age.
        return {1: float(total)}
    raw = {age: float(total) * w / tw for age, w in weights}
    if integer_units:
        return largest_remainder_round(raw, target_total=float(total))
    return raw


# ── Cohort key utilities ─────────────────────────────────────────────────────


def non_age_dimensions(dims: list[DimensionDef]) -> list[DimensionDef]:
    return [d for d in dims if not d.is_age]


def all_cohort_keys(dims: list[DimensionDef]) -> list[str]:
    """Cartesian product of labels of every non-age dimension."""
    nads = non_age_dimensions(dims)
    if not nads:
        return [""]
    label_lists = [d.labels for d in nads]
    return [COHORT_SEP.join(combo) for combo in product(*label_lists)]


def cohort_key_to_dict(cohort_key: str, dims: list[DimensionDef]) -> dict[str, str]:
    nads = non_age_dimensions(dims)
    parts = cohort_key.split(COHORT_SEP) if cohort_key else []
    return {d.name: parts[i] for i, d in enumerate(nads) if i < len(parts)}


def dict_to_cohort_key(values: dict[str, str], dims: list[DimensionDef]) -> str:
    nads = non_age_dimensions(dims)
    return COHORT_SEP.join(values.get(d.name, "") for d in nads)


# ── Survival lookup ──────────────────────────────────────────────────────────


def _config_matches(cfg: SurvivalConfig, cohort_dict: dict[str, str]) -> bool:
    return all(cohort_dict.get(k) == v for k, v in cfg.dimension_filters.items())


def _config_specificity(cfg: SurvivalConfig) -> int:
    return len(cfg.dimension_filters)


def best_config_for_cohort(
    cohort_dict: dict[str, str], configs: list[SurvivalConfig]
) -> SurvivalConfig | None:
    """Most specific config whose dimension_filters match this cohort."""
    matching = [c for c in configs if _config_matches(c, cohort_dict)]
    if not matching:
        return None
    return max(matching, key=_config_specificity)


def best_mode_for_cohort(
    cohort_dict: dict[str, str], configs: list[ModeConfig]
) -> DSMMode:
    """Most specific :class:`ModeConfig` whose ``dimension_filters`` match.

    Defaults to ``"survival_inflow"`` when no config matches (legacy behavior).
    """
    matching = [
        c
        for c in configs
        if all(cohort_dict.get(k) == v for k, v in c.dimension_filters.items())
    ]
    if not matching:
        return "survival_inflow"
    return max(matching, key=lambda c: len(c.dimension_filters)).mode


def best_rule_for_cohort(
    cohort_dict: dict[str, str],
    rules: list[DSMScalingRule],
    target: ScalingTarget,
) -> DSMScalingRule | None:
    """Most specific :class:`DSMScalingRule` for this cohort and target.

    Same most-specific-filter-wins semantics as :func:`best_mode_for_cohort`.
    No stacking: a single rule matches per (cohort, target). Users combine
    factors inside one expression (e.g. ``adoption * growth``).
    """
    matching = [
        r
        for r in rules
        if r.applies_to == target
        and all(cohort_dict.get(k) == v for k, v in r.dimension_filters.items())
    ]
    if not matching:
        return None
    return max(matching, key=lambda r: len(r.dimension_filters))


# ── CSV / XLSX parsing ───────────────────────────────────────────────────────


SUPPORTED_EXTS = (".csv", ".xlsx", ".xls")


def _sniff_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:5])
    if sample.count(";") > sample.count(","):
        return ";"
    return ","


def _strip_bom(text: str) -> str:
    if text.startswith("\ufeff"):
        return text[1:]
    return text


def _validate_label(value: str, dim: DimensionDef) -> None:
    if value not in dim.labels:
        raise ValueError(
            f"Value '{value}' is not a valid label for dimension '{dim.name}'. "
            f"Allowed: {dim.labels}"
        )


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_rows_from_bytes(
    content: bytes, filename: str
) -> tuple[list[str], list[dict[str, str]]]:
    """Read a CSV or XLSX file into (headers, list-of-dicts).

    Cells are stringified + trimmed. Fully empty trailing rows are skipped.
    Raises ``ValueError`` on unsupported extensions or parse errors.
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        text = _strip_bom(text)
        delim = _sniff_delimiter(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        headers = [h.strip() for h in (reader.fieldnames or [])]
        rows: list[dict[str, str]] = []
        for raw in reader:
            row = {
                (k.strip() if k else ""): (v.strip() if isinstance(v, str) else ("" if v is None else str(v).strip()))
                for k, v in raw.items() if k
            }
            if any(row.values()):
                rows.append(row)
        return headers, rows
    if ext in (".xlsx", ".xls"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}") from e
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no sheets.")
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            raise ValueError("Excel sheet is empty.")
        headers = [_cell_to_str(c) for c in header_row]
        rows = []
        for raw in iterator:
            if not raw:
                continue
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _cell_to_str(raw[i]) if i < len(raw) else ""
            if any(row.values()):
                rows.append(row)
        return headers, rows
    raise ValueError(f"Unsupported file extension '{ext}'. Use .csv or .xlsx.")


def parse_stock_file(
    content: bytes, filename: str, dims: list[DimensionDef]
) -> tuple[dict[str, float], int]:
    """Parse the initial-stock CSV or XLSX.

    Expected columns: every non-age dimension name, ``age``, ``count``.
    Returns (initial_stock keyed by ``cohort_key|age``, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = [d.name for d in nads] + ["age", "count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Stock file missing required column(s): {missing}. Got headers: {headers}"
        )

    out: dict[str, float] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        try:
            age = int(float(row["age"]))
        except ValueError as e:
            raise ValueError(f"Invalid age '{row['age']}' on row {rows + 1}") from e
        if age < 1:
            raise ValueError(
                "Initial stock must contain only ages 1 and above. "
                "Age=0 represents new arrivals, which should be uploaded via "
                "the inflows CSV for the reference year."
            )
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        full_key = f"{cohort_key}|{age}"
        out[full_key] = out.get(full_key, 0.0) + count
        rows += 1
    return out, rows


def parse_dependent_stock_file(
    content: bytes, filename: str, dims: list[DimensionDef]
) -> tuple[dict[str, float], int]:
    """Parse a dependent-subsystem initial-stock CSV or XLSX.

    Expected columns: every non-age dimension name + ``count``. There is no
    ``age`` column — dependent subsystems don't carry an age distribution.
    Returns (stock keyed by cohort_key, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Stock file missing required column(s): {missing}. Got headers: {headers}"
        )

    out: dict[str, float] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        out[cohort_key] = out.get(cohort_key, 0.0) + count
        rows += 1
    return out, rows


def dependent_stock_template_csv(dims: list[DimensionDef]) -> str:
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        writer.writerow(list(combo) + [""])
    return buf.getvalue()


def parse_inflow_file(
    content: bytes, filename: str, dims: list[DimensionDef], horizon_years: list[int]
) -> tuple[list[InflowData], int]:
    """Parse the inflow CSV or XLSX.

    Expected columns: ``year``, every non-age dimension name, ``count``.
    Returns (list of InflowData per year, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = ["year"] + [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Inflow file missing required column(s): {missing}. Got headers: {headers}"
        )

    horizon = set(horizon_years)
    by_year: dict[int, dict[str, float]] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            year = int(float(row["year"]))
        except ValueError as e:
            raise ValueError(f"Invalid year '{row['year']}' on row {rows + 1}") from e
        if year not in horizon:
            raise ValueError(
                f"Year {year} on row {rows + 1} is outside the system's time horizon."
            )
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        bucket = by_year.setdefault(year, {})
        bucket[cohort_key] = bucket.get(cohort_key, 0.0) + count
        rows += 1

    out = [InflowData(year=y, counts=by_year[y]) for y in sorted(by_year)]
    return out, rows


def parse_stock_target_file(
    content: bytes, filename: str, dims: list[DimensionDef], horizon_years: list[int]
) -> tuple[list[StockTargetData], int]:
    """Parse a stock-target CSV or XLSX for Mode B (stock-driven).

    Expected columns: ``year``, every non-age dimension name, ``count``. Mirrors
    :func:`parse_inflow_file` for consistency — users can reuse the inflow
    template and just swap the meaning of the ``count`` column.
    Returns (list of StockTargetData per year, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = ["year"] + [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Stock target file missing required column(s): {missing}. Got headers: {headers}"
        )

    horizon = set(horizon_years)
    by_year: dict[int, dict[str, float]] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            year = int(float(row["year"]))
        except ValueError as e:
            raise ValueError(f"Invalid year '{row['year']}' on row {rows + 1}") from e
        if year not in horizon:
            raise ValueError(
                f"Year {year} on row {rows + 1} is outside the system's time horizon."
            )
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        if count < 0:
            raise ValueError(
                f"Negative stock target '{count}' on row {rows + 1} is not allowed."
            )
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        bucket = by_year.setdefault(year, {})
        bucket[cohort_key] = bucket.get(cohort_key, 0.0) + count
        rows += 1

    out = [StockTargetData(year=y, counts=by_year[y]) for y in sorted(by_year)]
    return out, rows


def parse_outflow_file(
    content: bytes,
    filename: str,
    dims: list[DimensionDef],
    horizon_years: list[int],
) -> tuple[list[OutflowData], int, bool]:
    """Parse a manual-mode outflow CSV or XLSX.

    Required columns: ``year``, every non-age dimension name, ``count``.
    Optional columns: ``age`` or ``birth_year`` — when present, the row's count
    is assigned to a specific cohort-age (cohort-specific outflow). When
    absent, the engine falls back to FIFO allocation from oldest ages.

    Returns ``(list of OutflowData per year, rows_parsed, cohort_specific)``.
    ``cohort_specific`` is True iff an age/birth_year column was recognised.
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = ["year"] + [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Outflow file missing required column(s): {missing}. Got headers: {headers}"
        )

    has_age = "age" in headers
    has_birth = "birth_year" in headers
    cohort_specific = has_age or has_birth

    by_year: dict[int, OutflowData] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            year = int(float(row["year"]))
        except ValueError as e:
            raise ValueError(f"Invalid year '{row['year']}' on row {rows + 1}") from e
        if year not in set(horizon_years):
            raise ValueError(
                f"Year {year} on row {rows + 1} is outside the system's time horizon."
            )
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        if count < 0:
            raise ValueError(
                f"Negative outflow '{count}' on row {rows + 1} is not allowed."
            )
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        bucket = by_year.setdefault(year, OutflowData(year=year))
        bucket.counts[cohort_key] = bucket.counts.get(cohort_key, 0.0) + count

        if cohort_specific:
            age: int | None = None
            if has_age and row.get("age"):
                try:
                    age = int(float(row["age"]))
                except ValueError as e:
                    raise ValueError(f"Invalid age '{row['age']}' on row {rows + 1}") from e
            elif has_birth and row.get("birth_year"):
                try:
                    by = int(float(row["birth_year"]))
                except ValueError as e:
                    raise ValueError(
                        f"Invalid birth_year '{row['birth_year']}' on row {rows + 1}"
                    ) from e
                age = year - by
            if age is not None:
                if age < 0:
                    raise ValueError(
                        f"birth_year > year on row {rows + 1} gives a negative age."
                    )
                key = f"{cohort_key}|{age}"
                bucket.cohort_age_counts[key] = (
                    bucket.cohort_age_counts.get(key, 0.0) + count
                )
        rows += 1

    out = [by_year[y] for y in sorted(by_year)]
    return out, rows, cohort_specific


def outflow_template_csv(
    dims: list[DimensionDef], horizon_years: list[int]
) -> str:
    """Template for manual-mode outflow uploads (same long format as inflows).

    The optional ``age`` column is documented in a trailing commented row but
    not emitted as an actual column — users who need cohort-specific outflows
    add it themselves.
    """
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            writer.writerow([year] + list(combo) + [""])
    return buf.getvalue()


def parse_aggregate_stock_file(
    content: bytes,
    filename: str,
    dims: list[DimensionDef],
    shape: float = DEFAULT_WEIBULL_SHAPE,
    scale: float = DEFAULT_WEIBULL_SCALE,
    max_age: int = 25,
) -> tuple[dict[str, float], int]:
    """Parse an aggregate-format initial-stock file (``dims..., count``) and
    spread each row across ages using :func:`weibull_reverse_age_decomposition`.

    Returns the same ``{cohort_key|age: count}`` mapping produced by
    :func:`parse_stock_file` so the engine doesn't need to distinguish the two.
    This is the Phase 4 entry point for users with aggregate-only data (e.g.,
    the wind-turbine case study).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Aggregate stock file missing required column(s): {missing}. Got headers: {headers}"
        )

    out: dict[str, float] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        if count <= 0:
            continue
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        decomposed = weibull_reverse_age_decomposition(count, shape, scale, max_age)
        for age, c in decomposed.items():
            full_key = f"{cohort_key}|{age}"
            out[full_key] = out.get(full_key, 0.0) + c
        rows += 1
    return out, rows


LABEL_FILENAME_RE = re.compile(r"^(?P<dim>[A-Za-z0-9_]+)_labels$")


def parse_label_file(
    content: bytes,
    filename: str,
    expected_dimension: str,
    valid_dimensions: list[str] | None = None,
) -> list[str]:
    """Parse a `{dimension}_labels.{csv,xlsx}` file for a specific dimension.

    Enforces:
      * filename basename matches `{expected_dimension}_labels` exactly
      * first-row first-column header equals `expected_dimension`
      * at least one data row after the header

    Returns deduped labels in source order.
    """
    base = os.path.basename(filename or "")
    stem, ext = os.path.splitext(base)
    ext = ext.lower()
    if ext not in (".csv", ".xlsx", ".xls"):
        raise ValueError(
            f"Unsupported file extension '{ext}'. Upload a .csv or .xlsx file named "
            f"'{expected_dimension}_labels{ext or '.csv'}'."
        )

    m = LABEL_FILENAME_RE.match(stem)
    if not m:
        raise ValueError(
            f"Filename must be '{expected_dimension}_labels{ext}' "
            f"(got '{base}')."
        )
    inferred = m.group("dim")
    if inferred != expected_dimension:
        if valid_dimensions:
            valid_list = ", ".join(sorted(set(valid_dimensions)))
            if inferred not in valid_dimensions:
                raise ValueError(
                    f"Unrecognized dimension '{inferred}' in filename '{base}'. "
                    f"Valid dimensions: {valid_list}."
                )
            raise ValueError(
                f"Filename '{base}' is for dimension '{inferred}', but this upload "
                f"targets '{expected_dimension}'. Valid dimensions: {valid_list}."
            )
        raise ValueError(
            f"Filename '{base}' is for dimension '{inferred}', but this upload "
            f"targets '{expected_dimension}'. Rename the file to "
            f"'{expected_dimension}_labels{ext}'."
        )

    values: list[str] = []
    if ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        text = _strip_bom(text)
        delim = _sniff_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        for row in reader:
            if not row:
                continue
            cell = (row[0] or "").strip()
            if cell:
                values.append(cell)
    else:
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}") from e
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no sheets.")
        for raw in ws.iter_rows(values_only=True):
            if not raw:
                continue
            cell = _cell_to_str(raw[0])
            if cell:
                values.append(cell)

    if not values:
        raise ValueError(f"File '{base}' is empty.")

    header = values[0]
    if header != expected_dimension:
        raise ValueError(
            f"Header in '{base}' must be '{expected_dimension}' (got '{header}'). "
            f"The first column header must exactly match the dimension name."
        )
    body = values[1:]
    if not body:
        raise ValueError(
            f"File '{base}' has a header but no label rows. Add at least one label."
        )

    seen: set[str] = set()
    out: list[str] = []
    for v in body:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


# ── CSV templates ────────────────────────────────────────────────────────────


def stock_template_csv(dims: list[DimensionDef], example_ages: int = 5) -> str:
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["age", "count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        for age in range(example_ages + 1):
            writer.writerow(list(combo) + [age, ""])
    return buf.getvalue()


def inflow_template_csv(dims: list[DimensionDef], horizon_years: list[int]) -> str:
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            writer.writerow([year] + list(combo) + [""])
    return buf.getvalue()


def stock_target_template_csv(
    dims: list[DimensionDef], horizon_years: list[int]
) -> str:
    """Template for Mode B stock-target uploads (long format, same shape as inflow)."""
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            writer.writerow([year] + list(combo) + [""])
    return buf.getvalue()


def aggregate_stock_template_csv(dims: list[DimensionDef]) -> str:
    """Template for aggregate-format initial stock (no ``age`` column).

    The server applies Weibull reverse decomposition to spread each row across
    synthetic age cohorts. Use this when age-resolved data for the cohort is
    not available.
    """
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        writer.writerow(list(combo) + [""])
    return buf.getvalue()


# ── XLSX templates with Instructions sheet ───────────────────────────────────


_AGE_CONVENTION_NOTES = [
    "Age convention:",
    "  • Initial stock contains pre-existing products (age 1 and above).",
    "  • New arrivals at the reference year t₀ are specified in the inflows CSV.",
    "  • Age=0 rows are NOT allowed in initial stock — uploads with age=0 are rejected.",
]


def _build_xlsx(
    template_sheet_name: str,
    template_headers: list[str],
    template_rows: list[list[object]],
    instructions_lines: list[str],
) -> bytes:
    """Build a two-sheet xlsx: a 'Template' sheet (active, parsed by upload
    endpoints — first sheet wins) and an 'Instructions' sheet documenting
    conventions. Returns the xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = template_sheet_name
    ws.append(template_headers)
    for row in template_rows:
        ws.append(row)
    info = wb.create_sheet("Instructions")
    for line in instructions_lines:
        info.append([line])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def stock_template_xlsx(dims: list[DimensionDef], example_ages: int = 5) -> bytes:
    """Format A initial-stock template (by-age). Example rows start at age=1
    (age=0 is reserved for the inflows CSV)."""
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["age", "count"]
    rows: list[list[object]] = []
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        for age in range(1, example_ages + 1):
            rows.append(list(combo) + [age, ""])
    instructions = [
        "Initial Stock — Format A (by-age)",
        "",
        f"Columns: {', '.join(headers)}",
        "  • One row per (cohort, age). 'count' is the number of units of that cohort alive at the reference year t₀ with that age.",
        "  • 'age' is integer years since manufacture/production. Valid ages are 1, 2, 3, ...",
        "  • A 'unit' is whatever your system tracks: a vehicle, a building, a wind turbine, a device, a machine, a kg of material, etc. Counts can be integers (discrete items) or fractional quantities (mass, energy, volume).",
        "",
    ] + _AGE_CONVENTION_NOTES + [
        "",
        "Examples (using your system's dimension labels in the rows above):",
        "  • To represent 200 units of a cohort produced 3 years ago → row with age=3, count=200.",
        "  • For new arrivals at t₀, do NOT add an age=0 row here — use the annual-inflows template with year=t₀ instead.",
        "",
        "Tips:",
        "  • Empty 'count' cells are ignored. Delete rows you don't need or leave them blank.",
        "  • Uploads containing any age=0 rows are rejected with a validation error.",
    ]
    return _build_xlsx("Template", headers, rows, instructions)


def aggregate_stock_template_xlsx(dims: list[DimensionDef]) -> bytes:
    """Format B aggregate-stock template (no age column)."""
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["count"]
    rows: list[list[object]] = []
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        rows.append(list(combo) + [""])
    instructions = [
        "Initial Stock — Format B (aggregate, no age column)",
        "",
        f"Columns: {', '.join(headers)}",
        "  • One row per cohort. 'count' is the total number of pre-existing units (age ≥ 1) of that cohort alive at the reference year t₀.",
        "  • A 'unit' is whatever your system tracks (vehicles, buildings, turbines, devices, kg of material, etc.).",
        "",
        "How the engine handles this format:",
        "  • The server applies a Weibull REVERSE age decomposition to spread each row across synthetic age cohorts (ages 1..max_age — age=0 is excluded).",
        "  • The full count remains pre-existing stock; manufacturing impacts are NOT counted for it.",
        "  • New arrivals at t₀ must be supplied via the annual-inflows template with year=t₀.",
        "",
    ] + _AGE_CONVENTION_NOTES + [
        "",
        "Use this format when you do not have age-resolved data for the cohort. Use Format A (by-age) when you do.",
    ]
    return _build_xlsx("Template", headers, rows, instructions)


def inflow_template_xlsx(
    dims: list[DimensionDef], horizon_years: list[int]
) -> bytes:
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    rows: list[list[object]] = []
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            rows.append([year] + list(combo) + [""])
    t0 = horizon_years[0] if horizon_years else "t₀"
    instructions = [
        "Annual Inflows",
        "",
        f"Columns: {', '.join(headers)}",
        "  • One row per (year, dimension combination). 'count' is the number of units manufactured (and entering stock) in that year.",
        "  • This is the SOLE source of new arrivals, including arrivals at the reference year t₀.",
        "",
        "Year column convention:",
        f"  • Include a row at year = t₀ ({t0}) for any cohort that has new arrivals at the reference year. Manufacturing impacts at t₀ come exclusively from this row.",
        "  • Initial stock (uploaded separately) only contains pre-existing products with age ≥ 1.",
        "",
    ] + _AGE_CONVENTION_NOTES
    return _build_xlsx("Template", headers, rows, instructions)


def stock_target_template_xlsx(
    dims: list[DimensionDef], horizon_years: list[int]
) -> bytes:
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    rows: list[list[object]] = []
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            rows.append([year] + list(combo) + [""])
    instructions = [
        "Stock Targets (Mode B — survival_stock)",
        "",
        f"Columns: {', '.join(headers)}",
        "  • One row per (year, dimension combination). 'count' is the desired total alive stock for that cohort at end-of-year.",
        "  • The engine back-calculates the inflows needed each year to reach these targets, given the Weibull survival hazard.",
        "",
        "Note on year-zero accounting:",
        "  • survival_stock mode does NOT inject inflows at t₀ — the initial stock is taken as given. Year-0 inflow is therefore zero in this mode.",
        "  • If you need manufacturing impacts attributed at t₀, switch the cohort to survival_inflow mode and supply the t₀ row in the annual-inflows template.",
        "",
    ] + _AGE_CONVENTION_NOTES
    return _build_xlsx("Template", headers, rows, instructions)


def outflow_template_xlsx(
    dims: list[DimensionDef], horizon_years: list[int]
) -> bytes:
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    rows: list[list[object]] = []
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            rows.append([year] + list(combo) + [""])
    instructions = [
        "Manual Outflows",
        "",
        f"Required columns: {', '.join(headers)}",
        "Optional columns: 'age' OR 'birth_year' for cohort-specific outflows.",
        "  • Without age/birth_year → engine retires units oldest-first across all ages.",
        "  • With age=k → retires units that are exactly k years old in that year.",
        "  • With birth_year=Y → retires units born in year Y (i.e. age = year − Y).",
        "",
    ] + _AGE_CONVENTION_NOTES
    return _build_xlsx("Template", headers, rows, instructions)


# ── Post-simulation diagnostics ──────────────────────────────────────────────


FLEET_DRIFT_THRESHOLD = 0.05  # ±5% per the spec


def _fleet_drift_warnings(years_out: list[YearResult], _unused: float) -> list[str]:
    """Return one warning if total fleet stock drifts more than ±5% from
    the first simulated year's post-hazard total.

    Uses year 0's simulated stock (not the raw initial_stock sum) as the
    baseline — otherwise Mode B's year-0 no-correction rule would register
    natural Weibull attrition as drift. Only the worst-drift year is surfaced.
    """
    if not years_out:
        return []
    baseline = float(sum(years_out[0].stock.values()))
    if baseline <= 0:
        return []
    worst_year: int | None = None
    worst_total: float = 0.0
    worst_drift: float = 0.0
    for yr in years_out[1:]:
        total = float(sum(yr.stock.values()))
        drift = (total - baseline) / baseline
        if abs(drift) > abs(worst_drift):
            worst_drift = drift
            worst_total = total
            worst_year = yr.year
    if worst_year is None or abs(worst_drift) <= FLEET_DRIFT_THRESHOLD:
        return []
    return [
        f"Total fleet stock drifted to {worst_total:,.0f} in {worst_year} "
        f"({worst_drift:+.1%} vs {baseline:,.0f} baseline). "
        "Verify that Mode A inflows and Mode B stock targets come from a consistent scenario."
    ]


# ── Engine ───────────────────────────────────────────────────────────────────


class DynamicStockModel:
    """Year-by-year cohort-tracking stock-flow model.

    Public surface:
        model = DynamicStockModel(system, state)
        result = model.simulate()
    """

    def __init__(
        self,
        system: SystemDefinition,
        state: MaterializedDSMState | DSMSystemState,
        parameter_engine: ParameterEngine | None = None,
    ) -> None:
        self.system = system
        # Accept either the persisted multi-scenario state (auto-materialize the
        # active or Base scenario) or a pre-resolved view handed in by the API.
        if isinstance(state, DSMSystemState):
            state = materialize_scenario(state)
        self.state = state
        self.parameter_engine = parameter_engine
        self.years: list[int] = system.time_horizon.years
        self.cohort_keys: list[str] = all_cohort_keys(system.dimensions)
        self.max_age: int = system.time_horizon.length  # safe upper bound

        # Pre-compute scaled base data once per simulate. Without an engine or
        # rules, the scaled dicts mirror the uploaded values (backward compat).
        self._scaled_inflows: dict[int, dict[str, float]] = self._scale_year_cohort(
            {i.year: dict(i.counts) for i in state.inflows}, "inflows"
        )
        self._scaled_targets: dict[int, dict[str, float]] = self._scale_year_cohort(
            {t.year: dict(t.counts) for t in state.stock_targets}, "stock_targets"
        )
        self._scaled_outflows: dict[int, OutflowData] = self._scale_outflows()

    # — Scaling —

    def _resolve_rule(self, rule: DSMScalingRule, base: float, year: int) -> float:
        """Evaluate ``rule.expression`` with ``base`` / ``year`` injected."""
        if self.parameter_engine is None:
            # Unreachable in practice — ``_scale_year_cohort`` / ``_scale_outflows``
            # short-circuit when no engine is set. Kept defensive.
            return base
        return self.parameter_engine.resolve(
            rule.expression, extra_vars={"base": base, "year": float(year)}
        )

    def _scale_year_cohort(
        self, base: dict[int, dict[str, float]], target: ScalingTarget
    ) -> dict[int, dict[str, float]]:
        if self.parameter_engine is None or not self.state.scaling_rules:
            return base
        rules = [r for r in self.state.scaling_rules if r.applies_to == target]
        if not rules:
            return base
        out: dict[int, dict[str, float]] = {}
        for year, cohorts in base.items():
            out[year] = {}
            for ck, count in cohorts.items():
                cohort_dict = cohort_key_to_dict(ck, self.system.dimensions)
                rule = best_rule_for_cohort(cohort_dict, rules, target)
                out[year][ck] = (
                    count if rule is None else self._resolve_rule(rule, count, year)
                )
        return out

    def _scale_outflows(self) -> dict[int, OutflowData]:
        base_map = {o.year: o for o in self.state.outflows}
        if self.parameter_engine is None or not self.state.scaling_rules:
            return base_map
        rules = [r for r in self.state.scaling_rules if r.applies_to == "outflows"]
        if not rules:
            return base_map
        out: dict[int, OutflowData] = {}
        for year, o in base_map.items():
            new_counts: dict[str, float] = {}
            for ck, count in o.counts.items():
                cohort_dict = cohort_key_to_dict(ck, self.system.dimensions)
                rule = best_rule_for_cohort(cohort_dict, rules, "outflows")
                new_counts[ck] = (
                    count if rule is None else self._resolve_rule(rule, count, year)
                )
            new_age: dict[str, float] = {}
            for key, count in o.cohort_age_counts.items():
                ck = key.rsplit("|", 1)[0]
                cohort_dict = cohort_key_to_dict(ck, self.system.dimensions)
                rule = best_rule_for_cohort(cohort_dict, rules, "outflows")
                new_age[key] = (
                    count if rule is None else self._resolve_rule(rule, count, year)
                )
            out[year] = OutflowData(year=year, counts=new_counts, cohort_age_counts=new_age)
        return out

    # — Survival lookup —

    def _survival_params(
        self, cohort_key: str
    ) -> tuple[str, float, float, list[tuple[int, float]] | None]:
        cohort_dict = cohort_key_to_dict(cohort_key, self.system.dimensions)
        cfg = best_config_for_cohort(cohort_dict, self.state.survival_configs)
        if cfg is None:
            return ("weibull", DEFAULT_WEIBULL_SHAPE, DEFAULT_WEIBULL_SCALE, None)
        if cfg.method == "custom" and cfg.custom_curve:
            curve = [(p.age, p.survival_rate) for p in cfg.custom_curve]
            return ("custom", 0.0, 0.0, curve)
        shape = cfg.weibull_shape if cfg.weibull_shape is not None else DEFAULT_WEIBULL_SHAPE
        scale = cfg.weibull_scale if cfg.weibull_scale is not None else DEFAULT_WEIBULL_SCALE
        return ("weibull", shape, scale, None)

    def hazard(self, cohort_key: str, age: int) -> float:
        method, shape, scale, curve = self._survival_params(cohort_key)
        if method == "custom" and curve is not None:
            return custom_curve_hazard(curve, age)
        return weibull_hazard(age, shape, scale)

    # — Initialization —

    def _initial_stock_by_age(self) -> dict[str, dict[int, float]]:
        out: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}
        for full_key, count in self.state.initial_stock.items():
            try:
                cohort_key, age_str = full_key.rsplit("|", 1)
                age = int(age_str)
            except ValueError:
                # Treat keys with no age as age 0.
                cohort_key, age = full_key, 0
            if cohort_key not in out:
                out[cohort_key] = {}
            out[cohort_key][age] = out[cohort_key].get(age, 0.0) + count
        if self.state.integer_units:
            # Round each cohort's age distribution to integers, preserving the
            # rounded cohort total. Covers aggregate decomposition and any
            # fractional initial values from scaled imports.
            for ck, ages in out.items():
                if ages:
                    out[ck] = largest_remainder_round(ages)
        return out

    def _inflow_for_year(self, year: int) -> dict[str, float]:
        return dict(self._scaled_inflows.get(year, {}))

    def _outflow_for_year(self, year: int) -> OutflowData | None:
        return self._scaled_outflows.get(year)

    def _mode_for(self, cohort_key: str) -> DSMMode:
        cohort_dict = cohort_key_to_dict(cohort_key, self.system.dimensions)
        return best_mode_for_cohort(cohort_dict, self.state.mode_configs)

    # — Main loop —

    def simulate(self) -> SimulationResult:
        stock_by_age = self._initial_stock_by_age()
        integer_units = bool(self.state.integer_units)

        # Resolve each cohort's mode once — configs don't change mid-run.
        cohort_modes: dict[str, DSMMode] = {
            ck: self._mode_for(ck) for ck in self.cohort_keys
        }
        targets_by_year: dict[int, dict[str, float]] = {
            y: dict(c) for y, c in self._scaled_targets.items()
        }

        years_out: list[YearResult] = []
        total_in = 0.0
        total_out = 0.0
        total_start = sum(sum(d.values()) for d in stock_by_age.values())
        manual_warnings: list[str] = []

        for idx, year in enumerate(self.years):
            year_inflow: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            year_natural: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            year_forced: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            year_manual: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            outflow_by_age: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}
            forced_by_age: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}

            # 1. Age every existing cohort by one year. Skipped in year 0: the
            # initial stock comes from the CSV already at its correct ages.
            if idx > 0:
                aged: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}
                for ck, ages in stock_by_age.items():
                    for a, count in ages.items():
                        aged.setdefault(ck, {})[a + 1] = aged.get(ck, {}).get(a + 1, 0.0) + count
                stock_by_age = aged

            # 2. Apply survival hazard age-by-age for survival-mode cohorts.
            #    Manual-mode cohorts bypass this step — their outflows come
            #    entirely from the user-uploaded schedule.
            for ck, ages in stock_by_age.items():
                if cohort_modes.get(ck) == "manual":
                    continue
                raw_out: dict[int, float] = {}
                for a, count in ages.items():
                    h = self.hazard(ck, a)
                    raw_out[a] = count * h
                if integer_units:
                    # Round each age's outflow so the cohort's annual natural
                    # outflow stays an integer. Survivors at age a are
                    # ``count[a] - rounded_out[a]`` → also integer since count
                    # is already integer (initial stock / prior-year rounded).
                    raw_out = largest_remainder_round(raw_out)
                new_ages: dict[int, float] = {}
                for a, count in ages.items():
                    out = raw_out.get(a, 0.0)
                    survived = count - out
                    if out > 0:
                        outflow_by_age[ck][a] = out
                        year_natural[ck] = year_natural.get(ck, 0.0) + out
                    if survived > 0:
                        new_ages[a] = survived
                stock_by_age[ck] = new_ages

            # 3. Per-cohort flow injection — dispatch by resolved mode.
            inflows = self._inflow_for_year(year)
            targets = targets_by_year.get(year, {})
            outflow_data = self._outflow_for_year(year)
            for ck in self.cohort_keys:
                mode = cohort_modes[ck]
                if mode == "survival_inflow":
                    count = inflows.get(ck, 0.0)
                    if integer_units:
                        count = float(round(count))
                    if count > 0:
                        if ck not in stock_by_age:
                            stock_by_age[ck] = {}
                        stock_by_age[ck][0] = stock_by_age[ck].get(0, 0.0) + count
                        year_inflow[ck] = year_inflow.get(ck, 0.0) + count
                elif mode == "survival_stock":
                    # Year 0 is the base year: initial stock is taken as given,
                    # no gap correction (avoids a phantom inflow that would be
                    # incorrectly counted as year-0 manufacturing).
                    if idx == 0:
                        continue
                    target = targets.get(ck, 0.0)
                    surviving = sum(stock_by_age.get(ck, {}).values())
                    gap = target - surviving
                    if integer_units:
                        # Round target to integer before gap computation so
                        # inflows/forced retirements stay whole-unit.
                        gap = float(round(target)) - surviving
                    if gap > 1e-9:
                        if ck not in stock_by_age:
                            stock_by_age[ck] = {}
                        stock_by_age[ck][0] = stock_by_age[ck].get(0, 0.0) + gap
                        year_inflow[ck] = year_inflow.get(ck, 0.0) + gap
                    elif gap < -1e-9:
                        # Forced retirement — FIFO from oldest age (= earliest
                        # birth year within this cohort_key). Snapshot ages
                        # first because we mutate the dict during iteration.
                        remaining = -gap
                        ages_desc = sorted(
                            stock_by_age.get(ck, {}).keys(), reverse=True
                        )
                        for age in ages_desc:
                            if remaining <= 1e-12:
                                break
                            available = stock_by_age[ck].get(age, 0.0)
                            take = min(available, remaining)
                            new_count = available - take
                            if new_count <= 1e-12:
                                del stock_by_age[ck][age]
                            else:
                                stock_by_age[ck][age] = new_count
                            forced_by_age[ck][age] = (
                                forced_by_age[ck].get(age, 0.0) + take
                            )
                            remaining -= take
                        year_forced[ck] = year_forced.get(ck, 0.0) + (-gap - remaining)
                else:  # manual
                    # 3a. Add user-provided inflow at age 0.
                    inflow_count = inflows.get(ck, 0.0)
                    if integer_units:
                        inflow_count = float(round(inflow_count))
                    if inflow_count > 0:
                        if ck not in stock_by_age:
                            stock_by_age[ck] = {}
                        stock_by_age[ck][0] = stock_by_age[ck].get(0, 0.0) + inflow_count
                        year_inflow[ck] = year_inflow.get(ck, 0.0) + inflow_count

                    # 3b. Apply user-provided outflow. Cohort-specific (by age)
                    # when the upload carried ``age``/``birth_year``; otherwise
                    # FIFO from the oldest age.
                    if outflow_data is None:
                        continue
                    cohort_total = outflow_data.counts.get(ck, 0.0)
                    if integer_units:
                        cohort_total = float(round(cohort_total))
                    if cohort_total <= 0:
                        continue
                    if ck not in stock_by_age:
                        stock_by_age[ck] = {}

                    removed_total = 0.0
                    age_specific = {
                        int(k.rsplit("|", 1)[1]): v
                        for k, v in outflow_data.cohort_age_counts.items()
                        if k.rsplit("|", 1)[0] == ck
                    }
                    if age_specific and integer_units:
                        # Keep integer allocations aligned with the rounded
                        # cohort total so mass balance holds.
                        age_specific = {
                            a: v for a, v in largest_remainder_round(
                                age_specific, target_total=cohort_total
                            ).items()
                        }
                    if age_specific:
                        for age, requested in sorted(age_specific.items()):
                            available = stock_by_age[ck].get(age, 0.0)
                            take = min(available, requested)
                            new_count = available - take
                            if new_count <= 1e-12:
                                stock_by_age[ck].pop(age, None)
                            else:
                                stock_by_age[ck][age] = new_count
                            if take > 0:
                                outflow_by_age[ck][age] = (
                                    outflow_by_age[ck].get(age, 0.0) + take
                                )
                            if take < requested - 1e-9:
                                manual_warnings.append(
                                    f"Year {year}: requested {requested:.4g} outflows at age {age} for "
                                    f"cohort '{ck}' but only {available:.4g} available."
                                )
                            removed_total += take
                    else:
                        remaining = cohort_total
                        ages_desc = sorted(stock_by_age[ck].keys(), reverse=True)
                        for age in ages_desc:
                            if remaining <= 1e-12:
                                break
                            available = stock_by_age[ck].get(age, 0.0)
                            take = min(available, remaining)
                            new_count = available - take
                            if new_count <= 1e-12:
                                stock_by_age[ck].pop(age, None)
                            else:
                                stock_by_age[ck][age] = new_count
                            if take > 0:
                                outflow_by_age[ck][age] = (
                                    outflow_by_age[ck].get(age, 0.0) + take
                                )
                            removed_total += take
                            remaining -= take
                        if remaining > 1e-6:
                            manual_warnings.append(
                                f"Year {year}: manual outflow of {cohort_total:.4g} for cohort "
                                f"'{ck}' exceeds available stock by {remaining:.4g} — excess ignored."
                            )
                    year_manual[ck] = year_manual.get(ck, 0.0) + removed_total

            # 4. Snapshot.
            year_outflow = {
                ck: (
                    year_natural.get(ck, 0.0)
                    + year_forced.get(ck, 0.0)
                    + year_manual.get(ck, 0.0)
                )
                for ck in self.cohort_keys
            }
            stock_total = {ck: sum(ages.values()) for ck, ages in stock_by_age.items()}
            total_in += sum(year_inflow.values())
            total_out += sum(year_outflow.values())

            years_out.append(
                YearResult(
                    year=year,
                    stock=stock_total,
                    stock_by_age={ck: dict(ages) for ck, ages in stock_by_age.items()},
                    inflow=year_inflow,
                    outflow=year_outflow,
                    outflow_by_age=outflow_by_age,
                    natural_outflow=dict(year_natural),
                    forced_retirement=dict(year_forced),
                    forced_retirement_by_age=forced_by_age,
                    manual_outflow=dict(year_manual),
                )
            )

        total_end = sum(sum(d.values()) for d in stock_by_age.values())
        warnings = _fleet_drift_warnings(years_out, total_start) + manual_warnings
        return SimulationResult(
            system_id=self.state.system_id,
            years=years_out,
            summary=SimulationSummary(
                total_stock_start=total_start,
                total_stock_end=total_end,
                total_inflows=total_in,
                total_outflows=total_out,
                warnings=warnings,
            ),
        )
