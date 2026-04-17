"""Pure-Python dynamic stock model engine for MApper's MFA module.

This module is intentionally framework-agnostic: no FastAPI, no global state,
no IO of its own beyond CSV/XLSX parsing helpers. The ``DynamicStockModel`` class
takes a system definition + state and returns a fully-resolved
``SimulationResult``.

Mass balance per year:
    stock(t+1) = stock(t) - outflow(t) + inflow(t)

Cohort tracking is done by keeping ``stock_by_age[cohort_key][age] = count``.
A cohort_key is the pipe-delimited concatenation of non-age dimension labels
in the order the dimensions were declared on the system (e.g. ``BEV|Small``).
"""
from __future__ import annotations

import csv
import io
import math
import os
from itertools import product

from openpyxl import load_workbook

from mapper.models.mfa_schemas import (
    DimensionDef,
    InflowData,
    MFASystemState,
    SimulationResult,
    SimulationSummary,
    SurvivalConfig,
    SurvivalPreviewPoint,
    SystemDefinition,
    YearResult,
)


COHORT_SEP = "|"
DEFAULT_WEIBULL_SHAPE = 4.0
DEFAULT_WEIBULL_SCALE = 15.0


# ── Math helpers ─────────────────────────────────────────────────────────────


def weibull_survival(age: float, shape: float, scale: float) -> float:
    """S(a) = exp(-(a/scale)^shape).

    S(0) = 1.0; S decreases monotonically as age grows.
    """
    if age <= 0:
        return 1.0
    if scale <= 0 or shape <= 0:
        return 1.0
    return math.exp(-((age / scale) ** shape))


def weibull_hazard(age: float, shape: float, scale: float) -> float:
    """Conditional probability of failure at integer ``age``, given survival to ``age - 1``.

    h(a) = 1 - S(a)/S(a-1). Returns 0 when ``age <= 0``.
    """
    if age <= 0:
        return 0.0
    s_prev = weibull_survival(age - 1, shape, scale)
    if s_prev <= 0:
        return 1.0
    s_curr = weibull_survival(age, shape, scale)
    h = 1.0 - (s_curr / s_prev)
    return max(0.0, min(1.0, h))


def survival_curve(shape: float, scale: float, max_age: int) -> list[SurvivalPreviewPoint]:
    """Return S(a) and h(a) at each integer age in [0, max_age]."""
    out: list[SurvivalPreviewPoint] = []
    for a in range(max_age + 1):
        out.append(
            SurvivalPreviewPoint(
                age=a,
                survival_rate=weibull_survival(a, shape, scale),
                hazard_rate=weibull_hazard(a, shape, scale),
            )
        )
    return out


def custom_curve_survival(curve: list[tuple[int, float]], age: int) -> float:
    """Linear interpolation between custom (age, survival_rate) points."""
    if not curve:
        return 1.0
    pts = sorted(curve, key=lambda p: p[0])
    if age <= pts[0][0]:
        return pts[0][1]
    if age >= pts[-1][0]:
        return pts[-1][1]
    for (a0, s0), (a1, s1) in zip(pts, pts[1:]):
        if a0 <= age <= a1:
            if a1 == a0:
                return s0
            t = (age - a0) / (a1 - a0)
            return s0 + t * (s1 - s0)
    return pts[-1][1]


def custom_curve_hazard(curve: list[tuple[int, float]], age: int) -> float:
    if age <= 0:
        return 0.0
    s_prev = custom_curve_survival(curve, age - 1)
    if s_prev <= 0:
        return 1.0
    s_curr = custom_curve_survival(curve, age)
    return max(0.0, min(1.0, 1.0 - (s_curr / s_prev)))


# ── Cohort key utilities ─────────────────────────────────────────────────────


def non_age_dimensions(dims: list[DimensionDef]) -> list[DimensionDef]:
    return [d for d in dims if not d.is_age]


def all_cohort_keys(dims: list[DimensionDef]) -> list[str]:
    """Cartesian product of labels of every non-age dimension."""
    nads = non_age_dimensions(dims)
    if not nads:
        return [""]
    label_lists = [d.labels for d in nads]
    return [COHORT_SEP.join(combo) for combo in product(*label_lists)]


def cohort_key_to_dict(cohort_key: str, dims: list[DimensionDef]) -> dict[str, str]:
    nads = non_age_dimensions(dims)
    parts = cohort_key.split(COHORT_SEP) if cohort_key else []
    return {d.name: parts[i] for i, d in enumerate(nads) if i < len(parts)}


def dict_to_cohort_key(values: dict[str, str], dims: list[DimensionDef]) -> str:
    nads = non_age_dimensions(dims)
    return COHORT_SEP.join(values.get(d.name, "") for d in nads)


# ── Survival lookup ──────────────────────────────────────────────────────────


def _config_matches(cfg: SurvivalConfig, cohort_dict: dict[str, str]) -> bool:
    return all(cohort_dict.get(k) == v for k, v in cfg.dimension_filters.items())


def _config_specificity(cfg: SurvivalConfig) -> int:
    return len(cfg.dimension_filters)


def best_config_for_cohort(
    cohort_dict: dict[str, str], configs: list[SurvivalConfig]
) -> SurvivalConfig | None:
    """Most specific config whose dimension_filters match this cohort."""
    matching = [c for c in configs if _config_matches(c, cohort_dict)]
    if not matching:
        return None
    return max(matching, key=_config_specificity)


# ── CSV / XLSX parsing ───────────────────────────────────────────────────────


SUPPORTED_EXTS = (".csv", ".xlsx", ".xls")


def _sniff_delimiter(text: str) -> str:
    sample = "\n".join(text.splitlines()[:5])
    if sample.count(";") > sample.count(","):
        return ";"
    return ","


def _strip_bom(text: str) -> str:
    if text.startswith("\ufeff"):
        return text[1:]
    return text


def _validate_label(value: str, dim: DimensionDef) -> None:
    if value not in dim.labels:
        raise ValueError(
            f"Value '{value}' is not a valid label for dimension '{dim.name}'. "
            f"Allowed: {dim.labels}"
        )


def _cell_to_str(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _read_rows_from_bytes(
    content: bytes, filename: str
) -> tuple[list[str], list[dict[str, str]]]:
    """Read a CSV or XLSX file into (headers, list-of-dicts).

    Cells are stringified + trimmed. Fully empty trailing rows are skipped.
    Raises ``ValueError`` on unsupported extensions or parse errors.
    """
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        text = _strip_bom(text)
        delim = _sniff_delimiter(text)
        reader = csv.DictReader(io.StringIO(text), delimiter=delim)
        headers = [h.strip() for h in (reader.fieldnames or [])]
        rows: list[dict[str, str]] = []
        for raw in reader:
            row = {
                (k.strip() if k else ""): (v.strip() if isinstance(v, str) else ("" if v is None else str(v).strip()))
                for k, v in raw.items() if k
            }
            if any(row.values()):
                rows.append(row)
        return headers, rows
    if ext in (".xlsx", ".xls"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}") from e
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no sheets.")
        iterator = ws.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            raise ValueError("Excel sheet is empty.")
        headers = [_cell_to_str(c) for c in header_row]
        rows = []
        for raw in iterator:
            if not raw:
                continue
            row: dict[str, str] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                row[h] = _cell_to_str(raw[i]) if i < len(raw) else ""
            if any(row.values()):
                rows.append(row)
        return headers, rows
    raise ValueError(f"Unsupported file extension '{ext}'. Use .csv or .xlsx.")


def parse_stock_file(
    content: bytes, filename: str, dims: list[DimensionDef]
) -> tuple[dict[str, float], int]:
    """Parse the initial-stock CSV or XLSX.

    Expected columns: every non-age dimension name, ``age``, ``count``.
    Returns (initial_stock keyed by ``cohort_key|age``, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = [d.name for d in nads] + ["age", "count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Stock file missing required column(s): {missing}. Got headers: {headers}"
        )

    out: dict[str, float] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        try:
            age = int(float(row["age"]))
        except ValueError as e:
            raise ValueError(f"Invalid age '{row['age']}' on row {rows + 1}") from e
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        full_key = f"{cohort_key}|{age}"
        out[full_key] = out.get(full_key, 0.0) + count
        rows += 1
    return out, rows


def parse_inflow_file(
    content: bytes, filename: str, dims: list[DimensionDef], horizon_years: list[int]
) -> tuple[list[InflowData], int]:
    """Parse the inflow CSV or XLSX.

    Expected columns: ``year``, every non-age dimension name, ``count``.
    Returns (list of InflowData per year, rows_parsed).
    """
    headers, rows_data = _read_rows_from_bytes(content, filename)

    nads = non_age_dimensions(dims)
    expected = ["year"] + [d.name for d in nads] + ["count"]
    missing = [c for c in expected if c not in headers]
    if missing:
        raise ValueError(
            f"Inflow file missing required column(s): {missing}. Got headers: {headers}"
        )

    horizon = set(horizon_years)
    by_year: dict[int, dict[str, float]] = {}
    rows = 0
    for row in rows_data:
        if not row.get("count"):
            continue
        try:
            year = int(float(row["year"]))
        except ValueError as e:
            raise ValueError(f"Invalid year '{row['year']}' on row {rows + 1}") from e
        if year not in horizon:
            raise ValueError(
                f"Year {year} on row {rows + 1} is outside the system's time horizon."
            )
        try:
            count = float(row["count"])
        except ValueError as e:
            raise ValueError(f"Invalid count '{row['count']}' on row {rows + 1}") from e
        cohort_values: dict[str, str] = {}
        for d in nads:
            v = row.get(d.name, "")
            _validate_label(v, d)
            cohort_values[d.name] = v
        cohort_key = dict_to_cohort_key(cohort_values, dims)
        bucket = by_year.setdefault(year, {})
        bucket[cohort_key] = bucket.get(cohort_key, 0.0) + count
        rows += 1

    out = [InflowData(year=y, counts=by_year[y]) for y in sorted(by_year)]
    return out, rows


def parse_first_column_labels(content: bytes, filename: str) -> list[str]:
    """Extract unique values from the first column of a CSV or XLSX file.

    Strips the header row if the first value looks non-numeric and matches a
    typical header pattern (lowercased, snake/kebab case). Since we can't
    reliably identify a header without context, we simply treat the first row
    as a header and drop it unless it turns out to be the only row.
    """
    ext = os.path.splitext(filename)[1].lower()
    values: list[str] = []
    if ext == ".csv":
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        text = _strip_bom(text)
        delim = _sniff_delimiter(text)
        reader = csv.reader(io.StringIO(text), delimiter=delim)
        for row in reader:
            if not row:
                continue
            cell = (row[0] or "").strip()
            if cell:
                values.append(cell)
    elif ext in (".xlsx", ".xls"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        except Exception as e:
            raise ValueError(f"Could not read Excel file: {e}") from e
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no sheets.")
        for raw in ws.iter_rows(values_only=True):
            if not raw:
                continue
            cell = _cell_to_str(raw[0])
            if cell:
                values.append(cell)
    else:
        raise ValueError(f"Unsupported file extension '{ext}'. Use .csv or .xlsx.")

    # Drop the header row when there is more than one row.
    if len(values) > 1:
        values = values[1:]

    # Preserve order while deduplicating.
    seen: set[str] = set()
    out: list[str] = []
    for v in values:
        if v not in seen:
            seen.add(v)
            out.append(v)
    return out


# ── CSV templates ────────────────────────────────────────────────────────────


def stock_template_csv(dims: list[DimensionDef], example_ages: int = 5) -> str:
    nads = non_age_dimensions(dims)
    headers = [d.name for d in nads] + ["age", "count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for combo in product(*(d.labels for d in nads)) if nads else [()]:
        for age in range(example_ages + 1):
            writer.writerow(list(combo) + [age, ""])
    return buf.getvalue()


def inflow_template_csv(dims: list[DimensionDef], horizon_years: list[int]) -> str:
    nads = non_age_dimensions(dims)
    headers = ["year"] + [d.name for d in nads] + ["count"]
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    for year in horizon_years:
        for combo in product(*(d.labels for d in nads)) if nads else [()]:
            writer.writerow([year] + list(combo) + [""])
    return buf.getvalue()


# ── Engine ───────────────────────────────────────────────────────────────────


class DynamicStockModel:
    """Year-by-year cohort-tracking stock-flow model.

    Public surface:
        model = DynamicStockModel(system, state)
        result = model.simulate()
    """

    def __init__(self, system: SystemDefinition, state: MFASystemState) -> None:
        self.system = system
        self.state = state
        self.years: list[int] = system.time_horizon.years
        self.cohort_keys: list[str] = all_cohort_keys(system.dimensions)
        self.max_age: int = system.time_horizon.length  # safe upper bound

    # — Survival lookup —

    def _survival_params(
        self, cohort_key: str
    ) -> tuple[str, float, float, list[tuple[int, float]] | None]:
        cohort_dict = cohort_key_to_dict(cohort_key, self.system.dimensions)
        cfg = best_config_for_cohort(cohort_dict, self.state.survival_configs)
        if cfg is None:
            return ("weibull", DEFAULT_WEIBULL_SHAPE, DEFAULT_WEIBULL_SCALE, None)
        if cfg.method == "custom" and cfg.custom_curve:
            curve = [(p.age, p.survival_rate) for p in cfg.custom_curve]
            return ("custom", 0.0, 0.0, curve)
        shape = cfg.weibull_shape if cfg.weibull_shape is not None else DEFAULT_WEIBULL_SHAPE
        scale = cfg.weibull_scale if cfg.weibull_scale is not None else DEFAULT_WEIBULL_SCALE
        return ("weibull", shape, scale, None)

    def hazard(self, cohort_key: str, age: int) -> float:
        method, shape, scale, curve = self._survival_params(cohort_key)
        if method == "custom" and curve is not None:
            return custom_curve_hazard(curve, age)
        return weibull_hazard(age, shape, scale)

    # — Initialization —

    def _initial_stock_by_age(self) -> dict[str, dict[int, float]]:
        out: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}
        for full_key, count in self.state.initial_stock.items():
            try:
                cohort_key, age_str = full_key.rsplit("|", 1)
                age = int(age_str)
            except ValueError:
                # Treat keys with no age as age 0.
                cohort_key, age = full_key, 0
            if cohort_key not in out:
                out[cohort_key] = {}
            out[cohort_key][age] = out[cohort_key].get(age, 0.0) + count
        return out

    def _inflow_for_year(self, year: int) -> dict[str, float]:
        for inf in self.state.inflows:
            if inf.year == year:
                return dict(inf.counts)
        return {}

    # — Main loop —

    def simulate(self) -> SimulationResult:
        stock_by_age = self._initial_stock_by_age()

        years_out: list[YearResult] = []
        total_in = 0.0
        total_out = 0.0
        total_start = sum(sum(d.values()) for d in stock_by_age.values())

        for idx, year in enumerate(self.years):
            year_inflow: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            year_outflow: dict[str, float] = {ck: 0.0 for ck in self.cohort_keys}
            outflow_by_age: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}

            # 1. Age every existing cohort by one year. Skipped in year 1:
            # the initial stock comes from the CSV already at its correct ages.
            if idx > 0:
                aged: dict[str, dict[int, float]] = {ck: {} for ck in self.cohort_keys}
                for ck, ages in stock_by_age.items():
                    for a, count in ages.items():
                        aged.setdefault(ck, {})[a + 1] = aged.get(ck, {}).get(a + 1, 0.0) + count
                stock_by_age = aged

            # 2. Apply survival hazard age-by-age (every year, including year 1
            # so that already-old initial stock can still retire).
            for ck, ages in stock_by_age.items():
                new_ages: dict[int, float] = {}
                for a, count in ages.items():
                    h = self.hazard(ck, a)
                    out = count * h
                    survived = count - out
                    if out > 0:
                        outflow_by_age[ck][a] = out
                        year_outflow[ck] = year_outflow.get(ck, 0.0) + out
                    if survived > 0:
                        new_ages[a] = survived
                stock_by_age[ck] = new_ages

            # 3. Inflows enter at age 0.
            inflows = self._inflow_for_year(year)
            for ck, count in inflows.items():
                if count <= 0:
                    continue
                if ck not in stock_by_age:
                    stock_by_age[ck] = {}
                stock_by_age[ck][0] = stock_by_age[ck].get(0, 0.0) + count
                year_inflow[ck] = year_inflow.get(ck, 0.0) + count

            # 4. Snapshot.
            stock_total = {ck: sum(ages.values()) for ck, ages in stock_by_age.items()}
            total_in += sum(year_inflow.values())
            total_out += sum(year_outflow.values())

            years_out.append(
                YearResult(
                    year=year,
                    stock=stock_total,
                    stock_by_age={ck: dict(ages) for ck, ages in stock_by_age.items()},
                    inflow=year_inflow,
                    outflow=year_outflow,
                    outflow_by_age=outflow_by_age,
                )
            )

        total_end = sum(sum(d.values()) for d in stock_by_age.values())
        return SimulationResult(
            system_id=self.state.system_id,
            years=years_out,
            summary=SimulationSummary(
                total_stock_start=total_start,
                total_stock_end=total_end,
                total_inflows=total_in,
                total_outflows=total_out,
            ),
        )
