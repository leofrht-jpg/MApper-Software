"""Round-trip export→import test for MaterialEvolution on the xlsx pipeline.

Builds an Archetype with one learning_rate material and one milestones
material, exports it via the in-memory Workbook builders, re-imports via
``_parse_bom_workbook``, then asserts evolution data survives byte-for-byte.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api.bom import _build_export_workbook, _parse_bom_workbook
from mapper.models.bom_schemas import (
    Archetype,
    BOMNode,
    EcoinventLink,
    MaterialEvolution,
    QuantityMilestone,
)


def _build_fixture() -> Archetype:
    link = EcoinventLink(database="ei-3.9", code="XYZ", name="steel", location="GLO")
    steel_fixed = BOMNode(
        id="m-steel", name="Steel frame", node_type="material",
        quantity=900, unit="kg", ecoinvent_activity=link,
    )
    alu_lr = BOMNode(
        id="m-alu", name="Aluminium panels", node_type="material",
        quantity=120, unit="kg", ecoinvent_activity=link,
        evolution=MaterialEvolution(method="learning_rate", learning_rate=-0.025, base_year=2025),
    )
    lfp_ms = BOMNode(
        id="m-lfp", name="LFP cells", node_type="material",
        quantity=300, unit="kg", ecoinvent_activity=link,
        evolution=MaterialEvolution(
            method="milestones",
            milestones=[
                QuantityMilestone(year=2025, quantity=300),
                QuantityMilestone(year=2035, quantity=230),
                QuantityMilestone(year=2050, quantity=180),
            ],
        ),
    )
    elec_rb = BOMNode(
        id="m-elec", name="Electricity (lifetime)", node_type="material",
        quantity=30000, unit="kWh", ecoinvent_activity=link,
        evolution=MaterialEvolution(
            method="rebound_effect",
            rebound_rate=0.015,
            base_year=2025,
            applies_to_stages=["Use Phase"],
        ),
    )
    body = BOMNode(id="s-body", name="Body", node_type="component", quantity=1, unit="piece",
                   children=[steel_fixed, alu_lr])
    battery = BOMNode(id="s-bat", name="Battery Pack", node_type="component", quantity=1, unit="piece",
                      children=[lfp_ms])
    use = BOMNode(id="s-use", name="Use Phase", node_type="component", quantity=1, unit="piece",
                  children=[elec_rb])
    return Archetype(
        id="arc-1", name="BEV-LFP", description="round-trip fixture",
        category=None, folder="Cars/BEVs", bom=[body, battery, use],
    )


def _find(nodes: list[BOMNode], name: str) -> BOMNode:
    for n in nodes:
        if n.name == name:
            return n
        if n.children:
            try:
                return _find(n.children, name)
            except KeyError:
                continue
    raise KeyError(name)


def main() -> None:
    arc = _build_fixture()
    wb = _build_export_workbook(arc)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    print("✓ export succeeded")

    reopened = load_workbook(buf, data_only=True)
    assert "BOM" in reopened.sheetnames
    assert "Timeline" in reopened.sheetnames, "Timeline sheet missing when evolution present"

    # Inspect the BOM sheet directly to confirm evolution columns are populated.
    ws = reopened["BOM"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c or "") for c in rows[0]]
    idx = {c: header.index(c) for c in header}
    by_name = {r[idx["Name"]]: r for r in rows[1:]}

    alu = by_name["Aluminium panels"]
    assert alu[idx["Evolution Method"]] == "learning_rate", alu[idx["Evolution Method"]]
    assert abs(float(alu[idx["Learning Rate"]]) - (-0.025)) < 1e-9
    assert alu[idx["Rebound Rate"]] in ("", None)
    assert int(alu[idx["Base Year"]]) == 2025
    print(f"✓ learning_rate row wrote: method={alu[idx['Evolution Method']]}, "
          f"rate={alu[idx['Learning Rate']]}, base={alu[idx['Base Year']]}")

    elec = by_name["Electricity (lifetime)"]
    assert elec[idx["Evolution Method"]] == "rebound_effect", elec[idx["Evolution Method"]]
    assert abs(float(elec[idx["Rebound Rate"]]) - 0.015) < 1e-9
    assert elec[idx["Learning Rate"]] in ("", None)
    assert int(elec[idx["Base Year"]]) == 2025
    assert elec[idx["Rebound Applies To Stages"]] == "Use Phase", elec[idx["Rebound Applies To Stages"]]
    print(f"✓ rebound_effect row wrote: method={elec[idx['Evolution Method']]}, "
          f"rate={elec[idx['Rebound Rate']]}, base={elec[idx['Base Year']]}, "
          f"stages={elec[idx['Rebound Applies To Stages']]}")

    lfp = by_name["LFP cells"]
    assert lfp[idx["Evolution Method"]] == "milestones"
    assert lfp[idx["Milestone Years"]] == "2025;2035;2050"
    assert lfp[idx["Milestone Values"]] == "300;230;180"
    print(f"✓ milestones row wrote: years={lfp[idx['Milestone Years']]}, "
          f"values={lfp[idx['Milestone Values']]}")

    steel = by_name["Steel frame"]
    assert steel[idx["Evolution Method"]] == "fixed"
    assert steel[idx["Learning Rate"]] in ("", None)
    print("✓ fixed material row wrote: method=fixed, others empty")

    # Component rows leave the column blank — only materials advertise it.
    body_row = by_name["Body"]
    assert body_row[idx["Evolution Method"]] in ("", None)
    print("✓ component row wrote empty Evolution Method")

    # Re-import via _parse_bom_workbook (single-archetype format path).
    buf.seek(0)
    wb2 = load_workbook(buf, data_only=True)
    roots, warnings = _parse_bom_workbook(wb2)
    if warnings:
        print("  import warnings:", warnings)
    assert not warnings, f"unexpected import warnings: {warnings}"

    alu2 = _find(roots, "Aluminium panels")
    assert alu2.evolution is not None
    assert alu2.evolution.method == "learning_rate"
    assert abs((alu2.evolution.learning_rate or 0) - (-0.025)) < 1e-9
    assert alu2.evolution.base_year == 2025
    print("✓ learning_rate round-trip: method/rate/base_year preserved")

    elec2 = _find(roots, "Electricity (lifetime)")
    assert elec2.evolution is not None, "rebound material lost evolution on re-import"
    assert elec2.evolution.method == "rebound_effect"
    assert abs((elec2.evolution.rebound_rate or 0) - 0.015) < 1e-9
    assert elec2.evolution.base_year == 2025
    assert elec2.evolution.applies_to_stages == ["Use Phase"], elec2.evolution.applies_to_stages
    print("✓ rebound_effect round-trip: method/rebound_rate/base_year/applies_to_stages preserved")

    lfp2 = _find(roots, "LFP cells")
    assert lfp2.evolution is not None
    assert lfp2.evolution.method == "milestones"
    assert lfp2.evolution.milestones is not None
    ms = [(m.year, m.quantity) for m in lfp2.evolution.milestones]
    assert ms == [(2025, 300.0), (2035, 230.0), (2050, 180.0)], ms
    print("✓ milestones round-trip: year/value pairs preserved and sorted")

    steel2 = _find(roots, "Steel frame")
    assert steel2.evolution is None, "material with 'fixed' should yield evolution=None"
    print("✓ fixed round-trip: evolution=None")

    print("\nRound-trip test PASSED — all evolution data survived export→import.")


if __name__ == "__main__":
    main()
