"""Tests for the N-layer downscaling chain, preset migration, and xlsx I/O."""
from __future__ import annotations

import io

from mapper.core.aesa_engine import (
    AESAEngine,
    BUILTIN_PRINCIPLES,
    build_carbon_budget,
    build_default_multi_d_config,
    build_default_sharing_preset,
    load_boundary_sets,
    migrate_multi_d_to_preset,
    resolve_sharing,
    suggest_method_mapping,
)
from mapper.models.aesa_schemas import (
    AESAConfiguration,
    CategoryAssignment,
    DownscalingChain,
    DownscalingLayer,
    PrincipleDefinition,
    SharingPreset,
)
from mapper.models.bom_schemas import DSMLCAResult, DSMLCASummary, DSMLCAYearResult


# ─── fixtures ────────────────────────────────────────────────────────────────


def _impact_climate(total: float = 1.0e10, year: int = 2025) -> DSMLCAResult:
    return DSMLCAResult(
        mfa_system_id="sys",
        method=["EF v3.1", "climate change", "global warming potential (GWP100)"],
        method_label="EF v3.1 › climate change › GWP100",
        scope="stock",
        unit="kg CO2 eq",
        years=[DSMLCAYearResult(
            year=year, total_impact=total, unit="kg CO2 eq",
            impact_by_cohort={}, impact_by_material={}, count_by_cohort={},
        )],
        summary=DSMLCASummary(total_impact=total, peak_year=year, peak_impact=total),
    )


def _impact_acidification(total: float = 1.0e8, year: int = 2025) -> DSMLCAResult:
    return DSMLCAResult(
        mfa_system_id="sys",
        method=["EF v3.1", "acidification", "accumulated exceedance (AE)"],
        method_label="EF v3.1 › acidification › AE",
        scope="stock",
        unit="mol H+ eq",
        years=[DSMLCAYearResult(
            year=year, total_impact=total, unit="mol H+ eq",
            impact_by_cohort={}, impact_by_material={}, count_by_cohort={},
        )],
        summary=DSMLCASummary(total_impact=total, peak_year=year, peak_impact=total),
    )


def _single_layer_preset(principle: str, sys_val: float, glob_val: float,
                         pb_id: str = "acidification") -> SharingPreset:
    return SharingPreset(
        id="single",
        name="Single layer",
        principles=[PrincipleDefinition(id=principle, name=principle)],
        category_assignments=[CategoryAssignment(pb_id=pb_id, principle_id=principle)],
        chain=DownscalingChain(layers=[DownscalingLayer(
            layer_number=1, name="Only layer", principle_mode="category_specific",
            data={principle: {2025: (sys_val, glob_val)}},
        )]),
    )


def _cfg_with_preset(preset: SharingPreset, mapping, *, with_carbon=False) -> AESAConfiguration:
    return AESAConfiguration(
        id="cfg",
        name="test",
        mfa_system_id="sys",
        boundary_set_id="Sala2020_EF",
        multi_d=None,
        sharing=preset,
        carbon_budget=build_carbon_budget() if with_carbon else None,
        method_mapping=mapping,
        created_at="2025-01-01T00:00:00Z",
    )


# ─── chain compute ───────────────────────────────────────────────────────────


def test_single_layer_chain() -> None:
    preset = _single_layer_preset("EpC", 100.0, 1000.0)
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    config = _cfg_with_preset(preset, mapping)
    impact = _impact_acidification(total=1.0e8)
    out = AESAEngine.compute([impact], config, bset)
    r = next(r for r in out.results if r.pb_id == "acidification")
    expected_factor = 0.1
    expected_allocated = 1.0e12 * expected_factor
    assert abs(r.total_sharing_factor - expected_factor) < 1e-9
    assert abs(r.allocated_sos - expected_allocated) < 1e-3
    assert len(r.layer_factors) == 1
    assert r.sharing_principle == "EpC"


def test_three_layer_chain_product() -> None:
    """3-layer chain: factor = 0.5 × 0.4 × 0.2 = 0.04."""
    preset = SharingPreset(
        id="three",
        name="3-layer",
        principles=[PrincipleDefinition(id="EpC", name="EpC"),
                    PrincipleDefinition(id="AR", name="AR")],
        category_assignments=[CategoryAssignment(pb_id="acidification", principle_id="EpC")],
        chain=DownscalingChain(layers=[
            DownscalingLayer(layer_number=1, name="L1",
                             principle_mode="category_specific",
                             data={"EpC": {2025: (0.5, 1.0)}}),
            DownscalingLayer(layer_number=2, name="L2", principle_mode="fixed",
                             fixed_principle="AR",
                             data={"AR": {2025: (0.4, 1.0)}}),
            DownscalingLayer(layer_number=3, name="L3", principle_mode="fixed",
                             fixed_principle="AR",
                             data={"AR": {2025: (0.2, 1.0)}}),
        ]),
    )
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    config = _cfg_with_preset(preset, mapping)
    out = AESAEngine.compute([_impact_acidification()], config, bset)
    r = out.results[0]
    assert abs(r.total_sharing_factor - 0.04) < 1e-12
    assert r.layer_factors == [0.5, 0.4, 0.2]
    # Legacy fields: l1 = first, l2 = product of the rest
    assert abs(r.sharing_factor_l1 - 0.5) < 1e-12
    assert abs(r.sharing_factor_l2 - 0.08) < 1e-12


def test_custom_principle() -> None:
    """Custom principle 'GDP' not in built-ins."""
    preset = SharingPreset(
        id="gdp",
        name="GDP",
        principles=[PrincipleDefinition(id="GDP", name="GDP share")],
        category_assignments=[CategoryAssignment(pb_id="acidification", principle_id="GDP")],
        chain=DownscalingChain(layers=[DownscalingLayer(
            layer_number=1, name="GDP layer", principle_mode="category_specific",
            data={"GDP": {2025: (4.0e11, 1.0e14)}},
        )]),
    )
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    out = AESAEngine.compute([_impact_acidification()],
                             _cfg_with_preset(preset, mapping), bset)
    assert out.results[0].sharing_principle == "GDP"
    assert abs(out.results[0].total_sharing_factor - 4.0e-3) < 1e-12


def test_time_varying_data_exact_and_nearest() -> None:
    preset = _single_layer_preset("EpC", 100.0, 1000.0)
    # Add a 2030 entry with a different value
    preset.chain.layers[0].data["EpC"][2030] = (200.0, 1000.0)
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    # Exact match at 2030 → factor 0.2
    out2030 = AESAEngine.compute(
        [_impact_acidification(year=2030)],
        _cfg_with_preset(preset, mapping), bset,
    )
    assert abs(out2030.results[0].total_sharing_factor - 0.2) < 1e-9
    # 2027 has no entry — nearest is 2025 (distance 2) vs 2030 (distance 3) → use 2025
    out2027 = AESAEngine.compute(
        [_impact_acidification(year=2027)],
        _cfg_with_preset(preset, mapping), bset,
    )
    assert abs(out2027.results[0].total_sharing_factor - 0.1) < 1e-9


def test_single_year_acts_as_constant() -> None:
    preset = _single_layer_preset("EpC", 10.0, 100.0)  # 0.1
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    # Year 2040, data only has 2025 → constant factor 0.1
    out = AESAEngine.compute(
        [_impact_acidification(year=2040)],
        _cfg_with_preset(preset, mapping), bset,
    )
    assert abs(out.results[0].total_sharing_factor - 0.1) < 1e-9


# ─── migration ───────────────────────────────────────────────────────────────


def test_migrate_multi_d_preserves_output() -> None:
    """A config using legacy multi_d must produce the same SR as the migrated
    equivalent using .sharing."""
    multi_d = build_default_multi_d_config()
    bset = load_boundary_sets()["Sala2020_EF"]
    methods = [["EF v3.1", "acidification", "accumulated exceedance (AE)"]]
    mapping = suggest_method_mapping(methods, bset)
    legacy = AESAConfiguration(
        id="legacy", name="legacy", mfa_system_id="sys",
        boundary_set_id="Sala2020_EF",
        multi_d=multi_d, sharing=None,
        method_mapping=mapping, created_at="2025-01-01T00:00:00Z",
    )
    migrated_preset = migrate_multi_d_to_preset(multi_d)
    new = AESAConfiguration(
        id="new", name="new", mfa_system_id="sys",
        boundary_set_id="Sala2020_EF",
        multi_d=None, sharing=migrated_preset,
        method_mapping=mapping, created_at="2025-01-01T00:00:00Z",
    )
    out_legacy = AESAEngine.compute([_impact_acidification()], legacy, bset)
    out_new = AESAEngine.compute([_impact_acidification()], new, bset)
    assert abs(out_legacy.results[0].sr - out_new.results[0].sr) < 1e-9
    assert out_legacy.results[0].sharing_principle == out_new.results[0].sharing_principle


def test_resolve_sharing_falls_back_to_default_when_empty() -> None:
    cfg = AESAConfiguration(
        id="empty", name="empty", mfa_system_id="sys",
        boundary_set_id="Sala2020_EF", multi_d=None, sharing=None,
        method_mapping=[], created_at="2025-01-01T00:00:00Z",
    )
    preset = resolve_sharing(cfg)
    assert preset.built_in is True
    assert len(preset.chain.layers) == 3


# ─── xlsx round-trip ─────────────────────────────────────────────────────────


def test_xlsx_round_trip() -> None:
    from mapper.api.aesa import _build_sharing_workbook, _parse_sharing_workbook
    from openpyxl import load_workbook

    preset = build_default_sharing_preset()
    wb = _build_sharing_workbook(preset, include_instructions=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = load_workbook(buf)
    parsed = _parse_sharing_workbook(wb2, "roundtrip")

    assert {p.id for p in parsed.principles} == {"EpC", "IN", "AGR", "LA", "AR"}
    assert len(parsed.category_assignments) == 16
    assert len(parsed.chain.layers) == 3
    # Layer 2 (fixed AR) matches
    l2 = parsed.chain.layers[1]
    assert l2.principle_mode == "fixed"
    assert l2.fixed_principle == "AR"


def test_xlsx_rejects_unknown_principle_reference() -> None:
    """Category assignment referencing a principle not declared must raise."""
    from openpyxl import Workbook
    from mapper.api.aesa import _parse_sharing_workbook

    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Principles")
    ws.append(["Principle ID", "Name", "Description"])
    ws.append(["EpC", "Equal per Capita", ""])
    ws = wb.create_sheet("Category Assignments")
    ws.append(["PB ID", "Principle ID", "Justification"])
    ws.append(["climate_change", "GDP", ""])  # GDP not declared
    ws = wb.create_sheet("Downscaling Chain")
    ws.append(["Layer", "Name", "Mode", "Fixed Principle", "Description"])
    ws.append([1, "L1", "category_specific", "", ""])
    ws = wb.create_sheet("Sharing Data")
    ws.append(["Layer", "Principle", "System Value", "Global Value", "Year", "Source"])
    ws.append([1, "EpC", 1, 10, 2025, ""])

    import pytest
    with pytest.raises(ValueError, match="unknown principle 'GDP'"):
        _parse_sharing_workbook(wb, "bad")


# ─── sensitivity with custom preset ──────────────────────────────────────────


def test_sensitivity_uses_preset_principles() -> None:
    """compute_with_sensitivity must iterate the preset's principles,
    not the built-in five."""
    # Preset with only EpC and a custom GDP principle; both have data at L1.
    preset = SharingPreset(
        id="sens",
        name="sens",
        principles=[
            PrincipleDefinition(id="EpC", name="EpC"),
            PrincipleDefinition(id="GDP", name="GDP"),
        ],
        category_assignments=[CategoryAssignment(pb_id="acidification", principle_id="EpC")],
        chain=DownscalingChain(layers=[DownscalingLayer(
            layer_number=1, name="L1", principle_mode="category_specific",
            data={
                "EpC": {2025: (100.0, 1000.0)},
                "GDP": {2025: (200.0, 1000.0)},
            },
        )]),
    )
    bset = load_boundary_sets()["Sala2020_EF"]
    mapping = suggest_method_mapping(
        [["EF v3.1", "acidification", "accumulated exceedance (AE)"]], bset,
    )
    config = _cfg_with_preset(preset, mapping)
    out = AESAEngine.compute_with_sensitivity([_impact_acidification()], config, bset)
    assert out.sensitivity is not None
    assert set(out.sensitivity.keys()) == {"EpC", "GDP"}
    # GDP variant → factor 0.2
    gdp_row = out.sensitivity["GDP"][0]
    assert abs(gdp_row.total_sharing_factor - 0.2) < 1e-9
    assert gdp_row.sharing_principle == "GDP"


if __name__ == "__main__":
    import traceback
    tests = [
        test_single_layer_chain,
        test_three_layer_chain_product,
        test_custom_principle,
        test_time_varying_data_exact_and_nearest,
        test_single_year_acts_as_constant,
        test_migrate_multi_d_preserves_output,
        test_resolve_sharing_falls_back_to_default_when_empty,
        test_xlsx_round_trip,
        test_xlsx_rejects_unknown_principle_reference,
        test_sensitivity_uses_preset_principles,
    ]
    n_pass = 0
    for t in tests:
        try:
            t()
            print(f"PASS  {t.__name__}")
            n_pass += 1
        except Exception:
            print(f"FAIL  {t.__name__}")
            traceback.print_exc()
    print(f"\n{n_pass}/{len(tests)} passed")
