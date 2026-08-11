# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Exporting a fresh configuration replaces the removed template download.

``GET /aesa/config/template`` is gone. It was a second route to the same
artefact: both it and ``POST /aesa/config/export`` called
``_build_sharing_workbook(preset, include_instructions=True, bundle=…)`` with
``kind="round_trip"``, and the two files differed in exactly two cells —

* ``sharing_preset_id``: blank in the template, recorded by the export;
* ``budget_basis``: ``CO2`` in the template, ``CO2e_GHG`` from a fresh draft.

The second is why keeping both was a liability rather than a convenience. The
CO2e basis is the correct one — it keeps the carbon budget (the SR denominator)
scope-consistent with the all-GHG GWP100 numerator, which is the entire reason
the conversion exists. The template had drifted to the wrong default and would
have kept drifting, because nothing exercised it.

These tests hold the replacement to the standard the template had to meet:
a user who exports from a clean start gets a complete, self-describing,
re-importable workbook.
"""

from __future__ import annotations

import io

import pytest
from openpyxl import load_workbook

from mapper.api.aesa import _build_sharing_workbook, _parse_aesa_config_workbook
from mapper.core.aesa_engine import build_carbon_budget, build_default_sharing_preset
from mapper.models.aesa_schemas import AESAConfigBundle

# Every sheet the removed template shipped. The export must carry all of them:
# with the template gone this is the only route to Instructions and Reference.
TEMPLATE_SHEETS = [
    "Configuration", "Principles", "Category Assignments", "Downscaling Chain",
    "Sharing Data", "Method Mapping", "Carbon Budget", "Instructions", "Reference",
]


def _fresh_draft_bundle() -> AESAConfigBundle:
    """What the UI seeds for a never-saved configuration.

    Mirrors ``draftFromDefaults``: the built-in sharing snapshot, no method
    mappings yet, and the default budget flipped to the CO2e basis.
    """
    preset = build_default_sharing_preset()
    budget = build_carbon_budget().model_copy(update={"budget_basis": "CO2e_GHG"})
    return AESAConfigBundle(
        boundary_set_id=preset.boundary_set_id,
        sharing_preset_id=preset.id,
        sharing=preset,
        method_mapping=[],
        carbon_budget=budget,
    )


@pytest.fixture(scope="module")
def fresh_export():
    bundle = _fresh_draft_bundle()
    wb = _build_sharing_workbook(
        bundle.sharing, include_instructions=True, bundle=bundle,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


# ── the template use case, end to end ───────────────────────────────────────


def test_export_from_a_fresh_draft_has_every_template_sheet(fresh_export):
    assert fresh_export.sheetnames == TEMPLATE_SHEETS


def test_instructions_and_reference_are_present_and_populated(fresh_export):
    # The user's explicit condition: the export is now the only route to these,
    # so "present" is not enough — they must actually carry their content.
    instructions = list(fresh_export["Instructions"].iter_rows(values_only=True))
    reference = list(fresh_export["Reference"].iter_rows(values_only=True))
    assert len(instructions) > 50, "Instructions sheet is present but near-empty"
    assert len(reference) > 15, "Reference sheet is present but near-empty"
    # Reference must still enumerate the values the engine accepts.
    fields = {r[0] for r in reference[1:] if r and r[0]}
    for field in ("boundary_set_id", "principle_id", "principle_mode",
                  "resolution", "carbon_budget option", "ssp_scenario", "budget_basis"):
        assert field in fields, f"Reference no longer documents {field}"


def test_reference_sheet_stays_locked(fresh_export):
    assert fresh_export["Reference"].protection.sheet is True


def test_reference_never_enumerated_sharing_preset_ids(fresh_export):
    # It did not before and must not start: a preset id is not selectable
    # anywhere in the UI, so listing them as "valid values" would document a
    # choice the user cannot make.
    fields = {r[0] for r in fresh_export["Reference"].iter_rows(values_only=True) if r and r[0]}
    assert "sharing_preset_id" not in fields


def test_fresh_export_reimports_and_reproduces_the_built_in_defaults(fresh_export):
    """The whole point: export from a clean start, import it back, get defaults."""
    bundle = _parse_aesa_config_workbook(fresh_export, "Imported")
    built_in = build_default_sharing_preset()

    assert bundle.boundary_set_id == built_in.boundary_set_id
    assert [p.id for p in bundle.sharing.principles] == [p.id for p in built_in.principles]
    assert bundle.sharing.assignments_map() == built_in.assignments_map()
    assert len(bundle.sharing.chain.layers) == len(built_in.chain.layers)
    for got, want in zip(bundle.sharing.chain.layers, built_in.chain.layers):
        assert got.principle_mode == want.principle_mode
        assert got.fixed_principle == want.fixed_principle


def test_the_exported_budget_basis_is_co2e(fresh_export):
    """Deliberate: the removed template said CO2, and that was the stale one.

    The numerator is EF v3.1 GWP100 — all greenhouse gases. A CO2-only budget
    in the denominator is a scope mismatch that inflates the climate SR. The
    CO2e basis is what makes the two comparable, which is why the conversion
    exists; a fresh draft has seeded it for some time and only the template
    lagged.
    """
    rows = {r[0]: r[1] for r in fresh_export["Carbon Budget"].iter_rows(values_only=True) if r and r[0]}
    assert rows["budget_basis"] == "CO2e_GHG"


# ── the re-lock door stays shut ─────────────────────────────────────────────


def test_import_can_never_set_built_in(fresh_export):
    """An imported workbook must not re-lock the editors that were unlocked.

    `built_in` was the read-only gate on the chain, principles and assignment
    editors. It is not a column on any sheet, and the parser hardcodes False —
    so the gate cannot return through the import door even if someone
    hand-edits the file. Pinned because "impossible by construction" and
    "impossible by accident" look identical right up until the parser changes.
    """
    bundle = _parse_aesa_config_workbook(fresh_export, "Imported")
    assert bundle.sharing.built_in is False


def test_no_sheet_carries_a_built_in_column(fresh_export):
    for name in fresh_export.sheetnames:
        for row in fresh_export[name].iter_rows(values_only=True):
            for cell in row:
                assert str(cell).strip().lower() != "built_in", (
                    f"'{name}' carries a built_in cell — the read-only gate must "
                    "not be reachable through a workbook"
                )


def test_sharing_preset_id_is_provenance_only(fresh_export):
    """Written, round-tripped, never resolved through."""
    rows = {r[0]: (r[1], r[2]) for r in fresh_export["Configuration"].iter_rows(values_only=True) if r and r[0]}
    value, note = rows["sharing_preset_id"]
    assert value == build_default_sharing_preset().id
    assert "ignore" in str(note).lower(), "the sheet should say compute ignores it"
    # And it survives a parse without being used to look anything up: the
    # sharing values come from the workbook's own sheets.
    bundle = _parse_aesa_config_workbook(fresh_export, "Imported")
    assert bundle.sharing_preset_id == value
    assert bundle.sharing.principles, "sharing came from the sheets, not a lookup"


# ── the route is gone ───────────────────────────────────────────────────────


def test_the_template_route_no_longer_exists():
    from mapper.api.aesa import router

    # Paths carry the router's /aesa prefix.
    paths = {getattr(r, "path", "") for r in router.routes}
    assert "/aesa/config/template" not in paths
    # The export it was folded into is still there.
    assert "/aesa/config/export" in paths
    assert "/aesa/config/import" in paths
