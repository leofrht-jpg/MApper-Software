# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""The AESACFG workbook: whole section-(2) configuration round-tripped via xlsx.

The headline guarantee is round-trip identity — export a configuration, import
it back, get the same configuration. That has to hold across all seven groups,
including boundary set, method mapping and carbon budget, which did not
round-trip before.

Also locked here: the object split. ``method_mapping`` and ``carbon_budget``
are config-level; a preset saved from an import must not silently acquire
fields ``SharingPreset`` does not model.
"""
import io

import pytest
from openpyxl import load_workbook

from mapper.api.aesa import (
    ImportRejected,
    _build_sharing_workbook,
    _parse_aesa_config_workbook,
)
from mapper.core.aesa_engine import build_carbon_budget, build_default_sharing_preset
from mapper.models.aesa_schemas import AESAConfigBundle, MethodPBMapping


def _bundle() -> AESAConfigBundle:
    preset = build_default_sharing_preset()
    return AESAConfigBundle(
        boundary_set_id=preset.boundary_set_id,
        sharing_preset_id="preset-abc",
        sharing=preset,
        method_mapping=[
            MethodPBMapping(
                method_tuple=["EF v3.1", "climate change, GWP100", "global warming"],
                pb_id="climate_change",
                conversion_factor=1.0,
            ),
            MethodPBMapping(
                method_tuple=["EF v3.1", "land use"],
                pb_id="land_system_change",
                conversion_factor=2.5,
            ),
        ],
        carbon_budget=build_carbon_budget(),
    )


def _roundtrip(bundle: AESAConfigBundle) -> AESAConfigBundle:
    wb = _build_sharing_workbook(bundle.sharing, include_instructions=True, bundle=bundle)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "roundtrip")


# ── The acceptance criterion ────────────────────────────────────────────────


def test_roundtrip_identity():
    """export -> import -> identical, across all seven groups."""
    original = _bundle()
    back = _roundtrip(original)

    assert back.boundary_set_id == original.boundary_set_id
    assert back.sharing_preset_id == original.sharing_preset_id
    assert back.sharing.name == original.sharing.name
    assert back.sharing.description == original.sharing.description

    assert [p.model_dump() for p in back.sharing.principles] == \
           [p.model_dump() for p in original.sharing.principles]
    assert [a.model_dump() for a in back.sharing.category_assignments] == \
           [a.model_dump() for a in original.sharing.category_assignments]
    assert back.sharing.chain.model_dump() == original.sharing.chain.model_dump()

    assert [m.model_dump() for m in back.method_mapping] == \
           [m.model_dump() for m in original.method_mapping]

    assert (back.carbon_budget is None) == (original.carbon_budget is None)
    if original.carbon_budget is not None:
        o = original.carbon_budget.model_dump()
        n = back.carbon_budget.model_dump()
        # Floats compare with a tolerance, and only floats. xlsx stores numbers
        # as ~15-significant-digit decimal text, so a float64 carrying binary
        # noise (3.5999999999999996) comes back as the value it was always
        # meant to be (3.6). No transport through Excel can preserve full
        # float64; the workbook is canonical, which is what
        # test_roundtrip_survives_a_second_pass pins exactly.
        assert n["projected_emissions"] == pytest.approx(o["projected_emissions"], rel=1e-12)
        if o["co2e_conversion"] is None:
            assert n["co2e_conversion"] is None
        else:
            assert n["co2e_conversion"]["factor"] == pytest.approx(
                o["co2e_conversion"]["factor"], rel=1e-12)
            assert n["co2e_conversion"]["kind"] == o["co2e_conversion"]["kind"]
            assert n["co2e_conversion"]["source"] == o["co2e_conversion"]["source"]
        for k in o:
            if k not in ("projected_emissions", "co2e_conversion"):
                assert n[k] == o[k], k


def test_roundtrip_survives_a_second_pass():
    # A serialiser that is not idempotent will drift on the second export.
    once = _roundtrip(_bundle())
    twice = _roundtrip(once)
    assert twice.model_dump() == once.model_dump()


def test_method_tuple_with_commas_survives():
    # Joined on " | " precisely because method names contain commas.
    b = _bundle()
    back = _roundtrip(b)
    assert back.method_mapping[0].method_tuple == [
        "EF v3.1", "climate change, GWP100", "global warming",
    ]


def test_roundtrip_with_no_carbon_budget():
    b = _bundle().model_copy(update={"carbon_budget": None})
    assert _roundtrip(b).carbon_budget is None


# ── Template + Reference ────────────────────────────────────────────────────


def test_template_has_every_sheet_and_field():
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    for sheet in ("Configuration", "Principles", "Category Assignments",
                  "Downscaling Chain", "Sharing Data", "Method Mapping",
                  "Carbon Budget", "Reference", "Instructions"):
        assert sheet in wb.sheetnames, sheet

    cfg = {r[0] for r in wb["Configuration"].iter_rows(values_only=True) if r[0]}
    assert {"boundary_set_id", "sharing_preset_id", "preset_name"} <= cfg

    budget = {r[0] for r in wb["Carbon Budget"].iter_rows(values_only=True) if r[0]}
    for f in ("initial_budget_gt", "budget_source", "start_year", "end_year",
              "ssp_scenario", "budget_basis", "provisional"):
        assert f in budget, f


def test_reference_sheet_is_populated_from_live_data_and_locked():
    from mapper.core.aesa_engine import load_boundary_sets, load_ssp_trajectories

    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ref = wb["Reference"]
    assert ref.protection.sheet is True, "Reference sheet must be locked"

    rows = [r for r in ref.iter_rows(values_only=True) if r[0]]
    by_field: dict[str, set] = {}
    for field, value, *_ in rows:
        by_field.setdefault(str(field), set()).add(str(value))

    # Live, not hardcoded: every boundary set the engine knows must appear.
    assert set(load_boundary_sets()) <= by_field["boundary_set_id"]
    assert {s["id"] for s in load_ssp_trajectories()} <= by_field["ssp_scenario"]
    assert by_field["budget_basis"] == {"CO2", "CO2e_GHG"}
    assert by_field["principle_mode"] == {"category_specific", "fixed"}


# ── Import rejects, and applies nothing ─────────────────────────────────────


def _wb_with_config_override(**overrides):
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Configuration"]
    for row in ws.iter_rows(min_row=2):
        key = row[0].value
        if key in overrides:
            row[1].value = overrides[key]
    return wb


def test_import_rejects_unknown_boundary_set():
    wb = _wb_with_config_override(boundary_set_id="NotARealSet_2099")
    with pytest.raises(ImportRejected) as ei:
        _parse_aesa_config_workbook(wb, "x")
    errs = ei.value.errors
    assert any(e["field"] == "boundary_set_id" for e in errs)
    assert any("NotARealSet_2099" in e["error"] for e in errs)


def test_import_rejects_unknown_principle_in_assignments():
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    wb["Category Assignments"].cell(row=2, column=2, value="NOPE")
    with pytest.raises(ImportRejected):
        _parse_aesa_config_workbook(wb, "x")


def test_import_rejects_invalid_budget_basis():
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    for row in wb["Carbon Budget"].iter_rows(min_row=2):
        if row[0].value == "budget_basis":
            row[1].value = "CH4_ONLY"
    with pytest.raises(ImportRejected) as ei:
        _parse_aesa_config_workbook(wb, "x")
    assert any("budget_basis" in e["field"] for e in ei.value.errors)


def test_rejection_reports_every_failure_not_just_the_first():
    wb = _wb_with_config_override(boundary_set_id="bogus")
    for row in wb["Carbon Budget"].iter_rows(min_row=2):
        if row[0].value == "budget_basis":
            row[1].value = "bogus_basis"
    with pytest.raises(ImportRejected) as ei:
        _parse_aesa_config_workbook(wb, "x")
    fields = {e["field"] for e in ei.value.errors}
    assert "boundary_set_id" in fields
    assert any("budget_basis" in f for f in fields)


def test_rejected_import_returns_a_bundle_for_nothing():
    # Nothing partially built escapes: the parser raises rather than returning.
    wb = _wb_with_config_override(boundary_set_id="bogus")
    with pytest.raises(ImportRejected):
        _parse_aesa_config_workbook(wb, "x")


# ── Object split: config-level fields must not leak onto a preset ───────────


def test_preset_from_bundle_does_not_carry_method_mapping():
    b = _bundle()
    preset = b.to_preset("Saved from import")
    assert not hasattr(preset, "method_mapping")
    assert "method_mapping" not in preset.model_dump()


def test_preset_from_bundle_keeps_the_patch2a_seeding_defaults():
    # boundary_set_id / carbon_budget DO belong on a preset — as creation-time
    # defaults — so they should be carried over.
    b = _bundle()
    preset = b.to_preset("Saved from import")
    assert preset.boundary_set_id == b.boundary_set_id
    assert preset.carbon_budget == b.carbon_budget
    assert preset.built_in is False
    assert preset.name == "Saved from import"


def test_instructions_document_the_12dp_rounding():
    """A reader must not mistake a cell for the exact in-memory float.

    The workbook is canonical and the engine is left untouched, so the loss of
    precision happens on write — that has to be stated where someone reading
    the file will see it, not only in the commit history.
    """
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    text = "\n".join(
        str(r[0]) for r in wb["Instructions"].iter_rows(values_only=True) if r and r[0]
    )
    assert "12 decimal places" in text
    assert "NOT necessarily the exact in-memory figure" in text
    # and the sheets added by this feature are described
    for sheet in ("Configuration", "Method Mapping", "Carbon Budget", "Reference"):
        assert f"Sheet: {sheet}" in text, sheet


# ── The demo project must not break the round trip ──────────────────────────


def test_demo_stamped_export_still_imports():
    """A config exported from a demo project must be re-importable.

    excel_response() stamps a warning row at the top of every sheet on a demo
    project. That row landed above the header row, so the importer rejected the
    very file MApper had just exported — and the demo project is exactly what a
    reviewer uses. Config exports are now marked round-trippable (prefix only),
    and the parser additionally tolerates a stamped row so files exported
    before that fix still load.
    """
    from mapper.api.cohort_export import stamp_demo_warning

    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    stamp_demo_warning(wb)          # simulate the pre-fix demo export

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    back = _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "demo")

    assert back.boundary_set_id == b.boundary_set_id
    assert [m.model_dump() for m in back.method_mapping] == \
           [m.model_dump() for m in b.method_mapping]
    assert back.carbon_budget is not None


def test_config_export_endpoint_is_marked_round_trippable():
    """Structural: the export must not be stamped, or the file cannot re-import."""
    import inspect

    from mapper.api import aesa

    src = inspect.getsource(aesa.post_config_export)
    assert 'kind="round_trip"' in src, (
        'config export must declare kind="round_trip" so excel_response does '
        "not stamp a warning row above the header row"
    )


def _all_error_text(exc: BaseException) -> str:
    """Every message an import rejection carries.

    ImportRejected deliberately reports EVERY problem at once, so its str() is
    just a count ("1 validation error(s)"); the human-readable reasons live in
    ``.errors``. Assert against those, not the summary.
    """
    parts = [str(exc)]
    for e in getattr(exc, "errors", []) or []:
        parts.extend(str(v) for v in e.values())
    return " | ".join(parts)


# ── Per-principle resolution mode on the Sharing Data sheet ─────────────────


def _series_bundle() -> AESAConfigBundle:
    """The shipped preset with a two-year EpC series added on layer 1, one
    principle set to interpolate and the rest left at the default. The shipped
    template itself is never modified — this is a copy."""
    b = _bundle()
    layers = list(b.sharing.chain.layers)
    first = layers[0]
    principle = next(iter(first.data))
    data = {**first.data, principle: {2025: (100.0, 1000.0), 2050: (200.0, 1000.0)}}
    layers[0] = first.model_copy(update={
        "data": data, "resolution": {principle: "interpolate"},
    })
    sharing = b.sharing.model_copy(update={
        "chain": b.sharing.chain.model_copy(update={"layers": layers}),
    })
    return b.model_copy(update={"sharing": sharing})


def test_resolution_mode_survives_the_round_trip():
    original = _series_bundle()
    back = _roundtrip(original)
    assert back.sharing.chain.model_dump() == original.sharing.chain.model_dump()
    # And specifically the mode, so a passing dump-comparison can't hide it.
    assert back.sharing.chain.layers[0].resolution == \
           original.sharing.chain.layers[0].resolution
    assert "interpolate" in back.sharing.chain.layers[0].resolution.values()


def test_resolution_roundtrip_survives_a_second_pass():
    once = _roundtrip(_series_bundle())
    twice = _roundtrip(once)
    assert twice.model_dump() == once.model_dump()


def test_sharing_data_sheet_carries_a_resolution_column():
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    headers = [str(c.value).strip().lower() if c.value else "" for c in wb["Sharing Data"][1]]
    assert "resolution" in headers
    # Written on every row, not only where it differs from the default: a user
    # editing the sheet should see the mode without reading the Instructions.
    col = headers.index("resolution")
    values = {r[col].value for r in wb["Sharing Data"].iter_rows(min_row=2) if r[0].value}
    assert values and None not in values


def test_a_workbook_with_no_resolution_column_still_imports_as_step():
    """The backward-compat path: every AESACFG workbook exported before the
    column existed has no such column, and must load unchanged."""
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Sharing Data"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ws.delete_cols(headers.index("resolution") + 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    back = _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "legacy")
    for layer in back.sharing.chain.layers:
        assert layer.resolution == {}
        for principle in layer.data:
            assert layer.resolution_for(principle) == "step"


def test_a_blank_resolution_cell_means_step():
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Sharing Data"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = headers.index("resolution") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col).value = None

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    back = _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "blank")
    assert all(ly.resolution == {} for ly in back.sharing.chain.layers)


def test_an_explicit_step_cell_imports_as_the_default_not_as_a_stored_value():
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Sharing Data"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = headers.index("resolution") + 1
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=col).value = "step"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    back = _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "stepped")
    assert all(ly.resolution == {} for ly in back.sharing.chain.layers)


def test_an_unknown_resolution_value_is_rejected_with_a_useful_message():
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Sharing Data"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    ws.cell(row=2, column=headers.index("resolution") + 1).value = "linear"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises((ValueError, ImportRejected)) as exc:
        _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "bad")
    assert "linear" in _all_error_text(exc.value)


def test_contradictory_resolution_rows_are_rejected_not_guessed():
    """Two rows of one principle disagreeing is the user writing two answers to
    one question. Picking one would silently change every ratio it produces."""
    b = _series_bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    ws = wb["Sharing Data"]
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    col = headers.index("resolution") + 1
    principle_col = headers.index("principle") + 1
    first_principle = ws.cell(row=2, column=principle_col).value
    flipped = 0
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=principle_col).value == first_principle:
            ws.cell(row=row, column=col).value = "step" if flipped else "interpolate"
            flipped += 1
    assert flipped >= 2, "fixture needs a principle with at least two year rows"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises((ValueError, ImportRejected)) as exc:
        _parse_aesa_config_workbook(load_workbook(buf, data_only=True), "conflict")
    assert "Resolution" in _all_error_text(exc.value)


def test_reference_sheet_lists_the_resolution_modes_including_blank():
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    rows = [(r[0].value, r[1].value) for r in wb["Reference"].iter_rows(min_row=2)]
    modes = {v for f, v in rows if f == "resolution"}
    assert {"step", "interpolate", "(blank)"} <= modes


def test_instructions_explain_time_varying_shares_with_a_worked_example():
    b = _bundle()
    wb = _build_sharing_workbook(b.sharing, include_instructions=True, bundle=b)
    text = "\n".join(str(r[0].value or "") for r in wb["Instructions"].iter_rows())
    assert "interpolate" in text and "step" in text
    assert "clamp" in text.lower()          # both ends held, no extrapolation
    assert "Worked example" in text
    assert "2037" in text                   # the example resolves a real year
# ── Export must work for an UNSAVED configuration ───────────────────────────


def _draft_body() -> dict:
    """Exactly what the sidebar posts: AESAConfigDraft, with no server fields."""
    import json as _json

    from mapper.core.aesa_engine import build_carbon_budget, build_default_sharing_preset

    return {
        "name": "My AESA config",
        "boundary_set_id": "Sala2020_EF",
        "sharing": _json.loads(build_default_sharing_preset().model_dump_json()),
        "sharing_preset_id": None,          # no preset selected
        "carbon_budget": _json.loads(build_carbon_budget().model_dump_json()),
        "method_mapping": [],
        "impact_mode": "static",
        "dsm_scenario_id": None,
    }


def test_export_accepts_an_unsaved_draft():
    """The Export settings button posts the live draft, which has no id.

    The endpoint took ``AESAConfiguration``, which requires ``id`` and
    ``created_at`` — fields the server only assigns on save — so the button
    returned 422 for every unsaved configuration. A frontend
    ``as unknown as`` cast hid the mismatch from the type checker.
    """
    from fastapi.testclient import TestClient

    from mapper.main import app

    with TestClient(app) as c:
        r = c.post("/api/aesa/config/export", json=_draft_body())
        assert r.status_code == 200, r.text
        assert r.content[:2] == b"PK", "not a valid xlsx"
        assert "My_AESA_config_AESACFG.xlsx" in r.headers["content-disposition"]


def test_export_needs_nothing_but_a_name():
    """Minimal body: no preset, no budget, no mapping, no system."""
    from fastapi.testclient import TestClient

    from mapper.main import app

    with TestClient(app) as c:
        r = c.post("/api/aesa/config/export", json={"name": "bare"})
        assert r.status_code == 200, r.text
        assert r.content[:2] == b"PK"


def test_export_endpoint_does_not_require_server_assigned_fields():
    """Structural: `id` / `created_at` must stay out of the export body model.

    Re-introducing AESAConfiguration here would restore the 422 for every
    unsaved draft, and the frontend cast that hid it is gone, so nothing else
    would catch it.
    """
    import typing

    from mapper.api import aesa
    from mapper.models.aesa_schemas import AESAConfigurationCreate

    # aesa.py uses `from __future__ import annotations`, so the raw annotation
    # is a string — resolve it rather than comparing text.
    hints = typing.get_type_hints(aesa.post_config_export)
    assert hints["body"] is AESAConfigurationCreate
    required = {n for n, f in AESAConfigurationCreate.model_fields.items() if f.is_required()}
    assert required == {"name"}, required


def test_unsaved_draft_export_still_round_trips():
    """The acceptance criterion, driven through the real endpoints."""
    from fastapi.testclient import TestClient

    from mapper.main import app

    body = _draft_body()
    with TestClient(app) as c:
        exported = c.post("/api/aesa/config/export", json=body)
        assert exported.status_code == 200, exported.text

        back = c.post(
            "/api/aesa/config/import",
            files={"file": ("cfg.xlsx", exported.content,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )
        assert back.status_code == 200, back.text
        got = back.json()

    assert got["boundary_set_id"] == body["boundary_set_id"]
    assert got["sharing"]["name"] == body["sharing"]["name"]
    assert len(got["sharing"]["principles"]) == len(body["sharing"]["principles"])
    assert got["carbon_budget"] is not None
    assert got["carbon_budget"]["budget_source"] == body["carbon_budget"]["budget_source"]


# ── Phase B — the template DESCRIBES the conversion correctly ────────────────
#
# The AESACFG workbook is what a user edits and what a reviewer reads. Two of
# its notes were wrong or absent: `initial_budget_gt` was annotated
# "Gt CO2 (or CO2e once basis-applied)", which is not true of the STORED value
# (it is always the pre-basis CO2 figure; the basis is applied at compute), and
# nothing on the sheet or in the Instructions said the factor is derived per
# TEMPERATURE TARGET rather than per SSP — the single most misreadable thing
# about it.


def _cb_notes(wb) -> dict:
    ws = wb["Carbon Budget"]
    return {str(r[0]): str(r[2] or "") for r in ws.iter_rows(values_only=True) if r[0]}


def _instructions(wb) -> str:
    return "\n".join(
        str(r[0] or "") for r in wb["Instructions"].iter_rows(values_only=True))


def test_initial_budget_note_says_the_stored_value_is_pre_basis():
    notes = _cb_notes(_build_sharing_workbook(_bundle().sharing, bundle=_bundle()))
    note = notes["initial_budget_gt"]
    assert "pre-basis" in note or "ALWAYS the pre-basis" in note
    assert "co2e_factor" in note or "applied at compute" in note
    # The old note claimed the cell might already be CO2e. It never is.
    assert "or CO2e once basis-applied" not in note


def test_co2e_factor_note_states_the_derivation_and_its_axis():
    notes = _cb_notes(_build_sharing_workbook(_bundle().sharing, bundle=_bundle()))
    note = notes["co2e_factor"]
    assert "PER TEMPERATURE TARGET" in note, "the per-target axis must be explicit"
    assert "not per SSP" in note, "the misreading must be named"
    assert "(m*x20 + b - C) / x25" in note, "the formula must be stated"
    assert "inert" in note and "reject" in note


def test_instructions_describe_the_conversion_and_the_invariance():
    text = _instructions(_build_sharing_workbook(_bundle().sharing, bundle=_bundle()))
    assert "f = (m*x20 + b - C) / x25" in text
    assert "PER TEMPERATURE TARGET" in text and "not per SSP" in text
    assert "C1+C2" in text and "C3+C4" in text
    # The two consequences a user needs: SR divided by f, depletion year fixed.
    assert "depletion year does not move" in text
    assert "divided by that factor" in text
    # And the pointer to the full derivation.
    assert "co2e_ratio/README.md" in text


def test_reference_sheet_still_locked_and_data_driven():
    """Phase B derived the base year here (B7); the sheet must stay read-only."""
    wb = _build_sharing_workbook(_bundle().sharing, bundle=_bundle())
    assert wb["Reference"].protection.sheet is True
