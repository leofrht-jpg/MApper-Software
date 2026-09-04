# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""`DownscalingLayer.sources` — provenance that survives into the workbook.

The exported AESACFG **Sharing Data** sheet has had a `Source` column since the
sheet existed and wrote `""` into it, so a reader of an exported configuration
could not tell where any number in it came from. The column is now filled from
`layer.sources` and read back on import.

Two properties matter and are asserted here:

* it **round-trips** — export, re-import, same provenance;
* it is **display-only** — nothing in the engine may read it, so a wrong or
  missing source can never move an SR.
"""
from __future__ import annotations

import pytest

from mapper.core.aesa_engine import (
    _DEFAULT_BASE_YEAR,
    build_default_sharing_preset,
    load_sharing_data,
)
from mapper.models.aesa_schemas import DownscalingLayer


def test_the_builtin_preset_carries_a_source_for_every_layer1_principle():
    """Provenance that stops at the JSON file is invisible to a user."""
    sharing = load_sharing_data()
    layer1 = build_default_sharing_preset().chain.layers[0]
    for pid, d in sharing["layer1_defaults"].items():
        assert layer1.sources.get(pid) == d["source"].strip(), (
            f"layer 1 lost or altered the source for {pid}")


def test_the_grandfathering_layers_carry_THEIR_OWN_source():
    """Layers 2 and 3 are the two factors of `layer2.sector_share`
    (0.25 x 0.60 = 0.15), and each carries its own source.

    They used to share one string, taken from `layer2.source` in the data
    file. That string describes the PRODUCT -- "passenger cars ~60% of Danish
    transport GHG x transport ~25% of national total = 0.15" -- so on the
    exported Sharing Data sheet a reader saw a sentence about 0.15 sitting
    next to a factor of 0.25, and the same sentence again next to 0.60. It
    described neither layer.
    """
    layers = build_default_sharing_preset().chain.layers
    l2, l3 = layers[1].sources["AR"], layers[2].sources["AR"]

    assert l2 != l3, "layers 2 and 3 are sharing one source string again"

    # Layer 2 is country -> sector: transport within the national total.
    assert "national" in l2.lower()
    assert "passenger car" not in l2.lower(), (
        "layer 2 is the transport share, not the passenger-car share")

    # Layer 3 is sector -> sub-sector: passenger cars within transport.
    assert "passenger car" in l3.lower()
    assert "transport" in l3.lower()

    # Neither may describe the product. `x`/`*` between two shares is how the
    # combined string read, and 0.15 is the product itself.
    for src in (l2, l3):
        assert "0.15" not in src, f"a layer source states the product: {src}"
        assert " x " not in src.lower(), (
            f"a layer source multiplies two shares together: {src}")


def test_neither_grandfathering_source_implies_a_citation():
    """Both values are round estimates with no dataset and no data year, and
    the strings have to say so.

    Danish EPA is named as where to REFINE them, never as where they came
    from. A source string that merely named a body would read as provenance
    for a number that has none.
    """
    layers = build_default_sharing_preset().chain.layers
    for ly in layers[1:]:
        src = ly.sources["AR"]
        assert "round estimate" in src.lower(), (
            f"layer {ly.layer_number} does not admit it is an estimate: {src}")
        assert "no dataset or data year is recorded" in src.lower(), (
            f"layer {ly.layer_number} does not say its provenance is absent: {src}")
        assert "refine" in src.lower(), (
            f"layer {ly.layer_number} does not say Danish EPA is a refinement "
            f"target rather than the origin: {src}")
        # No numerals: the value is in `data`, and a number repeated in prose
        # goes stale silently.
        assert not any(ch.isdigit() for ch in src), (
            f"layer {ly.layer_number} embeds a number in its source: {src}")


def test_the_combined_string_stays_where_it_is_still_correct():
    """`layer2.source` keeps describing the product, because the legacy
    2-layer MultiDConfig really does apply 0.15 as a single factor.

    The fix was to stop REUSING it on the split chain, not to delete it.
    """
    sharing = load_sharing_data()
    combined = sharing["layer2"]["source"].strip()
    assert combined, "the legacy 2-layer shape lost its source string"
    assert "0.15" in combined, (
        "layer2.source no longer describes the product it is attached to")
    assert sharing["layer2"]["sector_share"] == pytest.approx(0.15)


def test_sources_default_to_empty_so_older_presets_still_load():
    """Back-compat: a layer built without the field is valid and dumps `{}`."""
    ly = DownscalingLayer(layer_number=1, name="L", principle_mode="fixed",
                          fixed_principle="AR", data={"AR": {2025: (1.0, 2.0)}})
    assert ly.sources == {}
    assert ly.model_dump()["sources"] == {}


def test_sources_never_reach_a_chain_factor():
    """Mutate every source to garbage; every factor must be unmoved."""
    clean = build_default_sharing_preset()
    dirty = clean.model_copy(update={"chain": clean.chain.model_copy(update={
        "layers": [ly.model_copy(update={"sources": {k: "GARBAGE" for k in ly.sources}})
                   for ly in clean.chain.layers]})})
    assert any(ly.sources for ly in dirty.chain.layers), "nothing was mutated"

    assignments = {a.pb_id: a.principle_id for a in clean.category_assignments}
    assert assignments, "no assignments to check"
    for pb_id in assignments:
        assert (clean.chain.compute_factor(pb_id, _DEFAULT_BASE_YEAR, assignments)
                == dirty.chain.compute_factor(pb_id, _DEFAULT_BASE_YEAR, assignments)), (
            f"{pb_id}: a source string reached the chain factor")


# ── Workbook round-trip ─────────────────────────────────────────────────────

def _roundtrip(preset):
    """Export the preset to an AESACFG workbook and parse it back."""
    import io

    import openpyxl

    from mapper.api.aesa import _build_sharing_workbook, _parse_aesa_config_workbook

    wb = _build_sharing_workbook(preset, include_instructions=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _parse_aesa_config_workbook(openpyxl.load_workbook(buf), "round-trip").sharing


def test_the_sharing_data_sheet_writes_the_source_instead_of_blank():
    import io

    import openpyxl

    from mapper.api.aesa import _build_sharing_workbook

    preset = build_default_sharing_preset()
    buf = io.BytesIO()
    _build_sharing_workbook(preset, include_instructions=True).save(buf)
    buf.seek(0)
    ws = openpyxl.load_workbook(buf)["Sharing Data"]
    headers = [c.value for c in ws[1]]
    i_src = headers.index("Source")
    i_layer, i_pid = headers.index("Layer"), headers.index("Principle")

    seen = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[i_layer] is None:
            continue
        seen[(int(row[i_layer]), str(row[i_pid]))] = row[i_src]

    assert seen, "no Sharing Data rows were written"
    assert all(v for v in seen.values()), (
        f"a Sharing Data row still carries a blank Source: "
        f"{[k for k, v in seen.items() if not v]}")
    # Compared against the data file rather than a quoted phrase: AR's source
    # text is under methodological review, and a test that pins its wording
    # would fail for a reason that has nothing to do with the export column.
    assert seen[(1, "AR")] == load_sharing_data()["layer1_defaults"]["AR"]["source"].strip()


def test_source_round_trips_through_the_workbook():
    before = build_default_sharing_preset()
    after = _roundtrip(before)
    for a, b in zip(before.chain.layers, after.chain.layers):
        assert a.sources == b.sources, f"layer {a.layer_number} lost its sources"


def test_a_workbook_with_no_source_column_still_imports():
    """Workbooks exported before the column carried anything must keep working;
    the parser treats it as optional and yields empty provenance."""
    import io

    import openpyxl

    from mapper.api.aesa import _build_sharing_workbook, _parse_aesa_config_workbook

    buf = io.BytesIO()
    _build_sharing_workbook(build_default_sharing_preset(), include_instructions=True).save(buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Sharing Data"]
    ws.delete_cols([c.value for c in ws[1]].index("Source") + 1)

    parsed = _parse_aesa_config_workbook(wb, "no-source-column").sharing
    assert all(ly.sources == {} for ly in parsed.chain.layers)
    # and the numbers are untouched
    assert parsed.chain.layers[0].data["AR"] == \
        build_default_sharing_preset().chain.layers[0].data["AR"]


def test_a_disagreeing_source_row_does_not_reject_the_import():
    """Unlike Resolution, provenance is a comment field: the first non-blank
    wins rather than failing an import over a typo in one row."""
    import io

    import openpyxl

    from mapper.api.aesa import _build_sharing_workbook, _parse_aesa_config_workbook

    buf = io.BytesIO()
    _build_sharing_workbook(build_default_sharing_preset(), include_instructions=True).save(buf)
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb["Sharing Data"]
    headers = [c.value for c in ws[1]]
    i_src, i_layer = headers.index("Source") + 1, headers.index("Layer") + 1
    edited = False
    for row in ws.iter_rows(min_row=2):
        if row[i_layer - 1].value == 1:
            row[i_src - 1].value = "A DIFFERENT STORY"
            edited = True
            break
    assert edited, "no layer-1 row to edit"

    parsed = _parse_aesa_config_workbook(wb, "disagreeing").sharing   # must not raise
    assert parsed.chain.layers[0].sources
