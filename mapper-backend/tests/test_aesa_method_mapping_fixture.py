# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""The UI's mapping table and the workbook's Method Mapping sheet agree.

The expanded Method → PB section shows which method characterises which
boundary. The AESACFG workbook round-trips the same `method_mapping` on its
Method Mapping sheet, and a user may edit it there and come back to the UI. If
the two disagree about what is in the mapping, the workbook silently teaches
the user something false about their own configuration.

They can disagree in exactly two ways, both of which come from the UI being
driven by the wrong list:

* a mapping whose ``pb_id`` the active boundary set does not define — an
  orphan. Iterating boundaries would drop it; the workbook writes it.
* two mappings targeting one boundary. Iterating boundaries would show one;
  the workbook writes both, and both reach compute, where they collide on
  (year, pb_id).

So the UI is driven by the MAPPINGS, and this test pins the agreement by
generating the fixture the frontend asserts against — from the real workbook
writer, not a hand-typed copy of what it is believed to emit.

Sibling of test_carbon_budget_series_fixture.py, and the same tripwire logic:
the frontend test is only as good as its fixture, so this fails if the writer's
shape drifts from what the fixture records.
"""

from __future__ import annotations

import io
import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from mapper.api.aesa import _build_sharing_workbook
from mapper.core.aesa_engine import build_default_sharing_preset
from mapper.models.aesa_schemas import AESAConfigBundle, MethodPBMapping

FIXTURE = (
    Path(__file__).resolve().parents[2]
    / "mapper-frontend" / "tests" / "fixtures" / "aesaMethodMappingWorkbook.json"
)

# A realistic EF v3.1 mapping, plus the two divergence cases on purpose. Small
# enough to read; the shape is what matters, not the size.
MAPPINGS = [
    MethodPBMapping(
        method_tuple=["EF v3.1", "climate change", "global warming potential (GWP100)"],
        pb_id="climate_change",
    ),
    MethodPBMapping(
        method_tuple=["EF v3.1", "acidification", "accumulated exceedance (AE)"],
        pb_id="acidification",
        conversion_factor=1.5,          # a non-default factor must survive
    ),
    MethodPBMapping(
        method_tuple=["EF v3.1", "land use", "soil quality index"],
        pb_id="land_use",
    ),
    # Orphan: no such boundary in Sala2020_EF.
    MethodPBMapping(
        method_tuple=["EF v3.1", "some retired category", "x"],
        pb_id="not_a_boundary",
    ),
    # Duplicate: a second method aimed at an already-mapped boundary.
    MethodPBMapping(
        method_tuple=["EF v3.1", "climate change: fossil", "global warming potential (GWP100)"],
        pb_id="climate_change",
    ),
]


def _sheet_rows() -> list[dict]:
    """What the workbook writer actually emits, read back from the file."""
    bundle = AESAConfigBundle(
        boundary_set_id="Sala2020_EF",
        sharing_preset_id=None,
        sharing=build_default_sharing_preset(),
        method_mapping=MAPPINGS,
        carbon_budget=None,
    )
    wb = _build_sharing_workbook(
        build_default_sharing_preset(), include_instructions=False, bundle=bundle,
    )
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ws = load_workbook(buf)["Method Mapping"]

    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h or "").strip() for h in rows[0]]
    assert headers == ["Method Tuple", "PB ID", "Conversion Factor"], headers
    return [
        {
            "method_tuple": [s.strip() for s in str(r[0]).split("|") if s.strip()],
            "pb_id": str(r[1]),
            "conversion_factor": float(r[2]),
        }
        for r in rows[1:]
        if r[0] is not None
    ]


def test_workbook_writes_every_mapping_including_orphan_and_duplicate():
    """The writer's contract: one row per mapping, nothing filtered."""
    rows = _sheet_rows()
    assert len(rows) == len(MAPPINGS)
    assert [r["pb_id"] for r in rows] == [m.pb_id for m in MAPPINGS]
    # The orphan is written despite not being in the boundary set...
    assert any(r["pb_id"] == "not_a_boundary" for r in rows)
    # ...and both climate_change rows survive.
    assert sum(1 for r in rows if r["pb_id"] == "climate_change") == 2


def test_conversion_factor_round_trips():
    rows = _sheet_rows()
    acid = next(r for r in rows if r["pb_id"] == "acidification")
    assert acid["conversion_factor"] == 1.5
    climate = next(r for r in rows if r["pb_id"] == "climate_change")
    assert climate["conversion_factor"] == 1.0        # the default, written explicitly


def test_tuple_is_pipe_joined_so_commas_in_names_survive():
    # "human toxicity: carcinogenic, organics" contains a comma; a comma-joined
    # tuple could not be split back. The reader splits on "|" to match.
    rows = _sheet_rows()
    assert all(len(r["method_tuple"]) == 3 for r in rows)
    assert rows[0]["method_tuple"][0] == "EF v3.1"


def test_frontend_fixture_matches_this_workbook():
    """The frontend asserts its table against this; keep them in step.

    To regenerate after an intentional change to the writer::

        python - <<'PY'
        import json
        from tests.test_aesa_method_mapping_fixture import _sheet_rows, FIXTURE, MAPPINGS
        FIXTURE.write_text(json.dumps({
            "_notice": "Generated by mapper-backend/tests/test_aesa_method_mapping_fixture.py",
            "mappings": [m.model_dump() for m in MAPPINGS],
            "workbook_rows": _sheet_rows(),
        }, indent=2) + "\\n", encoding="utf-8")
        PY
    """
    if not FIXTURE.exists():
        pytest.skip(f"frontend fixture not present at {FIXTURE}")
    fixture = json.loads(FIXTURE.read_text(encoding="utf-8"))
    assert fixture["workbook_rows"] == _sheet_rows(), (
        "the workbook writer no longer emits what the frontend fixture records — "
        "regenerate the fixture (see this test's docstring) and re-check "
        "mapper-frontend/tests/aesaMethodMappingTable.test.tsx"
    )
    # And the mappings the fixture feeds the UI must be the same objects the
    # workbook was built from, or the frontend would be comparing two
    # different configurations and calling them equal.
    assert fixture["mappings"] == [m.model_dump() for m in MAPPINGS]
