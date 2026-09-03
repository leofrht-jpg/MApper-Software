# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""A Basis cell on a non-root row is accepted, applied, and FLAGGED.

Three things disagreed about where Basis belongs:

  * the Instructions sheet says "Set ONLY on the stage root row"
  * the parser takes the first Basis cell found ANYWHERE in the stage
  * a real generator wrote it on child rows, with stage roots blank

The parser stays tolerant -- tightening it would break workbooks already in
the wild, and tolerant reader / strict writer is the right pairing. What was
missing is the signal: the child-row form works only by accident, because the
moment two children of one stage carry DIFFERENT values, first-read silently
wins and the loser is invisible. That is the failure this warning exists to
make audible.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from mapper.api.bom import _parse_bom_workbook


def _wb(rows: list[dict]):
    """Minimal Archetypes + BOM workbook."""
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "Archetypes"
    a.append(["archetype_name", "folder", "description"])
    a.append(["A1", "", ""])
    b = wb.create_sheet("BOM")
    cols = ["archetype_name", "Stage", "Scope", "Basis", "Parent", "Name",
            "Type", "Quantity", "Unit"]
    b.append(cols)
    for r in rows:
        b.append([r.get(c) for c in cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


ROOT = {"archetype_name": "A1", "Stage": "Use Phase", "Parent": None,
        "Name": "Use Phase", "Type": "component", "Quantity": 1, "Unit": "piece"}
CHILD = {"archetype_name": "A1", "Stage": "Use Phase", "Parent": "Use Phase",
         "Name": "Fuel", "Type": "material", "Quantity": 1, "Unit": "kg"}


def _parse(rows):
    """`_parse_bom_workbook` returns (stage-root BOMNodes, warnings, val rows)."""
    roots, warnings, *_ = _parse_bom_workbook(
        openpyxl.load_workbook(_wb(rows), data_only=True)
    )
    return roots, warnings


def _basis_of(roots, stage="Use Phase"):
    for root in roots or []:
        if root.name == stage:
            return root.basis
    return None


def test_basis_on_the_stage_root_is_silent():
    """The documented form must not warn -- a warning that fires on correct
    usage is a warning people learn to ignore."""
    arcs, warnings = _parse([{**ROOT, "Basis": "per year"}, CHILD])
    assert _basis_of(arcs) == "per_year"
    assert not [w for w in warnings if "child row" in w]


def test_basis_on_a_child_row_is_APPLIED():
    """Tolerance is real even when an EXPLICIT stage-root row exists.

    `_ensure_stage_root` documents "first one wins", but it is only reached
    when the root is created IMPLICITLY. With an explicit root row -- which is
    what every real workbook has -- a child's Basis was dropped. The warning
    says "accepted and applied"; this is what makes that true.
    """
    arcs, _ = _parse([ROOT, {**CHILD, "Basis": "per year"}])
    assert _basis_of(arcs) == "per_year"


def test_an_explicit_root_row_beats_a_child_row():
    """First-read-wins, and the root row is read first."""
    arcs, _ = _parse([{**ROOT, "Basis": "per year"},
                      {**CHILD, "Basis": "per unit"}])
    assert _basis_of(arcs) == "per_year"


def test_basis_absent_everywhere_stays_undeclared():
    """Undeclared must remain undeclared -- it computes at x1 and is never
    guessed from Scope."""
    arcs, warnings = _parse([ROOT, CHILD])
    assert _basis_of(arcs) is None
    assert not [w for w in warnings if "child row" in w]


def test_basis_on_a_child_row_WARNS():
    arcs, warnings = _parse([ROOT, {**CHILD, "Basis": "per year"}])
    hits = [w for w in warnings if "child row" in w]
    assert hits, f"expected a non-root Basis warning, got {warnings}"
    assert "Use Phase" in hits[0]
    assert "stage root" in hits[0]


def test_the_silent_loser_is_what_this_warns_about():
    """Two children disagreeing: first wins, the second vanishes.

    This is the concrete failure the warning exists for, so pin the behaviour
    AND the fact that it is reported rather than silent.
    """
    arcs, warnings = _parse([
        ROOT,
        {**CHILD, "Name": "Fuel", "Basis": "per year"},
        {**CHILD, "Name": "Oil", "Basis": "per unit"},
    ])
    assert _basis_of(arcs) == "per_year"          # first one wins
    assert len([w for w in warnings if "child row" in w]) == 2


def test_the_stage_root_branch_assigns_scope_AND_basis():
    """The pair is the defect. Assert it structurally, not by behaviour.

    `scope` and `basis` sit on adjacent lines in the branch that turns an
    explicit stage-root row INTO the stage root. Only `scope` was assigned, so
    the Basis column was silently inert on every real workbook for as long as
    it existed -- a declared `per year` computed at x1, exactly like an
    undeclared stage, with no warning and nothing in the result to show it.

    A behavioural test would catch today's omission. This catches TOMORROW's:
    the next per-stage field added to that branch and forgotten fails here by
    name, rather than being quietly dropped for a month.
    """
    import ast
    import inspect

    from mapper.api import bom as bom_api

    src = inspect.getsource(bom_api)
    tree = ast.parse(src)

    # The branch is `if name == stage and stage not in stages:` -- find it and
    # collect every `node.<attr> = ...` inside.
    assigned: set[str] = set()
    for n in ast.walk(tree):
        if not isinstance(n, ast.If):
            continue
        test = ast.unparse(n.test)
        if "name == stage" not in test or "stages" not in test:
            continue
        for stmt in ast.walk(n):
            if isinstance(stmt, ast.Assign):
                for t in stmt.targets:
                    if (isinstance(t, ast.Attribute)
                            and isinstance(t.value, ast.Name)
                            and t.value.id == "node"):
                        assigned.add(t.attr)

    assert assigned, "could not locate the stage-root branch -- update this test"
    missing = {"scope", "basis"} - assigned
    assert not missing, (
        f"the stage-root branch assigns {sorted(assigned)} but not {sorted(missing)}. "
        "A per-stage field set on the row must be carried onto the stage root "
        "here -- _ensure_stage_root is NOT reached for a row that becomes the "
        "root. Forgetting one makes that column silently inert."
    )


def test_an_invalid_basis_still_warns_separately():
    """The new warning must not swallow the pre-existing validation."""
    _, warnings = _parse([{**ROOT, "Basis": "per fortnight"}])
    assert [w for w in warnings if "invalid Basis" in w]
