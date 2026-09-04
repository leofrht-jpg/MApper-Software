# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Two guarantees a reviewer depends on.

1. CORS accepts any loopback dev-server port. Vite auto-increments off 5173
   when it is busy; before this, the app just rendered empty with a bare
   ERR_FAILED and nothing naming CORS.

2. Every Excel export taken from the demo project carries a fictional-data
   warning INSIDE the workbook. The in-app banner and the "(FICTIONAL DATA)"
   name suffixes do not travel with a downloaded file.
"""
import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from mapper.api.cohort_export import (
    DEMO_EXPORT_WARNING,
    excel_response,
    excel_response_from_bytes,
    stamp_demo_warning,
)
from mapper.main import app

client = TestClient(app)


# ── CORS ────────────────────────────────────────────────────────────────────


@pytest.mark.parametrize("origin", [
    "http://localhost:5173",   # the documented port
    "http://localhost:5174",   # what Vite picks when 5173 is busy
    "http://localhost:3000",
    "http://127.0.0.1:5173",
    "http://127.0.0.1:8080",
    "http://localhost",        # no explicit port
])
def test_loopback_dev_origins_are_allowed(origin):
    res = client.get("/api/health", headers={"Origin": origin})
    assert res.status_code == 200
    assert res.headers.get("access-control-allow-origin") == origin


@pytest.mark.parametrize("origin", [
    "http://evil.example.com",
    "https://localhost:5173",          # https is not a local vite dev server
    "http://localhost.evil.com",       # suffix attack on the host match
    "http://notlocalhost:5173",
])
def test_non_loopback_origins_are_refused(origin):
    # Widening for dev must not have opened the API to arbitrary origins.
    res = client.get("/api/health", headers={"Origin": origin})
    assert res.headers.get("access-control-allow-origin") != origin


def test_tauri_webview_origins_still_allowed():
    # The packaged macOS/Windows webview origins must keep working.
    for origin in ("tauri://localhost", "http://tauri.localhost"):
        res = client.get("/api/health", headers={"Origin": origin})
        assert res.headers.get("access-control-allow-origin") == origin


def test_content_disposition_still_exposed():
    # Export filenames are built server-side; JS cannot read the header unless
    # it is explicitly exposed.
    res = client.get("/api/health", headers={"Origin": "http://localhost:5199"})
    assert "content-disposition" in res.headers.get("access-control-expose-headers", "").lower()


# ── Demo export stamping ────────────────────────────────────────────────────


def _wb_with(sheets: int = 2) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "First"
    ws.append(["Year", "Value"])
    ws.append([2020, 1.5])
    ws.freeze_panes = "A2"
    for i in range(1, sheets):
        extra = wb.create_sheet(f"Sheet{i + 1}")
        extra.append(["Cohort", "Mass"])
        extra.append(["BEV", 42.0])
    return wb


def test_stamp_puts_warning_in_row_one_of_every_sheet():
    wb = _wb_with(3)
    stamp_demo_warning(wb)
    for ws in wb.worksheets:
        assert ws.cell(row=1, column=1).value == DEMO_EXPORT_WARNING, ws.title
        assert ws.cell(row=1, column=1).font.bold


def test_stamp_shifts_data_down_without_losing_it():
    wb = _wb_with(1)
    stamp_demo_warning(wb)
    ws = wb["First"]
    assert [c.value for c in ws[2]] == ["Year", "Value"]
    assert [c.value for c in ws[3]] == [2020, 1.5]


def test_stamp_moves_freeze_pane_down():
    # Otherwise the frozen region would show the warning instead of the header.
    wb = _wb_with(1)
    assert wb["First"].freeze_panes == "A2"
    stamp_demo_warning(wb)
    assert wb["First"].freeze_panes == "A3"


def test_demo_export_is_stamped_and_prefixed():
    res = excel_response(_wb_with(2), "Fleet_DSM.xlsx", kind="data", is_demo=True)
    assert 'filename="DEMO_Fleet_DSM.xlsx"' in res.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(res.body))
    for ws in wb.worksheets:
        assert ws.cell(row=1, column=1).value == DEMO_EXPORT_WARNING


def test_real_export_is_untouched():
    # The real-data path must not gain a warning row or a renamed file.
    res = excel_response(_wb_with(2), "Fleet_DSM.xlsx", kind="data", is_demo=False)
    assert 'filename="Fleet_DSM.xlsx"' in res.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(res.body))
    assert wb["First"].cell(row=1, column=1).value == "Year"
    assert wb["First"].freeze_panes == "A2"


def test_bytes_path_stamps_when_demo_and_passes_through_when_not():
    buf = io.BytesIO()
    _wb_with(2).save(buf)
    raw = buf.getvalue()

    plain = excel_response_from_bytes(raw, "x.xlsx", kind="data", is_demo=False)
    assert plain.body == raw          # byte-identical on the real path
    assert 'filename="x.xlsx"' in plain.headers["content-disposition"]

    demo = excel_response_from_bytes(raw, "x.xlsx", kind="data", is_demo=True)
    assert 'filename="DEMO_x.xlsx"' in demo.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(demo.body))
    for ws in wb.worksheets:
        assert ws.cell(row=1, column=1).value == DEMO_EXPORT_WARNING


def test_prefix_is_not_doubled():
    res = excel_response(_wb_with(1), "DEMO_already.xlsx", kind="data", is_demo=True)
    assert 'filename="DEMO_already.xlsx"' in res.headers["content-disposition"]


def test_every_xlsx_export_routes_through_the_shared_exit():
    """No API module may build an xlsx Response by hand.

    This is the invariant that makes the warning unforgettable: a new export
    surface inherits it by construction. If this fails, someone added a
    hand-rolled Response — route it through excel_response* instead.
    """
    import pathlib

    # encoding="utf-8" is load-bearing: the API sources contain box-drawing
    # and em-dash characters, and Path.read_text() defaults to the locale
    # codec — cp1252 on the Windows runner, which raises UnicodeDecodeError.
    api = pathlib.Path(__file__).resolve().parent.parent / "mapper" / "api"
    offenders = []
    for f in api.glob("*.py"):
        if f.name == "cohort_export.py":
            continue
        if "spreadsheetml.sheet" in f.read_text(encoding="utf-8"):
            offenders.append(f.name)
    assert not offenders, f"hand-rolled xlsx Response in: {offenders}"


# ── Templates must survive the stamping ─────────────────────────────────────


def test_templates_are_prefixed_but_not_stamped():
    """A template is a blank scaffold the user fills in and uploads back.

    Stamping it puts the warning where the header row belongs, and the upload
    parser then rejects the file ("missing required column(s)"). Templates
    therefore get the DEMO_ filename prefix only. This test exists because the
    first implementation did stamp them and broke the documented
    download-fill-upload workflow.
    """
    res = excel_response(_wb_with(1), "inflow_template_x.xlsx",
                         kind="round_trip", is_demo=True)
    assert 'filename="DEMO_inflow_template_x.xlsx"' in res.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(res.body))
    # header still on row 1
    assert [c.value for c in wb["First"][1]] == ["Year", "Value"]
    assert wb["First"].freeze_panes == "A2"


def test_bytes_template_path_is_also_unstamped():
    buf = io.BytesIO()
    _wb_with(1).save(buf)
    raw = buf.getvalue()
    res = excel_response_from_bytes(raw, "t.xlsx", kind="round_trip", is_demo=True)
    assert res.body == raw          # untouched
    assert 'filename="DEMO_t.xlsx"' in res.headers["content-disposition"]


def test_every_export_declares_its_kind():
    """No export may inherit stamping behaviour by default.

    `kind` is a required argument on excel_response/_from_bytes precisely so a
    new endpoint cannot silently get the wrong one. It was previously
    `template: bool = False` — opt-in, defaulting to "stamp" — and two
    round-trippable exports inherited the stamp and broke on re-import (the DSM
    upload templates, then the AESA config export). Omitting `kind` now raises
    a TypeError at call time; this test catches it at review time instead, and
    on paths a test run might not exercise.

    AST, not a line regex. The first version matched `kind=` on the SAME LINE as
    the opening paren, so it reported two false positives the moment a call was
    reflowed across lines — while a genuinely missing `kind` on a one-line call
    several frames away would still have passed. A keyword argument is a
    property of the CALL, not of a line, so the call is what gets inspected.
    """
    import ast
    import pathlib

    api = pathlib.Path(__file__).resolve().parent.parent / "mapper" / "api"
    missing = []
    for f in sorted(api.glob("*.py")):
        if f.name == "cohort_export.py":
            continue
        tree = ast.parse(f.read_text(encoding="utf-8"))
        for node in ast.walk(tree):
            if not isinstance(node, ast.Call):
                continue
            fn = node.func
            name = fn.attr if isinstance(fn, ast.Attribute) else getattr(fn, "id", "")
            if name not in {"excel_response", "excel_response_from_bytes"}:
                continue
            kws = {k.arg for k in node.keywords}
            # `**kwargs` forwarding (arg is None) counts — the caller supplies it.
            if "kind" not in kws and None not in kws:
                missing.append(f"{f.name}:{node.lineno}: {name}(...)")
    assert not missing, (
        "excel_response call without an explicit kind=\n  " + "\n  ".join(missing)
    )


def test_the_kind_guard_is_not_vacuous():
    """The guard must fire on a missing kind, INCLUDING across lines.

    Both shapes, because the line-oriented version passed the multi-line one by
    accident (it never matched it at all) and failed the reflowed *correct*
    call. Asserted against a synthetic corpus so the sweep cannot go green by
    finding nothing.
    """
    import ast

    def offenders(src: str) -> list[str]:
        out = []
        for node in ast.walk(ast.parse(src)):
            if not isinstance(node, ast.Call):
                continue
            fn = node.func
            name = fn.attr if isinstance(fn, ast.Attribute) else getattr(fn, "id", "")
            if name not in {"excel_response", "excel_response_from_bytes"}:
                continue
            kws = {k.arg for k in node.keywords}
            if "kind" not in kws and None not in kws:
                out.append(name)
        return out

    assert offenders("excel_response(buf, 'a.xlsx')") == ["excel_response"]
    assert offenders(
        "excel_response_from_bytes(\n    data,\n    'a.xlsx',\n)"
    ) == ["excel_response_from_bytes"]
    # …and stays quiet on the correct forms, one-line and reflowed.
    assert offenders("excel_response(buf, 'a.xlsx', kind='data')") == []
    assert offenders(
        "excel_response_from_bytes(\n    data,\n    'a.xlsx',\n    kind='round_trip',\n)"
    ) == []
    assert offenders("excel_response(buf, 'a.xlsx', **opts)") == []


def test_export_kind_has_no_default():
    """Guards the inversion itself: giving `kind` a default would restore the
    old failure mode, where forgetting it silently picked a behaviour."""
    import inspect

    from mapper.api.cohort_export import excel_response, excel_response_from_bytes

    for fn in (excel_response, excel_response_from_bytes):
        param = inspect.signature(fn).parameters["kind"]
        assert param.default is inspect.Parameter.empty, (
            f"{fn.__name__}: `kind` must stay required — a default is exactly "
            "the hazard this replaced"
        )
        assert param.kind is inspect.Parameter.KEYWORD_ONLY, fn.__name__
