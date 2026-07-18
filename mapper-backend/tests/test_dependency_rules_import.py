# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Dependency-rules Excel template export + bulk import endpoints.

Run: python -m pytest mapper-backend/tests/test_dependency_rules_import.py -v
"""
from __future__ import annotations

import asyncio
import io
import json

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from starlette.datastructures import Headers, UploadFile

from mapper.api import dsm as dsm_api
from mapper.api import subsystems as sub_api
from mapper.models.dsm_schemas import DimensionDef, SystemDefinition, TimeHorizon
from mapper.models.subsystem_schemas import Subsystem

SYS_ID = "sys-dep-test"
SUB_ID = "sub-dep-test"


@pytest.fixture()
def system_and_subsystem():
    """Register a primary system (dims f, s) + a dependent subsystem (charger)
    directly in the API modules' in-memory state, keyed by the current project."""
    primary = SystemDefinition(
        id=SYS_ID,
        name="Test Fleet",
        time_horizon=TimeHorizon(start_year=2025, end_year=2030),
        dimensions=[
            DimensionDef(name="f", display_name="Fuel", labels=["BEV-LFP", "BEV-NCA", "BEV-NMC532"]),
            DimensionDef(name="s", display_name="Size", labels=["Small", "Sedan", "SUV"]),
            DimensionDef(name="age", display_name="Age", labels=[], is_age=True),
        ],
    )
    sub = Subsystem(
        id=SUB_ID,
        name="Chargers",
        type="dependent",
        dimensions=[DimensionDef(name="charger", display_name="Charger", labels=["home", "public"])],
        depends_on=SYS_ID,
        dependency_rules=[],
    )
    dsm_api._proj_systems()[SYS_ID] = primary
    sub_api._sys_subs(SYS_ID)[SUB_ID] = sub
    yield
    dsm_api._proj_systems().pop(SYS_ID, None)
    sub_api._sys_subs(SYS_ID).pop(SUB_ID, None)


def _rules_xlsx(rows: list[list], headers: list[str] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rules"
    ws.append(headers or ["rule_number", "dependent_archetype", "description", "filter_f", "filter_s", "expression"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(data: bytes, filename: str = "rules.xlsx") -> UploadFile:
    return UploadFile(
        file=io.BytesIO(data),
        filename=filename,
        headers=Headers({"content-type": "application/octet-stream"}),
    )


# ── Template export ─────────────────────────────────────────────────────────


def test_template_columns_examples_and_reference(system_and_subsystem):
    resp = asyncio.run(sub_api.dependency_rules_template(SYS_ID, SUB_ID))
    assert resp.status_code == 200
    assert "dependency_rules_template.xlsx" in resp.headers["content-disposition"]

    wb = load_workbook(io.BytesIO(resp.body))
    assert wb.sheetnames == ["Rules", "Reference"]

    ws = wb["Rules"]
    header = [c.value for c in ws[1]]
    assert header == ["rule_number", "dependent_archetype", "description", "filter_f", "filter_s", "expression"]
    # 2 example rows present.
    assert ws.max_row == 3
    assert ws.cell(row=2, column=2).value == "Residential AC Charger|Default"
    assert ws.cell(row=2, column=6).value == "filtered_stock"

    # Reference sheet auto-populated + locked, lists valid archetypes + dim values.
    ref = wb["Reference"]
    ref_text = "\n".join(str(c.value) for row in ref.iter_rows() for c in row if c.value is not None)
    assert "home" in ref_text and "public" in ref_text  # subsystem cohort keys
    assert "BEV-LFP" in ref_text and "SUV" in ref_text   # primary f/s values
    assert "filtered_stock" in ref_text                  # expression variable
    assert ref.protection.sheet is True                  # read-only


# ── Import: valid ───────────────────────────────────────────────────────────


def test_import_valid_returns_parsed_rules(system_and_subsystem):
    data = _rules_xlsx([
        [1, "home", "one per BEV", "BEV-LFP,BEV-NCA", "Small,Sedan", "filtered_stock"],
        [2, "public", "fast", "all", "SUV", "filtered_stock * 0.1"],
    ])
    resp = asyncio.run(sub_api.import_dependency_rules(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 200
    body = json.loads(resp.body)
    assert body["ok"] is True
    assert len(body["rules"]) == 2
    assert body["rules"][0]["dependent_archetype_id"] == "home"
    assert body["rules"][0]["driver_filter"] == {"f": ["BEV-LFP", "BEV-NCA"], "s": ["Small", "Sedan"]}
    # "all" filter → no filter on that dim.
    assert body["rules"][1]["driver_filter"] == {"s": ["SUV"]}


# ── Import: invalid archetype (full rejection) ──────────────────────────────


def test_import_invalid_archetype_rejected(system_and_subsystem):
    data = _rules_xlsx([
        [1, "home", "", "all", "all", "filtered_stock"],
        [2, "does_not_exist", "", "all", "all", "filtered_stock"],
    ])
    resp = asyncio.run(sub_api.import_dependency_rules(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert body["ok"] is False
    errs = body["errors"]
    assert any(e["row"] == 3 and e["field"] == "dependent_archetype" for e in errs)


# ── Import: invalid expression (full rejection) ─────────────────────────────


def test_import_invalid_expression_rejected(system_and_subsystem):
    data = _rules_xlsx([
        [1, "home", "", "all", "all", "filtered_stock +"],  # syntax error
    ])
    resp = asyncio.run(sub_api.import_dependency_rules(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert body["ok"] is False
    assert any(e["field"] == "expression" for e in body["errors"])


# ── Import: invalid filter value (full rejection) ───────────────────────────


def test_import_invalid_filter_value_rejected(system_and_subsystem):
    data = _rules_xlsx([
        [1, "home", "", "NOT-A-FUEL", "all", "filtered_stock"],
    ])
    resp = asyncio.run(sub_api.import_dependency_rules(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert any(e["field"] == "filter" for e in body["errors"])


# ── Import: wrong file type ─────────────────────────────────────────────────


def test_import_rejects_non_xlsx(system_and_subsystem):
    with pytest.raises(HTTPException) as ei:
        asyncio.run(sub_api.import_dependency_rules(SYS_ID, SUB_ID, _upload(b"a,b\n1,2", "rules.csv")))
    assert ei.value.status_code == 400
    assert ".csv" in ei.value.detail or "xlsx" in ei.value.detail.lower()


def test_import_unknown_subsystem_404(system_and_subsystem):
    with pytest.raises(HTTPException) as ei:
        asyncio.run(sub_api.import_dependency_rules(SYS_ID, "nope", _upload(_rules_xlsx([]))))
    assert ei.value.status_code == 404
