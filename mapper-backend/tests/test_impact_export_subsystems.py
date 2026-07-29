# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Impact Assessment fleet Excel export — subsystem integration + filename.

Covers:
- `_impact_export_filename`: zero-subsystem (unchanged), one/two subsystems,
  sanitisation, >100-char truncation.
- `_build_mfa_lca_workbook` with subsystem-prefixed cohorts: By cohort has a
  System column + no UUID + populated Archetype; Summary has "Subsystems
  included" + Archetypes>0; the "By subsystem" sheet appears (and is absent
  without subsystems); Cohort mappings gains a System column + subsystem rows.
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api.bom import _build_mfa_lca_workbook, _impact_export_filename
from mapper.models.bom_schemas import (
    Archetype,
    CohortMapping,
    CohortMappingEntry,
    DSMLCAResult,
    DSMLCASummary,
    DSMLCAYearResult,
)
from mapper.models.dsm_schemas import DimensionDef

SYS_ID = "e5442abf-fa89-4804-b192-667f6ecd08bf"
SUB_ID = "670be0bf-eb95-4479-b5f1-dea938d0e46f"
METHOD = ["EF v3.1", "climate change", "GWP100"]


# ── Filename ────────────────────────────────────────────────────────────────


def test_filename_zero_subsystems_unchanged():
    assert _impact_export_filename("Car Fleet", [], "all") == "Car_Fleet_impact_all.xlsx"


def test_filename_one_subsystem():
    assert _impact_export_filename("Car Fleet", ["Fueling Infrastructure"], "all") == \
        "Car_Fleet+Fueling_Infrastructure_impact_all.xlsx"


def test_filename_two_subsystems():
    assert _impact_export_filename("Car Fleet", ["A", "B"], "all") == "Car_Fleet+A+B_impact_all.xlsx"


def test_filename_strips_invalid_chars():
    fn = _impact_export_filename('Car/Fleet:*?', ['Sub"<>|'], "all")
    for bad in '/\\:*?"<>|':
        assert bad not in fn
    assert fn == "CarFleet+Sub_impact_all.xlsx"


def test_filename_truncates_over_100_chars():
    subs = [f"Very_Long_Subsystem_Name_Number_{i}" for i in range(6)]
    fn = _impact_export_filename("Primary", subs, "all")
    assert fn == "Primary+6_subsystems_impact_all.xlsx"
    assert len(fn) - len(".xlsx") <= 100


# ── Workbook with subsystems ─────────────────────────────────────────────────


def _result(impact_by_cohort: dict[str, float]) -> DSMLCAResult:
    yr = DSMLCAYearResult(
        year=2030, total_impact=sum(impact_by_cohort.values()),
        impact_by_cohort=impact_by_cohort, impact_by_material={}, unit="kg CO2eq",
    )
    return DSMLCAResult(
        mfa_system_id=SYS_ID, method=METHOD, scope="all", unit="kg CO2eq",
        years=[yr], summary=DSMLCASummary(total_impact=yr.total_impact, peak_year=2030, peak_impact=yr.total_impact),
        stages_included=["Manufacturing"],
    )


def _build_with_subsystem():
    # Aggregated (subsystem-present) results → ALL keys are prefixed.
    results = [_result({
        f"{SYS_ID}::BEV-LFP|SUV": 100.0,
        f"{SUB_ID}::CNG Station|Default": 20.0,
    })]
    primary_mapping = CohortMapping(
        mfa_system_id=SYS_ID,
        mappings=[CohortMappingEntry(cohort_key="BEV-LFP|SUV", archetype_id="arc-bev", scaling_factor=1.0)],
    )
    archetypes = {
        "arc-bev": Archetype(id="arc-bev", name="BEV-LFP SUV"),
        "arc-charge": Archetype(id="arc-charge", name="Charging Infrastructure"),
    }
    subsystems = [{
        "id": SUB_ID, "name": "Fueling Infrastructure",
        "mappings": {"CNG Station|Default": ("arc-charge", 2.0)},
    }]
    sim_counts = {2030: {f"{SYS_ID}::BEV-LFP|SUV": 10.0, f"{SUB_ID}::CNG Station|Default": 3.0}}
    wb = _build_mfa_lca_workbook(
        system_name="Car Fleet", results=results, scope="all", selected_year=None,
        cohort_mapping=primary_mapping, archetypes=archetypes, sim_counts=sim_counts,
        dims=[DimensionDef(name="fuel", display_name="Fuel", labels=["BEV-LFP"]),
              DimensionDef(name="size", display_name="Size", labels=["SUV"])],
        sim_result=None, system_id=SYS_ID, subsystems=subsystems,
    )
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return load_workbook(buf)


def _sheet_text(ws) -> str:
    return "\n".join(str(c.value) for row in ws.iter_rows() for c in row if c.value is not None)


def test_by_cohort_has_system_column_no_uuid_and_archetype():
    wb = _build_with_subsystem()
    ws = wb["By cohort"]
    header = [c.value for c in ws[1]]
    assert header[0] == "Year" and header[1] == "System"  # System first after Year
    text = _sheet_text(ws)
    assert SUB_ID not in text and SYS_ID not in text          # no UUID anywhere
    assert "Fueling Infrastructure" in text                   # subsystem name
    assert "Charging Infrastructure" in text                  # subsystem archetype resolved
    assert "BEV-LFP SUV" in text                              # primary archetype resolved (prefixed)


def test_summary_subsystems_and_archetype_count():
    wb = _build_with_subsystem()
    text = _sheet_text(wb["Summary"])
    assert "Subsystems included" in text
    assert "Fueling Infrastructure" in text
    # Archetypes count > 0 (both primary + subsystem archetypes resolved).
    ws = wb["Summary"]
    arch_row = next(r for r in ws.iter_rows(values_only=True) if r and r[0] == "Archetypes")
    assert arch_row[1] == 2


def test_by_subsystem_sheet_present_with_clean_names():
    wb = _build_with_subsystem()
    assert "By subsystem" in wb.sheetnames
    ws = wb["By subsystem"]
    header = [c.value for c in ws[1]]
    assert header[:6] == ["Year", "Subsystem", "Dependent archetype", "BOM archetype", "Scale", "Unit count"]
    text = _sheet_text(ws)
    assert SUB_ID not in text
    assert "CNG Station Default" in text or "CNG Station|Default" in text
    assert "Charging Infrastructure" in text


def test_cohort_mappings_has_system_column_and_subsystem_rows():
    wb = _build_with_subsystem()
    ws = wb["Cohort mappings"]
    assert [c.value for c in ws[1]][0] == "System"
    text = _sheet_text(ws)
    assert "Car Fleet" in text and "Fueling Infrastructure" in text


def test_by_subsystem_absent_without_subsystems():
    results = [_result({"BEV-LFP|SUV": 100.0})]
    primary_mapping = CohortMapping(
        mfa_system_id=SYS_ID,
        mappings=[CohortMappingEntry(cohort_key="BEV-LFP|SUV", archetype_id="arc-bev", scaling_factor=1.0)],
    )
    wb = _build_mfa_lca_workbook(
        system_name="Car Fleet", results=results, scope="all", selected_year=None,
        cohort_mapping=primary_mapping, archetypes={"arc-bev": Archetype(id="arc-bev", name="BEV-LFP SUV")},
        sim_counts={2030: {"BEV-LFP|SUV": 10.0}},
        dims=[DimensionDef(name="fuel", display_name="Fuel", labels=["BEV-LFP"]),
              DimensionDef(name="size", display_name="Size", labels=["SUV"])],
        sim_result=None, system_id=SYS_ID, subsystems=[],
    )
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    wb2 = load_workbook(buf)
    assert "By subsystem" not in wb2.sheetnames
    # Summary still reports "None" for subsystems.
    assert "None" in _sheet_text(wb2["Summary"])
