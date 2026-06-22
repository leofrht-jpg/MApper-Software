# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Patch 4G — Single-product Impact Assessment Excel exports.

Round-trip tests for the three new builders:
    _build_single_product_static_workbook
    _build_single_product_prospective_workbook
    _build_single_product_comparison_workbook

Drive the builders directly with synthesised ``ArchetypeLCACalculateResult``
objects; the wire-level routes get a smoke-level acceptance test that
proves the FastAPI plumbing pipes the bytes through correctly.

What we assert:

* Sheet inventory per workbook matches the Patch-4G spec.
* The Configuration sheet carries the load-bearing reproducibility info
  (archetype name, scope, stage amounts). Stage amounts is a non-trivial
  reproducibility input — analysts re-running an export six months later
  must be able to recover the per-stage weighting from the workbook
  alone.
* Stage breakdown sheet appears IFF ``scope == "all"`` AND at least one
  result carries a non-empty ``stage_breakdown``. Specific-stage scopes
  must not emit an empty/redundant Stage breakdown sheet.
* Comparison Δ math is ``P − S`` and Δ% normalisation handles the
  ``S == 0`` edge by writing an empty cell (never ``inf`` / ``NaN``).
* The route returns 400 on empty inputs (defence-in-depth — the frontend
  Disable button covers the happy path).
"""
from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from mapper.api import impact as impact_api
from mapper.main import app
from mapper.models.schemas import (
    ArchetypeLCACalculateResult,
    ArchetypeLCAMethodResult,
    SingleProductProspectiveRunPayload,
)


# ── Fixtures ────────────────────────────────────────────────────────────────


GWP_TUPLE = ["EF v3.1 (E,T)", "climate change", "global warming potential (GWP100)"]
ACID_TUPLE = ["EF v3.1 (E,T)", "acidification", "accumulated exceedance"]


def _method_result(tuple_: list[str], score: float, unit: str = "kg CO2-eq") -> ArchetypeLCAMethodResult:
    return ArchetypeLCAMethodResult(
        method=tuple_,
        method_label=" › ".join(tuple_),
        score=score,
        unit=unit,
        contributions=[],
    )


def _archetype_result(
    name: str = "BEV-LFP|Small",
    scope: str = "all",
    methods: list[tuple[list[str], float]] | None = None,
    stage_breakdown: dict[str, dict[str, float]] | None = None,
    stage_amounts: dict[str, float] | None = None,
    compute_database: str | None = None,
    parameter_scenario: str | None = None,
) -> ArchetypeLCACalculateResult:
    if methods is None:
        methods = [(GWP_TUPLE, 1234.5), (ACID_TUPLE, 6.78)]
    return ArchetypeLCACalculateResult(
        archetype_id="arc-1",
        archetype_name=name,
        scope=scope,
        amount=1.0,
        stage_amounts=stage_amounts or {
            "Manufacturing": 1.0,
            "Use Phase": 15.0,
            "Maintenance": 15.0,
            "End of Life": 1.0,
        },
        stages_included=["Manufacturing", "Use Phase", "Maintenance", "End of Life"],
        results=[_method_result(t, s) for t, s in methods],
        elapsed_seconds=2.34,
        compute_database=compute_database,
        parameter_scenario=parameter_scenario,
        warnings=[],
        stage_breakdown=stage_breakdown,
    )


def _load(wb) -> object:
    """Save → reload through openpyxl. Verifies the bytes are a valid
    .xlsx that survives a round-trip."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf)


# ── Static builder ──────────────────────────────────────────────────────────


class TestStaticBuilder:
    def test_single_scenario_basic_layout(self):
        result = _archetype_result()
        wb = impact_api._build_single_product_static_workbook(
            archetype_name=result.archetype_name,
            scope=result.scope,
            scenarios=[("Base", result)],
        )
        wb = _load(wb)

        # No Stage breakdown — the result didn't include one.
        assert "Configuration" in wb.sheetnames
        assert "Total impacts" in wb.sheetnames
        assert "Stage breakdown" not in wb.sheetnames

        # Configuration carries archetype + scope + stage amounts.
        cfg_text = "\n".join(
            f"{row[0].value}={row[1].value}" for row in wb["Configuration"].iter_rows(max_col=2)
            if row[0].value is not None
        )
        assert "Archetype=BEV-LFP|Small" in cfg_text
        assert "Full lifecycle" in cfg_text
        # Stage amounts line includes every stage with its weight.
        assert "Manufacturing 1" in cfg_text
        assert "Use Phase 15" in cfg_text
        assert "Maintenance 15" in cfg_text
        assert "End of Life 1" in cfg_text

        # Total impacts: single-scenario shape is (Indicator, Method path,
        # Score, Unit).
        ws_tot = wb["Total impacts"]
        header = [c.value for c in ws_tot[1]]
        assert header == ["Indicator", "Method path", "Score", "Unit"]
        # Two methods → two rows.
        assert ws_tot.max_row == 3
        # GWP score in column 3.
        assert ws_tot.cell(row=2, column=3).value == pytest.approx(1234.5)
        assert ws_tot.cell(row=3, column=3).value == pytest.approx(6.78)

    def test_multi_scenario_adds_columns_per_case(self):
        base = _archetype_result(parameter_scenario=None)
        opt = _archetype_result(
            parameter_scenario="Optimistic",
            methods=[(GWP_TUPLE, 800.0), (ACID_TUPLE, 5.0)],
        )
        pess = _archetype_result(
            parameter_scenario="Pessimistic",
            methods=[(GWP_TUPLE, 1500.0), (ACID_TUPLE, 9.0)],
        )
        wb = impact_api._build_single_product_static_workbook(
            archetype_name="BEV-LFP|Small",
            scope="all",
            scenarios=[
                ("Base", base),
                ("Optimistic", opt),
                ("Pessimistic", pess),
            ],
        )
        wb = _load(wb)

        ws_tot = wb["Total impacts"]
        header = [c.value for c in ws_tot[1]]
        # Multi-case header: Indicator, Method path, Unit, then one col per scenario.
        assert header == ["Indicator", "Method path", "Unit", "Base", "Optimistic", "Pessimistic"]
        # Two methods → two rows.
        assert ws_tot.max_row == 3
        gwp_row = next(
            row for row in ws_tot.iter_rows(min_row=2, values_only=True)
            if "GWP" in str(row[0])
        )
        assert gwp_row[3] == pytest.approx(1234.5)
        assert gwp_row[4] == pytest.approx(800.0)
        assert gwp_row[5] == pytest.approx(1500.0)

    def test_stage_breakdown_sheet_only_when_scope_all_and_present(self):
        # scope="all" and breakdown set → sheet is created.
        breakdown = {
            " › ".join(GWP_TUPLE): {
                "Manufacturing": 400.0,
                "Use Phase": 800.0,
                "Maintenance": 30.0,
                "End of Life": 4.5,
            },
            " › ".join(ACID_TUPLE): {
                "Manufacturing": 1.0, "Use Phase": 4.0,
                "Maintenance": 1.5, "End of Life": 0.28,
            },
        }
        wb = impact_api._build_single_product_static_workbook(
            archetype_name="BEV",
            scope="all",
            scenarios=[("Base", _archetype_result(stage_breakdown=breakdown))],
        )
        wb = _load(wb)
        assert "Stage breakdown" in wb.sheetnames

        ws_sb = wb["Stage breakdown"]
        header = [c.value for c in ws_sb[1]]
        # Single-scenario header: Indicator, Method path, Unit, <stages…>, Total
        assert header[:3] == ["Indicator", "Method path", "Unit"]
        assert "Manufacturing" in header
        assert "End of Life" in header
        assert header[-1] == "Total"

    def test_stage_breakdown_omitted_for_specific_scope(self):
        # scope="stock" → even with a breakdown payload, no sheet (would
        # be redundant — single stage = single column).
        breakdown = {
            " › ".join(GWP_TUPLE): {"Use Phase": 800.0},
        }
        wb = impact_api._build_single_product_static_workbook(
            archetype_name="BEV",
            scope="stock",
            scenarios=[("Base", _archetype_result(scope="stock", stage_breakdown=breakdown))],
        )
        wb = _load(wb)
        assert "Stage breakdown" not in wb.sheetnames


# ── Prospective builder ─────────────────────────────────────────────────────


class TestProspectiveBuilder:
    def test_wide_and_long_sheets_present(self):
        runs = [
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp1_2030",
                year=2030, iam="remind", ssp="SSP1-PkBudg500",
                result=_archetype_result(
                    methods=[(GWP_TUPLE, 1100.0), (ACID_TUPLE, 6.0)],
                    compute_database="ei310_remind_ssp1_2030",
                ),
            ),
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp1_2050",
                year=2050, iam="remind", ssp="SSP1-PkBudg500",
                result=_archetype_result(
                    methods=[(GWP_TUPLE, 600.0), (ACID_TUPLE, 4.0)],
                    compute_database="ei310_remind_ssp1_2050",
                ),
            ),
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp2_2030",
                year=2030, iam="remind", ssp="SSP2-PkBudg1150",
                result=_archetype_result(
                    methods=[(GWP_TUPLE, 1300.0), (ACID_TUPLE, 7.0)],
                    compute_database="ei310_remind_ssp2_2030",
                ),
            ),
        ]
        wb = impact_api._build_single_product_prospective_workbook(
            archetype_name="BEV-LFP|Small",
            scope="all",
            runs=runs,
        )
        wb = _load(wb)

        assert "Configuration" in wb.sheetnames
        assert "Time series (wide)" in wb.sheetnames
        assert "Time series (long)" in wb.sheetnames

        # Wide layout: years are columns. Two trajectories × two methods
        # = 4 data rows. Years header should include both 2030 and 2050.
        ws_w = wb["Time series (wide)"]
        wide_header = [c.value for c in ws_w[1]]
        assert "2030" in wide_header
        assert "2050" in wide_header

        # Long layout: one row per (db × year × method). 3 runs × 2 methods = 6.
        ws_l = wb["Time series (long)"]
        assert ws_l.max_row == 1 + 3 * 2  # header + data

    def test_stage_breakdown_by_year_sheet_only_when_scope_all(self):
        breakdown = {
            " › ".join(GWP_TUPLE): {
                "Manufacturing": 400.0, "Use Phase": 600.0,
                "Maintenance": 28.0, "End of Life": 4.0,
            },
        }
        runs = [
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp1_2030",
                year=2030, iam="remind", ssp="SSP1",
                result=_archetype_result(
                    methods=[(GWP_TUPLE, 1032.0)],
                    stage_breakdown=breakdown,
                ),
            ),
        ]
        wb = impact_api._build_single_product_prospective_workbook(
            archetype_name="BEV", scope="all", runs=runs,
        )
        wb = _load(wb)
        assert "Stage breakdown by year" in wb.sheetnames

        # Same payload but scope="inflows" — no breakdown sheet.
        wb2 = impact_api._build_single_product_prospective_workbook(
            archetype_name="BEV", scope="inflows", runs=runs,
        )
        wb2 = _load(wb2)
        assert "Stage breakdown by year" not in wb2.sheetnames


# ── Comparison builder ──────────────────────────────────────────────────────


class TestComparisonBuilder:
    def test_delta_math_is_p_minus_s(self):
        static = _archetype_result(
            methods=[(GWP_TUPLE, 1000.0), (ACID_TUPLE, 10.0)],
        )
        runs = [
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp1_2030",
                year=2030, iam="remind", ssp="SSP1",
                result=_archetype_result(
                    methods=[(GWP_TUPLE, 800.0), (ACID_TUPLE, 12.0)],
                ),
            ),
        ]
        wb = impact_api._build_single_product_comparison_workbook(
            archetype_name="BEV", scope="all",
            static_result=static, projected_runs=runs,
        )
        wb = _load(wb)
        ws = wb["Comparison data"]

        # Header has S, P, Δ, Δ% in that order.
        header = [c.value for c in ws[1]]
        assert "Static (S)" in header
        assert "Projected (P)" in header
        assert "Δ (P − S)" in header
        assert "Δ %" in header

        # Locate the GWP row and confirm Δ = -200 (improvement) and
        # Δ% = -20.
        for row in ws.iter_rows(min_row=2, values_only=True):
            method_label = row[4]
            if method_label and "GWP" in method_label:
                s_idx = header.index("Static (S)")
                p_idx = header.index("Projected (P)")
                d_idx = header.index("Δ (P − S)")
                pct_idx = header.index("Δ %")
                assert row[s_idx] == pytest.approx(1000.0)
                assert row[p_idx] == pytest.approx(800.0)
                assert row[d_idx] == pytest.approx(-200.0)
                assert row[pct_idx] == pytest.approx(-20.0)
                break
        else:
            raise AssertionError("GWP row not found in Comparison data")

    def test_pct_handles_zero_static(self):
        # S = 0 → Δ% is empty (not inf, not NaN).
        static = _archetype_result(
            methods=[(GWP_TUPLE, 0.0)],
        )
        runs = [
            SingleProductProspectiveRunPayload(
                db_name="ei310_remind_ssp1_2030",
                year=2030, iam="remind", ssp="SSP1",
                result=_archetype_result(methods=[(GWP_TUPLE, 5.0)]),
            ),
        ]
        wb = impact_api._build_single_product_comparison_workbook(
            archetype_name="BEV", scope="all",
            static_result=static, projected_runs=runs,
        )
        wb = _load(wb)
        ws = wb["Comparison data"]
        header = [c.value for c in ws[1]]
        pct_idx = header.index("Δ %")
        for row in ws.iter_rows(min_row=2, values_only=True):
            assert row[pct_idx] in ("", None)

    def test_cumulative_summary_per_trajectory(self):
        static = _archetype_result(methods=[(GWP_TUPLE, 1000.0)])
        runs = [
            SingleProductProspectiveRunPayload(
                db_name="db1", year=2030, iam="remind", ssp="SSP1",
                result=_archetype_result(methods=[(GWP_TUPLE, 900.0)]),
            ),
            SingleProductProspectiveRunPayload(
                db_name="db2", year=2050, iam="remind", ssp="SSP1",
                result=_archetype_result(methods=[(GWP_TUPLE, 600.0)]),
            ),
        ]
        wb = impact_api._build_single_product_comparison_workbook(
            archetype_name="BEV", scope="all",
            static_result=static, projected_runs=runs,
        )
        wb = _load(wb)
        ws = wb["Cumulative summary per traj."]
        # One trajectory × one method = one data row.
        assert ws.max_row == 2
        header = [c.value for c in ws[1]]
        avg_idx = header.index("Average Δ")
        cum_idx = header.index("Cumulative Δ")
        peak_y_idx = header.index("Peak Δ year")
        # Δ values: 900-1000 = -100, 600-1000 = -400. Avg = -250, cum = -500.
        # Peak Δ (most positive) is the less-negative one → -100 at year 2030.
        row = next(ws.iter_rows(min_row=2, values_only=True))
        assert row[avg_idx] == pytest.approx(-250.0)
        assert row[cum_idx] == pytest.approx(-500.0)
        assert row[peak_y_idx] == 2030


# ── Route smoke tests ───────────────────────────────────────────────────────


class TestExportRoutes:
    def test_static_route_returns_xlsx(self):
        client = TestClient(app)
        body = {
            "archetype_name": "BEV-LFP|Small",
            "scope": "all",
            "scenarios": [
                {
                    "label": "Base",
                    "result": _archetype_result().model_dump(),
                },
            ],
        }
        resp = client.post("/api/impact/export-single-product-static", json=body)
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "MApper_Impact_SingleProduct_Static_" in resp.headers["content-disposition"]
        # Round-trip the body bytes through openpyxl.
        wb = load_workbook(io.BytesIO(resp.content))
        assert "Configuration" in wb.sheetnames
        assert "Total impacts" in wb.sheetnames

    def test_static_route_400s_on_empty(self):
        client = TestClient(app)
        resp = client.post(
            "/api/impact/export-single-product-static",
            json={"archetype_name": "BEV", "scope": "all", "scenarios": []},
        )
        assert resp.status_code == 400

    def test_prospective_route_returns_xlsx(self):
        client = TestClient(app)
        run = SingleProductProspectiveRunPayload(
            db_name="ei310_remind_ssp1_2030",
            year=2030, iam="remind", ssp="SSP1",
            result=_archetype_result(),
        ).model_dump()
        body = {
            "archetype_name": "BEV-LFP|Small",
            "scope": "all",
            "runs": [run],
        }
        resp = client.post("/api/impact/export-single-product-prospective", json=body)
        assert resp.status_code == 200
        assert "MApper_Impact_SingleProduct_Prospective_" in resp.headers["content-disposition"]

    def test_comparison_route_400s_on_empty_runs(self):
        client = TestClient(app)
        body = {
            "archetype_name": "BEV",
            "scope": "all",
            "static_result": _archetype_result().model_dump(),
            "projected_runs": [],
        }
        resp = client.post("/api/impact/export-single-product-comparison", json=body)
        assert resp.status_code == 400

    def test_comparison_route_returns_xlsx(self):
        client = TestClient(app)
        run = SingleProductProspectiveRunPayload(
            db_name="ei310_remind_ssp1_2030",
            year=2030, iam="remind", ssp="SSP1",
            result=_archetype_result(),
        ).model_dump()
        body = {
            "archetype_name": "BEV-LFP|Small",
            "scope": "all",
            "static_result": _archetype_result().model_dump(),
            "projected_runs": [run],
        }
        resp = client.post("/api/impact/export-single-product-comparison", json=body)
        assert resp.status_code == 200
        assert "MApper_Impact_SingleProduct_Comparison_" in resp.headers["content-disposition"]
