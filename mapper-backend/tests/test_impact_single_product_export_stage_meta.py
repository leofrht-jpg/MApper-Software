"""Patch 5K+ — single-product exports record stage-amount provenance
(preset + lifetime) in the Configuration block, closing the 5J divergence
(multi-item already records preset/lifetime/amounts).

Per-stage amounts continue to come from the result echo
(``ArchetypeLCACalculateResult.stage_amounts``); preset + lifetime are
threaded via ``stage_amounts_meta`` (reusing 5J's ``StageAmountsMeta``).
Optional: when absent (old client) preset/lifetime render as "—" and the
amounts still appear. Labels match the multi-item sheet (Preset / Lifetime (yr)).
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api import impact as impact_api
from mapper.models.schemas import (
    ArchetypeLCACalculateResult,
    ArchetypeLCAMethodResult,
    SingleProductProspectiveRunPayload,
    StageAmountsMeta,
)

STAGE_AMOUNTS = {"Manufacturing": 1.0, "Use Phase": 15.0, "Maintenance": 15.0, "End of Life": 1.0}
META = StageAmountsMeta(preset="lifetime", lifetime=15, amounts=STAGE_AMOUNTS)


def _result(scope: str = "all") -> ArchetypeLCACalculateResult:
    return ArchetypeLCACalculateResult(
        archetype_id="arc-1", archetype_name="BEV-LFP|Small", scope=scope,
        amount=1.0, stage_amounts=STAGE_AMOUNTS,
        stages_included=list(STAGE_AMOUNTS.keys()),
        results=[ArchetypeLCAMethodResult(
            method=["EF v3.1", "climate change", "GWP100"],
            method_label="EF v3.1 › climate change › GWP100",
            score=1234.5, unit="kg CO2-eq", contributions=[],
        )],
        elapsed_seconds=1.0, warnings=[], stage_breakdown=None,
    )


def _run(db: str = "ei310-remind-ssp2-2030") -> SingleProductProspectiveRunPayload:
    return SingleProductProspectiveRunPayload(
        db_name=db, year=2030, iam="remind", ssp="SSP2-PkBudg1150", result=_result(),
    )


def _cfg(wb) -> dict:
    """{label: value} from the Configuration sheet's two-column rows."""
    buf = io.BytesIO()
    wb.save(buf)
    loaded = load_workbook(io.BytesIO(buf.getvalue()))
    ws = loaded["Configuration"]
    return {r[0].value: r[1].value for r in ws.iter_rows(max_col=2) if r[0].value is not None}


# ── Static builder ──────────────────────────────────────────────────


def test_static_records_preset_lifetime_and_amounts():
    cfg = _cfg(impact_api._build_single_product_static_workbook(
        archetype_name="BEV-LFP|Small", scope="all",
        scenarios=[("Base", _result())], stage_amounts_meta=META,
    ))
    assert cfg["Preset"] == "lifetime"
    assert cfg["Lifetime (yr)"] == 15
    # Per-stage amounts still in the one-liner (from the result echo).
    assert "Use Phase 15" in cfg["Stage amounts"]
    assert "Manufacturing 1" in cfg["Stage amounts"]


def test_static_default_one_year_exports_validly():
    meta = StageAmountsMeta(preset="1year", lifetime=15,
                            amounts={"Manufacturing": 1, "Use Phase": 1, "End of Life": 1})
    res = _result()
    res.stage_amounts = {"Manufacturing": 1, "Use Phase": 1, "End of Life": 1}
    cfg = _cfg(impact_api._build_single_product_static_workbook(
        archetype_name="A", scope="all", scenarios=[("Base", res)], stage_amounts_meta=meta,
    ))
    assert cfg["Preset"] == "1year"
    assert cfg["Lifetime (yr)"] == 15


def test_static_no_meta_falls_back_to_dash():
    cfg = _cfg(impact_api._build_single_product_static_workbook(
        archetype_name="A", scope="all", scenarios=[("Base", _result())],  # no meta
    ))
    assert cfg["Preset"] == "—"
    assert cfg["Lifetime (yr)"] == "—"
    # Amounts still recorded from the result echo.
    assert "Use Phase 15" in cfg["Stage amounts"]


# ── Prospective builder ─────────────────────────────────────────────


def test_prospective_records_preset_lifetime():
    cfg = _cfg(impact_api._build_single_product_prospective_workbook(
        archetype_name="BEV-LFP|Small", scope="all", runs=[_run()], stage_amounts_meta=META,
    ))
    assert cfg["Preset"] == "lifetime"
    assert cfg["Lifetime (yr)"] == 15
    assert "Use Phase 15" in cfg["Stage amounts"]


def test_prospective_no_meta_falls_back():
    cfg = _cfg(impact_api._build_single_product_prospective_workbook(
        archetype_name="A", scope="all", runs=[_run()],
    ))
    assert cfg["Preset"] == "—" and cfg["Lifetime (yr)"] == "—"


# ── Comparison builder ──────────────────────────────────────────────


def test_comparison_records_preset_lifetime():
    cfg = _cfg(impact_api._build_single_product_comparison_workbook(
        archetype_name="BEV-LFP|Small", scope="all",
        static_result=_result(), projected_runs=[_run()], stage_amounts_meta=META,
    ))
    assert cfg["Preset"] == "lifetime"
    assert cfg["Lifetime (yr)"] == 15
    assert "Use Phase 15" in cfg["Stage amounts"]


def test_comparison_no_meta_falls_back():
    cfg = _cfg(impact_api._build_single_product_comparison_workbook(
        archetype_name="A", scope="all", static_result=_result(), projected_runs=[_run()],
    ))
    assert cfg["Preset"] == "—" and cfg["Lifetime (yr)"] == "—"


# ── Regression: existing Configuration content preserved ────────────


def test_existing_configuration_content_unchanged():
    cfg = _cfg(impact_api._build_single_product_static_workbook(
        archetype_name="BEV-LFP|Small", scope="all",
        scenarios=[("Base", _result())], stage_amounts_meta=META,
    ))
    # Pre-5K rows still present (additive change).
    assert cfg["Archetype"] == "BEV-LFP|Small"
    assert "Stage amounts" in cfg
    assert "Stages included" in cfg
    assert "Indicators" in cfg
