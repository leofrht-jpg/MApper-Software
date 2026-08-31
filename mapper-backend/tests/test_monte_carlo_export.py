# SPDX-License-Identifier: MPL-2.0
"""The Monte Carlo workbook.

Routed through the shared helpers rather than a private builder:
``build_export_filename`` for the one filename scheme and ``excel_response``
for the demo stamping, so this export cannot drift from the others or forget
the warning.
"""

from __future__ import annotations

import io

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from mapper.api.monte_carlo import MC_DOMAIN, _build_monte_carlo_workbook
from mapper.main import app
from mapper.models.schemas import (
    ArchetypeLCAMethodDistribution,
    MonteCarloResult,
    PedigreeCoverage,
    ScoredInput,
    UnscoredMaterial,
    VarianceContributor,
)

client = TestClient(app)


def _dist(label="climate change | GWP100", det=12597.0, samples=None):
    return ArchetypeLCAMethodDistribution(
        method=["EF v3.1", "climate change", "GWP100"], method_label=label,
        unit="kg CO2-eq", deterministic=det, median=14250.0, mean=14400.0,
        p2_5=11500.0, p25=13400.0, p75=15200.0, p97_5=17800.0, gsd2=1.235,
        n_iterations=1000, seed=4242, samples=samples,
    )


def _result(**over) -> MonteCarloResult:
    base = dict(
        archetype_id="a1", archetype_name="A - Circular EV", scope="all",
        n_iterations=1000, seed=4242, elapsed_seconds=75.3,
        compute_database=None, parameter_scenario=None,
        distributions=[_dist()],
        contributors=[VarianceContributor(name="d_annual", kind="parameter", share=0.514, gsd2=1.236)],
        rows_with_uncertainty=1, rows_inherited=1, parameters_with_uncertainty=2,
        scored_inputs=[
            ScoredInput(name="Steel frame", kind="row",
                        pedigree={"reliability": 4, "completeness": 3},
                        basic_variance=0.0006, gsd2=1.215, inherited=True),
            ScoredInput(name="d_annual", kind="parameter",
                        pedigree={"reliability": 3, "temporal correlation": 4},
                        basic_variance=0.0006, gsd2=1.236),
        ],
        warnings=[],
    )
    base.update(over)
    return MonteCarloResult(**base)


def _coverage() -> PedigreeCoverage:
    return PedigreeCoverage(
        materials_total=148, materials_scored=47,
        archetype_materials_total=36, archetype_materials_scored=12,
        impact_share=0.82, method_label="climate change | GWP100",
        unit="kg CO2-eq",
        top_unscored=[UnscoredMaterial(name="Aluminium panels", share=0.189, impact=2100.0)],
    )


def _sheet(wb, name):
    return {c: [wb[name].cell(r, c).value for r in range(1, wb[name].max_row + 1)]
            for c in range(1, wb[name].max_column + 1)}


def _flat(wb, name) -> str:
    return "\n".join(
        " | ".join("" if v is None else str(v) for v in row)
        for row in wb[name].iter_rows(values_only=True)
    )


# ── the sheets ────────────────────────────────────────────────────────────────


def test_the_expected_sheets_are_present():
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    assert wb.sheetnames == [
        "Summary", "Distributions", "Variance contribution",
        "Pedigree scores", "Samples",
    ]


def test_the_seed_is_in_the_workbook():
    """A Monte Carlo result nobody can reproduce is not a research output."""
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    assert "4242" in _flat(wb, "Summary")
    # And beside every indicator, so a single sheet is enough to reproduce it.
    assert "4242" in _flat(wb, "Distributions")


def test_summary_carries_the_configuration():
    wb = _build_monte_carlo_workbook(
        _result(parameter_scenario="sa_early", compute_database="ei310_ssp2_2040"), _coverage()
    )
    txt = _flat(wb, "Summary")
    for expected in ("A - Circular EV", "sa_early", "Full lifecycle",
                     "ei310_ssp2_2040", "1000"):
        assert expected in txt, expected


def test_summary_carries_the_scoring_provenance():
    """A distribution is not interpretable without knowing what share of the
    foreground carried uncertainty."""
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    txt = _flat(wb, "Summary")
    assert "47 of 148" in txt          # materials scored, project
    assert "12 of 36" in txt           # ...this archetype
    assert "82.0%" in txt              # impact-weighted
    assert "not row count" in txt      # what the percentage means


def test_summary_states_the_lower_bound():
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    txt = _flat(wb, "Summary")
    assert "LOWER BOUND" in txt
    assert "12%" in txt
    assert "FIXED" in txt


def test_summary_says_when_nothing_was_scored():
    wb = _build_monte_carlo_workbook(
        _result(rows_with_uncertainty=0, rows_inherited=0,
                parameters_with_uncertainty=0, scored_inputs=[], contributors=[]),
        _coverage(),
    )
    txt = _flat(wb, "Summary")
    assert "BACKGROUND only" in txt


def test_summary_says_when_coverage_was_not_recorded():
    wb = _build_monte_carlo_workbook(_result(), None)
    assert "not recorded" in _flat(wb, "Summary")


def test_distributions_carry_the_deterministic_score_and_the_ratio():
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    hdr = [c.value for c in wb["Distributions"][1]]
    for col in ("Deterministic", "Median", "Median / deterministic",
                "GSD2 = exp(2*sigma)", "95% dispersion factor = p97.5 / median",
                "Seed"):
        assert col in hdr, col
    row = [c.value for c in wb["Distributions"][2]]
    assert row[hdr.index("Deterministic")] == 12597.0
    assert row[hdr.index("Median")] == 14250.0
    assert row[hdr.index("Median / deterministic")] == pytest.approx(14250 / 12597, rel=1e-4)


def test_a_zero_deterministic_does_not_divide_by_zero():
    wb = _build_monte_carlo_workbook(_result(distributions=[_dist(det=0.0)]), _coverage())
    hdr = [c.value for c in wb["Distributions"][1]]
    assert [c.value for c in wb["Distributions"][2]][hdr.index("Median / deterministic")] == "n/a"


def test_variance_contribution_matches_the_on_screen_chart():
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    txt = _flat(wb, "Variance contribution")
    assert "d_annual" in txt and "parameter" in txt
    assert "approximate attribution" in txt


def test_pedigree_sheet_records_the_six_values_and_the_gsd():
    """So the run is reproducible from the workbook."""
    wb = _build_monte_carlo_workbook(_result(), _coverage())
    hdr = [c.value for c in wb["Pedigree scores"][1]]
    for col in ("Reliability", "Completeness", "Temporal", "Geographical",
                "Technological", "Basic variance", "GSD2 = exp(2*sigma)", "Source"):
        assert col in hdr, col
    row = [c.value for c in wb["Pedigree scores"][2]]
    assert row[hdr.index("Reliability")] == 4
    assert row[hdr.index("Completeness")] == 3
    assert row[hdr.index("Temporal")] == ""       # unset stays blank, not 1
    assert row[hdr.index("GSD2 = exp(2*sigma)")] == pytest.approx(1.215)
    assert row[hdr.index("Source")] == "material library"


def test_pedigree_sheet_says_so_when_nothing_was_scored():
    wb = _build_monte_carlo_workbook(_result(scored_inputs=[]), _coverage())
    assert "Nothing was scored" in _flat(wb, "Pedigree scores")


# ── samples ───────────────────────────────────────────────────────────────────


def test_samples_sheet_holds_one_column_per_indicator():
    wb = _build_monte_carlo_workbook(
        _result(distributions=[
            _dist(label="climate change", samples=[1.0, 2.0, 3.0]),
            _dist(label="acidification", samples=[4.0, 5.0, 6.0]),
        ]),
        _coverage(),
    )
    hdr = [c.value for c in wb["Samples"][1]]
    assert hdr == ["Iteration", "climate change", "acidification"]
    assert [c.value for c in wb["Samples"][2]] == [1, 1.0, 4.0]
    assert wb["Samples"].max_row == 4          # header + 3 iterations


def test_samples_sheet_is_a_NOTE_when_draws_were_not_retained():
    """Written, not omitted. An absent sheet is ambiguous with 'this build does
    not produce one', and a reader comparing two workbooks could not tell."""
    wb = _build_monte_carlo_workbook(_result(distributions=[_dist(samples=None)]), _coverage())
    assert "Samples" in wb.sheetnames
    txt = _flat(wb, "Samples")
    assert "not retained" in txt
    assert "keep_samples" in txt
    assert "percentiles on Distributions are unaffected" in txt


def test_ragged_sample_lengths_do_not_truncate_the_longest():
    wb = _build_monte_carlo_workbook(
        _result(distributions=[
            _dist(label="a", samples=[1.0, 2.0, 3.0]),
            _dist(label="b", samples=[4.0]),
        ]),
        _coverage(),
    )
    assert wb["Samples"].max_row == 4
    assert [c.value for c in wb["Samples"][4]] == [3, 3.0, None]


# ── the shared helpers ────────────────────────────────────────────────────────


def test_the_filename_uses_the_shared_scheme_with_an_MC_token():
    from mapper.api.bom import build_export_filename

    assert MC_DOMAIN == "MC"
    # Single-product: the archetype takes the system slot, never a subsystem.
    assert build_export_filename("A - Circular EV", [], MC_DOMAIN) == "A_-_Circular_EV_MC.xlsx"


def test_the_route_returns_a_workbook_through_excel_response():
    r = client.post(
        "/api/lca/monte-carlo/export",
        json={"result": _result().model_dump(), "coverage": _coverage().model_dump()},
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert "A_-_Circular_EV_MC.xlsx" in r.headers["content-disposition"]
    wb = load_workbook(io.BytesIO(r.content))
    assert "Summary" in wb.sheetnames


def test_a_demo_project_export_is_stamped_and_prefixed():
    """Routed through excel_response, so the demo warning is by construction."""
    from mapper.api import cohort_export

    wb = _build_monte_carlo_workbook(_result(), _coverage())
    resp = cohort_export.excel_response(wb, "x_MC.xlsx", kind="data", is_demo=True)
    assert "DEMO_x_MC.xlsx" in resp.headers["content-disposition"]
    got = load_workbook(io.BytesIO(resp.body))
    assert "SYNTHETIC DEMO DATA" in str(got["Summary"].cell(1, 1).value)
