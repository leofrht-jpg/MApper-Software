# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Patch 4AG.4 — backend tests for the multi-product LCA comparison
export workbook.

Coverage:
  - Workbook builds for archetype-only + scope='all' → emits per-item
    stage-breakdown sheets
  - Workbook builds for activity-only → NO stage-breakdown sheets
    (activities have no lifecycle stages)
  - Mixed-type workbook: archetype items get stage sheets; activity
    items don't
  - Errors sheet present when partial success; absent when full
    success
  - Configuration sheet captures all metadata (scope, compute_database,
    timestamp, items list)
  - Comparison (wide) failed-item row shows "—" for method values
    PLUS error_message in trailing column
  - Comparison (long) skips failed items entirely
  - Empty items → 400 from the endpoint
  - All items failed → workbook still produces, with empty comparison
    bodies and a populated Errors sheet

Uses openpyxl to inspect the produced workbook directly — no real
LCA needed; we synthesize MultiProductLCAResult objects.
"""
from __future__ import annotations

import asyncio
import io

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from mapper.api.impact import _build_multi_product_workbook, post_export_multi_product
from mapper.models.schemas import (
    ActivityContribution,
    ActivityLCAMethodResult,
    ActivityLCAResult,
    ArchetypeLCACalculateResult,
    ArchetypeLCAMethodResult,
    MultiProductExportRequest,
    MultiProductItemResult,
    MultiProductLCAResult,
)


def _arc_result(arc_id: str, name: str, score: float, with_stages: bool = True) -> ArchetypeLCACalculateResult:
    return ArchetypeLCACalculateResult(
        archetype_id=arc_id, archetype_name=name,
        scope="all", amount=1.0, stage_amounts={},
        stages_included=["Manufacturing", "Use Phase"] if with_stages else [],
        results=[
            ArchetypeLCAMethodResult(
                method=["EF v3.1", "climate change", "GWP100"],
                method_label="EF v3.1 › climate change › GWP100",
                score=score, unit="kg CO2 eq", contributions=[],
            ),
            ArchetypeLCAMethodResult(
                method=["EF v3.1", "water use", "deprivation"],
                method_label="EF v3.1 › water use › deprivation",
                score=score * 0.001, unit="m3 depriv.", contributions=[],
            ),
        ],
        stage_breakdown={
            "EF v3.1 › climate change › GWP100": {"Manufacturing": score * 0.7, "Use Phase": score * 0.3},
            "EF v3.1 › water use › deprivation": {"Manufacturing": score * 0.0005, "Use Phase": score * 0.0005},
        } if with_stages else None,
        elapsed_seconds=0.1,
    )


def _act_result(score: float) -> ActivityLCAResult:
    return ActivityLCAResult(
        results=[
            ActivityLCAMethodResult(
                method=["EF v3.1", "climate change", "GWP100"],
                method_label="EF v3.1 › climate change › GWP100",
                score=score, unit="kg CO2 eq",
                contributions=[ActivityContribution(
                    name="x", location="GLO", database="ei", code="c1",
                    demand_amount=1.0, demand_unit="kg", impact=score, percentage=100.0,
                )],
            ),
        ],
        elapsed_seconds=0.05,
    )


def _build_workbook_bytes(body: MultiProductExportRequest) -> bytes:
    wb = _build_multi_product_workbook(body)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _load(workbook_bytes: bytes):
    return load_workbook(io.BytesIO(workbook_bytes))


# ── Sheet inventory ────────────────────────────────────────────────


def test_archetype_only_with_stages_emits_stage_breakdown_sheets() -> None:
    """Each archetype item carrying stage_breakdown gets its own
    'SB_<label>' sheet. Activities never do."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="arc-bev", label="BEV-LFP",
                status="success", archetype_result=_arc_result("arc-bev", "BEV-LFP", 1000.0),
            ),
            MultiProductItemResult(
                type="archetype", item_id="arc-icev", label="ICEV petrol",
                status="success", archetype_result=_arc_result("arc-icev", "ICEV petrol", 2000.0),
            ),
        ],
        elapsed_seconds=0.5, success_count=2, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all", compute_database=None)
    wb = _load(_build_workbook_bytes(body))
    assert "Configuration" in wb.sheetnames
    assert "Comparison (wide)" in wb.sheetnames
    assert "Comparison (long)" in wb.sheetnames
    # Per-item stage breakdown sheets, prefixed SB_.
    sb_sheets = [n for n in wb.sheetnames if n.startswith("SB_")]
    assert len(sb_sheets) == 2
    assert "Errors" not in wb.sheetnames  # full success


def test_activity_only_omits_stage_breakdown_sheets() -> None:
    """Activities have no lifecycle stages → no SB_ sheets."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="activity", item_id="ei|c1", label="battery",
                status="success", activity_result=_act_result(500.0),
            ),
        ],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    assert "Configuration" in wb.sheetnames
    assert "Comparison (wide)" in wb.sheetnames
    assert "Comparison (long)" in wb.sheetnames
    assert not any(n.startswith("SB_") for n in wb.sheetnames)
    assert "Errors" not in wb.sheetnames


def test_mixed_type_partial_stages_only_for_archetype_items() -> None:
    """Mixed-type request: archetype item gets SB sheet; activity
    item does NOT. Single SB_ sheet expected."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="arc-bev", label="BEV-LFP",
                status="success", archetype_result=_arc_result("arc-bev", "BEV-LFP", 1000.0),
            ),
            MultiProductItemResult(
                type="activity", item_id="ei|c1", label="battery",
                status="success", activity_result=_act_result(500.0),
            ),
        ],
        elapsed_seconds=0.3, success_count=2, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    sb_sheets = [n for n in wb.sheetnames if n.startswith("SB_")]
    assert len(sb_sheets) == 1
    # The single SB sheet is for the archetype item, not the activity.
    assert "BEV-LFP" in sb_sheets[0]


def test_archetype_without_stage_breakdown_omits_sb_sheet() -> None:
    """Specific-scope archetype items have no stage_breakdown (the
    backend returns None for non-'all' scopes). No SB sheet for them."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="arc-x", label="X",
                status="success",
                archetype_result=_arc_result("arc-x", "X", 100.0, with_stages=False),
            ),
        ],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="inflows")
    wb = _load(_build_workbook_bytes(body))
    assert not any(n.startswith("SB_") for n in wb.sheetnames)


# ── Errors sheet ───────────────────────────────────────────────────


def test_errors_sheet_present_on_partial_success() -> None:
    """When at least one item failed, an Errors sheet appears with
    per-item rows (label, type, identifier, message)."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="arc-ok", label="OK",
                status="success", archetype_result=_arc_result("arc-ok", "OK", 100.0),
            ),
            MultiProductItemResult(
                type="archetype", item_id="arc-bad", label="bad arc",
                status="error", error_message="archetype not found",
            ),
        ],
        elapsed_seconds=0.2, success_count=1, error_count=1,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    assert "Errors" in wb.sheetnames
    err = wb["Errors"]
    # Row 1 = header, row 2 = the failed item.
    header = [c.value for c in err[1]]
    assert header == ["Item", "Type", "Identifier", "Error message"]
    row2 = [c.value for c in err[2]]
    assert row2[0] == "bad arc"
    assert row2[1] == "archetype"
    assert row2[2] == "arc-bad"
    assert "not found" in row2[3]


def test_errors_sheet_absent_on_full_success() -> None:
    res = MultiProductLCAResult(
        items=[MultiProductItemResult(
            type="archetype", item_id="x", label="X", status="success",
            archetype_result=_arc_result("x", "X", 100.0),
        )],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    assert "Errors" not in wb.sheetnames


def test_all_failed_workbook_still_produced_with_errors_sheet() -> None:
    """Degenerate case: every item failed. Workbook still produces;
    Comparison sheets render with empty body; Errors sheet lists
    every failed item."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="a1", label="a1", status="error",
                error_message="boom",
            ),
            MultiProductItemResult(
                type="archetype", item_id="a2", label="a2", status="error",
                error_message="kaboom",
            ),
        ],
        elapsed_seconds=0.1, success_count=0, error_count=2,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    assert "Errors" in wb.sheetnames
    err = wb["Errors"]
    assert err.max_row == 3  # header + 2 failed items
    # Comparison (wide) has rows for both items but with "—" for method values.
    wide = wb["Comparison (wide)"]
    assert wide.max_row == 3  # header + 2 items
    # Comparison (long) is empty body (only header) — failed items contribute nothing.
    long = wb["Comparison (long)"]
    assert long.max_row == 1


# ── Configuration sheet ────────────────────────────────────────────


def test_configuration_sheet_captures_metadata() -> None:
    res = MultiProductLCAResult(
        items=[MultiProductItemResult(
            type="archetype", item_id="x", label="X", status="success",
            archetype_result=_arc_result("x", "X", 100.0),
        )],
        elapsed_seconds=0.5, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(
        result=res, scope="inflows",
        compute_database="ecoinvent-3.10_remind_SSP2_2030",
        computed_at="2026-05-12T10:30:00+00:00",
    )
    wb = _load(_build_workbook_bytes(body))
    cfg = wb["Configuration"]
    # Read all cell values for inspection.
    text = "\n".join(
        " | ".join(str(c.value) if c.value is not None else "" for c in row)
        for row in cfg.iter_rows(values_only=False)
    )
    assert "2026-05-12T10:30:00+00:00" in text
    assert "inflows" in text
    assert "remind_SSP2_2030" in text
    assert "Items succeeded" in text
    assert "Compute elapsed" in text


def test_configuration_items_table_lists_all_items_with_status() -> None:
    """The Configuration sheet's 'Items' table includes one row per
    item — success AND failure — so users get a roll-call of what
    was attempted alongside what succeeded."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="a1", label="A1", status="success",
                archetype_result=_arc_result("a1", "A1", 100.0),
            ),
            MultiProductItemResult(
                type="archetype", item_id="a2", label="A2", status="error",
                error_message="x",
            ),
            MultiProductItemResult(
                type="activity", item_id="ei|c1", label="battery", status="success",
                activity_result=_act_result(50.0),
            ),
        ],
        elapsed_seconds=0.2, success_count=2, error_count=1,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    cfg = wb["Configuration"]
    # Walk rows, find the header "#" row and collect the items table.
    items_seen: list[str] = []
    in_items = False
    for row in cfg.iter_rows(values_only=True):
        if row and row[0] == "#":
            in_items = True
            continue
        if in_items and row and isinstance(row[0], int):
            items_seen.append(str(row[3]))  # column 4 = Label
    assert items_seen == ["A1", "A2", "battery"]


# ── Comparison (wide) ──────────────────────────────────────────────


def test_wide_sheet_failed_items_show_dash_and_error_message() -> None:
    """Failed items still appear in the wide sheet (so users see the
    full roster) but with '—' for method values and the
    error_message in the trailing 'Error' column."""
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="ok", label="OK", status="success",
                archetype_result=_arc_result("ok", "OK", 1000.0),
            ),
            MultiProductItemResult(
                type="archetype", item_id="bad", label="bad", status="error",
                error_message="archetype not found",
            ),
        ],
        elapsed_seconds=0.2, success_count=1, error_count=1,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    wide = wb["Comparison (wide)"]
    rows = list(wide.iter_rows(values_only=True))
    # Header row 0; OK row 1; bad row 2.
    bad_row = rows[2]
    # Method-score cells (4 onwards) are "—" for the failed item.
    assert bad_row[3] == "—"
    # Trailing column = error message.
    assert "not found" in str(bad_row[-1])


def test_wide_sheet_method_scores_in_scientific_format() -> None:
    res = MultiProductLCAResult(
        items=[MultiProductItemResult(
            type="archetype", item_id="x", label="X", status="success",
            archetype_result=_arc_result("x", "X", 1234.5),
        )],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    wide = wb["Comparison (wide)"]
    # Row 2, column 4 = first method's score for item X. Number format
    # must be scientific (Patch 4AG.4 spec).
    cell = wide.cell(row=2, column=4)
    assert cell.number_format == "0.000E+00"
    assert isinstance(cell.value, (int, float))


# ── Comparison (long) ──────────────────────────────────────────────


def test_long_sheet_one_row_per_item_method_pair_skipping_failed() -> None:
    res = MultiProductLCAResult(
        items=[
            MultiProductItemResult(
                type="archetype", item_id="a", label="A", status="success",
                archetype_result=_arc_result("a", "A", 100.0),
            ),
            MultiProductItemResult(
                type="archetype", item_id="b", label="B", status="error",
                error_message="x",
            ),
        ],
        elapsed_seconds=0.1, success_count=1, error_count=1,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    wb = _load(_build_workbook_bytes(body))
    long = wb["Comparison (long)"]
    rows = list(long.iter_rows(values_only=True))
    # Header + 2 rows (one per method) for item A; failed B contributes none.
    assert long.max_row == 3
    # Both rows are for item A.
    for row in rows[1:]:
        assert row[0] == "A"


# ── Endpoint validation ────────────────────────────────────────────


def test_endpoint_rejects_empty_items() -> None:
    res = MultiProductLCAResult(items=[], elapsed_seconds=0.0, success_count=0, error_count=0)
    body = MultiProductExportRequest(result=res, scope="all")
    with pytest.raises(HTTPException) as exc:
        asyncio.run(post_export_multi_product(body))
    assert exc.value.status_code == 400


def test_endpoint_produces_xlsx_filename_with_date() -> None:
    import datetime
    res = MultiProductLCAResult(
        items=[MultiProductItemResult(
            type="archetype", item_id="x", label="X", status="success",
            archetype_result=_arc_result("x", "X", 100.0),
        )],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")
    response = asyncio.run(post_export_multi_product(body))
    date_tag = datetime.date.today().isoformat()
    cd = response.headers["content-disposition"]
    assert "MultiProduct_Comparison" in cd
    assert date_tag in cd
    assert cd.endswith('.xlsx"')


# ── Imports needed for Font (used in builder) ──────────────────────


def test_imports_smoke() -> None:
    """The builder uses openpyxl.styles.Font (imported inside the
    function). Smoke test to ensure the import chain works."""
    from openpyxl.styles import Font
    assert Font is not None
