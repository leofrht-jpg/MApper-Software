# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Patch 5J — the multi-product comparison export records per-item stage
amounts (preset + lifetime + per-stage values) so a run is reproducible from
the export alone.

The "Stage amounts" sheet reads `stage_amounts_meta` (threaded from the
frontend store) for preset + lifetime; per-stage amounts come from that meta,
falling back to the result-echoed `stage_amounts` when no meta is supplied.

Synthesizes MultiProductLCAResult + MultiProductExportRequest; inspects the
workbook with openpyxl (no real LCA).
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api.impact import _build_multi_product_workbook
from mapper.models.schemas import (
    ActivityLCAMethodResult,
    ActivityLCAResult,
    ArchetypeLCACalculateResult,
    ArchetypeLCAMethodResult,
    MultiProductExportRequest,
    MultiProductItemResult,
    MultiProductLCAResult,
    StageAmountsMeta,
)


def _arc_result(arc_id: str, name: str, score: float, stage_amounts: dict | None = None) -> ArchetypeLCACalculateResult:
    return ArchetypeLCACalculateResult(
        archetype_id=arc_id, archetype_name=name,
        scope="all", amount=1.0, stage_amounts=stage_amounts or {},
        stages_included=["Manufacturing", "Use Phase", "End of Life"],
        results=[ArchetypeLCAMethodResult(
            method=["EF v3.1", "climate change", "GWP100"],
            method_label="EF v3.1 › climate change › GWP100",
            score=score, unit="kg CO2 eq", contributions=[],
        )],
        stage_breakdown=None,
        elapsed_seconds=0.1,
    )


def _arc_item(arc_id: str, name: str, score: float, stage_amounts: dict | None = None) -> MultiProductItemResult:
    return MultiProductItemResult(
        type="archetype", item_id=arc_id, label=name, status="success",
        archetype_result=_arc_result(arc_id, name, score, stage_amounts),
    )


def _build(body: MultiProductExportRequest):
    wb = _build_multi_product_workbook(body)
    buf = io.BytesIO()
    wb.save(buf)
    return load_workbook(io.BytesIO(buf.getvalue()))


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def test_stage_amounts_sheet_records_preset_lifetime_and_per_stage_values() -> None:
    """BEV-LFP: lifetime / Use 15; BEV-NCA: lifetime / Use 10 — preset,
    lifetime, and per-stage amounts all recorded per item."""
    res = MultiProductLCAResult(
        items=[
            _arc_item("arc-lfp", "BEV-LFP", 1000.0),
            _arc_item("arc-nca", "BEV-NCA", 1200.0),
        ],
        elapsed_seconds=0.5, success_count=2, error_count=0,
    )
    body = MultiProductExportRequest(
        result=res, scope="all",
        stage_amounts_meta={
            "arc-lfp": StageAmountsMeta(preset="lifetime", lifetime=15,
                                        amounts={"Manufacturing": 1, "Use Phase": 15, "End of Life": 1}),
            "arc-nca": StageAmountsMeta(preset="lifetime", lifetime=10,
                                        amounts={"Manufacturing": 1, "Use Phase": 10, "End of Life": 1}),
        },
    )
    wb = _build(body)
    assert "Stage amounts" in wb.sheetnames
    rows = _rows(wb["Stage amounts"])
    # Row 1 title, row 2 blank, row 3 header, rows 4+ data.
    header = rows[2]
    assert header[:4] == ("#", "Item", "Preset", "Lifetime (yr)")
    stage_cols = list(header[4:])
    assert stage_cols == ["Manufacturing", "Use Phase", "End of Life"]

    by_item = {r[1]: r for r in rows[3:]}
    lfp, nca = by_item["BEV-LFP"], by_item["BEV-NCA"]
    # preset + lifetime
    assert lfp[2] == "lifetime" and lfp[3] == 15
    assert nca[2] == "lifetime" and nca[3] == 10
    # per-stage amounts (Use Phase scales; Manufacturing/EoL one-time)
    use_idx = 4 + stage_cols.index("Use Phase")
    mfg_idx = 4 + stage_cols.index("Manufacturing")
    assert lfp[use_idx] == 15 and nca[use_idx] == 10
    assert lfp[mfg_idx] == 1 and nca[mfg_idx] == 1


def test_default_one_year_config_exports_validly() -> None:
    """Default 1year / all-ones items export with the defaults shown."""
    res = MultiProductLCAResult(
        items=[_arc_item("arc-a", "Product A", 500.0)],
        elapsed_seconds=0.2, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(
        result=res, scope="all",
        stage_amounts_meta={
            "arc-a": StageAmountsMeta(preset="1year", lifetime=15,
                                      amounts={"Manufacturing": 1, "Use Phase": 1, "End of Life": 1}),
        },
    )
    wb = _build(body)
    rows = _rows(wb["Stage amounts"])
    data = rows[3]
    assert data[1] == "Product A"
    assert data[2] == "1year"
    assert tuple(data[4:]) == (1, 1, 1)  # all-ones


def test_falls_back_to_result_amounts_when_no_meta() -> None:
    """Backward compat: no stage_amounts_meta → per-stage amounts come from
    the result-echoed stage_amounts; preset/lifetime show as '—'."""
    res = MultiProductLCAResult(
        items=[_arc_item("arc-a", "Product A", 500.0,
                         stage_amounts={"Manufacturing": 1, "Use Phase": 12})],
        elapsed_seconds=0.2, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(result=res, scope="all")  # no meta
    wb = _build(body)
    assert "Stage amounts" in wb.sheetnames
    rows = _rows(wb["Stage amounts"])
    header, data = rows[2], rows[3]
    assert data[2] == "—" and data[3] == "—"  # preset/lifetime unknown
    use_idx = 4 + list(header[4:]).index("Use Phase")
    assert data[use_idx] == 12  # amounts recovered from the result


def test_no_stage_amounts_sheet_for_activity_only() -> None:
    """Activities have no BOM stages → no Stage amounts sheet."""
    res = MultiProductLCAResult(
        items=[MultiProductItemResult(
            type="activity", item_id="ei|c1", label="electricity",
            status="success",
            activity_result=ActivityLCAResult(results=[ActivityLCAMethodResult(
                method=["EF v3.1", "climate change", "GWP100"],
                method_label="EF v3.1 › climate change › GWP100",
                score=3.0, unit="kg CO2 eq", contributions=[],
            )], elapsed_seconds=0.05),
        )],
        elapsed_seconds=0.1, success_count=1, error_count=0,
    )
    wb = _build(MultiProductExportRequest(result=res, scope="all"))
    assert "Stage amounts" not in wb.sheetnames


def test_existing_sheets_unchanged_by_amounts_addition() -> None:
    """Regression: the pre-5J sheets still exist with their content; the
    Stage amounts sheet is purely additive."""
    res = MultiProductLCAResult(
        items=[_arc_item("arc-a", "Product A", 500.0)],
        elapsed_seconds=0.2, success_count=1, error_count=0,
    )
    body = MultiProductExportRequest(
        result=res, scope="all",
        stage_amounts_meta={"arc-a": StageAmountsMeta(preset="1year", lifetime=15,
                                                      amounts={"Manufacturing": 1})},
    )
    wb = _build(body)
    for name in ("Configuration", "Comparison (wide)", "Comparison (long)"):
        assert name in wb.sheetnames
    # Configuration items table still has the item.
    cfg_text = "\n".join(str(r) for r in _rows(wb["Configuration"]))
    assert "Product A" in cfg_text
    # Wide sheet still carries the method score.
    wide_text = "\n".join(str(r) for r in _rows(wb["Comparison (wide)"]))
    assert "Product A" in wide_text
