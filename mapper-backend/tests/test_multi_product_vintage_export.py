# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Per-item-vintage activity export — the multi-product workbook records each
activity item's vintage (database / SSP / year) on a "Vintages" sheet, so a
run comparing an activity across ecoinvent + premise SSP×year is reproducible
from the export alone. Mirrors the Stage amounts provenance sheet (5J).
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api.impact import _build_multi_product_workbook
from mapper.models.schemas import (
    ActivityLCAMethodResult,
    ActivityLCAResult,
    ActivityVintageMeta,
    MultiProductExportRequest,
    MultiProductItemResult,
    MultiProductLCAResult,
)

BASE_DB = "ecoinvent-3.10-cutoff"
SSP1_DB = f"{BASE_DB}_premise_remind_ssp1-pkbudg1150_2040"
SSP5_DB = f"{BASE_DB}_premise_remind_ssp5-pkbudg1150_2040"
CODE = "elec"


def _act_result(score: float) -> ActivityLCAResult:
    return ActivityLCAResult(
        results=[ActivityLCAMethodResult(
            method=["EF v3.1", "climate change", "GWP100"],
            method_label="EF v3.1 › climate change › GWP100",
            score=score, unit="kg CO2 eq", contributions=[],
        )],
        elapsed_seconds=0.05,
    )


def _vintage_item(db: str, vintage_label: str, score: float) -> MultiProductItemResult:
    return MultiProductItemResult(
        type="activity", item_id=f"{db}|{CODE}",
        label=f"electricity, low voltage [{vintage_label}]",
        status="success", activity_result=_act_result(score),
    )


def _load(body: MultiProductExportRequest):
    wb = _build_multi_product_workbook(body)
    buf = io.BytesIO()
    wb.save(buf)
    return load_workbook(io.BytesIO(buf.getvalue()))


def _result() -> MultiProductLCAResult:
    # Same activity, three vintages — each its own DB. SSP1 > SSP5 (audited).
    return MultiProductLCAResult(
        items=[
            _vintage_item(BASE_DB, "ecoinvent", 0.5),
            _vintage_item(SSP1_DB, "SSP1 2040", 0.031),
            _vintage_item(SSP5_DB, "SSP5 2040", 0.020),
        ],
        elapsed_seconds=0.1, success_count=3, error_count=0,
    )


def _vintage_meta() -> dict[str, ActivityVintageMeta]:
    return {
        f"{BASE_DB}|{CODE}": ActivityVintageMeta(label="ecoinvent", database=BASE_DB, base_database=BASE_DB),
        f"{SSP1_DB}|{CODE}": ActivityVintageMeta(label="SSP1 2040", database=SSP1_DB, base_database=BASE_DB, iam="remind", ssp="SSP1-PkBudg1150", year=2040),
        f"{SSP5_DB}|{CODE}": ActivityVintageMeta(label="SSP5 2040", database=SSP5_DB, base_database=BASE_DB, iam="remind", ssp="SSP5-PkBudg1150", year=2040),
    }


def test_vintages_sheet_records_each_activity_items_database_and_vintage() -> None:
    body = MultiProductExportRequest(
        result=_result(), scope="all", activity_vintage_meta=_vintage_meta(),
    )
    wb = _load(body)
    assert "Vintages" in wb.sheetnames
    ws = wb["Vintages"]
    rows = list(ws.iter_rows(values_only=True))
    # Header row is the 3rd (title, blank, header).
    header = rows[2]
    assert header == ("#", "Item", "Vintage", "Database", "Base database", "IAM", "SSP", "Year")
    # Flatten the data rows into a {database: row} lookup.
    data = {r[3]: r for r in rows[3:]}
    assert set(data.keys()) == {BASE_DB, SSP1_DB, SSP5_DB}
    # Each item records its own DB + SSP + year.
    assert data[SSP1_DB][2] == "SSP1 2040"      # Vintage label
    assert data[SSP1_DB][6] == "SSP1-PkBudg1150"  # SSP
    assert data[SSP1_DB][7] == 2040               # Year
    assert data[BASE_DB][2] == "ecoinvent"
    assert data[BASE_DB][7] == "—"                # static has no year


def test_vintages_sheet_absent_when_no_activity_items() -> None:
    """Archetype-only runs don't get a Vintages sheet."""
    from mapper.models.schemas import ArchetypeLCACalculateResult, ArchetypeLCAMethodResult
    arc = ArchetypeLCACalculateResult(
        archetype_id="arc", archetype_name="BEV", scope="all", amount=1.0,
        stage_amounts={}, stages_included=[],
        results=[ArchetypeLCAMethodResult(
            method=["m"], method_label="M", score=1.0, unit="kg", contributions=[])],
        stage_breakdown=None, elapsed_seconds=0.1,
    )
    body = MultiProductExportRequest(
        result=MultiProductLCAResult(
            items=[MultiProductItemResult(type="archetype", item_id="arc", label="BEV", status="success", archetype_result=arc)],
            success_count=1, error_count=0,
        ),
        scope="all",
    )
    wb = _load(body)
    assert "Vintages" not in wb.sheetnames


def test_vintages_sheet_falls_back_to_db_from_item_id_without_meta() -> None:
    """Older clients that don't thread activity_vintage_meta still get a
    Vintages sheet, with the database derived from the item_id."""
    body = MultiProductExportRequest(result=_result(), scope="all")  # no meta
    wb = _load(body)
    assert "Vintages" in wb.sheetnames
    ws = wb["Vintages"]
    rows = list(ws.iter_rows(values_only=True))
    databases = {r[3] for r in rows[3:]}
    assert databases == {BASE_DB, SSP1_DB, SSP5_DB}
