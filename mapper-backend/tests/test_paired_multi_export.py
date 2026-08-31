# SPDX-License-Identifier: MPL-2.0
"""The multi-item workbook: an Item column, and pairwise differences on their
own sheet."""
from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from mapper.api.monte_carlo import _build_monte_carlo_multi_workbook
from mapper.main import app
from mapper.models.schemas import (
    ArchetypeLCAMethodDistribution, ItemDistribution,
    MonteCarloMultiResult, PairwiseDifference,
)

client = TestClient(app)


def _dist(label="climate change | GWP100", det=10.0, med=11.0, samples=None):
    return ArchetypeLCAMethodDistribution(
        method=["EF v3.1", "climate change", "GWP100"], method_label=label,
        unit="kg CO2-eq", deterministic=det, median=med, mean=med,
        p2_5=med * 0.9, p25=med * 0.95, p75=med * 1.05, p97_5=med * 1.1,
        gsd2=1.2, n_iterations=200, seed=77, samples=samples,
    )


def _result(**over):
    base = dict(
        scope="all", n_iterations=200, seed=77, elapsed_seconds=35.2,
        compute_database=None, parameter_scenario=None,
        items=[
            ItemDistribution(archetype_id="a", archetype_name="A - Circular EV",
                             distributions=[_dist(det=10.0, med=11.0, samples=[1.0, 2.0])]),
            ItemDistribution(archetype_id="b", archetype_name="A0 - Reference EV",
                             distributions=[_dist(det=12.0, med=13.0, samples=[3.0, 4.0])]),
        ],
        differences=[PairwiseDifference(
            method=["EF v3.1", "climate change", "GWP100"],
            method_label="climate change | GWP100", unit="kg CO2-eq",
            a_id="a", a_name="A - Circular EV", b_id="b", b_name="A0 - Reference EV",
            deterministic=-2.0, median=-2.1, mean=-2.1,
            p2_5=-2.4, p25=-2.2, p75=-2.0, p97_5=-1.8,
            fraction_a_lower=1.0, correlation=0.9861,
        )],
        warnings=[],
    )
    base.update(over)
    return MonteCarloMultiResult(**base)


def _flat(wb, name):
    return "\n".join(" | ".join("" if v is None else str(v) for v in row)
                     for row in wb[name].iter_rows(values_only=True))


def test_the_sheets():
    wb = _build_monte_carlo_multi_workbook(_result())
    assert wb.sheetnames == ["Summary", "Distributions", "Pairwise differences", "Samples"]


def _data_rows(ws) -> list[list]:
    """Data rows only, stopping at the blank line before the trailing notes.

    The Distributions sheet ends with a blank row and then the GSD2 /
    dispersion / migration definitions, so that a reader of the workbook has
    the formulas beside the numbers. Reading to the last populated row would
    swallow them as data.
    """
    out = []
    for r in ws.iter_rows(min_row=2):
        vals = [c.value for c in r]
        if all(v is None for v in vals):
            break
        out.append(vals)
    return out


def test_distributions_carry_an_Item_column_beside_Sensitivity_case():
    """Matching how every other multi-axis export carries its discriminator."""
    wb = _build_monte_carlo_multi_workbook(_result())
    hdr = [c.value for c in wb["Distributions"][1]]
    assert hdr[0] == "Item"
    assert hdr[1] == "Sensitivity case"
    rows = _data_rows(wb["Distributions"])
    assert [r[0] for r in rows] == ["A - Circular EV", "A0 - Reference EV"]
    assert all(r[1] == "Base" for r in rows)


def test_items_keep_comparison_order():
    wb = _build_monte_carlo_multi_workbook(_result())
    rows = _data_rows(wb["Distributions"])
    assert rows[0][0] == "A - Circular EV"       # not sorted by value or name


def test_pairwise_sheet_carries_the_correlation_and_the_claim():
    wb = _build_monte_carlo_multi_workbook(_result())
    hdr = [c.value for c in wb["Pairwise differences"][1]]
    for col in ("A", "B", "Deterministic (A-B)", "Median (A-B)", "A lower in", "Correlation(A,B)"):
        assert col in hdr, col
    row = [c.value for c in wb["Pairwise differences"][2]]
    assert row[hdr.index("A lower in")] == "100.0%"
    assert row[hdr.index("Correlation(A,B)")] == 0.9861
    # Stated as information, not a caveat.
    assert "INFORMATIVE, not a warning" in _flat(wb, "Pairwise differences")


def test_summary_says_the_sampling_is_paired_and_marginals_are_unaffected():
    txt = _flat(_build_monte_carlo_multi_workbook(_result()), "Summary")
    assert "PAIRED" in txt
    assert "Marginals are unaffected" in txt
    assert "77" in txt                     # the seed
    assert "LOWER BOUND" in txt


def test_a_single_item_run_says_so_rather_than_an_empty_pair_sheet():
    r = _result(differences=[], items=[_result().items[0]])
    assert "no pairwise difference" in _flat(_build_monte_carlo_multi_workbook(r), "Pairwise differences")


def test_samples_note_when_not_retained():
    r = _result(items=[
        ItemDistribution(archetype_id="a", archetype_name="A", distributions=[_dist(samples=None)]),
    ])
    assert "not retained" in _flat(_build_monte_carlo_multi_workbook(r), "Samples")


def test_the_route_returns_a_workbook_named_for_the_comparison():
    r = client.post("/api/lca/monte-carlo/multi/export", json={"result": _result().model_dump()})
    assert r.status_code == 200
    # build_export_filename: first item + the rest, MC token.
    assert "A_-_Circular_EV+A0_-_Reference_EV_MC.xlsx" in r.headers["content-disposition"]
    assert "Pairwise differences" in load_workbook(io.BytesIO(r.content)).sheetnames
