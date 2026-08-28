# SPDX-License-Identifier: MPL-2.0
"""The two scoring surfaces: the parameter editor's payload, and the BOM
workbook's pedigree columns.

The load-bearing property throughout is that adding the surfaces changes no
existing number. An unscored row and an unscored parameter contribute no
foreground variance, exactly as before -- so a project that scores nothing
must produce byte-identical results.
"""

import io

import pytest
from openpyxl import Workbook

from mapper.api.bom import _BOM_COLUMNS, _PEDIGREE_COLUMNS, _parse_bom_workbook, _walk_for_export
from mapper.core.monte_carlo_engine import collect_param_draws, collect_row_draws, sigma_of
from mapper.core.pedigree import INDICATORS, UNCERTAINTY_FACTORS
from mapper.models.bom_schemas import BOMNode, EcoinventLink, FlattenedMaterial, RowUncertainty
from mapper.models.parameter_schemas import Parameter, ParameterTable, ParamUncertainty

HEADER = {c: i for i, c in enumerate(_BOM_COLUMNS)}


# ── one table, not two ────────────────────────────────────────────────────────


def test_the_served_table_is_the_engine_s_own():
    """The UI needs the factors for a live GSD² preview. Serving them keeps one
    table -- a hard-coded copy in the frontend would drift silently, because
    both copies would keep producing plausible numbers."""
    from fastapi.testclient import TestClient

    from mapper.main import app

    body = TestClient(app).get("/api/lca/pedigree").json()
    assert body["indicators"] == list(INDICATORS)
    for ind, factors in UNCERTAINTY_FACTORS.items():
        assert body["factors"][ind] == list(factors)
    # And the convention travels with it, so the /2 is stated where someone
    # reading a score can see it.
    assert "/ 2" in body["convention"] or "/2" in body["convention"]


def test_the_columns_map_onto_the_engine_s_indicators():
    """A workbook column that named an indicator the engine does not know would
    be silently dropped by `pedigree_variance`, narrowing the spread."""
    assert set(_PEDIGREE_COLUMNS.values()) == set(INDICATORS)


# ── the workbook route ────────────────────────────────────────────────────────


def _sheet(rows: list[dict]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(_BOM_COLUMNS)
    for r in rows:
        ws.append([r.get(c, "") for c in _BOM_COLUMNS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _row(**over):
    base = {
        "Stage": "Manufacturing", "Scope": "inflows", "Parent": "",
        "Name": "Manufacturing", "Type": "component", "Quantity": 1, "Unit": "piece",
    }
    base.update(over)
    return base


def _material(name, **over):
    base = {
        "Stage": "Manufacturing", "Parent": "Manufacturing", "Name": name,
        "Type": "material", "Quantity": 10, "Unit": "kg",
        "Ecoinvent Database": "db", "Ecoinvent Code": f"code-{name}",
    }
    base.update(over)
    return base


def _parse(rows):
    from openpyxl import load_workbook
    return _parse_bom_workbook(load_workbook(_sheet(rows)))


def _find(tree, name):
    for root in tree:
        stack = [root]
        while stack:
            n = stack.pop()
            if n.name == name:
                return n
            stack.extend(n.children or [])
    return None


def test_pedigree_columns_import():
    res = _parse([
        _row(),
        _material("Steel", **{
            "Pedigree Reliability": 3,
            "Pedigree Completeness": 2,
            "Pedigree Temporal": 4,
            "Basic Variance": 0.01,
        }),
    ])
    node = _find(res[0], "Steel")
    assert node.uncertainty is not None
    assert node.uncertainty.pedigree == {
        "reliability": 3, "completeness": 2, "temporal correlation": 4,
    }
    assert node.uncertainty.basic_variance == 0.01


def test_blank_columns_leave_the_row_unscored():
    """Adding a column must never turn an unscored row into a scored one."""
    node = _find(_parse([_row(), _material("Steel")])[0], "Steel")
    assert node.uncertainty is None
    assert sigma_of(node.uncertainty) == 0.0


def test_score_one_is_not_recorded():
    """Score 1 adds no uncertainty, so storing it would only make an unscored
    row LOOK scored in the UI and the export."""
    node = _find(_parse([_row(), _material("Steel", **{"Pedigree Reliability": 1})])[0], "Steel")
    assert node.uncertainty is None


@pytest.mark.parametrize("bad", [0, 6, -1, "abc"])
def test_out_of_range_scores_warn_and_are_ignored_without_dropping_the_row(bad):
    tree, warnings = _parse([_row(), _material("Steel", **{"Pedigree Reliability": bad})])[:2]
    node = _find(tree, "Steel")
    assert node is not None, "the row itself must survive a bad score"
    assert node.quantity == 10
    assert node.uncertainty is None
    assert any("Pedigree Reliability" in w for w in warnings)


def test_an_expression_row_s_pedigree_columns_are_refused_at_IMPORT():
    """The engine already rejects the combination at compute time. Importing it
    silently would leave the user with a workbook whose whole point fails only
    when they press Run -- so the bulk route refuses it too, loudly."""
    tree, warnings = _parse([
        _row(Stage="Use Phase", Scope="stock", Name="Use Phase"),
        _material("Electricity", **{
            "Stage": "Use Phase", "Parent": "Use Phase",
            "Quantity": "d_annual * p_bev",
            "Pedigree Reliability": 3,
        }),
    ])[:2]
    node = _find(tree, "Electricity")
    assert node.quantity_expression == "d_annual * p_bev"
    assert node.uncertainty is None, "an expression row must not carry its own"
    assert any("expression" in w.lower() for w in warnings)
    assert any("parameters" in w.lower() for w in warnings)


def test_pedigree_round_trips_so_an_in_app_edit_survives_a_re_import():
    """Re-importing is what orphaned the WP5 cohort mapping once. A score set in
    the UI must come back out in the export, or the bulk route silently wipes
    the precise route."""
    rows: list[list] = []
    root = BOMNode(
        id="s", name="Manufacturing", node_type="component", scope="inflows",
        children=[BOMNode(
            id="m", name="Steel", node_type="material", quantity=10.0, unit="kg",
            ecoinvent_activity=EcoinventLink(database="db", code="c", name="Steel"),
            uncertainty=RowUncertainty(
                pedigree={"reliability": 3, "further technological correlation": 4},
                basic_variance=0.02,
            ),
        )],
    )
    _walk_for_export(root, root.name, "", rows, stage_scope="inflows")

    steel = rows[1]
    assert steel[HEADER["Pedigree Reliability"]] == 3
    assert steel[HEADER["Pedigree Technological"]] == 4
    assert steel[HEADER["Pedigree Completeness"]] == ""   # unset stays blank
    assert steel[HEADER["Basic Variance"]] == 0.02

    # ...and back in.
    back = _find(_parse([
        _row(),
        _material("Steel", **{
            "Pedigree Reliability": 3,
            "Pedigree Technological": 4,
            "Basic Variance": 0.02,
        }),
    ])[0], "Steel")
    assert back.uncertainty.pedigree == root.children[0].uncertainty.pedigree
    assert back.uncertainty.basic_variance == 0.02


def test_an_unscored_row_exports_blanks_not_zeros():
    """A zero would import back as an out-of-range score and warn on every row."""
    rows: list[list] = []
    root = BOMNode(
        id="s", name="Manufacturing", node_type="component",
        children=[BOMNode(id="m", name="Steel", node_type="material", quantity=1.0, unit="kg")],
    )
    _walk_for_export(root, root.name, "", rows)
    for col in list(_PEDIGREE_COLUMNS) + ["Basic Variance"]:
        assert rows[1][HEADER[col]] == ""


def test_an_older_workbook_without_the_columns_still_imports():
    """`col()` returns a default for an absent column, so a pre-pedigree file
    imports unchanged -- as unscored."""
    from openpyxl import load_workbook

    legacy_cols = [c for c in _BOM_COLUMNS if c not in _PEDIGREE_COLUMNS and c != "Basic Variance"]
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(legacy_cols)
    for r in (_row(), _material("Steel")):
        ws.append([r.get(c, "") for c in legacy_cols])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    tree = _parse_bom_workbook(load_workbook(buf))[0]
    node = _find(tree, "Steel")
    assert node is not None
    assert node.uncertainty is None


# ── the parameter surface ─────────────────────────────────────────────────────


def test_a_scored_parameter_is_collected_only_when_the_bom_references_it():
    """A project-wide table may score parameters this archetype never touches;
    listing those as zero-share contributors is noise."""
    table = ParameterTable(parameters={
        "d_annual": Parameter(
            name="d_annual", base_value=15000.0,
            uncertainty=ParamUncertainty(pedigree={"reliability": 3}),
        ),
        "unused": Parameter(
            name="unused", base_value=1.0,
            uncertainty=ParamUncertainty(pedigree={"reliability": 5}),
        ),
    })
    drawn = collect_param_draws(table, referenced={"d_annual"})
    assert [d.name for d in drawn] == ["d_annual"]
    assert drawn[0].sigma > 0


def test_an_unscored_parameter_contributes_nothing():
    table = ParameterTable(parameters={
        "d_annual": Parameter(name="d_annual", base_value=15000.0),
    })
    assert collect_param_draws(table, referenced={"d_annual"}) == []


# ── the invariant that matters most ───────────────────────────────────────────


def test_a_project_that_scores_nothing_is_unchanged():
    """Adding the UI must not change any existing number.

    Every path that could introduce foreground variance is checked to be inert
    when nothing is scored: rows, parameters, and the sigma each would produce.
    """
    rows = [
        FlattenedMaterial(
            node_id=f"n{i}", name=f"Row {i}", quantity=1.0, unit="kg",
            ecoinvent_activity=EcoinventLink(database="db", code=f"c{i}", name="x"),
        )
        for i in range(20)
    ]
    assert collect_row_draws(rows) == []

    table = ParameterTable(parameters={
        f"p{i}": Parameter(name=f"p{i}", base_value=float(i)) for i in range(44)
    })
    assert collect_param_draws(table, referenced={f"p{i}" for i in range(44)}) == []

    # And a row that IS scored is the only thing that produces a factor.
    assert sigma_of(None) == 0.0
    scored = RowUncertainty(pedigree={"reliability": 3})
    assert sigma_of(scored) > 0.0
