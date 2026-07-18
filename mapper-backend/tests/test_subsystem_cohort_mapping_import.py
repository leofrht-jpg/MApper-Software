# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Subsystem cohort-mapping Excel template export + validate-only import.

Mirrors the primary system's cohort-mapping Template/Upload and the sibling
dependency-rules import: populated .xlsx template (cohort keys pre-filled) +
reject-the-whole-file-on-any-error import (no partial). The client confirms the
destructive replace and saves via the existing subsystem update endpoint.

Run: python -m pytest mapper-backend/tests/test_subsystem_cohort_mapping_import.py -v
"""
from __future__ import annotations

import asyncio
import io
import json

import pytest
from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from starlette.datastructures import Headers, UploadFile

from mapper.api import bom as bom_api
from mapper.api import dsm as dsm_api
from mapper.api import subsystems as sub_api
from mapper.models.bom_schemas import Archetype
from mapper.models.dsm_schemas import DimensionDef, SystemDefinition, TimeHorizon
from mapper.models.subsystem_schemas import Subsystem

SYS_ID = "sys-cm-test"
SUB_ID = "sub-cm-test"
ARC_ID = "arc-cm-test"
ARC_NAME = "Charging Station BOM"


@pytest.fixture()
def system_subsystem_archetype():
    """Primary system + a dependent subsystem (charger dims) + one project
    archetype, registered in the API modules' in-memory state."""
    primary = SystemDefinition(
        id=SYS_ID,
        name="Test Fleet",
        time_horizon=TimeHorizon(start_year=2025, end_year=2030),
        dimensions=[
            DimensionDef(name="f", display_name="Fuel", labels=["CNG", "Diesel"]),
            DimensionDef(name="age", display_name="Age", labels=[], is_age=True),
        ],
    )
    sub = Subsystem(
        id=SUB_ID,
        name="Fuel Infrastructure",
        type="dependent",
        dimensions=[DimensionDef(name="station", display_name="Station", labels=["Default", "Large"])],
        depends_on=SYS_ID,
        dependency_rules=[],
    )
    arc = Archetype(id=ARC_ID, name=ARC_NAME)
    dsm_api._proj_systems()[SYS_ID] = primary
    sub_api._sys_subs(SYS_ID)[SUB_ID] = sub
    bom_api._proj_archetypes()[ARC_ID] = arc
    yield sub
    dsm_api._proj_systems().pop(SYS_ID, None)
    sub_api._sys_subs(SYS_ID).pop(SUB_ID, None)
    bom_api._proj_archetypes().pop(ARC_ID, None)


def _map_xlsx(rows: list[list], headers: list[str] | None = None, sheet: str = "Mappings") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers or ["dependent_archetype", "bom_archetype", "scale"])
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _upload(data: bytes, filename: str = "cohort_mapping.xlsx") -> UploadFile:
    return UploadFile(
        file=io.BytesIO(data),
        filename=filename,
        headers=Headers({"content-type": "application/octet-stream"}),
    )


# ── Template export ─────────────────────────────────────────────────────────


def test_template_filename_no_uuid(system_subsystem_archetype):
    resp = asyncio.run(sub_api.cohort_mapping_template(SYS_ID, SUB_ID))
    assert resp.status_code == 200
    cd = resp.headers["content-disposition"]
    # Static + human-readable: spaces→_, lowered. No UUID / system id.
    assert "cohort_mapping_fuel_infrastructure_template.xlsx" in cd
    assert SUB_ID not in cd and SYS_ID not in cd


def test_template_columns_and_prepopulated_cohort_keys(system_subsystem_archetype):
    resp = asyncio.run(sub_api.cohort_mapping_template(SYS_ID, SUB_ID))
    wb = load_workbook(io.BytesIO(resp.body))
    assert wb.sheetnames == ["Mappings", "Reference"]

    ws = wb["Mappings"]
    header = [c.value for c in ws[1]]
    assert header == ["dependent_archetype", "bom_archetype", "scale"]
    # Two cohort keys (Default, Large) pre-populated; bom_archetype + scale blank.
    dep_col = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    assert set(dep_col) == {"Default", "Large"}
    assert all(ws.cell(row=r, column=2).value in (None, "") for r in range(2, ws.max_row + 1))

    # Reference sheet lists valid BOM archetype names + is locked.
    ref = wb["Reference"]
    ref_text = "\n".join(str(c.value) for row in ref.iter_rows() for c in row if c.value is not None)
    assert ARC_NAME in ref_text
    assert ref.protection.sheet is True


# ── Import: valid ───────────────────────────────────────────────────────────


def test_import_valid_returns_mapping(system_subsystem_archetype):
    data = _map_xlsx([
        ["Default", ARC_NAME, 1.5],
        ["Large", ARC_NAME, ""],   # blank scale → default 1.0
    ])
    resp = asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 200
    body = json.loads(resp.body)
    assert body["ok"] is True
    assert body["mappings"] == {
        "Default": {"archetype_id": ARC_ID, "scaling_factor": 1.5},
        "Large": {"archetype_id": ARC_ID, "scaling_factor": 1.0},
    }


def test_import_blank_bom_is_unmapped_not_error(system_subsystem_archetype):
    data = _map_xlsx([
        ["Default", ARC_NAME, 1.0],
        ["Large", "", ""],  # unmapped — valid, omitted from result
    ])
    resp = asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 200
    body = json.loads(resp.body)
    assert body["ok"] is True
    assert set(body["mappings"].keys()) == {"Default"}


# ── Import: full rejection on any invalid row ───────────────────────────────


def test_import_unknown_cohort_key_rejected(system_subsystem_archetype):
    data = _map_xlsx([
        ["Default", ARC_NAME, 1.0],
        ["NotACohort", ARC_NAME, 1.0],
    ])
    resp = asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert body["ok"] is False
    assert any(e["row"] == 3 and e["field"] == "dependent_archetype" for e in body["errors"])


def test_import_unknown_bom_archetype_rejected(system_subsystem_archetype):
    data = _map_xlsx([["Default", "No Such Archetype", 1.0]])
    resp = asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert any(e["field"] == "bom_archetype" for e in body["errors"])


def test_import_non_positive_scale_rejected(system_subsystem_archetype):
    data = _map_xlsx([["Default", ARC_NAME, -2]])
    resp = asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(data)))
    assert resp.status_code == 422
    body = json.loads(resp.body)
    assert any(e["field"] == "scale" for e in body["errors"])


def test_import_rejects_non_xlsx(system_subsystem_archetype):
    with pytest.raises(HTTPException) as ei:
        asyncio.run(sub_api.import_cohort_mapping(SYS_ID, SUB_ID, _upload(b"a,b\n1,2", "m.csv")))
    assert ei.value.status_code == 400
    assert ".csv" in ei.value.detail or "xlsx" in ei.value.detail.lower()


def test_import_unknown_subsystem_404(system_subsystem_archetype):
    with pytest.raises(HTTPException) as ei:
        asyncio.run(sub_api.import_cohort_mapping(SYS_ID, "nope", _upload(_map_xlsx([]))))
    assert ei.value.status_code == 404


def test_template_unknown_subsystem_404(system_subsystem_archetype):
    with pytest.raises(HTTPException) as ei:
        asyncio.run(sub_api.cohort_mapping_template(SYS_ID, "nope"))
    assert ei.value.status_code == 404
