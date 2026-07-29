# SPDX-License-Identifier: MPL-2.0
# This Source Code Form is subject to the terms of the Mozilla Public
# License, v. 2.0. If a copy of the MPL was not distributed with this
# file, You can obtain one at https://mozilla.org/MPL/2.0/.
#
# © Copyright 2026 Technical University of Denmark
# Lead developer: Leonardo Ferhati

"""Subsystem DSM (stock/flow) Excel export — `_build_subsystem_dsm_workbook`.

Covers both scopes (combined = leading System column; subsystem = subsystem
rows only) and both modes (rules → Dependency rules sheet present; manual →
absent). No raw archetype UUID may appear in any cell (archetype NAMES are
resolved, cohort keys are the readable dependent-archetype labels).
"""
from __future__ import annotations

import io

from openpyxl import load_workbook

from mapper.api.subsystems import _build_subsystem_dsm_workbook
from mapper.models.bom_schemas import Archetype, CohortMapping, CohortMappingEntry
from mapper.models.dsm_schemas import (
    DimensionDef,
    SimulationResult,
    SimulationSummary,
    SystemDefinition,
    TimeHorizon,
    YearResult,
)
from mapper.models.subsystem_schemas import (
    DependencyRule,
    Subsystem,
    SubsystemCohortMapping,
)

SUB_ARC_UUID = "arc-uuid-charger-9f3a"
PRIM_ARC_UUID = "arc-uuid-bev-1122"


def _yr(year, stock, inflow, outflow):
    return YearResult(
        year=year, stock=stock, stock_by_age={k: {0: v} for k, v in stock.items()},
        inflow=inflow, outflow=outflow, outflow_by_age={},
    )


def _sub_result():
    return SimulationResult(
        system_id="sub-1",
        years=[
            _yr(2030, {"CNG Station": 10.0}, {"CNG Station": 10.0}, {}),
            _yr(2031, {"CNG Station": 18.0}, {"CNG Station": 8.0}, {"CNG Station": 0.0}),
        ],
        summary=SimulationSummary(total_stock_start=10.0, total_stock_end=18.0, total_inflows=18.0, total_outflows=0.0),
    )


def _primary_result():
    return SimulationResult(
        system_id="sys-1",
        years=[
            _yr(2030, {"BEV-LFP|SUV": 100.0}, {"BEV-LFP|SUV": 100.0}, {}),
            _yr(2031, {"BEV-LFP|SUV": 150.0}, {"BEV-LFP|SUV": 50.0}, {}),
        ],
        summary=SimulationSummary(total_stock_start=100.0, total_stock_end=150.0, total_inflows=150.0, total_outflows=0.0),
    )


def _subsystem(mode="rules"):
    return Subsystem(
        id="sub-1", name="Fueling Infrastructure", type="dependent", depends_on="sys-1",
        dimensions=[DimensionDef(name="infrastructure_type", display_name="Infrastructure Type", labels=["CNG Station"])],
        mode=mode,
        dependency_rules=[DependencyRule(id="r1", dependent_archetype_id="CNG Station",
                                         driver_filter={"fuel": ["CNG"]}, expression="filtered_stock * 0.1")]
        if mode == "rules" else [],
        manual_inflows={"CNG Station": {2030: 10.0, 2031: 8.0}} if mode == "manual" else {},
        initial_stock={"CNG Station": 5.0},
        cohort_mappings={"CNG Station": SubsystemCohortMapping(archetype_id=SUB_ARC_UUID, scaling_factor=2.0, color="#60a5fa")},
        unit_name="chargers",
    )


def _primary_def():
    return SystemDefinition(
        id="sys-1", name="Car Fleet",
        time_horizon=TimeHorizon(start_year=2030, end_year=2031),
        dimensions=[DimensionDef(name="fuel", display_name="Fuel", labels=["BEV-LFP"]),
                    DimensionDef(name="size", display_name="Size", labels=["SUV"])],
    )


def _archetypes():
    return {
        SUB_ARC_UUID: Archetype(id=SUB_ARC_UUID, name="Charging Infrastructure"),
        PRIM_ARC_UUID: Archetype(id=PRIM_ARC_UUID, name="BEV-LFP SUV"),
    }


def _primary_cohort_mapping():
    return CohortMapping(
        mfa_system_id="sys-1",
        mappings=[CohortMappingEntry(cohort_key="BEV-LFP|SUV", archetype_id=PRIM_ARC_UUID, scaling_factor=1.0)],
        row_colors={"BEV-LFP|SUV": "#22c55e"},
    )


def _build(scope, mode="rules"):
    content = _build_subsystem_dsm_workbook(
        scope=scope,
        primary_def=_primary_def(),
        primary_result=_primary_result() if scope == "combined" else None,
        primary_initial_stock={"BEV-LFP|SUV": 20.0} if scope == "combined" else {},
        primary_cohort_mapping=_primary_cohort_mapping(),
        subsystem=_subsystem(mode),
        subsystem_result=_sub_result(),
        archetypes=_archetypes(),
    )
    return load_workbook(io.BytesIO(content))


def _all_text(wb_or_ws) -> str:
    sheets = getattr(wb_or_ws, "worksheets", None) or [wb_or_ws]
    out = []
    for ws in sheets:
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is not None:
                    out.append(str(v))
    return "\n".join(out)


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


# ── Scope (b): subsystem only ────────────────────────────────────────────────


def test_subsystem_scope_sheets_and_no_system_column():
    wb = _build("subsystem")
    assert set(wb.sheetnames) == {
        "Summary", "Stock over time", "Inflows and outflows",
        "Initial stock", "Cohort mappings", "Dependency rules",
    }
    # No System column in subsystem-only scope.
    assert _rows(wb["Stock over time"])[0] == ("Year", "Cohort", "Archetype", "Count")
    # Stock rows are subsystem cohorts, archetype NAME resolved.
    text = _all_text(wb["Stock over time"])
    assert "CNG Station" in text and "Charging Infrastructure" in text
    # Aggregate flow sheet: two series (inflow/outflow totals).
    assert _rows(wb["Inflows and outflows"])[0] == ("Year", "Total inflow", "Total outflow")


def test_subsystem_scope_no_primary_rows():
    wb = _build("subsystem")
    # Primary archetype / cohort must NOT appear in a subsystem-only export.
    text = _all_text(wb)
    assert "BEV-LFP" not in text and "Car Fleet" not in text


# ── Scope (a): main system + subsystem ───────────────────────────────────────


def test_combined_scope_has_system_column_and_both_rows():
    wb = _build("combined")
    assert _rows(wb["Stock over time"])[0] == ("System", "Year", "Cohort", "Archetype", "Count")
    text = _all_text(wb["Stock over time"])
    # Both primary and subsystem rows present, labelled by System.
    assert "Car Fleet" in text and "Fueling Infrastructure" in text
    assert "BEV-LFP SUV" in text and "Charging Infrastructure" in text
    # Flows + initial stock + cohort mappings also carry System.
    assert _rows(wb["Inflows and outflows"])[0][0] == "System"
    assert _rows(wb["Initial stock"])[0][0] == "System"
    assert _rows(wb["Cohort mappings"])[0][0] == "System"


def test_combined_cohort_mappings_include_colour():
    wb = _build("combined")
    text = _all_text(wb["Cohort mappings"])
    assert "#60a5fa" in text  # subsystem colour
    assert "#22c55e" in text  # primary row colour


# ── No UUID in any cell (both scopes) ────────────────────────────────────────


def test_no_archetype_uuid_in_any_cell():
    for scope in ("subsystem", "combined"):
        wb = _build(scope)
        text = _all_text(wb)
        assert SUB_ARC_UUID not in text, f"{scope}: subsystem archetype UUID leaked"
        assert PRIM_ARC_UUID not in text, f"{scope}: primary archetype UUID leaked"


# ── Both modes ───────────────────────────────────────────────────────────────


def test_rules_mode_has_dependency_rules_sheet():
    wb = _build("subsystem", mode="rules")
    assert "Dependency rules" in wb.sheetnames
    text = _all_text(wb["Dependency rules"])
    assert "CNG Station" in text and "filtered_stock" in text


def test_manual_mode_omits_dependency_rules_sheet():
    wb = _build("subsystem", mode="manual")
    assert "Dependency rules" not in wb.sheetnames
    # Stock/flows still populate in manual mode.
    assert len(_rows(wb["Stock over time"])) > 1


def test_summary_reports_mode_and_scope():
    assert "Subsystem only" in _all_text(_build("subsystem")["Summary"])
    assert "Main system + subsystem" in _all_text(_build("combined")["Summary"])
    assert "manual" in _all_text(_build("subsystem", mode="manual")["Summary"])
